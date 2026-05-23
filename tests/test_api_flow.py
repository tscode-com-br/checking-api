import asyncio
import http.cookiejar
import io
import logging
import math
import os
import json
import pytest
import shutil
import socket
import subprocess
import sys
import threading
import time
import urllib.parse
import urllib.request
import uuid
from contextlib import closing, contextmanager
from datetime import date, datetime, timedelta
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch
from zoneinfo import ZoneInfo

from cryptography.fernet import Fernet
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
from pydantic import ValidationError
from sqlalchemy import delete, select, text
import uvicorn

# Override settings before app import.
os.environ["APP_ENV"] = "development"
os.environ["DATABASE_URL"] = "sqlite+pysqlite:///./test_checking.db"
os.environ["FORMS_URL"] = "https://example.com/form"
os.environ["DEVICE_SHARED_KEY"] = "device-test-key"
os.environ["MOBILE_APP_SHARED_KEY"] = "mobile-test-key"
os.environ["PROVIDER_SHARED_KEY"] = "PETROBRASP80P82P83"
os.environ["ADMIN_SESSION_SECRET"] = "test-admin-session-secret"
os.environ["BOOTSTRAP_ADMIN_KEY"] = "HR70"
os.environ["BOOTSTRAP_ADMIN_NAME"] = "Tamer Salmem"
os.environ["BOOTSTRAP_ADMIN_PASSWORD"] = "eAcacdLe2"
os.environ["FORMS_QUEUE_ENABLED"] = "false"
os.environ["TRANSPORT_EXPORTS_DIR"] = "./test_transport_exports"

test_db = Path("test_checking.db")
if test_db.exists():
    test_db.unlink()

test_transport_exports_dir = Path("test_transport_exports")
if test_transport_exports_dir.exists():
    shutil.rmtree(test_transport_exports_dir)

from fastapi.testclient import TestClient

from sistema.app.main import app, should_serve_static_site
from sistema.app import database as database_module
from sistema.app import http_runtime as http_runtime_module
from sistema.app.core.config import settings
from sistema.app.database import Base, SessionLocal, engine
from sistema.app.models import (
    AdminAccessRequest,
    CheckEvent,
    CheckingHistory,
    FormsSubmission,
    ManagedLocation,
    Project,
    TransportCurrencyOption,
    TransportAssignment,
    TransportRequest,
    TransportVehicleSchedule,
    TransportVehicleScheduleException,
    User,
    UserProjectMembership,
    UserSyncEvent,
    Vehicle,
    Workplace,
)
from sistema.app.routers import admin as admin_router
from sistema.app.services.admin_updates import AdminUpdatesBroker, admin_updates_broker, notify_transport_data_changed
from sistema.app.services.transport_ai_llm_settings import TransportAILlmSettingsEncryptionError
from sistema.app.services.forms_worker import FormsWorker
from sistema.app.services.forms_queue import process_forms_submission_queue_once
from sistema.app.services import forms_queue as forms_queue_module
from sistema.app import forms_worker_healthcheck as forms_worker_healthcheck_module
from sistema.app.services import forms_worker as forms_worker_module
from sistema.app.services import location_settings as location_settings_module
from sistema.app.services import transport as transport_service_module
from sistema.app.services import transport_proposals as transport_proposal_service_module
from sistema.app.services import transport_reevaluation_events as transport_reevaluation_module
from sistema.app.services import user_activity as user_activity_module
from sistema.app.services.admin_project_scope import (
    admin_monitors_project,
    dump_admin_monitored_projects,
    extract_admin_monitored_projects,
    normalize_admin_monitored_project_names,
    resolve_effective_admin_monitored_projects,
)
from sistema.app.schemas import (
    AdminManagementRow,
    AdminProfileUpdateRequest,
    HealthComponentResponse,
    TransportProposalDecision,
)
from sistema.app.services.admin_auth import ensure_default_admin, profile_can_view_activity_time, user_can_view_activity_time
from sistema.app.services.managed_locations import dump_location_projects
from sistema.app.services.passwords import hash_password, verify_password
from sistema.app.services.project_catalog import build_project_fields_for_country, seed_default_projects
from sistema.app.services.time_utils import build_timezone_label, now_sgt
from sistema.app.services.user_projects import (
    add_user_project_membership,
    ensure_user_active_project_is_member,
    list_user_project_names,
)
from sistema.app.services.user_sync import find_user_by_chave, find_user_by_rfid, normalize_event_time
from sistema.app.routers import web_check as web_check_router


ADMIN_LOGIN_CHAVE = "HR70"
ADMIN_LOGIN_SENHA = "eAcacdLe2"
MOBILE_HEADERS = {"x-mobile-shared-key": "mobile-test-key"}
PROVIDER_HEADERS = {"x-provider-shared-key": "PETROBRASP80P82P83"}

Base.metadata.create_all(bind=engine)
seed_default_projects()


def login_admin(client: TestClient, *, chave: str = ADMIN_LOGIN_CHAVE, senha: str = ADMIN_LOGIN_SENHA):
    return client.post("/api/admin/auth/login", json={"chave": chave, "senha": senha})


def ensure_admin_session(client: TestClient) -> None:
    session_response = client.get("/api/admin/auth/session")
    assert session_response.status_code == 200
    if not session_response.json().get("authenticated"):
        login_response = login_admin(client)
        assert login_response.status_code == 200, login_response.text

    with SessionLocal() as db:
        admin = get_user_by_chave(db, ADMIN_LOGIN_CHAVE)
        grant_user_project_memberships(
            db,
            admin,
            db.execute(select(Project.name).order_by(Project.name)).scalars().all(),
        )
        db.commit()

    transport_session_response = client.get("/api/transport/auth/session")
    assert transport_session_response.status_code == 200
    if transport_session_response.json().get("authenticated"):
        return

    transport_login_response = client.post(
        "/api/transport/auth/verify",
        json={"chave": ADMIN_LOGIN_CHAVE, "senha": ADMIN_LOGIN_SENHA},
    )
    assert transport_login_response.status_code == 200, transport_login_response.text
    assert transport_login_response.json()["authenticated"] is True


def extract_transport_structured_detail(response) -> dict:
    payload = response.json()
    detail = payload.get("detail")
    if isinstance(detail, dict):
        return detail
    normalized_detail = str(detail or "").strip()
    return {
        "message": normalized_detail,
        "message_key": None,
        "message_params": {},
        "error_code": None,
        "issues": [],
        "technical_detail": normalized_detail,
    }


def ensure_web_user_exists(*, chave: str, projeto: str = "P80", nome: str = "Oriundo da Web") -> None:
    with SessionLocal() as db:
        existing = find_user_by_chave(db, chave)
        if existing is not None:
            return

        db.add(
            User(
                rfid=None,
                nome=nome,
                chave=chave,
                projeto=projeto,
                workplace=None,
                placa=None,
                end_rua=None,
                zip=None,
                email=None,
                local=None,
                checkin=None,
                time=None,
                last_active_at=now_sgt(),
                inactivity_days=0,
            )
        )
        db.commit()


def ensure_web_transport_address(
    *,
    chave: str,
    end_rua: str = "10 Marina Boulevard",
    zip_code: str = "123456",
) -> None:
    with SessionLocal() as db:
        user = get_user_by_chave(db, chave)
        assert user is not None
        user.end_rua = end_rua
        user.zip = zip_code
        db.commit()


def ensure_project_exists(project_name: str) -> None:
    normalized_name = str(project_name).strip().upper()
    assert normalized_name

    with SessionLocal() as db:
        existing = db.execute(select(Project).where(Project.name == normalized_name)).scalar_one_or_none()
        if existing is not None:
            return

        db.add(Project(name=normalized_name, **build_project_fields_for_country()))
        bootstrap_admin = find_user_by_chave(db, settings.bootstrap_admin_key)
        if bootstrap_admin is not None:
            ensure_user_active_project_is_member(db, bootstrap_admin)
            add_user_project_membership(db, bootstrap_admin, normalized_name)
        db.commit()


def build_rectangle_coordinates(
    latitude: float,
    longitude: float,
    *,
    latitude_delta: float = 0.0002,
    longitude_delta: float = 0.0002,
) -> list[dict[str, float]]:
    base_latitude = float(latitude)
    base_longitude = float(longitude)
    return [
        {"latitude": base_latitude, "longitude": base_longitude},
        {"latitude": base_latitude + latitude_delta, "longitude": base_longitude},
        {"latitude": base_latitude + latitude_delta, "longitude": base_longitude + longitude_delta},
        {"latitude": base_latitude, "longitude": base_longitude + longitude_delta},
    ]


def register_web_password(
    client: TestClient,
    *,
    chave: str,
    senha: str = "abc123",
    projeto: str = "P80",
    ensure_user_exists: bool = True,
):
    if ensure_user_exists:
        ensure_web_user_exists(chave=chave, projeto=projeto)

    return client.post(
        "/api/web/auth/register-password",
        json={"chave": chave, "senha": senha},
    )


def login_web_password(client: TestClient, *, chave: str, senha: str):
    return client.post(
        "/api/web/auth/login",
        json={"chave": chave, "senha": senha},
    )


def clear_forms_queue_backlog() -> None:
    with SessionLocal() as db:
        db.execute(delete(FormsSubmission).where(FormsSubmission.status.in_(("pending", "processing"))))
        db.commit()


def submit_web_check(
    client: TestClient,
    *,
    chave: str,
    projeto: str,
    action: str,
    event_time: datetime,
    client_event_id: str | None = None,
    local: str | None = None,
    informe: str = "normal",
):
    payload = {
        "chave": chave,
        "projeto": projeto,
        "action": action,
        "informe": informe,
        "event_time": event_time.isoformat(),
        "client_event_id": client_event_id or f"web-check-{uuid.uuid4().hex}",
    }
    if local is not None:
        payload["local"] = local
    return client.post("/api/web/check", json=payload)


def get_admin_presence_row_by_chave(client: TestClient, *, endpoint: str, chave: str) -> dict[str, object]:
    response = client.get(f"/api/admin/{endpoint}")
    assert response.status_code == 200, response.text
    return next(row for row in response.json() if row["chave"] == chave)


def mocked_successful_forms_submit_with_statuses(
    self,
    action,
    chave,
    projeto,
    ontime=True,
    status_callback=None,
    **_kwargs,
):
    if status_callback is not None:
        for status_label in ("URL Carregada", "Preenchendo...", "Preenchido", "Enviado"):
            status_callback(status_label)
    return {
        "success": True,
        "message": f"mocked {action}",
        "retry_count": 0,
        "audit_events": [
            {
                "source": "forms",
                "action": "forms",
                "status": "completed",
                "message": "Forms submission completed",
                "details": f"chave={chave}; projeto={projeto}; ontime={ontime}",
            }
        ],
    }


def get_http_request_logs(caplog: pytest.LogCaptureFixture) -> list[dict[str, object]]:
    logs: list[dict[str, object]] = []
    for record in caplog.records:
        if record.name != "checking.http":
            continue
        try:
            payload = json.loads(record.getMessage())
        except json.JSONDecodeError:
            continue
        if payload.get("event") == "http_request":
            logs.append(payload)
    return logs


def get_forms_queue_logs(caplog: pytest.LogCaptureFixture) -> list[dict[str, object]]:
    logs: list[dict[str, object]] = []
    for record in caplog.records:
        if record.name != "checking.forms_queue":
            continue
        try:
            payload = json.loads(record.getMessage())
        except json.JSONDecodeError:
            continue
        logs.append(payload)
    return logs


def get_database_logs(caplog: pytest.LogCaptureFixture) -> list[dict[str, object]]:
    logs: list[dict[str, object]] = []
    for record in caplog.records:
        if record.name != "checking.db":
            continue
        try:
            payload = json.loads(record.getMessage())
        except json.JSONDecodeError:
            continue
        logs.append(payload)
    return logs


def make_test_key(prefix: str) -> str:
    normalized_prefix = str(prefix or "T").strip().upper()[:1]
    assert normalized_prefix and normalized_prefix.isalnum()
    return f"{normalized_prefix}{uuid.uuid4().hex[:3].upper()}"


def reserve_tcp_port() -> int:
    with closing(socket.socket(socket.AF_INET, socket.SOCK_STREAM)) as sock:
        sock.bind(("127.0.0.1", 0))
        sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        return int(sock.getsockname()[1])


def test_admin_project_scope_normalizes_and_serializes_unique_sorted_projects():
    assert normalize_admin_monitored_project_names([" p83 ", "P80", "p83", "P82 "]) == ["P80", "P82", "P83"]
    assert dump_admin_monitored_projects([" p83 ", "P80", "p83", "P82 "]) == '["P80","P82","P83"]'


def test_admin_project_scope_extracts_none_for_blank_and_invalid_payloads():
    assert extract_admin_monitored_projects(SimpleNamespace(admin_monitored_projects_json=None)) is None
    assert extract_admin_monitored_projects(SimpleNamespace(admin_monitored_projects_json="   ")) is None
    assert extract_admin_monitored_projects(SimpleNamespace(admin_monitored_projects_json="{invalid")) is None
    assert extract_admin_monitored_projects(SimpleNamespace(admin_monitored_projects_json='{"projects":["P80"]}')) is None


def test_admin_project_scope_extracts_explicit_projects_and_filters_unknown_catalog_entries():
    admin_user = SimpleNamespace(admin_monitored_projects_json='[" P83 ", "P80", "P83", " ", null]')
    assert extract_admin_monitored_projects(admin_user) == ["P80", "P83"]
    assert resolve_effective_admin_monitored_projects(admin_user, ["P83", "P80", "P99"]) == ["P80", "P83"]

    restricted_user = SimpleNamespace(admin_monitored_projects_json='["P80", "P99", "P80"]')
    assert resolve_effective_admin_monitored_projects(restricted_user, ["P83", "P80"]) == ["P80"]


def test_admin_project_scope_all_mode_and_explicit_mode_project_matching():
    unrestricted_user = SimpleNamespace(admin_monitored_projects_json=None)
    assert resolve_effective_admin_monitored_projects(unrestricted_user, ["P80", "P83"]) is None
    assert admin_monitors_project(unrestricted_user, "p80", ["P80", "P83"]) is True

    restricted_user = SimpleNamespace(admin_monitored_projects_json='["P80"]')
    assert admin_monitors_project(restricted_user, "P80", ["P80", "P83"]) is True
    assert admin_monitors_project(restricted_user, "P83", ["P80", "P83"]) is False
    assert admin_monitors_project(restricted_user, "", ["P80", "P83"]) is False


def test_admin_profile_update_request_accepts_backward_compatible_profile_only_payload():
    payload = AdminProfileUpdateRequest.model_validate({"perfil": 7})
    assert payload.perfil == 7
    assert payload.projects is None


def test_admin_profile_update_request_normalizes_and_deduplicates_projects():
    payload = AdminProfileUpdateRequest.model_validate(
        {"perfil": 2, "projects": [" p83 ", "P80", "p83"]}
    )
    assert payload.perfil == 2
    assert payload.projects == ["P80", "P83"]


def test_admin_profile_update_request_rejects_empty_or_non_list_projects():
    with pytest.raises(ValidationError, match="Selecione ao menos um projeto para o administrador"):
        AdminProfileUpdateRequest.model_validate({"perfil": 1, "projects": []})

    with pytest.raises(ValidationError, match="Os projetos do administrador devem ser enviados como lista"):
        AdminProfileUpdateRequest.model_validate({"perfil": 1, "projects": "P80"})


def test_admin_profile_update_request_can_validate_against_catalog_context():
    payload = AdminProfileUpdateRequest.model_validate(
        {"perfil": 3, "projects": ["P80", "P83"]},
        context={"allowed_project_names": ["P83", "P80", "P90"]},
    )
    assert payload.projects == ["P80", "P83"]

    with pytest.raises(ValidationError, match="Um ou mais projetos do administrador nao existem no catalogo atual"):
        AdminProfileUpdateRequest.model_validate(
            {"perfil": 3, "projects": ["P80", "P99"]},
            context={"allowed_project_names": ["P83", "P80", "P90"]},
        )


def test_admin_management_row_exposes_empty_projects_by_default():
    row = AdminManagementRow(
        id=1,
        row_type="admin",
        chave="HR70",
        nome="Tamer Salmem",
        perfil=999,
        status="active",
        status_label="Administrador",
        can_revoke=True,
        can_approve=False,
        can_reject=False,
        can_set_password=False,
    )
    assert row.projects == []


def test_admin_management_row_normalizes_projects_when_provided():
    row = AdminManagementRow(
        id=2,
        row_type="admin",
        chave="HR71",
        nome="Admin Secundario",
        perfil=100,
        projects=[" p83 ", "P80", "p83"],
        status="active",
        status_label="Administrador",
        can_revoke=True,
        can_approve=False,
        can_reject=False,
        can_set_password=False,
    )
    assert row.projects == ["P80", "P83"]


def test_profile_can_view_activity_time_requires_exact_profile_nine():
    assert profile_can_view_activity_time(9) is True
    assert profile_can_view_activity_time("9") is True
    assert profile_can_view_activity_time(1) is False
    assert profile_can_view_activity_time(0) is False
    assert profile_can_view_activity_time(999) is False
    assert profile_can_view_activity_time(None) is False


def test_user_can_view_activity_time_requires_exact_profile_nine():
    assert user_can_view_activity_time(SimpleNamespace(perfil=9)) is True
    assert user_can_view_activity_time(SimpleNamespace(perfil=999)) is False
    assert user_can_view_activity_time(SimpleNamespace(perfil=1)) is False
    assert user_can_view_activity_time(None) is False


def test_admin_update_profile_route_rejects_unknown_projects_from_catalog():
    with SessionLocal() as db:
        target_admin = User(
            rfid=None,
            chave="HA72",
            senha=hash_password("abc123"),
            perfil=100,
            nome="Admin Contrato Invalido",
            projeto="P80",
            workplace=None,
            placa=None,
            end_rua=None,
            zip=None,
            email=None,
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
        )
        db.add(target_admin)
        db.commit()
        db.refresh(target_admin)
        admin_id = target_admin.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            f"/api/admin/administrators/{admin_id}/profile",
            json={"perfil": 101, "projects": ["P80", "PX99"]},
        )

    assert response.status_code == 422, response.text
    payload = response.json()
    assert payload["detail"][0]["msg"] == "Value error, Um ou mais projetos do administrador nao existem no catalogo atual"


def test_administrators_endpoint_returns_real_membership_projects_for_admin_rows_only():
    with SessionLocal() as db:
        unrestricted_admin = User(
            rfid=None,
            chave="HA81",
            senha=hash_password("abc123"),
            perfil=1,
            nome="Admin Todos",
            projeto="P80",
            workplace=None,
            placa=None,
            end_rua=None,
            zip=None,
            email=None,
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
            admin_monitored_projects_json=dump_admin_monitored_projects(["P83"]),
        )
        restricted_admin = User(
            rfid=None,
            chave="HA82",
            senha=hash_password("abc123"),
            perfil=1,
            nome="Admin Restrito",
            projeto="P83",
            workplace=None,
            placa=None,
            end_rua=None,
            zip=None,
            email=None,
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
            admin_monitored_projects_json=None,
        )
        pending_request = AdminAccessRequest(
            chave="HA83",
            nome_completo="Admin Pendente",
            password_hash=hash_password("abc123"),
            requested_profile=1,
            requested_at=now_sgt(),
        )
        db.add_all([unrestricted_admin, restricted_admin, pending_request])
        db.flush()
        grant_user_project_memberships(db, unrestricted_admin, ["P80", "P82"])
        grant_user_project_memberships(db, restricted_admin, ["P83"])
        db.commit()

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.get("/api/admin/administrators")

    assert response.status_code == 200, response.text
    rows_by_key = {row["chave"]: row for row in response.json() if row["chave"] in {"HA81", "HA82", "HA83"}}
    assert rows_by_key["HA81"]["row_type"] == "admin"
    assert rows_by_key["HA81"]["projects"] == ["P80", "P82"]
    assert rows_by_key["HA82"]["row_type"] == "admin"
    assert rows_by_key["HA82"]["projects"] == ["P83"]
    assert rows_by_key["HA83"]["row_type"] == "request"
    assert rows_by_key["HA83"]["projects"] == []


def test_administrators_endpoint_does_not_fallback_to_legacy_active_project_without_real_memberships():
    with SessionLocal() as db:
        legacy_only_admin = User(
            rfid=None,
            chave="HA84",
            senha=hash_password("abc123"),
            perfil=1,
            nome="Admin Sem Membership Materializada",
            projeto="P83",
            workplace=None,
            placa=None,
            end_rua=None,
            zip=None,
            email=None,
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
            admin_monitored_projects_json=dump_admin_monitored_projects(["P80", "P83"]),
        )
        db.add(legacy_only_admin)
        db.commit()

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.get("/api/admin/administrators")

    assert response.status_code == 200, response.text
    row = next(
        item for item in response.json() if item["row_type"] == "admin" and item["chave"] == "HA84"
    )
    assert row["projects"] == []


def test_admin_update_profile_route_persists_real_memberships_and_audit_details():
    with SessionLocal() as db:
        target_admin = User(
            rfid=None,
            chave="HA73",
            senha=hash_password("abc123"),
            perfil=1,
            nome="Admin Persistencia",
            projeto="P80",
            workplace=None,
            placa=None,
            end_rua=None,
            zip=None,
            email=None,
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
            admin_monitored_projects_json=dump_admin_monitored_projects(["P82"]),
        )
        db.add(target_admin)
        db.flush()
        grant_user_project_memberships(db, target_admin, ["P80"])
        db.commit()
        db.refresh(target_admin)
        admin_id = target_admin.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            f"/api/admin/administrators/{admin_id}/profile",
            json={"perfil": 2, "projects": ["P83", " p80 "]},
        )

    assert response.status_code == 200, response.text
    assert response.json() == {
        "ok": True,
        "message": "Configuracoes do administrador atualizadas com sucesso.",
    }

    with SessionLocal() as db:
        refreshed_admin = db.get(User, admin_id)
        refreshed_projects = list_user_project_names(db, refreshed_admin)
        audit_event = db.execute(
            select(CheckEvent)
            .where(CheckEvent.request_path == f"/api/admin/administrators/{admin_id}/profile")
            .order_by(CheckEvent.id.desc())
        ).scalar_one()

    assert refreshed_admin is not None
    assert refreshed_admin.perfil == 3
    assert refreshed_projects == ["P80", "P83"]
    assert refreshed_admin.admin_monitored_projects_json is None
    assert audit_event.message == "Administrator configuration updated"
    assert "old_profile=1" in audit_event.details
    assert "new_profile=3" in audit_event.details
    assert 'old_projects=["P80"]' in audit_event.details
    assert 'new_projects=["P80","P83"]' in audit_event.details
    assert "old_active_project=P80" in audit_event.details
    assert "new_active_project=P80" in audit_event.details


def test_admin_update_profile_route_preserves_existing_memberships_when_projects_are_omitted():
    with SessionLocal() as db:
        legacy_admin = User(
            rfid=None,
            chave="HA74",
            senha=hash_password("abc123"),
            perfil=1,
            nome="Admin Compatibilidade",
            projeto="P80",
            workplace=None,
            placa=None,
            end_rua=None,
            zip=None,
            email=None,
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
            admin_monitored_projects_json=dump_admin_monitored_projects(["P80"]),
        )
        db.add(legacy_admin)
        db.flush()
        grant_user_project_memberships(db, legacy_admin, ["P80", "P82"])
        db.commit()
        db.refresh(legacy_admin)
        admin_id = legacy_admin.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        profile_only_response = client.post(
            f"/api/admin/administrators/{admin_id}/profile",
            json={"perfil": 2},
        )

    assert profile_only_response.status_code == 200, profile_only_response.text

    with SessionLocal() as db:
        refreshed_admin = db.get(User, admin_id)
        refreshed_projects = list_user_project_names(db, refreshed_admin)
        latest_audit_event = db.execute(
            select(CheckEvent)
            .where(CheckEvent.request_path == f"/api/admin/administrators/{admin_id}/profile")
            .order_by(CheckEvent.id.desc())
        ).scalars().first()

    assert refreshed_admin is not None
    assert latest_audit_event is not None
    assert refreshed_admin.perfil == 3
    assert refreshed_projects == ["P80", "P82"]
    assert refreshed_admin.admin_monitored_projects_json is None
    assert 'old_projects=["P80","P82"]' in latest_audit_event.details
    assert 'new_projects=["P80","P82"]' in latest_audit_event.details


def test_admin_approval_seeds_initial_membership_and_subsequent_profile_edit_updates_real_memberships():
    with SessionLocal() as db:
        requested_user = User(
            rfid=None,
            chave="HA91",
            senha=hash_password("abc123"),
            perfil=0,
            nome="Admin Novo Fase 6",
            projeto="P80",
            workplace=None,
            placa=None,
            end_rua=None,
            zip=None,
            email=None,
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
        )
        pending_request = AdminAccessRequest(
            chave="HA91",
            nome_completo="Admin Novo Fase 6",
            password_hash=hash_password("abc123"),
            requested_profile=1,
            requested_at=now_sgt(),
        )
        db.add_all([requested_user, pending_request])
        db.commit()
        db.refresh(requested_user)
        db.refresh(pending_request)
        request_id = pending_request.id
        user_id = requested_user.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        approval_response = client.post(
            f"/api/admin/administrators/requests/{request_id}/approve",
            json={"perfil": 1, "projects": ["P83"]},
        )
        update_response = client.post(
            f"/api/admin/administrators/{user_id}/profile",
            json={"perfil": 1, "projects": ["P80", "P83"]},
        )
        rows_response = client.get("/api/admin/administrators")

    assert approval_response.status_code == 200, approval_response.text
    assert update_response.status_code == 200, update_response.text
    assert rows_response.status_code == 200, rows_response.text

    with SessionLocal() as db:
        approved_admin = db.get(User, user_id)
        pending_request = db.get(AdminAccessRequest, request_id)
        approved_projects = list_user_project_names(db, approved_admin)

    assert approved_admin is not None
    assert approved_admin.admin_monitored_projects_json is None
    assert pending_request is None
    assert approved_projects == ["P80", "P83"]
    approved_row = next(
        row for row in rows_response.json() if row["row_type"] == "admin" and row["chave"] == "HA91"
    )
    assert approved_row["projects"] == ["P80", "P83"]


def test_admin_approval_preserves_existing_seed_project_and_transport_password_rule():
    original_password = "keep123"
    requested_password = "new123"

    with SessionLocal() as db:
        existing_user = User(
            rfid=None,
            chave="HA92",
            senha=hash_password(original_password),
            perfil=2,
            nome="Admin Transporte Existente",
            projeto="P80",
            workplace=None,
            placa=None,
            end_rua=None,
            zip=None,
            email=None,
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
            admin_monitored_projects_json=dump_admin_monitored_projects(["P80"]),
        )
        pending_request = AdminAccessRequest(
            chave="HA92",
            nome_completo="Admin Transporte Existente",
            password_hash=hash_password(requested_password),
            requested_profile=2,
            requested_at=now_sgt(),
        )
        db.add_all([existing_user, pending_request])
        db.commit()
        db.refresh(existing_user)
        db.refresh(pending_request)
        request_id = pending_request.id
        user_id = existing_user.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            f"/api/admin/administrators/requests/{request_id}/approve",
            json={"perfil": 2, "projects": ["P83"]},
        )

    assert response.status_code == 200, response.text

    with SessionLocal() as db:
        approved_admin = db.get(User, user_id)
        approved_projects = list_user_project_names(db, approved_admin)

    assert approved_admin is not None
    assert approved_admin.perfil == 3
    assert approved_admin.admin_monitored_projects_json is None
    assert approved_projects == ["P80"]
    assert approved_admin.senha is not None
    assert verify_password(original_password, approved_admin.senha) is False
    assert verify_password(requested_password, approved_admin.senha) is True


def test_admin_revocation_does_not_depend_on_monitored_scope_storage():
    with SessionLocal() as db:
        target_admin = User(
            rfid=None,
            chave="HA93",
            senha=hash_password("abc123"),
            perfil=1,
            nome="Admin Revogavel",
            projeto="P80",
            workplace=None,
            placa=None,
            end_rua=None,
            zip=None,
            email=None,
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
            admin_monitored_projects_json="{legacy-invalid-json",
        )
        db.add(target_admin)
        db.commit()
        db.refresh(target_admin)
        admin_id = target_admin.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(f"/api/admin/administrators/{admin_id}/revoke")

    assert response.status_code == 200, response.text

    with SessionLocal() as db:
        revoked_admin = db.get(User, admin_id)

    assert revoked_admin is not None
    assert revoked_admin.perfil == 0
    assert revoked_admin.admin_monitored_projects_json == "{legacy-invalid-json"


def test_bootstrap_admin_seed_resets_monitored_scope_to_all_mode():
    bootstrap_key = "HB70"

    with patch.object(settings, "bootstrap_admin_key", bootstrap_key), patch.object(
        settings,
        "bootstrap_admin_name",
        "Bootstrap Fase 6",
    ), patch.object(settings, "bootstrap_admin_password", "SeedPass123"):
        with SessionLocal() as db:
            bootstrap_user = User(
                rfid=None,
                chave=bootstrap_key,
                senha=hash_password("oldpass"),
                perfil=9,
                nome="Bootstrap Fase 6",
                projeto="P80",
                workplace=None,
                placa=None,
                end_rua=None,
                zip=None,
                email=None,
                local=None,
                checkin=None,
                time=None,
                last_active_at=now_sgt(),
                inactivity_days=0,
                admin_monitored_projects_json=dump_admin_monitored_projects(["P80"]),
            )
            db.add(bootstrap_user)
            db.commit()

            ensured_admin = ensure_default_admin(db)
            db.refresh(ensured_admin)

        assert ensured_admin.chave == bootstrap_key
        assert ensured_admin.admin_monitored_projects_json is None


def test_restricted_admin_scope_filters_presence_tables_by_monitored_projects():
    recent_time = now_sgt() - timedelta(hours=1)
    stale_time = now_sgt() - timedelta(hours=26)

    with SessionLocal() as db:
        restricted_admin = User(
            rfid=None,
            chave="S7A0",
            senha=hash_password("scope123"),
            perfil=1,
            nome="Admin Restrito Fase 7",
            projeto="P80",
            workplace=None,
            placa=None,
            end_rua=None,
            zip=None,
            email=None,
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
            admin_monitored_projects_json=dump_admin_monitored_projects(["P80"]),
        )
        rows = [
            User(
                rfid=None,
                chave="S7CI",
                nome="Checkin P80",
                projeto="P80",
                local="Porta 80",
                checkin=True,
                time=recent_time,
                last_active_at=recent_time,
                inactivity_days=0,
            ),
            User(
                rfid=None,
                chave="S7CJ",
                nome="Checkin P83",
                projeto="P83",
                local="Porta 83",
                checkin=True,
                time=recent_time,
                last_active_at=recent_time,
                inactivity_days=0,
            ),
            User(
                rfid=None,
                chave="S7CO",
                nome="Checkout P80",
                projeto="P80",
                local="Saida 80",
                checkin=False,
                time=recent_time,
                last_active_at=recent_time,
                inactivity_days=0,
            ),
            User(
                rfid=None,
                chave="S7CP",
                nome="Checkout P83",
                projeto="P83",
                local="Saida 83",
                checkin=False,
                time=recent_time,
                last_active_at=recent_time,
                inactivity_days=0,
            ),
            User(
                rfid=None,
                chave="S7IA",
                nome="Inativo P80",
                projeto="P80",
                local="Inativo 80",
                checkin=True,
                time=stale_time,
                last_active_at=stale_time,
                inactivity_days=1,
            ),
            User(
                rfid=None,
                chave="S7IB",
                nome="Inativo P83",
                projeto="P83",
                local="Inativo 83",
                checkin=False,
                time=stale_time,
                last_active_at=stale_time,
                inactivity_days=1,
            ),
        ]
        db.add(restricted_admin)
        db.add_all(rows)
        db.flush()
        grant_user_project_memberships(db, restricted_admin, ["P80"])
        db.commit()

    with TestClient(app) as client:
        login_response = login_admin(client, chave="S7A0", senha="scope123")
        assert login_response.status_code == 200, login_response.text

        checkin_response = client.get("/api/admin/checkin")
        checkout_response = client.get("/api/admin/checkout")
        inactive_response = client.get("/api/admin/inactive")

    assert checkin_response.status_code == 200, checkin_response.text
    assert checkout_response.status_code == 200, checkout_response.text
    assert inactive_response.status_code == 200, inactive_response.text

    checkin_keys = {row["chave"] for row in checkin_response.json()}
    checkout_keys = {row["chave"] for row in checkout_response.json()}
    inactive_keys = {row["chave"] for row in inactive_response.json()}

    assert "S7CI" in checkin_keys
    assert "S7CJ" not in checkin_keys
    assert "S7CO" in checkout_keys
    assert "S7CP" not in checkout_keys
    assert "S7IA" in inactive_keys
    assert "S7IB" not in inactive_keys


def test_unrestricted_admin_scope_keeps_all_presence_rows_visible():
    recent_time = now_sgt() - timedelta(hours=1)
    stale_time = now_sgt() - timedelta(hours=26)

    with SessionLocal() as db:
        unrestricted_admin = User(
            rfid=None,
            chave="S7A1",
            senha=hash_password("scope123"),
            perfil=1,
            nome="Admin Sem Restricao Fase 7",
            projeto="P80",
            workplace=None,
            placa=None,
            end_rua=None,
            zip=None,
            email=None,
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
            admin_monitored_projects_json=None,
        )
        rows = [
            User(
                rfid=None,
                chave="S7UI",
                nome="Checkin Livre P80",
                projeto="P80",
                local="Porta U80",
                checkin=True,
                time=recent_time,
                last_active_at=recent_time,
                inactivity_days=0,
            ),
            User(
                rfid=None,
                chave="S7UJ",
                nome="Checkin Livre P83",
                projeto="P83",
                local="Porta U83",
                checkin=True,
                time=recent_time,
                last_active_at=recent_time,
                inactivity_days=0,
            ),
            User(
                rfid=None,
                chave="S7UO",
                nome="Checkout Livre P80",
                projeto="P80",
                local="Saida U80",
                checkin=False,
                time=recent_time,
                last_active_at=recent_time,
                inactivity_days=0,
            ),
            User(
                rfid=None,
                chave="S7UP",
                nome="Checkout Livre P83",
                projeto="P83",
                local="Saida U83",
                checkin=False,
                time=recent_time,
                last_active_at=recent_time,
                inactivity_days=0,
            ),
            User(
                rfid=None,
                chave="S7UA",
                nome="Inativo Livre P80",
                projeto="P80",
                local="Inativo U80",
                checkin=True,
                time=stale_time,
                last_active_at=stale_time,
                inactivity_days=1,
            ),
            User(
                rfid=None,
                chave="S7UB",
                nome="Inativo Livre P83",
                projeto="P83",
                local="Inativo U83",
                checkin=False,
                time=stale_time,
                last_active_at=stale_time,
                inactivity_days=1,
            ),
        ]
        db.add(unrestricted_admin)
        db.add_all(rows)
        db.flush()
        grant_user_project_memberships(db, unrestricted_admin, ["P80", "P83"])
        db.commit()

    with TestClient(app) as client:
        login_response = login_admin(client, chave="S7A1", senha="scope123")
        assert login_response.status_code == 200, login_response.text

        checkin_response = client.get("/api/admin/checkin")
        checkout_response = client.get("/api/admin/checkout")
        inactive_response = client.get("/api/admin/inactive")

    assert checkin_response.status_code == 200, checkin_response.text
    assert checkout_response.status_code == 200, checkout_response.text
    assert inactive_response.status_code == 200, inactive_response.text

    checkin_keys = {row["chave"] for row in checkin_response.json()}
    checkout_keys = {row["chave"] for row in checkout_response.json()}
    inactive_keys = {row["chave"] for row in inactive_response.json()}

    assert {"S7UI", "S7UJ"}.issubset(checkin_keys)
    assert {"S7UO", "S7UP"}.issubset(checkout_keys)
    assert {"S7UA", "S7UB"}.issubset(inactive_keys)


@contextmanager
def live_app_server():
    port = reserve_tcp_port()
    config = uvicorn.Config(app, host="127.0.0.1", port=port, log_level="warning")
    server = uvicorn.Server(config)
    server.install_signal_handlers = lambda: None
    thread = threading.Thread(target=server.run, daemon=True)
    thread.start()

    deadline = time.monotonic() + 10
    while time.monotonic() < deadline:
        if getattr(server, "started", False):
            break
        try:
            with closing(socket.create_connection(("127.0.0.1", port), timeout=0.2)):
                break
        except OSError:
            time.sleep(0.05)
    else:
        server.should_exit = True
        thread.join(timeout=5)
        raise RuntimeError("Nao foi possivel iniciar o servidor HTTP para o teste E2E do admin.")

    try:
        yield f"http://127.0.0.1:{port}"
    finally:
        server.should_exit = True
        thread.join(timeout=10)


def build_cookie_opener() -> urllib.request.OpenerDirector:
    cookie_jar = http.cookiejar.CookieJar()
    return urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cookie_jar))


def perform_live_json_request(
    base_url: str,
    path: str,
    *,
    method: str = "GET",
    payload: dict[str, object] | None = None,
    query: dict[str, object] | None = None,
    headers: dict[str, str] | None = None,
    opener: urllib.request.OpenerDirector | None = None,
    timeout_seconds: float = 5.0,
) -> tuple[int, object, int]:
    resolved_url = f"{base_url}{path}"
    if query:
        resolved_url = f"{resolved_url}?{urllib.parse.urlencode(query, doseq=True)}"

    request_body = None
    if payload is not None:
        request_body = json.dumps(payload).encode("utf-8")

    request = urllib.request.Request(resolved_url, data=request_body, method=method.upper())
    request.add_header("Accept", "application/json")
    if request_body is not None:
        request.add_header("Content-Type", "application/json")
    for header_name, header_value in (headers or {}).items():
        request.add_header(header_name, header_value)

    started_at = time.perf_counter()
    opener_to_use = opener or urllib.request.build_opener()
    with opener_to_use.open(request, timeout=timeout_seconds) as response:
        raw_body = response.read()
        status_code = int(getattr(response, "status", response.getcode()))
    latency_ms = int(round((time.perf_counter() - started_at) * 1000))

    decoded_body = raw_body.decode("utf-8") if raw_body else ""
    try:
        parsed_body: object = json.loads(decoded_body) if decoded_body else {}
    except json.JSONDecodeError:
        parsed_body = decoded_body
    return status_code, parsed_body, latency_ms


def wait_for_condition(
    predicate,
    *,
    timeout_seconds: float,
    interval_seconds: float = 0.05,
    description: str,
):
    deadline = time.monotonic() + timeout_seconds
    last_value = None
    while time.monotonic() < deadline:
        last_value = predicate()
        if last_value:
            return last_value
        time.sleep(interval_seconds)
    raise AssertionError(f"Timed out waiting for {description}. Last observed value: {last_value!r}")


def summarize_latency_samples(samples: list[int]) -> dict[str, int]:
    assert samples
    ordered = sorted(int(sample) for sample in samples)
    p95_index = max(math.ceil(len(ordered) * 0.95) - 1, 0)
    return {
        "average_ms": int(round(sum(ordered) / len(ordered))),
        "count": len(ordered),
        "max_ms": ordered[-1],
        "min_ms": ordered[0],
        "p95_ms": ordered[p95_index],
    }


def start_controlled_slow_forms_worker_subprocess(*, processing_delay_seconds: float) -> subprocess.Popen[str]:
    repo_root = Path(__file__).resolve().parents[1]
    worker_script = f"""
import time

from sqlalchemy import func, select

from sistema.app.database import SessionLocal
from sistema.app.models import FormsSubmission
from sistema.app.services.forms_queue import process_forms_submission_queue_once
from sistema.app.services.forms_worker import FormsWorker


def slow_submit(self, action, chave, projeto, ontime=True):
    time.sleep({processing_delay_seconds!r})
    return {{
        \"success\": True,
        \"message\": \"controlled slow queue success\",
        \"retry_count\": 1,
        \"audit_events\": [
            {{
                \"source\": \"forms\",
                \"action\": \"forms\",
                \"status\": \"completed\",
                \"message\": \"Controlled slow queue success\",
                \"details\": \"sleep_seconds={processing_delay_seconds!r}\",
            }}
        ],
    }}


FormsWorker.submit_with_retries = slow_submit

empty_observations = 0
deadline = time.monotonic() + 60
while time.monotonic() < deadline:
    processed = process_forms_submission_queue_once(max_items=1)
    if processed:
        empty_observations = 0
        continue

    with SessionLocal() as db:
        backlog_count = db.execute(
            select(func.count())
            .select_from(FormsSubmission)
            .where(FormsSubmission.status.in_((\"pending\", \"processing\")))
        ).scalar_one()

    if backlog_count == 0:
        empty_observations += 1
        if empty_observations >= 5:
            break
    else:
        empty_observations = 0
    time.sleep(0.05)
"""
    return subprocess.Popen(
        [sys.executable, "-c", worker_script],
        cwd=str(repo_root),
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
    )


def get_user_by_rfid(db, rfid: str) -> User:
    user = find_user_by_rfid(db, rfid)
    assert user is not None
    return user


def get_user_by_chave(db, chave: str) -> User:
    user = find_user_by_chave(db, chave)
    assert user is not None
    return user


def get_materialized_user_project_names(db, user_id: int) -> list[str]:
    return db.execute(
        select(Project.name)
        .join(UserProjectMembership, UserProjectMembership.project_id == Project.id)
        .where(UserProjectMembership.user_id == user_id)
        .order_by(Project.name, Project.id, UserProjectMembership.id)
    ).scalars().all()


def grant_user_project_memberships(db, user: User, project_names: list[str]) -> None:
    normalized_project_names = sorted(
        {str(name).strip().upper() for name in project_names if str(name).strip()}
    )
    existing_project_names = set(
        db.execute(select(Project.name).where(Project.name.in_(normalized_project_names))).scalars().all()
    )
    for project_name in normalized_project_names:
        if project_name in existing_project_names:
            continue
        db.add(Project(name=project_name, **build_project_fields_for_country()))
    db.flush()

    ensure_user_active_project_is_member(db, user)
    for project_name in normalized_project_names:
        add_user_project_membership(db, user, project_name)


def add_transport_schedule(
    db,
    *,
    vehicle: Vehicle,
    service_scope: str,
    route_kind: str,
    recurrence_kind: str,
    service_date: date | None = None,
    weekday: int | None = None,
    departure_time: str | None = None,
) -> TransportVehicleSchedule:
    timestamp = now_sgt()
    schedule = TransportVehicleSchedule(
        vehicle_id=vehicle.id,
        service_scope=service_scope,
        route_kind=route_kind,
        recurrence_kind=recurrence_kind,
        service_date=service_date,
        weekday=weekday,
        departure_time=departure_time,
        is_active=True,
        created_at=timestamp,
        updated_at=timestamp,
    )
    db.add(schedule)
    db.flush()
    return schedule


def make_test_project_name(prefix: str = "TP") -> str:
    normalized_prefix = "".join(character for character in str(prefix or "TP").upper() if character.isalnum())[:4]
    if len(normalized_prefix) < 2:
        normalized_prefix = f"{normalized_prefix}P"
    return f"{normalized_prefix}{uuid.uuid4().hex[:4].upper()}"


def create_transport_planning_project(
    db,
    *,
    name: str,
    country_code: str = "SG",
    address: str,
    zip_code: str,
) -> Project:
    project = Project(
        name=str(name).strip().upper(),
        address=str(address).strip(),
        zip_code=str(zip_code).strip(),
        **build_project_fields_for_country(country_code),
    )
    db.add(project)
    db.flush()
    return project


def create_transport_planning_user_with_request(
    db,
    *,
    chave: str,
    nome: str,
    projeto: str,
    request_kind: str,
    requested_time: str,
    home_address: str,
    home_zip: str,
    service_date: date,
    workplace: str | None = None,
    timestamp: datetime | None = None,
) -> tuple[User, TransportRequest]:
    effective_timestamp = timestamp or now_sgt()
    user = User(
        rfid=None,
        nome=nome,
        chave=chave,
        projeto=str(projeto).strip().upper(),
        workplace=workplace,
        end_rua=str(home_address).strip(),
        zip=str(home_zip).strip(),
        local=None,
        checkin=None,
        time=None,
        last_active_at=effective_timestamp,
        inactivity_days=0,
    )
    db.add(user)
    db.flush()

    if request_kind == "regular":
        recurrence_kind = "weekday"
        selected_weekdays_json = json.dumps([0, 1, 2, 3, 4])
        single_date = None
    elif request_kind == "weekend":
        recurrence_kind = "weekend"
        selected_weekdays_json = None
        single_date = None
    elif request_kind == "extra":
        recurrence_kind = "single_date"
        selected_weekdays_json = None
        single_date = service_date
    else:
        raise ValueError(f"Unsupported request kind for planning fixture: {request_kind}")

    transport_request = TransportRequest(
        user_id=user.id,
        request_kind=request_kind,
        recurrence_kind=recurrence_kind,
        requested_time=requested_time,
        selected_weekdays_json=selected_weekdays_json,
        single_date=single_date,
        created_via="test",
        status="active",
        created_at=effective_timestamp,
        updated_at=effective_timestamp,
        cancelled_at=None,
    )
    db.add(transport_request)
    db.flush()
    return user, transport_request


def create_transport_planning_vehicle_with_schedules(
    db,
    *,
    plate: str,
    service_scope: str,
    service_date: date,
    vehicle_type: str | None = None,
    color: str = "White",
    seats: int | None = None,
    tolerance: int | None = None,
    weekday: int | None = None,
    route_kinds: list[str] | None = None,
    departure_time: str | None = None,
) -> tuple[Vehicle, list[TransportVehicleSchedule]]:
    resolved_vehicle_type = vehicle_type or ("carro" if service_scope == "extra" else "van")
    resolved_seats = seats if seats is not None else (4 if service_scope == "extra" else 10)
    resolved_tolerance = tolerance if tolerance is not None else 5
    resolved_route_kinds = list(route_kinds or (["home_to_work"] if service_scope == "extra" else ["home_to_work", "work_to_home"]))

    vehicle = Vehicle(
        placa=plate,
        tipo=resolved_vehicle_type,
        color=color,
        lugares=resolved_seats,
        tolerance=resolved_tolerance,
        service_scope=service_scope,
    )
    db.add(vehicle)
    db.flush()

    schedules: list[TransportVehicleSchedule] = []
    for route_kind in resolved_route_kinds:
        if service_scope == "regular":
            recurrence_kind = "weekday"
            schedule_service_date = None
            schedule_weekday = None
            schedule_departure_time = None
        elif service_scope == "weekend":
            recurrence_kind = "matching_weekday"
            schedule_service_date = None
            schedule_weekday = weekday if weekday is not None else 5
            schedule_departure_time = None
        elif service_scope == "extra":
            recurrence_kind = "single_date"
            schedule_service_date = service_date
            schedule_weekday = None
            schedule_departure_time = departure_time or "07:00"
        else:
            raise ValueError(f"Unsupported service scope for planning fixture: {service_scope}")

        schedules.append(
            add_transport_schedule(
                db,
                vehicle=vehicle,
                service_scope=service_scope,
                route_kind=route_kind,
                recurrence_kind=recurrence_kind,
                service_date=schedule_service_date,
                weekday=schedule_weekday,
                departure_time=schedule_departure_time,
            )
        )

    return vehicle, schedules


def clone_transport_settings_payload(db) -> dict[str, object]:
    return json.loads(json.dumps(location_settings_module.get_transport_settings_payload(db)))


def configure_transport_planning_settings(
    db,
    *,
    arrive_at_work_time: str,
    work_to_home_time: str,
    last_update_time: str,
    default_car_seats: int,
    default_minivan_seats: int,
    default_van_seats: int,
    default_bus_seats: int,
    default_tolerance_minutes: int,
    price_currency_code: str | None,
    price_currency_label: str | None,
    price_rate_unit: str,
    default_car_price: float | None,
    default_minivan_price: float | None,
    default_van_price: float | None,
    default_bus_price: float | None,
) -> dict[str, object]:
    previous_settings = clone_transport_settings_payload(db)
    created_currency_code = None
    normalized_currency_code = str(price_currency_code or "").strip().upper() or None

    if normalized_currency_code is not None:
        existing_currency = db.execute(
            select(TransportCurrencyOption).where(TransportCurrencyOption.code == normalized_currency_code)
        ).scalar_one_or_none()
        if existing_currency is None:
            location_settings_module.create_transport_currency_option(
                db,
                code=normalized_currency_code,
                display_label=price_currency_label,
            )
            created_currency_code = normalized_currency_code

    location_settings_module.upsert_transport_arrive_at_work_time(
        db,
        arrive_at_work_time=arrive_at_work_time,
    )
    location_settings_module.upsert_transport_work_to_home_time(
        db,
        work_to_home_time=work_to_home_time,
    )
    location_settings_module.upsert_transport_last_update_time(
        db,
        last_update_time=last_update_time,
    )
    location_settings_module.upsert_transport_vehicle_default_seat_counts(
        db,
        default_car_seats=default_car_seats,
        default_minivan_seats=default_minivan_seats,
        default_van_seats=default_van_seats,
        default_bus_seats=default_bus_seats,
        default_tolerance_minutes=default_tolerance_minutes,
    )
    location_settings_module.upsert_transport_pricing_settings(
        db,
        price_currency_code=normalized_currency_code,
        price_rate_unit=price_rate_unit,
        default_car_price=default_car_price,
        default_minivan_price=default_minivan_price,
        default_van_price=default_van_price,
        default_bus_price=default_bus_price,
    )

    return {
        "previous": previous_settings,
        "current": clone_transport_settings_payload(db),
        "created_currency_code": created_currency_code,
    }


def restore_transport_planning_settings(db, settings_payload: dict[str, object] | None) -> None:
    payload = dict(settings_payload or {})
    normalized_currency_code = str(payload.get("price_currency_code") or "").strip().upper() or None
    available_currencies = payload.get("available_currencies") or []
    currency_row = next(
        (row for row in available_currencies if row.get("code") == normalized_currency_code),
        None,
    )

    if normalized_currency_code is not None:
        existing_currency = db.execute(
            select(TransportCurrencyOption).where(TransportCurrencyOption.code == normalized_currency_code)
        ).scalar_one_or_none()
        if existing_currency is None:
            location_settings_module.create_transport_currency_option(
                db,
                code=normalized_currency_code,
                display_label=currency_row.get("display_label") if currency_row else None,
            )

    location_settings_module.upsert_transport_arrive_at_work_time(
        db,
        arrive_at_work_time=str(payload.get("arrive_at_work_time") or "07:45"),
    )
    location_settings_module.upsert_transport_work_to_home_time(
        db,
        work_to_home_time=str(payload.get("work_to_home_time") or "16:45"),
    )
    location_settings_module.upsert_transport_last_update_time(
        db,
        last_update_time=str(payload.get("last_update_time") or "16:00"),
    )
    location_settings_module.upsert_transport_vehicle_default_seat_counts(
        db,
        default_car_seats=int(payload.get("default_car_seats") or 3),
        default_minivan_seats=int(payload.get("default_minivan_seats") or 6),
        default_van_seats=int(payload.get("default_van_seats") or 10),
        default_bus_seats=int(payload.get("default_bus_seats") or 40),
        default_tolerance_minutes=int(payload.get("default_tolerance_minutes") or 5),
    )
    location_settings_module.upsert_transport_pricing_settings(
        db,
        price_currency_code=normalized_currency_code,
        price_rate_unit=str(payload.get("price_rate_unit") or "day"),
        default_car_price=payload.get("default_car_price"),
        default_minivan_price=payload.get("default_minivan_price"),
        default_van_price=payload.get("default_van_price"),
        default_bus_price=payload.get("default_bus_price"),
    )


def create_transport_planning_fixture_bundle(
    db,
    *,
    service_date: date,
) -> dict[str, object]:
    scenario_token = uuid.uuid4().hex[:4].upper()
    weekend_service_date = service_date + timedelta(days=(5 - service_date.weekday()) % 7)
    timestamp = now_sgt()

    regular_project = create_transport_planning_project(
        db,
        name=make_test_project_name(f"RG{scenario_token}"),
        country_code="SG",
        address=f"{scenario_token} Regular Plant Road",
        zip_code=f"10{scenario_token}",
    )
    weekend_project = create_transport_planning_project(
        db,
        name=make_test_project_name(f"WK{scenario_token}"),
        country_code="MY",
        address=f"{scenario_token} Weekend Terminal Avenue",
        zip_code=f"50{scenario_token}",
    )
    extra_project = create_transport_planning_project(
        db,
        name=make_test_project_name(f"EX{scenario_token}"),
        country_code="BR",
        address=f"{scenario_token} Extra Logistics Boulevard",
        zip_code=f"01{scenario_token}",
    )

    settings_context = configure_transport_planning_settings(
        db,
        arrive_at_work_time="07:35",
        work_to_home_time="17:25",
        last_update_time="16:35",
        default_car_seats=4,
        default_minivan_seats=7,
        default_van_seats=11,
        default_bus_seats=45,
        default_tolerance_minutes=8,
        price_currency_code=f"TP{scenario_token}",
        price_currency_label=f"Transport Planning {scenario_token}",
        price_rate_unit="week",
        default_car_price=120.5,
        default_minivan_price=155.75,
        default_van_price=240.0,
        default_bus_price=515.25,
    )

    regular_vehicle, regular_schedules = create_transport_planning_vehicle_with_schedules(
        db,
        plate=f"RG{scenario_token}01",
        service_scope="regular",
        service_date=service_date,
        vehicle_type="van",
        color="Silver",
        seats=11,
        tolerance=8,
    )
    weekend_vehicle, weekend_schedules = create_transport_planning_vehicle_with_schedules(
        db,
        plate=f"WK{scenario_token}01",
        service_scope="weekend",
        service_date=service_date,
        vehicle_type="minivan",
        color="Black",
        seats=7,
        tolerance=10,
        weekday=weekend_service_date.weekday(),
    )
    extra_vehicle, extra_schedules = create_transport_planning_vehicle_with_schedules(
        db,
        plate=f"EX{scenario_token}01",
        service_scope="extra",
        service_date=service_date,
        vehicle_type="carro",
        color="Red",
        seats=4,
        tolerance=6,
        departure_time="07:05",
    )

    regular_user, regular_request = create_transport_planning_user_with_request(
        db,
        chave=make_test_key("R"),
        nome=f"Regular Planning Rider {scenario_token}",
        projeto=regular_project.name,
        request_kind="regular",
        requested_time="07:10",
        home_address=f"{scenario_token} Regular Home Street",
        home_zip=f"90{scenario_token}",
        service_date=service_date,
        timestamp=timestamp,
    )
    weekend_user, weekend_request = create_transport_planning_user_with_request(
        db,
        chave=make_test_key("W"),
        nome=f"Weekend Planning Rider {scenario_token}",
        projeto=weekend_project.name,
        request_kind="weekend",
        requested_time="08:15",
        home_address=f"{scenario_token} Weekend Home Street",
        home_zip=f"91{scenario_token}",
        service_date=service_date,
        timestamp=timestamp,
    )
    extra_user, extra_request = create_transport_planning_user_with_request(
        db,
        chave=make_test_key("E"),
        nome=f"Extra Planning Rider {scenario_token}",
        projeto=extra_project.name,
        request_kind="extra",
        requested_time="06:55",
        home_address=f"{scenario_token} Extra Home Street",
        home_zip=f"92{scenario_token}",
        service_date=service_date,
        timestamp=timestamp,
    )

    return {
        "service_date": service_date.isoformat(),
        "weekend_service_date": weekend_service_date.isoformat(),
        "project_ids": [regular_project.id, weekend_project.id, extra_project.id],
        "user_ids": [regular_user.id, weekend_user.id, extra_user.id],
        "request_ids": [regular_request.id, weekend_request.id, extra_request.id],
        "vehicle_ids": [regular_vehicle.id, weekend_vehicle.id, extra_vehicle.id],
        "schedule_ids": [
            *(schedule.id for schedule in regular_schedules),
            *(schedule.id for schedule in weekend_schedules),
            *(schedule.id for schedule in extra_schedules),
        ],
        "projects": {
            "regular": {
                "id": regular_project.id,
                "name": regular_project.name,
                "country_code": regular_project.country_code,
                "country_name": regular_project.country_name,
                "address": regular_project.address,
                "zip_code": regular_project.zip_code,
            },
            "weekend": {
                "id": weekend_project.id,
                "name": weekend_project.name,
                "country_code": weekend_project.country_code,
                "country_name": weekend_project.country_name,
                "address": weekend_project.address,
                "zip_code": weekend_project.zip_code,
            },
            "extra": {
                "id": extra_project.id,
                "name": extra_project.name,
                "country_code": extra_project.country_code,
                "country_name": extra_project.country_name,
                "address": extra_project.address,
                "zip_code": extra_project.zip_code,
            },
        },
        "requests": {
            "regular": {
                "id": regular_request.id,
                "chave": regular_user.chave,
                "nome": regular_user.nome,
                "projeto": regular_user.projeto,
                "end_rua": regular_user.end_rua,
                "zip": regular_user.zip,
                "service_date": service_date.isoformat(),
            },
            "weekend": {
                "id": weekend_request.id,
                "chave": weekend_user.chave,
                "nome": weekend_user.nome,
                "projeto": weekend_user.projeto,
                "end_rua": weekend_user.end_rua,
                "zip": weekend_user.zip,
                "service_date": weekend_service_date.isoformat(),
            },
            "extra": {
                "id": extra_request.id,
                "chave": extra_user.chave,
                "nome": extra_user.nome,
                "projeto": extra_user.projeto,
                "end_rua": extra_user.end_rua,
                "zip": extra_user.zip,
                "service_date": service_date.isoformat(),
            },
        },
        "vehicles": {
            "regular": {
                "vehicle_id": regular_vehicle.id,
                "placa": regular_vehicle.placa,
                "service_scope": regular_vehicle.service_scope,
            },
            "weekend": {
                "vehicle_id": weekend_vehicle.id,
                "placa": weekend_vehicle.placa,
                "service_scope": weekend_vehicle.service_scope,
            },
            "extra": {
                "vehicle_id": extra_vehicle.id,
                "placa": extra_vehicle.placa,
                "service_scope": extra_vehicle.service_scope,
            },
        },
        "settings_context": settings_context,
    }


def cleanup_transport_planning_fixture_bundle(db, fixture_bundle: dict[str, object] | None) -> None:
    bundle = dict(fixture_bundle or {})
    request_ids = [int(value) for value in bundle.get("request_ids") or []]
    schedule_ids = [int(value) for value in bundle.get("schedule_ids") or []]
    user_ids = [int(value) for value in bundle.get("user_ids") or []]
    vehicle_ids = [int(value) for value in bundle.get("vehicle_ids") or []]
    project_ids = [int(value) for value in bundle.get("project_ids") or []]
    project_names = [project_payload["name"] for project_payload in (bundle.get("projects") or {}).values()]
    settings_context = bundle.get("settings_context") or {}

    restore_transport_planning_settings(db, settings_context.get("previous"))

    if request_ids:
        for assignment in db.execute(
            select(TransportAssignment).where(TransportAssignment.request_id.in_(request_ids))
        ).scalars().all():
            db.delete(assignment)

        for transport_request in db.execute(
            select(TransportRequest).where(TransportRequest.id.in_(request_ids))
        ).scalars().all():
            db.delete(transport_request)

    if schedule_ids:
        for exception_row in db.execute(
            select(TransportVehicleScheduleException).where(
                TransportVehicleScheduleException.vehicle_schedule_id.in_(schedule_ids)
            )
        ).scalars().all():
            db.delete(exception_row)

        for schedule_row in db.execute(
            select(TransportVehicleSchedule).where(TransportVehicleSchedule.id.in_(schedule_ids))
        ).scalars().all():
            db.delete(schedule_row)

    if user_ids:
        for user_row in db.execute(select(User).where(User.id.in_(user_ids))).scalars().all():
            db.delete(user_row)

    if vehicle_ids:
        for vehicle_row in db.execute(select(Vehicle).where(Vehicle.id.in_(vehicle_ids))).scalars().all():
            db.delete(vehicle_row)

    if project_ids:
        for project_row in db.execute(select(Project).where(Project.id.in_(project_ids))).scalars().all():
            db.delete(project_row)

    created_currency_code = settings_context.get("created_currency_code")
    if created_currency_code:
        currency_row = db.execute(
            select(TransportCurrencyOption).where(TransportCurrencyOption.code == created_currency_code)
        ).scalar_one_or_none()
        if currency_row is not None:
            db.delete(currency_row)

    db.commit()


def set_user_checkin_state(*, chave: str, event_time: datetime, local: str = "Web") -> None:
    with SessionLocal() as db:
        user = get_user_by_chave(db, chave)
        user.checkin = True
        user.time = event_time
        user.local = local
        db.commit()


def test_health():
    with TestClient(app) as client:
        res = client.get("/api/health")
        assert res.status_code == 200
        payload = res.json()
        assert payload["status"] == "ok"
        assert payload["ready"] is True
        assert payload["overall_status"] == "ok"
        assert payload["components"]["database"]["status"] == "ok"
        assert payload["components"]["static_sites"]["status"] == "ok"
        assert payload["components"]["transport_ai_operational_readiness"]["status"] in {"ok", "disabled"}
        assert payload["components"]["transport_ai_settings_encryption"]["status"] in {"ok", "disabled"}
        assert payload["components"]["forms_worker"]["status"] == "disabled"


def test_health_live():
    with TestClient(app) as client:
        res = client.get("/api/health/live")
        assert res.status_code == 200
        assert res.json() == {"status": "ok", "app": settings.app_name}


def test_health_ready():
    with TestClient(app) as client:
        res = client.get("/api/health/ready")
        assert res.status_code == 200
        payload = res.json()
        assert payload["status"] == "ok"
        assert payload["ready"] is True
        assert payload["overall_status"] == "ok"
        assert payload["components"]["database"]["status"] == "ok"
        assert payload["components"]["static_sites"]["status"] == "ok"
        assert payload["components"]["transport_ai_operational_readiness"]["status"] in {"ok", "disabled"}
        assert payload["components"]["transport_ai_settings_encryption"]["status"] in {"ok", "disabled"}
        assert "forms_worker" not in payload["components"]


def test_health_ready_returns_503_when_database_is_unavailable(monkeypatch):
    monkeypatch.setattr(
        "sistema.app.routers.health._build_database_component",
        lambda: HealthComponentResponse(status="failed", detail="database unavailable"),
    )

    with TestClient(app) as client:
        res = client.get("/api/health/ready")

    assert res.status_code == 503
    payload = res.json()
    assert payload["status"] == "unready"
    assert payload["ready"] is False
    assert payload["overall_status"] == "unready"
    assert payload["components"]["database"]["status"] == "failed"


def test_health_ready_returns_503_when_transport_ai_settings_encryption_is_unavailable(monkeypatch):
    monkeypatch.setattr(settings, "transport_ai_enabled", True)
    monkeypatch.setattr(settings, "transport_ai_agent_mode", "agent")
    monkeypatch.setattr(settings, "transport_ai_operational_approval_evidence", "phase8-loadtest-2026-05-05")
    monkeypatch.setattr(settings, "transport_ai_max_concurrent_runs", 1)
    monkeypatch.setattr(settings, "transport_ai_max_runtime_seconds", 180)
    monkeypatch.setattr(
        "sistema.app.routers.health.validate_transport_ai_settings_encryption_availability",
        lambda: (_ for _ in ()).throw(
            TransportAILlmSettingsEncryptionError("Transport AI settings encryption key is not configured.")
        ),
    )

    with TestClient(app) as client:
        res = client.get("/api/health/ready")

    assert res.status_code == 503
    payload = res.json()
    assert payload["status"] == "unready"
    assert payload["ready"] is False
    assert payload["overall_status"] == "unready"
    assert payload["components"]["transport_ai_settings_encryption"]["status"] == "failed"
    assert payload["components"]["transport_ai_settings_encryption"]["detail"] == "Transport AI settings encryption key is not configured."


def test_health_ready_returns_503_when_transport_ai_operational_readiness_is_unavailable(monkeypatch):
    monkeypatch.setattr(settings, "transport_ai_enabled", True)
    monkeypatch.setattr(settings, "transport_ai_agent_mode", "deterministic")
    monkeypatch.setattr(settings, "transport_ai_operational_approval_evidence", None)
    monkeypatch.setattr(settings, "transport_ai_max_concurrent_runs", 1)
    monkeypatch.setattr(settings, "transport_ai_max_runtime_seconds", 180)

    with TestClient(app) as client:
        res = client.get("/api/health/ready")

    assert res.status_code == 503
    payload = res.json()
    assert payload["status"] == "unready"
    assert payload["ready"] is False
    assert payload["overall_status"] == "unready"
    assert payload["components"]["transport_ai_operational_readiness"]["status"] == "failed"
    assert "transport_ai_operational_approval_missing" in payload["components"]["transport_ai_operational_readiness"]["detail"]


def test_health_surfaces_forms_worker_degradation_without_failing_api(monkeypatch):
    monkeypatch.setattr(
        "sistema.app.routers.health._build_forms_worker_component",
        lambda: HealthComponentResponse(status="degraded", detail="forms worker heartbeat stale"),
    )

    with TestClient(app) as client:
        res = client.get("/api/health")

    assert res.status_code == 200
    payload = res.json()
    assert payload["status"] == "ok"
    assert payload["ready"] is True
    assert payload["overall_status"] == "degraded"
    assert payload["components"]["forms_worker"]["status"] == "degraded"


def test_http_request_logging_middleware_emits_structured_fields_and_request_id(caplog: pytest.LogCaptureFixture):
    with caplog.at_level(logging.INFO, logger="checking.http"):
        with TestClient(app) as client:
            caplog.clear()
            health_response = client.get("/api/health")
            assert health_response.status_code == 200
            health_request_id = health_response.headers.get("X-Request-ID")
            assert health_request_id

            health_logs = get_http_request_logs(caplog)
            assert health_logs
            health_log = health_logs[-1]
            assert health_log == {
                "authenticated_kind": None,
                "client_surface": "health",
                "event": "http_request",
                "is_critical_route": True,
                "latency_ms": health_log["latency_ms"],
                "method": "GET",
                "path": "/api/health",
                "request_id": health_request_id,
                "status_code": 200,
            }
            assert isinstance(health_log["latency_ms"], int)
            assert health_log["latency_ms"] >= 0

            caplog.clear()
            login_response = login_admin(client)
            assert login_response.status_code == 200, login_response.text
            session_response = client.get("/api/admin/auth/session")
            assert session_response.status_code == 200
            session_request_id = session_response.headers.get("X-Request-ID")
            assert session_request_id

            session_logs = get_http_request_logs(caplog)
            assert session_logs
            session_log = next(payload for payload in reversed(session_logs) if payload["path"] == "/api/admin/auth/session")
            assert session_log["method"] == "GET"
            assert session_log["path"] == "/api/admin/auth/session"
            assert session_log["status_code"] == 200
            assert session_log["client_surface"] == "admin"
            assert session_log["authenticated_kind"] == "admin_session"
            assert session_log["is_critical_route"] is False
            assert session_log["request_id"] == session_request_id
            assert isinstance(session_log["latency_ms"], int)
            assert session_log["latency_ms"] >= 0


def test_vehicle_schema_and_user_transport_fields_persist_expected_values():
    with SessionLocal() as db:
        db.add(Workplace(workplace="Innovation Hub", address="1 Harbour Front", zip="098632", country="Singapore"))
        vehicle = Vehicle(placa="SGX1234A", tipo="van", color="White", lugares=18, tolerance=12, service_scope="regular")
        db.add(vehicle)
        db.flush()

        user = User(
            rfid=None,
            nome="Usuario Transporte",
            chave="TR01",
            projeto="P80",
            workplace="Innovation Hub",
            vehicle_id=vehicle.id,
            placa="SGX1234A",
            end_rua="123 Harbour Road",
            zip="0012345678",
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
        )
        db.add(user)
        db.commit()
        db.refresh(user)

        persisted_vehicle = db.execute(select(Vehicle).where(Vehicle.placa == "SGX1234A")).scalar_one()
        persisted_user = db.execute(select(User).where(User.chave == "TR01")).scalar_one()

        assert persisted_vehicle.tipo == "van"
        assert persisted_vehicle.color == "White"
        assert persisted_vehicle.lugares == 18
        assert persisted_vehicle.tolerance == 12
        assert persisted_vehicle.service_scope == "regular"
        assert persisted_user.workplace == "Innovation Hub"
        assert persisted_user.vehicle_id == persisted_vehicle.id
        assert persisted_user.placa == "SGX1234A"
        assert persisted_user.end_rua == "123 Harbour Road"
        assert persisted_user.zip == "0012345678"


def test_mobile_sync_records_checkinghistory_entry():
    event_time = now_sgt().replace(microsecond=0)

    with TestClient(app) as client:
        response = client.post(
            "/api/mobile/events/sync",
            headers=MOBILE_HEADERS,
            json={
                "chave": "HC01",
                "projeto": "P82",
                "action": "checkin",
                "event_time": event_time.isoformat(),
                "client_event_id": f"android-{uuid.uuid4().hex}",
            },
        )
        assert response.status_code == 200

    with SessionLocal() as db:
        row = db.execute(
            select(CheckingHistory)
            .where(CheckingHistory.chave == "HC01")
            .order_by(CheckingHistory.id.desc())
            .limit(1)
        ).scalar_one()

    assert row.atividade == "check-in"
    assert row.projeto == "P82"
    assert row.informe == "normal"
    assert row.time.replace(tzinfo=ZoneInfo(settings.tz_name)) == event_time


def test_admin_users_support_extended_profile_fields_and_key_updates_history():
    with SessionLocal() as db:
        db.add(Workplace(workplace="Refinery West", address="2 Harbour Road", zip="112233", country="Singapore"))
        db.add(Vehicle(placa="TRP1234A", tipo="van", color="Blue", lugares=18, tolerance=10, service_scope="regular"))
        db.commit()

    with TestClient(app) as client:
        ensure_admin_session(client)
        created = client.post(
            "/api/admin/users",
            json={
                "rfid": "USR9001",
                "nome": "Usuario Completo",
                "chave": "UF01",
                "projeto": "P82",
                "workplace": "Refinery West",
                "placa": "TRP1234A",
                "end_rua": "123 Harbour Road",
                "zip": "0012345678",
                "email": "USER@EXAMPLE.COM",
            },
        )
        assert created.status_code == 200

        sync_res = client.post(
            "/api/mobile/events/sync",
            headers=MOBILE_HEADERS,
            json={
                "chave": "UF01",
                "projeto": "P82",
                "action": "checkin",
                "event_time": now_sgt().isoformat(),
                "client_event_id": f"android-{uuid.uuid4().hex}",
            },
        )
        assert sync_res.status_code == 200

        users = client.get("/api/admin/users")
        assert users.status_code == 200
        user_row = next(row for row in users.json() if row["chave"] == "UF01")
        assert user_row["vehicle_id"] is not None
        assert user_row["workplace"] == "Refinery West"
        assert user_row["placa"] == "TRP1234A"
        assert user_row["end_rua"] == "123 Harbour Road"
        assert user_row["zip"] == "0012345678"
        assert user_row["email"] == "user@example.com"
        assert user_row["projetos"] == ["P82"]
        assert user_row["projeto_ativo"] == "P82"

        updated = client.post(
            "/api/admin/users",
            json={
                "user_id": user_row["id"],
                "rfid": "USR9002",
                "nome": "Usuario Ajustado",
                "chave": "UF02",
                "projeto": "P83",
                "workplace": "Refinery West",
                "vehicle_id": user_row["vehicle_id"],
                "end_rua": "456 Harbour Road",
                "zip": "0000001234",
                "email": "final@example.com",
            },
        )
        assert updated.status_code == 200

        users_after = client.get("/api/admin/users")
        assert users_after.status_code == 200
        updated_row = next(row for row in users_after.json() if row["id"] == user_row["id"])
        assert updated_row["rfid"] == "USR9002"
        assert updated_row["nome"] == "Usuario Ajustado"
        assert updated_row["chave"] == "UF02"
        assert updated_row["projeto"] == "P83"
        assert updated_row["vehicle_id"] == user_row["vehicle_id"]
        assert updated_row["workplace"] == "Refinery West"
        assert updated_row["placa"] == "TRP1234A"
        assert updated_row["end_rua"] == "456 Harbour Road"
        assert updated_row["zip"] == "0000001234"
        assert updated_row["email"] == "final@example.com"
        assert updated_row["projetos"] == ["P83"]
        assert updated_row["projeto_ativo"] == "P83"

    with SessionLocal() as db:
        sync_rows = db.execute(
            select(UserSyncEvent)
            .where(UserSyncEvent.user_id == user_row["id"])
            .order_by(UserSyncEvent.id)
        ).scalars().all()
        history_rows = db.execute(
            select(CheckingHistory)
            .where(CheckingHistory.chave == "UF02")
            .order_by(CheckingHistory.id)
        ).scalars().all()

    assert sync_rows
    assert all(row.chave == "UF02" for row in sync_rows)
    assert history_rows
    assert all(row.chave == "UF02" for row in history_rows)


def test_admin_projects_catalog_lists_creates_and_blocks_linked_user_deletion():
    with TestClient(app) as client:
        ensure_admin_session(client)

        listed = client.get("/api/admin/projects")
        assert listed.status_code == 200
        listed_payload = listed.json()
        listed_names = {row["name"] for row in listed_payload}
        assert {"P80", "P82", "P83"}.issubset(listed_names)
        listed_p80 = next(row for row in listed_payload if row["name"] == "P80")
        assert listed_p80["country_code"] == "SG"
        assert listed_p80["country_name"] == "Singapura"
        assert listed_p80["timezone_name"] == "Asia/Singapore"
        assert listed_p80["timezone_label"] == "Singapura (+8)"
        assert listed_p80["address"] == ""
        assert listed_p80["zip_code"] == ""

        created = client.post(
            "/api/admin/projects",
            json={
                "name": "P90",
                "country_name": "China",
                "timezone_name": "Asia/Shanghai",
                "address": "1 Jurong West Central 2",
                "zip_code": "648886",
            },
        )
        assert created.status_code == 200, created.text
        project_payload = created.json()
        assert project_payload["name"] == "P90"
        assert project_payload["country_code"] == "CN"
        assert project_payload["country_name"] == "China"
        assert project_payload["timezone_name"] == "Asia/Shanghai"
        assert project_payload["timezone_label"] == "China (+8)"
        assert project_payload["address"] == "1 Jurong West Central 2"
        assert project_payload["zip_code"] == "648886"

        with SessionLocal() as db:
            db.add(
                User(
                    rfid=None,
                    nome="Usuario Vinculado",
                    chave="PJ90",
                    projeto="P90",
                    perfil=0,
                    local=None,
                    checkin=None,
                    time=None,
                    last_active_at=now_sgt(),
                    inactivity_days=0,
                )
            )
            db.commit()

        blocked = client.delete(f"/api/admin/projects/{project_payload['id']}")
        assert blocked.status_code == 409
        assert blocked.json()["detail"] == "Nao e possivel remover um projeto com usuarios vinculados."

        with SessionLocal() as db:
            linked_user = get_user_by_chave(db, "PJ90")
            db.delete(linked_user)
            db.commit()

        removed = client.delete(f"/api/admin/projects/{project_payload['id']}")
        assert removed.status_code == 200
        assert removed.json()["ok"] is True

        listed_after = client.get("/api/admin/projects")
        assert listed_after.status_code == 200
        listed_after_names = {row["name"] for row in listed_after.json()}
        assert "P90" not in listed_after_names


def test_admin_project_update_changes_country_metadata_without_renaming():
    with TestClient(app) as client:
        ensure_admin_session(client)
        created = client.post(
            "/api/admin/projects",
            json={
                "name": "P96",
                "country_name": "Singapura",
                "timezone_name": "Asia/Singapore",
                "address": "10 Bayfront Avenue",
                "zip_code": "018956",
            },
        )
        assert created.status_code == 200, created.text
        project_payload = created.json()

        updated = client.put(
            f"/api/admin/projects/{project_payload['id']}",
            json={
                "name": "P96",
                "country_name": "Brasil",
                "timezone_name": "America/Sao_Paulo",
                "address": "Avenida Paulista 1000",
                "zip_code": "01310-100",
            },
        )
        assert updated.status_code == 200, updated.text
        updated_payload = updated.json()
        assert updated_payload["id"] == project_payload["id"]
        assert updated_payload["name"] == "P96"
        assert updated_payload["country_code"] == "BR"
        assert updated_payload["country_name"] == "Brasil"
        assert updated_payload["timezone_name"] == "America/Sao_Paulo"
        assert updated_payload["timezone_label"] == "Brasil (-3)"
        assert updated_payload["address"] == "Avenida Paulista 1000"
        assert updated_payload["zip_code"] == "01310-100"

    with SessionLocal() as db:
        project = db.execute(select(Project).where(Project.name == "P96")).scalar_one()

    assert project.country_code == "BR"
    assert project.country_name == "Brasil"
    assert project.timezone_name == "America/Sao_Paulo"
    assert project.address == "Avenida Paulista 1000"
    assert project.zip_code == "01310-100"


def test_admin_project_update_renames_live_links_and_exposes_transport_project_payload():
    source_project_name = f"Q{uuid.uuid4().hex[:3].upper()}"
    target_project_name = f"R{uuid.uuid4().hex[:3].upper()}"
    linked_user_key = make_test_key("U")
    scoped_admin_key = make_test_key("S")
    location_name = f"Projeto {source_project_name} Base"
    dashboard_service_date = date(2026, 5, 11)

    with TestClient(app) as client:
        ensure_admin_session(client)
        created = client.post(
            "/api/admin/projects",
            json={
                "name": source_project_name,
                "country_name": "Singapura",
                "timezone_name": "Asia/Singapore",
                "address": "Old Address 98",
                "zip_code": "100098",
                "minimum_checkout_distance_meters": 4321,
            },
        )
        assert created.status_code == 200, created.text
        project_payload = created.json()

        with SessionLocal() as db:
            admin = get_user_by_chave(db, ADMIN_LOGIN_CHAVE)
            grant_user_project_memberships(
                db,
                admin,
                db.execute(select(Project.name).order_by(Project.name)).scalars().all(),
            )
            db.commit()

        with SessionLocal() as db:
            linked_user = User(
                rfid=None,
                nome="Usuario Projeto Renomeado",
                chave=linked_user_key,
                projeto=source_project_name,
                perfil=0,
                local=None,
                checkin=None,
                time=None,
                last_active_at=now_sgt(),
                inactivity_days=0,
            )
            db.add_all(
                [
                    linked_user,
                    User(
                        rfid=None,
                        nome="Admin Escopo Projeto Renomeado",
                        chave=scoped_admin_key,
                        projeto="P80",
                        perfil=1,
                        local=None,
                        checkin=None,
                        time=None,
                        last_active_at=now_sgt(),
                        inactivity_days=0,
                        admin_monitored_projects_json=dump_admin_monitored_projects(["P80", source_project_name]),
                    ),
                ]            )
            db.flush()
            ensure_user_active_project_is_member(db, linked_user)
            add_user_project_membership(db, linked_user, "P80")
            db.add(
                TransportRequest(
                    user_id=linked_user.id,
                    request_kind="extra",
                    recurrence_kind="single_date",
                    requested_time="18:15",
                    single_date=dashboard_service_date,
                    created_via="bot",
                    status="active",
                    created_at=now_sgt(),
                    updated_at=now_sgt(),
                    cancelled_at=None,
                )
            )
            db.commit()

        create_location = client.post(
            "/api/admin/locations",
            json={
                "local": location_name,
                "coordinates": build_rectangle_coordinates(1.255936, 103.611066),
                "projects": ["P80", source_project_name],
                "tolerance_meters": 150,
            },
        )
        assert create_location.status_code == 200, create_location.text

        updated = client.put(
            f"/api/admin/projects/{project_payload['id']}",
            json={
                "name": target_project_name,
                "country_name": "Malásia",
                "timezone_name": "Asia/Kuala_Lumpur",
                "address": "New Address 99",
                "zip_code": "500099",
            },
        )
        assert updated.status_code == 200, updated.text
        updated_payload = updated.json()
        assert updated_payload["id"] == project_payload["id"]
        assert updated_payload["name"] == target_project_name
        assert updated_payload["country_code"] == "MY"
        assert updated_payload["country_name"] == "Malásia"
        assert updated_payload["timezone_name"] == "Asia/Kuala_Lumpur"
        assert updated_payload["address"] == "New Address 99"
        assert updated_payload["zip_code"] == "500099"

        transport_projects = client.get("/api/transport/projects")
        assert transport_projects.status_code == 200, transport_projects.text
        transport_project = next(row for row in transport_projects.json() if row["id"] == project_payload["id"])
        assert transport_project["name"] == target_project_name
        assert transport_project["country_name"] == "Malásia"
        assert transport_project["address"] == "New Address 99"
        assert transport_project["zip_code"] == "500099"

        locations_response = client.get("/api/admin/locations")
        assert locations_response.status_code == 200, locations_response.text
        updated_location = next(row for row in locations_response.json()["items"] if row["local"] == location_name)
        assert source_project_name not in updated_location["projects"]
        assert set(updated_location["projects"]) == {"P80", target_project_name}

        dashboard_response = client.get(
            "/api/transport/dashboard",
            params={"service_date": dashboard_service_date.isoformat(), "route_kind": "home_to_work"},
        )
        assert dashboard_response.status_code == 200, dashboard_response.text
        dashboard_payload = dashboard_response.json()
        matching_rows = [
            row for row in dashboard_payload["extra_requests"] if row["chave"] == linked_user_key
        ]
        assert len(matching_rows) == 1
        assert matching_rows[0]["projeto"] == target_project_name
        assert matching_rows[0]["projects"] == ["P80", target_project_name]

    with SessionLocal() as db:
        renamed_project = db.get(Project, project_payload["id"])
        legacy_project = db.execute(select(Project).where(Project.name == source_project_name)).scalar_one_or_none()
        renamed_user = find_user_by_chave(db, linked_user_key)
        scoped_admin = find_user_by_chave(db, scoped_admin_key)

    assert renamed_project is not None
    assert renamed_project.name == target_project_name
    assert renamed_project.minimum_checkout_distance_meters == 4321
    assert renamed_project.address == "New Address 99"
    assert renamed_project.zip_code == "500099"
    assert legacy_project is None
    assert renamed_user.projeto == target_project_name
    assert list_user_project_names(db, renamed_user) == ["P80", target_project_name]
    assert scoped_admin.admin_monitored_projects_json is None


def test_admin_project_create_accepts_custom_country_and_timezone_pair():
    with TestClient(app) as client:
        ensure_admin_session(client)

        created = client.post(
            "/api/admin/projects",
            json={"name": "P97", "country_name": "Chile", "timezone_name": "America/Santiago"},
        )

    assert created.status_code == 200, created.text
    payload = created.json()
    assert payload["name"] == "P97"
    assert payload["country_code"] == "CH"
    assert payload["country_name"] == "Chile"
    assert payload["timezone_name"] == "America/Santiago"
    assert payload["timezone_label"] == "Chile (-4)"


def test_admin_project_delete_reassigns_admin_only_users_to_fallback_project():
    with TestClient(app) as client:
        ensure_admin_session(client)
        created = client.post("/api/admin/projects", json={"name": "P91"})
        assert created.status_code == 200, created.text
        project_payload = created.json()

        with SessionLocal() as db:
            db.add(
                User(
                    rfid=None,
                    nome="Admin Vinculado",
                    chave="PA91",
                    projeto="P91",
                    perfil=1,
                    local=None,
                    checkin=None,
                    time=None,
                    last_active_at=now_sgt(),
                    inactivity_days=0,
                )
            )
            db.commit()

        removed = client.delete(f"/api/admin/projects/{project_payload['id']}")
        assert removed.status_code == 200, removed.text
        assert removed.json()["ok"] is True

    with SessionLocal() as db:
        admin_user = get_user_by_chave(db, "PA91")
        removed_project = db.execute(select(Project).where(Project.name == "P91")).scalar_one_or_none()

    assert removed_project is None
    assert admin_user.projeto in {"P80", "P82", "P83"}
    assert admin_user.projeto != "P91"


def test_admin_plural_user_contract_returns_memberships_and_active_project():
    with TestClient(app) as client:
        ensure_admin_session(client)
        created = client.post(
            "/api/admin/users",
            json={
                "rfid": "USRPLURAL1",
                "nome": "Usuario Multi Projeto",
                "chave": "UP11",
                "projetos": ["P83", "P82"],
            },
        )
        assert created.status_code == 200, created.text

        users = client.get("/api/admin/users")
        assert users.status_code == 200, users.text
        user_row = next(row for row in users.json() if row["chave"] == "UP11")
        assert user_row["projetos"] == ["P82", "P83"]
        assert user_row["projeto"] == "P82"
        assert user_row["projeto_ativo"] == "P82"

        updated = client.post(
            "/api/admin/users",
            json={
                "user_id": user_row["id"],
                "rfid": "USRPLURAL1",
                "nome": "Usuario Multi Projeto Atualizado",
                "chave": "UP11",
                "projetos": ["P80", "P83"],
            },
        )
        assert updated.status_code == 200, updated.text

        users_after = client.get("/api/admin/users")
        assert users_after.status_code == 200, users_after.text
        updated_row = next(row for row in users_after.json() if row["id"] == user_row["id"])
        assert updated_row["projetos"] == ["P80", "P83"]
        assert updated_row["projeto"] == "P80"
        assert updated_row["projeto_ativo"] == "P80"

    with SessionLocal() as db:
        stored_user = get_user_by_chave(db, "UP11")
        assert list_user_project_names(db, stored_user) == ["P80", "P83"]
        assert stored_user.projeto == "P80"


def test_admin_project_delete_preserves_users_with_remaining_memberships():
    project_name = f"Q{uuid.uuid4().hex[:3].upper()}"
    dashboard_service_date = date(2026, 5, 12)

    with TestClient(app) as client:
        ensure_admin_session(client)
        created = client.post("/api/admin/projects", json={"name": project_name})
        assert created.status_code == 200, created.text
        project_payload = created.json()

        with SessionLocal() as db:
            linked_user = User(
                rfid=None,
                nome="Usuario Multi Vinculo",
                chave="PM01",
                projeto=project_name,
                perfil=0,
                local=None,
                checkin=None,
                time=None,
                last_active_at=now_sgt(),
                inactivity_days=0,
            )
            db.add(linked_user)
            db.flush()
            ensure_user_active_project_is_member(db, linked_user)
            add_user_project_membership(db, linked_user, "P82")
            db.add(
                TransportRequest(
                    user_id=linked_user.id,
                    request_kind="extra",
                    recurrence_kind="single_date",
                    requested_time="19:00",
                    single_date=dashboard_service_date,
                    created_via="bot",
                    status="active",
                    created_at=now_sgt(),
                    updated_at=now_sgt(),
                    cancelled_at=None,
                )
            )
            db.commit()

        removed = client.delete(f"/api/admin/projects/{project_payload['id']}")
        assert removed.status_code == 200, removed.text
        assert removed.json()["ok"] is True

        dashboard_response = client.get(
            "/api/transport/dashboard",
            params={"service_date": dashboard_service_date.isoformat(), "route_kind": "home_to_work"},
        )
        assert dashboard_response.status_code == 200, dashboard_response.text
        dashboard_payload = dashboard_response.json()
        matching_rows = [
            row for row in dashboard_payload["extra_requests"] if row["chave"] == "PM01"
        ]
        assert len(matching_rows) == 1
        assert matching_rows[0]["projeto"] == "P82"
        assert matching_rows[0]["projects"] == ["P82"]
        assert all(project_row["name"] != project_name for project_row in dashboard_payload["projects"])

    with SessionLocal() as db:
        linked_user = get_user_by_chave(db, "PM01")
        removed_project = db.execute(select(Project).where(Project.id == project_payload["id"])).scalar_one_or_none()

        assert removed_project is None
        assert linked_user.projeto == "P82"
        assert list_user_project_names(db, linked_user) == ["P82"]


def test_admin_project_delete_rewrites_explicit_admin_scopes_and_resets_empty_scope_to_all_mode():
    project_name = f"Q{uuid.uuid4().hex[:3].upper()}"

    with TestClient(app) as client:
        ensure_admin_session(client)
        created = client.post("/api/admin/projects", json={"name": project_name})
        assert created.status_code == 200, created.text
        project_payload = created.json()

        restricted_key = make_test_key("S")
        emptied_key = make_test_key("S")
        unrestricted_key = make_test_key("S")
        legacy_full_key = make_test_key("S")

        with SessionLocal() as db:
            db.add_all(
                [
                    User(
                        rfid=None,
                        nome="Admin Restrito Projeto Removido",
                        chave=restricted_key,
                        projeto="P80",
                        perfil=1,
                        local=None,
                        checkin=None,
                        time=None,
                        last_active_at=now_sgt(),
                        inactivity_days=0,
                        admin_monitored_projects_json=dump_admin_monitored_projects(["P80", project_name]),
                    ),
                    User(
                        rfid=None,
                        nome="Admin Escopo Esvaziado",
                        chave=emptied_key,
                        projeto="P80",
                        perfil=1,
                        local=None,
                        checkin=None,
                        time=None,
                        last_active_at=now_sgt(),
                        inactivity_days=0,
                        admin_monitored_projects_json=dump_admin_monitored_projects([project_name]),
                    ),
                    User(
                        rfid=None,
                        nome="Admin Todos",
                        chave=unrestricted_key,
                        projeto="P80",
                        perfil=1,
                        local=None,
                        checkin=None,
                        time=None,
                        last_active_at=now_sgt(),
                        inactivity_days=0,
                        admin_monitored_projects_json=None,
                    ),
                    User(
                        rfid=None,
                        nome="Admin Todos Legado",
                        chave=legacy_full_key,
                        projeto="P80",
                        perfil=1,
                        local=None,
                        checkin=None,
                        time=None,
                        last_active_at=now_sgt(),
                        inactivity_days=0,
                        admin_monitored_projects_json=dump_admin_monitored_projects(["P80", "P82", "P83", project_name]),
                    ),
                ]
            )
            db.commit()

        removed = client.delete(f"/api/admin/projects/{project_payload['id']}")
        assert removed.status_code == 200, removed.text
        assert removed.json()["ok"] is True

        administrators_response = client.get("/api/admin/administrators")
        assert administrators_response.status_code == 200, administrators_response.text
        rows_by_key = {row["chave"]: row for row in administrators_response.json()}

    with SessionLocal() as db:
        restricted_admin = get_user_by_chave(db, restricted_key)
        emptied_admin = get_user_by_chave(db, emptied_key)
        unrestricted_admin = get_user_by_chave(db, unrestricted_key)
        legacy_full_admin = get_user_by_chave(db, legacy_full_key)
        removal_event = db.execute(
            select(CheckEvent)
            .where(
                CheckEvent.request_path == f"/api/admin/projects/{project_payload['id']}",
                CheckEvent.message == "Project removed via admin",
            )
            .order_by(CheckEvent.id.desc())
        ).scalars().first()

    resolved_all_projects = rows_by_key[unrestricted_key]["projects"]

    assert restricted_admin.admin_monitored_projects_json is None
    assert emptied_admin.admin_monitored_projects_json is None
    assert unrestricted_admin.admin_monitored_projects_json is None
    assert legacy_full_admin.admin_monitored_projects_json is None

    assert rows_by_key[restricted_key]["projects"] == ["P80"]
    assert rows_by_key[emptied_key]["projects"] == resolved_all_projects
    assert rows_by_key[unrestricted_key]["projects"] == resolved_all_projects
    assert rows_by_key[legacy_full_key]["projects"] == ["P80"]

    assert removal_event is not None
    assert f"updated_admin_scopes=3" in removal_event.details


def test_administrators_endpoint_keeps_all_mode_dynamic_when_new_project_is_created():
    project_name = f"Q{uuid.uuid4().hex[:3].upper()}"
    admin_key = make_test_key("S")

    with SessionLocal() as db:
        db.add(
            User(
                rfid=None,
                nome="Admin Todos Dinamico",
                chave=admin_key,
                projeto="P80",
                perfil=1,
                local=None,
                checkin=None,
                time=None,
                last_active_at=now_sgt(),
                inactivity_days=0,
                admin_monitored_projects_json=None,
            )
        )
        db.commit()

    with TestClient(app) as client:
        ensure_admin_session(client)

        created = client.post("/api/admin/projects", json={"name": project_name})
        assert created.status_code == 200, created.text

        administrators_response = client.get("/api/admin/administrators")
        assert administrators_response.status_code == 200, administrators_response.text
        row = next(row for row in administrators_response.json() if row["chave"] == admin_key)

    assert project_name in row["projects"]
    assert set({"P80", "P82", "P83", project_name}).issubset(set(row["projects"]))
    assert row["projects"] == sorted(row["projects"])


def test_web_projects_endpoint_lists_catalog_and_authenticated_user_projects_update_persists():
    ensure_project_exists("P95")
    ensure_project_exists("P96")
    ensure_web_user_exists(chave="WP95", projeto="P80", nome="Usuario Projeto Web")

    with SessionLocal() as db:
        user = get_user_by_chave(db, "WP95")
        grant_user_project_memberships(db, user, ["P80", "P95"])
        db.commit()

    with TestClient(app) as client:
        listed = client.get("/api/web/projects")
        assert listed.status_code == 200
        listed_payload = listed.json()
        listed_names = {row["name"] for row in listed_payload}
        assert {"P80", "P82", "P83", "P95"}.issubset(listed_names)
        listed_p95 = next(row for row in listed_payload if row["name"] == "P95")
        assert listed_p95["country_code"] == "SG"
        assert listed_p95["country_name"] == "Singapura"
        assert listed_p95["timezone_name"] == "Asia/Singapore"
        assert listed_p95["timezone_label"] == "Singapura (+8)"

        register_response = register_web_password(
            client,
            chave="WP95",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert register_response.status_code == 200, register_response.text

        user_projects = client.get("/api/web/user-projects")
        assert user_projects.status_code == 200, user_projects.text
        assert user_projects.json() == {
            "projects": ["P80", "P95"],
            "active_project": "P80",
        }

        updated = client.put("/api/web/user-projects", json={"projects": ["P96", "P95"]})
        assert updated.status_code == 200, updated.text
        assert updated.json()["ok"] is True
        assert updated.json()["projects"] == ["P95", "P96"]
        assert updated.json()["active_project"] == "P95"

        history_state = client.get("/api/web/check/state", params={"chave": "WP95"})
        assert history_state.status_code == 200, history_state.text
        assert history_state.json()["projeto"] == "P95"

    with SessionLocal() as db:
        user = get_user_by_chave(db, "WP95")

    assert user.projeto == "P95"


def test_web_legacy_project_update_route_switches_only_the_active_project_within_memberships():
    ensure_project_exists("P96")
    ensure_web_user_exists(chave="WP96", projeto="P80", nome="Usuario Projeto Legado Web")

    with SessionLocal() as db:
        user = get_user_by_chave(db, "WP96")
        grant_user_project_memberships(db, user, ["P80", "P96"])
        db.commit()

    with TestClient(app) as client:
        register_response = register_web_password(
            client,
            chave="WP96",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert register_response.status_code == 200, register_response.text

        updated = client.put("/api/web/project", json={"project": "P96"})
        assert updated.status_code == 200, updated.text
        assert updated.json() == {
            "ok": True,
            "message": "Projeto ativo atualizado com sucesso.",
            "project": "P96",
            "projects": ["P80", "P96"],
            "active_project": "P96",
        }

        user_projects = client.get("/api/web/user-projects")
        assert user_projects.status_code == 200, user_projects.text
        assert user_projects.json() == {
            "projects": ["P80", "P96"],
            "active_project": "P96",
        }

    with SessionLocal() as db:
        user = get_user_by_chave(db, "WP96")
        assert user.projeto == "P96"
        assert get_materialized_user_project_names(db, user.id) == ["P80", "P96"]


def test_web_check_submit_rejects_projects_outside_authenticated_user_memberships():
    ensure_project_exists("P96A")
    ensure_web_user_exists(chave="WC96", projeto="P80", nome="Usuario Escopo Check Web")

    with SessionLocal() as db:
        user = get_user_by_chave(db, "WC96")
        grant_user_project_memberships(db, user, ["P80", "P83"])
        db.commit()

    with TestClient(app) as client:
        register_response = register_web_password(
            client,
            chave="WC96",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert register_response.status_code == 200, register_response.text

        submit = client.post(
            "/api/web/check",
            json={
                "chave": "WC96",
                "projeto": "P96A",
                "action": "checkin",
                "informe": "normal",
                "event_time": now_sgt().isoformat(),
                "client_event_id": f"web-membership-guard-{uuid.uuid4().hex}",
            },
        )

    assert submit.status_code == 409
    assert submit.json()["detail"] == "O projeto informado nao pertence aos projetos cadastrados do usuario."

    with SessionLocal() as db:
        user = get_user_by_chave(db, "WC96")
        assert user.projeto == "P80"
        assert get_materialized_user_project_names(db, user.id) == ["P80", "P83"]


def test_provider_endpoint_requires_valid_shared_key():
    with TestClient(app) as client:
        response = client.post(
            "/api/provider/updaterecords",
            json={
                "chave": "PV01",
                "nome": "Usuario Provider",
                "projeto": "P80",
                "atividade": "check-in",
                "informe": "normal",
                "data": "17/04/2026",
                "hora": "08:00:00",
            },
        )

    assert response.status_code == 401
    assert response.json()["detail"] == "Invalid provider shared key"


def test_provider_endpoint_never_writes_check_events_for_failed_duplicate_or_successful_requests():
    with SessionLocal() as db:
        check_event_ids_before = db.execute(
            select(CheckEvent.id)
            .where(CheckEvent.request_path == "/api/provider/updaterecords")
            .order_by(CheckEvent.id)
        ).scalars().all()

    with TestClient(app) as client:
        invalid = client.post(
            "/api/provider/updaterecords",
            json={
                "chave": "PV14",
                "nome": "USUARIO PROVIDER SEM LOG",
                "projeto": "P80",
                "atividade": "check-in",
                "informe": "normal",
                "data": "18/04/2026",
                "hora": "08:00:00",
            },
        )
        assert invalid.status_code == 401, invalid.text

        created = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": "PV14",
                "nome": "USUARIO PROVIDER SEM LOG",
                "projeto": "P80",
                "atividade": "check-in",
                "informe": "normal",
                "data": "18/04/2026",
                "hora": "08:00:00",
            },
        )
        assert created.status_code == 200, created.text
        assert created.json()["duplicate"] is False

        duplicate = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": "PV14",
                "nome": "USUARIO PROVIDER SEM LOG",
                "projeto": "P80",
                "atividade": "check-in",
                "informe": "normal",
                "data": "18/04/2026",
                "hora": "08:00:00",
            },
        )
        assert duplicate.status_code == 200, duplicate.text
        assert duplicate.json()["duplicate"] is True

    with SessionLocal() as db:
        check_event_ids_after = db.execute(
            select(CheckEvent.id)
            .where(CheckEvent.request_path == "/api/provider/updaterecords")
            .order_by(CheckEvent.id)
        ).scalars().all()
        history_rows = db.execute(
            select(CheckingHistory)
            .where(CheckingHistory.chave == "PV14")
            .order_by(CheckingHistory.id)
        ).scalars().all()
        provider_events = db.execute(
            select(UserSyncEvent)
            .where(UserSyncEvent.chave == "PV14", UserSyncEvent.source == "provider")
            .order_by(UserSyncEvent.id)
        ).scalars().all()
        user = get_user_by_chave(db, "PV14")

    assert check_event_ids_after == check_event_ids_before
    assert user is not None
    assert len(history_rows) == 1
    assert history_rows[0].atividade == "check-in"
    assert len(provider_events) == 1


def test_provider_endpoint_creates_user_and_history_with_normalized_name():
    with TestClient(app) as client:
        response = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": "PV11",
                "nome": "ADRIANO JOSE DA SILVA",
                "projeto": "P82",
                "atividade": "check-in",
                "informe": "normal",
                "data": "17/04/2026",
                "hora": "07:26:00",
            },
        )
        assert response.status_code == 200
        payload = response.json()
        assert payload["created_user"] is True
        assert payload["updated_current_state"] is True

    with SessionLocal() as db:
        user = get_user_by_chave(db, "PV11")
        history_rows = db.execute(
            select(CheckingHistory)
            .where(CheckingHistory.chave == "PV11")
            .order_by(CheckingHistory.id)
        ).scalars().all()
        provider_events = db.execute(
            select(UserSyncEvent)
            .where(UserSyncEvent.chave == "PV11", UserSyncEvent.source == "provider")
            .order_by(UserSyncEvent.id)
        ).scalars().all()
        forms_rows = db.execute(
            select(FormsSubmission)
            .where(FormsSubmission.chave == "PV11")
            .order_by(FormsSubmission.id)
        ).scalars().all()

    assert user.nome == "Adriano Jose da Silva"
    assert user.projeto == "P82"
    assert user.checkin is True
    assert user.time.strftime("%d/%m/%Y %H:%M:%S") == "17/04/2026 07:26:00"
    assert len(history_rows) == 1
    assert history_rows[0].atividade == "check-in"
    assert history_rows[0].informe == "normal"
    assert len(provider_events) == 1
    assert forms_rows == []


def test_provider_endpoint_never_enqueues_forms_even_for_multiple_events():
    with TestClient(app) as client:
        first = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": "PV21",
                "nome": "USUARIO FORMS",
                "projeto": "P80",
                "atividade": "check-in",
                "informe": "normal",
                "data": "17/04/2026",
                "hora": "07:30:00",
            },
        )
        second = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": "PV21",
                "nome": "USUARIO FORMS",
                "projeto": "P80",
                "atividade": "check-out",
                "informe": "normal",
                "data": "17/04/2026",
                "hora": "18:00:00",
            },
        )
        assert first.status_code == 200
        assert second.status_code == 200

    with SessionLocal() as db:
        forms_rows = db.execute(
            select(FormsSubmission)
            .where(FormsSubmission.chave == "PV21")
            .order_by(FormsSubmission.id)
        ).scalars().all()
        provider_log_rows = db.execute(
            select(CheckEvent)
            .where(CheckEvent.source == "provider", CheckEvent.request_path == "/api/provider/updaterecords")
            .order_by(CheckEvent.id.desc())
        ).scalars().all()

    assert forms_rows == []
    assert provider_log_rows == []


def test_provider_endpoint_updates_project_but_keeps_existing_name():
    with SessionLocal() as db:
        db.add(
            User(
                rfid=None,
                chave="PV12",
                nome="Nome Original",
                projeto="P80",
                local=None,
                checkin=None,
                time=None,
                last_active_at=now_sgt(),
                inactivity_days=0,
            )
        )
        db.commit()

    with TestClient(app) as client:
        response = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": "PV12",
                "nome": "NOME NAO DEVE TROCAR",
                "projeto": "P83",
                "atividade": "check-out",
                "informe": "retroativo",
                "data": "17/04/2026",
                "hora": "19:40:00",
            },
        )
        assert response.status_code == 200
        payload = response.json()
        assert payload["created_user"] is False
        assert payload["updated_project"] is True

    with SessionLocal() as db:
        user = get_user_by_chave(db, "PV12")
        history_row = db.execute(
            select(CheckingHistory)
            .where(CheckingHistory.chave == "PV12")
            .order_by(CheckingHistory.id.desc())
            .limit(1)
        ).scalar_one()

    assert user.nome == "Nome Original"
    assert user.projeto == "P83"
    assert user.checkin is False
    assert history_row.atividade == "check-out"
    assert history_row.informe == "retroativo"


def test_admin_provider_forms_rows_include_timezone_metadata_for_non_singapore_project():
    with SessionLocal() as db:
        project = db.execute(select(Project).where(Project.name == "P98")).scalar_one_or_none()
        if project is None:
            db.add(Project(name="P98", **build_project_fields_for_country("JP")))
            db.commit()

    with TestClient(app) as client:
        response = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": "PV98",
                "nome": "USUARIO JAPAO",
                "projeto": "P98",
                "atividade": "check-in",
                "informe": "normal",
                "data": "17/04/2026",
                "hora": "07:26:00",
            },
        )
        assert response.status_code == 200, response.text

        ensure_admin_session(client)
        forms_response = client.get("/api/admin/forms")
        assert forms_response.status_code == 200, forms_response.text
        form_row = next(row for row in forms_response.json() if row["chave"] == "PV98")
        assert form_row["projeto"] == "P98"
        assert form_row["timezone_name"] == "Asia/Tokyo"
        assert form_row["timezone_label"] == "Japão (+9)"
        assert form_row["recebimento"] is not None
        assert form_row["recebimento_date_label"] == "17/04/2026"
        assert form_row["recebimento_time_label"] == "08:26:00"
        assert form_row["hora"] == "08:26:00"

    with SessionLocal() as db:
        user = get_user_by_chave(db, "PV98")

    assert user is not None
    assert user.time is not None
    assert normalize_event_time(user.time, timezone_name="Asia/Tokyo").strftime("%d/%m/%Y %H:%M:%S") == "17/04/2026 07:26:00"


def test_provider_endpoint_keeps_newer_current_user_state_while_recording_older_history():
    newer_time = datetime(2026, 4, 18, 9, 0, tzinfo=ZoneInfo(settings.tz_name))
    with SessionLocal() as db:
        db.add(
            User(
                rfid=None,
                chave="PV13",
                nome="Usuario Atual",
                projeto="P82",
                local="main",
                checkin=False,
                time=newer_time,
                last_active_at=newer_time,
                inactivity_days=0,
            )
        )
        db.commit()

    with TestClient(app) as client:
        response = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": "PV13",
                "nome": "IGNORADO",
                "projeto": "P82",
                "atividade": "check-in",
                "informe": "normal",
                "data": "17/04/2026",
                "hora": "08:00:00",
            },
        )
        assert response.status_code == 200
        payload = response.json()
        assert payload["updated_current_state"] is False

        duplicate = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": "PV13",
                "nome": "IGNORADO",
                "projeto": "P82",
                "atividade": "check-in",
                "informe": "normal",
                "data": "17/04/2026",
                "hora": "08:00:00",
            },
        )
        assert duplicate.status_code == 200
        assert duplicate.json()["duplicate"] is True

    with SessionLocal() as db:
        user = get_user_by_chave(db, "PV13")
        history_rows = db.execute(
            select(CheckingHistory)
            .where(CheckingHistory.chave == "PV13")
            .order_by(CheckingHistory.id)
        ).scalars().all()
        provider_events = db.execute(
            select(UserSyncEvent)
            .where(UserSyncEvent.chave == "PV13", UserSyncEvent.source == "provider")
            .order_by(UserSyncEvent.id)
        ).scalars().all()

    assert normalize_event_time(user.time) == newer_time
    assert user.checkin is False
    assert user.local == "main"
    assert len(history_rows) == 2
    assert history_rows[0].atividade == "check-out"
    assert history_rows[1].atividade == "check-in"
    assert len(provider_events) == 1


def test_provider_same_day_events_do_not_override_web_state_and_are_reported_in_forms(monkeypatch):
    web_checkin_time = datetime(2026, 4, 20, 7, 10, 0, tzinfo=ZoneInfo(settings.tz_name))
    web_checkout_time = datetime(2026, 4, 20, 21, 7, 14, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(admin_router, "now_sgt", lambda: web_checkout_time)

    with TestClient(app) as client:
        registered = register_web_password(client, chave="PV31", senha="web123", projeto="P80")
        assert registered.status_code == 200, registered.text

        web_checkin = client.post(
            "/api/web/check",
            json={
                "chave": "PV31",
                "projeto": "P80",
                "action": "checkin",
                "informe": "normal",
                "local": "Web",
                "event_time": web_checkin_time.isoformat(),
                "client_event_id": f"web-checkin-{uuid.uuid4().hex}",
            },
        )
        assert web_checkin.status_code == 200, web_checkin.text

        provider_checkin = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": "PV31",
                "nome": "USUARIO WEB FORMS",
                "projeto": "P80",
                "atividade": "check-in",
                "informe": "normal",
                "data": "20/04/2026",
                "hora": "07:12:33",
            },
        )
        assert provider_checkin.status_code == 200, provider_checkin.text
        assert provider_checkin.json()["updated_current_state"] is False

        ensure_admin_session(client)

        checkin_rows = client.get("/api/admin/checkin")
        assert checkin_rows.status_code == 200, checkin_rows.text
        checkin_row = next(row for row in checkin_rows.json() if row["chave"] == "PV31")
        assert normalize_event_time(datetime.fromisoformat(checkin_row["time"])) == web_checkin_time
        assert checkin_row["local"] == "Web"

        web_checkout = client.post(
            "/api/web/check",
            json={
                "chave": "PV31",
                "projeto": "P80",
                "action": "checkout",
                "informe": "normal",
                "local": "Web",
                "event_time": web_checkout_time.isoformat(),
                "client_event_id": f"web-checkout-{uuid.uuid4().hex}",
            },
        )
        assert web_checkout.status_code == 200, web_checkout.text

        provider_checkout = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": "PV31",
                "nome": "USUARIO WEB FORMS",
                "projeto": "P80",
                "atividade": "check-out",
                "informe": "normal",
                "data": "20/04/2026",
                "hora": "21:09:37",
            },
        )
        assert provider_checkout.status_code == 200, provider_checkout.text
        assert provider_checkout.json()["updated_current_state"] is False

        checkout_rows = client.get("/api/admin/checkout")
        assert checkout_rows.status_code == 200, checkout_rows.text
        checkout_row = next(row for row in checkout_rows.json() if row["chave"] == "PV31")
        assert normalize_event_time(datetime.fromisoformat(checkout_row["time"])) == web_checkout_time
        assert checkout_row["local"] == "Web"

        history_state = client.get("/api/web/check/state", params={"chave": "PV31"})
        assert history_state.status_code == 200, history_state.text
        assert normalize_event_time(datetime.fromisoformat(history_state.json()["last_checkout_at"])) == web_checkout_time

        forms_rows = client.get("/api/admin/forms")
        assert forms_rows.status_code == 200, forms_rows.text
        provider_rows = [row for row in forms_rows.json() if row["chave"] == "PV31"]
        assert any(row["atividade"] == "check-in" and row["hora"] == "07:12:33" for row in provider_rows)
        assert any(row["atividade"] == "check-out" and row["hora"] == "21:09:37" for row in provider_rows)


def test_provider_current_state_uses_forms_as_local_when_provider_event_wins(monkeypatch):
    provider_time = datetime(2026, 4, 21, 18, 5, 0, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(admin_router, "now_sgt", lambda: provider_time)

    with TestClient(app) as client:
        ensure_admin_session(client)

        response = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": "PV32",
                "nome": "USUARIO PROVIDER LOCAL",
                "projeto": "P82",
                "atividade": "check-out",
                "informe": "retroativo",
                "data": "21/04/2026",
                "hora": "18:05:00",
            },
        )
        assert response.status_code == 200, response.text
        assert response.json()["updated_current_state"] is True

        checkout_rows = client.get("/api/admin/checkout")
        assert checkout_rows.status_code == 200, checkout_rows.text
        checkout_row = next(row for row in checkout_rows.json() if row["chave"] == "PV32")
        assert normalize_event_time(datetime.fromisoformat(checkout_row["time"])) == provider_time
        assert checkout_row["local"] == "Forms"
        assert checkout_row["assiduidade"] == "Retroativo"

        forms_rows = client.get("/api/admin/forms")
        assert forms_rows.status_code == 200, forms_rows.text
        forms_row = next(row for row in forms_rows.json() if row["chave"] == "PV32")
        assert forms_row["nome"] == "USUARIO PROVIDER LOCAL"
        assert forms_row["atividade"] == "check-out"
        assert forms_row["informe"] == "retroativo"
        assert forms_row["recebimento"] is not None
        assert forms_row["recebimento_date_label"] == "21/04/2026"
        assert forms_row["recebimento_time_label"] == "18:05:00"
        assert forms_row["hora"] == "18:05:00"


def test_admin_provider_forms_hide_sensitive_time_for_profile_one():
    with SessionLocal() as db:
        admin = find_user_by_chave(db, "P101")
        if admin is None:
            admin = User(
                rfid=None,
                nome="Perfil Um Forms",
                chave="P101",
                projeto="P80",
                senha=hash_password("adm123"),
                perfil=1,
                workplace=None,
                placa=None,
                end_rua=None,
                zip=None,
                email=None,
                local=None,
                checkin=None,
                time=None,
                last_active_at=now_sgt(),
                inactivity_days=0,
            )
            db.add(admin)
        else:
            admin.nome = "Perfil Um Forms"
            admin.projeto = "P80"
            admin.senha = hash_password("adm123")
            admin.perfil = 1
            admin.last_active_at = now_sgt()
            admin.inactivity_days = 0
        db.commit()

    with TestClient(app) as client:
        response = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": "PF41",
                "nome": "USUARIO FORMS PERFIL UM",
                "projeto": "P80",
                "atividade": "check-in",
                "informe": "normal",
                "data": "22/04/2026",
                "hora": "14:32:10",
            },
        )
        assert response.status_code == 200, response.text

        login_response = login_admin(client, chave="P101", senha="adm123")
        assert login_response.status_code == 200, login_response.text

        forms_response = client.get("/api/admin/forms")
        assert forms_response.status_code == 200, forms_response.text
        form_row = next(row for row in forms_response.json() if row["chave"] == "PF41")
        assert form_row["nome"] == "USUARIO FORMS PERFIL UM"
        assert form_row["atividade"] == "check-in"
        assert form_row["informe"] == "normal"
        assert form_row["data"] == "22/04/2026"
        assert form_row["recebimento"] is None
        assert form_row["recebimento_date_label"] == "22/04/2026"
        assert form_row["recebimento_time_label"] is None
        assert form_row["hora"] is None


def test_web_check_ignores_provider_checkout_when_deciding_same_day_submission():
    first_event_time = datetime(2026, 4, 22, 8, 0, 0, tzinfo=ZoneInfo(settings.tz_name))
    second_event_time = datetime(2026, 4, 22, 11, 0, 0, tzinfo=ZoneInfo(settings.tz_name))

    with TestClient(app) as client:
        registered = register_web_password(client, chave="PV33", senha="web123", projeto="P80")
        assert registered.status_code == 200, registered.text

        first = client.post(
            "/api/web/check",
            json={
                "chave": "PV33",
                "projeto": "P80",
                "action": "checkin",
                "informe": "normal",
                "local": "Web",
                "event_time": first_event_time.isoformat(),
                "client_event_id": f"web-provider-ignore-1-{uuid.uuid4().hex}",
            },
        )
        assert first.status_code == 200, first.text
        assert first.json()["queued_forms"] is True

        provider_checkout = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": "PV33",
                "nome": "USUARIO WEB IGNORA PROVIDER",
                "projeto": "P80",
                "atividade": "check-out",
                "informe": "normal",
                "data": "22/04/2026",
                "hora": "10:00:00",
            },
        )
        assert provider_checkout.status_code == 200, provider_checkout.text
        assert provider_checkout.json()["updated_current_state"] is True

        second = client.post(
            "/api/web/check",
            json={
                "chave": "PV33",
                "projeto": "P80",
                "action": "checkout",
                "informe": "normal",
                "local": "Web",
                "event_time": second_event_time.isoformat(),
                "client_event_id": f"web-provider-ignore-2-{uuid.uuid4().hex}",
            },
        )
        assert second.status_code == 200, second.text
        assert second.json()["queued_forms"] is True

        with SessionLocal() as db:
            queued = db.execute(
                select(FormsSubmission).where(FormsSubmission.chave == "PV33").order_by(FormsSubmission.id)
            ).scalars().all()
            assert len(queued) == 2


def test_admin_forms_clear_route_removes_only_provider_sync_rows_and_keeps_legacy_audit_rows():
    timestamp = now_sgt().replace(microsecond=0)
    provider_event_key = uuid.uuid4().hex
    device_event_key = uuid.uuid4().hex

    with SessionLocal() as db:
        user = User(
            rfid="rfid-clear-forms-user",
            nome="Usuario Forms Limpar",
            chave="PF99",
            projeto="P80",
            local="Forms",
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        db.add_all(
            [
                UserSyncEvent(
                    user_id=user.id,
                    chave=user.chave,
                    rfid=user.rfid,
                    source="provider",
                    action="checkin",
                    projeto="P80",
                    local="Forms",
                    ontime=True,
                    event_time=timestamp,
                    created_at=timestamp,
                    source_request_id=f"forms-clear-provider-{uuid.uuid4().hex}",
                    device_id="FORMS-CLEAR-01",
                ),
                CheckEvent(
                    idempotency_key=provider_event_key,
                    source="provider",
                    rfid=None,
                    action="checkin",
                    status="success",
                    message="Entrada Forms registrada",
                    details="chave=PF99; nome=USUARIO LIMPAR FORMS; projeto=P80; atividade=check-in; informe=normal; data=22/04/2026; hora=08:00:00",
                    project="P80",
                    device_id="FORMS-CLEAR-01",
                    local="Forms",
                    request_path="/api/provider/updaterecords",
                    http_status=200,
                    ontime=True,
                    event_time=timestamp,
                    submitted_at=timestamp,
                    retry_count=0,
                ),
                CheckEvent(
                    idempotency_key=device_event_key,
                    source="device",
                    rfid="rfid-clear-forms",
                    action="checkin",
                    status="success",
                    message="Entrada por leitor",
                    details="reader=gate-clear",
                    project="P82",
                    device_id="ESP-CLEAR-01",
                    local="Portaria",
                    request_path="/api/scan",
                    http_status=200,
                    ontime=True,
                    event_time=timestamp,
                    submitted_at=timestamp,
                    retry_count=0,
                ),
            ]
        )
        db.commit()

    with TestClient(app) as client:
        ensure_admin_session(client)

        forms_before = client.get("/api/admin/forms")
        assert forms_before.status_code == 200, forms_before.text
        assert any(row["chave"] == "PF99" for row in forms_before.json())

        cleared = client.delete("/api/admin/forms")
        assert cleared.status_code == 200, cleared.text
        assert "Forms removido" in cleared.json()["message"]

        forms_after = client.get("/api/admin/forms")
        assert forms_after.status_code == 200, forms_after.text
        assert all(row["chave"] != "PF99" for row in forms_after.json())

    with SessionLocal() as db:
        provider_sync_rows = db.execute(
            select(UserSyncEvent).where(UserSyncEvent.chave == "PF99", UserSyncEvent.source == "provider")
        ).scalars().all()
        provider_event = db.execute(
            select(CheckEvent).where(CheckEvent.idempotency_key == provider_event_key)
        ).scalar_one_or_none()
        device_event = db.execute(
            select(CheckEvent).where(CheckEvent.idempotency_key == device_event_key)
        ).scalar_one_or_none()

    assert provider_sync_rows == []
    assert provider_event is not None
    assert device_event is not None


def test_admin_reports_events_returns_history_by_chave_in_desc_order():
    with SessionLocal() as db:
        project = db.execute(select(Project).where(Project.name == "P98")).scalar_one_or_none()
        if project is None:
            db.add(Project(name="P98", **build_project_fields_for_country("JP")))
            db.flush()

        user = User(
            rfid="RPT1001",
            chave="RP41",
            nome="Usuario Relatorio Chave",
            projeto="P80",
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        older_time = datetime(2026, 4, 24, 18, 0, 0, tzinfo=ZoneInfo(settings.tz_name))
        same_time = datetime(2026, 4, 25, 9, 30, 0, tzinfo=ZoneInfo("Asia/Tokyo"))
        db.add_all(
            [
                UserSyncEvent(
                    user_id=user.id,
                    chave=user.chave,
                    rfid=user.rfid,
                    source="web",
                    action="checkin",
                    projeto="P80",
                    local="Web",
                    ontime=True,
                    event_time=older_time,
                    created_at=now_sgt(),
                    source_request_id=f"report-web-{uuid.uuid4().hex}",
                    device_id=None,
                ),
                UserSyncEvent(
                    user_id=user.id,
                    chave=user.chave,
                    rfid=user.rfid,
                    source="provider",
                    action="checkin",
                    projeto="P98",
                    local="Forms",
                    ontime=False,
                    event_time=same_time,
                    created_at=now_sgt(),
                    source_request_id=f"report-provider-1-{uuid.uuid4().hex}",
                    device_id="provider",
                ),
                UserSyncEvent(
                    user_id=user.id,
                    chave=user.chave,
                    rfid=user.rfid,
                    source="provider",
                    action="checkout",
                    projeto="P98",
                    local="Forms",
                    ontime=True,
                    event_time=same_time,
                    created_at=now_sgt(),
                    source_request_id=f"report-provider-2-{uuid.uuid4().hex}",
                    device_id="provider",
                ),
            ]
        )
        db.commit()

    with TestClient(app) as client:
        ensure_admin_session(client)

        response = client.get("/api/admin/reports/events", params={"chave": "rp41"})
        assert response.status_code == 200, response.text
        payload = response.json()

    assert payload["person"]["chave"] == "RP41"
    assert payload["person"]["nome"] == "Usuario Relatorio Chave"
    assert payload["person"]["projeto"] == "P80"
    assert payload["person"]["projetos"] == ["P80"]
    assert len(payload["events"]) == 3
    assert [row["action"] for row in payload["events"]] == ["checkout", "checkin", "checkin"]
    assert payload["events"][0]["source"] == "provider"
    assert payload["events"][0]["source_label"] == "Forms"
    assert payload["events"][0]["projeto"] == "P98"
    assert payload["events"][0]["timezone_name"] == "Asia/Tokyo"
    assert payload["events"][0]["timezone_label"] == "Japão (+9)"
    assert payload["events"][0]["event_date"] == "25/04/2026"
    assert payload["events"][0]["event_time"].startswith("2026-04-25T09:30:00")
    assert payload["events"][0]["event_time_label"] == "10:30:00"
    assert payload["events"][0]["action_label"] == "Check-Out"
    assert payload["events"][0]["local_label"] == "Forms"
    assert payload["events"][0]["assiduidade"] == "Normal"
    assert payload["events"][1]["assiduidade"] == "Retroativo"
    assert payload["events"][2]["source"] == "web"
    assert payload["events"][2]["source_label"] == "Aplicativo"
    assert payload["events"][2]["event_time"].startswith("2026-04-24T18:00:00")


def test_admin_reports_events_returns_history_by_unique_nome():
    with SessionLocal() as db:
        user = User(
            rfid=None,
            chave="RP42",
            nome="Relatorio Nome Unico",
            projeto="P82",
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
        )
        db.add(user)
        db.flush()
        db.add(
            UserSyncEvent(
                user_id=user.id,
                chave=user.chave,
                rfid=user.rfid,
                source="android",
                action="checkin",
                projeto="P82",
                local="App",
                ontime=True,
                event_time=datetime(2026, 4, 25, 7, 0, 0, tzinfo=ZoneInfo(settings.tz_name)),
                created_at=now_sgt(),
                source_request_id=f"report-android-{uuid.uuid4().hex}",
                device_id="android-app",
            )
        )
        db.commit()

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.get("/api/admin/reports/events", params={"nome": "  RELATORIO   NOME UNICO  "})
        assert response.status_code == 200, response.text
        payload = response.json()

    assert payload["person"]["chave"] == "RP42"
    assert payload["person"]["nome"] == "Relatorio Nome Unico"
    assert len(payload["events"]) == 1
    assert payload["events"][0]["source"] == "android"
    assert payload["events"][0]["source_label"] == "android"
    assert payload["events"][0]["local"] == "App"
    assert payload["events"][0]["local_label"] == "App"


def test_admin_reports_events_hide_sensitive_time_for_profile_one():
    with SessionLocal() as db:
        project = db.execute(select(Project).where(Project.name == "P98")).scalar_one_or_none()
        if project is None:
            db.add(Project(name="P98", **build_project_fields_for_country("JP")))
            db.flush()

        admin = find_user_by_chave(db, "P100")
        if admin is None:
            admin = User(
                rfid=None,
                nome="Perfil Um",
                chave="P100",
                projeto="P80",
                senha=hash_password("adm123"),
                perfil=1,
                workplace=None,
                placa=None,
                end_rua=None,
                zip=None,
                email=None,
                local=None,
                checkin=None,
                time=None,
                last_active_at=now_sgt(),
                inactivity_days=0,
            )
            db.add(admin)
        else:
            admin.nome = "Perfil Um"
            admin.projeto = "P80"
            admin.senha = hash_password("adm123")
            admin.perfil = 1
            admin.last_active_at = now_sgt()
            admin.inactivity_days = 0
        db.flush()
        grant_user_project_memberships(db, admin, ["P80", "P98"])

        user = User(
            rfid="RPT1002",
            chave="RP48",
            nome="Usuario Relatorio Restrito",
            projeto="P98",
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        db.add(
            UserSyncEvent(
                user_id=user.id,
                chave=user.chave,
                rfid=user.rfid,
                source="provider",
                action="checkout",
                projeto="P98",
                local="Forms",
                ontime=False,
                event_time=datetime(2026, 4, 25, 9, 30, 0, tzinfo=ZoneInfo("Asia/Tokyo")),
                created_at=now_sgt(),
                source_request_id=f"report-provider-hidden-{uuid.uuid4().hex}",
                device_id="provider",
            )
        )
        db.commit()

    with TestClient(app) as client:
        login_response = login_admin(client, chave="P100", senha="adm123")
        assert login_response.status_code == 200

        response = client.get("/api/admin/reports/events", params={"chave": "RP48"})
        assert response.status_code == 200, response.text
        payload = response.json()

    assert payload["person"]["chave"] == "RP48"
    assert len(payload["events"]) == 1
    assert payload["events"][0]["timezone_name"] == "Asia/Tokyo"
    assert payload["events"][0]["timezone_label"] == "Japão (+9)"
    assert payload["events"][0]["event_date"] == "25/04/2026"
    assert payload["events"][0]["event_time"] is None
    assert payload["events"][0]["event_time_label"] is None
    assert payload["events"][0]["source_label"] == "Forms"
    assert payload["events"][0]["action_label"] == "Check-Out"
    assert payload["events"][0]["assiduidade"] == "Retroativo"


def test_admin_reports_events_export_builds_xlsx_download_with_display_labels():
    fixed_now = datetime(2026, 4, 22, 15, 26, 45, tzinfo=ZoneInfo(settings.tz_name))

    with SessionLocal() as db:
        project = db.execute(select(Project).where(Project.name == "P80")).scalar_one_or_none()
        if project is None:
            db.add(Project(name="P80", **build_project_fields_for_country("SG")))
            db.flush()

        user = User(
            rfid="RPT1003",
            chave="RP49",
            nome="Usuario Relatorio Export",
            projeto="P80",
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
        )
        db.add(user)
        db.flush()
        db.add(
            UserSyncEvent(
                user_id=user.id,
                chave=user.chave,
                rfid=user.rfid,
                source="device",
                action="checkout",
                projeto="P80",
                local="main",
                ontime=False,
                event_time=datetime(2026, 4, 25, 7, 8, 9, tzinfo=ZoneInfo(settings.tz_name)),
                created_at=now_sgt(),
                source_request_id=f"report-device-export-{uuid.uuid4().hex}",
                device_id="esp32-box-0001",
            )
        )
        db.commit()

    with patch("sistema.app.routers.admin.now_sgt", return_value=fixed_now):
        with TestClient(app) as client:
            ensure_admin_session(client)
            exported = client.get("/api/admin/reports/events/export", params={"chave": "RP49"})

    assert exported.status_code == 200, exported.text
    assert exported.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert (
        exported.headers["content-disposition"]
        == 'attachment; filename="Relatorio - RP49 - 20260422 - 152645.xlsx"'
    )

    workbook = load_workbook(io.BytesIO(exported.content))
    worksheet = workbook.active
    assert worksheet.title == "Relatório"
    assert worksheet["A1"].value == "Usuario Relatorio Export (RP49)"
    assert worksheet["A2"].value == "Projeto atual: P80 | RFID: RPT1003 | Fuso horário: Singapura (+8) | 1 evento"
    assert {str(cell_range) for cell_range in worksheet.merged_cells.ranges} == {"A1:H1", "A2:H2"}
    assert [worksheet["A4"].value, worksheet["B4"].value, worksheet["C4"].value, worksheet["D4"].value, worksheet["E4"].value, worksheet["F4"].value, worksheet["G4"].value, worksheet["H4"].value] == [
        "Data",
        "Horário",
        "Ação",
        "Origem",
        "Local",
        "Projeto",
        "Fuso Horário",
        "Assiduidade",
    ]
    assert [worksheet["A5"].value, worksheet["B5"].value, worksheet["C5"].value, worksheet["D5"].value, worksheet["E5"].value, worksheet["F5"].value, worksheet["G5"].value, worksheet["H5"].value] == [
        "25/04/2026",
        "07:08:09",
        "Check-Out",
        "Box ESP32-0001",
        "Escritório Principal",
        "P80",
        "Singapura (+8)",
        "Retroativo",
    ]
    workbook.close()


def test_admin_reports_events_export_hides_time_column_for_profile_one():
    fixed_now = datetime(2026, 4, 22, 15, 40, 45, tzinfo=ZoneInfo(settings.tz_name))

    with SessionLocal() as db:
        project = db.execute(select(Project).where(Project.name == "P80")).scalar_one_or_none()
        if project is None:
            db.add(Project(name="P80", **build_project_fields_for_country("SG")))
            db.flush()

        admin = find_user_by_chave(db, "P100")
        if admin is None:
            admin = User(
                rfid=None,
                nome="Perfil Um",
                chave="P100",
                projeto="P80",
                senha=hash_password("adm123"),
                perfil=1,
                workplace=None,
                placa=None,
                end_rua=None,
                zip=None,
                email=None,
                local=None,
                checkin=None,
                time=None,
                last_active_at=now_sgt(),
                inactivity_days=0,
            )
            db.add(admin)
        else:
            admin.nome = "Perfil Um"
            admin.projeto = "P80"
            admin.senha = hash_password("adm123")
            admin.perfil = 1
            admin.last_active_at = now_sgt()
            admin.inactivity_days = 0
        db.flush()
        grant_user_project_memberships(db, admin, ["P80"])

        user = User(
            rfid="RPT1004",
            chave="RP50",
            nome="Usuario Relatorio Export Restrito",
            projeto="P80",
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
        )
        db.add(user)
        db.flush()
        db.add(
            UserSyncEvent(
                user_id=user.id,
                chave=user.chave,
                rfid=user.rfid,
                source="device",
                action="checkout",
                projeto="P80",
                local="main",
                ontime=False,
                event_time=datetime(2026, 4, 25, 7, 8, 9, tzinfo=ZoneInfo(settings.tz_name)),
                created_at=now_sgt(),
                source_request_id=f"report-device-export-hidden-{uuid.uuid4().hex}",
                device_id="esp32-box-0001",
            )
        )
        db.commit()

    with patch("sistema.app.routers.admin.now_sgt", return_value=fixed_now):
        with TestClient(app) as client:
            login_response = login_admin(client, chave="P100", senha="adm123")
            assert login_response.status_code == 200

            exported = client.get("/api/admin/reports/events/export", params={"chave": "RP50"})

    assert exported.status_code == 200, exported.text
    assert exported.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert (
        exported.headers["content-disposition"]
        == 'attachment; filename="Relatorio - RP50 - 20260422 - 154045.xlsx"'
    )

    workbook = load_workbook(io.BytesIO(exported.content))
    worksheet = workbook.active
    assert worksheet.title == "Relatório"
    assert worksheet["A1"].value == "Usuario Relatorio Export Restrito (RP50)"
    assert worksheet["A2"].value == "Projeto atual: P80 | RFID: RPT1004 | Fuso horário: Singapura (+8) | 1 evento"
    assert {str(cell_range) for cell_range in worksheet.merged_cells.ranges} == {"A1:G1", "A2:G2"}
    assert [worksheet["A4"].value, worksheet["B4"].value, worksheet["C4"].value, worksheet["D4"].value, worksheet["E4"].value, worksheet["F4"].value, worksheet["G4"].value] == [
        "Data",
        "Ação",
        "Origem",
        "Local",
        "Projeto",
        "Fuso Horário",
        "Assiduidade",
    ]
    assert worksheet["H4"].value is None
    assert [worksheet["A5"].value, worksheet["B5"].value, worksheet["C5"].value, worksheet["D5"].value, worksheet["E5"].value, worksheet["F5"].value, worksheet["G5"].value] == [
        "25/04/2026",
        "Check-Out",
        "Box ESP32-0001",
        "Escritório Principal",
        "P80",
        "Singapura (+8)",
        "Retroativo",
    ]
    assert worksheet["H5"].value is None
    assert not any(cell.value == "07:08:09" for row in worksheet.iter_rows() for cell in row)
    workbook.close()


def test_admin_reports_events_export_all_builds_xlsx_download_for_all_users():
    fixed_now = datetime(2026, 4, 22, 16, 45, 30, tzinfo=ZoneInfo(settings.tz_name))
    project_name_a = "P910"
    project_name_b = "P911"

    with SessionLocal() as db:
        project_a = db.execute(select(Project).where(Project.name == project_name_a)).scalar_one_or_none()
        if project_a is None:
            db.add(Project(name=project_name_a, **build_project_fields_for_country("SG")))
        project_b = db.execute(select(Project).where(Project.name == project_name_b)).scalar_one_or_none()
        if project_b is None:
            db.add(Project(name=project_name_b, **build_project_fields_for_country("BR")))
        db.flush()

        user_a = User(
            rfid="RPT2001",
            chave="RA11",
            nome="Ana Relatorio",
            projeto=project_name_a,
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
        )
        user_b = User(
            rfid="RPT2002",
            chave="RB22",
            nome="Bruno Relatorio",
            projeto=project_name_b,
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
        )
        db.add_all([user_a, user_b])
        db.flush()
        db.add_all(
            [
                UserSyncEvent(
                    user_id=user_a.id,
                    chave=user_a.chave,
                    rfid=user_a.rfid,
                    source="web",
                    action="checkin",
                        projeto=project_name_a,
                    local="main",
                    ontime=True,
                    event_time=datetime(2026, 4, 25, 7, 15, 0, tzinfo=ZoneInfo(settings.tz_name)),
                    created_at=now_sgt(),
                    source_request_id=f"report-export-all-a-{uuid.uuid4().hex}",
                    device_id=None,
                ),
                UserSyncEvent(
                    user_id=user_b.id,
                    chave=user_b.chave,
                    rfid=user_b.rfid,
                    source="provider",
                    action="checkout",
                        projeto=project_name_b,
                    local="Forms",
                    ontime=False,
                    event_time=datetime(2026, 4, 25, 8, 30, 0, tzinfo=ZoneInfo(settings.tz_name)),
                    created_at=now_sgt(),
                    source_request_id=f"report-export-all-b-{uuid.uuid4().hex}",
                    device_id="provider",
                ),
            ]
        )
        db.commit()

    with patch("sistema.app.routers.admin.now_sgt", return_value=fixed_now):
        with TestClient(app) as client:
            ensure_admin_session(client)
            exported = client.get("/api/admin/reports/events/export-all")

    assert exported.status_code == 200, exported.text
    assert exported.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert (
        exported.headers["content-disposition"]
        == 'attachment; filename="Relatorio - Todos - 20260422 - 164530.xlsx"'
    )

    workbook = load_workbook(io.BytesIO(exported.content))
    worksheet = workbook.active
    assert worksheet.title == "Relatório"
    assert [worksheet["A1"].value, worksheet["B1"].value, worksheet["C1"].value, worksheet["D1"].value, worksheet["E1"].value, worksheet["F1"].value, worksheet["G1"].value, worksheet["H1"].value, worksheet["I1"].value] == [
        "Nome",
        "Data",
        "Horário",
        "Ação",
        "Origem",
        "Local",
        "Projeto",
        "Fuso Horário",
        "Assiduidade",
    ]
    rows_by_name = {
        row[0]: row
        for row in worksheet.iter_rows(min_row=2, max_col=9, values_only=True)
        if row[0] in {"Ana Relatorio", "Bruno Relatorio"}
    }
    assert rows_by_name["Ana Relatorio"] == (
        "Ana Relatorio",
        "25/04/2026",
        "07:15:00",
        "Check-In",
        "Aplicativo",
        "Escritório Principal",
        project_name_a,
        "Singapura (+8)",
        "Normal",
    )
    assert rows_by_name["Bruno Relatorio"] == (
        "Bruno Relatorio",
        "24/04/2026",
        "21:30:00",
        "Check-Out",
        "Forms",
        "Forms",
        project_name_b,
        "Brasil (-3)",
        "Retroativo",
    )
    workbook.close()


def test_admin_reports_events_export_all_hides_time_column_for_profile_one():
    fixed_now = datetime(2026, 4, 22, 16, 55, 30, tzinfo=ZoneInfo(settings.tz_name))
    project_name_a = "P912"
    project_name_b = "P913"

    with SessionLocal() as db:
        project_a = db.execute(select(Project).where(Project.name == project_name_a)).scalar_one_or_none()
        if project_a is None:
            db.add(Project(name=project_name_a, **build_project_fields_for_country("SG")))
        project_b = db.execute(select(Project).where(Project.name == project_name_b)).scalar_one_or_none()
        if project_b is None:
            db.add(Project(name=project_name_b, **build_project_fields_for_country("BR")))

        admin = find_user_by_chave(db, "P100")
        if admin is None:
            admin = User(
                rfid=None,
                nome="Perfil Um",
                chave="P100",
                projeto="P80",
                senha=hash_password("adm123"),
                perfil=1,
                workplace=None,
                placa=None,
                end_rua=None,
                zip=None,
                email=None,
                local=None,
                checkin=None,
                time=None,
                last_active_at=now_sgt(),
                inactivity_days=0,
            )
            db.add(admin)
        else:
            admin.nome = "Perfil Um"
            admin.projeto = "P80"
            admin.senha = hash_password("adm123")
            admin.perfil = 1
            admin.last_active_at = now_sgt()
            admin.inactivity_days = 0

        db.flush()
        grant_user_project_memberships(db, admin, ["P80", project_name_a, project_name_b])

        user_a = User(
            rfid="RPT3001",
            chave="RC11",
            nome="Carla Relatorio",
            projeto=project_name_a,
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
        )
        user_b = User(
            rfid="RPT3002",
            chave="RD22",
            nome="Daniel Relatorio",
            projeto=project_name_b,
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
        )
        db.add_all([user_a, user_b])
        db.flush()
        db.add_all(
            [
                UserSyncEvent(
                    user_id=user_a.id,
                    chave=user_a.chave,
                    rfid=user_a.rfid,
                    source="web",
                    action="checkin",
                    projeto=project_name_a,
                    local="main",
                    ontime=True,
                    event_time=datetime(2026, 4, 25, 7, 15, 0, tzinfo=ZoneInfo(settings.tz_name)),
                    created_at=now_sgt(),
                    source_request_id=f"report-export-all-hidden-a-{uuid.uuid4().hex}",
                    device_id=None,
                ),
                UserSyncEvent(
                    user_id=user_b.id,
                    chave=user_b.chave,
                    rfid=user_b.rfid,
                    source="provider",
                    action="checkout",
                    projeto=project_name_b,
                    local="Forms",
                    ontime=False,
                    event_time=datetime(2026, 4, 25, 8, 30, 0, tzinfo=ZoneInfo(settings.tz_name)),
                    created_at=now_sgt(),
                    source_request_id=f"report-export-all-hidden-b-{uuid.uuid4().hex}",
                    device_id="provider",
                ),
            ]
        )
        db.commit()

    with patch("sistema.app.routers.admin.now_sgt", return_value=fixed_now):
        with TestClient(app) as client:
            login_response = login_admin(client, chave="P100", senha="adm123")
            assert login_response.status_code == 200

            exported = client.get("/api/admin/reports/events/export-all")

    assert exported.status_code == 200, exported.text
    assert exported.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert (
        exported.headers["content-disposition"]
        == 'attachment; filename="Relatorio - Todos - 20260422 - 165530.xlsx"'
    )

    workbook = load_workbook(io.BytesIO(exported.content))
    worksheet = workbook.active
    assert worksheet.title == "Relatório"
    assert [worksheet["A1"].value, worksheet["B1"].value, worksheet["C1"].value, worksheet["D1"].value, worksheet["E1"].value, worksheet["F1"].value, worksheet["G1"].value, worksheet["H1"].value] == [
        "Nome",
        "Data",
        "Ação",
        "Origem",
        "Local",
        "Projeto",
        "Fuso Horário",
        "Assiduidade",
    ]
    assert worksheet["I1"].value is None
    rows_by_name = {
        row[0]: row
        for row in worksheet.iter_rows(min_row=2, max_col=9, values_only=True)
        if row[0] in {"Carla Relatorio", "Daniel Relatorio"}
    }
    assert rows_by_name["Carla Relatorio"] == (
        "Carla Relatorio",
        "25/04/2026",
        "Check-In",
        "Aplicativo",
        "Escritório Principal",
        project_name_a,
        "Singapura (+8)",
        "Normal",
        None,
    )
    assert rows_by_name["Daniel Relatorio"] == (
        "Daniel Relatorio",
        "24/04/2026",
        "Check-Out",
        "Forms",
        "Forms",
        project_name_b,
        "Brasil (-3)",
        "Retroativo",
        None,
    )
    assert not any(cell.value in {"07:15:00", "21:30:00"} for row in worksheet.iter_rows() for cell in row)
    workbook.close()


def test_admin_reports_events_rejects_ambiguous_nome():
    duplicated_name = "Relatorio Nome Duplicado"

    with SessionLocal() as db:
        db.add_all(
            [
                User(
                    rfid=None,
                    chave="RP43",
                    nome=duplicated_name,
                    projeto="P80",
                    local=None,
                    checkin=None,
                    time=None,
                    last_active_at=now_sgt(),
                    inactivity_days=0,
                ),
                User(
                    rfid=None,
                    chave="RP44",
                    nome=duplicated_name,
                    projeto="P82",
                    local=None,
                    checkin=None,
                    time=None,
                    last_active_at=now_sgt(),
                    inactivity_days=0,
                ),
            ]
        )
        db.commit()

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.get("/api/admin/reports/events", params={"nome": duplicated_name.upper()})
        assert response.status_code == 409, response.text
        assert "Use a chave" in response.json()["detail"]


def test_admin_reports_events_require_exactly_one_search_criterion():
    with TestClient(app) as client:
        ensure_admin_session(client)

        missing = client.get("/api/admin/reports/events")
        assert missing.status_code == 400, missing.text
        assert "Informe chave ou nome" in missing.json()["detail"]

        invalid = client.get("/api/admin/reports/events", params={"chave": "RP41", "nome": "Usuario"})
        assert invalid.status_code == 400, invalid.text
        assert "Informe apenas chave ou nome" in invalid.json()["detail"]


def test_admin_reports_events_route_is_restricted_to_full_admin():
    with SessionLocal() as db:
        user = find_user_by_chave(db, "PZ00")
        if user is None:
            user = User(
                rfid=None,
                nome="Perfil Zero",
                chave="PZ00",
                projeto="P80",
                senha=hash_password("lim123"),
                perfil=0,
                workplace=None,
                placa=None,
                end_rua=None,
                zip=None,
                email=None,
                local=None,
                checkin=None,
                time=None,
                last_active_at=now_sgt(),
                inactivity_days=0,
            )
            db.add(user)
        else:
            user.nome = "Perfil Zero"
            user.projeto = "P80"
            user.senha = hash_password("lim123")
            user.perfil = 0
            user.last_active_at = now_sgt()
            user.inactivity_days = 0
        db.commit()

    with TestClient(app) as client:
        login_response = login_admin(client, chave="PZ00", senha="lim123")
        assert login_response.status_code == 200

        denied = client.get("/api/admin/reports/events", params={"chave": "RP41"})
        assert denied.status_code == 403, denied.text
        assert denied.json()["detail"] == "Este usuario nao possui permissao para esta area do Admin."


def test_admin_events_keep_raw_and_safe_activity_time_for_profile_nine():
    with SessionLocal() as db:
        project = db.execute(select(Project).where(Project.name == "P98")).scalar_one_or_none()
        if project is None:
            db.add(Project(name="P98", **build_project_fields_for_country("JP")))
            db.flush()

        user = User(
            rfid="EVT9001",
            chave="EV91",
            nome="Usuario Evento Perfil Nove",
            projeto="P98",
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        event_time = datetime(2026, 4, 25, 9, 30, 0, tzinfo=ZoneInfo("Asia/Tokyo"))
        event = CheckEvent(
            idempotency_key=f"event-profile-nine-{uuid.uuid4().hex}",
            source="device",
            rfid=user.rfid,
            action="checkin",
            status="success",
            message="Entrada registrada",
            details="reader=tokyo-gate",
            project="P98",
            device_id="ESP-TOKYO-01",
            local="Portaria",
            request_path="/api/scan",
            http_status=200,
            ontime=True,
            event_time=event_time,
            submitted_at=event_time,
            retry_count=0,
        )
        db.add(event)
        db.commit()
        event_id = event.id

    with TestClient(app) as client:
        ensure_admin_session(client)

        response = client.get("/api/admin/events")
        assert response.status_code == 200, response.text
        payload = response.json()

    row = next(item for item in payload if item["id"] == event_id)
    assert row["source"] == "device"
    assert row["chave"] == "EV91"
    assert row["timezone_name"] == "Asia/Tokyo"
    assert row["timezone_label"] == "Japão (+9)"
    assert row["event_time"].startswith("2026-04-25T09:30:00")
    assert row["event_date_label"] == "25/04/2026"
    assert row["event_time_label"] == "10:30:00"
    assert row["request_path"] == "/api/scan"


def test_admin_events_hide_sensitive_time_for_profile_one():
    with SessionLocal() as db:
        project = db.execute(select(Project).where(Project.name == "P911")).scalar_one_or_none()
        if project is None:
            db.add(Project(name="P911", **build_project_fields_for_country("BR")))
            db.flush()

        admin = find_user_by_chave(db, "P100")
        if admin is None:
            admin = User(
                rfid=None,
                nome="Perfil Um",
                chave="P100",
                projeto="P80",
                senha=hash_password("adm123"),
                perfil=1,
                workplace=None,
                placa=None,
                end_rua=None,
                zip=None,
                email=None,
                local=None,
                checkin=None,
                time=None,
                last_active_at=now_sgt(),
                inactivity_days=0,
            )
            db.add(admin)
        else:
            admin.nome = "Perfil Um"
            admin.projeto = "P80"
            admin.senha = hash_password("adm123")
            admin.perfil = 1
            admin.last_active_at = now_sgt()
            admin.inactivity_days = 0
        db.flush()
        grant_user_project_memberships(db, admin, ["P80", "P911"])

        user = User(
            rfid="EVT1001",
            chave="EV11",
            nome="Usuario Evento Perfil Um",
            projeto="P911",
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        event_time = datetime(2026, 4, 25, 0, 30, 0, tzinfo=ZoneInfo(settings.tz_name))
        event = CheckEvent(
            idempotency_key=f"event-profile-one-{uuid.uuid4().hex}",
            source="provider",
            rfid=user.rfid,
            action="checkout",
            status="success",
            message="Saída Forms registrada",
            details="source=forms",
            project="P911",
            device_id="FORMS-BR-01",
            local="Forms",
            request_path="/api/provider/updaterecords",
            http_status=200,
            ontime=False,
            event_time=event_time,
            submitted_at=event_time,
            retry_count=1,
        )
        db.add(event)
        db.commit()
        event_id = event.id

    with TestClient(app) as client:
        login_response = login_admin(client, chave="P100", senha="adm123")
        assert login_response.status_code == 200

        response = client.get("/api/admin/events")
        assert response.status_code == 200, response.text
        payload = response.json()

    row = next(item for item in payload if item["id"] == event_id)
    assert row["source"] == "provider"
    assert row["chave"] == "EV11"
    assert row["timezone_name"] == "America/Sao_Paulo"
    assert row["timezone_label"] == "Brasil (-3)"
    assert row["event_time"] is None
    assert row["event_date_label"] == "24/04/2026"
    assert row["event_time_label"] is None
    assert row["request_path"] == "/api/provider/updaterecords"
    assert row["retry_count"] == 1


def test_admin_stream_requires_valid_session():
    with TestClient(app) as client:
        forbidden = client.get("/api/admin/stream")
        assert forbidden.status_code == 401


def test_admin_routes_do_not_accept_legacy_header_only():
    with TestClient(app) as client:
        forbidden = client.get("/api/admin/pending", headers={"x-admin-key": "admin-test-key"})
        assert forbidden.status_code == 401


def test_admin_updates_broker_publishes_payload():
    broker = AdminUpdatesBroker()
    subscriber_id, queue = broker.subscribe()

    try:
        broker.publish("pending")
        payload = queue.get_nowait()
        assert '"reason": "pending"' in payload
        assert '"emitted_at":' in payload
    finally:
        broker.unsubscribe(subscriber_id)


def test_http_runtime_builds_gunicorn_command_from_environment(monkeypatch):
    monkeypatch.setenv("APP_HOST", "127.0.0.1")
    monkeypatch.setenv("APP_PORT", "18080")
    monkeypatch.setenv("APP_WORKERS", "2")
    monkeypatch.setenv("APP_KEEPALIVE_SECONDS", "5")
    monkeypatch.setenv("APP_TIMEOUT_SECONDS", "90")
    monkeypatch.setenv("APP_GRACEFUL_TIMEOUT_SECONDS", "30")
    monkeypatch.setenv("APP_MAX_REQUESTS", "1000")
    monkeypatch.setenv("APP_MAX_REQUESTS_JITTER", "100")

    command = http_runtime_module.build_http_server_command()

    assert command == [
        sys.executable,
        "-m",
        "gunicorn.app.wsgiapp",
        "sistema.app.main:app",
        "--worker-class",
        "uvicorn.workers.UvicornWorker",
        "--workers",
        "2",
        "--bind",
        "127.0.0.1:18080",
        "--keep-alive",
        "5",
        "--timeout",
        "90",
        "--graceful-timeout",
        "30",
        "--max-requests",
        "1000",
        "--max-requests-jitter",
        "100",
    ]


def test_http_runtime_execs_server_without_migration_preflight(monkeypatch):
    startup_calls: list[tuple[str, object]] = []

    def fake_execv(executable, argv):
        startup_calls.append(("execv", (executable, argv)))
        raise SystemExit(0)

    monkeypatch.setattr(
        http_runtime_module,
        "build_http_server_command",
        lambda: [
            sys.executable,
            "-m",
            "gunicorn.app.wsgiapp",
            "sistema.app.main:app",
        ],
    )
    monkeypatch.setattr(http_runtime_module.os, "execv", fake_execv)

    with pytest.raises(SystemExit):
        http_runtime_module.main()

    assert startup_calls == [
        (
            "execv",
            (
                sys.executable,
                [
                    sys.executable,
                    "-m",
                    "gunicorn.app.wsgiapp",
                    "sistema.app.main:app",
                ],
            ),
        ),
    ]


def test_admin_updates_broker_dispatches_cross_worker_payload_without_local_duplicates(monkeypatch):
    source_broker = AdminUpdatesBroker("checking_admin_updates")
    target_broker = AdminUpdatesBroker("checking_admin_updates")
    source_subscriber_id, source_queue = source_broker.subscribe()
    target_subscriber_id, target_queue = target_broker.subscribe()
    published_payloads: list[str] = []

    def fake_publish_payload_to_postgres(payload: str) -> bool:
        published_payloads.append(payload)
        target_broker._dispatch_remote_payload(payload)
        source_broker._dispatch_remote_payload(payload)
        return True

    monkeypatch.setattr(source_broker, "_supports_cross_worker", lambda: True)
    monkeypatch.setattr(source_broker, "_publish_payload_to_postgres", fake_publish_payload_to_postgres)

    try:
        source_broker.publish("event", metadata={"scope": "admin"})

        source_payload = source_queue.get_nowait()
        target_payload = target_queue.get_nowait()
        source_event = json.loads(source_payload)
        target_event = json.loads(target_payload)

        assert published_payloads == [source_payload]
        assert source_event["reason"] == "event"
        assert source_event["scope"] == "admin"
        assert target_event["reason"] == "event"
        assert target_event["scope"] == "admin"
        assert target_event["event_id"] == source_event["event_id"]

        with pytest.raises(asyncio.QueueEmpty):
            source_queue.get_nowait()
    finally:
        source_broker.unsubscribe(source_subscriber_id)
        target_broker.unsubscribe(target_subscriber_id)


def test_pending_registration_flow():
    with TestClient(app) as client:
        ensure_admin_session(client)
        heartbeat = client.post(
            "/api/device/heartbeat",
            json={"device_id": "ESP32-TEST", "shared_key": "device-test-key"},
        )
        assert heartbeat.status_code == 200

        scan_pending = client.post(
            "/api/scan",
            json={
                "local": "main",
                "rfid": "ABC12345",
                "action": "checkin",
                "device_id": "ESP32-TEST",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_pending.status_code == 200
        assert scan_pending.json()["outcome"] == "pending_registration"
        assert scan_pending.json()["led"] == "orange_4s"

        pending_list = client.get("/api/admin/pending")
        assert pending_list.status_code == 200
        assert len(pending_list.json()) >= 1

        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "ABC12345", "nome": "Usuario Teste", "chave": "UT70", "projeto": "P80"},
        )
        assert save_user.status_code == 200

        checkin_rows = client.get("/api/admin/checkin")
        assert checkin_rows.status_code == 200
        assert all(row["rfid"] != "ABC12345" for row in checkin_rows.json())

        checkout_rows = client.get("/api/admin/checkout")
        assert checkout_rows.status_code == 200
        assert all(row["rfid"] != "ABC12345" for row in checkout_rows.json())


def test_unknown_rfid_goes_pending():
    with TestClient(app) as client:
        ensure_admin_session(client)
        scan_unknown = client.post(
            "/api/scan",
            json={
                "local": "main",
                "rfid": "ZZZ99999",
                "action": "checkout",
                "device_id": "ESP32-TEST",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_unknown.status_code == 200
        assert scan_unknown.json()["outcome"] == "pending_registration"
        assert scan_unknown.json()["led"] == "orange_4s"

        events = client.get("/api/admin/events")
        assert events.status_code == 200
        assert any(event["status"] == "received" and event["request_path"] == "/api/scan" for event in events.json())
        assert any(event["status"] == "pending" and event["source"] == "device" for event in events.json())


def test_explicit_checkin_and_checkout_flow(monkeypatch):
    monkeypatch.setattr(
        FormsWorker,
        "submit_with_retries",
        lambda self, action, chave, projeto, ontime=True, **_kwargs: {
            "success": True,
            "message": f"mocked {action}",
            "retry_count": 0,
        },
    )

    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "CARD1000", "nome": "Usuario Fluxo", "chave": "AB12", "projeto": "P83"},
        )
        assert save_user.status_code == 200

        scan_checkin = client.post(
            "/api/scan",
            json={
                "local": "main",
                "rfid": "CARD1000",
                "action": "checkin",
                "device_id": "ESP32-TEST",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_checkin.status_code == 200
        assert scan_checkin.json()["outcome"] == "submitted"
        assert scan_checkin.json()["led"] == "green_1s"

        processed_first = process_forms_submission_queue_once()
        assert processed_first >= 1

        checkin_rows = client.get("/api/admin/checkin")
        assert checkin_rows.status_code == 200
        assert any(row["rfid"] == "CARD1000" and row["local"] == "main" for row in checkin_rows.json())

        scan_checkin_again = client.post(
            "/api/scan",
            json={
                "local": "un83",
                "rfid": "CARD1000",
                "action": "checkin",
                "device_id": "ESP32-TEST",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_checkin_again.status_code == 200
        assert scan_checkin_again.json()["outcome"] == "local_updated"
        assert scan_checkin_again.json()["led"] == "green_blink_3x_1s"

        checkin_rows_updated = client.get("/api/admin/checkin")
        assert checkin_rows_updated.status_code == 200
        assert any(row["rfid"] == "CARD1000" and row["local"] == "un83" for row in checkin_rows_updated.json())

        checkout_rows_after_checkin = client.get("/api/admin/checkout")
        assert checkout_rows_after_checkin.status_code == 200
        assert all(row["rfid"] != "CARD1000" for row in checkout_rows_after_checkin.json())

        scan_checkout = client.post(
            "/api/scan",
            json={
                "local": "co83",
                "rfid": "CARD1000",
                "action": "checkout",
                "device_id": "ESP32-TEST",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_checkout.status_code == 200
        assert scan_checkout.json()["outcome"] == "submitted"
        assert scan_checkout.json()["led"] == "green_1s"

        processed_second = process_forms_submission_queue_once()
        assert processed_second >= 1

        checkout_rows = client.get("/api/admin/checkout")
        assert checkout_rows.status_code == 200
        assert any(row["rfid"] == "CARD1000" and row["local"] == "co83" for row in checkout_rows.json())

        checkin_rows_after = client.get("/api/admin/checkin")
        assert checkin_rows_after.status_code == 200
        assert all(row["rfid"] != "CARD1000" for row in checkin_rows_after.json())

        events = client.get("/api/admin/events")
        assert events.status_code == 200
        assert any(
            event["rfid"] == "CARD1000"
            and event["chave"] == "AB12"
            and event["request_path"] == "/api/scan"
            for event in events.json()
        )


def test_checkout_without_checkin_returns_red_2s(monkeypatch):
    monkeypatch.setattr(
        FormsWorker,
        "submit_with_retries",
        lambda self, action, chave, projeto, ontime=True, **_kwargs: {
            "success": True,
            "message": f"mocked {action}",
            "retry_count": 0,
        },
    )

    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "CARD2000", "nome": "Usuario Sem Checkin", "chave": "EF56", "projeto": "P80"},
        )
        assert save_user.status_code == 200

        scan_checkout = client.post(
            "/api/scan",
            json={
                "local": "main",
                "rfid": "CARD2000",
                "action": "checkout",
                "device_id": "ESP32-TEST",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_checkout.status_code == 200
        assert scan_checkout.json()["outcome"] == "failed"
        assert scan_checkout.json()["led"] == "red_2s"
        assert "Check-in not found" in scan_checkout.json()["message"]

        events = client.get("/api/admin/events")
        assert events.status_code == 200
        assert any(event["rfid"] == "CARD2000" and event["status"] == "blocked" for event in events.json())


def test_repeated_same_day_checkout_updates_state_without_forms_submission(monkeypatch):
    monkeypatch.setattr(
        FormsWorker,
        "submit_with_retries",
        lambda self, action, chave, projeto, ontime=True, **_kwargs: {
            "success": True,
            "message": f"mocked {action}",
            "retry_count": 0,
        },
    )

    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "CARD2001", "nome": "Usuario Checkout Repetido", "chave": "EG57", "projeto": "P80"},
        )
        assert save_user.status_code == 200

        scan_checkin = client.post(
            "/api/scan",
            json={
                "local": "main",
                "rfid": "CARD2001",
                "action": "checkin",
                "device_id": "ESP32-TEST",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_checkin.status_code == 200
        assert scan_checkin.json()["outcome"] == "submitted"

        first_checkout = client.post(
            "/api/scan",
            json={
                "local": "co80-a",
                "rfid": "CARD2001",
                "action": "checkout",
                "device_id": "ESP32-TEST",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert first_checkout.status_code == 200
        assert first_checkout.json()["outcome"] == "submitted"

        second_checkout = client.post(
            "/api/scan",
            json={
                "local": "co80-b",
                "rfid": "CARD2001",
                "action": "checkout",
                "device_id": "ESP32-TEST",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert second_checkout.status_code == 200
        assert second_checkout.json()["outcome"] == "local_updated"
        assert second_checkout.json()["led"] == "green_blink_3x_1s"

        with SessionLocal() as db:
            user = get_user_by_rfid(db, "CARD2001")
            queued = db.execute(select(FormsSubmission).where(FormsSubmission.rfid == "CARD2001")).scalars().all()

            assert user.checkin is False
            assert user.local == "co80-b"
            assert len(queued) == 3
            assert queued[-1].status == "skipped"
            assert queued[-1].display_status == "not_realized"


def test_device_checkout_ignores_provider_only_history():
    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "CARDPROV1", "nome": "Usuario Provider Ignorado", "chave": "EG59", "projeto": "P80"},
        )
        assert save_user.status_code == 200, save_user.text

        provider_checkin = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": "EG59",
                "nome": "USUARIO DEVICE IGNORA PROVIDER",
                "projeto": "P80",
                "atividade": "check-in",
                "informe": "normal",
                "data": "22/04/2026",
                "hora": "08:00:00",
            },
        )
        assert provider_checkin.status_code == 200, provider_checkin.text
        assert provider_checkin.json()["updated_current_state"] is True

        scan_checkout = client.post(
            "/api/scan",
            json={
                "local": "main",
                "rfid": "CARDPROV1",
                "action": "checkout",
                "device_id": "ESP32-TEST",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_checkout.status_code == 200, scan_checkout.text
        assert scan_checkout.json()["outcome"] == "failed"
        assert scan_checkout.json()["led"] == "red_2s"
        assert "Check-in not found" in scan_checkout.json()["message"]

        with SessionLocal() as db:
            queued = db.execute(select(FormsSubmission).where(FormsSubmission.rfid == "CARDPROV1")).scalars().all()
            assert queued == []


def test_repeated_checkout_after_singapore_midnight_is_not_sent_to_forms_again(monkeypatch):
    monkeypatch.setattr(
        FormsWorker,
        "submit_with_retries",
        lambda self, action, chave, projeto, ontime=True, **_kwargs: {
            "success": True,
            "message": f"mocked {action}",
            "retry_count": 0,
        },
    )

    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "CARD2002", "nome": "Usuario Midnight", "chave": "EG58", "projeto": "P83"},
        )
        assert save_user.status_code == 200

        with SessionLocal() as db:
            user = get_user_by_rfid(db, "CARD2002")
            prior_checkout = now_sgt() - timedelta(days=1)
            user.checkin = False
            user.local = "co83-old"
            user.time = prior_checkout
            user.last_active_at = prior_checkout
            db.commit()

        scan_checkout = client.post(
            "/api/scan",
            json={
                "local": "co83-new",
                "rfid": "CARD2002",
                "action": "checkout",
                "device_id": "ESP32-TEST",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_checkout.status_code == 200
        assert scan_checkout.json()["outcome"] == "local_updated"
        assert scan_checkout.json()["led"] == "green_blink_3x_1s"

        with SessionLocal() as db:
            user = get_user_by_rfid(db, "CARD2002")
            queued = db.execute(select(FormsSubmission).where(FormsSubmission.rfid == "CARD2002")).scalars().all()

            assert user.checkin is False
            assert user.local == "co83-new"
            assert len(queued) == 1
            assert queued[0].status == "skipped"
            assert queued[0].display_status == "not_realized"


def test_forms_step_timeout_returns_red_blink_pattern(monkeypatch):
    monkeypatch.setattr(
        FormsWorker,
        "submit_with_retries",
        lambda self, action, chave, projeto, ontime=True, **_kwargs: {
            "success": False,
            "message": "Step 'digitar_chave' not found within 10 seconds",
            "retry_count": 0,
            "error_code": "forms_step_timeout",
            "failed_step": "digitar_chave",
            "audit_events": [
                {
                    "source": "forms",
                    "action": "forms",
                    "status": "failed",
                    "message": "Forms step timeout",
                    "details": "step=digitar_chave; timeout=10",
                }
            ],
        },
    )

    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "CARDFAIL", "nome": "Usuario Falha", "chave": "CD34", "projeto": "P80"},
        )
        assert save_user.status_code == 200

        scan_failed = client.post(
            "/api/scan",
            json={
                "local": "co80",
                "rfid": "CARDFAIL",
                "action": "checkin",
                "device_id": "ESP32-TEST",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_failed.status_code == 200
        assert scan_failed.json()["outcome"] == "submitted"
        assert scan_failed.json()["led"] == "green_1s"

        processed = process_forms_submission_queue_once()
        assert processed >= 1

        events = client.get("/api/admin/events")
        assert events.status_code == 200
        assert any(event["rfid"] == "CARDFAIL" and event["status"] == "failed" for event in events.json())
        assert any(event["source"] == "forms" and event["status"] == "failed" for event in events.json())
        assert all(
            not (event["source"] == "forms" and event["message"] == "Forms step timeout")
            for event in events.json()
        )


def test_valid_scan_updates_user_before_forms_processing(monkeypatch):
    monkeypatch.setattr(
        FormsWorker,
        "submit_with_retries",
        lambda self, action, chave, projeto, ontime=True, **_kwargs: {
            "success": True,
            "message": f"mocked {action}",
            "retry_count": 0,
        },
    )

    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "CARDFAST", "nome": "Usuario Rapido", "chave": "QP12", "projeto": "P80"},
        )
        assert save_user.status_code == 200

        scan_response = client.post(
            "/api/scan",
            json={
                "local": "main",
                "rfid": "CARDFAST",
                "action": "checkin",
                "device_id": "ESP32-FAST",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_response.status_code == 200
        assert scan_response.json()["outcome"] == "submitted"
        assert scan_response.json()["led"] == "green_1s"

        with SessionLocal() as db:
            user = get_user_by_rfid(db, "CARDFAST")
            queued = db.execute(select(FormsSubmission).where(FormsSubmission.rfid == "CARDFAST")).scalar_one()
            assert user.checkin is True
            assert user.local == "main"
            assert queued.status == "pending"


def test_forms_queue_processing_persists_failure_state(monkeypatch):
    monkeypatch.setattr(
        FormsWorker,
        "submit_with_retries",
        lambda self, action, chave, projeto, ontime=True, **_kwargs: {
            "success": False,
            "message": "mocked queue failure",
            "retry_count": 2,
            "error_code": "forms_runtime_error",
            "failed_step": "botao_enviar",
            "audit_events": [
                {
                    "source": "forms",
                    "action": "forms",
                    "status": "failed",
                    "message": "Forms runtime error",
                    "details": "mocked queue failure",
                }
            ],
        },
    )

    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "QUEUEFAIL", "nome": "Fila Falha", "chave": "LM34", "projeto": "P83"},
        )
        assert save_user.status_code == 200

        scan_response = client.post(
            "/api/scan",
            json={
                "local": "line-2",
                "rfid": "QUEUEFAIL",
                "action": "checkin",
                "device_id": "ESP32-QUEUE",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_response.status_code == 200
        assert scan_response.json()["outcome"] == "submitted"

        processed = process_forms_submission_queue_once()
        assert processed >= 1

        with SessionLocal() as db:
            queued = db.execute(select(FormsSubmission).where(FormsSubmission.rfid == "QUEUEFAIL")).scalar_one()
            assert queued.status == "failed"
            assert queued.retry_count == 2
            assert queued.last_error == "mocked queue failure"


def test_forms_queue_diagnostics_endpoint_reports_backlog_and_recent_processing_metrics():
    reference_time = now_sgt()

    with SessionLocal() as db:
        db.execute(delete(FormsSubmission))
        db.add_all(
            [
                FormsSubmission(
                    request_id=f"diag-pending-{uuid.uuid4().hex}",
                    rfid="DIAGP1",
                    action="checkin",
                    chave="DP01",
                    projeto="P80",
                    device_id="ESP32-DIAG",
                    local="main",
                    ontime=True,
                    status="pending",
                    retry_count=0,
                    last_error=None,
                    created_at=reference_time - timedelta(seconds=90),
                    updated_at=reference_time - timedelta(seconds=90),
                    processed_at=None,
                ),
                FormsSubmission(
                    request_id=f"diag-processing-{uuid.uuid4().hex}",
                    rfid="DIAGP2",
                    action="checkout",
                    chave="DP02",
                    projeto="P80",
                    device_id="ESP32-DIAG",
                    local="main",
                    ontime=True,
                    status="processing",
                    retry_count=1,
                    last_error=None,
                    created_at=reference_time - timedelta(seconds=30),
                    updated_at=reference_time - timedelta(seconds=5),
                    processed_at=None,
                ),
                FormsSubmission(
                    request_id=f"diag-success-{uuid.uuid4().hex}",
                    rfid="DIAGS1",
                    action="checkin",
                    chave="DS01",
                    projeto="P80",
                    device_id="ESP32-DIAG",
                    local="main",
                    ontime=True,
                    status="success",
                    retry_count=0,
                    last_error=None,
                    created_at=reference_time - timedelta(seconds=80),
                    updated_at=reference_time - timedelta(seconds=20),
                    processed_at=reference_time - timedelta(seconds=20),
                ),
                FormsSubmission(
                    request_id=f"diag-failed-{uuid.uuid4().hex}",
                    rfid="DIAGF1",
                    action="checkout",
                    chave="DF01",
                    projeto="P80",
                    device_id="ESP32-DIAG",
                    local="main",
                    ontime=True,
                    status="failed",
                    retry_count=2,
                    last_error="timeout",
                    created_at=reference_time - timedelta(seconds=40),
                    updated_at=reference_time - timedelta(seconds=10),
                    processed_at=reference_time - timedelta(seconds=10),
                ),
            ]
        )
        db.commit()

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.get("/api/admin/forms/queue/diagnostics")

    assert response.status_code == 200, response.text
    payload = response.json()
    assert payload["backlog_count"] == 2
    assert payload["pending_count"] == 1
    assert payload["processing_count"] == 1
    assert payload["success_count"] == 1
    assert payload["failed_count"] == 1
    assert payload["oldest_backlog_age_seconds"] >= 90
    assert payload["oldest_pending_age_seconds"] >= 90
    assert payload["oldest_processing_age_seconds"] >= 30
    assert payload["recent_average_processing_ms"] == 45000
    assert payload["recent_processed_sample_size"] == 2
    assert payload["worker"]["enabled"] is False
    assert payload["worker"]["running"] is False
    assert payload["worker"]["status"] == "stopped"
    assert payload["worker"]["poll_interval_seconds"] == 0.25


def test_forms_queue_processing_emits_structured_logs(monkeypatch, caplog: pytest.LogCaptureFixture):
    with SessionLocal() as db:
        db.execute(delete(FormsSubmission))
        db.add(
            FormsSubmission(
                request_id=f"diag-log-{uuid.uuid4().hex}",
                rfid="DIAGLOG1",
                action="checkin",
                chave="DL01",
                projeto="P80",
                device_id="ESP32-DIAG",
                local="main",
                ontime=True,
                status="pending",
                retry_count=0,
                last_error=None,
                created_at=now_sgt() - timedelta(seconds=5),
                updated_at=now_sgt() - timedelta(seconds=5),
                processed_at=None,
            )
        )
        db.commit()

    monkeypatch.setattr(
        FormsWorker,
        "submit_with_retries",
        lambda self, action, chave, projeto, ontime=True, **_kwargs: {
            "success": True,
            "message": "mocked queue success",
            "retry_count": 1,
            "audit_events": [
                {
                    "source": "forms",
                    "action": "forms",
                    "status": "completed",
                    "message": "Microsoft Forms completed",
                    "details": "steps=ok",
                }
            ],
        },
    )

    with caplog.at_level(logging.INFO, logger="checking.forms_queue"):
        caplog.clear()
        processed = process_forms_submission_queue_once(max_items=1)

    assert processed == 1
    queue_logs = get_forms_queue_logs(caplog)
    assert queue_logs

    reserved_log = next(payload for payload in queue_logs if payload["event"] == "forms_queue_reserved")
    processed_log = next(payload for payload in queue_logs if payload["event"] == "forms_queue_processed")
    assert reserved_log["status"] == "processing"
    assert isinstance(reserved_log["submission_id"], int)
    assert processed_log["status"] == "success"
    assert processed_log["action"] == "checkin"
    assert processed_log["retry_count"] == 1
    assert processed_log["error_code"] is None
    assert processed_log["turnaround_ms"] >= 0


def test_http_app_lifespan_does_not_start_forms_worker(monkeypatch):
    start_calls: list[str] = []
    stop_calls: list[str] = []

    monkeypatch.setattr(
        "sistema.app.services.forms_queue.forms_submission_worker.start",
        lambda: start_calls.append("start"),
    )
    monkeypatch.setattr(
        "sistema.app.services.forms_queue.forms_submission_worker.stop",
        lambda: stop_calls.append("stop"),
    )

    with TestClient(app) as client:
        response = client.get("/api/health")

    assert response.status_code == 200, response.text
    assert start_calls == []
    assert stop_calls == []


def test_http_app_lifespan_starts_and_stops_realtime_brokers(monkeypatch):
    lifecycle_calls: list[str] = []

    monkeypatch.setattr("sistema.app.main.start_realtime_brokers", lambda: lifecycle_calls.append("start"))
    monkeypatch.setattr("sistema.app.main.stop_realtime_brokers", lambda: lifecycle_calls.append("stop"))

    with TestClient(app) as client:
        response = client.get("/api/health")

    assert response.status_code == 200, response.text
    assert lifecycle_calls == ["start", "stop"]


def test_forms_queue_reservation_retries_when_candidate_was_claimed(monkeypatch):
    selected_rows = iter(
        [
            SimpleNamespace(id=101, request_id="forms-race-1"),
            SimpleNamespace(id=202, request_id="forms-race-2"),
        ]
    )
    claim_calls: list[int] = []

    monkeypatch.setattr(
        forms_queue_module,
        "_select_next_pending_submission_row",
        lambda db: next(selected_rows, None),
    )

    def fake_claim(db, *, submission_id: int, updated_at):
        claim_calls.append(submission_id)
        return submission_id == 202

    monkeypatch.setattr(forms_queue_module, "_claim_submission_for_processing", fake_claim)

    reserved_id = forms_queue_module._reserve_next_submission_id()

    assert reserved_id == 202
    assert claim_calls == [101, 202]


def test_forms_queue_diagnostics_uses_persisted_worker_health_snapshot(monkeypatch, tmp_path: Path):
    monkeypatch.setattr(settings, "event_archives_dir", str(tmp_path))
    reference_time = now_sgt()
    forms_queue_module._write_forms_worker_health_snapshot(
        {
            "enabled": True,
            "running": True,
            "status": "running",
            "thread_name": "forms-submission-worker",
            "process_id": 4321,
            "started_at": reference_time - timedelta(seconds=30),
            "last_heartbeat_at": reference_time,
            "last_loop_started_at": reference_time - timedelta(seconds=2),
            "last_loop_completed_at": reference_time - timedelta(seconds=1),
            "last_loop_processed_count": 4,
            "consecutive_error_count": 0,
            "current_backoff_seconds": 0.0,
            "restart_count": 2,
            "last_error": None,
        }
    )

    with SessionLocal() as db:
        diagnostics = forms_queue_module.get_forms_queue_diagnostics(db=db)

    assert diagnostics["worker"]["enabled"] is True
    assert diagnostics["worker"]["running"] is True
    assert diagnostics["worker"]["status"] == "running"
    assert diagnostics["worker"]["process_id"] == 4321
    assert diagnostics["worker"]["restart_count"] == 2
    assert diagnostics["worker"]["heartbeat_age_seconds"] == 0
    assert diagnostics["worker"]["stale"] is False


def test_forms_worker_healthcheck_reports_stale_snapshot_as_unhealthy(monkeypatch, tmp_path: Path, capsys: pytest.CaptureFixture[str]):
    monkeypatch.setattr(settings, "event_archives_dir", str(tmp_path))
    stale_time = now_sgt() - timedelta(seconds=settings.forms_worker_health_stale_seconds + 5)
    forms_queue_module._write_forms_worker_health_snapshot(
        {
            "enabled": True,
            "running": True,
            "status": "running",
            "thread_name": "forms-submission-worker",
            "process_id": 999,
            "started_at": stale_time - timedelta(seconds=10),
            "last_heartbeat_at": stale_time,
            "last_loop_started_at": stale_time,
            "last_loop_completed_at": stale_time,
            "last_loop_processed_count": 0,
            "consecutive_error_count": 0,
            "current_backoff_seconds": 0.0,
            "restart_count": 0,
            "last_error": None,
        }
    )

    exit_code = forms_worker_healthcheck_module.main()

    assert exit_code == 1
    payload = json.loads(capsys.readouterr().out)
    assert payload["status"] == "unhealthy"
    assert payload["reason"] == "forms worker heartbeat stale"


def test_forms_worker_error_backoff_grows_exponentially_and_caps():
    assert forms_queue_module._compute_exponential_backoff_seconds(base_seconds=1.0, max_seconds=15.0, attempt=1) == 1.0
    assert forms_queue_module._compute_exponential_backoff_seconds(base_seconds=1.0, max_seconds=15.0, attempt=2) == 2.0
    assert forms_queue_module._compute_exponential_backoff_seconds(base_seconds=1.0, max_seconds=15.0, attempt=4) == 8.0
    assert forms_queue_module._compute_exponential_backoff_seconds(base_seconds=1.0, max_seconds=15.0, attempt=6) == 15.0


def test_run_forms_submission_worker_forever_restarts_after_unexpected_thread_exit(monkeypatch, tmp_path: Path):
    monkeypatch.setattr(settings, "event_archives_dir", str(tmp_path))

    class DeadThread:
        name = "forms-submission-worker"

        def is_alive(self) -> bool:
            return False

    class FakeWorker:
        def __init__(self) -> None:
            self._stop_event = threading.Event()
            self._thread = None
            self._consumer_threads: list = []  # NOVO
            self.start_calls = 0
            self.stop_calls = 0
            self.backoff_waits: list[float] = []

        def start(self) -> None:
            self.start_calls += 1
            self._thread = DeadThread()
            self._consumer_threads = [DeadThread()]  # NOVO
            if self.start_calls >= 2:
                self._stop_event.set()

        def stop(self) -> None:
            self.stop_calls += 1
            self._stop_event.set()
            self._thread = None
            self._consumer_threads = []  # NOVO

        def stop_requested(self) -> bool:
            return self._stop_event.is_set()

        def has_alive_consumers(self) -> bool:  # NOVO
            return False

        def mark_supervisor_restart_wait(self, *, backoff_seconds: float) -> None:
            self.backoff_waits.append(backoff_seconds)

        def snapshot(self) -> dict[str, object]:
            return {
                "running": False,
                "status": "stopped",
                "thread_name": "forms-submission-worker",
                "started_at": now_sgt(),
                "last_loop_started_at": None,
                "last_loop_completed_at": None,
                "last_loop_processed_count": 0,
                "consecutive_error_count": 0,
                "current_backoff_seconds": 0.0,
                "restart_count": max(self.start_calls - 1, 0),
                "last_error": "thread exited unexpectedly",
                "concurrency": 1,              # NOVO
                "consumer_threads_alive": 0,   # NOVO
            }

    fake_worker = FakeWorker()
    monkeypatch.setattr(forms_queue_module, "forms_submission_worker", fake_worker)
    monkeypatch.setattr(
        forms_queue_module,
        "_compute_exponential_backoff_seconds",
        lambda **kwargs: 0.0,
    )

    forms_queue_module.run_forms_submission_worker_forever()

    assert fake_worker.start_calls >= 2
    assert fake_worker.stop_calls == 1
    assert fake_worker.backoff_waits == [0.0]


def test_forms_backlog_pressure_keeps_http_routes_responsive():
    backlog_size = 20
    processing_delay_seconds = 0.15
    web_key = make_test_key("W")
    mobile_key = make_test_key("M")
    web_password = "iso1234"

    ensure_web_user_exists(chave=web_key, projeto="P80", nome="Isolacao Web")
    ensure_web_user_exists(chave=mobile_key, projeto="P80", nome="Isolacao Mobile")

    with TestClient(app) as client:
        register_response = register_web_password(
            client,
            chave=web_key,
            senha=web_password,
            projeto="P80",
            ensure_user_exists=False,
        )
        assert register_response.status_code == 200, register_response.text

    with SessionLocal() as db:
        db.execute(delete(FormsSubmission))
        reference_time = now_sgt()
        db.add_all(
            [
                FormsSubmission(
                    request_id=f"forms-isolation-{index}-{uuid.uuid4().hex}",
                    rfid=f"ISO{index:03d}",
                    action="checkin" if index % 2 == 0 else "checkout",
                    chave=f"Q{index:03d}",
                    projeto="P80",
                    device_id="ESP32-ISOLATION",
                    local="controlled-backlog",
                    ontime=True,
                    status="pending",
                    retry_count=0,
                    last_error=None,
                    created_at=reference_time - timedelta(seconds=backlog_size - index),
                    updated_at=reference_time - timedelta(seconds=backlog_size - index),
                    processed_at=None,
                )
                for index in range(backlog_size)
            ]
        )
        db.commit()
        initial_diagnostics = forms_queue_module.get_forms_queue_diagnostics(db=db)

    assert initial_diagnostics["backlog_count"] == backlog_size
    assert initial_diagnostics["pending_count"] == backlog_size

    worker_process: subprocess.Popen[str] | None = None
    worker_stdout = ""
    worker_stderr = ""
    try:
        with live_app_server() as base_url:
            worker_process = start_controlled_slow_forms_worker_subprocess(
                processing_delay_seconds=processing_delay_seconds,
            )

            def read_diagnostics() -> dict[str, object]:
                with SessionLocal() as db:
                    return forms_queue_module.get_forms_queue_diagnostics(db=db)

            pressure_diagnostics = wait_for_condition(
                lambda: (
                    snapshot
                    if (snapshot := read_diagnostics())["processing_count"] >= 1
                    and snapshot["backlog_count"] >= backlog_size - 2
                    else None
                ),
                timeout_seconds=10,
                description="forms backlog pressure to begin",
            )

            route_samples: dict[str, list[int]] = {
                "admin_checkin": [],
                "admin_login": [],
                "admin_projects": [],
                "health": [],
                "mobile_state": [],
                "web_login": [],
                "web_state": [],
            }

            for _ in range(5):
                health_status, health_payload, health_latency = perform_live_json_request(base_url, "/api/health")
                assert health_status == 200
                assert health_payload["status"] == "ok"
                route_samples["health"].append(health_latency)

                web_opener = build_cookie_opener()
                web_login_status, web_login_payload, web_login_latency = perform_live_json_request(
                    base_url,
                    "/api/web/auth/login",
                    method="POST",
                    payload={"chave": web_key, "senha": web_password},
                    opener=web_opener,
                )
                assert web_login_status == 200
                assert web_login_payload["authenticated"] is True
                route_samples["web_login"].append(web_login_latency)

                web_state_status, web_state_payload, web_state_latency = perform_live_json_request(
                    base_url,
                    "/api/web/check/state",
                    query={"chave": web_key},
                    opener=web_opener,
                )
                assert web_state_status == 200
                assert web_state_payload["chave"] == web_key
                route_samples["web_state"].append(web_state_latency)

                admin_opener = build_cookie_opener()
                admin_login_status, admin_login_payload, admin_login_latency = perform_live_json_request(
                    base_url,
                    "/api/admin/auth/login",
                    method="POST",
                    payload={"chave": ADMIN_LOGIN_CHAVE, "senha": ADMIN_LOGIN_SENHA},
                    opener=admin_opener,
                )
                assert admin_login_status == 200
                assert admin_login_payload["ok"] is True
                route_samples["admin_login"].append(admin_login_latency)

                admin_projects_status, admin_projects_payload, admin_projects_latency = perform_live_json_request(
                    base_url,
                    "/api/admin/projects",
                    opener=admin_opener,
                )
                assert admin_projects_status == 200
                assert isinstance(admin_projects_payload, list)
                route_samples["admin_projects"].append(admin_projects_latency)

                admin_checkin_status, admin_checkin_payload, admin_checkin_latency = perform_live_json_request(
                    base_url,
                    "/api/admin/checkin",
                    opener=admin_opener,
                )
                assert admin_checkin_status == 200
                assert isinstance(admin_checkin_payload, list)
                route_samples["admin_checkin"].append(admin_checkin_latency)

                mobile_state_status, mobile_state_payload, mobile_state_latency = perform_live_json_request(
                    base_url,
                    "/api/mobile/state",
                    query={"chave": mobile_key},
                    headers=MOBILE_HEADERS,
                )
                assert mobile_state_status == 200
                assert mobile_state_payload["chave"] == mobile_key
                route_samples["mobile_state"].append(mobile_state_latency)

            during_diagnostics = read_diagnostics()
            drained_diagnostics = wait_for_condition(
                lambda: (
                    snapshot
                    if (snapshot := read_diagnostics())["backlog_count"] == 0
                    else None
                ),
                timeout_seconds=20,
                description="forms backlog drain to finish",
            )

        if worker_process is not None:
            worker_stdout, worker_stderr = worker_process.communicate(timeout=10)
    finally:
        if worker_process is not None and worker_process.poll() is None:
            worker_process.terminate()
            worker_stdout, worker_stderr = worker_process.communicate(timeout=10)
        with SessionLocal() as db:
            db.execute(delete(FormsSubmission))
            db.commit()

    assert worker_process is not None
    assert worker_process.returncode == 0, worker_stderr or worker_stdout
    assert pressure_diagnostics["processing_count"] >= 1
    assert during_diagnostics["backlog_count"] > 0
    assert drained_diagnostics["backlog_count"] == 0

    latency_summary = {route_name: summarize_latency_samples(samples) for route_name, samples in route_samples.items()}
    latency_budgets_ms = {
        "admin_checkin": 1500,
        "admin_login": 1500,
        "admin_projects": 1500,
        "health": 500,
        "mobile_state": 1000,
        "web_login": 1500,
        "web_state": 1000,
    }
    for route_name, budget_ms in latency_budgets_ms.items():
        assert latency_summary[route_name]["max_ms"] < budget_ms, {
            "route": route_name,
            "budget_ms": budget_ms,
            **latency_summary[route_name],
        }

    print(
        json.dumps(
            {
                "event": "forms_isolation_validation",
                "initial_backlog": initial_diagnostics["backlog_count"],
                "pressure_backlog": pressure_diagnostics["backlog_count"],
                "pressure_processing": pressure_diagnostics["processing_count"],
                "backlog_after_requests": during_diagnostics["backlog_count"],
                "backlog_after_drain": drained_diagnostics["backlog_count"],
                "latency_summary": latency_summary,
            },
            separators=(",", ":"),
            sort_keys=True,
        )
    )


def test_database_diagnostics_endpoint_reports_query_and_pool_metrics():
    with TestClient(app) as client:
        ensure_admin_session(client)

        checkin_response = client.get("/api/admin/checkin")
        assert checkin_response.status_code == 200, checkin_response.text

        diagnostics_response = client.get("/api/admin/diagnostics/database")
        assert diagnostics_response.status_code == 200, diagnostics_response.text

    payload = diagnostics_response.json()
    assert payload["pool"]["dialect"] == "sqlite"
    assert payload["pool"]["pool_class"]
    assert payload["pool"]["configured_pool_timeout_seconds"] is None
    assert payload["pool"]["configured_pool_recycle_seconds"] is None
    assert payload["pool"]["pool_pre_ping"] is True
    assert payload["pool"]["checked_out"] >= 1
    assert payload["pool"]["total_connect_events"] >= 1
    assert payload["latency"]["query_count_total"] >= 1
    assert payload["latency"]["recent_query_sample_size"] >= 1
    assert payload["latency"]["recent_average_query_ms"] is not None
    hot_path = next(item for item in payload["latency"]["hot_paths"] if item["path"] == "/api/admin/checkin")
    assert hot_path["recent_query_count"] >= 1
    assert hot_path["total_query_count"] >= 1
    assert payload["server_connections"]["source"] == "unsupported"
    assert payload["recommended_alert_thresholds"] == {
        "pool_usage_warning_ratio": 0.8,
        "pool_usage_critical_ratio": 1.0,
        "recent_query_p95_warning_ms": 150,
        "recent_query_p95_critical_ms": 300,
        "slow_query_log_threshold_ms": 250,
        "postgres_active_connections_warning": 24,
        "postgres_active_connections_critical": 32,
        "postgres_waiting_connections_warning": 1,
        "postgres_waiting_connections_critical": 3,
        "postgres_idle_in_transaction_warning": 1,
    }


def test_database_engine_kwargs_apply_explicit_queue_pool_settings_for_postgres():
    pool_config = database_module.resolve_database_pool_config(
        database_url="postgresql+psycopg://postgres:postgres@localhost:5432/checking",
        pool_size=6,
        max_overflow=2,
        pool_timeout_seconds=5,
        pool_recycle_seconds=1800,
    )

    sample_engine = database_module.create_engine(
        "postgresql+psycopg://postgres:postgres@localhost:5432/checking",
        **database_module.build_database_engine_kwargs(pool_config),
    )
    try:
        assert type(sample_engine.pool).__name__ == "QueuePool"
        assert getattr(sample_engine.pool, "_pool").maxsize == 6
        assert getattr(sample_engine.pool, "_max_overflow") == 2
        assert getattr(sample_engine.pool, "_timeout") == 5
        assert getattr(sample_engine.pool, "_recycle") == 1800
    finally:
        sample_engine.dispose()


def test_database_engine_kwargs_leave_sqlite_pool_defaults():
    pool_config = database_module.resolve_database_pool_config(
        database_url="sqlite:///./checking.db",
        pool_size=6,
        max_overflow=2,
        pool_timeout_seconds=5,
        pool_recycle_seconds=1800,
    )

    assert database_module.build_database_engine_kwargs(pool_config) == {
        "pool_pre_ping": True,
    }


def test_slow_database_queries_emit_structured_logs(monkeypatch, caplog: pytest.LogCaptureFixture):
    monkeypatch.setattr(database_module, "DATABASE_SLOW_QUERY_LOG_THRESHOLD_MS", 0)

    with caplog.at_level(logging.WARNING, logger="checking.db"):
        with TestClient(app) as client:
            ensure_admin_session(client)
            caplog.clear()
            checkin_response = client.get("/api/admin/checkin")
            assert checkin_response.status_code == 200, checkin_response.text

    database_logs = get_database_logs(caplog)
    assert database_logs
    slow_query_log = next(payload for payload in database_logs if payload["event"] == "db_query_slow")
    assert slow_query_log["path"] == "/api/admin/checkin"
    assert slow_query_log["request_id"]
    assert slow_query_log["database_dialect"] == "sqlite"
    assert slow_query_log["latency_ms"] >= 0
    assert slow_query_log["sql_operation"] == "SELECT"
    assert slow_query_log["failed"] is False


def test_forms_success_generates_single_final_forms_event(monkeypatch):
    monkeypatch.setattr(
        FormsWorker,
        "submit_with_retries",
        lambda self, action, chave, projeto, ontime=True, **_kwargs: {
            "success": True,
            "message": "Form submitted successfully",
            "retry_count": 0,
            "audit_events": [
                {
                    "source": "forms",
                    "action": "forms",
                    "status": "opened",
                    "message": "Microsoft Forms opened",
                    "details": None,
                },
                {
                    "source": "forms",
                    "action": "forms",
                    "status": "completed",
                    "message": "Microsoft Forms completed",
                    "details": "steps=ok; success_xpath_visible=true",
                },
            ],
        },
    )

    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "FORMSOK", "nome": "Forms Final", "chave": "ZX12", "projeto": "P80"},
        )
        assert save_user.status_code == 200

        scan_response = client.post(
            "/api/scan",
            json={
                "local": "main",
                "rfid": "FORMSOK",
                "action": "checkin",
                "device_id": "ESP32-FORMS",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_response.status_code == 200

        processed = process_forms_submission_queue_once()
        assert processed >= 1

        events = client.get("/api/admin/events")
        assert events.status_code == 200
        forms_events = [event for event in events.json() if event["rfid"] == "FORMSOK" and event["source"] == "forms"]
        assert len(forms_events) == 1
        assert forms_events[0]["status"] == "success"
        assert "forms_details=steps=ok; success_xpath_visible=true" in (forms_events[0]["details"] or "")


def test_forms_worker_requires_success_xpath_after_submit(tmp_path, monkeypatch):
    xpath_dir = tmp_path / "xpath"
    xpath_dir.mkdir(parents=True)
    xpaths = {
        "digitar_chave.txt": "//digitar_chave",
        "confirmar_chave.txt": "//confirmar_chave",
        "botao_normal.txt": "//botao_normal",
        "botao_retroativo.txt": "//botao_retroativo",
        "botao_checkin.txt": "//botao_checkin",
        "botao_checkout.txt": "//botao_checkout",
        "botao_enviar.txt": "//botao_enviar",
        "sucesso.txt": "//sucesso",
        "botao_projeto_P80.txt": "//botao_projeto_P80",
        "botao_projeto_P82.txt": "//botao_projeto_P82",
        "botao_projeto_P83.txt": "//botao_projeto_P83",
    }
    for name, content in xpaths.items():
        (xpath_dir / name).write_text(content, encoding="utf-8")

    send_selector = "xpath=//botao_enviar"
    success_selector = "xpath=//sucesso"

    class FakeLocator:
        def __init__(self, page, selector: str):
            self.page = page
            self.selector = selector

        def fill(self, value: str) -> None:
            self.page.filled[self.selector] = value

        def click(self) -> None:
            self.page.clicked.append(self.selector)
            self.page.checked.add(self.selector)
            if self.selector == send_selector:
                self.page.success_visible = True
                self.page.url = f"{self.page.url}#submitted"

        def input_value(self) -> str:
            return self.page.filled.get(self.selector, "")

        def is_checked(self) -> bool:
            return self.selector in self.page.checked

        def inner_text(self) -> str:
            if self.selector == success_selector:
                return "Sua resposta foi enviada."
            return ""

    class FakePage:
        def __init__(self):
            self.url = ""
            self.success_visible = False
            self.filled = {}
            self.clicked = []
            self.checked = set()
            self.visible_selectors = {
                "xpath=//digitar_chave",
                "xpath=//confirmar_chave",
                "xpath=//botao_normal",
                "xpath=//botao_retroativo",
                "xpath=//botao_checkin",
                "xpath=//botao_checkout",
                send_selector,
                "xpath=//botao_projeto_P80",
                "xpath=//botao_projeto_P82",
                "xpath=//botao_projeto_P83",
            }

        def goto(self, url: str, timeout: int) -> None:
            self.url = url

        def wait_for_selector(self, selector: str, state: str = "visible", timeout: int = 0):
            if selector == success_selector and self.success_visible:
                return True
            if selector in self.visible_selectors:
                return True
            raise forms_worker_module.PlaywrightTimeoutError("timeout")

        def locator(self, selector: str) -> FakeLocator:
            return FakeLocator(self, selector)

    class FakeBrowser:
        def __init__(self, page: FakePage):
            self.page = page

        def new_page(self) -> FakePage:
            return self.page

        def close(self) -> None:
            return None

    class FakePlaywright:
        def __init__(self, page: FakePage):
            self.chromium = SimpleNamespace(launch=lambda headless=True: FakeBrowser(page))

    class FakePlaywrightContext:
        def __init__(self, page: FakePage):
            self.page = page

        def __enter__(self):
            return FakePlaywright(self.page)

        def __exit__(self, exc_type, exc, tb):
            return False

    fake_page = FakePage()
    monkeypatch.setattr(forms_worker_module, "sync_playwright", lambda: FakePlaywrightContext(fake_page))

    worker = FormsWorker(assets_dir=tmp_path)
    result = worker.submit_with_retries(action="checkin", chave="HR70", projeto="P80")

    assert result["success"] is True
    completed_event = next(event for event in result["audit_events"] if event["status"] == "completed")
    assert "steps=digitar_chave:filled+verified,confirmar_chave:filled+verified,botao_normal:clicked+verified,botao_checkin:clicked+verified,botao_projeto_P80:clicked+verified,botao_enviar:clicked,sucesso:visible" in completed_event["details"]
    assert "success_xpath_visible=true" in completed_event["details"]
    assert "submit_to_success_ms=" in completed_event["details"]
    assert "success_text=Sua resposta foi enviada." in completed_event["details"]


def test_forms_worker_rejects_success_xpath_visible_before_submit(tmp_path, monkeypatch):
    xpath_dir = tmp_path / "xpath"
    xpath_dir.mkdir(parents=True)
    for name in [
        "digitar_chave.txt",
        "confirmar_chave.txt",
        "botao_normal.txt",
        "botao_retroativo.txt",
        "botao_checkin.txt",
        "botao_checkout.txt",
        "botao_enviar.txt",
        "sucesso.txt",
        "botao_projeto_P80.txt",
            "botao_projeto_P82.txt",
        "botao_projeto_P83.txt",
    ]:
        (xpath_dir / name).write_text(f"//{name}", encoding="utf-8")

    success_selector = "xpath=//sucesso.txt"

    class FakeLocator:
        def __init__(self, selector: str):
            self.selector = selector

        def fill(self, value: str) -> None:
            return None

        def click(self) -> None:
            return None

        def input_value(self) -> str:
            return "HR70"

        def is_checked(self) -> bool:
            return True

        def inner_text(self) -> str:
            return ""

    class FakePage:
        url = "https://example.com/form"

        def goto(self, url: str, timeout: int) -> None:
            self.url = url

        def wait_for_selector(self, selector: str, state: str = "visible", timeout: int = 0):
            if selector == success_selector:
                return True
            return True

        def locator(self, selector: str) -> FakeLocator:
            return FakeLocator(selector)

    class FakeBrowser:
        def new_page(self) -> FakePage:
            return FakePage()

        def close(self) -> None:
            return None

    class FakePlaywright:
        chromium = SimpleNamespace(launch=lambda headless=True: FakeBrowser())

    class FakePlaywrightContext:
        def __enter__(self):
            return FakePlaywright()

        def __exit__(self, exc_type, exc, tb):
            return False

    monkeypatch.setattr(forms_worker_module, "sync_playwright", lambda: FakePlaywrightContext())

    worker = FormsWorker(assets_dir=tmp_path)
    result = worker.submit_with_retries(action="checkin", chave="HR70", projeto="P80")

    assert result["success"] is False
    assert result["error_code"] == "forms_validation_error"
    assert "XPath de sucesso ja estava visivel antes do envio" in result["message"]


def test_forms_worker_fails_when_first_field_is_not_confirmed(tmp_path, monkeypatch):
    xpath_dir = tmp_path / "xpath"
    xpath_dir.mkdir(parents=True)
    xpaths = {
        "digitar_chave.txt": "//digitar_chave",
        "confirmar_chave.txt": "//confirmar_chave",
        "botao_normal.txt": "//botao_normal",
        "botao_retroativo.txt": "//botao_retroativo",
        "botao_checkin.txt": "//botao_checkin",
        "botao_checkout.txt": "//botao_checkout",
        "botao_enviar.txt": "//botao_enviar",
        "sucesso.txt": "//sucesso",
        "botao_projeto_P80.txt": "//botao_projeto_P80",
        "botao_projeto_P82.txt": "//botao_projeto_P82",
        "botao_projeto_P83.txt": "//botao_projeto_P83",
    }
    for name, content in xpaths.items():
        (xpath_dir / name).write_text(content, encoding="utf-8")

    class FakeLocator:
        def __init__(self, selector: str):
            self.selector = selector

        def fill(self, value: str) -> None:
            return None

        def click(self) -> None:
            return None

        def input_value(self) -> str:
            return ""

        def is_checked(self) -> bool:
            return False

        def inner_text(self) -> str:
            return ""

    class FakePage:
        url = "https://example.com/form"

        def goto(self, url: str, timeout: int) -> None:
            self.url = url

        def wait_for_selector(self, selector: str, state: str = "visible", timeout: int = 0):
            return True

        def locator(self, selector: str) -> FakeLocator:
            return FakeLocator(selector)

    class FakeBrowser:
        def new_page(self) -> FakePage:
            return FakePage()

        def close(self) -> None:
            return None

    class FakePlaywright:
        chromium = SimpleNamespace(launch=lambda headless=True: FakeBrowser())

    class FakePlaywrightContext:
        def __enter__(self):
            return FakePlaywright()

        def __exit__(self, exc_type, exc, tb):
            return False

    monkeypatch.setattr(forms_worker_module, "sync_playwright", lambda: FakePlaywrightContext())

    worker = FormsWorker(assets_dir=tmp_path)
    result = worker.submit_with_retries(action="checkin", chave="HR70", projeto="P80")

    assert result["success"] is False
    assert result["error_code"] == "forms_step_validation_failed"
    assert result["failed_step"] == "digitar_chave"
    assert "expected_value=HR70" in result["audit_events"][0]["details"]


def test_heartbeat_success_is_not_logged_in_events():
    with TestClient(app) as client:
        ensure_admin_session(client)
        heartbeat = client.post(
            "/api/device/heartbeat",
            json={"device_id": "ESP32-TEST", "shared_key": "device-test-key"},
        )
        assert heartbeat.status_code == 200

        events = client.get("/api/admin/events")
        assert events.status_code == 200
        assert all(
            not (
                event["action"] == "heartbeat"
                and event["status"] == "success"
                and event["device_id"] == "ESP32-TEST"
            )
            for event in events.json()
        )


def test_heartbeat_failure_is_logged_in_events():
    with TestClient(app) as client:
        ensure_admin_session(client)
        heartbeat = client.post(
            "/api/device/heartbeat",
            json={"device_id": "ESP32-BAD", "shared_key": "wrong-key"},
        )
        assert heartbeat.status_code == 200
        assert heartbeat.json()["ok"] is False

        events = client.get("/api/admin/events")
        assert events.status_code == 200
        assert any(
            event["action"] == "heartbeat"
            and event["status"] == "failed"
            and event["device_id"] == "ESP32-BAD"
            for event in events.json()
        )


def test_remove_pending_registration():
    with TestClient(app) as client:
        ensure_admin_session(client)
        scan_unknown = client.post(
            "/api/scan",
            json={
                "local": "main",
                "rfid": "PENDING01",
                "action": "checkin",
                "device_id": "ESP32-TEST",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_unknown.status_code == 200

        pending_list = client.get("/api/admin/pending")
        assert pending_list.status_code == 200
        pending_id = next(row["id"] for row in pending_list.json() if row["rfid"] == "PENDING01")

        remove_res = client.delete(f"/api/admin/pending/{pending_id}")
        assert remove_res.status_code == 200

        pending_list_after = client.get("/api/admin/pending")
        assert pending_list_after.status_code == 200
        assert all(row["id"] != pending_id for row in pending_list_after.json())


def test_list_and_remove_registered_user():
    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "USERDEL1", "nome": "Usuario Cadastro", "chave": "GH78", "projeto": "P83"},
        )
        assert save_user.status_code == 200

        checkout_before = client.get("/api/admin/checkout")
        assert checkout_before.status_code == 200
        assert all(row["rfid"] != "USERDEL1" for row in checkout_before.json())

        checkin_before = client.get("/api/admin/checkin")
        assert checkin_before.status_code == 200
        assert all(row["rfid"] != "USERDEL1" for row in checkin_before.json())

        events_before = client.get("/api/admin/events")
        assert events_before.status_code == 200
        assert any(event["rfid"] == "USERDEL1" for event in events_before.json())

        users_before = client.get("/api/admin/users")
        assert users_before.status_code == 200
        assert any(row["rfid"] == "USERDEL1" and row["nome"] == "Usuario Cadastro" for row in users_before.json())

        user_id = next(row["id"] for row in users_before.json() if row["rfid"] == "USERDEL1")

        remove_user = client.delete(f"/api/admin/users/{user_id}")
        assert remove_user.status_code == 200

        users_after = client.get("/api/admin/users")
        assert users_after.status_code == 200
        assert all(row["rfid"] != "USERDEL1" for row in users_after.json())

        checkout_after = client.get("/api/admin/checkout")
        assert checkout_after.status_code == 200
        assert all(row["rfid"] != "USERDEL1" for row in checkout_after.json())

        checkin_after = client.get("/api/admin/checkin")
        assert checkin_after.status_code == 200
        assert all(row["rfid"] != "USERDEL1" for row in checkin_after.json())

        events_after = client.get("/api/admin/events")
        assert events_after.status_code == 200
        assert any(event["rfid"] == "USERDEL1" for event in events_after.json())


def test_presence_moves_users_between_checkin_checkout_and_inactive_after_24_hours(monkeypatch):
    fixed_now = datetime(2024, 4, 8, 12, 0, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(admin_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(user_activity_module, "now_sgt", lambda: fixed_now)

    with TestClient(app) as client:
        ensure_admin_session(client)
        save_a = client.post(
            "/api/admin/users",
            json={"rfid": "INA001", "nome": "Zelda Ativa", "chave": "AA11", "projeto": "P80"},
        )
        save_b = client.post(
            "/api/admin/users",
            json={"rfid": "INA002", "nome": "Ana Inativa", "chave": "BB22", "projeto": "P83"},
        )
        save_c = client.post(
            "/api/admin/users",
            json={"rfid": "INA003", "nome": "Bruno Inativo", "chave": "CC33", "projeto": "P80"},
        )
        save_d = client.post(
            "/api/admin/users",
            json={"rfid": "INA004", "nome": "Carlos Checkout", "chave": "DD55", "projeto": "P82"},
        )
        save_e = client.post(
            "/api/admin/users",
            json={"rfid": "INA005", "nome": "Eva Sem Checkout", "chave": "EE66", "projeto": "P83"},
        )
        assert save_a.status_code == 200
        assert save_b.status_code == 200
        assert save_c.status_code == 200
        assert save_d.status_code == 200
        assert save_e.status_code == 200

        with SessionLocal() as db:
            user_active = get_user_by_rfid(db, "INA001")
            user_active.checkin = True
            user_active.local = "main"
            user_active.time = fixed_now - timedelta(hours=3)
            user_active.last_active_at = fixed_now - timedelta(hours=3)
            user_active.inactivity_days = 0

            user_two = get_user_by_rfid(db, "INA002")
            user_two.checkin = True
            user_two.local = "co83"
            user_two.time = fixed_now - timedelta(hours=25)
            user_two.last_active_at = user_two.time
            user_two.inactivity_days = 0

            user_three = get_user_by_rfid(db, "INA003")
            user_three.checkin = False
            user_three.local = "main"
            user_three.time = fixed_now - timedelta(hours=27)
            user_three.last_active_at = user_three.time
            user_three.inactivity_days = 0

            user_four = get_user_by_rfid(db, "INA004")
            user_four.checkin = False
            user_four.local = "co80"
            user_four.time = fixed_now - timedelta(hours=8)
            user_four.last_active_at = user_four.time
            user_four.inactivity_days = 0

            user_five = get_user_by_rfid(db, "INA005")
            user_five.checkin = True
            user_five.local = "un83"
            user_five.time = fixed_now - timedelta(hours=23, minutes=30)
            user_five.last_active_at = user_five.time
            user_five.inactivity_days = 0
            db.commit()

        inactive_rows = client.get("/api/admin/inactive")
        assert inactive_rows.status_code == 200
        inactive_payload = inactive_rows.json()
        assert all(isinstance(row["id"], int) and row["id"] > 0 for row in inactive_payload)
        assert [row["nome"] for row in inactive_payload] == ["Ana Inativa", "Bruno Inativo"]
        assert [row["projetos"] for row in inactive_payload] == [["P83"], ["P80"]]
        assert [row["inactivity_days"] for row in inactive_payload] == [1, 1]
        assert [row["latest_action"] for row in inactive_payload] == ["checkin", "checkout"]
        assert inactive_payload[0]["latest_time"].startswith("2024-04-07T11:00:00")
        assert inactive_payload[1]["latest_time"].startswith("2024-04-07T09:00:00")

        checkin_rows = client.get("/api/admin/checkin")
        assert checkin_rows.status_code == 200
        checkin_payload = checkin_rows.json()
        checkin_rows_by_rfid = {row["rfid"]: row for row in checkin_payload}
        assert any(row["rfid"] == "INA001" and row["id"] > 0 for row in checkin_payload)
        assert all(row["rfid"] != "INA002" for row in checkin_payload)
        assert any(row["rfid"] == "INA005" and row["id"] > 0 for row in checkin_payload)
        assert checkin_rows_by_rfid["INA001"]["projetos"] == ["P80"]
        assert checkin_rows_by_rfid["INA005"]["projetos"] == ["P83"]

        missing_checkout_rows = client.get("/api/admin/missing-checkout")
        assert missing_checkout_rows.status_code == 200
        assert missing_checkout_rows.json() == []

        checkout_rows = client.get("/api/admin/checkout")
        assert checkout_rows.status_code == 200
        checkout_payload = checkout_rows.json()
        checkout_rows_by_rfid = {row["rfid"]: row for row in checkout_payload}
        assert all(row["rfid"] != "INA003" for row in checkout_payload)
        assert any(row["rfid"] == "INA004" and row["id"] > 0 for row in checkout_payload)
        assert checkout_rows_by_rfid["INA004"]["projetos"] == ["P82"]


def test_checkin_remains_visible_until_24_hours_even_across_singapore_midnight(monkeypatch):
    fixed_now = datetime(2024, 4, 7, 12, 0, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(admin_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(user_activity_module, "now_sgt", lambda: fixed_now)

    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "INA010", "nome": "Fabio Fim Semana", "chave": "DD44", "projeto": "P82"},
        )
        assert save_user.status_code == 200

        with SessionLocal() as db:
            weekend_user = get_user_by_rfid(db, "INA010")
            weekend_user.checkin = True
            weekend_user.local = "main"
            weekend_user.time = fixed_now - timedelta(hours=23, minutes=45)
            weekend_user.last_active_at = weekend_user.time
            weekend_user.inactivity_days = 0
            db.commit()

        inactive_rows = client.get("/api/admin/inactive")
        assert inactive_rows.status_code == 200
        assert inactive_rows.json() == []

        checkin_rows = client.get("/api/admin/checkin")
        assert checkin_rows.status_code == 200
        assert any(row["rfid"] == "INA010" for row in checkin_rows.json())

        missing_checkout_rows = client.get("/api/admin/missing-checkout")
        assert missing_checkout_rows.status_code == 200
        assert missing_checkout_rows.json() == []

        with SessionLocal() as db:
            weekend_user = get_user_by_rfid(db, "INA010")
            assert weekend_user.inactivity_days == 0


def test_users_move_to_inactive_after_24_hours_even_on_weekends(monkeypatch):
    fixed_now = datetime(2024, 4, 7, 12, 0, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(admin_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(user_activity_module, "now_sgt", lambda: fixed_now)

    with TestClient(app) as client:
        ensure_admin_session(client)
        save_checkin_user = client.post(
            "/api/admin/users",
            json={"rfid": "INA011", "nome": "Gil Checkin Weekend", "chave": "GF11", "projeto": "P80"},
        )
        save_checkout_user = client.post(
            "/api/admin/users",
            json={"rfid": "INA012", "nome": "Helena Checkout Weekend", "chave": "HF12", "projeto": "P82"},
        )
        assert save_checkin_user.status_code == 200
        assert save_checkout_user.status_code == 200

        with SessionLocal() as db:
            weekend_checkin_user = get_user_by_rfid(db, "INA011")
            weekend_checkin_user.checkin = True
            weekend_checkin_user.local = "main"
            weekend_checkin_user.time = fixed_now - timedelta(hours=25)
            weekend_checkin_user.last_active_at = weekend_checkin_user.time
            weekend_checkin_user.inactivity_days = 0

            weekend_checkout_user = get_user_by_rfid(db, "INA012")
            weekend_checkout_user.checkin = False
            weekend_checkout_user.local = "co80"
            weekend_checkout_user.time = fixed_now - timedelta(hours=26)
            weekend_checkout_user.last_active_at = weekend_checkout_user.time
            weekend_checkout_user.inactivity_days = 0
            db.commit()

        inactive_rows = client.get("/api/admin/inactive")
        assert inactive_rows.status_code == 200
        inactive_payload = inactive_rows.json()
        assert [row["nome"] for row in inactive_payload] == ["Gil Checkin Weekend", "Helena Checkout Weekend"]
        assert [row["inactivity_days"] for row in inactive_payload] == [1, 1]
        assert [row["latest_action"] for row in inactive_payload] == ["checkin", "checkout"]

        checkin_rows = client.get("/api/admin/checkin")
        assert checkin_rows.status_code == 200
        assert all(row["rfid"] != "INA011" for row in checkin_rows.json())

        checkout_rows = client.get("/api/admin/checkout")
        assert checkout_rows.status_code == 200
        assert all(row["rfid"] != "INA012" for row in checkout_rows.json())

        missing_checkout_rows = client.get("/api/admin/missing-checkout")
        assert missing_checkout_rows.status_code == 200
        assert missing_checkout_rows.json() == []

        with SessionLocal() as db:
            weekend_checkin_user = get_user_by_rfid(db, "INA011")
            weekend_checkout_user = get_user_by_rfid(db, "INA012")
            assert weekend_checkin_user.inactivity_days == 1
            assert weekend_checkout_user.inactivity_days == 1


def test_admin_presence_lists_follow_latest_activity_even_when_current_state_is_missing_or_stale():
    with SessionLocal() as db:
        stale_user = User(
            rfid="LATE001",
            chave="LT01",
            nome="Usuario Estado Antigo",
            projeto="P80",
            local="main",
            checkin=True,
            time=now_sgt() - timedelta(days=5),
            last_active_at=now_sgt() - timedelta(days=5),
            inactivity_days=0,
        )
        check_event_only_user = User(
            rfid="LATE002",
            chave="LT02",
            nome="Usuario Historico RFID",
            projeto="P83",
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt() - timedelta(days=3),
            inactivity_days=0,
        )
        no_activity_user = User(
            rfid="LATE003",
            chave="LT03",
            nome="Usuario Sem Atividade",
            projeto="P82",
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
        )
        db.add_all([stale_user, check_event_only_user, no_activity_user])
        db.flush()

        db.add(
            UserSyncEvent(
                user_id=stale_user.id,
                chave=stale_user.chave,
                rfid=stale_user.rfid,
                source="android",
                action="checkout",
                projeto=stale_user.projeto,
                local="co80",
                event_time=now_sgt() - timedelta(hours=6),
                created_at=now_sgt(),
                source_request_id=f"android-{uuid.uuid4().hex}",
                device_id=None,
            )
        )
        db.add(
            CheckEvent(
                idempotency_key=f"late-checkout-{uuid.uuid4().hex}",
                source="device",
                rfid="LATE002",
                action="checkout",
                status="queued",
                message="checkout antigo",
                details=None,
                project="P83",
                device_id="ESP32-LATE",
                local="main",
                request_path="/api/scan",
                http_status=202,
                event_time=now_sgt() - timedelta(days=2),
                submitted_at=None,
                retry_count=0,
            )
        )
        db.add(
            CheckEvent(
                idempotency_key=f"late-checkin-{uuid.uuid4().hex}",
                source="device",
                rfid="LATE002",
                action="checkin",
                status="queued",
                message="checkin recente",
                details=None,
                project="P83",
                device_id="ESP32-LATE",
                local="co83",
                request_path="/api/scan",
                http_status=202,
                event_time=now_sgt() - timedelta(hours=3),
                submitted_at=None,
                retry_count=0,
            )
        )
        db.commit()

    with TestClient(app) as client:
        ensure_admin_session(client)

        checkin_rows = client.get("/api/admin/checkin")
        assert checkin_rows.status_code == 200
        checkin_payload = checkin_rows.json()

        checkout_rows = client.get("/api/admin/checkout")
        assert checkout_rows.status_code == 200
        checkout_payload = checkout_rows.json()

        assert all(row["chave"] != "LT01" for row in checkin_payload)
        stale_checkout = next(row for row in checkout_payload if row["chave"] == "LT01")
        assert stale_checkout["local"] == "co80"
        assert stale_checkout["checkin"] is False

        fallback_checkin = next(row for row in checkin_payload if row["chave"] == "LT02")
        assert fallback_checkin["local"] == "co83"
        assert fallback_checkin["checkin"] is True
        assert all(row["chave"] != "LT02" for row in checkout_payload)

        assert all(row["chave"] != "LT03" for row in checkin_payload)
        assert all(row["chave"] != "LT03" for row in checkout_payload)


def test_admin_presence_lists_include_assiduidade_labels():
    with TestClient(app) as client:
        ensure_admin_session(client)

        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "ASSI01", "nome": "Usuario RFID", "chave": "AQ11", "projeto": "P80"},
        )
        assert save_user.status_code == 200

        rfid_checkin = client.post(
            "/api/scan",
            json={
                "local": "main",
                "rfid": "ASSI01",
                "action": "checkin",
                "device_id": "ESP32-ASSI",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert rfid_checkin.status_code == 200

        retroativo_checkout = client.post(
            "/api/mobile/events/forms-submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AQ22",
                "projeto": "P82",
                "action": "checkout",
                "informe": "Retroativo",
                "event_time": now_sgt().isoformat(),
                "client_event_id": f"android-forms-assiduidade-{uuid.uuid4().hex}",
            },
        )
        assert retroativo_checkout.status_code == 200

        checkin_rows = client.get("/api/admin/checkin")
        assert checkin_rows.status_code == 200
        checkin_match = next(row for row in checkin_rows.json() if row["chave"] == "AQ11")
        assert checkin_match["assiduidade"] == "Normal"

        checkout_rows = client.get("/api/admin/checkout")
        assert checkout_rows.status_code == 200
        checkout_match = next(row for row in checkout_rows.json() if row["chave"] == "AQ22")
        assert checkout_match["assiduidade"] == "Retroativo"


def test_mobile_sync_autocreates_user_and_updates_state():
    with TestClient(app) as client:
        event_time = now_sgt().isoformat()
        response = client.post(
            "/api/mobile/events/sync",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AP11",
                "projeto": "P80",
                "action": "checkin",
                "event_time": event_time,
                "client_event_id": f"android-{uuid.uuid4().hex}",
            },
        )
        assert response.status_code == 200
        payload = response.json()
        assert payload["ok"] is True
        assert payload["duplicate"] is False
        assert payload["state"]["found"] is True
        assert payload["state"]["last_checkin_at"] is not None

        with SessionLocal() as db:
            user = get_user_by_chave(db, "AP11")
            assert user.nome == "Oriundo do Aplicativo"
            assert user.rfid is None
            assert user.checkin is True


def test_mobile_submit_autocreates_user_with_app_origin_name():
    client_event_id = f"android-submit-{uuid.uuid4().hex}"

    with TestClient(app) as client:
        response = client.post(
            "/api/mobile/events/submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AS11",
                "projeto": "P83",
                "action": "checkout",
                "event_time": now_sgt().isoformat(),
                "client_event_id": client_event_id,
            },
        )
        assert response.status_code == 200
        payload = response.json()
        assert payload["ok"] is True
        assert payload["duplicate"] is False
        assert payload["state"]["found"] is True
        assert payload["state"]["projeto"] == "P83"

        with SessionLocal() as db:
            user = get_user_by_chave(db, "AS11")
            assert user.nome == "Oriundo do Aplicativo"
            assert user.rfid is None
            assert user.projeto == "P83"


def test_mobile_forms_submit_autocreates_user_with_app_origin_name():
    client_event_id = f"android-forms-create-{uuid.uuid4().hex}"

    with TestClient(app) as client:
        response = client.post(
            "/api/mobile/events/forms-submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AF11",
                "projeto": "P82",
                "action": "checkin",
                "informe": "normal",
                "event_time": now_sgt().isoformat(),
                "client_event_id": client_event_id,
            },
        )
        assert response.status_code == 200
        payload = response.json()
        assert payload["ok"] is True
        assert payload["duplicate"] is False
        assert payload["state"]["found"] is True
        assert payload["state"]["projeto"] == "P82"

        with SessionLocal() as db:
            user = get_user_by_chave(db, "AF11")
            assert user.nome == "Oriundo do Aplicativo"
            assert user.rfid is None
            assert user.projeto == "P82"


def test_mobile_forms_submit_skips_same_day_repeated_action():
    first_event_time = now_sgt()
    second_event_time = first_event_time + timedelta(minutes=5)

    with TestClient(app) as client:
        first = client.post(
            "/api/mobile/events/forms-submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AF12",
                "projeto": "P82",
                "action": "checkin",
                "local": "Area A",
                "informe": "normal",
                "event_time": first_event_time.isoformat(),
                "client_event_id": f"android-forms-same-day-1-{uuid.uuid4().hex}",
            },
        )
        second = client.post(
            "/api/mobile/events/forms-submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AF12",
                "projeto": "P82",
                "action": "checkin",
                "local": "Area B",
                "informe": "normal",
                "event_time": second_event_time.isoformat(),
                "client_event_id": f"android-forms-same-day-2-{uuid.uuid4().hex}",
            },
        )

        assert first.status_code == 200
        assert first.json()["queued_forms"] is True
        assert second.status_code == 200
        assert second.json()["ok"] is True
        assert second.json()["duplicate"] is False
        assert second.json()["queued_forms"] is False

        with SessionLocal() as db:
            user = get_user_by_chave(db, "AF12")
            queued = db.execute(
                select(FormsSubmission)
                .where(FormsSubmission.chave == "AF12")
                .order_by(FormsSubmission.id)
            ).scalars().all()
            sync_events = db.execute(
                select(UserSyncEvent).where(UserSyncEvent.chave == "AF12", UserSyncEvent.action == "checkin")
            ).scalars().all()

            assert user.checkin is True
            assert user.local == "Area B"
            assert len(queued) == 2
            assert queued[0].status == "pending"
            assert queued[-1].status == "skipped"
            assert queued[-1].display_status == "not_realized"
            assert queued[-1].last_error == "repeated_same_action_same_day"
            assert len(sync_events) == 2


def test_mobile_submit_ignores_provider_checkout_when_evaluating_same_day_repeat():
    first_event_time = datetime(2026, 4, 22, 8, 0, 0, tzinfo=ZoneInfo(settings.tz_name))
    second_event_time = datetime(2026, 4, 22, 11, 0, 0, tzinfo=ZoneInfo(settings.tz_name))

    with TestClient(app) as client:
        first = client.post(
            "/api/mobile/events/submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AS12",
                "projeto": "P82",
                "action": "checkin",
                "local": "Area A",
                "event_time": first_event_time.isoformat(),
                "client_event_id": f"mobile-provider-ignore-1-{uuid.uuid4().hex}",
            },
        )
        assert first.status_code == 200, first.text
        assert first.json()["queued_forms"] is True

        provider_checkout = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": "AS12",
                "nome": "USUARIO MOBILE IGNORA PROVIDER",
                "projeto": "P82",
                "atividade": "check-out",
                "informe": "normal",
                "data": "22/04/2026",
                "hora": "10:00:00",
            },
        )
        assert provider_checkout.status_code == 200, provider_checkout.text
        assert provider_checkout.json()["updated_current_state"] is True

        second = client.post(
            "/api/mobile/events/submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AS12",
                "projeto": "P82",
                "action": "checkin",
                "local": "Area B",
                "event_time": second_event_time.isoformat(),
                "client_event_id": f"mobile-provider-ignore-2-{uuid.uuid4().hex}",
            },
        )
        assert second.status_code == 200, second.text
        assert second.json()["queued_forms"] is False

        with SessionLocal() as db:
            user = get_user_by_chave(db, "AS12")
            queued = db.execute(
                select(FormsSubmission).where(FormsSubmission.chave == "AS12").order_by(FormsSubmission.id)
            ).scalars().all()

            assert user.checkin is True
            assert user.local == "Area B"
            assert len(queued) == 1


def test_mobile_forms_submit_requeues_same_action_after_singapore_midnight():
    first_event_time = now_sgt() - timedelta(days=1)
    second_event_time = now_sgt()

    with TestClient(app) as client:
        first = client.post(
            "/api/mobile/events/forms-submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AF13",
                "projeto": "P83",
                "action": "checkin",
                "local": "Area A",
                "informe": "normal",
                "event_time": first_event_time.isoformat(),
                "client_event_id": f"android-forms-next-day-1-{uuid.uuid4().hex}",
            },
        )
        second = client.post(
            "/api/mobile/events/forms-submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AF13",
                "projeto": "P83",
                "action": "checkin",
                "local": "Area C",
                "informe": "normal",
                "event_time": second_event_time.isoformat(),
                "client_event_id": f"android-forms-next-day-2-{uuid.uuid4().hex}",
            },
        )

        assert first.status_code == 200
        assert second.status_code == 200
        assert second.json()["ok"] is True
        assert second.json()["duplicate"] is False
        assert second.json()["queued_forms"] is True

        with SessionLocal() as db:
            queued = db.execute(select(FormsSubmission).where(FormsSubmission.chave == "AF13")).scalars().all()
            user = get_user_by_chave(db, "AF13")

            assert len(queued) == 2
            assert user.checkin is True
            assert user.local == "Area C"


def test_mobile_forms_submit_requeues_same_action_after_project_local_midnight():
    with SessionLocal() as db:
        project = db.execute(select(Project).where(Project.name == "P99")).scalar_one_or_none()
        if project is None:
            db.add(Project(name="P99", **build_project_fields_for_country("JP")))
            db.commit()

    first_event_time = datetime(2026, 4, 17, 23, 30, 0, tzinfo=ZoneInfo("Asia/Tokyo"))
    second_event_time = datetime(2026, 4, 18, 0, 30, 0, tzinfo=ZoneInfo("Asia/Tokyo"))

    with TestClient(app) as client:
        first = client.post(
            "/api/mobile/events/forms-submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AF99",
                "projeto": "P99",
                "action": "checkin",
                "local": "Area Tokyo A",
                "informe": "normal",
                "event_time": first_event_time.isoformat(),
                "client_event_id": f"android-forms-project-midnight-1-{uuid.uuid4().hex}",
            },
        )
        second = client.post(
            "/api/mobile/events/forms-submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AF99",
                "projeto": "P99",
                "action": "checkin",
                "local": "Area Tokyo B",
                "informe": "normal",
                "event_time": second_event_time.isoformat(),
                "client_event_id": f"android-forms-project-midnight-2-{uuid.uuid4().hex}",
            },
        )

        assert first.status_code == 200
        assert first.json()["queued_forms"] is True
        assert second.status_code == 200
        assert second.json()["ok"] is True
        assert second.json()["duplicate"] is False
        assert second.json()["queued_forms"] is True

        with SessionLocal() as db:
            queued = db.execute(select(FormsSubmission).where(FormsSubmission.chave == "AF99")).scalars().all()
            user = get_user_by_chave(db, "AF99")

            assert len(queued) == 2
            assert user.checkin is True
            assert user.local == "Area Tokyo B"


def test_web_and_mobile_accept_naive_event_time_in_project_timezone_outside_singapore():
    with SessionLocal() as db:
        project = db.execute(select(Project).where(Project.name == "P99")).scalar_one_or_none()
        if project is None:
            db.add(Project(name="P99", **build_project_fields_for_country("JP")))
            db.commit()

    expected_web_time = datetime(2026, 4, 18, 8, 15, 0, tzinfo=ZoneInfo("Asia/Tokyo"))
    expected_mobile_time = datetime(2026, 4, 18, 8, 45, 0, tzinfo=ZoneInfo("Asia/Tokyo"))

    with TestClient(app) as client:
        registered = register_web_password(client, chave="TZ99", senha="web123", projeto="P99")
        assert registered.status_code == 200, registered.text

        web_response = client.post(
            "/api/web/check",
            json={
                "chave": "TZ99",
                "projeto": "P99",
                "action": "checkin",
                "informe": "normal",
                "local": "Tokyo Web",
                "event_time": "2026-04-18T08:15:00",
                "client_event_id": f"web-naive-project-timezone-{uuid.uuid4().hex}",
            },
        )
        assert web_response.status_code == 200, web_response.text

        mobile_response = client.post(
            "/api/mobile/events/sync",
            headers=MOBILE_HEADERS,
            json={
                "chave": "TZ98",
                "projeto": "P99",
                "action": "checkin",
                "event_time": "2026-04-18T08:45:00",
                "client_event_id": f"mobile-naive-project-timezone-{uuid.uuid4().hex}",
            },
        )
        assert mobile_response.status_code == 200, mobile_response.text

        web_history = client.get("/api/web/check/state", params={"chave": "TZ99"})
        assert web_history.status_code == 200, web_history.text
        web_state = web_history.json()
        assert web_state["found"] is True
        assert web_state["projeto"] == "P99"
        assert web_state["current_action"] == "checkin"
        assert normalize_event_time(datetime.fromisoformat(web_state["last_checkin_at"]), timezone_name="Asia/Tokyo") == expected_web_time

        mobile_state_response = client.get("/api/mobile/state?chave=TZ98", headers=MOBILE_HEADERS)
        assert mobile_state_response.status_code == 200, mobile_state_response.text
        mobile_state = mobile_state_response.json()
        assert mobile_state["found"] is True
        assert mobile_state["current_action"] == "checkin"
        assert normalize_event_time(datetime.fromisoformat(mobile_state["last_checkin_at"]), timezone_name="Asia/Tokyo") == expected_mobile_time

    with SessionLocal() as db:
        web_user = get_user_by_chave(db, "TZ99")
        mobile_user = get_user_by_chave(db, "TZ98")

    assert normalize_event_time(web_user.time, timezone_name="Asia/Tokyo") == expected_web_time
    assert normalize_event_time(mobile_user.time, timezone_name="Asia/Tokyo") == expected_mobile_time


def test_mobile_forms_submit_surfaces_non_singapore_timezone_in_checkin_and_checkout_rows():
    with SessionLocal() as db:
        project = db.execute(select(Project).where(Project.name == "P99")).scalar_one_or_none()
        if project is None:
            db.add(Project(name="P99", **build_project_fields_for_country("JP")))
            db.commit()

    reference_time = now_sgt().astimezone(ZoneInfo("Asia/Tokyo")).replace(microsecond=0)
    checkin_time = reference_time - timedelta(hours=2)
    checkout_time = reference_time - timedelta(minutes=30)

    with TestClient(app) as client:
        checkin_response = client.post(
            "/api/mobile/events/forms-submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "TZ97",
                "projeto": "P99",
                "action": "checkin",
                "local": "Tokyo Area A",
                "informe": "normal",
                "event_time": checkin_time.isoformat(),
                "client_event_id": f"mobile-tokyo-checkin-{uuid.uuid4().hex}",
            },
        )
        assert checkin_response.status_code == 200, checkin_response.text

        ensure_admin_session(client)
        checkin_rows = client.get("/api/admin/checkin")
        assert checkin_rows.status_code == 200, checkin_rows.text
        checkin_row = next(row for row in checkin_rows.json() if row["chave"] == "TZ97")
        assert checkin_row["projeto"] == "P99"
        assert checkin_row["timezone_name"] == "Asia/Tokyo"
        assert checkin_row["timezone_label"] == "Japão (+9)"
        assert normalize_event_time(datetime.fromisoformat(checkin_row["time"]), timezone_name="Asia/Tokyo") == checkin_time

        checkout_response = client.post(
            "/api/mobile/events/forms-submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "TZ97",
                "projeto": "P99",
                "action": "checkout",
                "local": "Tokyo Area B",
                "informe": "normal",
                "event_time": checkout_time.isoformat(),
                "client_event_id": f"mobile-tokyo-checkout-{uuid.uuid4().hex}",
            },
        )
        assert checkout_response.status_code == 200, checkout_response.text

        checkout_rows = client.get("/api/admin/checkout")
        assert checkout_rows.status_code == 200, checkout_rows.text
        checkout_row = next(row for row in checkout_rows.json() if row["chave"] == "TZ97")
        assert checkout_row["projeto"] == "P99"
        assert checkout_row["timezone_name"] == "Asia/Tokyo"
        assert checkout_row["timezone_label"] == "Japão (+9)"
        assert normalize_event_time(datetime.fromisoformat(checkout_row["time"]), timezone_name="Asia/Tokyo") == checkout_time


def test_mobile_state_returns_current_location_for_checked_in_user():
    with TestClient(app) as client:
        submit_response = client.post(
            "/api/mobile/events/forms-submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AF14",
                "projeto": "P80",
                "action": "checkin",
                "local": "Base P80",
                "informe": "normal",
                "event_time": now_sgt().isoformat(),
                "client_event_id": f"android-forms-state-{uuid.uuid4().hex}",
            },
        )
        assert submit_response.status_code == 200

        state_response = client.get(
            "/api/mobile/state?chave=AF14",
            headers=MOBILE_HEADERS,
        )
        assert state_response.status_code == 200
        payload = state_response.json()
        assert payload["found"] is True
        assert payload["current_action"] == "checkin"
        assert payload["current_local"] == "Base P80"


def test_admin_checkin_list_accepts_mobile_user_without_rfid():
    with TestClient(app) as client:
        created = client.post(
            "/api/mobile/events/sync",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AP15",
                "projeto": "P80",
                "action": "checkin",
                "event_time": now_sgt().isoformat(),
                "client_event_id": f"android-{uuid.uuid4().hex}",
            },
        )
        assert created.status_code == 200

        ensure_admin_session(client)
        checkin_rows = client.get("/api/admin/checkin")
        assert checkin_rows.status_code == 200
        payload = checkin_rows.json()
        matched = next(row for row in payload if row["chave"] == "AP15")
        assert matched["rfid"] is None
        assert matched["checkin"] is True


def test_mobile_sync_notifies_admin_realtime_subscribers():
    subscriber_id, queue = admin_updates_broker.subscribe()

    try:
        with TestClient(app) as client:
            response = client.post(
                "/api/mobile/events/sync",
                headers=MOBILE_HEADERS,
                json={
                    "chave": "AP16",
                    "projeto": "P83",
                    "action": "checkout",
                    "event_time": now_sgt().isoformat(),
                    "client_event_id": f"android-{uuid.uuid4().hex}",
                },
            )

        assert response.status_code == 200
        payloads = [queue.get_nowait()]
        while True:
            try:
                payloads.append(queue.get_nowait())
            except asyncio.QueueEmpty:
                break

        assert any('"reason": "checkout"' in payload for payload in payloads)
    finally:
        admin_updates_broker.unsubscribe(subscriber_id)


def test_mobile_sync_is_idempotent_for_same_event_id():
    with TestClient(app) as client:
        client_event_id = f"android-{uuid.uuid4().hex}"
        payload = {
            "chave": "AP12",
            "projeto": "P83",
            "action": "checkout",
            "event_time": now_sgt().isoformat(),
            "client_event_id": client_event_id,
        }
        first = client.post("/api/mobile/events/sync", headers=MOBILE_HEADERS, json=payload)
        second = client.post("/api/mobile/events/sync", headers=MOBILE_HEADERS, json=payload)
        assert first.status_code == 200
        assert second.status_code == 200
        assert second.json()["duplicate"] is True


def test_admin_can_edit_mobile_created_user_without_rfid():
    with TestClient(app) as client:
        created = client.post(
            "/api/mobile/events/sync",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AP13",
                "projeto": "P80",
                "action": "checkin",
                "event_time": now_sgt().isoformat(),
                "client_event_id": f"android-{uuid.uuid4().hex}",
            },
        )
        assert created.status_code == 200

        ensure_admin_session(client)
        users = client.get("/api/admin/users")
        assert users.status_code == 200
        created_user = next(row for row in users.json() if row["chave"] == "AP13")
        assert created_user["rfid"] is None

        updated = client.post(
            "/api/admin/users",
            json={
                "user_id": created_user["id"],
                "nome": "Nome Corrigido",
                "chave": "AP13",
                "projeto": "P83",
            },
        )
        assert updated.status_code == 200

        users_after = client.get("/api/admin/users")
        assert users_after.status_code == 200
        updated_user = next(row for row in users_after.json() if row["id"] == created_user["id"])
        assert updated_user["nome"] == "Nome Corrigido"
        assert updated_user["projeto"] == "P83"


def test_admin_can_attach_rfid_to_mobile_created_user_by_unique_chave():
    with TestClient(app) as client:
        created = client.post(
            "/api/mobile/events/sync",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AP14",
                "projeto": "P82",
                "action": "checkin",
                "event_time": now_sgt().isoformat(),
                "client_event_id": f"android-{uuid.uuid4().hex}",
            },
        )
        assert created.status_code == 200

        ensure_admin_session(client)
        attached = client.post(
            "/api/admin/users",
            json={
                "rfid": "APPRFID1",
                "nome": "Nome Ajustado",
                "chave": "AP14",
                "projeto": "P82",
            },
        )
        assert attached.status_code == 200
        assert attached.json()["linked_existing_user"] is True

        users = client.get("/api/admin/users")
        assert users.status_code == 200
        matched = [row for row in users.json() if row["chave"] == "AP14"]
        assert len(matched) == 1
        assert matched[0]["rfid"] == "APPRFID1"
        assert matched[0]["nome"] == "Nome Ajustado"
        assert matched[0]["projeto"] == "P82"


def test_pending_registration_links_rfid_to_existing_mobile_user_by_chave():
    with TestClient(app) as client:
        created = client.post(
            "/api/mobile/events/sync",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AP17",
                "projeto": "P80",
                "action": "checkin",
                "event_time": now_sgt().isoformat(),
                "client_event_id": f"android-{uuid.uuid4().hex}",
            },
        )
        assert created.status_code == 200

        ensure_admin_session(client)
        scan_pending = client.post(
            "/api/scan",
            json={
                "local": "main",
                "rfid": "LINKRFID17",
                "action": "checkin",
                "device_id": "ESP32-LINK",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_pending.status_code == 200
        assert scan_pending.json()["outcome"] == "pending_registration"

        attached = client.post(
            "/api/admin/users",
            json={
                "rfid": "LINKRFID17",
                "nome": "Nome Vindo da Pendência",
                "chave": "AP17",
                "projeto": "P83",
            },
        )
        assert attached.status_code == 200
        assert attached.json()["linked_existing_user"] is True

        users = client.get("/api/admin/users")
        assert users.status_code == 200
        matched = [row for row in users.json() if row["chave"] == "AP17"]
        assert len(matched) == 1
        assert matched[0]["rfid"] == "LINKRFID17"
        assert matched[0]["nome"] == "Nome Vindo da Pendência"
        assert matched[0]["projeto"] == "P83"

        pending = client.get("/api/admin/pending")
        assert pending.status_code == 200
        assert all(row["rfid"] != "LINKRFID17" for row in pending.json())


def test_mobile_state_reflects_rfid_scan_history(monkeypatch):
    monkeypatch.setattr(
        FormsWorker,
        "submit_with_retries",
        lambda self, action, chave, projeto, ontime=True, **_kwargs: {
            "success": True,
            "message": f"mocked {action}",
            "retry_count": 0,
        },
    )

    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "RFIDSYNC1", "nome": "Usuario Sync", "chave": "SY11", "projeto": "P80"},
        )
        assert save_user.status_code == 200

        scan_response = client.post(
            "/api/scan",
            json={
                "local": "main",
                "rfid": "RFIDSYNC1",
                "action": "checkin",
                "device_id": "ESP32-SYNC",
                "request_id": f"req-{uuid.uuid4().hex}",
                "shared_key": "device-test-key",
            },
        )
        assert scan_response.status_code == 200

        mobile_state = client.get("/api/mobile/state?chave=SY11", headers=MOBILE_HEADERS)
        assert mobile_state.status_code == 200
        payload = mobile_state.json()
        assert payload["found"] is True
        assert payload["current_action"] == "checkin"
        assert payload["last_checkin_at"] is not None


def test_mobile_forms_submit_accepts_retroativo_and_persists_ontime_false(monkeypatch):
    monkeypatch.setattr(
        FormsWorker,
        "submit_with_retries",
        lambda self, action, chave, projeto, ontime=True, **_kwargs: {
            "success": True,
            "message": f"mocked {action}",
            "retry_count": 0,
        },
    )

    client_event_id = f"android-forms-{uuid.uuid4().hex}"

    with TestClient(app) as client:
        response = client.post(
            "/api/mobile/events/forms-submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "RT11",
                "projeto": "P82",
                "action": "checkin",
                "informe": "Retroativo",
                "event_time": now_sgt().isoformat(),
                "client_event_id": client_event_id,
            },
        )
        assert response.status_code == 200
        payload = response.json()
        assert payload["ok"] is True
        assert payload["duplicate"] is False
        assert payload["queued_forms"] is True
        assert payload["state"]["current_action"] == "checkin"

        with SessionLocal() as db:
            queued = db.execute(
                select(FormsSubmission).where(FormsSubmission.request_id == client_event_id)
            ).scalar_one()
            sync_event = db.execute(
                select(UserSyncEvent).where(
                    UserSyncEvent.source == "android_forms",
                    UserSyncEvent.source_request_id == client_event_id,
                )
            ).scalar_one()
            assert queued.ontime is False
            assert sync_event.ontime is False

        processed = process_forms_submission_queue_once(max_items=1000)
        assert processed >= 1

        with SessionLocal() as db:
            processed_submission = db.execute(
                select(FormsSubmission).where(FormsSubmission.request_id == client_event_id)
            ).scalar_one()
            assert processed_submission.status == "success"

        ensure_admin_session(client)
        events = client.get("/api/admin/events")
        assert events.status_code == 200
        assert any(
            event["request_path"] == "/api/mobile/events/forms-submit"
            and event["ontime"] is False
            and event["status"] == "queued"
            and event["chave"] == "RT11"
            for event in events.json()
        )
        assert any(
            event["source"] == "forms"
            and event["ontime"] is False
            and event["status"] == "success"
            and event["chave"] == "RT11"
            for event in events.json()
        )


def test_mobile_forms_submit_is_idempotent_for_same_event_id():
    with TestClient(app) as client:
        client_event_id = f"android-forms-{uuid.uuid4().hex}"
        payload = {
            "chave": "RT12",
            "projeto": "P80",
            "action": "checkout",
            "informe": "normal",
            "event_time": now_sgt().isoformat(),
            "client_event_id": client_event_id,
        }
        first = client.post("/api/mobile/events/forms-submit", headers=MOBILE_HEADERS, json=payload)
        second = client.post("/api/mobile/events/forms-submit", headers=MOBILE_HEADERS, json=payload)
        assert first.status_code == 200
        assert second.status_code == 200
        assert second.json()["duplicate"] is True


def test_mobile_check_page_is_served_on_user_path():
    with TestClient(app) as client:
        response = client.get("/user")
        assert response.status_code == 200
        assert "Registrar" in response.text
        assert "Senha" in response.text
        assert "Local" in response.text
        assert "Atualizar local" in response.text
        assert "Chave" in response.text
        assert "Último Check-In" in response.text
        assert "Último Check-Out" in response.text
        assert "/api/web/check" in response.text
        assert "/api/web/auth/status" in response.text
        assert "/api/web/auth/register-password" in response.text
        assert "/api/web/auth/register-user" in response.text
        assert "/api/web/auth/login" in response.text
        assert "/api/web/auth/change-password" in response.text
        assert "/api/web/auth/logout" in response.text
        assert "/api/web/check/state" in response.text
        assert "/api/web/check/location" in response.text
        assert "Solicitar Cadastro" in response.text


def test_static_site_mount_flags_default_to_enabled():
    assert should_serve_static_site("admin", settings_obj=settings) is True
    assert should_serve_static_site("user", settings_obj=settings) is True
    assert should_serve_static_site("transport", settings_obj=settings) is True


def test_static_site_mount_flags_can_disable_api_serving_per_site():
    custom_settings = SimpleNamespace(
        serve_admin_site_in_api=False,
        serve_user_site_in_api=True,
        serve_transport_site_in_api=False,
    )

    assert should_serve_static_site("admin", settings_obj=custom_settings) is False
    assert should_serve_static_site("user", settings_obj=custom_settings) is True
    assert should_serve_static_site("transport", settings_obj=custom_settings) is False


def test_transport_page_is_served_on_transport_path():
    with TestClient(app) as client:
        response = client.get("/transport")
        assert response.status_code == 200
        assert "User List" in response.text
        assert 'data-toggle-request-section="extra"' in response.text
        assert 'data-toggle-request-section="weekend"' in response.text
        assert 'data-toggle-request-section="regular"' in response.text
        assert "Regular Transport List" in response.text
        assert "Weekend Transport List" in response.text
        assert "Extra Transport List" in response.text
        assert "System Support" in response.text
        assert "Tamer Salmem (HR70)" in response.text
        assert "Home to Work" in response.text
        assert "Work to Home" in response.text
        assert 'id="tela01menu"' in response.text
        assert 'id="tela01main_dir_down"' in response.text
        assert 'data-date-link' in response.text
        assert 'data-date-today' not in response.text


def test_transport_dashboard_groups_requests_by_selected_date_and_assignment_status():
    friday = date(2026, 4, 17)
    saturday = date(2026, 4, 18)
    timestamp = now_sgt()
    regular_key = make_test_key("T")
    weekend_key = make_test_key("T")
    extra_key = make_test_key("T")

    with SessionLocal() as db:
        regular_vehicle = Vehicle(placa="REG9001", tipo="van", color="Silver", lugares=12, tolerance=8, service_scope="regular")
        weekend_vehicle = Vehicle(placa="WKD9001", tipo="van", color="Black", lugares=10, tolerance=15, service_scope="weekend")
        extra_vehicle = Vehicle(placa="EXT9001", tipo="carro", color="Red", lugares=4, tolerance=6, service_scope="extra")
        db.add_all(
            [
                Workplace(workplace="Transport Hub Alpha", address="1 Harbour Front", zip="111111", country="Singapore"),
                regular_vehicle,
                weekend_vehicle,
                extra_vehicle,
            ]
        )
        db.flush()
        add_transport_schedule(
            db,
            vehicle=regular_vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=regular_vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=weekend_vehicle,
            service_scope="weekend",
            route_kind="home_to_work",
            recurrence_kind="matching_weekday",
            weekday=saturday.weekday(),
        )
        add_transport_schedule(
            db,
            vehicle=weekend_vehicle,
            service_scope="weekend",
            route_kind="work_to_home",
            recurrence_kind="matching_weekday",
            weekday=saturday.weekday(),
        )
        add_transport_schedule(
            db,
            vehicle=extra_vehicle,
            service_scope="extra",
            route_kind="home_to_work",
            recurrence_kind="single_date",
            service_date=friday,
        )

        regular_user = User(
            rfid=None,
            nome="Regular Rider",
            chave=regular_key,
            projeto="P80",
            workplace="Transport Hub Alpha",
            end_rua="10 Regular Street",
            zip="900001",
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        weekend_user = User(
            rfid=None,
            nome="Weekend Rider",
            chave=weekend_key,
            projeto="P82",
            workplace="Transport Hub Alpha",
            end_rua="20 Weekend Street",
            zip="900002",
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        extra_user = User(
            rfid=None,
            nome="Extra Rider",
            chave=extra_key,
            projeto="P83",
            workplace="Transport Hub Alpha",
            end_rua="30 Extra Street",
            zip="900003",
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        db.add_all([regular_user, weekend_user, extra_user])
        db.flush()

        regular_request = TransportRequest(
            user_id=regular_user.id,
            request_kind="regular",
            recurrence_kind="weekday",
            requested_time="07:30",
            single_date=None,
            created_via="bot",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
            cancelled_at=None,
        )
        weekend_request = TransportRequest(
            user_id=weekend_user.id,
            request_kind="weekend",
            recurrence_kind="weekend",
            requested_time="08:10",
            single_date=None,
            created_via="bot",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
            cancelled_at=None,
        )
        extra_request = TransportRequest(
            user_id=extra_user.id,
            request_kind="extra",
            recurrence_kind="single_date",
            requested_time="09:15",
            single_date=friday,
            created_via="bot",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
            cancelled_at=None,
        )
        db.add_all([regular_request, weekend_request, extra_request])
        db.flush()
        db.add(
            TransportAssignment(
                request_id=regular_request.id,
                service_date=friday,
                route_kind="home_to_work",
                vehicle_id=regular_vehicle.id,
                status="confirmed",
                response_message=None,
                boarding_time="07:05",
                assigned_by_admin_id=None,
                created_at=timestamp,
                updated_at=timestamp,
                notified_at=None,
            )
        )
        db.commit()

    with TestClient(app) as client:
        ensure_admin_session(client)
        friday_response = client.get("/api/transport/dashboard", params={"service_date": friday.isoformat()})
        saturday_response = client.get("/api/transport/dashboard", params={"service_date": saturday.isoformat()})
        friday_work_to_home_response = client.get(
            "/api/transport/dashboard",
            params={"service_date": friday.isoformat(), "route_kind": "work_to_home"},
        )

    assert friday_response.status_code == 200
    assert saturday_response.status_code == 200
    assert friday_work_to_home_response.status_code == 200

    friday_payload = friday_response.json()
    saturday_payload = saturday_response.json()
    friday_work_to_home_payload = friday_work_to_home_response.json()

    friday_regular_row = next(row for row in friday_payload["regular_requests"] if row["chave"] == regular_key)
    friday_extra_row = next(row for row in friday_payload["extra_requests"] if row["chave"] == extra_key)
    friday_weekend_row = next(row for row in friday_payload["weekend_requests"] if row["chave"] == weekend_key)

    assert friday_regular_row["assignment_status"] == "confirmed"
    assert friday_regular_row["assigned_vehicle"]["placa"] == "REG9001"
    assert friday_regular_row["boarding_time"] == "07:05"
    assert friday_extra_row["assignment_status"] == "pending"
    assert friday_extra_row["boarding_time"] is None
    assert friday_weekend_row["assignment_status"] == "pending"
    assert friday_weekend_row["boarding_time"] is None
    assert friday_weekend_row["service_date"] == saturday.isoformat()
    assert any(row["placa"] == "REG9001" for row in friday_payload["regular_vehicles"])
    assert any(row["placa"] == "EXT9001" and row["route_kind"] == "home_to_work" for row in friday_payload["extra_vehicles"])

    friday_regular_registry_row = next(row for row in friday_payload["regular_vehicle_registry"] if row["placa"] == "REG9001")
    friday_weekend_registry_row = next(row for row in friday_payload["weekend_vehicle_registry"] if row["placa"] == "WKD9001")
    friday_extra_registry_row = next(row for row in friday_payload["extra_vehicle_registry"] if row["placa"] == "EXT9001")

    assert friday_regular_registry_row["assigned_count"] == 1
    assert friday_weekend_registry_row["assigned_count"] == 0
    assert friday_extra_registry_row["service_date"] == friday.isoformat()
    assert friday_extra_registry_row["route_kind"] == "home_to_work"
    assert friday_payload["selected_route"] == "home_to_work"
    assert datetime.fromisoformat(friday_payload["dashboard_generated_at"]).tzinfo is not None
    assert friday_payload["arrive_at_work_time"] == "07:45"

    saturday_regular_row = next(row for row in saturday_payload["regular_requests"] if row["chave"] == regular_key)
    saturday_weekend_row = next(row for row in saturday_payload["weekend_requests"] if row["chave"] == weekend_key)
    saturday_extra_row = next(row for row in saturday_payload["extra_requests"] if row["chave"] == extra_key)

    assert saturday_regular_row["assignment_status"] == "pending"
    assert saturday_regular_row["assigned_vehicle"] is None
    assert saturday_regular_row["boarding_time"] is None
    assert saturday_weekend_row["chave"] == weekend_key
    assert saturday_weekend_row["boarding_time"] is None
    assert saturday_extra_row["assignment_status"] == "pending"
    assert saturday_extra_row["boarding_time"] is None
    assert saturday_extra_row["service_date"] == friday.isoformat()
    assert any(row["placa"] == "WKD9001" for row in saturday_payload["weekend_vehicles"])
    assert datetime.fromisoformat(saturday_payload["dashboard_generated_at"]).tzinfo is not None
    assert saturday_payload["arrive_at_work_time"] == "07:45"

    friday_work_to_home_regular_row = next(
        row for row in friday_work_to_home_payload["regular_requests"] if row["chave"] == regular_key
    )

    assert any(row["placa"] == "REG9001" for row in friday_work_to_home_payload["regular_vehicles"])
    assert any(
        row["placa"] == "EXT9001" and row["route_kind"] == "home_to_work"
        for row in friday_work_to_home_payload["extra_vehicles"]
    )
    assert any(row["placa"] == "EXT9001" for row in friday_work_to_home_payload["extra_vehicle_registry"])
    assert friday_work_to_home_regular_row["assignment_status"] == "confirmed"
    assert friday_work_to_home_regular_row["assigned_vehicle"]["placa"] == "REG9001"
    assert friday_work_to_home_regular_row["boarding_time"] is None
    assert datetime.fromisoformat(friday_work_to_home_payload["dashboard_generated_at"]).tzinfo is not None
    assert friday_work_to_home_payload["arrive_at_work_time"] == "07:45"


def _create_transport_boarding_time_update_context(
    *,
    service_date: date,
    route_kind: str = "home_to_work",
    assignment_status: str = "confirmed",
) -> dict[str, int | date | str]:
    timestamp = now_sgt()
    request_key = make_test_key("BT")

    with SessionLocal() as db:
        workplace = Workplace(
            workplace=f"Boarding Time Hub {request_key}",
            address="90 Boarding Street",
            zip="900090",
            country="Singapore",
        )
        db.add(workplace)
        db.flush()

        vehicle = Vehicle(
            placa=f"BT{request_key[-4:]}",
            tipo="van",
            color="blue",
            lugares=4,
            tolerance=10,
            service_scope="regular",
        )
        db.add(vehicle)
        db.flush()

        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind=route_kind,
            recurrence_kind="weekday",
        )

        user = User(
            rfid=None,
            nome=f"Boarding Time Rider {request_key}",
            chave=request_key,
            projeto="P90",
            workplace=workplace.workplace,
            placa="BT9000",
            end_rua="90 Boarding Street",
            zip="900090",
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        request_row = TransportRequest(
            user_id=user.id,
            request_kind="regular",
            recurrence_kind="weekday",
            requested_time="07:30",
            single_date=None,
            created_via="bot",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
            cancelled_at=None,
        )
        db.add(request_row)
        db.flush()

        db.add(
            TransportAssignment(
                request_id=request_row.id,
                service_date=service_date,
                route_kind=route_kind,
                vehicle_id=vehicle.id if assignment_status == "confirmed" else None,
                status=assignment_status,
                response_message=None,
                boarding_time=None,
                assigned_by_admin_id=None,
                created_at=timestamp,
                updated_at=timestamp,
                notified_at=None,
            )
        )
        db.commit()

        return {
            "request_id": request_row.id,
            "service_date": service_date,
            "request_key": request_key,
        }


def test_transport_assignment_boarding_time_update_roundtrips_dashboard_and_supports_clear():
    friday = date(2026, 4, 24)
    context = _create_transport_boarding_time_update_context(service_date=friday)

    with TestClient(app) as client:
        ensure_admin_session(client)

        save_response = client.put(
            "/api/transport/assignments/boarding-time",
            json={
                "request_id": context["request_id"],
                "service_date": friday.isoformat(),
                "route_kind": "home_to_work",
                "boarding_time": "07:05",
            },
        )
        saved_dashboard_response = client.get(
            "/api/transport/dashboard",
            params={"service_date": friday.isoformat(), "route_kind": "home_to_work"},
        )

        clear_response = client.put(
            "/api/transport/assignments/boarding-time",
            json={
                "request_id": context["request_id"],
                "service_date": friday.isoformat(),
                "route_kind": "home_to_work",
                "boarding_time": None,
            },
        )
        cleared_dashboard_response = client.get(
            "/api/transport/dashboard",
            params={"service_date": friday.isoformat(), "route_kind": "home_to_work"},
        )

    assert save_response.status_code == 200, save_response.text
    assert save_response.json()["ok"] is True
    assert save_response.json()["message"] == "Transport boarding time saved successfully."
    assert save_response.json()["message_key"] == "status.boardingTimeSaved"
    assert save_response.json()["error_code"] is None

    assert saved_dashboard_response.status_code == 200, saved_dashboard_response.text
    saved_payload = saved_dashboard_response.json()
    saved_row = next(row for row in saved_payload["regular_requests"] if row["id"] == context["request_id"])
    assert saved_row["assignment_status"] == "confirmed"
    assert saved_row["boarding_time"] == "07:05"

    assert clear_response.status_code == 200, clear_response.text
    assert clear_response.json()["ok"] is True
    assert clear_response.json()["message"] == "Transport boarding time saved successfully."
    assert clear_response.json()["message_key"] == "status.boardingTimeSaved"
    assert clear_response.json()["error_code"] is None

    assert cleared_dashboard_response.status_code == 200, cleared_dashboard_response.text
    cleared_payload = cleared_dashboard_response.json()
    cleared_row = next(row for row in cleared_payload["regular_requests"] if row["id"] == context["request_id"])
    assert cleared_row["assignment_status"] == "confirmed"
    assert cleared_row["boarding_time"] is None


def test_transport_assignment_boarding_time_update_rejects_invalid_time_format():
    friday = date(2026, 4, 24)
    context = _create_transport_boarding_time_update_context(service_date=friday)

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.put(
            "/api/transport/assignments/boarding-time",
            json={
                "request_id": context["request_id"],
                "service_date": friday.isoformat(),
                "route_kind": "home_to_work",
                "boarding_time": "7:5",
            },
        )

    assert response.status_code == 422
    assert "O horario deve estar no formato hh:mm" in response.text


def test_transport_assignment_boarding_time_update_rejects_non_confirmed_and_non_eta_assignments():
    friday = date(2026, 4, 24)
    pending_context = _create_transport_boarding_time_update_context(
        service_date=friday,
        route_kind="home_to_work",
        assignment_status="pending",
    )
    work_to_home_context = _create_transport_boarding_time_update_context(
        service_date=friday,
        route_kind="work_to_home",
        assignment_status="confirmed",
    )

    with TestClient(app) as client:
        ensure_admin_session(client)
        pending_response = client.put(
            "/api/transport/assignments/boarding-time",
            json={
                "request_id": pending_context["request_id"],
                "service_date": friday.isoformat(),
                "route_kind": "home_to_work",
                "boarding_time": "07:05",
            },
        )
        work_to_home_response = client.put(
            "/api/transport/assignments/boarding-time",
            json={
                "request_id": work_to_home_context["request_id"],
                "service_date": friday.isoformat(),
                "route_kind": "work_to_home",
                "boarding_time": "18:05",
            },
        )

    assert pending_response.status_code == 409
    pending_detail = extract_transport_structured_detail(pending_response)
    assert pending_detail["message_key"] == "warnings.boardingTimeRequiresConfirmedAssignment"
    assert pending_detail["error_code"] == "transport_boarding_time_confirmed_required"
    assert pending_detail["technical_detail"] == "A confirmed transport assignment is required to update boarding_time."
    assert work_to_home_response.status_code == 409
    work_to_home_detail = extract_transport_structured_detail(work_to_home_response)
    assert work_to_home_detail["message_key"] == "warnings.boardingTimeEtaOnly"
    assert work_to_home_detail["error_code"] == "transport_boarding_time_eta_only"
    assert work_to_home_detail["technical_detail"] == (
        "Manual boarding_time is only available for confirmed home_to_work assignments."
    )


def test_transport_dashboard_exposes_membership_projects_without_duplicating_request_rows():
    monday = date(2026, 4, 20)
    timestamp = now_sgt()
    user_key = make_test_key("M")

    with SessionLocal() as db:
        user = User(
            rfid=None,
            nome="Multi Project Dashboard Rider",
            chave=user_key,
            projeto="P80",
            workplace=None,
            end_rua="80 Membership Street",
            zip="800080",
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()
        grant_user_project_memberships(db, user, ["P80", "P83"])

        transport_request = TransportRequest(
            user_id=user.id,
            request_kind="regular",
            recurrence_kind="weekday",
            requested_time="07:40",
            single_date=None,
            created_via="bot",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
            cancelled_at=None,
        )
        db.add(transport_request)
        db.commit()
        request_id = transport_request.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.get("/api/transport/dashboard", params={"service_date": monday.isoformat()})

    assert response.status_code == 200, response.text
    payload = response.json()
    matching_rows = [
        row for row in payload["regular_requests"] if row["id"] == request_id
    ]

    assert len(matching_rows) == 1
    assert matching_rows[0]["projeto"] == "P80"
    assert matching_rows[0]["projects"] == ["P80", "P83"]


def test_transport_dashboard_reflects_extra_override_for_multi_project_user_and_preserves_opposite_route_after_reload():
    monday = date(2026, 4, 20)
    timestamp = now_sgt()
    unique_suffix = make_test_key("T")
    workplace_name = f"Override Dashboard Hub {unique_suffix}"
    regular_plate = f"REG{unique_suffix}"
    extra_plate = f"EXT{unique_suffix}"
    user_key = make_test_key("T")

    with SessionLocal() as db:
        workplace = Workplace(
            workplace=workplace_name,
            address="70 Dashboard Street",
            zip="700070",
            country="Singapore",
        )
        regular_vehicle = Vehicle(
            placa=regular_plate,
            tipo="van",
            color="Silver",
            lugares=12,
            tolerance=8,
            service_scope="regular",
        )
        extra_vehicle = Vehicle(
            placa=extra_plate,
            tipo="carro",
            color="Red",
            lugares=4,
            tolerance=6,
            service_scope="extra",
        )
        db.add_all([workplace, regular_vehicle, extra_vehicle])
        db.flush()

        add_transport_schedule(
            db,
            vehicle=regular_vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=regular_vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=extra_vehicle,
            service_scope="extra",
            route_kind="home_to_work",
            recurrence_kind="single_date",
            service_date=monday,
        )

        user = User(
            rfid=None,
            nome="Override Dashboard Rider",
            chave=user_key,
            projeto="P80",
            workplace=workplace.workplace,
            end_rua="50 Override Road",
            zip="500050",
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()
        grant_user_project_memberships(db, user, ["P80", "P83"])

        regular_request = TransportRequest(
            user_id=user.id,
            request_kind="regular",
            recurrence_kind="weekday",
            requested_time="07:25",
            single_date=None,
            created_via="bot",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
            cancelled_at=None,
        )
        extra_request = TransportRequest(
            user_id=user.id,
            request_kind="extra",
            recurrence_kind="single_date",
            requested_time="07:55",
            single_date=monday,
            created_via="bot",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
            cancelled_at=None,
        )
        db.add_all([regular_request, extra_request])
        db.commit()

        regular_request_id = regular_request.id
        extra_request_id = extra_request.id
        regular_vehicle_id = regular_vehicle.id
        extra_vehicle_id = extra_vehicle.id

    with TestClient(app) as client:
        ensure_admin_session(client)

        regular_confirm_response = client.post(
            "/api/transport/assignments",
            json={
                "request_id": regular_request_id,
                "service_date": monday.isoformat(),
                "route_kind": "home_to_work",
                "status": "confirmed",
                "vehicle_id": regular_vehicle_id,
            },
        )
        assert regular_confirm_response.status_code == 200, regular_confirm_response.text

        extra_confirm_response = client.post(
            "/api/transport/assignments",
            json={
                "request_id": extra_request_id,
                "service_date": monday.isoformat(),
                "route_kind": "home_to_work",
                "status": "confirmed",
                "vehicle_id": extra_vehicle_id,
            },
        )
        assert extra_confirm_response.status_code == 200, extra_confirm_response.text

        blocked_reconfirm_response = client.post(
            "/api/transport/assignments",
            json={
                "request_id": regular_request_id,
                "service_date": monday.isoformat(),
                "route_kind": "home_to_work",
                "status": "confirmed",
                "vehicle_id": regular_vehicle_id,
            },
        )

        home_dashboard_response = client.get(
            "/api/transport/dashboard",
            params={"service_date": monday.isoformat(), "route_kind": "home_to_work"},
        )
        work_dashboard_response = client.get(
            "/api/transport/dashboard",
            params={"service_date": monday.isoformat(), "route_kind": "work_to_home"},
        )

    assert blocked_reconfirm_response.status_code == 409
    blocked_reconfirm_detail = extract_transport_structured_detail(blocked_reconfirm_response)
    assert blocked_reconfirm_detail["message_key"] == "status.couldNotUpdateAllocation"
    assert blocked_reconfirm_detail["error_code"] == "transport_assignment_save_failed"
    assert blocked_reconfirm_detail["technical_detail"] == (
        "The user already has a confirmed extra transport override for this date and route: home_to_work."
    )

    assert home_dashboard_response.status_code == 200, home_dashboard_response.text
    assert work_dashboard_response.status_code == 200, work_dashboard_response.text

    home_dashboard_payload = home_dashboard_response.json()
    work_dashboard_payload = work_dashboard_response.json()

    home_regular_request_rows = [
        row for row in home_dashboard_payload["regular_requests"] if row["id"] == regular_request_id
    ]
    home_extra_request_rows = [
        row for row in home_dashboard_payload["extra_requests"] if row["id"] == extra_request_id
    ]
    work_regular_request_rows = [
        row for row in work_dashboard_payload["regular_requests"] if row["id"] == regular_request_id
    ]
    assert len(home_regular_request_rows) == 1
    assert len(home_extra_request_rows) == 1
    assert len(work_regular_request_rows) == 1

    home_regular_request_row = home_regular_request_rows[0]
    home_extra_request_row = home_extra_request_rows[0]
    work_regular_request_row = work_regular_request_rows[0]
    extra_vehicle_registry_row = next(
        row for row in home_dashboard_payload["extra_vehicle_registry"] if row["vehicle_id"] == extra_vehicle_id
    )

    assert home_regular_request_row["projeto"] == "P80"
    assert home_regular_request_row["projects"] == ["P80", "P83"]
    assert home_regular_request_row["assignment_status"] == "pending"
    assert home_regular_request_row["assigned_vehicle"] is None
    assert all(
        row["assigned_vehicle"] is None or row["assigned_vehicle"]["placa"] != regular_plate
        for row in home_dashboard_payload["regular_requests"]
    )
    assert home_extra_request_row["projects"] == ["P80", "P83"]
    assert home_extra_request_row["assignment_status"] == "confirmed"
    assert home_extra_request_row["assigned_vehicle"]["placa"] == extra_plate
    assert extra_vehicle_registry_row["assigned_count"] == 1

    assert work_regular_request_row["projects"] == ["P80", "P83"]
    assert work_regular_request_row["assignment_status"] == "confirmed"
    assert work_regular_request_row["assigned_vehicle"]["placa"] == regular_plate

    with SessionLocal() as db:
        home_assignment = db.execute(
            select(TransportAssignment).where(
                TransportAssignment.request_id == regular_request_id,
                TransportAssignment.service_date == monday,
                TransportAssignment.route_kind == "home_to_work",
            )
        ).scalar_one()
        work_assignment = db.execute(
            select(TransportAssignment).where(
                TransportAssignment.request_id == regular_request_id,
                TransportAssignment.service_date == monday,
                TransportAssignment.route_kind == "work_to_home",
            )
        ).scalar_one()
        extra_assignment = db.execute(
            select(TransportAssignment).where(
                TransportAssignment.request_id == extra_request_id,
                TransportAssignment.service_date == monday,
                TransportAssignment.route_kind == "home_to_work",
            )
        ).scalar_one()

    assert home_assignment.status == "pending"
    assert home_assignment.vehicle_id is None
    assert work_assignment.status == "confirmed"
    assert work_assignment.vehicle_id == regular_vehicle_id
    assert extra_assignment.status == "confirmed"
    assert extra_assignment.vehicle_id == extra_vehicle_id


def test_transport_planning_fixture_bundle_supports_snapshot_and_settings_contracts():
    friday = date(2026, 5, 1)

    with SessionLocal() as db:
        fixture_bundle = create_transport_planning_fixture_bundle(db, service_date=friday)
        db.commit()

    try:
        with TestClient(app) as client:
            ensure_admin_session(client)
            snapshot_response = client.get(
                "/api/transport/operational-snapshot",
                params={"service_date": friday.isoformat(), "route_kind": "home_to_work"},
            )
            settings_response = client.get("/api/transport/settings")

        assert snapshot_response.status_code == 200, snapshot_response.text
        assert settings_response.status_code == 200, settings_response.text

        snapshot_payload = snapshot_response.json()
        settings_payload = settings_response.json()

        for project_payload in fixture_bundle["projects"].values():
            project_row = next(
                row for row in snapshot_payload["projects"] if row["id"] == project_payload["id"]
            )
            assert project_row["name"] == project_payload["name"]
            assert project_row["country_code"] == project_payload["country_code"]
            assert project_row["country_name"] == project_payload["country_name"]
            assert project_row["address"] == project_payload["address"]
            assert project_row["zip_code"] == project_payload["zip_code"]

        regular_request = next(
            row for row in snapshot_payload["regular_requests"] if row["id"] == fixture_bundle["requests"]["regular"]["id"]
        )
        weekend_request = next(
            row for row in snapshot_payload["weekend_requests"] if row["id"] == fixture_bundle["requests"]["weekend"]["id"]
        )
        extra_request = next(
            row for row in snapshot_payload["extra_requests"] if row["id"] == fixture_bundle["requests"]["extra"]["id"]
        )

        assert regular_request["projeto"] == fixture_bundle["requests"]["regular"]["projeto"]
        assert regular_request["end_rua"] == fixture_bundle["requests"]["regular"]["end_rua"]
        assert regular_request["zip"] == fixture_bundle["requests"]["regular"]["zip"]

        assert weekend_request["projeto"] == fixture_bundle["requests"]["weekend"]["projeto"]
        assert weekend_request["end_rua"] == fixture_bundle["requests"]["weekend"]["end_rua"]
        assert weekend_request["zip"] == fixture_bundle["requests"]["weekend"]["zip"]

        assert extra_request["projeto"] == fixture_bundle["requests"]["extra"]["projeto"]
        assert extra_request["end_rua"] == fixture_bundle["requests"]["extra"]["end_rua"]
        assert extra_request["zip"] == fixture_bundle["requests"]["extra"]["zip"]

        assert any(
            row["vehicle_id"] == fixture_bundle["vehicles"]["regular"]["vehicle_id"]
            for row in snapshot_payload["regular_vehicle_registry"]
        )
        assert any(
            row["vehicle_id"] == fixture_bundle["vehicles"]["weekend"]["vehicle_id"]
            for row in snapshot_payload["weekend_vehicle_registry"]
        )
        assert any(
            row["vehicle_id"] == fixture_bundle["vehicles"]["extra"]["vehicle_id"]
            for row in snapshot_payload["extra_vehicle_registry"]
        )

        settings_current = fixture_bundle["settings_context"]["current"]
        assert snapshot_payload["dashboard_generated_at"] == snapshot_payload["captured_at"]
        assert snapshot_payload["arrive_at_work_time"] == settings_current["arrive_at_work_time"]
        assert settings_payload["arrive_at_work_time"] == settings_current["arrive_at_work_time"]
        assert settings_payload["default_car_seats"] == settings_current["default_car_seats"]
        assert settings_payload["default_minivan_seats"] == settings_current["default_minivan_seats"]
        assert settings_payload["default_van_seats"] == settings_current["default_van_seats"]
        assert settings_payload["default_bus_seats"] == settings_current["default_bus_seats"]
        assert settings_payload["default_car_price"] == settings_current["default_car_price"]
        assert settings_payload["default_minivan_price"] == settings_current["default_minivan_price"]
        assert settings_payload["default_van_price"] == settings_current["default_van_price"]
        assert settings_payload["default_bus_price"] == settings_current["default_bus_price"]
        assert settings_payload["price_currency_code"] == settings_current["price_currency_code"]
        assert settings_payload["price_rate_unit"] == settings_current["price_rate_unit"]
    finally:
        with SessionLocal() as db:
            cleanup_transport_planning_fixture_bundle(db, fixture_bundle)


def test_transport_operational_proposal_represents_decisions_without_applying_assignments():
    friday = date(2026, 4, 17)
    captured_at = datetime(2026, 4, 16, 21, 0, tzinfo=ZoneInfo(settings.tz_name))
    created_at = datetime(2026, 4, 16, 21, 5, tzinfo=ZoneInfo(settings.tz_name))

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="18:10")
        workplace = Workplace(
            workplace="Proposal Hub",
            address="41 Proposal Road",
            zip="941001",
            country="Singapore",
        )
        db.add(workplace)
        db.flush()

        vehicle = Vehicle(placa="PRP4101", tipo="van", color="White", lugares=10, tolerance=8, service_scope="regular")
        db.add(vehicle)
        db.flush()

        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )

        user = User(
            rfid=None,
            nome="Proposal Rider",
            chave="PR41",
            projeto="P41",
            workplace=workplace.workplace,
            placa=vehicle.placa,
            end_rua="41 Proposal Avenue",
            zip="941002",
            local=None,
            checkin=None,
            time=None,
            last_active_at=created_at,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        transport_request = TransportRequest(
            user_id=user.id,
            request_kind="regular",
            recurrence_kind="weekday",
            requested_time="07:30",
            selected_weekdays_json=json.dumps([0, 1, 2, 3, 4]),
            single_date=None,
            created_via="admin",
            status="active",
            created_at=created_at,
            updated_at=created_at,
        )
        db.add(transport_request)
        db.commit()

        request_id = transport_request.id
        vehicle_id = vehicle.id

    with SessionLocal() as db:
        snapshot = transport_proposal_service_module.build_transport_operational_snapshot(
            db,
            service_date=friday,
            route_kind="home_to_work",
            captured_at=captured_at,
        )
        proposal = transport_proposal_service_module.build_transport_operational_proposal(
            snapshot=snapshot,
            origin="manual",
            created_at=created_at,
            decisions=[
                TransportProposalDecision(
                    request_id=request_id,
                    request_kind="regular",
                    service_date=friday,
                    route_kind="home_to_work",
                    suggested_status="confirmed",
                    vehicle_id=vehicle_id,
                    rationale="Manual planning draft for weekday route.",
                )
            ],
        )
        persisted_assignments = db.execute(
            select(TransportAssignment).where(TransportAssignment.request_id == request_id)
        ).scalars().all()

    assert snapshot.service_date == friday
    assert snapshot.route_kind == "home_to_work"
    assert snapshot.snapshot_key.startswith("transport-snapshot:2026-04-17:home_to_work:")
    snapshot_request = next(row for row in snapshot.regular_requests if row.id == request_id)
    assert snapshot_request.assignment_status == "pending"
    assert snapshot_request.id == request_id

    assert proposal.proposal_status == "draft"
    assert proposal.origin == "manual"
    assert proposal.created_at == created_at
    assert proposal.snapshot.snapshot_key == snapshot.snapshot_key
    assert proposal.summary.total_snapshot_requests == (
        len(snapshot.regular_requests) + len(snapshot.weekend_requests) + len(snapshot.extra_requests)
    )
    assert proposal.summary.total_snapshot_vehicles == (
        len(snapshot.regular_vehicles)
        + len(snapshot.weekend_vehicles)
        + len(snapshot.extra_vehicles)
    )
    assert proposal.summary.total_decisions == 1
    assert proposal.summary.confirmed_decisions == 1
    assert proposal.summary.rejected_decisions == 0
    assert proposal.summary.pending_decisions == 0
    assert proposal.decisions[0].vehicle_id == vehicle_id
    assert proposal.decisions[0].suggested_status == "confirmed"
    assert persisted_assignments == []


def test_transport_proposal_validate_and_approve_commands_record_audit_without_applying_assignments():
    friday = date(2026, 4, 17)
    captured_at = datetime(2026, 4, 16, 21, 0, tzinfo=ZoneInfo(settings.tz_name))
    created_at = datetime(2026, 4, 16, 21, 5, tzinfo=ZoneInfo(settings.tz_name))

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="18:10")
        workplace = Workplace(
            workplace="Proposal Review Hub",
            address="42 Review Road",
            zip="941101",
            country="Singapore",
        )
        db.add(workplace)
        db.flush()

        vehicle = Vehicle(placa="PRP4201", tipo="van", color="Silver", lugares=10, tolerance=8, service_scope="regular")
        db.add(vehicle)
        db.flush()

        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )

        user = User(
            rfid=None,
            nome="Review Rider",
            chave="PR42",
            projeto="P42",
            workplace=workplace.workplace,
            placa=vehicle.placa,
            end_rua="42 Review Avenue",
            zip="941102",
            local=None,
            checkin=None,
            time=None,
            last_active_at=created_at,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        transport_request = TransportRequest(
            user_id=user.id,
            request_kind="regular",
            recurrence_kind="weekday",
            requested_time="07:40",
            selected_weekdays_json=json.dumps([0, 1, 2, 3, 4]),
            single_date=None,
            created_via="admin",
            status="active",
            created_at=created_at,
            updated_at=created_at,
        )
        db.add(transport_request)
        db.commit()

        proposal = transport_proposal_service_module.build_transport_operational_proposal(
            snapshot=transport_proposal_service_module.build_transport_operational_snapshot(
                db,
                service_date=friday,
                route_kind="home_to_work",
                captured_at=captured_at,
            ),
            origin="manual",
            created_at=created_at,
            decisions=[
                TransportProposalDecision(
                    request_id=transport_request.id,
                    request_kind="regular",
                    service_date=friday,
                    route_kind="home_to_work",
                    suggested_status="confirmed",
                    vehicle_id=vehicle.id,
                    rationale="Planner approved this pairing.",
                )
            ],
        )

    with TestClient(app) as client:
        ensure_admin_session(client)

        validate_response = client.post(
            "/api/transport/proposals/validate",
            json=proposal.model_dump(mode="json"),
        )
        approve_response = client.post(
            "/api/transport/proposals/approve",
            json=proposal.model_dump(mode="json"),
        )

    assert validate_response.status_code == 200, validate_response.text
    assert approve_response.status_code == 200, approve_response.text

    validate_payload = validate_response.json()
    approve_payload = approve_response.json()

    assert validate_payload["ok"] is True
    assert validate_payload["message"] == "Proposal validation passed without blocking issues."
    assert validate_payload["proposal"]["proposal_status"] == "draft"
    assert validate_payload["proposal"]["validation_issues"] == []
    assert validate_payload["proposal"]["audit_trail"][0]["action"] == "validated"
    assert validate_payload["proposal"]["audit_trail"][0]["outcome"] == "passed"
    assert validate_payload["proposal"]["audit_trail"][0]["actor"]["chave"] == ADMIN_LOGIN_CHAVE

    assert approve_payload["ok"] is True
    assert approve_payload["message"] == "Proposal approved without applying assignments."
    assert approve_payload["proposal"]["proposal_status"] == "approved"
    assert approve_payload["proposal"]["validation_issues"] == []
    assert [entry["action"] for entry in approve_payload["proposal"]["audit_trail"]] == ["validated", "approved"]
    assert [entry["outcome"] for entry in approve_payload["proposal"]["audit_trail"]] == ["passed", "approved"]

    with SessionLocal() as db:
        persisted_assignments = db.execute(
            select(TransportAssignment).where(TransportAssignment.request_id == transport_request.id)
        ).scalars().all()

    assert persisted_assignments == []


def test_transport_proposal_approval_reports_blocking_inconsistencies():
    friday = date(2026, 4, 17)
    captured_at = datetime(2026, 4, 16, 21, 0, tzinfo=ZoneInfo(settings.tz_name))
    created_at = datetime(2026, 4, 16, 21, 5, tzinfo=ZoneInfo(settings.tz_name))

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="18:10")
        workplace = Workplace(
            workplace="Proposal Drift Hub",
            address="43 Drift Road",
            zip="941201",
            country="Singapore",
        )
        db.add(workplace)
        db.flush()

        vehicle = Vehicle(placa="PRP4202", tipo="van", color="Black", lugares=8, tolerance=8, service_scope="regular")
        db.add(vehicle)
        db.flush()

        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )

        user = User(
            rfid=None,
            nome="Drift Rider",
            chave="PR43",
            projeto="P43",
            workplace=workplace.workplace,
            placa=vehicle.placa,
            end_rua="43 Drift Avenue",
            zip="941202",
            local=None,
            checkin=None,
            time=None,
            last_active_at=created_at,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        transport_request = TransportRequest(
            user_id=user.id,
            request_kind="regular",
            recurrence_kind="weekday",
            requested_time="07:50",
            selected_weekdays_json=json.dumps([0, 1, 2, 3, 4]),
            single_date=None,
            created_via="admin",
            status="active",
            created_at=created_at,
            updated_at=created_at,
        )
        db.add(transport_request)
        db.commit()

        proposal = transport_proposal_service_module.build_transport_operational_proposal(
            snapshot=transport_proposal_service_module.build_transport_operational_snapshot(
                db,
                service_date=friday,
                route_kind="home_to_work",
                captured_at=captured_at,
            ),
            origin="manual",
            created_at=created_at,
            decisions=[
                TransportProposalDecision(
                    request_id=transport_request.id,
                    request_kind="regular",
                    service_date=friday,
                    route_kind="home_to_work",
                    suggested_status="confirmed",
                    vehicle_id=vehicle.id,
                    rationale="Vehicle looked available when the snapshot was captured.",
                )
            ],
        )

        schedule = db.execute(
            select(TransportVehicleSchedule).where(
                TransportVehicleSchedule.vehicle_id == vehicle.id,
                TransportVehicleSchedule.route_kind == "home_to_work",
            )
        ).scalar_one()
        db.delete(schedule)
        db.commit()

    with TestClient(app) as client:
        ensure_admin_session(client)
        approve_response = client.post(
            "/api/transport/proposals/approve",
            json=proposal.model_dump(mode="json"),
        )

    assert approve_response.status_code == 200, approve_response.text
    approve_payload = approve_response.json()

    assert approve_payload["ok"] is False
    assert approve_payload["message"] == "Proposal approval was blocked by validation issues."
    assert approve_payload["proposal"]["proposal_status"] == "draft"
    assert approve_payload["proposal"]["validation_issues"][0]["code"] == "vehicle_unavailable"
    assert "not currently available" in approve_payload["proposal"]["validation_issues"][0]["message"]
    assert [entry["action"] for entry in approve_payload["proposal"]["audit_trail"]] == ["validated", "approved"]
    assert [entry["outcome"] for entry in approve_payload["proposal"]["audit_trail"]] == ["blocked", "blocked"]
    assert approve_payload["proposal"]["audit_trail"][1]["result"]["validation_issue_codes"] == [
        "vehicle_unavailable"
    ]


def test_transport_proposal_build_contract_returns_fresh_snapshot_backed_draft():
    friday = date(2026, 4, 17)
    created_at = datetime(2026, 4, 16, 21, 5, tzinfo=ZoneInfo(settings.tz_name))
    captured_at = datetime(2026, 4, 16, 21, 0, tzinfo=ZoneInfo(settings.tz_name))

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="18:10")
        workplace = Workplace(
            workplace="Contract Build Hub",
            address="62 Contract Lane",
            zip="941621",
            country="Singapore",
        )
        db.add(workplace)
        db.flush()

        vehicle = Vehicle(placa="CNT6201", tipo="van", color="Silver", lugares=10, tolerance=8, service_scope="regular")
        db.add(vehicle)
        db.flush()

        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )

        user = User(
            rfid=None,
            nome="Contract Builder",
            chave="CB62",
            projeto="P62",
            workplace=workplace.workplace,
            placa=vehicle.placa,
            end_rua="62 Contract Avenue",
            zip="941622",
            local=None,
            checkin=None,
            time=None,
            last_active_at=created_at,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        transport_request = TransportRequest(
            user_id=user.id,
            request_kind="regular",
            recurrence_kind="weekday",
            requested_time="07:35",
            selected_weekdays_json=json.dumps([0, 1, 2, 3, 4]),
            single_date=None,
            created_via="admin",
            status="active",
            created_at=created_at,
            updated_at=created_at,
        )
        db.add(transport_request)
        db.commit()
        vehicle_id = vehicle.id
        request_id = transport_request.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        build_response = client.post(
            "/api/transport/proposals/build",
            json={
                "service_date": friday.isoformat(),
                "route_kind": "home_to_work",
                "origin": "manual",
                "replaces_proposal_key": "transport-proposal:previous-draft-62",
                "captured_at": captured_at.isoformat(),
                "created_at": created_at.isoformat(),
                "decisions": [
                    {
                        "request_id": request_id,
                        "request_kind": "regular",
                        "service_date": friday.isoformat(),
                        "route_kind": "home_to_work",
                        "suggested_status": "confirmed",
                        "vehicle_id": vehicle_id,
                        "rationale": "Build a contract-backed draft from a stable snapshot.",
                    }
                ],
            },
        )

    assert build_response.status_code == 200, build_response.text
    payload = build_response.json()
    assert payload["proposal_status"] == "draft"
    assert payload["snapshot"]["service_date"] == friday.isoformat()
    assert payload["snapshot"]["route_kind"] == "home_to_work"
    snapshot_request = next(row for row in payload["snapshot"]["regular_requests"] if row["id"] == request_id)
    assert snapshot_request["id"] == request_id
    assert payload["summary"]["total_snapshot_requests"] == (
        len(payload["snapshot"]["regular_requests"])
        + len(payload["snapshot"]["weekend_requests"])
        + len(payload["snapshot"]["extra_requests"])
    )
    assert payload["summary"]["confirmed_decisions"] == 1
    assert payload["validation_issues"] == []
    assert payload["replaces_proposal_key"] == "transport-proposal:previous-draft-62"
    assert [entry["action"] for entry in payload["audit_trail"]] == ["generated"]
    assert payload["audit_trail"][0]["outcome"] == "generated"
    assert payload["audit_trail"][0]["actor"]["chave"] == ADMIN_LOGIN_CHAVE
    assert payload["audit_trail"][0]["context"]["proposal_origin"] == "manual"
    assert payload["audit_trail"][0]["context"]["proposal_snapshot_key"] == payload["snapshot"]["snapshot_key"]
    assert payload["audit_trail"][0]["context"]["evaluation_snapshot_key"] == payload["snapshot"]["snapshot_key"]
    assert payload["audit_trail"][0]["context"]["decision_request_ids"] == [request_id]
    assert payload["audit_trail"][0]["context"]["decision_vehicle_ids"] == [vehicle_id]
    assert payload["audit_trail"][0]["context"]["replaces_proposal_key"] == "transport-proposal:previous-draft-62"
    assert payload["audit_trail"][0]["result"]["proposal_status"] == "draft"
    assert payload["audit_trail"][0]["result"]["validation_issue_count"] == 0
    assert payload["audit_trail"][0]["result"]["applied_assignment_count"] == 0


def test_transport_operational_snapshot_endpoint_returns_contract_for_selected_date_and_route():
    friday = date(2026, 4, 17)
    created_at = datetime(2026, 4, 16, 21, 5, tzinfo=ZoneInfo(settings.tz_name))

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="18:10")
        workplace = Workplace(
            workplace="Snapshot Contract Hub",
            address="66 Contract Lane",
            zip="941661",
            country="Singapore",
        )
        db.add(workplace)
        db.flush()

        vehicle = Vehicle(placa="SNP6601", tipo="van", color="White", lugares=10, tolerance=8, service_scope="regular")
        db.add(vehicle)
        db.flush()

        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )

        user = User(
            rfid=None,
            nome="Snapshot Rider",
            chave="SN66",
            projeto="P66",
            workplace=workplace.workplace,
            placa=vehicle.placa,
            end_rua="66 Snapshot Avenue",
            zip="941662",
            local=None,
            checkin=None,
            time=None,
            last_active_at=created_at,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        transport_request = TransportRequest(
            user_id=user.id,
            request_kind="regular",
            recurrence_kind="weekday",
            requested_time="07:10",
            selected_weekdays_json=json.dumps([0, 1, 2, 3, 4]),
            single_date=None,
            created_via="admin",
            status="active",
            created_at=created_at,
            updated_at=created_at,
        )
        db.add(transport_request)
        db.commit()
        request_id = transport_request.id
        vehicle_id = vehicle.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        snapshot_response = client.get(
            "/api/transport/operational-snapshot",
            params={"service_date": friday.isoformat(), "route_kind": "home_to_work"},
        )

    assert snapshot_response.status_code == 200, snapshot_response.text
    payload = snapshot_response.json()
    assert payload["service_date"] == friday.isoformat()
    assert payload["route_kind"] == "home_to_work"
    assert payload["snapshot_key"].startswith("transport-snapshot:2026-04-17:home_to_work:")
    snapshot_request = next(row for row in payload["regular_requests"] if row["id"] == request_id)
    assert snapshot_request["assignment_status"] == "pending"
    assert any(row["vehicle_id"] == vehicle_id for row in payload["regular_vehicle_registry"])
    assert any(row["workplace"] == "Snapshot Contract Hub" for row in payload["workplaces"])


def test_transport_proposal_apply_contract_persists_assignments_after_approval():
    transport_reevaluation_module.clear_transport_reevaluation_events()
    friday = date(2026, 4, 17)
    created_at = datetime(2026, 4, 16, 21, 5, tzinfo=ZoneInfo(settings.tz_name))
    captured_at = datetime(2026, 4, 16, 21, 0, tzinfo=ZoneInfo(settings.tz_name))
    contract_key = make_test_key("C")

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="18:10")
        workplace = Workplace(
            workplace=f"Contract Apply Hub {contract_key}",
            address="63 Contract Lane",
            zip="941631",
            country="Singapore",
        )
        db.add(workplace)
        db.flush()

        vehicle = Vehicle(
            placa=f"CNT{contract_key}",
            tipo="van",
            color="White",
            lugares=10,
            tolerance=8,
            service_scope="regular",
        )
        db.add(vehicle)
        db.flush()

        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )

        user = User(
            rfid=None,
            nome="Contract Applier",
            chave=contract_key,
            projeto="P63",
            workplace=workplace.workplace,
            placa=vehicle.placa,
            end_rua="63 Contract Avenue",
            zip="941632",
            local=None,
            checkin=None,
            time=None,
            last_active_at=created_at,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        transport_request = TransportRequest(
            user_id=user.id,
            request_kind="regular",
            recurrence_kind="weekday",
            requested_time="07:25",
            selected_weekdays_json=json.dumps([0, 1, 2, 3, 4]),
            single_date=None,
            created_via="admin",
            status="active",
            created_at=created_at,
            updated_at=created_at,
        )
        db.add(transport_request)
        db.commit()
        vehicle_id = vehicle.id
        request_id = transport_request.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        build_response = client.post(
            "/api/transport/proposals/build",
            json={
                "service_date": friday.isoformat(),
                "route_kind": "home_to_work",
                "origin": "manual",
                "captured_at": captured_at.isoformat(),
                "created_at": created_at.isoformat(),
                "decisions": [
                    {
                        "request_id": request_id,
                        "request_kind": "regular",
                        "service_date": friday.isoformat(),
                        "route_kind": "home_to_work",
                        "suggested_status": "confirmed",
                        "vehicle_id": vehicle_id,
                        "boarding_time": "07:05",
                        "response_message": "Assigned through proposal application.",
                        "rationale": "Use the contract apply flow after review.",
                    }
                ],
            },
        )
        assert build_response.status_code == 200, build_response.text

        approve_response = client.post(
            "/api/transport/proposals/approve",
            json=build_response.json(),
        )
        assert approve_response.status_code == 200, approve_response.text

        apply_response = client.post(
            "/api/transport/proposals/apply",
            json={"proposal": approve_response.json()["proposal"]},
        )
        events_response = client.get("/api/transport/reevaluation-events", params={"limit": 3})

    assert apply_response.status_code == 200, apply_response.text
    assert events_response.status_code == 200, events_response.text
    payload = apply_response.json()
    events_payload = events_response.json()
    assert payload["ok"] is True
    assert payload["message"] == "Proposal applied to transport assignments."
    assert payload["proposal"]["proposal_status"] == "applied"
    assert payload["applied_assignments"][0]["request_id"] == request_id
    assert payload["applied_assignments"][0]["status"] == "confirmed"
    assert payload["applied_assignments"][0]["vehicle_id"] == vehicle_id
    assert [entry["action"] for entry in payload["proposal"]["audit_trail"]] == ["generated", "validated", "approved", "validated", "applied"]
    assert [entry["outcome"] for entry in payload["proposal"]["audit_trail"]] == ["generated", "passed", "approved", "passed", "applied"]
    assert payload["proposal"]["audit_trail"][1]["context"]["evaluation_snapshot_key"].startswith(
        "transport-snapshot:2026-04-17:home_to_work:"
    )
    assert payload["proposal"]["audit_trail"][-1]["result"]["applied_assignment_count"] == len(
        payload["applied_assignments"]
    )
    assert payload["proposal"]["audit_trail"][-1]["result"]["applied_assignment_ids"] == [
        row["assignment_id"] for row in payload["applied_assignments"]
    ]
    assert payload["proposal"]["audit_trail"][-1]["result"]["proposal_status"] == "applied"
    assert events_payload["recent_events"][0]["event_type"] == "transport_assignment_changed"
    assert events_payload["recent_events"][0]["source"] == "transport_proposal"
    assert events_payload["recent_events"][0]["proposal_key"] == payload["proposal"]["proposal_key"]
    assert events_payload["recent_events"][0]["route_kind"] == "home_to_work"

    with SessionLocal() as db:
        persisted_assignments = db.execute(
            select(TransportAssignment).where(TransportAssignment.request_id == request_id)
        ).scalars().all()

    assert len(persisted_assignments) == 2
    assert {assignment.route_kind for assignment in persisted_assignments} == {"home_to_work", "work_to_home"}
    assert all(assignment.status == "confirmed" for assignment in persisted_assignments)
    assert all(assignment.vehicle_id == vehicle_id for assignment in persisted_assignments)
    home_assignment = next(assignment for assignment in persisted_assignments if assignment.route_kind == "home_to_work")
    work_assignment = next(assignment for assignment in persisted_assignments if assignment.route_kind == "work_to_home")
    assert home_assignment.boarding_time == "07:05"
    assert work_assignment.boarding_time is None


def test_transport_proposal_apply_contract_blocks_draft_proposal():
    friday = date(2026, 4, 17)
    created_at = datetime(2026, 4, 16, 21, 5, tzinfo=ZoneInfo(settings.tz_name))

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="18:10")
        workplace = Workplace(
            workplace="Contract Draft Guard Hub",
            address="64 Contract Lane",
            zip="941641",
            country="Singapore",
        )
        db.add(workplace)
        db.flush()

        vehicle = Vehicle(placa="CNT6401", tipo="van", color="Blue", lugares=10, tolerance=8, service_scope="regular")
        db.add(vehicle)
        db.flush()

        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )

        user = User(
            rfid=None,
            nome="Draft Guard Rider",
            chave="DG64",
            projeto="P64",
            workplace=workplace.workplace,
            placa=vehicle.placa,
            end_rua="64 Contract Avenue",
            zip="941642",
            local=None,
            checkin=None,
            time=None,
            last_active_at=created_at,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        transport_request = TransportRequest(
            user_id=user.id,
            request_kind="regular",
            recurrence_kind="weekday",
            requested_time="07:20",
            selected_weekdays_json=json.dumps([0, 1, 2, 3, 4]),
            single_date=None,
            created_via="admin",
            status="active",
            created_at=created_at,
            updated_at=created_at,
        )
        db.add(transport_request)
        db.commit()
        request_id = transport_request.id
        vehicle_id = vehicle.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        build_response = client.post(
            "/api/transport/proposals/build",
            json={
                "service_date": friday.isoformat(),
                "route_kind": "home_to_work",
                "origin": "manual",
                "decisions": [
                    {
                        "request_id": request_id,
                        "request_kind": "regular",
                        "service_date": friday.isoformat(),
                        "route_kind": "home_to_work",
                        "suggested_status": "confirmed",
                        "vehicle_id": vehicle_id,
                        "rationale": "Draft proposals must not bypass approval.",
                    }
                ],
            },
        )
        assert build_response.status_code == 200, build_response.text

        apply_response = client.post(
            "/api/transport/proposals/apply",
            json={"proposal": build_response.json()},
        )

    assert apply_response.status_code == 200, apply_response.text
    payload = apply_response.json()
    assert payload["ok"] is False
    assert payload["message"] == "Proposal application was blocked by validation issues."
    assert payload["proposal"]["proposal_status"] == "draft"
    assert payload["applied_assignments"] == []
    assert payload["proposal"]["validation_issues"][-1]["code"] == "proposal_not_approved_for_application"
    assert [entry["action"] for entry in payload["proposal"]["audit_trail"]] == ["generated", "validated", "applied"]
    assert payload["proposal"]["audit_trail"][-1]["result"]["validation_issue_codes"][-1] == (
        "proposal_not_approved_for_application"
    )
    assert payload["proposal"]["audit_trail"][-1]["result"]["applied_assignment_count"] == 0

    with SessionLocal() as db:
        persisted_assignments = db.execute(
            select(TransportAssignment).where(TransportAssignment.request_id == request_id)
        ).scalars().all()

    assert persisted_assignments == []


def test_transport_proposal_apply_contract_revalidates_after_operational_drift():
    friday = date(2026, 4, 17)
    created_at = datetime(2026, 4, 16, 21, 5, tzinfo=ZoneInfo(settings.tz_name))
    captured_at = datetime(2026, 4, 16, 21, 0, tzinfo=ZoneInfo(settings.tz_name))

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="18:10")
        workplace = Workplace(
            workplace="Drift Apply Hub",
            address="67 Contract Lane",
            zip="941671",
            country="Singapore",
        )
        db.add(workplace)
        db.flush()

        vehicle = Vehicle(placa="DRF6701", tipo="van", color="Black", lugares=10, tolerance=8, service_scope="regular")
        db.add(vehicle)
        db.flush()

        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )

        user = User(
            rfid=None,
            nome="Drift Apply Rider",
            chave="DA67",
            projeto="P67",
            workplace=workplace.workplace,
            placa=vehicle.placa,
            end_rua="67 Drift Avenue",
            zip="941672",
            local=None,
            checkin=None,
            time=None,
            last_active_at=created_at,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        transport_request = TransportRequest(
            user_id=user.id,
            request_kind="regular",
            recurrence_kind="weekday",
            requested_time="07:18",
            selected_weekdays_json=json.dumps([0, 1, 2, 3, 4]),
            single_date=None,
            created_via="admin",
            status="active",
            created_at=created_at,
            updated_at=created_at,
        )
        db.add(transport_request)
        db.commit()
        request_id = transport_request.id
        vehicle_id = vehicle.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        build_response = client.post(
            "/api/transport/proposals/build",
            json={
                "service_date": friday.isoformat(),
                "route_kind": "home_to_work",
                "origin": "manual",
                "captured_at": captured_at.isoformat(),
                "created_at": created_at.isoformat(),
                "decisions": [
                    {
                        "request_id": request_id,
                        "request_kind": "regular",
                        "service_date": friday.isoformat(),
                        "route_kind": "home_to_work",
                        "suggested_status": "confirmed",
                        "vehicle_id": vehicle_id,
                        "rationale": "Apply should revalidate after approval.",
                    }
                ],
            },
        )
        assert build_response.status_code == 200, build_response.text

        approve_response = client.post(
            "/api/transport/proposals/approve",
            json=build_response.json(),
        )
        assert approve_response.status_code == 200, approve_response.text
        approved_payload = approve_response.json()["proposal"]

    with SessionLocal() as db:
        schedule = db.execute(
            select(TransportVehicleSchedule).where(
                TransportVehicleSchedule.vehicle_id == vehicle_id,
                TransportVehicleSchedule.route_kind == "home_to_work",
            )
        ).scalar_one()
        db.delete(schedule)
        db.commit()

    with TestClient(app) as client:
        ensure_admin_session(client)
        apply_response = client.post(
            "/api/transport/proposals/apply",
            json={"proposal": approved_payload},
        )

    assert apply_response.status_code == 200, apply_response.text
    payload = apply_response.json()
    assert payload["ok"] is False
    assert payload["message"] == "Proposal application was blocked by validation issues."
    assert payload["proposal"]["proposal_status"] == "approved"
    assert payload["applied_assignments"] == []
    assert payload["proposal"]["validation_issues"][0]["code"] == "vehicle_unavailable"
    assert payload["proposal"]["audit_trail"][-1]["action"] == "applied"
    assert payload["proposal"]["audit_trail"][-1]["outcome"] == "blocked"
    assert payload["proposal"]["audit_trail"][-1]["result"]["validation_issue_codes"] == [
        "vehicle_unavailable"
    ]

    with SessionLocal() as db:
        persisted_assignments = db.execute(
            select(TransportAssignment).where(TransportAssignment.request_id == request_id)
        ).scalars().all()

    assert persisted_assignments == []


def test_transport_proposal_reject_records_structured_audit_metadata():
    friday = date(2026, 4, 17)
    created_at = datetime(2026, 4, 16, 21, 5, tzinfo=ZoneInfo(settings.tz_name))
    captured_at = datetime(2026, 4, 16, 21, 0, tzinfo=ZoneInfo(settings.tz_name))

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="18:10")
        workplace = Workplace(
            workplace="Contract Reject Hub",
            address="65 Contract Lane",
            zip="941651",
            country="Singapore",
        )
        db.add(workplace)
        db.flush()

        vehicle = Vehicle(placa="CNT6501", tipo="van", color="Gray", lugares=10, tolerance=8, service_scope="regular")
        db.add(vehicle)
        db.flush()

        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )

        user = User(
            rfid=None,
            nome="Contract Reject Rider",
            chave="CR65",
            projeto="P65",
            workplace=workplace.workplace,
            placa=vehicle.placa,
            end_rua="65 Contract Avenue",
            zip="941652",
            local=None,
            checkin=None,
            time=None,
            last_active_at=created_at,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        transport_request = TransportRequest(
            user_id=user.id,
            request_kind="regular",
            recurrence_kind="weekday",
            requested_time="07:15",
            selected_weekdays_json=json.dumps([0, 1, 2, 3, 4]),
            single_date=None,
            created_via="admin",
            status="active",
            created_at=created_at,
            updated_at=created_at,
        )
        db.add(transport_request)
        db.commit()
        request_id = transport_request.id
        vehicle_id = vehicle.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        build_response = client.post(
            "/api/transport/proposals/build",
            json={
                "service_date": friday.isoformat(),
                "route_kind": "home_to_work",
                "origin": "manual",
                "captured_at": captured_at.isoformat(),
                "created_at": created_at.isoformat(),
                "decisions": [
                    {
                        "request_id": request_id,
                        "request_kind": "regular",
                        "service_date": friday.isoformat(),
                        "route_kind": "home_to_work",
                        "suggested_status": "confirmed",
                        "vehicle_id": vehicle_id,
                        "rationale": "This proposal will be rejected during review.",
                    }
                ],
            },
        )
        assert build_response.status_code == 200, build_response.text

        reject_response = client.post(
            "/api/transport/proposals/reject",
            json={
                "proposal": build_response.json(),
                "message": "Rejected after supervisor review.",
            },
        )

    assert reject_response.status_code == 200, reject_response.text
    payload = reject_response.json()
    assert payload["ok"] is True
    assert payload["proposal"]["proposal_status"] == "rejected"
    assert [entry["action"] for entry in payload["proposal"]["audit_trail"]] == ["generated", "rejected"]
    assert payload["proposal"]["audit_trail"][-1]["context"]["decision_request_ids"] == [request_id]
    assert payload["proposal"]["audit_trail"][-1]["context"]["decision_vehicle_ids"] == [vehicle_id]
    assert payload["proposal"]["audit_trail"][-1]["result"]["proposal_status"] == "rejected"
    assert payload["proposal"]["audit_trail"][-1]["result"]["validation_issue_count"] == 0


def test_transport_vehicle_registration_creates_route_aware_schedules():
    friday = date(2026, 4, 17)
    saturday = date(2026, 4, 18)
    sunday = date(2026, 4, 19)

    with TestClient(app) as client:
        ensure_admin_session(client)

        extra_response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "extra",
                "service_date": friday.isoformat(),
                "route_kind": "work_to_home",
                "departure_time": "17:45",
                "tipo": "carro",
                "placa": "EXT7001",
                "color": "Red",
                "lugares": 4,
                "tolerance": 6,
            },
        )
        weekend_response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "weekend",
                "service_date": sunday.isoformat(),
                "every_saturday": True,
                "every_sunday": True,
                "tipo": "van",
                "placa": "WKD7001",
                "color": "Black",
                "lugares": 10,
                "tolerance": 12,
            },
        )
        regular_response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "regular",
                "service_date": friday.isoformat(),
                "tipo": "minivan",
                "placa": "REG7001",
                "color": "White",
                "lugares": 7,
                "tolerance": 9,
            },
        )

    assert extra_response.status_code == 200
    assert weekend_response.status_code == 200
    assert regular_response.status_code == 200

    with SessionLocal() as db:
        extra_vehicle = db.execute(select(Vehicle).where(Vehicle.placa == "EXT7001")).scalar_one()
        weekend_vehicle = db.execute(select(Vehicle).where(Vehicle.placa == "WKD7001")).scalar_one()
        regular_vehicle = db.execute(select(Vehicle).where(Vehicle.placa == "REG7001")).scalar_one()
        active_schedule_rows = transport_service_module._list_active_transport_schedule_rows(db)
        friday_work_to_home_departure_time = transport_service_module.get_transport_work_to_home_time_for_date(
            db,
            service_date=friday,
        )
        saturday_work_to_home_departure_time = transport_service_module.get_transport_work_to_home_time_for_date(
            db,
            service_date=saturday,
        )

        extra_schedules = db.execute(
            select(TransportVehicleSchedule)
            .where(TransportVehicleSchedule.vehicle_id == extra_vehicle.id)
            .order_by(TransportVehicleSchedule.id)
        ).scalars().all()
        weekend_schedules = db.execute(
            select(TransportVehicleSchedule)
            .where(TransportVehicleSchedule.vehicle_id == weekend_vehicle.id)
            .order_by(TransportVehicleSchedule.route_kind)
        ).scalars().all()
        regular_schedules = db.execute(
            select(TransportVehicleSchedule)
            .where(TransportVehicleSchedule.vehicle_id == regular_vehicle.id)
            .order_by(TransportVehicleSchedule.route_kind)
        ).scalars().all()
        friday_vehicle_rows, _, _ = transport_service_module._build_vehicle_rows_for_dashboard(
            db,
            service_date=friday,
            route_kind="work_to_home",
            work_to_home_departure_time=friday_work_to_home_departure_time,
        )
        saturday_vehicle_rows, _, _ = transport_service_module._build_vehicle_rows_for_dashboard(
            db,
            service_date=saturday,
            route_kind="work_to_home",
            work_to_home_departure_time=saturday_work_to_home_departure_time,
        )
        friday_registry_rows = transport_service_module._build_transport_vehicle_registry_rows(
            active_schedule_rows=active_schedule_rows,
            request_kind_by_id={},
            recurring_assignment_templates={},
            explicit_assignments=[],
            service_date=friday,
            route_kind="work_to_home",
            work_to_home_departure_time=friday_work_to_home_departure_time,
        )

    assert len(extra_schedules) == 1
    assert extra_schedules[0].route_kind == "work_to_home"
    assert extra_schedules[0].recurrence_kind == "single_date"
    assert extra_schedules[0].service_date == friday
    friday_extra_rows = [row for row in friday_vehicle_rows["extra"] if row.placa == "EXT7001"]
    assert len(friday_extra_rows) == 1
    assert friday_extra_rows[0].route_kind == "work_to_home"
    assert all(row.placa != "EXT7001" for row in saturday_vehicle_rows["extra"])
    friday_registry_extra_rows = [row for row in friday_registry_rows["extra"] if row.placa == "EXT7001"]
    assert len(friday_registry_extra_rows) == 1
    assert friday_registry_extra_rows[0].service_date == friday
    assert friday_registry_extra_rows[0].route_kind == "work_to_home"

    assert len(weekend_schedules) == 4
    assert {row.route_kind for row in weekend_schedules} == {"home_to_work", "work_to_home"}
    assert all(row.recurrence_kind == "matching_weekday" for row in weekend_schedules)
    assert {row.weekday for row in weekend_schedules} == {5, 6}
    assert all(row.service_date == sunday for row in weekend_schedules)

    assert len(regular_schedules) == 2
    assert {row.route_kind for row in regular_schedules} == {"home_to_work", "work_to_home"}
    assert all(row.recurrence_kind == "weekday" for row in regular_schedules)
    assert all(row.service_date == friday for row in regular_schedules)


def test_transport_vehicle_registration_accepts_long_plate_with_special_characters():
    friday = date(2026, 4, 17)
    long_plate = "SG-1234.ABC-789"

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "regular",
                "service_date": friday.isoformat(),
                "tipo": "van",
                "placa": long_plate,
                "color": "Gray",
                "lugares": 11,
                "tolerance": 9,
            },
        )

    assert response.status_code == 200, response.text

    with SessionLocal() as db:
        persisted_vehicle = db.execute(select(Vehicle).where(Vehicle.placa == long_plate)).scalar_one()
        assert persisted_vehicle.placa == long_plate


def test_transport_extra_vehicle_registration_allows_partial_base_data():
    friday = date(2026, 4, 17)

    with SessionLocal() as db:
        previous_vehicle_id = db.execute(select(Vehicle.id).order_by(Vehicle.id.desc())).scalars().first() or 0

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "extra",
                "service_date": friday.isoformat(),
                "route_kind": "work_to_home",
                "departure_time": "17:45",
            },
        )
        dashboard_response = client.get(
            "/api/transport/dashboard",
            params={"service_date": friday.isoformat(), "route_kind": "work_to_home"},
        )

    assert response.status_code == 200, response.text
    assert dashboard_response.status_code == 200, dashboard_response.text

    with SessionLocal() as db:
        persisted_vehicle = db.execute(
            select(Vehicle).where(Vehicle.id > previous_vehicle_id).order_by(Vehicle.id.desc())
        ).scalar_one()
        schedules = db.execute(
            select(TransportVehicleSchedule)
            .where(TransportVehicleSchedule.vehicle_id == persisted_vehicle.id)
            .order_by(TransportVehicleSchedule.id)
        ).scalars().all()

    assert persisted_vehicle.placa is None
    assert persisted_vehicle.tipo is None
    assert persisted_vehicle.color is None
    assert persisted_vehicle.lugares is None
    assert persisted_vehicle.tolerance is None
    assert len(schedules) == 1
    assert schedules[0].service_scope == "extra"
    assert schedules[0].route_kind == "work_to_home"
    assert schedules[0].service_date == friday
    assert schedules[0].departure_time == "17:45"

    dashboard_payload = dashboard_response.json()
    extra_vehicle_row = next(row for row in dashboard_payload["extra_vehicles"] if row["id"] == persisted_vehicle.id)
    extra_registry_row = next(
        row for row in dashboard_payload["extra_vehicle_registry"] if row["vehicle_id"] == persisted_vehicle.id
    )

    assert extra_vehicle_row["placa"] is None
    assert extra_vehicle_row["tipo"] is None
    assert extra_vehicle_row["lugares"] is None
    assert extra_vehicle_row["tolerance"] is None
    assert extra_vehicle_row["pending_fields"] == ["tipo", "placa", "color", "lugares", "tolerance"]
    assert extra_vehicle_row["is_ready_for_allocation"] is False
    assert extra_vehicle_row["route_kind"] == "work_to_home"
    assert extra_registry_row["placa"] is None
    assert extra_registry_row["pending_fields"] == ["tipo", "placa", "color", "lugares", "tolerance"]
    assert extra_registry_row["is_ready_for_allocation"] is False


def test_transport_regular_vehicle_registration_supports_selected_weekdays():
    saturday = date(2026, 4, 18)

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "regular",
                "service_date": saturday.isoformat(),
                "every_monday": True,
                "every_thursday": True,
                "tipo": "minivan",
                "placa": "REG7021",
                "color": "Silver",
                "lugares": 7,
                "tolerance": 9,
            },
        )

    assert response.status_code == 200

    with SessionLocal() as db:
        regular_vehicle = db.execute(select(Vehicle).where(Vehicle.placa == "REG7021")).scalar_one()
        regular_schedules = db.execute(
            select(TransportVehicleSchedule)
            .where(TransportVehicleSchedule.vehicle_id == regular_vehicle.id)
            .order_by(TransportVehicleSchedule.route_kind, TransportVehicleSchedule.weekday)
        ).scalars().all()

    assert len(regular_schedules) == 4
    assert {row.route_kind for row in regular_schedules} == {"home_to_work", "work_to_home"}
    assert all(row.recurrence_kind == "matching_weekday" for row in regular_schedules)
    assert {row.weekday for row in regular_schedules} == {0, 3}
    assert all(row.service_date == saturday for row in regular_schedules)


def test_transport_extra_vehicle_registration_requires_departure_time():
    friday = date(2026, 4, 17)

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "extra",
                "service_date": friday.isoformat(),
                "route_kind": "work_to_home",
                "tipo": "carro",
                "placa": "EXT7011",
                "color": "Gray",
                "lugares": 4,
                "tolerance": 6,
            },
        )

    assert response.status_code == 422
    assert "departure_time is required for extra vehicles" in json.dumps(response.json()["detail"])


def test_transport_weekend_vehicle_registration_requires_persistent_weekday_selection():
    saturday = date(2026, 4, 18)

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "weekend",
                "service_date": saturday.isoformat(),
                "tipo": "van",
                "placa": "WKD7011",
                "color": "Gray",
                "lugares": 8,
                "tolerance": 10,
            },
        )

    assert response.status_code == 422
    assert "Weekend vehicles must be persistent" in json.dumps(response.json()["detail"])


def test_transport_weekend_vehicle_registration_can_start_from_weekday_dashboard_date():
    wednesday = date(2026, 4, 22)
    friday = date(2026, 4, 24)
    saturday = date(2026, 4, 25)

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "weekend",
                "service_date": wednesday.isoformat(),
                "every_saturday": True,
                "tipo": "van",
                "placa": "WKD7022",
                "color": "Blue",
                "lugares": 10,
                "tolerance": 10,
            },
        )
        wednesday_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": wednesday.isoformat(), "route_kind": "home_to_work"},
        )
        friday_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": friday.isoformat(), "route_kind": "home_to_work"},
        )
        saturday_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": saturday.isoformat(), "route_kind": "home_to_work"},
        )

    assert response.status_code == 200
    assert wednesday_dashboard.status_code == 200
    assert friday_dashboard.status_code == 200
    assert saturday_dashboard.status_code == 200
    assert all(row["placa"] != "WKD7022" for row in wednesday_dashboard.json()["weekend_vehicles"])
    assert all(row["placa"] != "WKD7022" for row in friday_dashboard.json()["weekend_vehicles"])
    assert any(row["placa"] == "WKD7022" for row in saturday_dashboard.json()["weekend_vehicles"])


def test_transport_regular_vehicle_registration_can_start_from_weekend_dashboard_date():
    saturday = date(2026, 4, 18)
    monday = date(2026, 4, 20)
    tuesday = date(2026, 4, 21)
    wednesday = date(2026, 4, 22)
    thursday = date(2026, 4, 23)

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "regular",
                "service_date": saturday.isoformat(),
                "every_monday": True,
                "every_tuesday": True,
                "every_thursday": True,
                "tipo": "carro",
                "placa": "REG7022",
                "color": "White",
                "lugares": 4,
                "tolerance": 7,
            },
        )
        saturday_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": saturday.isoformat(), "route_kind": "home_to_work"},
        )
        monday_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": monday.isoformat(), "route_kind": "home_to_work"},
        )
        tuesday_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": tuesday.isoformat(), "route_kind": "home_to_work"},
        )
        wednesday_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": wednesday.isoformat(), "route_kind": "home_to_work"},
        )
        thursday_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": thursday.isoformat(), "route_kind": "home_to_work"},
        )

    assert response.status_code == 200
    assert saturday_dashboard.status_code == 200
    assert monday_dashboard.status_code == 200
    assert tuesday_dashboard.status_code == 200
    assert wednesday_dashboard.status_code == 200
    assert thursday_dashboard.status_code == 200
    assert all(row["placa"] != "REG7022" for row in saturday_dashboard.json()["regular_vehicles"])
    assert any(row["placa"] == "REG7022" for row in monday_dashboard.json()["regular_vehicles"])
    assert any(row["placa"] == "REG7022" for row in tuesday_dashboard.json()["regular_vehicles"])
    assert all(row["placa"] != "REG7022" for row in wednesday_dashboard.json()["regular_vehicles"])
    assert any(row["placa"] == "REG7022" for row in thursday_dashboard.json()["regular_vehicles"])


def test_transport_vehicle_registration_conflict_messages_are_in_english():
    friday = date(2026, 4, 17)

    with TestClient(app) as client:
        ensure_admin_session(client)

        first_response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "extra",
                "service_date": friday.isoformat(),
                "route_kind": "work_to_home",
                "departure_time": "17:45",
                "tipo": "carro",
                "placa": "EXT7010",
                "color": "Red",
                "lugares": 4,
                "tolerance": 6,
            },
        )
        duplicate_response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "regular",
                "service_date": friday.isoformat(),
                "tipo": "carro",
                "placa": "EXT7010",
                "color": "Red",
                "lugares": 4,
                "tolerance": 6,
            },
        )

    assert first_response.status_code == 200
    assert duplicate_response.status_code == 409
    duplicate_detail = extract_transport_structured_detail(duplicate_response)
    assert duplicate_detail["message_key"] == "status.couldNotSaveVehicle"
    assert duplicate_detail["error_code"] == "transport_vehicle_create_failed"
    assert duplicate_detail["technical_detail"] == (
        "A vehicle with this plate already exists in another list: "
        "Extra list (Work to Home on 2026-04-17)."
    )


def test_transport_vehicle_registration_reuses_plate_after_past_single_date_schedule():
    friday = date(2026, 4, 17)
    monday = date(2026, 4, 20)

    with TestClient(app) as client:
        ensure_admin_session(client)

        first_response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "extra",
                "service_date": friday.isoformat(),
                "route_kind": "work_to_home",
                "departure_time": "17:45",
                "tipo": "carro",
                "placa": "AAA0000A",
                "color": "Gray",
                "lugares": 4,
                "tolerance": 6,
            },
        )
        reused_response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "regular",
                "service_date": monday.isoformat(),
                "tipo": "minivan",
                "placa": "AAA0000A",
                "color": "Gray",
                "lugares": 7,
                "tolerance": 10,
            },
        )

    assert first_response.status_code == 200
    assert reused_response.status_code == 200

    with SessionLocal() as db:
        vehicle = db.execute(select(Vehicle).where(Vehicle.placa == "AAA0000A")).scalar_one()
        schedules = db.execute(
            select(TransportVehicleSchedule)
            .where(TransportVehicleSchedule.vehicle_id == vehicle.id)
            .order_by(TransportVehicleSchedule.id)
        ).scalars().all()

    assert vehicle.service_scope == "regular"
    assert vehicle.tipo == "minivan"
    assert len(schedules) == 3
    assert [row.is_active for row in schedules] == [False, True, True]
    assert schedules[0].recurrence_kind == "single_date"
    assert schedules[0].service_date == friday
    assert {row.route_kind for row in schedules[1:] if row.is_active} == {"home_to_work", "work_to_home"}


def test_transport_vehicle_registration_allows_partial_regular_vehicle_and_exposes_readiness_metadata():
    monday = date(2026, 4, 20)

    with SessionLocal() as db:
        previous_vehicle_id = db.execute(select(Vehicle.id).order_by(Vehicle.id.desc())).scalars().first() or 0

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "regular",
                "service_date": monday.isoformat(),
                "every_monday": True,
            },
        )
        dashboard_response = client.get(
            "/api/transport/dashboard",
            params={"service_date": monday.isoformat(), "route_kind": "home_to_work"},
        )

    assert response.status_code == 200, response.text
    assert dashboard_response.status_code == 200, dashboard_response.text

    with SessionLocal() as db:
        persisted_vehicle = db.execute(
            select(Vehicle).where(Vehicle.id > previous_vehicle_id).order_by(Vehicle.id.desc())
        ).scalar_one()

    assert persisted_vehicle.placa is None
    assert persisted_vehicle.tipo is None
    assert persisted_vehicle.color is None
    assert persisted_vehicle.lugares is None
    assert persisted_vehicle.tolerance is None

    dashboard_payload = dashboard_response.json()
    vehicle_row = next(row for row in dashboard_payload["regular_vehicles"] if row["id"] == persisted_vehicle.id)
    registry_row = next(
        row for row in dashboard_payload["regular_vehicle_registry"] if row["vehicle_id"] == persisted_vehicle.id
    )

    assert vehicle_row["placa"] is None
    assert vehicle_row["tipo"] is None
    assert vehicle_row["lugares"] is None
    assert vehicle_row["pending_fields"] == ["tipo", "placa", "color", "lugares", "tolerance"]
    assert vehicle_row["is_ready_for_allocation"] is False
    assert registry_row["placa"] is None
    assert registry_row["tipo"] is None
    assert registry_row["lugares"] is None
    assert registry_row["pending_fields"] == ["tipo", "placa", "color", "lugares", "tolerance"]
    assert registry_row["is_ready_for_allocation"] is False


def test_transport_vehicle_registration_allows_partial_weekend_vehicle_and_exposes_readiness_metadata():
    saturday = date(2026, 4, 18)

    with SessionLocal() as db:
        previous_vehicle_id = db.execute(select(Vehicle.id).order_by(Vehicle.id.desc())).scalars().first() or 0

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "weekend",
                "service_date": saturday.isoformat(),
                "every_sunday": True,
            },
        )
        dashboard_response = client.get(
            "/api/transport/dashboard",
            params={"service_date": saturday.isoformat(), "route_kind": "home_to_work"},
        )

    assert response.status_code == 200, response.text
    assert dashboard_response.status_code == 200, dashboard_response.text

    with SessionLocal() as db:
        persisted_vehicle = db.execute(
            select(Vehicle).where(Vehicle.id > previous_vehicle_id).order_by(Vehicle.id.desc())
        ).scalar_one()

    assert persisted_vehicle.placa is None
    assert persisted_vehicle.tipo is None
    assert persisted_vehicle.color is None
    assert persisted_vehicle.lugares is None
    assert persisted_vehicle.tolerance is None

    dashboard_payload = dashboard_response.json()
    registry_row = next(
        row for row in dashboard_payload["weekend_vehicle_registry"] if row["vehicle_id"] == persisted_vehicle.id
    )

    assert registry_row["placa"] is None
    assert registry_row["tipo"] is None
    assert registry_row["lugares"] is None
    assert registry_row["pending_fields"] == ["tipo", "placa", "color", "lugares", "tolerance"]
    assert registry_row["is_ready_for_allocation"] is False


def test_transport_regular_vehicle_registration_requires_explicit_weekday_selection_when_fields_are_present():
    monday = date(2026, 4, 20)

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "regular",
                "service_date": monday.isoformat(),
                "every_monday": False,
                "every_tuesday": False,
                "every_wednesday": False,
                "every_thursday": False,
                "every_friday": False,
            },
        )

    assert response.status_code == 422
    assert "Regular vehicles must be persistent" in json.dumps(response.json()["detail"])


def test_transport_assignment_rejects_vehicle_that_is_not_ready_for_allocation():
    friday = date(2026, 4, 17)
    timestamp = now_sgt()

    with SessionLocal() as db:
        workplace = Workplace(
            workplace="Pending Allocation Hub",
            address="55 Pending Road",
            zip="550055",
            country="Singapore",
        )
        db.add(workplace)
        db.flush()

        vehicle = Vehicle(placa=None, tipo=None, color=None, lugares=None, tolerance=None, service_scope="regular")
        db.add(vehicle)
        db.flush()

        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )

        user = User(
            rfid=None,
            nome="Pending Allocation Rider",
            chave="PA55",
            projeto="P55",
            workplace=workplace.workplace,
            placa="PA5500",
            end_rua="55 Pending Road",
            zip="550055",
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        request_row = TransportRequest(
            user_id=user.id,
            request_kind="regular",
            recurrence_kind="weekday",
            requested_time="08:10",
            single_date=None,
            created_via="bot",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
            cancelled_at=None,
        )
        db.add(request_row)
        db.commit()

        request_id = request_row.id
        vehicle_id = vehicle.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            "/api/transport/assignments",
            json={
                "request_id": request_id,
                "service_date": friday.isoformat(),
                "route_kind": "home_to_work",
                "status": "confirmed",
                "vehicle_id": vehicle_id,
            },
        )

    assert response.status_code == 409
    detail = extract_transport_structured_detail(response)
    assert detail["message_key"] == "warnings.vehiclePendingAllocation"
    assert detail["error_code"] == "transport_vehicle_not_ready_for_allocation"
    assert detail["technical_detail"] == "The selected vehicle is not ready for allocation."

    with SessionLocal() as db:
        assignment = db.execute(
            select(TransportAssignment).where(
                TransportAssignment.request_id == request_id,
                TransportAssignment.service_date == friday,
                TransportAssignment.route_kind == "home_to_work",
            )
        ).scalar_one_or_none()

    assert assignment is None


def test_transport_assignment_allows_vehicle_missing_plate_and_color_when_operational_fields_exist():
    friday = date(2026, 4, 17)
    timestamp = now_sgt()

    with SessionLocal() as db:
        workplace = Workplace(
            workplace="Partial Allocation Hub",
            address="56 Pending Road",
            zip="560056",
            country="Singapore",
        )
        db.add(workplace)
        db.flush()

        vehicle = Vehicle(
            placa=None,
            tipo="van",
            color=None,
            lugares=10,
            tolerance=8,
            service_scope="regular",
        )
        db.add(vehicle)
        db.flush()

        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )

        user = User(
            rfid=None,
            nome="Partial Allocation Rider",
            chave="PA56",
            projeto="P56",
            workplace=workplace.workplace,
            placa="PA5600",
            end_rua="56 Pending Road",
            zip="560056",
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        request_row = TransportRequest(
            user_id=user.id,
            request_kind="regular",
            recurrence_kind="weekday",
            requested_time="08:10",
            single_date=None,
            created_via="bot",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
            cancelled_at=None,
        )
        db.add(request_row)
        db.commit()

        request_id = request_row.id
        vehicle_id = vehicle.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        dashboard_response = client.get(
            "/api/transport/dashboard",
            params={"service_date": friday.isoformat(), "route_kind": "home_to_work"},
        )
        response = client.post(
            "/api/transport/assignments",
            json={
                "request_id": request_id,
                "service_date": friday.isoformat(),
                "route_kind": "home_to_work",
                "status": "confirmed",
                "vehicle_id": vehicle_id,
            },
        )

    assert dashboard_response.status_code == 200
    dashboard_payload = dashboard_response.json()
    vehicle_row = next(row for row in dashboard_payload["regular_vehicles"] if row["id"] == vehicle_id)
    registry_row = next(
        row for row in dashboard_payload["regular_vehicle_registry"] if row["vehicle_id"] == vehicle_id
    )

    assert vehicle_row["placa"] is None
    assert vehicle_row["color"] is None
    assert vehicle_row["pending_fields"] == ["placa", "color"]
    assert vehicle_row["is_ready_for_allocation"] is True
    assert registry_row["pending_fields"] == ["placa", "color"]
    assert registry_row["is_ready_for_allocation"] is True

    assert response.status_code == 200

    with SessionLocal() as db:
        assignment = db.execute(
            select(TransportAssignment).where(
                TransportAssignment.request_id == request_id,
                TransportAssignment.service_date == friday,
                TransportAssignment.route_kind == "home_to_work",
            )
        ).scalar_one()

    assert assignment.status == "confirmed"
    assert assignment.vehicle_id == vehicle_id


def test_transport_vehicle_registration_without_plate_always_creates_new_vehicle():
    monday = date(2026, 4, 20)
    saturday = date(2026, 4, 25)

    with SessionLocal() as db:
        previous_vehicle_id = db.execute(select(Vehicle.id).order_by(Vehicle.id.desc())).scalars().first() or 0

    with TestClient(app) as client:
        ensure_admin_session(client)
        first_response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "regular",
                "service_date": monday.isoformat(),
                "every_monday": True,
            },
        )
        second_response = client.post(
            "/api/transport/vehicles",
            json={
                "service_scope": "weekend",
                "service_date": saturday.isoformat(),
                "every_saturday": True,
            },
        )

    assert first_response.status_code == 200, first_response.text
    assert second_response.status_code == 200, second_response.text

    with SessionLocal() as db:
        new_vehicle_ids = db.execute(
            select(Vehicle.id).where(Vehicle.id > previous_vehicle_id).order_by(Vehicle.id)
        ).scalars().all()

    assert len(new_vehicle_ids) == 2


def test_transport_vehicle_update_blocks_ready_vehicle_from_becoming_incomplete_with_future_confirmed_assignments():
    future_monday = now_sgt().date() + timedelta(days=1)
    while future_monday.weekday() != 0:
        future_monday += timedelta(days=1)
    timestamp = now_sgt()

    with SessionLocal() as db:
        workplace = Workplace(
            workplace="Vehicle Readiness Hub",
            address="41 Readiness Road",
            zip="841001",
            country="Singapore",
        )
        db.add(workplace)
        db.flush()

        vehicle = Vehicle(placa="RDY4101", tipo="van", color="White", lugares=2, tolerance=8, service_scope="regular")
        db.add(vehicle)
        db.flush()

        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )

        user = User(
            rfid=None,
            nome="Vehicle Readiness Rider",
            chave="RD41",
            projeto="P41",
            workplace=workplace.workplace,
            placa=vehicle.placa,
            end_rua="41 Readiness Avenue",
            zip="841002",
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        transport_request = TransportRequest(
            user_id=user.id,
            request_kind="regular",
            recurrence_kind="weekday",
            requested_time="07:30",
            selected_weekdays_json=json.dumps([0, 1, 2, 3, 4]),
            single_date=None,
            created_via="admin",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
        )
        db.add(transport_request)
        db.flush()

        db.add(
            TransportAssignment(
                request_id=transport_request.id,
                service_date=future_monday,
                route_kind="home_to_work",
                vehicle_id=vehicle.id,
                status="confirmed",
                response_message="Assigned",
                acknowledged_by_user=False,
                assigned_by_admin_id=None,
                created_at=timestamp,
                updated_at=timestamp,
            )
        )
        db.commit()

        vehicle_id = vehicle.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.put(
            f"/api/transport/vehicles/{vehicle_id}",
            json={
                "placa": "RDY4101",
                "tipo": "van",
                "color": "White",
                "lugares": None,
                "tolerance": 8,
            },
        )

    assert response.status_code == 409
    detail = extract_transport_structured_detail(response)
    assert detail["message_key"] == "status.couldNotUpdateVehicle"
    assert detail["error_code"] == "transport_vehicle_update_failed"
    assert "Cannot make the vehicle incomplete because future confirmed assignments exist" in detail["technical_detail"]

    with SessionLocal() as db:
        persisted_vehicle = db.get(Vehicle, vehicle_id)

    assert persisted_vehicle is not None
    assert persisted_vehicle.placa == "RDY4101"


def test_transport_vehicle_update_updates_base_fields_without_recreating_schedules_or_assignments():
    friday = date(2026, 4, 17)
    timestamp = now_sgt()

    with SessionLocal() as db:
        workplace = Workplace(
            workplace="Vehicle Update Hub",
            address="1 Update Road",
            zip="750001",
            country="Singapore",
        )
        db.add(workplace)
        db.flush()

        vehicle = Vehicle(placa="UPD2301", tipo="van", color="Blue", lugares=11, tolerance=8, service_scope="regular")
        db.add(vehicle)
        db.flush()

        first_schedule = add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        second_schedule = add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )

        user = User(
            rfid=None,
            nome="Vehicle Update Rider",
            chave="VU23",
            projeto="P23",
            workplace=workplace.workplace,
            placa=vehicle.placa,
            end_rua="23 Update Avenue",
            zip="760001",
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        transport_request = TransportRequest(
            user_id=user.id,
            request_kind="regular",
            recurrence_kind="weekday",
            requested_time="07:30",
            selected_weekdays_json=json.dumps([0, 1, 2, 3, 4]),
            single_date=None,
            created_via="admin",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
        )
        db.add(transport_request)
        db.flush()

        assignment = TransportAssignment(
            request_id=transport_request.id,
            service_date=friday,
            route_kind="home_to_work",
            vehicle_id=vehicle.id,
            status="confirmed",
            response_message="Assigned",
            acknowledged_by_user=False,
            assigned_by_admin_id=None,
            created_at=timestamp,
            updated_at=timestamp,
        )
        db.add(assignment)
        db.commit()

        vehicle_id = vehicle.id
        schedule_ids = {first_schedule.id, second_schedule.id}
        assignment_id = assignment.id
        user_id = user.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.put(
            f"/api/transport/vehicles/{vehicle_id}",
            json={
                "placa": "UPD2301B",
                "tipo": "onibus",
                "color": "White",
                "lugares": 40,
                "tolerance": 15,
            },
        )

    assert response.status_code == 200, response.text

    with SessionLocal() as db:
        persisted_vehicle = db.get(Vehicle, vehicle_id)
        persisted_user = db.get(User, user_id)
        persisted_assignment = db.get(TransportAssignment, assignment_id)
        schedules = db.execute(
            select(TransportVehicleSchedule)
            .where(TransportVehicleSchedule.vehicle_id == vehicle_id)
            .order_by(TransportVehicleSchedule.id)
        ).scalars().all()

    assert persisted_vehicle is not None
    assert persisted_vehicle.placa == "UPD2301B"
    assert persisted_vehicle.tipo == "onibus"
    assert persisted_vehicle.color == "White"
    assert persisted_vehicle.lugares == 40
    assert persisted_vehicle.tolerance == 15
    assert persisted_vehicle.service_scope == "regular"
    assert persisted_user is not None
    assert persisted_user.vehicle_id == vehicle_id
    assert persisted_user.placa == "UPD2301B"
    assert persisted_assignment is not None
    assert persisted_assignment.vehicle_id == vehicle_id
    assert len(schedules) == 2
    assert {schedule.id for schedule in schedules} == schedule_ids


def test_transport_vehicle_update_can_complete_pending_vehicle_and_expose_readiness_metadata():
    monday = date(2026, 4, 20)

    with SessionLocal() as db:
        pending_vehicle = Vehicle(
            placa=None,
            tipo=None,
            color=None,
            lugares=None,
            tolerance=None,
            service_scope="regular",
        )
        db.add(pending_vehicle)
        db.flush()

        first_schedule = add_transport_schedule(
            db,
            vehicle=pending_vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        second_schedule = add_transport_schedule(
            db,
            vehicle=pending_vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )
        db.commit()

        vehicle_id = pending_vehicle.id
        schedule_ids = {first_schedule.id, second_schedule.id}

    with TestClient(app) as client:
        ensure_admin_session(client)
        update_response = client.put(
            f"/api/transport/vehicles/{vehicle_id}",
            json={
                "placa": "CMP8401",
                "tipo": "minivan",
                "color": "Silver",
                "lugares": 7,
                "tolerance": 9,
            },
        )
        dashboard_response = client.get(
            "/api/transport/dashboard",
            params={"service_date": monday.isoformat(), "route_kind": "home_to_work"},
        )

    assert update_response.status_code == 200, update_response.text
    assert dashboard_response.status_code == 200, dashboard_response.text

    with SessionLocal() as db:
        persisted_vehicle = db.get(Vehicle, vehicle_id)
        schedules = db.execute(
            select(TransportVehicleSchedule)
            .where(TransportVehicleSchedule.vehicle_id == vehicle_id)
            .order_by(TransportVehicleSchedule.id)
        ).scalars().all()

    assert persisted_vehicle is not None
    assert persisted_vehicle.placa == "CMP8401"
    assert persisted_vehicle.tipo == "minivan"
    assert persisted_vehicle.color == "Silver"
    assert persisted_vehicle.lugares == 7
    assert persisted_vehicle.tolerance == 9
    assert len(schedules) == 2
    assert {schedule.id for schedule in schedules} == schedule_ids

    dashboard_payload = dashboard_response.json()
    vehicle_row = next(row for row in dashboard_payload["regular_vehicles"] if row["id"] == vehicle_id)
    registry_row = next(
        row for row in dashboard_payload["regular_vehicle_registry"] if row["vehicle_id"] == vehicle_id
    )

    assert vehicle_row["placa"] == "CMP8401"
    assert vehicle_row["tipo"] == "minivan"
    assert vehicle_row["color"] == "Silver"
    assert vehicle_row["lugares"] == 7
    assert vehicle_row["tolerance"] == 9
    assert vehicle_row["pending_fields"] == []
    assert vehicle_row["is_ready_for_allocation"] is True
    assert registry_row["placa"] == "CMP8401"
    assert registry_row["pending_fields"] == []
    assert registry_row["is_ready_for_allocation"] is True


def test_transport_vehicle_update_rejects_duplicate_plate_from_another_vehicle():
    with SessionLocal() as db:
        first_vehicle = Vehicle(placa="UPD2302", tipo="van", color="Blue", lugares=12, tolerance=9, service_scope="regular")
        second_vehicle = Vehicle(placa="UPD2303", tipo="carro", color="Gray", lugares=4, tolerance=6, service_scope="extra")
        db.add_all([first_vehicle, second_vehicle])
        db.commit()
        second_vehicle_id = second_vehicle.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.put(
            f"/api/transport/vehicles/{second_vehicle_id}",
            json={
                "placa": "UPD2302",
                "tipo": "carro",
                "color": "Gray",
                "lugares": 4,
                "tolerance": 6,
            },
        )

    assert response.status_code == 409
    detail = extract_transport_structured_detail(response)
    assert detail["message_key"] == "status.couldNotUpdateVehicle"
    assert detail["error_code"] == "transport_vehicle_update_failed"
    assert detail["technical_detail"] == "A vehicle with this plate already exists."


def test_transport_vehicle_update_blocks_seat_reduction_when_future_confirmed_assignments_exceed_new_capacity():
    future_monday = now_sgt().date() + timedelta(days=1)
    while future_monday.weekday() != 0:
        future_monday += timedelta(days=1)
    timestamp = now_sgt()

    with SessionLocal() as db:
        workplace = Workplace(
            workplace="Vehicle Capacity Hub",
            address="31 Capacity Road",
            zip="831001",
            country="Singapore",
        )
        db.add(workplace)
        db.flush()

        vehicle = Vehicle(placa="CAP3301", tipo="van", color="White", lugares=2, tolerance=8, service_scope="regular")
        db.add(vehicle)
        db.flush()

        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )

        request_ids: list[int] = []
        for suffix in ("A", "B"):
            user = User(
                rfid=None,
                nome=f"Capacity Rider {suffix}",
                chave=f"C3{suffix}1",
                projeto="P33",
                workplace=workplace.workplace,
                placa=vehicle.placa,
                end_rua=f"31 Capacity Avenue {suffix}",
                zip="831002",
                local=None,
                checkin=None,
                time=None,
                last_active_at=timestamp,
                inactivity_days=0,
            )
            db.add(user)
            db.flush()

            transport_request = TransportRequest(
                user_id=user.id,
                request_kind="regular",
                recurrence_kind="weekday",
                requested_time="07:30",
                selected_weekdays_json=json.dumps([0, 1, 2, 3, 4]),
                single_date=None,
                created_via="admin",
                status="active",
                created_at=timestamp,
                updated_at=timestamp,
            )
            db.add(transport_request)
            db.flush()
            request_ids.append(transport_request.id)

        for request_id in request_ids:
            db.add(
                TransportAssignment(
                    request_id=request_id,
                    service_date=future_monday,
                    route_kind="home_to_work",
                    vehicle_id=vehicle.id,
                    status="confirmed",
                    response_message="Assigned",
                    acknowledged_by_user=False,
                    assigned_by_admin_id=None,
                    created_at=timestamp,
                    updated_at=timestamp,
                )
            )
        db.commit()

        vehicle_id = vehicle.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.put(
            f"/api/transport/vehicles/{vehicle_id}",
            json={
                "placa": "CAP3301",
                "tipo": "van",
                "color": "White",
                "lugares": 1,
                "tolerance": 8,
            },
        )

    assert response.status_code == 409
    detail = extract_transport_structured_detail(response)
    assert detail["message_key"] == "status.couldNotUpdateVehicle"
    assert detail["error_code"] == "transport_vehicle_update_failed"
    assert "confirmed assignments would exceed the new capacity" in detail["technical_detail"]

    with SessionLocal() as db:
        persisted_vehicle = db.get(Vehicle, vehicle_id)

    assert persisted_vehicle is not None
    assert persisted_vehicle.lugares == 2


def test_transport_vehicle_schedule_update_changes_extra_availability_without_recreating_vehicle():
    original_date = now_sgt().date() + timedelta(days=10)
    updated_date = original_date + timedelta(days=1)
    timestamp = now_sgt()

    with SessionLocal() as db:
        vehicle = Vehicle(placa="SCH2401", tipo="carro", color="Black", lugares=4, tolerance=6, service_scope="extra")
        db.add(vehicle)
        db.flush()

        schedule = add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="extra",
            route_kind="work_to_home",
            recurrence_kind="single_date",
            service_date=original_date,
            departure_time="18:10",
        )
        db.add(
            TransportVehicleScheduleException(
                vehicle_schedule_id=schedule.id,
                service_date=original_date,
                created_at=timestamp,
            )
        )
        db.commit()

        schedule_id = schedule.id
        vehicle_id = vehicle.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.put(
            f"/api/transport/vehicle-schedules/{schedule_id}",
            json={
                "service_scope": "extra",
                "route_kind": "home_to_work",
                "recurrence_kind": "single_date",
                "service_date": updated_date.isoformat(),
                "departure_time": "19:00",
                "is_active": True,
            },
        )
        old_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": original_date.isoformat(), "route_kind": "work_to_home"},
        )
        new_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": updated_date.isoformat(), "route_kind": "home_to_work"},
        )

    assert response.status_code == 200, response.text
    assert old_dashboard.status_code == 200
    assert new_dashboard.status_code == 200
    assert all(row["placa"] != "SCH2401" for row in old_dashboard.json()["extra_vehicles"])
    assert any(row["placa"] == "SCH2401" for row in new_dashboard.json()["extra_vehicles"])

    with SessionLocal() as db:
        persisted_schedule = db.get(TransportVehicleSchedule, schedule_id)
        persisted_vehicle = db.get(Vehicle, vehicle_id)
        persisted_exceptions = db.execute(
            select(TransportVehicleScheduleException).where(
                TransportVehicleScheduleException.vehicle_schedule_id == schedule_id
            )
        ).scalars().all()

    assert persisted_schedule is not None
    assert persisted_schedule.vehicle_id == vehicle_id
    assert persisted_schedule.service_scope == "extra"
    assert persisted_schedule.route_kind == "home_to_work"
    assert persisted_schedule.recurrence_kind == "single_date"
    assert persisted_schedule.service_date == updated_date
    assert persisted_schedule.departure_time == "19:00"
    assert persisted_schedule.is_active is True
    assert persisted_vehicle is not None
    assert persisted_vehicle.placa == "SCH2401"
    assert persisted_exceptions == []


def test_transport_vehicle_schedule_update_rejects_conflict_with_existing_active_schedule():
    with SessionLocal() as db:
        vehicle = Vehicle(placa="SCH2402", tipo="van", color="Blue", lugares=10, tolerance=8, service_scope="regular")
        db.add(vehicle)
        db.flush()

        home_schedule = add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        work_schedule = add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )
        db.commit()

        work_schedule_id = work_schedule.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.put(
            f"/api/transport/vehicle-schedules/{work_schedule_id}",
            json={
                "service_scope": "regular",
                "route_kind": "home_to_work",
                "recurrence_kind": "weekday",
                "is_active": True,
            },
        )

    assert response.status_code == 409
    detail = extract_transport_structured_detail(response)
    assert detail["message_key"] == "status.couldNotUpdateVehicle"
    assert detail["error_code"] == "transport_vehicle_schedule_update_failed"
    assert detail["technical_detail"] == "Another active schedule already exists for the selected list and recurrence pattern."


def test_transport_vehicle_schedule_update_blocks_when_confirmed_assignments_would_become_unavailable():
    future_monday = now_sgt().date() + timedelta(days=1)
    while future_monday.weekday() != 0:
        future_monday += timedelta(days=1)
    timestamp = now_sgt()

    with SessionLocal() as db:
        workplace = Workplace(
            workplace="Schedule Update Hub",
            address="24 Schedule Road",
            zip="824001",
            country="Singapore",
        )
        db.add(workplace)
        db.flush()

        vehicle = Vehicle(placa="SCH2403", tipo="van", color="White", lugares=11, tolerance=9, service_scope="regular")
        db.add(vehicle)
        db.flush()

        home_schedule = add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )

        user = User(
            rfid=None,
            nome="Schedule Update Rider",
            chave="SU24",
            projeto="P24",
            workplace=workplace.workplace,
            placa=vehicle.placa,
            end_rua="24 Update Avenue",
            zip="834001",
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        transport_request = TransportRequest(
            user_id=user.id,
            request_kind="regular",
            recurrence_kind="weekday",
            requested_time="07:30",
            selected_weekdays_json=json.dumps([0, 1, 2, 3, 4]),
            single_date=None,
            created_via="admin",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
        )
        db.add(transport_request)
        db.flush()

        db.add(
            TransportAssignment(
                request_id=transport_request.id,
                service_date=future_monday,
                route_kind="home_to_work",
                vehicle_id=vehicle.id,
                status="confirmed",
                response_message="Assigned",
                acknowledged_by_user=False,
                assigned_by_admin_id=None,
                created_at=timestamp,
                updated_at=timestamp,
            )
        )
        db.commit()

        home_schedule_id = home_schedule.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.put(
            f"/api/transport/vehicle-schedules/{home_schedule_id}",
            json={
                "service_scope": "regular",
                "route_kind": "home_to_work",
                "recurrence_kind": "weekday",
                "is_active": False,
            },
        )

    assert response.status_code == 409
    detail = extract_transport_structured_detail(response)
    assert detail["message_key"] == "status.couldNotUpdateVehicle"
    assert detail["error_code"] == "transport_vehicle_schedule_update_failed"
    assert "confirmed assignments would become unavailable" in detail["technical_detail"]

    with SessionLocal() as db:
        persisted_schedule = db.get(TransportVehicleSchedule, home_schedule_id)

    assert persisted_schedule is not None
    assert persisted_schedule.is_active is True


def test_transport_vehicle_delete_purges_vehicle_and_returns_requests_to_pending():
    friday = date(2026, 4, 17)
    timestamp = now_sgt()

    with SessionLocal() as db:
        db.add(Workplace(workplace="Delete Hub", address="7 Delete Road", zip="707070", country="Singapore"))
        vehicle = Vehicle(placa="DEL1700", tipo="carro", color="Red", lugares=4, tolerance=8, service_scope="regular")
        db.add(vehicle)
        db.flush()
        first_schedule = add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        second_schedule = add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )
        db.flush()
        db.add(
            TransportVehicleScheduleException(
                vehicle_schedule_id=second_schedule.id,
                service_date=friday,
                created_at=timestamp,
            )
        )

        user = User(
            rfid=None,
            nome="Delete Route Rider",
            chave="TD17",
            projeto="P80",
            workplace="Delete Hub",
            placa="DEL1700",
            end_rua="17 Delete Street",
            zip="170170",
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        request_row = TransportRequest(
            user_id=user.id,
            request_kind="regular",
            recurrence_kind="weekday",
            requested_time="08:15",
            single_date=None,
            created_via="bot",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
            cancelled_at=None,
        )
        db.add(request_row)
        db.commit()
        request_id = request_row.id
        user_id = user.id
        vehicle_id = vehicle.id
        schedule_id = first_schedule.id
        schedule_ids = [first_schedule.id, second_schedule.id]

    with TestClient(app) as client:
        ensure_admin_session(client)
        assigned = client.post(
            "/api/transport/assignments",
            json={
                "request_id": request_id,
                "service_date": friday.isoformat(),
                "route_kind": "home_to_work",
                "status": "confirmed",
                "vehicle_id": vehicle_id,
            },
        )
        assert assigned.status_code == 200

        removed = client.delete(
            f"/api/transport/vehicles/{schedule_id}",
            params={"service_date": friday.isoformat()},
        )
        assert removed.status_code == 200

        refreshed_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": friday.isoformat(), "route_kind": "home_to_work"},
        )
        paired_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": friday.isoformat(), "route_kind": "work_to_home"},
        )

    assert refreshed_dashboard.status_code == 200
    assert paired_dashboard.status_code == 200
    request_row = next(row for row in refreshed_dashboard.json()["regular_requests"] if row["id"] == request_id)
    paired_request_row = next(row for row in paired_dashboard.json()["regular_requests"] if row["id"] == request_id)
    assert request_row["assignment_status"] == "pending"
    assert paired_request_row["assignment_status"] == "pending"
    assert request_row["assigned_vehicle"] is None
    assert paired_request_row["assigned_vehicle"] is None
    assert all(row["placa"] != "DEL1700" for row in refreshed_dashboard.json()["regular_vehicles"])
    assert all(row["placa"] != "DEL1700" for row in paired_dashboard.json()["regular_vehicles"])

    with SessionLocal() as db:
        user_row = db.get(User, user_id)
        vehicle_row = db.get(Vehicle, vehicle_id)
        schedules = db.execute(
            select(TransportVehicleSchedule).where(TransportVehicleSchedule.vehicle_id == vehicle_id)
        ).scalars().all()
        schedule_exceptions = db.execute(
            select(TransportVehicleScheduleException).where(
                TransportVehicleScheduleException.vehicle_schedule_id.in_(schedule_ids)
            )
        ).scalars().all()
        assignments = db.execute(
            select(TransportAssignment).where(TransportAssignment.request_id == request_id)
        ).scalars().all()

    assert user_row is not None
    assert user_row.vehicle_id is None
    assert user_row.placa is None
    assert vehicle_row is None
    assert schedules == []
    assert schedule_exceptions == []
    assert assignments == []


def test_transport_vehicle_delete_clears_user_plate_references_for_all_vehicle_lists():
    scenarios = [
        {
            "scope": "regular",
            "plate": "CLR1701",
            "user_key": "VR17",
            "selected_date": date(2026, 4, 17),
            "dashboard_key": "regular_vehicles",
            "schedule_specs": [
                {"route_kind": "home_to_work", "recurrence_kind": "weekday"},
                {"route_kind": "work_to_home", "recurrence_kind": "weekday"},
            ],
        },
        {
            "scope": "weekend",
            "plate": "CLW1901",
            "user_key": "VW19",
            "selected_date": date(2026, 4, 19),
            "dashboard_key": "weekend_vehicles",
            "schedule_specs": [
                {"route_kind": "home_to_work", "recurrence_kind": "matching_weekday", "weekday": 6},
                {"route_kind": "work_to_home", "recurrence_kind": "matching_weekday", "weekday": 6},
            ],
        },
        {
            "scope": "extra",
            "plate": "CLX2101",
            "user_key": "VX21",
            "selected_date": date(2026, 4, 21),
            "dashboard_key": "extra_vehicles",
            "schedule_specs": [
                {"route_kind": "home_to_work", "recurrence_kind": "single_date", "service_date": date(2026, 4, 21)},
            ],
        },
    ]

    created_rows = []
    with SessionLocal() as db:
        for index, scenario in enumerate(scenarios, start=1):
            workplace = Workplace(
                workplace=f"Delete Scope Hub {index}",
                address=f"{index} Scope Road",
                zip=f"70{index:04d}",
                country="Singapore",
            )
            db.add(workplace)
            db.flush()

            vehicle = Vehicle(
                placa=scenario["plate"],
                tipo="carro",
                color="Blue",
                lugares=4,
                tolerance=6,
                service_scope=scenario["scope"],
            )
            db.add(vehicle)
            db.flush()

            first_schedule_id = None
            for schedule_spec in scenario["schedule_specs"]:
                schedule = add_transport_schedule(
                    db,
                    vehicle=vehicle,
                    service_scope=scenario["scope"],
                    route_kind=schedule_spec["route_kind"],
                    recurrence_kind=schedule_spec["recurrence_kind"],
                    service_date=schedule_spec.get("service_date"),
                    weekday=schedule_spec.get("weekday"),
                )
                if first_schedule_id is None:
                    first_schedule_id = schedule.id

            user = User(
                rfid=None,
                nome=f"Delete Scope Rider {index}",
                chave=scenario["user_key"],
                projeto="P80",
                workplace=workplace.workplace,
                placa=scenario["plate"],
                end_rua=f"{index} Delete Avenue",
                zip=f"80{index:04d}",
                local=None,
                checkin=None,
                time=None,
                last_active_at=now_sgt(),
                inactivity_days=0,
            )
            db.add(user)
            db.flush()

            created_rows.append(
                {
                    "dashboard_key": scenario["dashboard_key"],
                    "plate": scenario["plate"],
                    "schedule_id": first_schedule_id,
                    "selected_date": scenario["selected_date"],
                    "user_id": user.id,
                    "vehicle_id": vehicle.id,
                }
            )

        db.commit()

    with TestClient(app) as client:
        ensure_admin_session(client)

        for created_row in created_rows:
            removed = client.delete(
                f"/api/transport/vehicles/{created_row['schedule_id']}",
                params={"service_date": created_row["selected_date"].isoformat()},
            )
            assert removed.status_code == 200

            dashboard = client.get(
                "/api/transport/dashboard",
                params={
                    "service_date": created_row["selected_date"].isoformat(),
                    "route_kind": "home_to_work",
                },
            )
            assert dashboard.status_code == 200
            assert all(
                row["placa"] != created_row["plate"]
                for row in dashboard.json()[created_row["dashboard_key"]]
            )

    with SessionLocal() as db:
        for created_row in created_rows:
            user_row = db.get(User, created_row["user_id"])
            vehicle_row = db.get(Vehicle, created_row["vehicle_id"])
            schedule_rows = db.execute(
                select(TransportVehicleSchedule).where(
                    TransportVehicleSchedule.vehicle_id == created_row["vehicle_id"]
                )
            ).scalars().all()

            assert user_row is not None
            assert user_row.placa is None
            assert vehicle_row is None
            assert schedule_rows == []


def test_transport_vehicle_delete_removes_legacy_notifications_before_assignments():
    friday = date(2026, 4, 17)
    timestamp = now_sgt()

    with SessionLocal() as db:
        workplace = Workplace(
            workplace="Legacy Notification Hub",
            address="9 Legacy Avenue",
            zip="919191",
            country="Singapore",
        )
        vehicle = Vehicle(placa="LEG1701", tipo="onibus", color="White", lugares=40, tolerance=12, service_scope="regular")
        db.add_all([workplace, vehicle])
        db.flush()

        schedule = add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )

        user = User(
            rfid=None,
            nome="Legacy Notification Rider",
            chave="LN17",
            projeto="P80",
            workplace=workplace.workplace,
            placa=vehicle.placa,
            end_rua="17 Legacy Lane",
            zip="171717",
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        request_row = TransportRequest(
            user_id=user.id,
            request_kind="regular",
            recurrence_kind="weekday",
            requested_time="08:30",
            single_date=None,
            created_via="admin",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
            cancelled_at=None,
        )
        db.add(request_row)
        db.flush()

        assignment = TransportAssignment(
            request_id=request_row.id,
            service_date=friday,
            route_kind="home_to_work",
            vehicle_id=vehicle.id,
            status="confirmed",
            response_message=None,
            assigned_by_admin_id=None,
            created_at=timestamp,
            updated_at=timestamp,
            notified_at=None,
        )
        db.add(assignment)
        db.commit()

        user_id = user.id
        vehicle_id = vehicle.id
        request_id = request_row.id
        assignment_id = assignment.id
        schedule_id = schedule.id

    with engine.begin() as conn:
        conn.execute(text("DROP TABLE IF EXISTS transport_notifications"))
        conn.execute(
            text(
                """
                CREATE TABLE transport_notifications (
                    id INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,
                    user_id INTEGER NOT NULL REFERENCES users(id),
                    chat_id VARCHAR(120),
                    request_id INTEGER REFERENCES transport_requests(id),
                    assignment_id INTEGER REFERENCES transport_assignments(id),
                    message VARCHAR(500) NOT NULL,
                    status VARCHAR(16) NOT NULL DEFAULT 'pending',
                    created_at DATETIME NOT NULL,
                    sent_at DATETIME
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO transport_notifications (
                    user_id,
                    chat_id,
                    request_id,
                    assignment_id,
                    message,
                    status,
                    created_at,
                    sent_at
                ) VALUES (
                    :user_id,
                    NULL,
                    :request_id,
                    :assignment_id,
                    :message,
                    'pending',
                    :created_at,
                    NULL
                )
                """
            ),
            {
                "user_id": user_id,
                "request_id": request_id,
                "assignment_id": assignment_id,
                "message": "Legacy notification row",
                "created_at": timestamp.isoformat(),
            },
        )

    try:
        with TestClient(app) as client:
            ensure_admin_session(client)
            removed = client.delete(
                f"/api/transport/vehicles/{schedule_id}",
                params={"service_date": friday.isoformat()},
            )
            assert removed.status_code == 200, removed.text

        with engine.begin() as conn:
            remaining_notifications = conn.execute(
                text("SELECT COUNT(*) FROM transport_notifications WHERE assignment_id = :assignment_id"),
                {"assignment_id": assignment_id},
            ).scalar_one()

        with SessionLocal() as db:
            user_row = db.get(User, user_id)
            vehicle_row = db.get(Vehicle, vehicle_id)
            assignment_row = db.get(TransportAssignment, assignment_id)

        assert remaining_notifications == 0
        assert user_row is not None
        assert user_row.placa is None
        assert vehicle_row is None
        assert assignment_row is None
    finally:
        with engine.begin() as conn:
            conn.execute(text("DROP TABLE IF EXISTS transport_notifications"))


def test_transport_vehicle_delete_purges_generic_legacy_foreign_key_dependencies():
    friday = date(2026, 4, 17)
    timestamp = now_sgt()

    with SessionLocal() as db:
        workplace = Workplace(
            workplace="Generic Legacy Hub",
            address="8 Generic Avenue",
            zip="818181",
            country="Singapore",
        )
        vehicle = Vehicle(placa="GLG1701", tipo="onibus", color="Blue", lugares=36, tolerance=10, service_scope="regular")
        db.add_all([workplace, vehicle])
        db.flush()

        schedule = add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )

        user = User(
            rfid=None,
            nome="Generic Legacy Rider",
            chave="GL17",
            projeto="P80",
            workplace=workplace.workplace,
            placa=vehicle.placa,
            end_rua="17 Generic Road",
            zip="171818",
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        request_row = TransportRequest(
            user_id=user.id,
            request_kind="regular",
            recurrence_kind="weekday",
            requested_time="08:50",
            single_date=None,
            created_via="admin",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
            cancelled_at=None,
        )
        db.add(request_row)
        db.flush()

        assignment = TransportAssignment(
            request_id=request_row.id,
            service_date=friday,
            route_kind="home_to_work",
            vehicle_id=vehicle.id,
            status="confirmed",
            response_message=None,
            assigned_by_admin_id=None,
            created_at=timestamp,
            updated_at=timestamp,
            notified_at=None,
        )
        db.add(assignment)
        db.commit()

        user_id = user.id
        vehicle_id = vehicle.id
        schedule_id = schedule.id
        assignment_id = assignment.id

    with engine.begin() as conn:
        conn.execute(text("DROP TABLE IF EXISTS legacy_vehicle_links"))
        conn.execute(text("DROP TABLE IF EXISTS legacy_schedule_links"))
        conn.execute(text("DROP TABLE IF EXISTS legacy_assignment_links"))
        conn.execute(
            text(
                """
                CREATE TABLE legacy_vehicle_links (
                    id INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,
                    vehicle_id INTEGER NOT NULL REFERENCES vehicles(id),
                    note VARCHAR(120) NOT NULL
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE legacy_schedule_links (
                    id INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,
                    vehicle_schedule_id INTEGER NOT NULL REFERENCES transport_vehicle_schedules(id),
                    note VARCHAR(120) NOT NULL
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE legacy_assignment_links (
                    id INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,
                    assignment_id INTEGER NOT NULL REFERENCES transport_assignments(id),
                    note VARCHAR(120) NOT NULL
                )
                """
            )
        )
        conn.execute(
            text("INSERT INTO legacy_vehicle_links (vehicle_id, note) VALUES (:vehicle_id, 'legacy vehicle ref')"),
            {"vehicle_id": vehicle_id},
        )
        conn.execute(
            text(
                "INSERT INTO legacy_schedule_links (vehicle_schedule_id, note) VALUES (:schedule_id, 'legacy schedule ref')"
            ),
            {"schedule_id": schedule_id},
        )
        conn.execute(
            text(
                "INSERT INTO legacy_assignment_links (assignment_id, note) VALUES (:assignment_id, 'legacy assignment ref')"
            ),
            {"assignment_id": assignment_id},
        )

    try:
        with TestClient(app) as client:
            ensure_admin_session(client)
            removed = client.delete(
                f"/api/transport/vehicles/{schedule_id}",
                params={"service_date": friday.isoformat()},
            )
            assert removed.status_code == 200, removed.text

        with engine.begin() as conn:
            remaining_vehicle_links = conn.execute(text("SELECT COUNT(*) FROM legacy_vehicle_links")).scalar_one()
            remaining_schedule_links = conn.execute(text("SELECT COUNT(*) FROM legacy_schedule_links")).scalar_one()
            remaining_assignment_links = conn.execute(text("SELECT COUNT(*) FROM legacy_assignment_links")).scalar_one()

        with SessionLocal() as db:
            user_row = db.get(User, user_id)
            vehicle_row = db.get(Vehicle, vehicle_id)
            assignment_row = db.get(TransportAssignment, assignment_id)

        assert remaining_vehicle_links == 0
        assert remaining_schedule_links == 0
        assert remaining_assignment_links == 0
        assert user_row is not None
        assert user_row.placa is None
        assert vehicle_row is None
        assert assignment_row is None
    finally:
        with engine.begin() as conn:
            conn.execute(text("DROP TABLE IF EXISTS legacy_vehicle_links"))
            conn.execute(text("DROP TABLE IF EXISTS legacy_schedule_links"))
            conn.execute(text("DROP TABLE IF EXISTS legacy_assignment_links"))


def test_transport_regular_assignment_persists_across_weekdays_and_routes():
    friday = date(2026, 4, 17)
    monday = date(2026, 4, 20)
    timestamp = now_sgt()

    with SessionLocal() as db:
        db.add(Workplace(workplace="Mirror Hub", address="2 Mirror Road", zip="222222", country="Singapore"))
        vehicle = Vehicle(placa="REG8001", tipo="van", color="Blue", lugares=11, tolerance=8, service_scope="regular")
        db.add(vehicle)
        db.flush()
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )

        user = User(
            rfid=None,
            nome="Mirror Rider",
            chave="MR01",
            projeto="P80",
            workplace="Mirror Hub",
            end_rua="22 Mirror Street",
            zip="220022",
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        request_row = TransportRequest(
            user_id=user.id,
            request_kind="regular",
            recurrence_kind="weekday",
            requested_time="07:00",
            single_date=None,
            created_via="bot",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
            cancelled_at=None,
        )
        db.add(request_row)
        db.commit()
        request_id = request_row.id
        vehicle_id = vehicle.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            "/api/transport/assignments",
            json={
                "request_id": request_id,
                "service_date": friday.isoformat(),
                "route_kind": "home_to_work",
                "status": "confirmed",
                "vehicle_id": vehicle_id,
            },
        )
        monday_home_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": monday.isoformat(), "route_kind": "home_to_work"},
        )
        monday_work_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": monday.isoformat(), "route_kind": "work_to_home"},
        )

    assert response.status_code == 200
    assert monday_home_dashboard.status_code == 200
    assert monday_work_dashboard.status_code == 200

    with SessionLocal() as db:
        assignments = db.execute(
            select(TransportAssignment)
            .where(TransportAssignment.request_id == request_id)
            .order_by(TransportAssignment.route_kind)
        ).scalars().all()

    assert [row.route_kind for row in assignments] == ["home_to_work", "work_to_home"]
    assert all(row.status == "confirmed" for row in assignments)
    assert all(row.vehicle_id == vehicle_id for row in assignments)
    monday_home_request = next(
        row for row in monday_home_dashboard.json()["regular_requests"] if row["id"] == request_id
    )
    monday_work_request = next(
        row for row in monday_work_dashboard.json()["regular_requests"] if row["id"] == request_id
    )
    assert monday_home_request["assignment_status"] == "confirmed"
    assert monday_work_request["assignment_status"] == "confirmed"
    assert monday_home_request["assigned_vehicle"]["placa"] == "REG8001"
    assert monday_work_request["assigned_vehicle"]["placa"] == "REG8001"


def test_transport_regular_vehicle_registry_ignores_non_selected_weekdays_for_assigned_count():
    monday = date(2026, 4, 20)
    tuesday = date(2026, 4, 21)
    wednesday = date(2026, 4, 22)
    timestamp = now_sgt()

    with SessionLocal() as db:
        db.add(Workplace(workplace="Selected Weekday Hub", address="3 Weekday Road", zip="333333", country="Singapore"))
        vehicle = Vehicle(placa="REG8101", tipo="van", color="Gray", lugares=10, tolerance=7, service_scope="regular")
        db.add(vehicle)
        db.flush()
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )

        user = User(
            rfid=None,
            nome="Weekday Filter Rider",
            chave="WF01",
            projeto="P80",
            workplace="Selected Weekday Hub",
            end_rua="33 Filter Street",
            zip="330033",
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        request_row = TransportRequest(
            user_id=user.id,
            request_kind="regular",
            recurrence_kind="weekday",
            requested_time="07:10",
            selected_weekdays_json=json.dumps([0, 2, 4], ensure_ascii=True, separators=(",", ":")),
            single_date=None,
            created_via="bot",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
            cancelled_at=None,
        )
        db.add(request_row)
        db.commit()
        request_id = request_row.id
        vehicle_id = vehicle.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            "/api/transport/assignments",
            json={
                "request_id": request_id,
                "service_date": monday.isoformat(),
                "route_kind": "home_to_work",
                "status": "confirmed",
                "vehicle_id": vehicle_id,
            },
        )
        tuesday_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": tuesday.isoformat(), "route_kind": "home_to_work"},
        )
        wednesday_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": wednesday.isoformat(), "route_kind": "home_to_work"},
        )

    assert response.status_code == 200
    assert tuesday_dashboard.status_code == 200
    assert wednesday_dashboard.status_code == 200

    tuesday_payload = tuesday_dashboard.json()
    wednesday_payload = wednesday_dashboard.json()
    tuesday_request = next(row for row in tuesday_payload["regular_requests"] if row["id"] == request_id)
    wednesday_request = next(row for row in wednesday_payload["regular_requests"] if row["id"] == request_id)
    tuesday_vehicle_registry_row = next(
        row for row in tuesday_payload["regular_vehicle_registry"] if row["placa"] == "REG8101"
    )
    wednesday_vehicle_registry_row = next(
        row for row in wednesday_payload["regular_vehicle_registry"] if row["placa"] == "REG8101"
    )

    assert tuesday_request["service_date"] == wednesday.isoformat()
    assert tuesday_request["assignment_status"] == "confirmed"
    assert tuesday_request["assigned_vehicle"]["placa"] == "REG8101"
    assert tuesday_vehicle_registry_row["assigned_count"] == 0
    assert wednesday_vehicle_registry_row["assigned_count"] == 1
    assert wednesday_request["service_date"] == wednesday.isoformat()
    assert wednesday_request["assignment_status"] == "confirmed"
    assert wednesday_request["assigned_vehicle"]["placa"] == "REG8101"


def test_transport_weekend_vehicle_registry_ignores_non_selected_days_for_assigned_count():
    saturday = date(2026, 4, 18)
    sunday = date(2026, 4, 19)
    timestamp = now_sgt()

    with SessionLocal() as db:
        db.add(Workplace(workplace="Weekend Selected Hub", address="7 Weekend Lane", zip="777777", country="Singapore"))
        vehicle = Vehicle(placa="WKD8101", tipo="van", color="Silver", lugares=10, tolerance=8, service_scope="weekend")
        db.add(vehicle)
        db.flush()
        for weekday in (saturday.weekday(), sunday.weekday()):
            add_transport_schedule(
                db,
                vehicle=vehicle,
                service_scope="weekend",
                route_kind="home_to_work",
                recurrence_kind="matching_weekday",
                weekday=weekday,
            )
            add_transport_schedule(
                db,
                vehicle=vehicle,
                service_scope="weekend",
                route_kind="work_to_home",
                recurrence_kind="matching_weekday",
                weekday=weekday,
            )

        user = User(
            rfid=None,
            nome="Weekend Selected Rider",
            chave="WS01",
            projeto="P80",
            workplace="Weekend Selected Hub",
            end_rua="71 Weekend Street",
            zip="770071",
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        request_row = TransportRequest(
            user_id=user.id,
            request_kind="weekend",
            recurrence_kind="weekend",
            requested_time="09:10",
            selected_weekdays_json=json.dumps([6], ensure_ascii=True, separators=(",", ":")),
            single_date=None,
            created_via="bot",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
            cancelled_at=None,
        )
        db.add(request_row)
        db.commit()
        request_id = request_row.id
        vehicle_id = vehicle.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            "/api/transport/assignments",
            json={
                "request_id": request_id,
                "service_date": sunday.isoformat(),
                "route_kind": "home_to_work",
                "status": "confirmed",
                "vehicle_id": vehicle_id,
            },
        )
        saturday_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": saturday.isoformat(), "route_kind": "home_to_work"},
        )
        sunday_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": sunday.isoformat(), "route_kind": "home_to_work"},
        )

    assert response.status_code == 200
    assert saturday_dashboard.status_code == 200
    assert sunday_dashboard.status_code == 200

    saturday_payload = saturday_dashboard.json()
    sunday_payload = sunday_dashboard.json()
    saturday_request = next(row for row in saturday_payload["weekend_requests"] if row["id"] == request_id)
    sunday_request = next(row for row in sunday_payload["weekend_requests"] if row["id"] == request_id)
    saturday_vehicle_registry_row = next(
        row for row in saturday_payload["weekend_vehicle_registry"] if row["placa"] == "WKD8101"
    )
    sunday_vehicle_registry_row = next(
        row for row in sunday_payload["weekend_vehicle_registry"] if row["placa"] == "WKD8101"
    )

    assert saturday_request["service_date"] == sunday.isoformat()
    assert saturday_request["assignment_status"] == "confirmed"
    assert saturday_request["assigned_vehicle"]["placa"] == "WKD8101"
    assert saturday_vehicle_registry_row["assigned_count"] == 0
    assert sunday_vehicle_registry_row["assigned_count"] == 1
    assert sunday_request["service_date"] == sunday.isoformat()
    assert sunday_request["assignment_status"] == "confirmed"
    assert sunday_request["assigned_vehicle"]["placa"] == "WKD8101"


def test_transport_weekend_assignment_respects_selected_persistent_weekdays():
    saturday = date(2026, 4, 18)
    sunday = date(2026, 4, 19)
    timestamp = now_sgt()

    with SessionLocal() as db:
        db.add(Workplace(workplace="Weekend Persist Hub", address="9 Weekend Road", zip="919191", country="Singapore"))
        vehicle = Vehicle(placa="WKD8801", tipo="van", color="Blue", lugares=12, tolerance=11, service_scope="weekend")
        db.add(vehicle)
        db.flush()
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="weekend",
            route_kind="home_to_work",
            recurrence_kind="matching_weekday",
            weekday=saturday.weekday(),
        )
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="weekend",
            route_kind="work_to_home",
            recurrence_kind="matching_weekday",
            weekday=saturday.weekday(),
        )

        user = User(
            rfid=None,
            nome="Weekend Persist Rider",
            chave="WP88",
            projeto="P80",
            workplace="Weekend Persist Hub",
            end_rua="88 Weekend Street",
            zip="880088",
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        request_row = TransportRequest(
            user_id=user.id,
            request_kind="weekend",
            recurrence_kind="weekend",
            requested_time="09:00",
            single_date=None,
            created_via="bot",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
            cancelled_at=None,
        )
        db.add(request_row)
        db.commit()
        request_id = request_row.id
        vehicle_id = vehicle.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            "/api/transport/assignments",
            json={
                "request_id": request_id,
                "service_date": saturday.isoformat(),
                "route_kind": "home_to_work",
                "status": "confirmed",
                "vehicle_id": vehicle_id,
            },
        )
        saturday_home_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": saturday.isoformat(), "route_kind": "home_to_work"},
        )
        saturday_work_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": saturday.isoformat(), "route_kind": "work_to_home"},
        )
        sunday_home_dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": sunday.isoformat(), "route_kind": "home_to_work"},
        )

    assert response.status_code == 200
    assert saturday_home_dashboard.status_code == 200
    assert saturday_work_dashboard.status_code == 200
    assert sunday_home_dashboard.status_code == 200

    saturday_home_request = next(
        row for row in saturday_home_dashboard.json()["weekend_requests"] if row["id"] == request_id
    )
    saturday_work_request = next(
        row for row in saturday_work_dashboard.json()["weekend_requests"] if row["id"] == request_id
    )
    sunday_home_request = next(
        row for row in sunday_home_dashboard.json()["weekend_requests"] if row["id"] == request_id
    )
    assert saturday_home_request["assignment_status"] == "confirmed"
    assert saturday_work_request["assignment_status"] == "confirmed"
    assert sunday_home_request["assignment_status"] == "pending"
    assert saturday_home_request["assigned_vehicle"]["placa"] == "WKD8801"
    assert saturday_work_request["assigned_vehicle"]["placa"] == "WKD8801"


def test_transport_dashboard_surfaces_extra_assignment_from_any_route():
    friday = date(2026, 4, 17)
    timestamp = now_sgt()

    with SessionLocal() as db:
        db.add(Workplace(workplace="Extra Route Hub", address="14 Route Way", zip="414141", country="Singapore"))
        vehicle = Vehicle(placa="EXT8101", tipo="carro", color="Black", lugares=4, tolerance=5, service_scope="extra")
        db.add(vehicle)
        db.flush()
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="extra",
            route_kind="work_to_home",
            recurrence_kind="single_date",
            service_date=friday,
            departure_time="18:10",
        )

        user = User(
            rfid=None,
            nome="Extra Route Rider",
            chave="ER01",
            projeto="P80",
            workplace="Extra Route Hub",
            end_rua="14 Route Street",
            zip="410014",
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        request_row = TransportRequest(
            user_id=user.id,
            request_kind="extra",
            recurrence_kind="single_date",
            requested_time="18:10",
            selected_weekdays_json=None,
            single_date=friday,
            created_via="bot",
            status="active",
            created_at=timestamp,
            updated_at=timestamp,
            cancelled_at=None,
        )
        db.add(request_row)
        db.commit()
        request_id = request_row.id
        vehicle_id = vehicle.id

    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            "/api/transport/assignments",
            json={
                "request_id": request_id,
                "service_date": friday.isoformat(),
                "route_kind": "work_to_home",
                "status": "confirmed",
                "vehicle_id": vehicle_id,
            },
        )
        dashboard = client.get(
            "/api/transport/dashboard",
            params={"service_date": friday.isoformat(), "route_kind": "home_to_work"},
        )

    assert response.status_code == 200
    assert dashboard.status_code == 200

    payload = dashboard.json()
    extra_request = next(row for row in payload["extra_requests"] if row["id"] == request_id)

    assert extra_request["assignment_status"] == "confirmed"
    assert extra_request["assigned_vehicle"]["placa"] == "EXT8101"
    assert extra_request["assigned_vehicle"]["route_kind"] == "work_to_home"


def test_admin_page_is_served_on_admin_path():
    with TestClient(app) as client:
        response = client.get("/admin")
        assert response.status_code == 200
        assert "Checking Admin" in response.text
        assert "Acesso Administrativo" in response.text
    assert "Projetos" in response.text
    assert 'id="projectsBody"' in response.text
    assert 'id="addProjectButton"' in response.text


def test_gerencia_page_is_not_served_anymore():
    with TestClient(app) as client:
        response = client.get("/gerencia")
    assert response.status_code == 404


def test_gerencia_trailing_slash_is_not_served_anymore():
    with TestClient(app) as client:
        response = client.get("/gerencia/", follow_redirects=False)
    assert response.status_code == 404


def test_database_events_endpoint_filters_and_paginates_check_events():
    primary_key = f"D{uuid.uuid4().hex[:3].upper()}"
    secondary_key = f"E{uuid.uuid4().hex[:3].upper()}"
    primary_rfid = f"rfid-db-{uuid.uuid4().hex[:8]}"
    secondary_rfid = f"rfid-db-{uuid.uuid4().hex[:8]}"
    timestamp = now_sgt().replace(microsecond=0)
    idempotency_keys = [uuid.uuid4().hex for _ in range(4)]

    try:
        with SessionLocal() as db:
            db.add_all(
                [
                    User(
                        rfid=primary_rfid,
                        nome="Usuario Banco Primario",
                        chave=primary_key,
                        projeto="P80",
                        workplace=None,
                        placa=None,
                        end_rua=None,
                        zip=None,
                        email=None,
                        local=None,
                        checkin=None,
                        time=None,
                        last_active_at=timestamp,
                        inactivity_days=0,
                    ),
                    User(
                        rfid=secondary_rfid,
                        nome="Usuario Banco Secundario",
                        chave=secondary_key,
                        projeto="P82",
                        workplace=None,
                        placa=None,
                        end_rua=None,
                        zip=None,
                        email=None,
                        local=None,
                        checkin=None,
                        time=None,
                        last_active_at=timestamp,
                        inactivity_days=0,
                    ),
                ]
            )
            db.add_all(
                [
                    CheckEvent(
                        idempotency_key=idempotency_keys[0],
                        source="device",
                        rfid=primary_rfid,
                        action="checkin",
                        status="success",
                        message="Entrada liberada",
                        details="reader=gate-1",
                        project="P80",
                        device_id="ESP-DB-01",
                        local="Portaria Norte",
                        request_path="/api/scan",
                        http_status=200,
                        ontime=True,
                        event_time=timestamp - timedelta(minutes=10),
                        submitted_at=timestamp - timedelta(minutes=10),
                        retry_count=0,
                    ),
                    CheckEvent(
                        idempotency_key=idempotency_keys[1],
                        source="mobile",
                        rfid=primary_rfid,
                        action="checkout",
                        status="queued",
                        message="Saida enviada pelo app",
                        details="client_event_id=db-checkout-1",
                        project="P80",
                        device_id="APP-DB-01",
                        local="Portaria Norte",
                        request_path="/api/mobile/events/sync",
                        http_status=202,
                        ontime=False,
                        event_time=timestamp - timedelta(minutes=2),
                        submitted_at=timestamp - timedelta(minutes=2),
                        retry_count=1,
                    ),
                    CheckEvent(
                        idempotency_key=idempotency_keys[2],
                        source="forms",
                        rfid=secondary_rfid,
                        action="checkin",
                        status="success",
                        message="Entrada confirmada",
                        details="provider=meta",
                        project="P82",
                        device_id="FORMS-DB-01",
                        local="Oficina Sul",
                        request_path="/api/provider/forms",
                        http_status=200,
                        ontime=True,
                        event_time=timestamp - timedelta(minutes=1),
                        submitted_at=timestamp - timedelta(minutes=1),
                        retry_count=0,
                    ),
                    CheckEvent(
                        idempotency_key=idempotency_keys[3],
                        source="admin",
                        rfid=None,
                        action="admin_request",
                        status="success",
                        message="Solicitacao administrativa",
                        details="chave=HR70",
                        project=None,
                        device_id=None,
                        local=None,
                        request_path="/api/admin/administrators/request",
                        http_status=200,
                        ontime=None,
                        event_time=timestamp,
                        submitted_at=timestamp,
                        retry_count=0,
                    ),
                ]
            )
            db.commit()

        with TestClient(app) as client:
            ensure_admin_session(client)

            first_page = client.get(
                "/api/admin/database-events",
                params={"chave": primary_key, "page_size": 1},
            )
            assert first_page.status_code == 200, first_page.text
            first_payload = first_page.json()
            assert first_payload["total"] == 2
            assert first_payload["page"] == 1
            assert first_payload["page_size"] == 1
            assert first_payload["total_pages"] == 2
            filter_options = first_payload["filter_options"]
            assert set(filter_options["action"]).issuperset({"checkin", "checkout"})
            assert set(filter_options["chave"]).issuperset({primary_key, secondary_key})
            assert set(filter_options["rfid"]).issuperset({primary_rfid, secondary_rfid})
            assert set(filter_options["project"]).issuperset({"P80", "P82"})
            assert set(filter_options["source"]).issuperset({"device", "forms", "mobile"})
            assert set(filter_options["status"]).issuperset({"queued", "success"})
            assert len(first_payload["items"]) == 1
            assert first_payload["items"][0]["chave"] == primary_key
            assert first_payload["items"][0]["action"] == "checkout"

            second_page = client.get(
                "/api/admin/database-events",
                params={"chave": primary_key, "page_size": 1, "page": 2},
            )
            assert second_page.status_code == 200, second_page.text
            second_payload = second_page.json()
            assert second_payload["page"] == 2
            assert len(second_payload["items"]) == 1
            assert second_payload["items"][0]["action"] == "checkin"

            key_sort_asc = client.get(
                "/api/admin/database-events",
                params={"sort_by": "chave", "sort_direction": "asc", "page_size": 200},
            )
            assert key_sort_asc.status_code == 200, key_sort_asc.text
            key_sort_asc_payload = key_sort_asc.json()
            key_sort_asc_chaves = [
                item["chave"]
                for item in key_sort_asc_payload["items"]
                if item["chave"] in {primary_key, secondary_key}
            ]
            assert key_sort_asc_chaves == [primary_key, primary_key, secondary_key]

            key_sort_desc = client.get(
                "/api/admin/database-events",
                params={"sort_by": "chave", "sort_direction": "desc", "page_size": 200},
            )
            assert key_sort_desc.status_code == 200, key_sort_desc.text
            key_sort_desc_payload = key_sort_desc.json()
            key_sort_desc_chaves = [
                item["chave"]
                for item in key_sort_desc_payload["items"]
                if item["chave"] in {primary_key, secondary_key}
            ]
            assert key_sort_desc_chaves == [secondary_key, primary_key, primary_key]

            message_sort_asc = client.get(
                "/api/admin/database-events",
                params={"chave": primary_key, "sort_by": "message", "sort_direction": "asc", "page_size": 10},
            )
            assert message_sort_asc.status_code == 200, message_sort_asc.text
            message_sort_asc_payload = message_sort_asc.json()
            assert [item["message"] for item in message_sort_asc_payload["items"]] == [
                "Entrada liberada",
                "Saida enviada pelo app",
            ]

            search_response = client.get(
                "/api/admin/database-events",
                params={"search": "oficina", "project": "P82"},
            )
            assert search_response.status_code == 200, search_response.text
            search_payload = search_response.json()
            assert search_payload["total"] == 1
            assert search_payload["items"][0]["chave"] == secondary_key
            assert search_payload["items"][0]["local"] == "Oficina Sul"
            assert search_payload["items"][0]["action"] == "checkin"
    finally:
        with SessionLocal() as db:
            for row in db.execute(select(CheckEvent).where(CheckEvent.idempotency_key.in_(idempotency_keys))).scalars().all():
                db.delete(row)
            for user in db.execute(select(User).where(User.chave.in_([primary_key, secondary_key]))).scalars().all():
                db.delete(user)
            db.commit()


def test_database_events_endpoint_rejects_invalid_date_ranges():
    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.get(
            "/api/admin/database-events",
            params={"from_date": "2025-01-10", "to_date": "2025-01-09"},
        )
        assert response.status_code == 400
        assert response.json()["detail"] == "Intervalo de datas invalido para a consulta de eventos."


def test_database_events_endpoint_rejects_invalid_sort_parameters():
    with TestClient(app) as client:
        ensure_admin_session(client)

        invalid_column = client.get(
            "/api/admin/database-events",
            params={"sort_by": "unexpected_column"},
        )
        assert invalid_column.status_code == 400
        assert invalid_column.json()["detail"] == "Coluna invalida para ordenacao de eventos."

        invalid_direction = client.get(
            "/api/admin/database-events",
            params={"sort_direction": "sideways"},
        )
        assert invalid_direction.status_code == 400
        assert invalid_direction.json()["detail"] == "Direcao invalida para ordenacao de eventos."


def test_web_password_registration_requires_existing_key_and_hashes_password():
    with TestClient(app) as client:
        response = register_web_password(client, chave="WB11", senha="s#123", projeto="P82")
        assert response.status_code == 200
        payload = response.json()
        assert payload["ok"] is True
        assert payload["authenticated"] is True
        assert payload["has_password"] is True

        status = client.get("/api/web/auth/status", params={"chave": "WB11"})
        assert status.status_code == 200
        assert status.json() == {
            "found": True,
            "chave": "WB11",
            "has_password": True,
            "authenticated": True,
            "message": "Aplicacao liberada.",
        }

        with SessionLocal() as db:
            user = get_user_by_chave(db, "WB11")
            assert user.nome == "Oriundo da Web"
            assert user.rfid is None
            assert user.projeto == "P82"
            assert user.senha is not None
            assert user.senha != "s#123"
            assert user.senha.startswith("pbkdf2_sha256$")
            assert verify_password("s#123", user.senha) is True


def test_web_password_registration_returns_not_found_for_unknown_key():
    with TestClient(app) as client:
        response = register_web_password(
            client,
            chave="ZZ91",
            senha="s#123",
            projeto="P82",
            ensure_user_exists=False,
        )

        assert response.status_code == 404
        assert response.json()["detail"] == "A chave do usuario nao esta cadastrada"


def test_web_user_self_registration_creates_common_user_and_authenticates_web_session():
    with TestClient(app) as client:
        response = client.post(
            "/api/web/auth/register-user",
            json={
                "chave": "WU11",
                "nome": "maria jose da silva",
                "projetos": ["P83"],
                "email": "maria.jose@petrobras.com.br",
                "senha": "cad123",
                "confirmar_senha": "cad123",
            },
        )

        assert response.status_code == 201
        payload = response.json()
        assert payload["ok"] is True
        assert payload["authenticated"] is True
        assert payload["has_password"] is True
        assert payload["message"] == "Cadastro concluido com sucesso."
        assert payload["projects"] == ["P83"]
        assert payload["active_project"] == "P83"

        status = client.get("/api/web/auth/status", params={"chave": "WU11"})
        assert status.status_code == 200
        assert status.json() == {
            "found": True,
            "chave": "WU11",
            "has_password": True,
            "authenticated": True,
            "message": "Aplicacao liberada.",
        }

    with SessionLocal() as db:
        user = get_user_by_chave(db, "WU11")
        pending = db.execute(select(AdminAccessRequest).where(AdminAccessRequest.chave == "WU11")).scalar_one_or_none()
        assert user is not None
        assert user.nome == "Maria Jose da Silva"
        assert user.projeto == "P83"
        assert user.perfil == 0
        assert user.end_rua is None
        assert user.zip is None
        assert user.email == "maria.jose@petrobras.com.br"
        assert user.senha is not None
        assert user.senha != "cad123"
        assert verify_password("cad123", user.senha) is True
        assert pending is None

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        administrators = admin_client.get("/api/admin/administrators")
        assert administrators.status_code == 200
        assert not any(
            row["row_type"] == "request" and row["chave"] == "WU11"
            for row in administrators.json()
        )

    with TestClient(app) as transport_client:
        denied_before_login = transport_client.get("/api/transport/auth/session")
        assert denied_before_login.status_code == 200
        assert denied_before_login.json()["authenticated"] is False

        transport_login = transport_client.post(
            "/api/transport/auth/verify",
            json={"chave": "WU11", "senha": "cad123"},
        )
        assert transport_login.status_code == 200
        assert transport_login.json()["authenticated"] is False
        assert transport_login.json()["message"] == "This user does not have transport access."
        assert transport_login.json()["message_key"] == "auth.noAccess"
        assert transport_login.json()["error_code"] == "transport_auth_access_denied"


def test_web_user_self_registration_accepts_plural_memberships_and_seeds_active_project():
    with TestClient(app) as client:
        response = client.post(
            "/api/web/auth/register-user",
            json={
                "chave": "WU13",
                "nome": "ana multi projeto",
                "projetos": ["P83", "P80"],
                "email": "ana.multi@petrobras.com.br",
                "senha": "cad456",
                "confirmar_senha": "cad456",
            },
        )

        assert response.status_code == 201, response.text
        assert response.json()["ok"] is True
        assert response.json()["authenticated"] is True
        assert response.json()["has_password"] is True
        assert response.json()["projects"] == ["P80", "P83"]
        assert response.json()["active_project"] == "P80"

        user_projects = client.get("/api/web/user-projects")
        assert user_projects.status_code == 200, user_projects.text
        assert user_projects.json() == {
            "projects": ["P80", "P83"],
            "active_project": "P80",
        }

    with SessionLocal() as db:
        user = get_user_by_chave(db, "WU13")
        assert user is not None
        assert user.projeto == "P80"
        assert list_user_project_names(db, user) == ["P80", "P83"]



def test_web_user_self_registration_accepts_optional_email_blank():
    with TestClient(app) as client:
        response = client.post(
            "/api/web/auth/register-user",
            json={
                "chave": "WU12",
                "nome": "joao sem email",
                "projetos": ["P82"],
                "senha": "cad321",
                "confirmar_senha": "cad321",
            },
        )

        assert response.status_code == 201

    with SessionLocal() as db:
        user = get_user_by_chave(db, "WU12")
        assert user is not None
        assert user.email is None


def test_web_check_endpoints_require_authenticated_password_session():
    payload = {
        "chave": "WB90",
        "projeto": "P80",
        "action": "checkin",
        "informe": "normal",
        "event_time": now_sgt().isoformat(),
        "client_event_id": f"web-check-auth-{uuid.uuid4().hex}",
    }

    with TestClient(app) as client:
        submit_response = client.post("/api/web/check", json=payload)
        history_response = client.get("/api/web/check/state", params={"chave": "WB90"})
        locations_response = client.get("/api/web/check/locations")
        location_response = client.post(
            "/api/web/check/location",
            json={
                "latitude": 1.255936,
                "longitude": 103.611066,
                "accuracy_meters": 8,
            },
        )

        assert submit_response.status_code == 401
        assert history_response.status_code == 401
        assert locations_response.status_code == 401
        assert location_response.status_code == 401


def test_web_password_change_replaces_previous_password():
    with TestClient(app) as client:
        registered = register_web_password(client, chave="WB15", senha="abc123", projeto="P80")
        assert registered.status_code == 200

        wrong_change = client.post(
            "/api/web/auth/change-password",
            json={
                "chave": "WB15",
                "senha_antiga": "000000",
                "nova_senha": "n0va#1",
            },
        )
        assert wrong_change.status_code == 401

        changed = client.post(
            "/api/web/auth/change-password",
            json={
                "chave": "WB15",
                "senha_antiga": "abc123",
                "nova_senha": "n0va#1",
            },
        )
        assert changed.status_code == 200
        assert changed.json()["authenticated"] is True
        assert changed.json()["has_password"] is True

        old_login = login_web_password(client, chave="WB15", senha="abc123")
        assert old_login.status_code == 401

        new_login = login_web_password(client, chave="WB15", senha="n0va#1")
        assert new_login.status_code == 200
        assert new_login.json()["authenticated"] is True
        assert new_login.json()["has_password"] is True


def test_web_transport_vehicle_request_returns_pending_state_without_same_day_checkin_requirement(monkeypatch):
    fixed_now = datetime(2026, 4, 17, 7, 30, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    ensure_web_user_exists(chave="WT11", projeto="P80", nome="Transport Web Rider")
    with SessionLocal() as db:
        user = get_user_by_chave(db, "WT11")
        user.end_rua = "10 Marina Boulevard"
        user.zip = "123456"
        db.commit()

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WT11",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        created = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT11", "request_kind": "regular"},
        )
        assert created.status_code == 200
        payload = created.json()
        assert payload["ok"] is True
        assert payload["state"]["status"] == "pending"
        assert payload["state"]["request_kind"] == "regular"
        assert payload["state"]["requested_time"] == "07:30"
        assert payload["state"]["confirmation_deadline_time"] == "07:30"
        assert payload["state"]["end_rua"] == "10 Marina Boulevard"
        assert payload["state"]["zip"] == "123456"

    with SessionLocal() as db:
        user = get_user_by_chave(db, "WT11")
        request_row = db.execute(
            select(TransportRequest)
            .where(
                TransportRequest.user_id == user.id,
                TransportRequest.request_kind == "regular",
                TransportRequest.status == "active",
            )
            .order_by(TransportRequest.id.desc())
            .limit(1)
        ).scalar_one()

    assert request_row.created_via == "web"
    assert request_row.requested_time == "07:30"

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        dashboard = admin_client.get(
            "/api/transport/dashboard",
            params={"service_date": fixed_now.date().isoformat(), "route_kind": "home_to_work"},
        )

    assert dashboard.status_code == 200
    assert any(row["chave"] == "WT11" for row in dashboard.json()["regular_requests"])


def test_web_transport_vehicle_request_rejects_incomplete_address(monkeypatch):
    fixed_now = datetime(2026, 4, 17, 7, 45, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    ensure_web_user_exists(chave="WT10", projeto="P80", nome="Missing Address Rider")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WT10",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        created = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT10", "request_kind": "regular"},
        )

    assert created.status_code == 400
    assert created.json()["detail"] == "Cadastre um endereco completo antes de solicitar o transporte."

    with SessionLocal() as db:
        user = get_user_by_chave(db, "WT10")
        request_rows = db.execute(
            select(TransportRequest).where(TransportRequest.user_id == user.id)
        ).scalars().all()

    assert request_rows == []


def test_web_transport_vehicle_request_rejects_missing_extra_date_or_time(monkeypatch):
    fixed_now = datetime(2026, 4, 18, 9, 30, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    ensure_web_user_exists(chave="WT15", projeto="P83", nome="Incomplete Extra Rider")
    ensure_web_transport_address(chave="WT15")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WT15",
            senha="abc123",
            projeto="P83",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        missing_date = client.post(
            "/api/web/transport/vehicle-request",
            json={
                "chave": "WT15",
                "request_kind": "extra",
                "requested_time": "18:10",
            },
        )
        missing_time = client.post(
            "/api/web/transport/vehicle-request",
            json={
                "chave": "WT15",
                "request_kind": "extra",
                "requested_date": fixed_now.date().isoformat(),
            },
        )

    assert missing_date.status_code == 400
    assert missing_date.json()["detail"] == "Informe a data do transporte extra."
    assert missing_time.status_code == 400
    assert missing_time.json()["detail"] == "Informe o horario do transporte extra."

    with SessionLocal() as db:
        user = get_user_by_chave(db, "WT15")
        request_rows = db.execute(
            select(TransportRequest).where(TransportRequest.user_id == user.id)
        ).scalars().all()

    assert request_rows == []


def test_web_transport_stream_requires_authenticated_matching_session():
    ensure_web_user_exists(chave="WT12", projeto="P80", nome="Transport Stream Guard")

    with TestClient(app) as client:
        response = client.get("/api/web/transport/stream", params={"chave": "WT12"})

    assert response.status_code == 401
    assert response.json()["detail"] == "Sessao do usuario invalida ou expirada"


def test_web_transport_stream_emits_connected_and_transport_events():
    ensure_web_user_exists(chave="WT13", projeto="P80", nome="Transport Stream Rider")

    with SessionLocal() as db:
        user = get_user_by_chave(db, "WT13")
        user.senha = hash_password("abc123")
        db.commit()

    class DummyWebTransportStreamRequest:
        def __init__(self, session):
            self.session = session

        async def is_disconnected(self):
            return False

    request = DummyWebTransportStreamRequest({web_check_router.WEB_USER_SESSION_KEY: "WT13"})

    with SessionLocal() as db:
        response = asyncio.run(
            web_check_router.stream_web_transport_updates(
                request,
                chave="WT13",
                db=db,
            )
        )
        assert response.media_type == "text/event-stream"
        assert response.headers["cache-control"] == "no-cache"

        first_chunk = asyncio.run(asyncio.wait_for(anext(response.body_iterator), timeout=1))
        first_payload = json.loads(first_chunk.removeprefix("data: ").strip())
        assert first_payload["reason"] == "connected"

        notify_transport_data_changed("event")

        second_chunk = asyncio.run(asyncio.wait_for(anext(response.body_iterator), timeout=1))
        second_payload = json.loads(second_chunk.removeprefix("data: ").strip())
        assert second_payload["reason"] == "event"

        asyncio.run(response.body_iterator.aclose())


def test_transport_reevaluation_catalog_tracks_recent_request_trigger(monkeypatch):
    fixed_now = datetime(2026, 4, 28, 9, 5, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)
    transport_reevaluation_module.clear_transport_reevaluation_events()

    ensure_web_user_exists(chave="WT61", projeto="P80", nome="Reevaluation Request Rider")
    ensure_web_transport_address(chave="WT61")
    set_user_checkin_state(chave="WT61", event_time=fixed_now, local="Reevaluation Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WT61",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        created = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT61", "request_kind": "regular"},
        )
        assert created.status_code == 200, created.text
        request_id = created.json()["state"]["request_id"]

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        catalog_response = admin_client.get("/api/transport/reevaluation-events", params={"limit": 5})

    assert catalog_response.status_code == 200, catalog_response.text
    payload = catalog_response.json()
    request_catalog_entry = next(row for row in payload["catalog"] if row["event_type"] == "transport_request_changed")
    latest_event = payload["recent_events"][0]

    assert request_catalog_entry["downstream_actions"] == [
        "refresh_snapshot",
        "revalidate_constraints",
        "rebuild_proposal",
        "regenerate_export",
        "refresh_transport_state",
    ]
    assert latest_event["event_type"] == "transport_request_changed"
    assert latest_event["source"] == "web_transport"
    assert latest_event["reason"] == "event"
    assert latest_event["request_id"] == request_id
    assert latest_event["downstream_actions"] == request_catalog_entry["downstream_actions"]


def test_transport_proposal_approval_emits_operational_review_trigger():
    transport_reevaluation_module.clear_transport_reevaluation_events()
    friday = date(2026, 4, 17)
    captured_at = datetime(2026, 4, 16, 21, 0, tzinfo=ZoneInfo(settings.tz_name))
    created_at = datetime(2026, 4, 16, 21, 5, tzinfo=ZoneInfo(settings.tz_name))

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="18:10")
        workplace = Workplace(
            workplace="Reevaluation Review Hub",
            address="51 Review Road",
            zip="941511",
            country="Singapore",
        )
        db.add(workplace)
        db.flush()

        vehicle = Vehicle(placa="REV6101", tipo="van", color="White", lugares=10, tolerance=8, service_scope="regular")
        db.add(vehicle)
        db.flush()

        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )

        user = User(
            rfid=None,
            nome="Reevaluation Review Rider",
            chave="RV61",
            projeto="P61",
            workplace=workplace.workplace,
            placa=vehicle.placa,
            end_rua="61 Review Avenue",
            zip="941512",
            local=None,
            checkin=None,
            time=None,
            last_active_at=created_at,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()

        transport_request = TransportRequest(
            user_id=user.id,
            request_kind="regular",
            recurrence_kind="weekday",
            requested_time="07:35",
            selected_weekdays_json=json.dumps([0, 1, 2, 3, 4]),
            single_date=None,
            created_via="admin",
            status="active",
            created_at=created_at,
            updated_at=created_at,
        )
        db.add(transport_request)
        db.commit()

        proposal = transport_proposal_service_module.build_transport_operational_proposal(
            snapshot=transport_proposal_service_module.build_transport_operational_snapshot(
                db,
                service_date=friday,
                route_kind="home_to_work",
                captured_at=captured_at,
            ),
            origin="manual",
            created_at=created_at,
            decisions=[
                TransportProposalDecision(
                    request_id=transport_request.id,
                    request_kind="regular",
                    service_date=friday,
                    route_kind="home_to_work",
                    suggested_status="confirmed",
                    vehicle_id=vehicle.id,
                    rationale="Review trigger validation.",
                )
            ],
        )

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        approve_response = admin_client.post(
            "/api/transport/proposals/approve",
            json=proposal.model_dump(mode="json"),
        )
        assert approve_response.status_code == 200, approve_response.text

        events_response = admin_client.get("/api/transport/reevaluation-events", params={"limit": 3})

    assert events_response.status_code == 200, events_response.text
    latest_event = events_response.json()["recent_events"][0]
    assert latest_event["event_type"] == "transport_operational_review_changed"
    assert latest_event["source"] == "transport_proposal"
    assert latest_event["proposal_key"] == approve_response.json()["proposal"]["proposal_key"]
    assert latest_event["service_date"] == friday.isoformat()
    assert latest_event["route_kind"] == "home_to_work"


def test_web_transport_vehicle_request_allows_weekend_and_extra_requests_for_the_same_service_date(monkeypatch):
    fixed_now = datetime(2026, 4, 18, 9, 15, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    ensure_web_user_exists(chave="WT14", projeto="P83", nome="Weekend Transport Rider")
    ensure_web_transport_address(chave="WT14")
    set_user_checkin_state(chave="WT14", event_time=fixed_now, local="Weekend Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WT14",
            senha="abc123",
            projeto="P83",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        weekend_created = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT14", "request_kind": "weekend"},
        )
        assert weekend_created.status_code == 200
        assert weekend_created.json()["state"]["status"] == "pending"
        assert weekend_created.json()["state"]["request_kind"] == "weekend"

        extra_created = client.post(
            "/api/web/transport/vehicle-request",
            json={
                "chave": "WT14",
                "request_kind": "extra",
                "requested_date": fixed_now.date().isoformat(),
                "requested_time": "18:10",
            },
        )
        assert extra_created.status_code == 200
        assert extra_created.json()["state"]["status"] == "pending"

    with SessionLocal() as db:
        user = get_user_by_chave(db, "WT14")
        request_kinds = db.execute(
            select(TransportRequest.request_kind)
            .where(
                TransportRequest.user_id == user.id,
                TransportRequest.status == "active",
            )
            .order_by(TransportRequest.id)
        ).scalars().all()

    assert request_kinds == ["weekend", "extra"]

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        dashboard = admin_client.get(
            "/api/transport/dashboard",
            params={"service_date": fixed_now.date().isoformat(), "route_kind": "home_to_work"},
        )

    assert dashboard.status_code == 200
    weekend_row = next(row for row in dashboard.json()["weekend_requests"] if row["chave"] == "WT14")
    extra_row = next(row for row in dashboard.json()["extra_requests"] if row["chave"] == "WT14")

    assert weekend_row["service_date"] == fixed_now.date().isoformat()
    assert weekend_row["assignment_status"] == "pending"
    assert extra_row["service_date"] == fixed_now.date().isoformat()
    assert extra_row["requested_time"] == "18:10"
    assert extra_row["assignment_status"] == "pending"


def test_web_transport_vehicle_request_allows_extra_request_for_the_same_regular_service_date(monkeypatch):
    fixed_now = datetime(2026, 4, 20, 9, 15, tzinfo=ZoneInfo(settings.tz_name))
    next_tuesday = fixed_now.date() + timedelta(days=1)
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    ensure_web_user_exists(chave="WT16", projeto="P80", nome="Regular Transport Rider")
    ensure_web_transport_address(chave="WT16")
    set_user_checkin_state(chave="WT16", event_time=fixed_now, local="Monday Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WT16",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        regular_created = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT16", "request_kind": "regular", "selected_weekdays": [1]},
        )
        assert regular_created.status_code == 200
        assert regular_created.json()["state"]["status"] == "pending"
        assert regular_created.json()["state"]["requests"][0]["service_date"] == next_tuesday.isoformat()

        extra_created = client.post(
            "/api/web/transport/vehicle-request",
            json={
                "chave": "WT16",
                "request_kind": "extra",
                "requested_date": next_tuesday.isoformat(),
                "requested_time": "18:10",
            },
        )
        assert extra_created.status_code == 200
        assert extra_created.json()["state"]["status"] == "pending"

    with SessionLocal() as db:
        user = get_user_by_chave(db, "WT16")
        request_kinds = db.execute(
            select(TransportRequest.request_kind)
            .where(
                TransportRequest.user_id == user.id,
                TransportRequest.status == "active",
            )
            .order_by(TransportRequest.id)
        ).scalars().all()

    assert request_kinds == ["regular", "extra"]

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        dashboard = admin_client.get(
            "/api/transport/dashboard",
            params={"service_date": next_tuesday.isoformat(), "route_kind": "home_to_work"},
        )

    assert dashboard.status_code == 200
    regular_row = next(row for row in dashboard.json()["regular_requests"] if row["chave"] == "WT16")
    extra_row = next(row for row in dashboard.json()["extra_requests"] if row["chave"] == "WT16")

    assert regular_row["service_date"] == next_tuesday.isoformat()
    assert regular_row["assignment_status"] == "pending"
    assert extra_row["service_date"] == next_tuesday.isoformat()
    assert extra_row["requested_time"] == "18:10"
    assert extra_row["assignment_status"] == "pending"


def test_web_transport_vehicle_request_allows_regular_request_for_the_same_extra_service_date(monkeypatch):
    fixed_now = datetime(2026, 4, 20, 10, 5, tzinfo=ZoneInfo(settings.tz_name))
    next_tuesday = fixed_now.date() + timedelta(days=1)
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    ensure_web_user_exists(chave="WT17", projeto="P80", nome="Extra First Transport Rider")
    ensure_web_transport_address(chave="WT17")
    set_user_checkin_state(chave="WT17", event_time=fixed_now, local="Monday Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WT17",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        extra_created = client.post(
            "/api/web/transport/vehicle-request",
            json={
                "chave": "WT17",
                "request_kind": "extra",
                "requested_date": next_tuesday.isoformat(),
                "requested_time": "18:10",
            },
        )
        assert extra_created.status_code == 200
        assert extra_created.json()["state"]["status"] == "pending"

        regular_created = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT17", "request_kind": "regular", "selected_weekdays": [1]},
        )
        assert regular_created.status_code == 200
        assert regular_created.json()["state"]["status"] == "pending"
        assert regular_created.json()["state"]["requests"][0]["service_date"] == next_tuesday.isoformat()

    with SessionLocal() as db:
        user = get_user_by_chave(db, "WT17")
        request_kinds = db.execute(
            select(TransportRequest.request_kind)
            .where(
                TransportRequest.user_id == user.id,
                TransportRequest.status == "active",
            )
            .order_by(TransportRequest.id)
        ).scalars().all()

    assert request_kinds == ["extra", "regular"]

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        dashboard = admin_client.get(
            "/api/transport/dashboard",
            params={"service_date": next_tuesday.isoformat(), "route_kind": "home_to_work"},
        )

    assert dashboard.status_code == 200
    regular_row = next(row for row in dashboard.json()["regular_requests"] if row["chave"] == "WT17")
    extra_row = next(row for row in dashboard.json()["extra_requests"] if row["chave"] == "WT17")

    assert regular_row["service_date"] == next_tuesday.isoformat()
    assert regular_row["assignment_status"] == "pending"
    assert extra_row["service_date"] == next_tuesday.isoformat()
    assert extra_row["requested_time"] == "18:10"
    assert extra_row["assignment_status"] == "pending"


def test_web_transport_vehicle_request_rejects_second_extra_request_for_the_same_service_date(monkeypatch):
    fixed_now = datetime(2026, 4, 22, 8, 45, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    ensure_web_user_exists(chave="WT18", projeto="P83", nome="Duplicate Extra Rider")
    ensure_web_transport_address(chave="WT18")
    set_user_checkin_state(chave="WT18", event_time=fixed_now, local="Morning Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WT18",
            senha="abc123",
            projeto="P83",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        extra_created = client.post(
            "/api/web/transport/vehicle-request",
            json={
                "chave": "WT18",
                "request_kind": "extra",
                "requested_date": fixed_now.date().isoformat(),
                "requested_time": "18:10",
            },
        )
        assert extra_created.status_code == 200
        assert extra_created.json()["state"]["status"] == "pending"

        duplicate_extra = client.post(
            "/api/web/transport/vehicle-request",
            json={
                "chave": "WT18",
                "request_kind": "extra",
                "requested_date": fixed_now.date().isoformat(),
                "requested_time": "18:30",
            },
        )

    assert duplicate_extra.status_code == 409
    assert duplicate_extra.json()["detail"] == "Ja existe uma solicitacao de transporte ativa para 22/04/2026."

    with SessionLocal() as db:
        user = get_user_by_chave(db, "WT18")
        active_requests = db.execute(
            select(TransportRequest.request_kind, TransportRequest.requested_time)
            .where(
                TransportRequest.user_id == user.id,
                TransportRequest.status == "active",
            )
            .order_by(TransportRequest.id)
        ).all()

    assert active_requests == [("extra", "18:10")]


def test_web_transport_weekend_and_extra_requests_remain_visible_before_their_target_date(monkeypatch):
    fixed_now = datetime(2026, 4, 17, 9, 15, tzinfo=ZoneInfo(settings.tz_name))
    saturday = fixed_now.date() + timedelta(days=1)
    sunday = fixed_now.date() + timedelta(days=2)
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    ensure_web_user_exists(chave="WT24", projeto="P83", nome="Upcoming Weekend Rider")
    ensure_web_transport_address(chave="WT24")
    set_user_checkin_state(chave="WT24", event_time=fixed_now, local="Friday Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WT24",
            senha="abc123",
            projeto="P83",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        weekend_created = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT24", "request_kind": "weekend", "selected_weekdays": [5]},
        )
        assert weekend_created.status_code == 200
        assert weekend_created.json()["state"]["status"] == "pending"

        extra_created = client.post(
            "/api/web/transport/vehicle-request",
            json={
                "chave": "WT24",
                "request_kind": "extra",
                "requested_date": sunday.isoformat(),
                "requested_time": "18:10",
            },
        )
        assert extra_created.status_code == 200
        assert extra_created.json()["state"]["status"] == "pending"

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        dashboard = admin_client.get(
            "/api/transport/dashboard",
            params={"service_date": fixed_now.date().isoformat(), "route_kind": "home_to_work"},
        )

    assert dashboard.status_code == 200
    weekend_row = next(row for row in dashboard.json()["weekend_requests"] if row["chave"] == "WT24")
    extra_row = next(row for row in dashboard.json()["extra_requests"] if row["chave"] == "WT24")

    assert weekend_row["service_date"] == saturday.isoformat()
    assert weekend_row["assignment_status"] == "pending"
    assert extra_row["service_date"] == sunday.isoformat()
    assert extra_row["requested_time"] == "18:10"
    assert extra_row["assignment_status"] == "pending"


def test_web_transport_regular_request_remains_visible_before_the_next_selected_weekday(monkeypatch):
    fixed_now = datetime(2026, 4, 20, 9, 15, tzinfo=ZoneInfo(settings.tz_name))
    tuesday = fixed_now.date() + timedelta(days=1)
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    ensure_web_user_exists(chave="WT33", projeto="P80", nome="Future Regular Rider")
    ensure_web_transport_address(chave="WT33")
    set_user_checkin_state(chave="WT33", event_time=fixed_now, local="Monday Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WT33",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        created = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT33", "request_kind": "regular", "selected_weekdays": [1]},
        )
        assert created.status_code == 200
        assert created.json()["state"]["status"] == "pending"

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        dashboard = admin_client.get(
            "/api/transport/dashboard",
            params={"service_date": fixed_now.date().isoformat(), "route_kind": "home_to_work"},
        )

    assert dashboard.status_code == 200
    regular_rows = [row for row in dashboard.json()["regular_requests"] if row["chave"] == "WT33"]
    assert len(regular_rows) == 1
    assert regular_rows[0]["service_date"] == tuesday.isoformat()


def test_web_transport_state_accumulates_requests_and_cancels_replaced_regular_assignments(monkeypatch):
    clock = {"now": datetime(2026, 4, 20, 7, 30, tzinfo=ZoneInfo(settings.tz_name))}
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: clock["now"])
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: clock["now"])

    with SessionLocal() as db:
        vehicle = Vehicle(placa="HST2501", tipo="van", color="White", lugares=10, tolerance=5, service_scope="regular")
        db.add(vehicle)
        db.flush()
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )
        db.commit()
        vehicle_id = vehicle.id

    ensure_web_user_exists(chave="WT25", projeto="P80", nome="History Transport Rider")
    ensure_web_transport_address(chave="WT25")
    set_user_checkin_state(chave="WT25", event_time=clock["now"], local="History Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WT25",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        first_created = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT25", "request_kind": "regular", "selected_weekdays": [0, 1]},
        )
        assert first_created.status_code == 200
        first_request_id = first_created.json()["state"]["request_id"]

        with TestClient(app) as admin_client:
            ensure_admin_session(admin_client)
            assigned = admin_client.post(
                "/api/transport/assignments",
                json={
                    "request_id": first_request_id,
                    "service_date": clock["now"].date().isoformat(),
                    "route_kind": "home_to_work",
                    "status": "confirmed",
                    "vehicle_id": vehicle_id,
                },
            )
            assert assigned.status_code == 200

        clock["now"] = clock["now"] + timedelta(minutes=5)

        second_created = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT25", "request_kind": "regular", "selected_weekdays": [2, 3]},
        )
        assert second_created.status_code == 200

        state_response = client.get("/api/web/transport/state", params={"chave": "WT25"})
        assert state_response.status_code == 200
        state_payload = state_response.json()

    assert [item["request_id"] for item in state_payload["requests"][:2]] == [
        second_created.json()["state"]["requests"][0]["request_id"],
        first_request_id,
    ]
    assert state_payload["requests"][0]["status"] == "pending"
    assert state_payload["requests"][0]["selected_weekdays"] == [2, 3]
    assert state_payload["requests"][1]["status"] == "cancelled"
    assert state_payload["requests"][1]["is_active"] is False

    with SessionLocal() as db:
        first_request = db.get(TransportRequest, first_request_id)
        assert first_request is not None
        assert first_request.status == "cancelled"

        first_assignments = db.execute(
            select(TransportAssignment)
            .where(TransportAssignment.request_id == first_request_id)
            .order_by(TransportAssignment.route_kind, TransportAssignment.service_date)
        ).scalars().all()

    assert first_assignments
    assert all(row.status == "cancelled" for row in first_assignments)
    assert all(row.vehicle_id is None for row in first_assignments)


def test_transport_dashboard_reject_marks_web_request_as_cancelled(monkeypatch):
    fixed_now = datetime(2026, 4, 21, 8, 10, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    ensure_web_user_exists(chave="WT26", projeto="P82", nome="Rejected Transport Rider")
    ensure_web_transport_address(chave="WT26")
    set_user_checkin_state(chave="WT26", event_time=fixed_now, local="Reject Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WT26",
            senha="abc123",
            projeto="P82",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        created = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT26", "request_kind": "regular"},
        )
        assert created.status_code == 200
        request_id = created.json()["state"]["request_id"]

        with TestClient(app) as admin_client:
            ensure_admin_session(admin_client)
            rejected = admin_client.post(
                "/api/transport/requests/reject",
                json={
                    "request_id": request_id,
                    "service_date": fixed_now.date().isoformat(),
                    "route_kind": "home_to_work",
                },
            )
            assert rejected.status_code == 200

            dashboard = admin_client.get(
                "/api/transport/dashboard",
                params={"service_date": fixed_now.date().isoformat(), "route_kind": "home_to_work"},
            )
            assert dashboard.status_code == 200

        state_response = client.get("/api/web/transport/state", params={"chave": "WT26"})
        assert state_response.status_code == 200
        state_payload = state_response.json()

    assert state_payload["status"] == "available"
    assert state_payload["requests"][0]["request_id"] == request_id
    assert state_payload["requests"][0]["status"] == "cancelled"
    assert state_payload["requests"][0]["is_active"] is False
    assert all(row["chave"] != "WT26" for row in dashboard.json()["regular_requests"])

    with SessionLocal() as db:
        request_row = db.get(TransportRequest, request_id)
        assert request_row is not None
        assert request_row.status == "cancelled"

        assignments = db.execute(
            select(TransportAssignment)
            .where(TransportAssignment.request_id == request_id)
            .order_by(TransportAssignment.route_kind, TransportAssignment.service_date)
        ).scalars().all()

    assert assignments
    assert all(row.status == "rejected" for row in assignments)
    assert all(row.vehicle_id is None for row in assignments)


def test_transport_dashboard_reject_marks_weekend_and_extra_web_requests_as_cancelled(monkeypatch):
    clock = {"now": datetime(2026, 4, 18, 8, 20, tzinfo=ZoneInfo(settings.tz_name))}
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: clock["now"])
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: clock["now"])

    scenarios = [
        {
            "chave": "WW27",
            "nome": "Weekend Reject Rider",
            "projeto": "P82",
            "local": "Weekend Gate",
            "now": datetime(2026, 4, 18, 8, 20, tzinfo=ZoneInfo(settings.tz_name)),
            "create_payload": {"request_kind": "weekend", "selected_weekdays": [5]},
            "dashboard_key": "weekend_requests",
        },
        {
            "chave": "WX27",
            "nome": "Extra Reject Rider",
            "projeto": "P83",
            "local": "Extra Gate",
            "now": datetime(2026, 4, 21, 9, 5, tzinfo=ZoneInfo(settings.tz_name)),
            "create_payload": {
                "request_kind": "extra",
                "requested_date": "2026-04-21",
                "requested_time": "18:10",
            },
            "dashboard_key": "extra_requests",
        },
    ]

    for scenario in scenarios:
        clock["now"] = scenario["now"]
        ensure_web_user_exists(chave=scenario["chave"], projeto=scenario["projeto"], nome=scenario["nome"])
        ensure_web_transport_address(chave=scenario["chave"])
        set_user_checkin_state(chave=scenario["chave"], event_time=clock["now"], local=scenario["local"])

        with TestClient(app) as client:
            registered = register_web_password(
                client,
                chave=scenario["chave"],
                senha="abc123",
                projeto=scenario["projeto"],
                ensure_user_exists=False,
            )
            assert registered.status_code == 200

            created = client.post(
                "/api/web/transport/vehicle-request",
                json={"chave": scenario["chave"], **scenario["create_payload"]},
            )
            assert created.status_code == 200
            request_id = created.json()["state"]["request_id"]

            with TestClient(app) as admin_client:
                ensure_admin_session(admin_client)
                rejected = admin_client.post(
                    "/api/transport/requests/reject",
                    json={
                        "request_id": request_id,
                        "service_date": clock["now"].date().isoformat(),
                        "route_kind": "home_to_work",
                    },
                )
                assert rejected.status_code == 200

                dashboard = admin_client.get(
                    "/api/transport/dashboard",
                    params={"service_date": clock["now"].date().isoformat(), "route_kind": "home_to_work"},
                )
                assert dashboard.status_code == 200

            state_response = client.get("/api/web/transport/state", params={"chave": scenario["chave"]})
            assert state_response.status_code == 200
            state_payload = state_response.json()

        assert state_payload["status"] == "available"
        assert state_payload["requests"][0]["request_id"] == request_id
        assert state_payload["requests"][0]["status"] == "cancelled"
        assert state_payload["requests"][0]["is_active"] is False
        assert all(row["chave"] != scenario["chave"] for row in dashboard.json()[scenario["dashboard_key"]])

        with SessionLocal() as db:
            request_row = db.get(TransportRequest, request_id)

        assert request_row is not None
        assert request_row.status == "cancelled"


def test_transport_dashboard_pending_assignment_returns_request_to_pending_in_dashboard_and_webapp(monkeypatch):
    fixed_now = datetime(2026, 4, 22, 8, 10, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    with SessionLocal() as db:
        vehicle = Vehicle(placa="PEN4222", tipo="carro", color="Blue", lugares=4, tolerance=6, service_scope="regular")
        db.add(vehicle)
        db.flush()
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )
        db.commit()
        vehicle_id = vehicle.id

    ensure_web_user_exists(chave="WT42", projeto="P82", nome="Pending Return Rider")
    ensure_web_transport_address(chave="WT42")
    set_user_checkin_state(chave="WT42", event_time=fixed_now, local="Pending Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WT42",
            senha="abc123",
            projeto="P82",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        created = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT42", "request_kind": "regular"},
        )
        assert created.status_code == 200
        request_id = created.json()["state"]["request_id"]

        with TestClient(app) as admin_client:
            ensure_admin_session(admin_client)
            assigned = admin_client.post(
                "/api/transport/assignments",
                json={
                    "request_id": request_id,
                    "service_date": fixed_now.date().isoformat(),
                    "route_kind": "home_to_work",
                    "status": "confirmed",
                    "vehicle_id": vehicle_id,
                },
            )
            assert assigned.status_code == 200

            returned_to_pending = admin_client.post(
                "/api/transport/assignments",
                json={
                    "request_id": request_id,
                    "service_date": fixed_now.date().isoformat(),
                    "route_kind": "home_to_work",
                    "status": "pending",
                },
            )
            assert returned_to_pending.status_code == 200

            dashboard_home = admin_client.get(
                "/api/transport/dashboard",
                params={"service_date": fixed_now.date().isoformat(), "route_kind": "home_to_work"},
            )
            dashboard_work = admin_client.get(
                "/api/transport/dashboard",
                params={"service_date": fixed_now.date().isoformat(), "route_kind": "work_to_home"},
            )

        state_response = client.get("/api/web/transport/state", params={"chave": "WT42"})
        assert state_response.status_code == 200
        state_payload = state_response.json()

    assert dashboard_home.status_code == 200
    assert dashboard_work.status_code == 200

    home_row = next(row for row in dashboard_home.json()["regular_requests"] if row["id"] == request_id)
    work_row = next(row for row in dashboard_work.json()["regular_requests"] if row["id"] == request_id)
    assert home_row["assignment_status"] == "pending"
    assert work_row["assignment_status"] == "pending"
    assert home_row["assigned_vehicle"] is None
    assert work_row["assigned_vehicle"] is None

    assert state_payload["status"] == "pending"
    assert state_payload["request_id"] == request_id
    assert state_payload["requests"][0]["request_id"] == request_id
    assert state_payload["requests"][0]["status"] == "pending"
    assert state_payload["requests"][0]["is_active"] is True

    with SessionLocal() as db:
        assignments = db.execute(
            select(TransportAssignment)
            .where(TransportAssignment.request_id == request_id)
            .order_by(TransportAssignment.route_kind, TransportAssignment.service_date)
        ).scalars().all()

    assert assignments
    assert all(row.status == "pending" for row in assignments)
    assert all(row.vehicle_id is None for row in assignments)


def test_web_transport_regular_request_stays_visible_on_weekend_dashboard(monkeypatch):
    fixed_now = datetime(2026, 4, 19, 9, 15, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    ensure_web_user_exists(chave="WT48", projeto="P80", nome="Regular Weekend Rider")

    with SessionLocal() as db:
        user = get_user_by_chave(db, "WT48")
        user.end_rua = "40 Weekend Avenue"
        user.zip = "654321"
        db.commit()

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WT48",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        created = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT48", "request_kind": "regular"},
        )
        assert created.status_code == 200
        assert created.json()["state"]["status"] == "pending"
        assert created.json()["state"]["request_kind"] == "regular"

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        dashboard = admin_client.get(
            "/api/transport/dashboard",
            params={"service_date": fixed_now.date().isoformat(), "route_kind": "home_to_work"},
        )

    assert dashboard.status_code == 200
    regular_rows = [row for row in dashboard.json()["regular_requests"] if row["chave"] == "WT48"]
    assert len(regular_rows) == 1
    assert regular_rows[0]["nome"] == "Regular Weekend Rider"


def test_web_transport_address_update_and_acknowledgement_reflect_on_admin_dashboard(monkeypatch):
    fixed_now = datetime(2026, 4, 17, 8, 0, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="16:45")
        db.commit()

    with SessionLocal() as db:
        vehicle = Vehicle(placa="TWA1234", tipo="van", color="Blue", lugares=12, tolerance=10, service_scope="regular")
        db.add(vehicle)
        db.flush()
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )
        db.commit()
        vehicle_id = vehicle.id

    ensure_web_user_exists(chave="WT47", projeto="P82", nome="Aware Transport Rider")
    set_user_checkin_state(chave="WT47", event_time=fixed_now, local="Main Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WT47",
            senha="abc123",
            projeto="P82",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        updated_address = client.post(
            "/api/web/transport/address",
            json={
                "chave": "WT47",
                "end_rua": "Block 3, Harbour Street 55",
                "zip": "654321",
            },
        )
        assert updated_address.status_code == 200
        assert updated_address.json()["state"]["end_rua"] == "Block 3, Harbour Street 55"
        assert updated_address.json()["state"]["zip"] == "654321"

        requested = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT47", "request_kind": "regular"},
        )
        assert requested.status_code == 200
        request_id = requested.json()["state"]["request_id"]

        with TestClient(app) as admin_client:
            ensure_admin_session(admin_client)
            assigned = admin_client.post(
                "/api/transport/assignments",
                json={
                    "request_id": request_id,
                    "service_date": fixed_now.date().isoformat(),
                    "route_kind": "home_to_work",
                    "status": "confirmed",
                    "vehicle_id": vehicle_id,
                },
            )
            assert assigned.status_code == 200

        confirmed_state = client.get("/api/web/transport/state", params={"chave": "WT47"})
        assert confirmed_state.status_code == 200
        assert confirmed_state.json()["status"] == "confirmed"
        assert confirmed_state.json()["awareness_confirmed"] is False
        assert confirmed_state.json()["route_kind"] == "work_to_home"
        assert confirmed_state.json()["boarding_time"] == "16:45"
        assert confirmed_state.json()["vehicle_plate"] == "TWA1234"

        acknowledged = client.post(
            "/api/web/transport/acknowledge",
            json={"chave": "WT47", "request_id": request_id},
        )
        assert acknowledged.status_code == 200
        assert acknowledged.json()["state"]["awareness_confirmed"] is True

    with SessionLocal() as db:
        user = get_user_by_chave(db, "WT47")
        assert user.end_rua == "Block 3, Harbour Street 55"
        assert user.zip == "654321"

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        home_dashboard = admin_client.get(
            "/api/transport/dashboard",
            params={"service_date": fixed_now.date().isoformat(), "route_kind": "home_to_work"},
        )
        paired_dashboard = admin_client.get(
            "/api/transport/dashboard",
            params={"service_date": fixed_now.date().isoformat(), "route_kind": "work_to_home"},
        )

    assert home_dashboard.status_code == 200
    assert paired_dashboard.status_code == 200
    home_row = next(row for row in home_dashboard.json()["regular_requests"] if row["chave"] == "WT47")
    paired_row = next(row for row in paired_dashboard.json()["regular_requests"] if row["chave"] == "WT47")
    assert home_row["awareness_status"] == "aware"
    assert paired_row["awareness_status"] == "aware"


def test_transport_export_endpoint_builds_xlsx_download_and_saves_server_copy(monkeypatch):
    fixed_now = datetime(2026, 4, 22, 15, 26, 45, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    export_dir = Path(settings.transport_exports_dir)
    if export_dir.exists():
        shutil.rmtree(export_dir)

    with SessionLocal() as db:
        vehicle = Vehicle(placa="EXP4521", tipo="van", color="White", lugares=9, tolerance=8, service_scope="regular")
        db.add(vehicle)
        db.flush()
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )

        export_home_user = User(
            nome="Export Home Rider",
            chave="EH01",
            projeto="P80",
            end_rua="10 Export Avenue",
            zip="111111",
            last_active_at=fixed_now,
            inactivity_days=0,
        )
        export_work_user = User(
            nome="Export Work Rider",
            chave="EW02",
            projeto="P81",
            end_rua="20 Return Road",
            zip="222222",
            last_active_at=fixed_now,
            inactivity_days=0,
        )
        export_pending_user = User(
            nome="Pending Export Rider",
            chave="EP03",
            projeto="P82",
            end_rua="30 Waiting Street",
            zip="333333",
            last_active_at=fixed_now,
            inactivity_days=0,
        )
        db.add_all([export_home_user, export_work_user, export_pending_user])
        db.flush()

        home_request, _ = transport_service_module.upsert_transport_request(
            db,
            user=export_home_user,
            request_kind="regular",
            requested_time="07:15",
            requested_date=None,
            created_via="web",
        )
        work_request, _ = transport_service_module.upsert_transport_request(
            db,
            user=export_work_user,
            request_kind="regular",
            requested_time="18:10",
            requested_date=None,
            created_via="web",
        )
        pending_request, _ = transport_service_module.upsert_transport_request(
            db,
            user=export_pending_user,
            request_kind="regular",
            requested_time="08:05",
            requested_date=None,
            created_via="web",
        )
        db.commit()

        home_request_id = home_request.id
        work_request_id = work_request.id
        pending_request_id = pending_request.id
        vehicle_id = vehicle.id

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        home_assignment = admin_client.post(
            "/api/transport/assignments",
            json={
                "request_id": home_request_id,
                "service_date": fixed_now.date().isoformat(),
                "route_kind": "home_to_work",
                "status": "confirmed",
                "vehicle_id": vehicle_id,
            },
        )
        assert home_assignment.status_code == 200

        work_assignment = admin_client.post(
            "/api/transport/assignments",
            json={
                "request_id": work_request_id,
                "service_date": fixed_now.date().isoformat(),
                "route_kind": "work_to_home",
                "status": "confirmed",
                "vehicle_id": vehicle_id,
            },
        )
        assert work_assignment.status_code == 200

        pending_assignment = admin_client.post(
            "/api/transport/assignments",
            json={
                "request_id": pending_request_id,
                "service_date": fixed_now.date().isoformat(),
                "route_kind": "home_to_work",
                "status": "pending",
                "vehicle_id": None,
            },
        )
        assert pending_assignment.status_code == 200

        exported = admin_client.get(
            "/api/transport/exports/transport-list",
            params={"service_date": fixed_now.date().isoformat(), "route_kind": "home_to_work"},
        )

    assert exported.status_code == 200
    assert exported.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert (
        exported.headers["content-disposition"]
        == 'attachment; filename="Transport List - 20260422 - 152645.xlsx"'
    )

    workbook = load_workbook(io.BytesIO(exported.content))
    worksheet = workbook.active
    assert worksheet.title == "Transport List"
    assert [worksheet["A1"].value, worksheet["B1"].value, worksheet["C1"].value, worksheet["D1"].value, worksheet["E1"].value, worksheet["F1"].value] == [
        "Nome/Name",
        "Chave/Key",
        "Projeto/Project",
        "Endereço/Address",
        "Data/Date",
        "Partida/Departure",
    ]
    assert [worksheet["A2"].value, worksheet["B2"].value, worksheet["C2"].value, worksheet["D2"].value, worksheet["E2"].value, worksheet["F2"].value] == [
        "Export Home Rider",
        "EH01",
        "P80",
        "10 Export Avenue",
        "2026-04-22",
        None,
    ]
    assert [worksheet["A3"].value, worksheet["B3"].value, worksheet["C3"].value, worksheet["D3"].value, worksheet["E3"].value, worksheet["F3"].value] == [
        "Export Work Rider",
        "EW02",
        "P81",
        "20 Return Road",
        "2026-04-22",
        None,
    ]
    assert worksheet.max_row == 3
    workbook.close()

    saved_exports = sorted(Path(settings.transport_exports_dir).glob("Transport List - 20260422 - 152645*.xlsx"))
    assert len(saved_exports) == 1
    assert saved_exports[0].is_file()


def test_transport_operational_plan_export_includes_proposal_review_tabs(monkeypatch):
    fixed_now = datetime(2026, 4, 22, 16, 5, 30, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    export_dir = Path(settings.transport_exports_dir)
    if export_dir.exists():
        shutil.rmtree(export_dir)

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="18:10")
        vehicle = Vehicle(placa="EXP4530", tipo="van", color="Blue", lugares=7, tolerance=8, service_scope="regular")
        db.add(vehicle)
        db.flush()

        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )

        rider = User(
            nome="Proposal Export Rider",
            chave="PX30",
            projeto="P83",
            end_rua="30 Proposal Export Way",
            zip="303030",
            last_active_at=fixed_now,
            inactivity_days=0,
        )
        db.add(rider)
        db.flush()

        request_row, _ = transport_service_module.upsert_transport_request(
            db,
            user=rider,
            request_kind="regular",
            requested_time="07:25",
            requested_date=None,
            created_via="web",
        )
        db.commit()

        draft_proposal = transport_proposal_service_module.build_transport_operational_proposal(
            snapshot=transport_proposal_service_module.build_transport_operational_snapshot(
                db,
                service_date=fixed_now.date(),
                route_kind="home_to_work",
                captured_at=fixed_now,
            ),
            origin="manual",
            created_at=fixed_now,
            decisions=[
                TransportProposalDecision(
                    request_id=request_row.id,
                    request_kind="regular",
                    service_date=fixed_now.date(),
                    route_kind="home_to_work",
                    suggested_status="confirmed",
                    vehicle_id=vehicle.id,
                    rationale="Approved for export review.",
                )
            ],
        )

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)

        approved_proposal_response = admin_client.post(
            "/api/transport/proposals/approve",
            json=draft_proposal.model_dump(mode="json"),
        )
        assert approved_proposal_response.status_code == 200, approved_proposal_response.text
        approved_proposal_payload = approved_proposal_response.json()["proposal"]

        exported = admin_client.post(
            "/api/transport/exports/operational-plan",
            json=approved_proposal_payload,
        )

    assert exported.status_code == 200
    assert exported.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert (
        exported.headers["content-disposition"]
        == 'attachment; filename="Transport Operational Plan - 20260422 - 160530.xlsx"'
    )

    workbook = load_workbook(io.BytesIO(exported.content))
    assert workbook.sheetnames == [
        "Transport List",
        "Executive Summary",
        "Vehicle Load",
        "Snapshot Requests",
        "Proposed Decisions",
        "Exceptions",
        "Audit Trail",
    ]

    transport_list_sheet = workbook["Transport List"]
    assert [
        transport_list_sheet["A1"].value,
        transport_list_sheet["B1"].value,
        transport_list_sheet["C1"].value,
        transport_list_sheet["D1"].value,
        transport_list_sheet["E1"].value,
        transport_list_sheet["F1"].value,
    ] == [
        "Nome/Name",
        "Chave/Key",
        "Projeto/Project",
        "Endereço/Address",
        "Data/Date",
        "Partida/Departure",
    ]
    transport_list_rows = list(transport_list_sheet.iter_rows(min_row=2, values_only=True))
    assert all(row[1] != "PX30" for row in transport_list_rows)

    executive_summary_sheet = workbook["Executive Summary"]
    assert [executive_summary_sheet["A1"].value, executive_summary_sheet["B1"].value] == [
        "Campo/Field",
        "Valor/Value",
    ]
    assert [executive_summary_sheet["A2"].value, executive_summary_sheet["B2"].value] == [
        "Modo/Mode",
        "proposal_review",
    ]
    assert [executive_summary_sheet["A13"].value, executive_summary_sheet["B13"].value] == [
        "Proposal Key",
        approved_proposal_payload["proposal_key"],
    ]
    assert [executive_summary_sheet["A14"].value, executive_summary_sheet["B14"].value] == [
        "Proposal Status",
        "approved",
    ]

    vehicle_load_sheet = workbook["Vehicle Load"]
    vehicle_load_rows = {
        row[0]: row
        for row in vehicle_load_sheet.iter_rows(min_row=2, values_only=True)
        if row[0] is not None
    }
    assert [vehicle_load_rows["EXP4530"][0], vehicle_load_rows["EXP4530"][3], vehicle_load_rows["EXP4530"][4]] == [
        "EXP4530",
        0,
        1,
    ]

    snapshot_requests_sheet = workbook["Snapshot Requests"]
    snapshot_request_rows = {
        row[0]: row
        for row in snapshot_requests_sheet.iter_rows(min_row=2, values_only=True)
        if row[0] is not None
    }
    assert [snapshot_request_rows[request_row.id][0], snapshot_request_rows[request_row.id][1], snapshot_request_rows[request_row.id][2]] == [
        request_row.id,
        "regular",
        "pending",
    ]

    proposed_decisions_sheet = workbook["Proposed Decisions"]
    assert [
        proposed_decisions_sheet["A2"].value,
        proposed_decisions_sheet["C2"].value,
        proposed_decisions_sheet["D2"].value,
        proposed_decisions_sheet["E2"].value,
        proposed_decisions_sheet["F2"].value,
    ] == [
        request_row.id,
        "PX30",
        "Proposal Export Rider",
        "confirmed",
        "EXP4530",
    ]

    exceptions_sheet = workbook["Exceptions"]
    exception_rows = list(exceptions_sheet.iter_rows(min_row=2, values_only=True))
    request_exception = next(
        row
        for row in exception_rows
        if row[0] == "snapshot_request" and row[1] == f"request:{request_row.id}"
    )
    assert [request_exception[0], request_exception[2]] == ["snapshot_request", "pending"]

    audit_trail_sheet = workbook["Audit Trail"]
    assert [audit_trail_sheet["A2"].value, audit_trail_sheet["B2"].value] == ["validated", "passed"]
    assert [audit_trail_sheet["A3"].value, audit_trail_sheet["B3"].value] == ["approved", "approved"]
    workbook.close()

    saved_exports = sorted(Path(settings.transport_exports_dir).glob("Transport Operational Plan - 20260422 - 160530*.xlsx"))
    assert len(saved_exports) == 1
    assert saved_exports[0].is_file()


def test_transport_operational_plan_export_includes_ai_suggestion_tabs_for_agent_proposal(monkeypatch):
    from sistema.app.models import TransportAIRun
    from sistema.app.schemas import TransportAgentPlan
    from sistema.app.services.transport_ai_runs import (
        create_transport_ai_suggestion_from_plan,
        ensure_transport_ai_actor_admin_user,
    )

    fixed_now = datetime(2026, 4, 22, 16, 12, 10, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    export_dir = Path(settings.transport_exports_dir)
    if export_dir.exists():
        shutil.rmtree(export_dir)

    with SessionLocal() as db:
        admin_user = ensure_transport_ai_actor_admin_user(
            db,
            chave=ADMIN_LOGIN_CHAVE,
            nome_completo="Transport Export Agent Admin",
            ensured_at=fixed_now,
        )

        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="18:10")
        vehicle = Vehicle(placa="EXP4710", tipo="van", color="White", lugares=7, tolerance=8, service_scope="regular")
        db.add(vehicle)
        db.flush()

        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )

        rider = User(
            nome="Agent Export Rider",
            chave="AX40",
            projeto="P85",
            end_rua="40 Agent Export Way",
            zip="404040",
            last_active_at=fixed_now,
            inactivity_days=0,
        )
        db.add(rider)
        db.flush()

        request_row, _ = transport_service_module.upsert_transport_request(
            db,
            user=rider,
            request_kind="regular",
            requested_time="07:10",
            requested_date=None,
            created_via="web",
        )
        db.commit()
        request_id = request_row.id
        vehicle_id = vehicle.id

        run = TransportAIRun(
            run_key="transport-ai-run:export-agent",
            service_date=fixed_now.date(),
            route_kind="home_to_work",
            status="proposed",
            actor_user_id=admin_user.id,
            earliest_boarding_time="07:00",
            arrival_at_work_time="08:00",
            llm_provider="openai",
            llm_model="gpt-5-2025-08-07",
            llm_reasoning_effort="high",
            openai_model="gpt-5-2025-08-07",
            route_provider="here",
            price_currency_code="SGD",
            price_rate_unit="trip",
            baseline_snapshot_json=None,
            baseline_assignments_json=None,
            baseline_vehicle_state_json=None,
            planning_input_json=json.dumps(
                {
                    "llm_runtime_projects": [
                        {
                            "project_id": 85,
                            "project_name": "P85",
                            "partition_keys": ["partition:export-agent"],
                            "provider": "deepseek",
                            "model_name": "deepseek-v4-pro",
                            "reasoning_effort": "high",
                        }
                    ]
                },
                ensure_ascii=True,
                sort_keys=True,
            ),
            planning_input_hash="0" * 64,
            preflight_issues_json=None,
            error_code=None,
            error_message=None,
            created_at=fixed_now,
            updated_at=fixed_now,
            completed_at=fixed_now,
        )
        db.add(run)
        db.flush()

        agent_plan = TransportAgentPlan.model_validate(
            {
                "plan_key": "transport-ai-plan:export-agent",
                "service_date": fixed_now.date().isoformat(),
                "route_kind": "home_to_work",
                "earliest_boarding_time": "07:00",
                "arrival_at_work_time": "08:00",
                "objective_summary": "Keep the existing van active and preserve the arrival window.",
                "vehicle_actions": [
                    {
                        "action_key": "action:keep-exp4710",
                        "action_type": "keep",
                        "service_scope": "regular",
                        "vehicle_id": vehicle_id,
                        "schedule_id": None,
                        "client_vehicle_key": "existing-exp4710",
                        "before": {"plate": "EXP4710", "assigned_count": 0},
                        "after": {"plate": "EXP4710", "assigned_count": 1},
                        "rationale": "Keep the current van assigned because it already matches the route.",
                        "cost_delta": 0,
                    }
                ],
                "passenger_allocations": [
                    {
                        "request_id": request_id,
                        "request_kind": "regular",
                        "service_date": fixed_now.date().isoformat(),
                        "route_kind": "home_to_work",
                        "vehicle_ref": f"existing:{vehicle_id}",
                        "user_id": rider.id,
                        "chave": "AX40",
                        "nome": "Agent Export Rider",
                        "project_name": "P85",
                        "pickup_order": 1,
                        "scheduled_pickup_time": "07:10",
                        "projected_arrival_time": "07:50",
                        "rationale": "Board the rider on the existing van without changing the plan.",
                    }
                ],
                "route_itineraries": [
                    {
                        "route_key": "route:exp4710",
                        "partition_key": "partition:export-agent",
                        "vehicle_ref": f"existing:{vehicle_id}",
                        "service_scope": "regular",
                        "route_kind": "home_to_work",
                        "vehicle_type": "van",
                        "vehicle_id": vehicle_id,
                        "schedule_id": None,
                        "client_vehicle_key": "existing-exp4710",
                        "plate": "EXP4710",
                        "project_name": "P85",
                        "country_code": "SG",
                        "country_name": "Singapore",
                        "estimated_cost": 18.5,
                        "total_duration_seconds": 2400,
                        "total_distance_meters": 12500,
                        "projected_arrival_time": "07:50",
                        "stops": [
                            {
                                "stop_order": 1,
                                "stop_type": "pickup",
                                "request_id": request_id,
                                "user_id": rider.id,
                                "passenger_name": "Agent Export Rider",
                                "project_name": "P85",
                                "address": "40 Agent Export Way",
                                "zip_code": "404040",
                                "country_code": "SG",
                                "longitude": 103.851959,
                                "latitude": 1.29027,
                                "scheduled_time": "07:10",
                                "duration_from_previous_seconds": 0,
                                "distance_from_previous_meters": 0,
                            },
                            {
                                "stop_order": 2,
                                "stop_type": "destination",
                                "request_id": None,
                                "user_id": None,
                                "passenger_name": None,
                                "project_name": "P85",
                                "address": "Singapore HQ",
                                "zip_code": "049213",
                                "country_code": "SG",
                                "longitude": 103.852,
                                "latitude": 1.286,
                                "scheduled_time": "07:50",
                                "duration_from_previous_seconds": 2400,
                                "distance_from_previous_meters": 12500,
                            },
                        ],
                    }
                ],
                "cost_summary": {
                    "price_currency_code": "SGD",
                    "price_rate_unit": "trip",
                    "current_total_estimated_cost": 18.5,
                    "suggested_total_estimated_cost": 18.5,
                    "estimated_cost_delta": 0,
                    "current_vehicle_count": 1,
                    "suggested_vehicle_count": 1,
                },
                "change_summary": {
                    "total_vehicle_actions": 1,
                    "keep_count": 1,
                    "create_count": 0,
                    "update_count": 0,
                    "remove_from_day_count": 0,
                    "by_vehicle_type": [
                        {
                            "vehicle_type": "van",
                            "keep_count": 1,
                            "create_count": 0,
                            "update_count": 0,
                            "remove_from_day_count": 0,
                            "total_count": 1,
                        }
                    ],
                },
                "validation_issues": [
                    {
                        "code": "soft_overlap_warning",
                        "message": "Pickup window overlaps the arrival buffer by a few minutes.",
                        "blocking": False,
                        "request_id": request_id,
                        "vehicle_id": vehicle_id,
                    }
                ],
            }
        )
        db.commit()

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)

        build_response = admin_client.post(
            "/api/transport/proposals/build",
            json={
                "service_date": fixed_now.date().isoformat(),
                "route_kind": "home_to_work",
                "origin": "agent",
                "captured_at": fixed_now.isoformat(),
                "created_at": fixed_now.isoformat(),
                "decisions": [
                    {
                        "request_id": request_id,
                        "request_kind": "regular",
                        "service_date": fixed_now.date().isoformat(),
                        "route_kind": "home_to_work",
                        "suggested_status": "confirmed",
                        "vehicle_id": vehicle_id,
                        "rationale": "Generated from the persisted AI suggestion.",
                    }
                ],
            },
        )
        assert build_response.status_code == 200, build_response.text

        approved_proposal_response = admin_client.post(
            "/api/transport/proposals/approve",
            json=build_response.json(),
        )
        assert approved_proposal_response.status_code == 200, approved_proposal_response.text
        approved_proposal_payload = approved_proposal_response.json()["proposal"]

    with SessionLocal() as db:
        run = db.execute(select(TransportAIRun).where(TransportAIRun.run_key == "transport-ai-run:export-agent")).scalar_one()
        suggestion = create_transport_ai_suggestion_from_plan(
            db,
            run=run,
            plan=agent_plan,
            prompt_version="transport-ai-export-v1",
            suggestion_key="transport-ai-suggestion:export-agent",
            proposal_key=approved_proposal_payload["proposal_key"],
            status="saved",
            created_at=fixed_now,
        )
        suggestion.transport_proposal_json = json.dumps(approved_proposal_payload, ensure_ascii=True, sort_keys=True)
        suggestion.updated_at = fixed_now
        suggestion.saved_at = fixed_now
        db.commit()

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        exported = admin_client.post(
            "/api/transport/exports/operational-plan",
            json=approved_proposal_payload,
        )

    assert exported.status_code == 200, exported.text
    workbook = load_workbook(io.BytesIO(exported.content))
    assert workbook.sheetnames == [
        "Transport List",
        "Executive Summary",
        "Vehicle Load",
        "Snapshot Requests",
        "Proposed Decisions",
        "Exceptions",
        "Audit Trail",
        "AI Summary",
        "AI Vehicle Actions",
        "AI Itineraries",
        "AI Issues",
    ]

    ai_summary_sheet = workbook["AI Summary"]
    ai_summary_rows = {
        row[0]: row[1]
        for row in ai_summary_sheet.iter_rows(min_row=2, values_only=True)
        if row[0] is not None
    }
    assert ai_summary_rows["Suggestion Key"] == "transport-ai-suggestion:export-agent"
    assert ai_summary_rows["LLM Provider"] == "deepseek"
    assert ai_summary_rows["LLM Model"] == "deepseek-v4-pro"
    assert ai_summary_rows["LLM Reasoning Effort"] == "high"
    assert ai_summary_rows["Objective Summary"] == "Keep the existing van active and preserve the arrival window."
    assert ai_summary_rows["Total Vehicle Actions"] == 1
    assert ai_summary_rows["Route Itineraries"] == 1
    assert ai_summary_rows["Validation Issues"] == 1

    ai_vehicle_actions_sheet = workbook["AI Vehicle Actions"]
    assert [
        ai_vehicle_actions_sheet["A2"].value,
        ai_vehicle_actions_sheet["B2"].value,
        ai_vehicle_actions_sheet["C2"].value,
        ai_vehicle_actions_sheet["H2"].value,
    ] == [
        "action:keep-exp4710",
        "keep",
        "regular",
        "Keep the current van assigned because it already matches the route.",
    ]

    ai_itineraries_sheet = workbook["AI Itineraries"]
    itinerary_rows = list(ai_itineraries_sheet.iter_rows(min_row=2, values_only=True))
    assert any(
        row[0] == "route:exp4710" and row[9] == 1 and row[10] == "pickup" and row[12] == "Agent Export Rider"
        for row in itinerary_rows
    )
    assert any(
        row[0] == "route:exp4710" and row[9] == 2 and row[10] == "destination" and row[13] == "Singapore HQ"
        for row in itinerary_rows
    )

    ai_issues_sheet = workbook["AI Issues"]
    assert [
        ai_issues_sheet["A2"].value,
        ai_issues_sheet["B2"].value,
        ai_issues_sheet["C2"].value,
        ai_issues_sheet["D2"].value,
        ai_issues_sheet["E2"].value,
    ] == [
        "soft_overlap_warning",
        False,
        request_id,
        vehicle_id,
        "Pickup window overlaps the arrival buffer by a few minutes.",
    ]
    workbook.close()


def test_transport_operational_plan_export_supports_contract_built_proposal_audit_trail(monkeypatch):
    fixed_now = datetime(2026, 4, 22, 16, 20, 45, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    export_dir = Path(settings.transport_exports_dir)
    if export_dir.exists():
        shutil.rmtree(export_dir)

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="18:10")
        vehicle = Vehicle(placa="EXP4620", tipo="van", color="Silver", lugares=7, tolerance=8, service_scope="regular")
        db.add(vehicle)
        db.flush()

        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )

        rider = User(
            nome="Contract Export Rider",
            chave="CX20",
            projeto="P84",
            end_rua="20 Contract Export Way",
            zip="202020",
            last_active_at=fixed_now,
            inactivity_days=0,
        )
        db.add(rider)
        db.flush()

        request_row, _ = transport_service_module.upsert_transport_request(
            db,
            user=rider,
            request_kind="regular",
            requested_time="07:40",
            requested_date=None,
            created_via="web",
        )
        db.commit()
        request_id = request_row.id
        vehicle_id = vehicle.id

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)

        build_response = admin_client.post(
            "/api/transport/proposals/build",
            json={
                "service_date": fixed_now.date().isoformat(),
                "route_kind": "home_to_work",
                "origin": "manual",
                "captured_at": fixed_now.isoformat(),
                "created_at": fixed_now.isoformat(),
                "decisions": [
                    {
                        "request_id": request_id,
                        "request_kind": "regular",
                        "service_date": fixed_now.date().isoformat(),
                        "route_kind": "home_to_work",
                        "suggested_status": "confirmed",
                        "vehicle_id": vehicle_id,
                        "rationale": "Built through the contract before export review.",
                    }
                ],
            },
        )
        assert build_response.status_code == 200, build_response.text

        approved_proposal_response = admin_client.post(
            "/api/transport/proposals/approve",
            json=build_response.json(),
        )
        assert approved_proposal_response.status_code == 200, approved_proposal_response.text
        approved_proposal_payload = approved_proposal_response.json()["proposal"]

        exported = admin_client.post(
            "/api/transport/exports/operational-plan",
            json=approved_proposal_payload,
        )

    assert exported.status_code == 200, exported.text
    workbook = load_workbook(io.BytesIO(exported.content))
    audit_trail_sheet = workbook["Audit Trail"]
    assert [audit_trail_sheet["A2"].value, audit_trail_sheet["B2"].value] == ["generated", "generated"]
    assert [audit_trail_sheet["A3"].value, audit_trail_sheet["B3"].value] == ["validated", "passed"]
    assert [audit_trail_sheet["A4"].value, audit_trail_sheet["B4"].value] == ["approved", "approved"]
    workbook.close()


def test_transport_settings_endpoint_updates_work_to_home_boarding_time(monkeypatch):
    fixed_now = datetime(2026, 4, 17, 8, 10, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="16:45")
        vehicle = Vehicle(placa="TWS1810", tipo="carro", color="Silver", lugares=4, tolerance=8, service_scope="regular")
        db.add(vehicle)
        db.flush()
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )
        db.commit()
        vehicle_id = vehicle.id

    ensure_web_user_exists(chave="WT19", projeto="P80", nome="Settings Boarding Rider")
    ensure_web_transport_address(chave="WT19")
    set_user_checkin_state(chave="WT19", event_time=fixed_now, local="Settings Gate")

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        current_settings = admin_client.get("/api/transport/settings")
        assert current_settings.status_code == 200
        assert current_settings.json()["arrive_at_work_time"] == "07:45"
        assert current_settings.json()["work_to_home_time"] == "16:45"
        assert current_settings.json()["last_update_time"] == "16:00"
        assert current_settings.json()["default_car_seats"] == 3
        assert current_settings.json()["default_minivan_seats"] == 6
        assert current_settings.json()["default_van_seats"] == 10
        assert current_settings.json()["default_bus_seats"] == 40
        assert current_settings.json()["default_tolerance_minutes"] == 5
        assert current_settings.json()["extra_car_tolerance_minutes"] == 30
        assert current_settings.json()["price_currency_code"] is None
        assert current_settings.json()["price_rate_unit"] == "day"
        assert current_settings.json()["default_car_price"] is None
        assert current_settings.json()["default_minivan_price"] is None
        assert current_settings.json()["default_van_price"] is None
        assert current_settings.json()["default_bus_price"] is None
        assert current_settings.json()["available_currencies"] == []

        created_currency = admin_client.post(
            "/api/transport/settings/currencies",
            json={
                "code": "sgd",
                "display_label": "Singapore Dollar",
            },
        )
        assert created_currency.status_code == 200
        assert created_currency.json() == {
            "code": "SGD",
            "display_label": "Singapore Dollar",
        }

        updated_settings = admin_client.put(
            "/api/transport/settings",
            json={
                "arrive_at_work_time": "07:25",
                "work_to_home_time": "18:10",
                "last_update_time": "16:20",
                "default_car_seats": 4,
                "default_minivan_seats": 7,
                "default_van_seats": 11,
                "default_bus_seats": 44,
                "default_tolerance_minutes": 9,
                "extra_car_tolerance_minutes": 45,
                "price_currency_code": "SGD",
                "price_rate_unit": "week",
                "default_car_price": 120.5,
                "default_minivan_price": 150.75,
                "default_van_price": 230.0,
                "default_bus_price": 510.25,
            },
        )
        dashboard_after_update = admin_client.get(
            "/api/transport/dashboard",
            params={"service_date": fixed_now.date().isoformat()},
        )
        assert updated_settings.status_code == 200
        assert dashboard_after_update.status_code == 200, dashboard_after_update.text
        assert updated_settings.json()["arrive_at_work_time"] == "07:25"
        assert updated_settings.json()["work_to_home_time"] == "18:10"
        assert updated_settings.json()["last_update_time"] == "16:20"
        assert updated_settings.json()["default_car_seats"] == 4
        assert updated_settings.json()["default_minivan_seats"] == 7
        assert updated_settings.json()["default_van_seats"] == 11
        assert updated_settings.json()["default_bus_seats"] == 44
        assert updated_settings.json()["default_tolerance_minutes"] == 9
        assert updated_settings.json()["extra_car_tolerance_minutes"] == 45
        assert updated_settings.json()["price_currency_code"] == "SGD"
        assert updated_settings.json()["price_rate_unit"] == "week"
        assert updated_settings.json()["default_car_price"] == 120.5
        assert updated_settings.json()["default_minivan_price"] == 150.75
        assert updated_settings.json()["default_van_price"] == 230.0
        assert updated_settings.json()["default_bus_price"] == 510.25
        assert updated_settings.json()["available_currencies"] == [
            {
                "code": "SGD",
                "display_label": "Singapore Dollar",
            }
        ]
        assert dashboard_after_update.json()["arrive_at_work_time"] == "07:25"

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WT19",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        requested = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT19", "request_kind": "regular"},
        )
        assert requested.status_code == 200
        request_id = requested.json()["state"]["request_id"]

        pending_state = client.get("/api/web/transport/state", params={"chave": "WT19"})
        assert pending_state.status_code == 200
        assert pending_state.json()["status"] == "pending"
        assert pending_state.json()["confirmation_deadline_time"] == "16:20"

        with TestClient(app) as admin_client:
            ensure_admin_session(admin_client)
            assigned = admin_client.post(
                "/api/transport/assignments",
                json={
                    "request_id": request_id,
                    "service_date": fixed_now.date().isoformat(),
                    "route_kind": "home_to_work",
                    "status": "confirmed",
                    "vehicle_id": vehicle_id,
                },
            )
            assert assigned.status_code == 200
        confirmed_state = client.get("/api/web/transport/state", params={"chave": "WT19"})
        assert confirmed_state.status_code == 200
        assert confirmed_state.json()["status"] == "confirmed"
        assert confirmed_state.json()["route_kind"] == "work_to_home"
        assert confirmed_state.json()["boarding_time"] == "18:10"
        assert confirmed_state.json()["confirmation_deadline_time"] == "16:20"

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="16:45")
        location_settings_module.upsert_transport_last_update_time(db, last_update_time="16:00")
        db.commit()


def test_transport_settings_endpoint_keeps_ai_secret_on_dedicated_contract(monkeypatch):
    monkeypatch.setattr(settings, "transport_ai_settings_encryption_key", Fernet.generate_key().decode("utf-8"))

    with SessionLocal() as db:
        project = create_transport_planning_project(
            db,
            name="AI SETTINGS CONTRACT PROJECT",
            address="100 Contract Avenue",
            zip_code="018994",
        )
        db.commit()
        project_id = project.id
        project_name = project.name

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)

        ai_settings_response = admin_client.put(
            "/api/transport/ai/settings",
            json={
                "project_id": project_id,
                "provider": "openai",
                "api_key": "sk-separated-1234",
            },
        )
        assert ai_settings_response.status_code == 200, ai_settings_response.text
        assert ai_settings_response.json() == {
            "project_id": project_id,
            "project_name": project_name,
            "provider": "openai",
            "resolved_model": "gpt-5.4-2026-03-05",
            "reasoning_effort": "high",
            "has_api_key": True,
            "api_key_hint": "***1234",
        }

        current_settings = admin_client.get("/api/transport/settings")
        assert current_settings.status_code == 200, current_settings.text
        current_settings_payload = current_settings.json()
        assert "provider" not in current_settings_payload
        assert "api_key" not in current_settings_payload
        assert "api_key_hint" not in current_settings_payload
        assert "resolved_model" not in current_settings_payload
        assert "reasoning_effort" not in current_settings_payload

        mixed_payload = {
            "arrive_at_work_time": current_settings_payload["arrive_at_work_time"],
            "work_to_home_time": current_settings_payload["work_to_home_time"],
            "last_update_time": current_settings_payload["last_update_time"],
            "default_car_seats": current_settings_payload["default_car_seats"],
            "default_minivan_seats": current_settings_payload["default_minivan_seats"],
            "default_van_seats": current_settings_payload["default_van_seats"],
            "default_bus_seats": current_settings_payload["default_bus_seats"],
            "default_tolerance_minutes": current_settings_payload["default_tolerance_minutes"],
            "extra_car_tolerance_minutes": current_settings_payload["extra_car_tolerance_minutes"],
            "price_currency_code": current_settings_payload["price_currency_code"],
            "price_rate_unit": current_settings_payload["price_rate_unit"],
            "default_car_price": current_settings_payload["default_car_price"],
            "default_minivan_price": current_settings_payload["default_minivan_price"],
            "default_van_price": current_settings_payload["default_van_price"],
            "default_bus_price": current_settings_payload["default_bus_price"],
            "provider": "deepseek",
            "api_key": "sk-should-not-be-accepted",
        }
        rejected = admin_client.put("/api/transport/settings", json=mixed_payload)
        assert rejected.status_code == 422, rejected.text
        rejection_fields = {
            tuple(item.get("loc") or ())[-1]
            for item in rejected.json().get("detail", [])
            if item.get("type") == "extra_forbidden"
        }
        assert rejection_fields == {"provider", "api_key"}

        ai_settings_after_rejection = admin_client.get(
            "/api/transport/ai/settings",
            params={"project_id": project_id},
        )
        assert ai_settings_after_rejection.status_code == 200, ai_settings_after_rejection.text
        assert ai_settings_after_rejection.json() == ai_settings_response.json()


def test_transport_settings_currency_endpoint_rejects_duplicate_currency_code():
    currency_code = f"T{uuid.uuid4().hex[:7].upper()}"

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)

        first_response = admin_client.post(
            "/api/transport/settings/currencies",
            json={
                "code": currency_code,
                "display_label": "Singapore Dollar",
            },
        )
        duplicate_response = admin_client.post(
            "/api/transport/settings/currencies",
            json={
                "code": currency_code,
                "display_label": "Singapore Dollar Duplicate",
            },
        )

    assert first_response.status_code == 200, first_response.text
    assert duplicate_response.status_code == 409
    duplicate_detail = extract_transport_structured_detail(duplicate_response)
    assert duplicate_detail["message_key"] == "warnings.currencyAlreadyExists"
    assert duplicate_detail["error_code"] == "transport_currency_code_duplicate"
    assert duplicate_detail["technical_detail"] == "Currency code already exists."

    with SessionLocal() as db:
        matching_rows = db.execute(
            select(TransportCurrencyOption).where(TransportCurrencyOption.code == currency_code)
        ).scalars().all()

    assert len(matching_rows) == 1
    assert matching_rows[0].display_label == "Singapore Dollar"


def test_transport_workplaces_support_operational_context_fields():
    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)

        created = admin_client.post(
            "/api/transport/workplaces",
            json={
                "workplace": "Operations Hub Prime",
                "address": "100 Operations Avenue",
                "zip": "543210",
                "country": "Singapore",
                "transport_group": "Jurong Cluster",
                "boarding_point": "Gate 4 - South Lobby",
                "transport_window_start": "07:10",
                "transport_window_end": "08:05",
                "service_restrictions": "Escort required after 20:00.",
                "transport_work_to_home_time": "17:50",
            },
        )
        assert created.status_code == 200, created.text
        created_payload = created.json()

        updated = admin_client.put(
            f"/api/transport/workplaces/{created_payload['id']}",
            json={
                "address": "101 Operations Avenue",
                "zip": "543211",
                "country": "Singapore",
                "transport_group": "Night Shift Cluster",
                "boarding_point": "Gate 6 - Loading Bay",
                "transport_window_start": "06:45",
                "transport_window_end": "07:30",
                "service_restrictions": "Badge check required before boarding.",
                "transport_work_to_home_time": "18:05",
            },
        )
        assert updated.status_code == 200, updated.text
        updated_payload = updated.json()

        listed = admin_client.get("/api/transport/workplaces")
        assert listed.status_code == 200
        listed_row = next(row for row in listed.json() if row["id"] == created_payload["id"])

    assert created_payload["transport_group"] == "Jurong Cluster"
    assert created_payload["boarding_point"] == "Gate 4 - South Lobby"
    assert created_payload["transport_window_start"] == "07:10"
    assert created_payload["transport_window_end"] == "08:05"
    assert created_payload["service_restrictions"] == "Escort required after 20:00."
    assert created_payload["transport_work_to_home_time"] == "17:50"

    assert updated_payload["address"] == "101 Operations Avenue"
    assert updated_payload["zip"] == "543211"
    assert updated_payload["transport_group"] == "Night Shift Cluster"
    assert updated_payload["boarding_point"] == "Gate 6 - Loading Bay"
    assert updated_payload["transport_window_start"] == "06:45"
    assert updated_payload["transport_window_end"] == "07:30"
    assert updated_payload["service_restrictions"] == "Badge check required before boarding."
    assert updated_payload["transport_work_to_home_time"] == "18:05"

    assert listed_row["workplace"] == "Operations Hub Prime"
    assert listed_row["address"] == "101 Operations Avenue"
    assert listed_row["transport_group"] == "Night Shift Cluster"
    assert listed_row["boarding_point"] == "Gate 6 - Loading Bay"
    assert listed_row["transport_window_start"] == "06:45"
    assert listed_row["transport_window_end"] == "07:30"
    assert listed_row["transport_work_to_home_time"] == "18:05"


def test_transport_work_to_home_time_policy_prefers_date_override_then_workplace_then_global(monkeypatch):
    fixed_now = datetime(2026, 4, 23, 8, 20, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="16:45")
        db.add(
            Workplace(
                workplace="Policy Context Hub",
                address="200 Policy Road",
                zip="654321",
                country="Singapore",
                transport_group="Policy Cluster",
                boarding_point="Gate A",
                transport_window_start="07:00",
                transport_window_end="08:00",
                service_restrictions="Avoid oversized cargo.",
                transport_work_to_home_time="17:40",
            )
        )
        db.commit()

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)

        workplace_policy = admin_client.get(
            "/api/transport/work-to-home-time-policy",
            params={"service_date": fixed_now.date().isoformat(), "workplace": "Policy Context Hub"},
        )
        assert workplace_policy.status_code == 200, workplace_policy.text
        workplace_policy_payload = workplace_policy.json()

        global_policy = admin_client.get(
            "/api/transport/work-to-home-time-policy",
            params={"service_date": fixed_now.date().isoformat()},
        )
        assert global_policy.status_code == 200, global_policy.text
        global_policy_payload = global_policy.json()

        date_override = admin_client.put(
            "/api/transport/date-settings",
            json={
                "service_date": fixed_now.date().isoformat(),
                "work_to_home_time": "18:10",
            },
        )
        assert date_override.status_code == 200, date_override.text

        date_policy = admin_client.get(
            "/api/transport/work-to-home-time-policy",
            params={"service_date": fixed_now.date().isoformat(), "workplace": "Policy Context Hub"},
        )
        assert date_policy.status_code == 200, date_policy.text
        date_policy_payload = date_policy.json()

    assert workplace_policy_payload["resolved_work_to_home_time"] == "17:40"
    assert workplace_policy_payload["source"] == "workplace_context"
    assert workplace_policy_payload["global_work_to_home_time"] == "16:45"
    assert workplace_policy_payload["workplace_work_to_home_time"] == "17:40"
    assert workplace_policy_payload["date_override_work_to_home_time"] is None
    assert workplace_policy_payload["transport_group"] == "Policy Cluster"
    assert workplace_policy_payload["boarding_point"] == "Gate A"
    assert workplace_policy_payload["transport_window_start"] == "07:00"
    assert workplace_policy_payload["transport_window_end"] == "08:00"

    assert global_policy_payload["resolved_work_to_home_time"] == "16:45"
    assert global_policy_payload["source"] == "global"
    assert global_policy_payload["workplace"] is None

    assert date_policy_payload["resolved_work_to_home_time"] == "18:10"
    assert date_policy_payload["source"] == "date_override"
    assert date_policy_payload["global_work_to_home_time"] == "16:45"
    assert date_policy_payload["workplace_work_to_home_time"] == "17:40"
    assert date_policy_payload["date_override_work_to_home_time"] == "18:10"


def test_web_transport_workplace_context_time_applies_before_date_override(monkeypatch):
    fixed_now = datetime(2026, 4, 24, 8, 15, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="16:45")
        db.add(
            Workplace(
                workplace="Context Boarding Hub",
                address="300 Context Road",
                zip="765432",
                country="Singapore",
                transport_group="Context Cluster",
                boarding_point="West Gate",
                transport_window_start="07:15",
                transport_window_end="08:10",
                service_restrictions="Use west gate after check-in.",
                transport_work_to_home_time="17:55",
            )
        )

        vehicle = Vehicle(placa="CTX1755", tipo="van", color="White", lugares=8, tolerance=8, service_scope="regular")
        db.add(vehicle)
        db.flush()
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )
        db.commit()
        vehicle_id = vehicle.id

    with SessionLocal() as db:
        db.add(
            User(
                rfid=None,
                nome="Context Boarding Rider",
                chave="CB55",
                projeto="P80",
                workplace="Context Boarding Hub",
                placa=None,
                end_rua="55 Context Avenue",
                zip="765433",
                local="Context Gate",
                checkin=True,
                time=fixed_now,
                last_active_at=fixed_now,
                inactivity_days=0,
            )
        )
        db.commit()

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="CB55",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        requested = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "CB55", "request_kind": "regular"},
        )
        assert requested.status_code == 200
        request_id = requested.json()["state"]["request_id"]

        with TestClient(app) as admin_client:
            ensure_admin_session(admin_client)
            assigned = admin_client.post(
                "/api/transport/assignments",
                json={
                    "request_id": request_id,
                    "service_date": fixed_now.date().isoformat(),
                    "route_kind": "home_to_work",
                    "status": "confirmed",
                    "vehicle_id": vehicle_id,
                },
            )
            assert assigned.status_code == 200

            initial_policy = admin_client.get(
                "/api/transport/work-to-home-time-policy",
                params={"service_date": fixed_now.date().isoformat(), "workplace": "Context Boarding Hub"},
            )
            assert initial_policy.status_code == 200

        confirmed_state = client.get("/api/web/transport/state", params={"chave": "CB55"})
        assert confirmed_state.status_code == 200
        assert confirmed_state.json()["boarding_time"] == "17:55"

        with TestClient(app) as admin_client:
            ensure_admin_session(admin_client)
            updated_settings = admin_client.put(
                "/api/transport/date-settings",
                json={
                    "service_date": fixed_now.date().isoformat(),
                    "work_to_home_time": "18:20",
                },
            )
            assert updated_settings.status_code == 200

        overridden_state = client.get("/api/web/transport/state", params={"chave": "CB55"})
        assert overridden_state.status_code == 200
        assert overridden_state.json()["boarding_time"] == "18:20"


def test_transport_date_settings_update_work_to_home_departure_for_selected_date_only(monkeypatch):
    friday_now = datetime(2026, 4, 17, 8, 35, tzinfo=ZoneInfo(settings.tz_name))
    saturday = friday_now.date() + timedelta(days=1)
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: friday_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: friday_now)

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="16:45")

        regular_vehicle = Vehicle(
            placa="DTD1710",
            tipo="carro",
            color="Graphite",
            lugares=4,
            tolerance=8,
            service_scope="regular",
        )
        weekend_vehicle = Vehicle(
            placa="DTD1720",
            tipo="van",
            color="Blue",
            lugares=8,
            tolerance=10,
            service_scope="weekend",
        )
        friday_extra_vehicle = Vehicle(
            placa="DTD1730",
            tipo="minivan",
            color="White",
            lugares=6,
            tolerance=12,
            service_scope="extra",
        )
        saturday_extra_vehicle = Vehicle(
            placa="DTD1740",
            tipo="carro",
            color="Black",
            lugares=4,
            tolerance=9,
            service_scope="extra",
        )
        db.add_all([regular_vehicle, weekend_vehicle, friday_extra_vehicle, saturday_extra_vehicle])
        db.flush()

        add_transport_schedule(
            db,
            vehicle=regular_vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=regular_vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=weekend_vehicle,
            service_scope="weekend",
            route_kind="home_to_work",
            recurrence_kind="matching_weekday",
            weekday=5,
        )
        add_transport_schedule(
            db,
            vehicle=weekend_vehicle,
            service_scope="weekend",
            route_kind="work_to_home",
            recurrence_kind="matching_weekday",
            weekday=5,
        )
        add_transport_schedule(
            db,
            vehicle=friday_extra_vehicle,
            service_scope="extra",
            route_kind="work_to_home",
            recurrence_kind="single_date",
            service_date=friday_now.date(),
        )
        add_transport_schedule(
            db,
            vehicle=saturday_extra_vehicle,
            service_scope="extra",
            route_kind="work_to_home",
            recurrence_kind="single_date",
            service_date=saturday,
        )
        db.commit()

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)

        updated_settings = admin_client.put(
            "/api/transport/date-settings",
            json={
                "service_date": friday_now.date().isoformat(),
                "work_to_home_time": "18:10",
            },
        )
        assert updated_settings.status_code == 200
        assert updated_settings.json() == {
            "service_date": friday_now.date().isoformat(),
            "work_to_home_time": "18:10",
        }

        friday_dashboard = admin_client.get(
            "/api/transport/dashboard",
            params={"service_date": friday_now.date().isoformat(), "route_kind": "work_to_home"},
        )
        saturday_dashboard = admin_client.get(
            "/api/transport/dashboard",
            params={"service_date": saturday.isoformat(), "route_kind": "work_to_home"},
        )

    assert friday_dashboard.status_code == 200
    assert saturday_dashboard.status_code == 200

    friday_payload = friday_dashboard.json()
    saturday_payload = saturday_dashboard.json()
    friday_regular_row = next(row for row in friday_payload["regular_vehicles"] if row["placa"] == "DTD1710")
    friday_extra_row = next(row for row in friday_payload["extra_vehicles"] if row["placa"] == "DTD1730")
    saturday_weekend_row = next(row for row in saturday_payload["weekend_vehicles"] if row["placa"] == "DTD1720")
    saturday_extra_row = next(row for row in saturday_payload["extra_vehicles"] if row["placa"] == "DTD1740")

    assert friday_payload["work_to_home_departure_time"] == "18:10"
    assert friday_regular_row["departure_time"] == "18:10"
    assert friday_extra_row["departure_time"] is None

    assert saturday_payload["work_to_home_departure_time"] == "16:45"
    assert saturday_weekend_row["departure_time"] == "16:45"
    assert saturday_extra_row["departure_time"] is None


def test_transport_dashboard_keeps_arrive_at_work_time_global_when_date_override_changes_100_seed_users(monkeypatch):
    fixed_now = datetime(2026, 6, 12, 17, 5, tzinfo=ZoneInfo(settings.tz_name))
    service_date = fixed_now.date()
    seed_keys = [f"{index:04d}" for index in range(1, 101)]
    cleanup_bundle = {
        "request_ids": [],
        "schedule_ids": [],
        "user_ids": [],
        "vehicle_ids": [],
        "settings_context": {},
    }

    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)

    ensure_project_exists("P80")

    with SessionLocal() as db:
        cleanup_bundle["settings_context"] = {"previous": clone_transport_settings_payload(db)}
        location_settings_module.upsert_transport_arrive_at_work_time(db, arrive_at_work_time="07:45")
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="16:45")

        regular_vehicle, schedules = create_transport_planning_vehicle_with_schedules(
            db,
            plate="REG0100",
            service_scope="regular",
            service_date=service_date,
            vehicle_type="van",
            seats=12,
            tolerance=6,
        )
        cleanup_bundle["vehicle_ids"].append(regular_vehicle.id)
        cleanup_bundle["schedule_ids"].extend(schedule.id for schedule in schedules)

        for index, chave in enumerate(seed_keys, start=1):
            requested_hour = 6 + ((index - 1) % 10)
            requested_minute = ((index - 1) % 6) * 10
            user_row, request_row = create_transport_planning_user_with_request(
                db,
                chave=chave,
                nome=f"Transport Seed User {chave}",
                projeto="P80",
                request_kind="regular",
                requested_time=f"{requested_hour:02d}:{requested_minute:02d}",
                home_address=f"{index} Seed Route Avenue",
                home_zip=f"{100000 + index:06d}",
                service_date=service_date,
                timestamp=fixed_now,
            )
            cleanup_bundle["user_ids"].append(user_row.id)
            cleanup_bundle["request_ids"].append(request_row.id)

        db.commit()

    try:
        with TestClient(app) as admin_client:
            ensure_admin_session(admin_client)

            current_settings = admin_client.get("/api/transport/settings")
            dashboard_before = admin_client.get(
                "/api/transport/dashboard",
                params={"service_date": service_date.isoformat(), "route_kind": "work_to_home"},
            )
            override_response = admin_client.put(
                "/api/transport/date-settings",
                json={
                    "service_date": service_date.isoformat(),
                    "work_to_home_time": "18:10",
                },
            )
            dashboard_after = admin_client.get(
                "/api/transport/dashboard",
                params={"service_date": service_date.isoformat(), "route_kind": "work_to_home"},
            )
            settings_after = admin_client.get("/api/transport/settings")

        assert current_settings.status_code == 200, current_settings.text
        assert dashboard_before.status_code == 200, dashboard_before.text
        assert override_response.status_code == 200, override_response.text
        assert dashboard_after.status_code == 200, dashboard_after.text
        assert settings_after.status_code == 200, settings_after.text

        current_settings_payload = current_settings.json()
        dashboard_before_payload = dashboard_before.json()
        dashboard_after_payload = dashboard_after.json()
        settings_after_payload = settings_after.json()

        seeded_requests_before = [
            row for row in dashboard_before_payload["regular_requests"] if row["chave"] in set(seed_keys)
        ]
        seeded_requests_after = [
            row for row in dashboard_after_payload["regular_requests"] if row["chave"] in set(seed_keys)
        ]
        regular_vehicle_before = next(
            row for row in dashboard_before_payload["regular_vehicles"] if row["placa"] == "REG0100"
        )
        regular_vehicle_after = next(
            row for row in dashboard_after_payload["regular_vehicles"] if row["placa"] == "REG0100"
        )
        regular_registry_after = next(
            row for row in dashboard_after_payload["regular_vehicle_registry"] if row["placa"] == "REG0100"
        )

        assert current_settings_payload["arrive_at_work_time"] == "07:45"
        assert current_settings_payload["work_to_home_time"] == "16:45"
        assert settings_after_payload["arrive_at_work_time"] == "07:45"
        assert settings_after_payload["work_to_home_time"] == "16:45"

        assert override_response.json() == {
            "service_date": service_date.isoformat(),
            "work_to_home_time": "18:10",
        }

        assert len(seeded_requests_before) == 100
        assert len(seeded_requests_after) == 100
        assert sorted(row["chave"] for row in seeded_requests_before) == seed_keys
        assert sorted(row["chave"] for row in seeded_requests_after) == seed_keys
        assert {row["assignment_status"] for row in seeded_requests_before} == {"pending"}
        assert {row["assignment_status"] for row in seeded_requests_after} == {"pending"}
        assert all(row["assigned_vehicle"] is None for row in seeded_requests_before)
        assert all(row["assigned_vehicle"] is None for row in seeded_requests_after)

        assert dashboard_before_payload["arrive_at_work_time"] == "07:45"
        assert dashboard_before_payload["work_to_home_departure_time"] == "16:45"
        assert datetime.fromisoformat(dashboard_before_payload["dashboard_generated_at"]).tzinfo is not None
        assert regular_vehicle_before["departure_time"] == "16:45"

        assert dashboard_after_payload["arrive_at_work_time"] == "07:45"
        assert dashboard_after_payload["work_to_home_departure_time"] == "18:10"
        assert datetime.fromisoformat(dashboard_after_payload["dashboard_generated_at"]).tzinfo is not None
        assert regular_vehicle_after["departure_time"] == "18:10"
        assert regular_registry_after["assigned_count"] == 0
    finally:
        with SessionLocal() as db:
            cleanup_transport_planning_fixture_bundle(db, cleanup_bundle)


def test_web_transport_date_override_applies_only_on_selected_day(monkeypatch):
    current_now = {"value": datetime(2026, 4, 17, 8, 45, tzinfo=ZoneInfo(settings.tz_name))}
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: current_now["value"])
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: current_now["value"])

    saturday = current_now["value"].date() + timedelta(days=1)
    regular_plate = f"DOWR{uuid.uuid4().hex[:5].upper()}"
    weekend_plate = f"DOWW{uuid.uuid4().hex[:5].upper()}"

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="16:45")

        regular_vehicle = Vehicle(
            placa=regular_plate,
            tipo="carro",
            color="Silver",
            lugares=4,
            tolerance=8,
            service_scope="regular",
        )
        weekend_vehicle = Vehicle(
            placa=weekend_plate,
            tipo="van",
            color="White",
            lugares=8,
            tolerance=10,
            service_scope="weekend",
        )
        db.add_all([regular_vehicle, weekend_vehicle])
        db.flush()

        add_transport_schedule(
            db,
            vehicle=regular_vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=regular_vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=weekend_vehicle,
            service_scope="weekend",
            route_kind="home_to_work",
            recurrence_kind="matching_weekday",
            weekday=5,
        )
        add_transport_schedule(
            db,
            vehicle=weekend_vehicle,
            service_scope="weekend",
            route_kind="work_to_home",
            recurrence_kind="matching_weekday",
            weekday=5,
        )
        db.commit()
        regular_vehicle_id = regular_vehicle.id
        weekend_vehicle_id = weekend_vehicle.id

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        updated_settings = admin_client.put(
            "/api/transport/date-settings",
            json={
                "service_date": current_now["value"].date().isoformat(),
                "work_to_home_time": "18:10",
            },
        )
        assert updated_settings.status_code == 200

    ensure_web_user_exists(chave="DW18", projeto="P80", nome="Date Override Friday Rider")
    ensure_web_transport_address(chave="DW18")
    set_user_checkin_state(chave="DW18", event_time=current_now["value"], local="Date Override Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="DW18",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        requested = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "DW18", "request_kind": "regular"},
        )
        assert requested.status_code == 200
        friday_request_id = requested.json()["state"]["request_id"]

        with TestClient(app) as admin_client:
            ensure_admin_session(admin_client)
            assigned = admin_client.post(
                "/api/transport/assignments",
                json={
                    "request_id": friday_request_id,
                    "service_date": current_now["value"].date().isoformat(),
                    "route_kind": "home_to_work",
                    "status": "confirmed",
                    "vehicle_id": regular_vehicle_id,
                },
            )
            assert assigned.status_code == 200

        friday_state = client.get("/api/web/transport/state", params={"chave": "DW18"})
        assert friday_state.status_code == 200
        assert friday_state.json()["status"] == "confirmed"
        assert friday_state.json()["route_kind"] == "work_to_home"
        assert friday_state.json()["boarding_time"] == "18:10"

    current_now["value"] = datetime(2026, 4, 18, 8, 50, tzinfo=ZoneInfo(settings.tz_name))
    ensure_web_user_exists(chave="DW16", projeto="P80", nome="Date Override Saturday Rider")
    ensure_web_transport_address(chave="DW16")
    set_user_checkin_state(chave="DW16", event_time=current_now["value"], local="Date Override Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="DW16",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        requested = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "DW16", "request_kind": "weekend"},
        )
        assert requested.status_code == 200
        saturday_request_id = requested.json()["state"]["request_id"]

        with TestClient(app) as admin_client:
            ensure_admin_session(admin_client)
            assigned = admin_client.post(
                "/api/transport/assignments",
                json={
                    "request_id": saturday_request_id,
                    "service_date": saturday.isoformat(),
                    "route_kind": "home_to_work",
                    "status": "confirmed",
                    "vehicle_id": weekend_vehicle_id,
                },
            )
            assert assigned.status_code == 200

        saturday_state = client.get("/api/web/transport/state", params={"chave": "DW16"})
        assert saturday_state.status_code == 200
        assert saturday_state.json()["status"] == "confirmed"
        assert saturday_state.json()["route_kind"] == "work_to_home"
        assert saturday_state.json()["boarding_time"] == "16:45"

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="16:45")
        db.commit()


def test_web_transport_state_marks_departed_confirmed_request_as_realized_and_exposes_vehicle_color(monkeypatch):
    fixed_now = datetime(2026, 4, 17, 19, 30, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="18:10")
        vehicle = Vehicle(placa="REL1730", tipo="carro", color="Silver", lugares=4, tolerance=7, service_scope="regular")
        db.add(vehicle)
        db.flush()
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )
        db.commit()
        vehicle_id = vehicle.id

    ensure_web_user_exists(chave="WR17", projeto="P80", nome="Realized History Rider")
    ensure_web_transport_address(chave="WR17")
    set_user_checkin_state(chave="WR17", event_time=fixed_now, local="Realized Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WR17",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        requested = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WR17", "request_kind": "regular"},
        )
        assert requested.status_code == 200
        request_id = requested.json()["state"]["request_id"]

        with TestClient(app) as admin_client:
            ensure_admin_session(admin_client)
            assigned = admin_client.post(
                "/api/transport/assignments",
                json={
                    "request_id": request_id,
                    "service_date": fixed_now.date().isoformat(),
                    "route_kind": "home_to_work",
                    "status": "confirmed",
                    "vehicle_id": vehicle_id,
                },
            )
            assert assigned.status_code == 200

        state_response = client.get("/api/web/transport/state", params={"chave": "WR17"})
        assert state_response.status_code == 200

    state_payload = state_response.json()
    assert state_payload["status"] == "realized"
    assert state_payload["route_kind"] == "work_to_home"
    assert state_payload["boarding_time"] == "18:10"
    assert state_payload["vehicle_color"] == "Silver"
    assert state_payload["requests"]
    assert state_payload["requests"][0]["request_id"] == request_id
    assert state_payload["requests"][0]["status"] == "realized"
    assert state_payload["requests"][0]["service_date"] == fixed_now.date().isoformat()
    assert state_payload["requests"][0]["vehicle_color"] == "Silver"


def test_web_transport_state_prefers_schedule_scope_over_legacy_vehicle_scope_mirror(monkeypatch):
    fixed_now = datetime(2026, 4, 17, 17, 30, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="18:10")
        vehicle = Vehicle(placa="SCP1730", tipo="carro", color="Silver", lugares=4, tolerance=7, service_scope="extra")
        db.add(vehicle)
        db.flush()
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )
        db.commit()
        vehicle_id = vehicle.id

    ensure_web_user_exists(chave="WS32", projeto="P80", nome="Scope Mirror Rider")
    ensure_web_transport_address(chave="WS32")
    set_user_checkin_state(chave="WS32", event_time=fixed_now, local="Scope Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WS32",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        requested = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WS32", "request_kind": "regular"},
        )
        assert requested.status_code == 200
        request_id = requested.json()["state"]["request_id"]

        with TestClient(app) as admin_client:
            ensure_admin_session(admin_client)
            assigned = admin_client.post(
                "/api/transport/assignments",
                json={
                    "request_id": request_id,
                    "service_date": fixed_now.date().isoformat(),
                    "route_kind": "home_to_work",
                    "status": "confirmed",
                    "vehicle_id": vehicle_id,
                },
            )
            assert assigned.status_code == 200

        state_response = client.get("/api/web/transport/state", params={"chave": "WS32"})
        assert state_response.status_code == 200

    state_payload = state_response.json()
    assert state_payload["status"] == "confirmed"
    assert state_payload["route_kind"] == "work_to_home"
    assert state_payload["boarding_time"] == "18:10"
    assert state_payload["requests"]
    assert state_payload["requests"][0]["boarding_time"] == "18:10"


def test_web_transport_request_history_uses_next_service_date_and_effective_departure_time(monkeypatch):
    current_now = {"value": datetime(2026, 4, 18, 9, 10, tzinfo=ZoneInfo(settings.tz_name))}
    monday = current_now["value"].date() + timedelta(days=2)
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: current_now["value"])
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: current_now["value"])

    with SessionLocal() as db:
        location_settings_module.upsert_transport_work_to_home_time(db, work_to_home_time="16:45")
        location_settings_module.upsert_transport_work_to_home_time_for_date(
            db,
            service_date=monday,
            work_to_home_time="18:10",
        )
        db.commit()

    ensure_web_user_exists(chave="WH18", projeto="P80", nome="History Departure Rider")
    ensure_web_transport_address(chave="WH18")
    set_user_checkin_state(chave="WH18", event_time=current_now["value"], local="Weekend Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WH18",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        created = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WH18", "request_kind": "regular"},
        )
        assert created.status_code == 200

        state_response = client.get("/api/web/transport/state", params={"chave": "WH18"})
        assert state_response.status_code == 200

    state_payload = state_response.json()
    assert state_payload["requests"]
    assert state_payload["requests"][0]["status"] == "pending"
    assert state_payload["requests"][0]["service_date"] == monday.isoformat()
    assert state_payload["requests"][0]["boarding_time"] == "18:10"


def test_web_transport_cancel_pending_request_marks_history_item_cancelled(monkeypatch):
    fixed_now = datetime(2026, 4, 17, 8, 35, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    ensure_web_user_exists(chave="WT31", projeto="P80", nome="Pending Cancel Rider")
    ensure_web_transport_address(chave="WT31")
    set_user_checkin_state(chave="WT31", event_time=fixed_now, local="Pending Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WT31",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        requested = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT31", "request_kind": "regular"},
        )
        assert requested.status_code == 200
        request_id = requested.json()["state"]["request_id"]

        cancelled = client.post(
            "/api/web/transport/cancel",
            json={"chave": "WT31", "request_id": request_id},
        )
        assert cancelled.status_code == 200

    cancelled_payload = cancelled.json()
    assert cancelled_payload["state"]["status"] == "available"
    assert cancelled_payload["state"]["requests"]
    assert cancelled_payload["state"]["requests"][0]["request_id"] == request_id
    assert cancelled_payload["state"]["requests"][0]["status"] == "cancelled"
    assert cancelled_payload["state"]["requests"][0]["is_active"] is False


def test_web_transport_cancel_future_regular_request_removes_row_from_dashboard(monkeypatch):
    fixed_now = datetime(2026, 4, 20, 8, 35, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    ensure_web_user_exists(chave="WT34", projeto="P80", nome="Future Cancel Rider")
    ensure_web_transport_address(chave="WT34")
    set_user_checkin_state(chave="WT34", event_time=fixed_now, local="Monday Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave="WT34",
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        requested = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": "WT34", "request_kind": "regular", "selected_weekdays": [1]},
        )
        assert requested.status_code == 200
        request_id = requested.json()["state"]["request_id"]

        cancelled = client.post(
            "/api/web/transport/cancel",
            json={"chave": "WT34", "request_id": request_id},
        )
        assert cancelled.status_code == 200

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        dashboard = admin_client.get(
            "/api/transport/dashboard",
            params={"service_date": fixed_now.date().isoformat(), "route_kind": "home_to_work"},
        )

    assert dashboard.status_code == 200
    assert all(row["chave"] != "WT34" for row in dashboard.json()["regular_requests"])


def test_transport_dashboard_reject_extra_request_does_not_store_transport_session_user_in_admin_fk(monkeypatch):
    fixed_now = datetime(2026, 4, 21, 7, 45, tzinfo=ZoneInfo(settings.tz_name))
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    with SessionLocal() as db:
        transport_operator = User(
            rfid=None,
            nome="Transport Operator",
            chave="TP41",
            senha=hash_password("tp1234"),
            perfil=2,
            projeto="P82",
            workplace=None,
            placa=None,
            end_rua=None,
            zip=None,
            email=None,
            local=None,
            checkin=None,
            time=None,
            last_active_at=fixed_now,
            inactivity_days=0,
        )
        rider = User(
            rfid=None,
            nome="Extra Reject Rider",
            chave="XR41",
            senha=None,
            perfil=0,
            projeto="P82",
            workplace=None,
            placa=None,
            end_rua=None,
            zip=None,
            email=None,
            local=None,
            checkin=True,
            time=fixed_now,
            last_active_at=fixed_now,
            inactivity_days=0,
        )
        db.add_all([transport_operator, rider])
        db.flush()

        request_row = TransportRequest(
            user_id=rider.id,
            request_kind="extra",
            recurrence_kind="single_date",
            requested_time="07:45",
            selected_weekdays_json=None,
            single_date=fixed_now.date(),
            created_via="web",
            status="active",
            created_at=fixed_now,
            updated_at=fixed_now,
            cancelled_at=None,
        )
        db.add(request_row)
        db.commit()
        request_id = request_row.id

    with TestClient(app) as transport_client:
        login_response = transport_client.post(
            "/api/transport/auth/verify",
            json={"chave": "TP41", "senha": "tp1234"},
        )
        assert login_response.status_code == 200
        assert login_response.json()["authenticated"] is True

        rejected = transport_client.post(
            "/api/transport/requests/reject",
            json={
                "request_id": request_id,
                "service_date": fixed_now.date().isoformat(),
                "route_kind": "home_to_work",
            },
        )

    assert rejected.status_code == 200

    with SessionLocal() as db:
        request_row = db.get(TransportRequest, request_id)
        assignment = db.execute(
            select(TransportAssignment).where(
                TransportAssignment.request_id == request_id,
                TransportAssignment.service_date == fixed_now.date(),
                TransportAssignment.route_kind == "home_to_work",
            )
        ).scalar_one()

    assert request_row is not None
    assert request_row.status == "cancelled"
    assert assignment.status == "rejected"
    assert assignment.assigned_by_admin_id is None


def test_web_transport_cancel_after_confirmation_removes_request_from_vehicle_dashboard(monkeypatch):
    fixed_now = datetime(2026, 4, 17, 8, 20, tzinfo=ZoneInfo(settings.tz_name))
    user_key = f"W{uuid.uuid4().hex[:3].upper()}"
    monkeypatch.setattr(web_check_router, "now_sgt", lambda: fixed_now)
    monkeypatch.setattr(transport_service_module, "now_sgt", lambda: fixed_now)

    with SessionLocal() as db:
        vehicle = Vehicle(placa="TWC1234", tipo="carro", color="White", lugares=4, tolerance=7, service_scope="regular")
        db.add(vehicle)
        db.flush()
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="home_to_work",
            recurrence_kind="weekday",
        )
        add_transport_schedule(
            db,
            vehicle=vehicle,
            service_scope="regular",
            route_kind="work_to_home",
            recurrence_kind="weekday",
        )
        db.commit()
        vehicle_id = vehicle.id

    ensure_web_user_exists(chave=user_key, projeto="P80", nome="Cancel Transport Rider")
    ensure_web_transport_address(chave=user_key)
    set_user_checkin_state(chave=user_key, event_time=fixed_now, local="North Gate")

    with TestClient(app) as client:
        registered = register_web_password(
            client,
            chave=user_key,
            senha="abc123",
            projeto="P80",
            ensure_user_exists=False,
        )
        assert registered.status_code == 200

        requested = client.post(
            "/api/web/transport/vehicle-request",
            json={"chave": user_key, "request_kind": "regular"},
        )
        assert requested.status_code == 200
        request_id = requested.json()["state"]["request_id"]

        with TestClient(app) as admin_client:
            ensure_admin_session(admin_client)
            assigned = admin_client.post(
                "/api/transport/assignments",
                json={
                    "request_id": request_id,
                    "service_date": fixed_now.date().isoformat(),
                    "route_kind": "home_to_work",
                    "status": "confirmed",
                    "vehicle_id": vehicle_id,
                },
            )
            assert assigned.status_code == 200

        cancelled = client.post(
            "/api/web/transport/cancel",
            json={"chave": user_key, "request_id": request_id},
        )
        assert cancelled.status_code == 200
        assert cancelled.json()["state"]["status"] == "available"

    with SessionLocal() as db:
        request_row = db.get(TransportRequest, request_id)
        assert request_row is not None
        assert request_row.status == "cancelled"
        assignments = db.execute(
            select(TransportAssignment)
            .where(TransportAssignment.request_id == request_id)
            .order_by(TransportAssignment.route_kind)
        ).scalars().all()

    assert assignments
    assert all(row.status == "cancelled" for row in assignments)
    assert all(row.vehicle_id is None for row in assignments)

    with TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        dashboard = admin_client.get(
            "/api/transport/dashboard",
            params={"service_date": fixed_now.date().isoformat(), "route_kind": "home_to_work"},
        )

    assert dashboard.status_code == 200
    assert all(row["chave"] != user_key for row in dashboard.json()["regular_requests"])


def test_admin_can_clear_registered_user_password_and_allow_new_registration():
    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "USRSPASS1", "nome": "Usuario Senha", "chave": "PW11", "projeto": "P80"},
        )
        assert save_user.status_code == 200

        users = client.get("/api/admin/users")
        assert users.status_code == 200
        user_id = next(row["id"] for row in users.json() if row["chave"] == "PW11")

        registered = register_web_password(client, chave="PW11", senha="abc123", projeto="P80", ensure_user_exists=False)
        assert registered.status_code == 200

        reset_response = client.post(f"/api/admin/users/{user_id}/reset-password")
        assert reset_response.status_code == 200
        assert "nova senha" in reset_response.json()["message"]

        with SessionLocal() as db:
            user = get_user_by_chave(db, "PW11")
            assert user.senha is None

        status = client.get("/api/web/auth/status", params={"chave": "PW11"})
        assert status.status_code == 200
        assert status.json() == {
            "found": True,
            "chave": "PW11",
            "has_password": False,
            "authenticated": False,
            "message": "Digite sua chave e crie uma senha.",
        }

        old_login = login_web_password(client, chave="PW11", senha="abc123")
        assert old_login.status_code == 404

        new_registration = register_web_password(client, chave="PW11", senha="nova123", projeto="P80", ensure_user_exists=False)
        assert new_registration.status_code == 200
        assert new_registration.json()["authenticated"] is True
        assert new_registration.json()["has_password"] is True

        with SessionLocal() as db:
            user = get_user_by_chave(db, "PW11")
            assert user.senha is not None
            assert verify_password("nova123", user.senha) is True

        events = client.get("/api/admin/events")
        assert events.status_code == 200
        assert any(
            event["action"] == "password"
            and event["request_path"] == f"/api/admin/users/{user_id}/reset-password"
            and event["status"] == "removed"
            for event in events.json()
        )


def test_web_password_status_returns_not_found_for_unknown_key():
    with TestClient(app) as client:
        response = client.get("/api/web/auth/status", params={"chave": "ZZ99"})

        assert response.status_code == 200
        assert response.json() == {
            "found": False,
            "chave": "ZZ99",
            "has_password": False,
            "authenticated": False,
            "message": "Digite sua chave e crie uma senha.",
        }


def test_web_password_login_accepts_partial_attempts_without_validation_error():
    with TestClient(app) as client:
        registered = register_web_password(client, chave="WB19", senha="abc123", projeto="P80")
        assert registered.status_code == 200

        logout = client.post("/api/web/auth/logout")
        assert logout.status_code == 200

        partial_attempt = client.post(
            "/api/web/auth/login",
            json={
                "chave": "WB19",
                "senha": "a",
            },
        )

        assert partial_attempt.status_code == 401
        assert partial_attempt.json()["detail"] == "Chave ou senha invalida"


def test_web_location_match_returns_known_location_when_accuracy_is_good():
    latitude = 1.365936
    longitude = 103.811066

    with TestClient(app) as client:
        ensure_admin_session(client)
        auth_response = register_web_password(client, chave="WL80", senha="loc123", projeto="P80")
        assert auth_response.status_code == 200

        create_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Web Match P80",
                "coordinates": build_rectangle_coordinates(latitude, longitude),
                "projects": ["P80"],
                "tolerance_meters": 150,
            },
        )
        assert create_location.status_code == 200

        update_settings = client.post(
            "/api/admin/locations/settings",
            json={"location_accuracy_threshold_meters": 25},
        )
        assert update_settings.status_code == 200

        match_response = client.post(
            "/api/web/check/location",
            json={
                "latitude": latitude,
                "longitude": longitude,
                "accuracy_meters": 8,
            },
        )

        assert match_response.status_code == 200
        payload = match_response.json()
        assert payload["matched"] is True
        assert payload["resolved_local"] == "Web Match P80"
        assert payload["label"] == "Web Match P80"
        assert payload["status"] == "matched"
        assert payload["accuracy_threshold_meters"] == 25


def test_web_location_endpoints_use_union_of_authenticated_user_projects():
    with TestClient(app) as client:
        ensure_project_exists("P90")
        ensure_project_exists("P91")
        ensure_admin_session(client)
        auth_response = register_web_password(client, chave="WL84", senha="loc123", projeto="P90")
        assert auth_response.status_code == 200

        with SessionLocal() as db:
            user = get_user_by_chave(db, "WL84")
            grant_user_project_memberships(db, user, ["P90", "P91"])
            db.commit()

        create_p80_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Projeto P80",
                "coordinates": build_rectangle_coordinates(1.255936, 103.611066),
                "projects": ["P90"],
                "tolerance_meters": 150,
            },
        )
        assert create_p80_location.status_code == 200

        create_p82_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Projeto P82",
                "coordinates": build_rectangle_coordinates(1.355936, 103.711066),
                "projects": ["P91"],
                "tolerance_meters": 150,
            },
        )
        assert create_p82_location.status_code == 200

        locations_response = client.get("/api/web/check/locations")
        assert locations_response.status_code == 200
        assert "Projeto P80" in locations_response.json()["items"]
        assert "Projeto P82" in locations_response.json()["items"]

        p80_match = client.post(
            "/api/web/check/location",
            json={
                "latitude": 1.255936,
                "longitude": 103.611066,
                "accuracy_meters": 8,
            },
        )
        assert p80_match.status_code == 200
        assert p80_match.json()["matched"] is True
        assert p80_match.json()["resolved_local"] == "Projeto P80"

        p82_match = client.post(
            "/api/web/check/location",
            json={
                "latitude": 1.355936,
                "longitude": 103.711066,
                "accuracy_meters": 8,
            },
        )
        assert p82_match.status_code == 200
        assert p82_match.json()["matched"] is True
        assert p82_match.json()["resolved_local"] == "Projeto P82"


def test_web_location_match_blocks_low_accuracy_before_matching():
    with TestClient(app) as client:
        ensure_admin_session(client)
        auth_response = register_web_password(client, chave="WL81", senha="loc123", projeto="P80")
        assert auth_response.status_code == 200

        create_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Web Accuracy P80",
                "coordinates": build_rectangle_coordinates(1.300001, 103.800001),
                "projects": ["P80"],
                "tolerance_meters": 120,
            },
        )
        assert create_location.status_code == 200

        update_settings = client.post(
            "/api/admin/locations/settings",
            json={"location_accuracy_threshold_meters": 15},
        )
        assert update_settings.status_code == 200

        match_response = client.post(
            "/api/web/check/location",
            json={
                "latitude": 1.300001,
                "longitude": 103.800001,
                "accuracy_meters": 44,
            },
        )

        assert match_response.status_code == 200
        payload = match_response.json()
        assert payload["matched"] is False
        assert payload["resolved_local"] is None
        assert payload["label"] == "Precisao insuficiente"
        assert payload["status"] == "accuracy_too_low"
        assert payload["accuracy_threshold_meters"] == 15


def test_web_location_match_returns_unregistered_location_without_message_within_two_km():
    with TestClient(app) as client:
        ensure_admin_session(client)
        auth_response = register_web_password(client, chave="WL82", senha="loc123", projeto="P80")
        assert auth_response.status_code == 200

        create_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Web Nearby P80",
                "coordinates": build_rectangle_coordinates(1.255936, 103.611066),
                "projects": ["P80"],
                "tolerance_meters": 120,
            },
        )
        assert create_location.status_code == 200

        update_settings = client.post(
            "/api/admin/locations/settings",
            json={"location_accuracy_threshold_meters": 25},
        )
        assert update_settings.status_code == 200

        match_response = client.post(
            "/api/web/check/location",
            json={
                "latitude": 1.260936,
                "longitude": 103.611066,
                "accuracy_meters": 8,
            },
        )

        assert match_response.status_code == 200
        payload = match_response.json()
        assert payload["matched"] is False
        assert payload["resolved_local"] is None
        assert payload["label"] == "Localização não Cadastrada"
        assert payload["status"] == "not_in_known_location"
        assert payload["message"] == ""
        assert payload["nearest_workplace_distance_meters"] < 2000


def test_web_location_match_returns_outside_workplace_without_message():
    with TestClient(app) as client:
        ensure_admin_session(client)
        auth_response = register_web_password(client, chave="WL83", senha="loc123", projeto="P80")
        assert auth_response.status_code == 200

        create_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Web Far P80",
                "coordinates": build_rectangle_coordinates(1.255936, 103.611066),
                "projects": ["P80"],
                "tolerance_meters": 120,
            },
        )
        assert create_location.status_code == 200

        update_settings = client.post(
            "/api/admin/locations/settings",
            json={"location_accuracy_threshold_meters": 25},
        )
        assert update_settings.status_code == 200

        match_response = client.post(
            "/api/web/check/location",
            json={
                "latitude": 1.285936,
                "longitude": 103.611066,
                "accuracy_meters": 8,
            },
        )

        assert match_response.status_code == 200
        payload = match_response.json()
        assert payload["matched"] is False
        assert payload["resolved_local"] is None
        assert payload["label"] == "Fora do Ambiente de Trabalho"
        assert payload["status"] == "outside_workplace"
        assert payload["message"] == ""
        assert payload["minimum_checkout_distance_meters"] == 2000
        assert payload["nearest_workplace_distance_meters"] > 2000


def test_web_location_match_ignores_checkout_zone_for_outside_workplace_threshold():
    with TestClient(app) as client:
        ensure_admin_session(client)
        ensure_project_exists("P80Z")
        auth_response = register_web_password(client, chave="WL8Z", senha="loc123", projeto="P80Z")
        assert auth_response.status_code == 200

        create_workplace = client.post(
            "/api/admin/locations",
            json={
                "local": "Workplace Far P80Z",
                "coordinates": build_rectangle_coordinates(1.285936, 103.611066),
                "projects": ["P80Z"],
                "tolerance_meters": 120,
            },
        )
        assert create_workplace.status_code == 200

        create_checkout = client.post(
            "/api/admin/locations",
            json={
                "local": "Zona de CheckOut",
                "coordinates": build_rectangle_coordinates(1.255936, 103.611066),
                "projects": ["P80Z"],
                "tolerance_meters": 20,
            },
        )
        assert create_checkout.status_code == 200

        update_settings = client.post(
            "/api/admin/locations/settings",
            json={"location_accuracy_threshold_meters": 25},
        )
        assert update_settings.status_code == 200

        match_response = client.post(
            "/api/web/check/location",
            json={
                "latitude": 1.257936,
                "longitude": 103.611066,
                "accuracy_meters": 8,
            },
        )

        assert match_response.status_code == 200
        payload = match_response.json()
        assert payload["matched"] is False
        assert payload["resolved_local"] is None
        assert payload["label"] == "Fora do Ambiente de Trabalho"
        assert payload["status"] == "outside_workplace"
        assert payload["minimum_checkout_distance_meters"] == 2000
        assert payload["nearest_workplace_distance_meters"] > 2000


def test_web_location_match_uses_project_specific_outside_workplace_threshold():
    with TestClient(app) as client:
        ensure_admin_session(client)
        auth_response = register_web_password(client, chave="WL85", senha="loc123", projeto="P82")
        assert auth_response.status_code == 200

        create_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Web Far P82",
                "coordinates": build_rectangle_coordinates(1.255936, 103.611066),
                "projects": ["P82"],
                "tolerance_meters": 120,
            },
        )
        assert create_location.status_code == 200

        update_settings = client.post(
            "/api/admin/locations/settings",
            json={"location_accuracy_threshold_meters": 25},
        )
        assert update_settings.status_code == 200

        update_checkout_threshold = client.post(
            "/api/admin/locations/auto-checkout-distances",
            json={
                "items": [
                    {
                        "project_name": "p82",
                        "minimum_checkout_distance_meters": 4000,
                    }
                ]
            },
        )
        assert update_checkout_threshold.status_code == 200

        match_response = client.post(
            "/api/web/check/location",
            json={
                "latitude": 1.285936,
                "longitude": 103.611066,
                "accuracy_meters": 8,
            },
        )

        assert match_response.status_code == 200
        payload = match_response.json()
        assert payload["matched"] is False
        assert payload["resolved_local"] is None
        assert payload["label"] == "Localização não Cadastrada"
        assert payload["status"] == "not_in_known_location"
        assert payload["minimum_checkout_distance_meters"] == 4000
        assert 2000 < payload["nearest_workplace_distance_meters"] < 4000


def test_web_location_match_uses_polygon_matching_when_location_has_valid_polygon():
    with TestClient(app) as client:
        ensure_admin_session(client)
        ensure_project_exists("P92")
        auth_response = register_web_password(client, chave="WP80", senha="poly123", projeto="P92")
        assert auth_response.status_code == 200

        create_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Area Poligonal P80",
                "coordinates": [
                    {"latitude": 1.255800, "longitude": 103.611000},
                    {"latitude": 1.256100, "longitude": 103.611000},
                    {"latitude": 1.256100, "longitude": 103.611400},
                    {"latitude": 1.255800, "longitude": 103.611400},
                ],
                "projects": ["P92"],
                "tolerance_meters": 45,
            },
        )
        assert create_location.status_code == 200

        update_settings = client.post(
            "/api/admin/locations/settings",
            json={"location_accuracy_threshold_meters": 25},
        )
        assert update_settings.status_code == 200

        inside_match_response = client.post(
            "/api/web/check/location",
            json={
                "latitude": 1.255950,
                "longitude": 103.611200,
                "accuracy_meters": 8,
            },
        )

        assert inside_match_response.status_code == 200
        inside_payload = inside_match_response.json()
        assert inside_payload["matched"] is True
        assert inside_payload["resolved_local"] == "Area Poligonal P80"
        assert inside_payload["label"] == "Area Poligonal P80"
        assert inside_payload["status"] == "matched"
        assert inside_payload["accuracy_threshold_meters"] == 25

        match_response = client.post(
            "/api/web/check/location",
            json={
                "latitude": 1.255760,
                "longitude": 103.610990,
                "accuracy_meters": 12,
            },
        )

        assert match_response.status_code == 200
        payload = match_response.json()
        assert payload["matched"] is True
        assert payload["resolved_local"] == "Area Poligonal P80"
        assert payload["label"] == "Area Poligonal P80"
        assert payload["status"] == "matched"
        assert payload["accuracy_threshold_meters"] == 25


def test_web_location_match_polygon_tie_breaks_by_nearest_vertex():
    with TestClient(app) as client:
        ensure_admin_session(client)
        ensure_project_exists("P93")
        auth_response = register_web_password(client, chave="WP81", senha="poly123", projeto="P93")
        assert auth_response.status_code == 200

        create_nearer = client.post(
            "/api/admin/locations",
            json={
                "local": "Area Mais Perto",
                "coordinates": [
                    {"latitude": 1.255800, "longitude": 103.611000},
                    {"latitude": 1.256100, "longitude": 103.611000},
                    {"latitude": 1.256100, "longitude": 103.611300},
                    {"latitude": 1.255800, "longitude": 103.611300},
                ],
                "projects": ["P93"],
                "tolerance_meters": 120,
            },
        )
        assert create_nearer.status_code == 200

        create_farther = client.post(
            "/api/admin/locations",
            json={
                "local": "Area Mais Longe",
                "coordinates": [
                    {"latitude": 1.255800, "longitude": 103.611350},
                    {"latitude": 1.256100, "longitude": 103.611350},
                    {"latitude": 1.256100, "longitude": 103.611650},
                    {"latitude": 1.255800, "longitude": 103.611650},
                ],
                "projects": ["P93"],
                "tolerance_meters": 120,
            },
        )
        assert create_farther.status_code == 200

        update_settings = client.post(
            "/api/admin/locations/settings",
            json={"location_accuracy_threshold_meters": 80},
        )
        assert update_settings.status_code == 200

        match_response = client.post(
            "/api/web/check/location",
            json={
                "latitude": 1.255790,
                "longitude": 103.611010,
                "accuracy_meters": 70,
            },
        )

        assert match_response.status_code == 200
        payload = match_response.json()
        assert payload["matched"] is True
        assert payload["resolved_local"] == "Area Mais Perto"
        assert payload["label"] == "Area Mais Perto"


def test_web_location_match_prefers_checkout_zone_and_logs_decision_details(caplog):
    with TestClient(app) as client:
        ensure_admin_session(client)
        ensure_project_exists("P93A")
        auth_response = register_web_password(client, chave="W93A", senha="poly123", projeto="P93A")
        assert auth_response.status_code == 200

        create_regular = client.post(
            "/api/admin/locations",
            json={
                "local": "Area Regular P93A",
                "coordinates": build_rectangle_coordinates(1.265936, 103.621066),
                "projects": ["P93A"],
                "tolerance_meters": 90,
            },
        )
        assert create_regular.status_code == 200

        create_checkout = client.post(
            "/api/admin/locations",
            json={
                "local": "Zona de CheckOut",
                "coordinates": build_rectangle_coordinates(1.265956, 103.621086),
                "projects": ["P93A"],
                "tolerance_meters": 90,
            },
        )
        assert create_checkout.status_code == 200

        update_settings = client.post(
            "/api/admin/locations/settings",
            json={"location_accuracy_threshold_meters": 25},
        )
        assert update_settings.status_code == 200

        with caplog.at_level(logging.DEBUG, logger="sistema.app.services.location_matching"):
            match_response = client.post(
                "/api/web/check/location",
                json={
                    "latitude": 1.266010,
                    "longitude": 103.621120,
                    "accuracy_meters": 8,
                },
            )

        assert match_response.status_code == 200
        payload = match_response.json()
        assert payload["matched"] is True
        assert payload["resolved_local"] == "Zona de CheckOut"
        assert payload["label"] == "Zona de Check-Out"
        assert payload["status"] == "matched"
        assert "location_match_decision" in caplog.text
        assert "selection_source=polygon_checkout" in caplog.text
        assert "matched_local=Zona de CheckOut" in caplog.text


def test_web_location_match_ignores_legacy_single_coordinate_locations():
    with TestClient(app) as client:
        ensure_admin_session(client)
        ensure_project_exists("P94")
        auth_response = register_web_password(client, chave="WP82", senha="poly123", projeto="P94")
        assert auth_response.status_code == 200

    with SessionLocal() as db:
        timestamp = now_sgt()
        db.add(
            ManagedLocation(
                local="Legado P80",
                latitude=1.255936,
                longitude=103.611066,
                coordinates_json=None,
                projects_json=dump_location_projects(["P94"]),
                tolerance_meters=150,
                created_at=timestamp,
                updated_at=timestamp,
            )
        )
        db.commit()

    with TestClient(app) as client:
        ensure_admin_session(client)
        update_settings = client.post(
            "/api/admin/locations/settings",
            json={"location_accuracy_threshold_meters": 25},
        )
        assert update_settings.status_code == 200

        login_response = login_web_password(client, chave="WP82", senha="poly123")
        assert login_response.status_code == 200

        match_response = client.post(
            "/api/web/check/location",
            json={
                "latitude": 1.255936,
                "longitude": 103.611066,
                "accuracy_meters": 8,
            },
        )

        assert match_response.status_code == 200
        payload = match_response.json()
    assert payload["matched"] is False
    assert payload["resolved_local"] is None
    assert payload["label"] == "Localização não Cadastrada"
    assert payload["status"] == "not_in_known_location"
    assert payload["nearest_workplace_distance_meters"] is None


def test_web_check_updates_user_local_when_location_is_provided():
    client_event_id = f"web-check-local-{uuid.uuid4().hex}"

    with TestClient(app) as client:
        auth_response = register_web_password(client, chave="WB14", senha="local1", projeto="P80")
        assert auth_response.status_code == 200

        response = client.post(
            "/api/web/check",
            json={
                "chave": "WB14",
                "projeto": "P80",
                "action": "checkin",
                "local": "Web Match P80",
                "informe": "normal",
                "event_time": now_sgt().isoformat(),
                "client_event_id": client_event_id,
            },
        )
        history = client.get("/api/web/check/state", params={"chave": "WB14"})

        assert response.status_code == 200
        assert response.json()["ok"] is True
        assert history.status_code == 200
        assert history.json() == {
            "found": True,
            "chave": "WB14",
            "projeto": "P80",
            "current_action": "checkin",
            "current_local": "Web Match P80",
            "has_current_day_checkin": True,
            "last_checkin_at": response.json()["state"]["last_checkin_at"],
            "last_checkout_at": None,
        }

        with SessionLocal() as db:
            user = get_user_by_chave(db, "WB14")
            assert user.local == "Web Match P80"


def test_web_check_rejects_unregistered_location_placeholder_for_web_submit():
    client_event_id = f"web-check-local-unregistered-{uuid.uuid4().hex}"

    with TestClient(app) as client:
        auth_response = register_web_password(client, chave="WB17", senha="local3", projeto="P80")
        assert auth_response.status_code == 200

        response = client.post(
            "/api/web/check",
            json={
                "chave": "WB17",
                "projeto": "P80",
                "action": "checkin",
                "local": "Localização não Cadastrada",
                "informe": "normal",
                "event_time": now_sgt().isoformat(),
                "client_event_id": client_event_id,
            },
        )

        assert response.status_code == 422
        assert response.json()["detail"] == (
            "O estado 'Localização não Cadastrada' nao e um local operacional valido para submit pela Web."
        )

    with SessionLocal() as db:
        user = get_user_by_chave(db, "WB17")
        assert user.local is None
        queued = db.execute(
            select(FormsSubmission).where(FormsSubmission.chave == "WB17")
        ).scalars().all()
        assert queued == []
        sync_events = db.execute(
            select(UserSyncEvent).where(
                UserSyncEvent.chave == "WB17",
                UserSyncEvent.source == "web_forms",
            )
        ).scalars().all()
        assert sync_events == []


def test_web_check_accepts_synthetic_accuracy_fallback_local():
    client_event_id = f"web-check-local-fallback-{uuid.uuid4().hex}"

    with TestClient(app) as client:
        auth_response = register_web_password(client, chave="WB16", senha="local2", projeto="P80")
        assert auth_response.status_code == 200

        response = client.post(
            "/api/web/check",
            json={
                "chave": "WB16",
                "projeto": "P80",
                "action": "checkin",
                "local": "Precisao Insuficiente",
                "informe": "normal",
                "event_time": now_sgt().isoformat(),
                "client_event_id": client_event_id,
            },
        )
        history = client.get("/api/web/check/state", params={"chave": "WB16"})

        assert response.status_code == 200
        assert response.json()["ok"] is True
        assert history.status_code == 200
        assert history.json() == {
            "found": True,
            "chave": "WB16",
            "projeto": "P80",
            "current_action": "checkin",
            "current_local": "Precisao Insuficiente",
            "has_current_day_checkin": True,
            "last_checkin_at": response.json()["state"]["last_checkin_at"],
            "last_checkout_at": None,
        }

        with SessionLocal() as db:
            user = get_user_by_chave(db, "WB16")
            assert user.local == "Precisao Insuficiente"


def test_web_check_reuses_flutter_like_hidden_project_for_checkout():
    first_event_time = now_sgt()
    second_event_time = first_event_time + timedelta(minutes=4)

    with TestClient(app) as client:
        auth_response = register_web_password(client, chave="WB12", senha="check1", projeto="P83")
        assert auth_response.status_code == 200

        first = client.post(
            "/api/web/check",
            json={
                "chave": "WB12",
                "projeto": "P83",
                "action": "checkout",
                "informe": "retroativo",
                "event_time": first_event_time.isoformat(),
                "client_event_id": f"web-check-1-{uuid.uuid4().hex}",
            },
        )
        second = client.post(
            "/api/web/check",
            json={
                "chave": "WB12",
                "projeto": "P83",
                "action": "checkout",
                "informe": "retroativo",
                "event_time": second_event_time.isoformat(),
                "client_event_id": f"web-check-2-{uuid.uuid4().hex}",
            },
        )

        assert first.status_code == 200
        assert first.json()["queued_forms"] is True
        assert second.status_code == 200
        assert second.json()["queued_forms"] is False

        with SessionLocal() as db:
            user = get_user_by_chave(db, "WB12")
            queued = db.execute(select(FormsSubmission).where(FormsSubmission.chave == "WB12")).scalars().all()
            sync_events = db.execute(
                select(UserSyncEvent).where(
                    UserSyncEvent.source == "web_forms",
                    UserSyncEvent.chave == "WB12",
                )
            ).scalars().all()
            request_events = db.execute(
                select(CheckEvent).where(CheckEvent.request_path == "/api/web/check", CheckEvent.rfid.is_(None))
            ).scalars().all()

            assert user.nome == "Oriundo da Web"
            assert user.projeto == "P83"
            assert user.checkin is False
            assert len(queued) == 2
            ordered_queued = sorted(queued, key=lambda submission: submission.id)
            assert ordered_queued[0].status == "pending"
            assert ordered_queued[-1].status == "skipped"
            assert ordered_queued[-1].display_status == "not_realized"
            assert ordered_queued[-1].last_error == "repeated_checkout"
            assert len(sync_events) == 2
            assert any(event.ontime is False and event.action == "checkout" for event in sync_events)
            assert any(event.action == "checkout" and event.status == "queued" for event in request_events)


def test_web_checkout_after_checkin_processes_forms_and_admin_checkout_shows_sent(monkeypatch):
    monkeypatch.setattr(
        FormsWorker,
        "submit_with_retries",
        mocked_successful_forms_submit_with_statuses,
    )

    clear_forms_queue_backlog()
    web_key = make_test_key("W")
    checkin_request_id = f"web-checkin-{uuid.uuid4().hex}"
    checkout_request_id = f"web-checkout-{uuid.uuid4().hex}"
    reference_time = now_sgt().replace(second=0, microsecond=0)
    checkin_at = reference_time - timedelta(hours=1)
    checkout_at = reference_time - timedelta(minutes=5)

    with TestClient(app) as web_client, TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        auth_response = register_web_password(web_client, chave=web_key, senha="web123", projeto="P80")
        assert auth_response.status_code == 200, auth_response.text

        checkin_response = submit_web_check(
            web_client,
            chave=web_key,
            projeto="P80",
            action="checkin",
            event_time=checkin_at,
            client_event_id=checkin_request_id,
            local="Portaria P80",
        )
        assert checkin_response.status_code == 200, checkin_response.text
        assert checkin_response.json()["queued_forms"] is True

        checkout_response = submit_web_check(
            web_client,
            chave=web_key,
            projeto="P80",
            action="checkout",
            event_time=checkout_at,
            client_event_id=checkout_request_id,
            local="Saida P80",
        )
        assert checkout_response.status_code == 200, checkout_response.text
        assert checkout_response.json()["queued_forms"] is True
        assert checkout_response.json()["state"]["current_action"] == "checkout"

        processed = process_forms_submission_queue_once(max_items=2)
        assert processed == 2

        checkout_row = get_admin_presence_row_by_chave(admin_client, endpoint="checkout", chave=web_key)
        assert checkout_row["forms_status"] == "sent"

        with SessionLocal() as db:
            queued = db.execute(
                select(FormsSubmission).where(FormsSubmission.request_id == checkout_request_id)
            ).scalar_one()
            assert queued.status == "success"
            assert queued.display_status == "sent"


def test_web_checkin_after_previous_checkout_processes_forms_and_admin_checkin_shows_sent(monkeypatch):
    monkeypatch.setattr(
        FormsWorker,
        "submit_with_retries",
        mocked_successful_forms_submit_with_statuses,
    )

    clear_forms_queue_backlog()
    web_key = make_test_key("W")
    first_checkin_request_id = f"web-checkin-initial-{uuid.uuid4().hex}"
    previous_checkout_request_id = f"web-checkout-previous-{uuid.uuid4().hex}"
    current_checkin_request_id = f"web-checkin-current-{uuid.uuid4().hex}"
    reference_time = now_sgt().replace(second=0, microsecond=0)
    current_checkin_at = reference_time - timedelta(minutes=5)
    previous_checkout_at = current_checkin_at - timedelta(hours=8)
    first_checkin_at = previous_checkout_at - timedelta(hours=8)

    with TestClient(app) as web_client, TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        auth_response = register_web_password(web_client, chave=web_key, senha="web123", projeto="P80")
        assert auth_response.status_code == 200, auth_response.text

        first_checkin_response = submit_web_check(
            web_client,
            chave=web_key,
            projeto="P80",
            action="checkin",
            event_time=first_checkin_at,
            client_event_id=first_checkin_request_id,
            local="Portaria P80",
        )
        assert first_checkin_response.status_code == 200, first_checkin_response.text
        assert first_checkin_response.json()["queued_forms"] is True

        previous_checkout_response = submit_web_check(
            web_client,
            chave=web_key,
            projeto="P80",
            action="checkout",
            event_time=previous_checkout_at,
            client_event_id=previous_checkout_request_id,
            local="Saida P80",
        )
        assert previous_checkout_response.status_code == 200, previous_checkout_response.text
        assert previous_checkout_response.json()["queued_forms"] is True

        current_checkin_response = submit_web_check(
            web_client,
            chave=web_key,
            projeto="P80",
            action="checkin",
            event_time=current_checkin_at,
            client_event_id=current_checkin_request_id,
            local="Retorno P80",
        )
        assert current_checkin_response.status_code == 200, current_checkin_response.text
        assert current_checkin_response.json()["queued_forms"] is True
        assert current_checkin_response.json()["state"]["current_action"] == "checkin"

        processed = process_forms_submission_queue_once(max_items=3)
        assert processed == 3

        checkin_row = get_admin_presence_row_by_chave(admin_client, endpoint="checkin", chave=web_key)
        assert checkin_row["forms_status"] == "sent"

        with SessionLocal() as db:
            queued = db.execute(
                select(FormsSubmission).where(FormsSubmission.request_id == current_checkin_request_id)
            ).scalar_one()
            assert queued.status == "success"
            assert queued.display_status == "sent"


def test_web_repeated_same_day_checkin_marks_admin_forms_as_not_realized():
    clear_forms_queue_backlog()
    web_key = make_test_key("W")
    first_request_id = f"web-checkin-first-{uuid.uuid4().hex}"
    repeated_request_id = f"web-checkin-repeat-{uuid.uuid4().hex}"
    reference_time = now_sgt().replace(second=0, microsecond=0)
    minutes_since_midnight = (reference_time.hour * 60) + reference_time.minute
    first_offset_minutes = max(min(minutes_since_midnight, 30), 2)
    repeated_offset_minutes = max(first_offset_minutes - 1, 1)
    first_checkin_at = reference_time - timedelta(minutes=first_offset_minutes)
    repeated_checkin_at = reference_time - timedelta(minutes=repeated_offset_minutes)

    with TestClient(app) as web_client, TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        auth_response = register_web_password(web_client, chave=web_key, senha="web123", projeto="P83")
        assert auth_response.status_code == 200, auth_response.text

        first_response = submit_web_check(
            web_client,
            chave=web_key,
            projeto="P83",
            action="checkin",
            event_time=first_checkin_at,
            client_event_id=first_request_id,
            local="Portaria P83",
        )
        assert first_response.status_code == 200, first_response.text
        assert first_response.json()["queued_forms"] is True

        repeated_response = submit_web_check(
            web_client,
            chave=web_key,
            projeto="P83",
            action="checkin",
            event_time=repeated_checkin_at,
            client_event_id=repeated_request_id,
            local="Portaria P83 B",
        )
        assert repeated_response.status_code == 200, repeated_response.text
        assert repeated_response.json()["queued_forms"] is False
        assert repeated_response.json()["state"]["current_action"] == "checkin"

        checkin_row = get_admin_presence_row_by_chave(admin_client, endpoint="checkin", chave=web_key)
        assert checkin_row["forms_status"] == "not_realized"

        with SessionLocal() as db:
            queued = db.execute(
                select(FormsSubmission).where(FormsSubmission.request_id == repeated_request_id)
            ).scalar_one()
            assert queued.status == "skipped"
            assert queued.display_status == "not_realized"
            assert queued.last_error == "repeated_same_action_same_day"


def test_web_repeated_checkout_next_day_marks_admin_forms_as_not_realized():
    clear_forms_queue_backlog()
    web_key = make_test_key("W")
    first_checkin_request_id = f"web-checkin-first-{uuid.uuid4().hex}"
    first_checkout_request_id = f"web-checkout-first-{uuid.uuid4().hex}"
    repeated_checkout_request_id = f"web-checkout-repeat-{uuid.uuid4().hex}"
    reference_time = now_sgt().replace(second=0, microsecond=0)
    repeated_checkout_at = reference_time - timedelta(minutes=5)
    first_checkout_at = repeated_checkout_at - timedelta(days=1)
    first_checkin_at = first_checkout_at - timedelta(hours=8)

    with TestClient(app) as web_client, TestClient(app) as admin_client:
        ensure_admin_session(admin_client)
        auth_response = register_web_password(web_client, chave=web_key, senha="web123", projeto="P80")
        assert auth_response.status_code == 200, auth_response.text

        first_checkin_response = submit_web_check(
            web_client,
            chave=web_key,
            projeto="P80",
            action="checkin",
            event_time=first_checkin_at,
            client_event_id=first_checkin_request_id,
            local="Portaria P80",
        )
        assert first_checkin_response.status_code == 200, first_checkin_response.text
        assert first_checkin_response.json()["queued_forms"] is True

        first_checkout_response = submit_web_check(
            web_client,
            chave=web_key,
            projeto="P80",
            action="checkout",
            event_time=first_checkout_at,
            client_event_id=first_checkout_request_id,
            local="Saida P80",
        )
        assert first_checkout_response.status_code == 200, first_checkout_response.text
        assert first_checkout_response.json()["queued_forms"] is True

        repeated_checkout_response = submit_web_check(
            web_client,
            chave=web_key,
            projeto="P80",
            action="checkout",
            event_time=repeated_checkout_at,
            client_event_id=repeated_checkout_request_id,
            local="Saida P80 Dia Seguinte",
        )
        assert repeated_checkout_response.status_code == 200, repeated_checkout_response.text
        assert repeated_checkout_response.json()["queued_forms"] is False
        assert repeated_checkout_response.json()["state"]["current_action"] == "checkout"

        checkout_row = get_admin_presence_row_by_chave(admin_client, endpoint="checkout", chave=web_key)
        assert checkout_row["forms_status"] == "not_realized"

        with SessionLocal() as db:
            queued = db.execute(
                select(FormsSubmission).where(FormsSubmission.request_id == repeated_checkout_request_id)
            ).scalar_one()
            assert queued.status == "skipped"
            assert queued.display_status == "not_realized"
            assert queued.last_error == "repeated_checkout"


def test_mobile_forms_submit_projects_sent_status_in_admin_checkin_after_queue(monkeypatch):
    monkeypatch.setattr(
        FormsWorker,
        "submit_with_retries",
        mocked_successful_forms_submit_with_statuses,
    )

    clear_forms_queue_backlog()
    mobile_key = make_test_key("M")
    client_event_id = f"mobile-admin-status-{uuid.uuid4().hex}"
    event_time = now_sgt().replace(second=0, microsecond=0) - timedelta(minutes=5)

    with TestClient(app) as client:
        ensure_admin_session(client)

        response = client.post(
            "/api/mobile/events/forms-submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": mobile_key,
                "projeto": "P80",
                "action": "checkin",
                "local": "Android Base P80",
                "informe": "normal",
                "event_time": event_time.isoformat(),
                "client_event_id": client_event_id,
            },
        )
        assert response.status_code == 200, response.text
        assert response.json()["queued_forms"] is True

        processed = process_forms_submission_queue_once(max_items=1)
        assert processed == 1

        checkin_row = get_admin_presence_row_by_chave(client, endpoint="checkin", chave=mobile_key)
        assert checkin_row["forms_status"] == "sent"

        with SessionLocal() as db:
            queued = db.execute(
                select(FormsSubmission).where(FormsSubmission.request_id == client_event_id)
            ).scalar_one()
            assert queued.request_path == "/api/mobile/events/forms-submit"
            assert queued.status == "success"
            assert queued.display_status == "sent"


def test_device_scan_projects_sent_status_in_admin_checkin_after_queue(monkeypatch):
    monkeypatch.setattr(
        FormsWorker,
        "submit_with_retries",
        mocked_successful_forms_submit_with_statuses,
    )

    clear_forms_queue_backlog()
    device_key = make_test_key("D")
    device_rfid = f"RF{uuid.uuid4().hex[:8].upper()}"
    request_id = f"device-admin-status-{uuid.uuid4().hex}"

    with TestClient(app) as client:
        ensure_admin_session(client)

        save_user = client.post(
            "/api/admin/users",
            json={"rfid": device_rfid, "nome": "Usuario Device Admin", "chave": device_key, "projeto": "P83"},
        )
        assert save_user.status_code == 200, save_user.text

        response = client.post(
            "/api/scan",
            json={
                "local": "Portaria Device P83",
                "rfid": device_rfid,
                "action": "checkin",
                "device_id": "ESP32-ADMIN-STATUS",
                "request_id": request_id,
                "shared_key": "device-test-key",
            },
        )
        assert response.status_code == 200, response.text
        assert response.json()["outcome"] == "submitted"

        processed = process_forms_submission_queue_once(max_items=1)
        assert processed == 1

        checkin_row = get_admin_presence_row_by_chave(client, endpoint="checkin", chave=device_key)
        assert checkin_row["forms_status"] == "sent"

        with SessionLocal() as db:
            queued = db.execute(
                select(FormsSubmission).where(FormsSubmission.request_id == request_id)
            ).scalar_one()
            assert queued.request_path == "/api/scan"
            assert queued.status == "success"
            assert queued.display_status == "sent"


def test_provider_current_state_keeps_presence_forms_status_empty_and_forms_panel_populated():
    clear_forms_queue_backlog()
    provider_key = make_test_key("P")
    provider_time = now_sgt().replace(second=0, microsecond=0) - timedelta(minutes=10)

    with TestClient(app) as client:
        ensure_admin_session(client)

        response = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": provider_key,
                "nome": "USUARIO PROVIDER PRESENCA",
                "projeto": "P80",
                "atividade": "check-out",
                "informe": "normal",
                "data": provider_time.strftime("%d/%m/%Y"),
                "hora": provider_time.strftime("%H:%M:%S"),
            },
        )
        assert response.status_code == 200, response.text
        assert response.json()["updated_current_state"] is True

        checkout_row = get_admin_presence_row_by_chave(client, endpoint="checkout", chave=provider_key)
        assert checkout_row["forms_status"] is None

        forms_rows = client.get("/api/admin/forms")
        assert forms_rows.status_code == 200, forms_rows.text
        provider_row = next(row for row in forms_rows.json() if row["chave"] == provider_key)
        assert provider_row["atividade"] == "check-out"

        with SessionLocal() as db:
            forms_submissions = db.execute(
                select(FormsSubmission).where(FormsSubmission.chave == provider_key)
            ).scalars().all()
            assert forms_submissions == []


def test_web_check_state_returns_latest_public_history():
    checkin_at = now_sgt() - timedelta(hours=1)
    checkout_at = now_sgt()

    with TestClient(app) as client:
        auth_response = register_web_password(client, chave="WB13", senha="state1", projeto="P80")
        assert auth_response.status_code == 200

        first = client.post(
            "/api/web/check",
            json={
                "chave": "WB13",
                "projeto": "P80",
                "action": "checkin",
                "informe": "normal",
                "event_time": checkin_at.isoformat(),
                "client_event_id": f"web-history-1-{uuid.uuid4().hex}",
            },
        )
        second = client.post(
            "/api/web/check",
            json={
                "chave": "WB13",
                "projeto": "P80",
                "action": "checkout",
                "informe": "normal",
                "event_time": checkout_at.isoformat(),
                "client_event_id": f"web-history-2-{uuid.uuid4().hex}",
            },
        )
        history = client.get("/api/web/check/state", params={"chave": "WB13"})

        assert first.status_code == 200
        assert second.status_code == 200
        assert history.status_code == 200

        payload = history.json()
        assert payload == {
            "found": True,
            "chave": "WB13",
            "projeto": "P80",
            "current_action": "checkout",
            "current_local": "Web",
            "has_current_day_checkin": True,
            "last_checkin_at": checkin_at.replace(tzinfo=None).isoformat(),
            "last_checkout_at": checkout_at.replace(tzinfo=None).isoformat(),
        }


def test_mobile_sync_accepts_project_p82():
    with TestClient(app) as client:
        response = client.post(
            "/api/mobile/events/sync",
            headers=MOBILE_HEADERS,
            json={
                "chave": "AP82",
                "projeto": "P82",
                "action": "checkin",
                "event_time": now_sgt().isoformat(),
                "client_event_id": f"android-{uuid.uuid4().hex}",
            },
        )
        assert response.status_code == 200
        assert response.json()["state"]["projeto"] == "P82"


def test_mobile_checkout_preserves_previous_checkin_history_without_existing_sync_events():
    previous_checkin_at = now_sgt() - timedelta(hours=2)
    checkout_at = now_sgt()

    with SessionLocal() as db:
        user = User(
            rfid=None,
            chave="LG11",
            nome="Legado Mobile",
            projeto="P80",
            local=None,
            checkin=True,
            time=previous_checkin_at,
            last_active_at=previous_checkin_at,
            inactivity_days=0,
        )
        db.add(user)
        db.commit()

    with TestClient(app) as client:
        response = client.post(
            "/api/mobile/events/sync",
            headers=MOBILE_HEADERS,
            json={
                "chave": "LG11",
                "projeto": "P80",
                "action": "checkout",
                "event_time": checkout_at.isoformat(),
                "client_event_id": f"android-{uuid.uuid4().hex}",
            },
        )

        assert response.status_code == 200
        state = response.json()["state"]
        assert state["current_action"] == "checkout"
        assert state["last_checkin_at"] is not None
        assert state["last_checkout_at"] is not None


def test_mobile_state_falls_back_to_check_events_history():
    checkin_at = now_sgt() - timedelta(hours=3)
    checkout_at = now_sgt() - timedelta(hours=1)

    with SessionLocal() as db:
        user = User(
            rfid="RFBACK1",
            chave="FB11",
            nome="Fallback Historico",
            projeto="P80",
            local="main",
            checkin=False,
            time=checkout_at,
            last_active_at=checkout_at,
            inactivity_days=0,
        )
        db.add(user)
        db.flush()
        db.add(
            CheckEvent(
                idempotency_key=f"fallback-checkin-{uuid.uuid4().hex}",
                source="device",
                rfid="RFBACK1",
                action="checkin",
                status="queued",
                message="checkin historico",
                details=None,
                project="P80",
                device_id="ESP32-FALLBACK",
                local="main",
                request_path="/api/scan",
                http_status=202,
                event_time=checkin_at,
                submitted_at=None,
                retry_count=0,
            )
        )
        db.add(
            CheckEvent(
                idempotency_key=f"fallback-checkout-{uuid.uuid4().hex}",
                source="device",
                rfid="RFBACK1",
                action="checkout",
                status="queued",
                message="checkout historico",
                details=None,
                project="P80",
                device_id="ESP32-FALLBACK",
                local="main",
                request_path="/api/scan",
                http_status=202,
                event_time=checkout_at,
                submitted_at=None,
                retry_count=0,
            )
        )
        db.commit()

    with TestClient(app) as client:
        response = client.get("/api/mobile/state?chave=FB11", headers=MOBILE_HEADERS)

        assert response.status_code == 200
        state = response.json()
        assert state["found"] is True
        assert state["current_action"] == "checkout"
        assert state["last_checkin_at"] is not None
        assert state["last_checkout_at"] is not None


def test_archive_events_creates_csv_clears_table_and_lists_downloads(tmp_path):
    settings.event_archives_dir = str(tmp_path / "event_archives")

    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "ARC1000", "nome": "Usuario Arquivo", "chave": "JK90", "projeto": "P80"},
        )
        assert save_user.status_code == 200

        archive_res = client.post("/api/admin/events/archive")
        assert archive_res.status_code == 200

        archive_payload = archive_res.json()
        assert archive_payload["created"] is True
        assert archive_payload["cleared_count"] >= 1
        assert archive_payload["archive"]["file_name"].endswith(".csv")
        assert " a " in archive_payload["archive"]["period"]
        assert archive_payload["archive"]["record_count"] >= 1
        assert archive_payload["archives"]["total"] >= 1
        assert archive_payload["archives"]["total_size_bytes"] >= archive_payload["archive"]["size_bytes"]
        assert archive_payload["archives"]["items"][0]["file_name"] == archive_payload["archive"]["file_name"]

        events_after = client.get("/api/admin/events")
        assert events_after.status_code == 200
        assert events_after.json() == []

        archives_list = client.get("/api/admin/events/archives")
        assert archives_list.status_code == 200
        assert archives_list.json()["items"][0]["file_name"] == archive_payload["archive"]["file_name"]
        assert archives_list.json()["page"] == 1
        assert archives_list.json()["page_size"] == 8

        download_res = client.get(
            f"/api/admin/events/archives/{archive_payload['archive']['file_name']}",
        )
        assert download_res.status_code == 200
        assert "attachment;" in download_res.headers["content-disposition"]
        assert "event_time" in download_res.text
        assert "ARC1000" in download_res.text

        download_all_res = client.get("/api/admin/events/archives/download-all")
        assert download_all_res.status_code == 200
        assert download_all_res.headers["content-type"].startswith("application/zip")
        assert len(download_all_res.content) > 0


def test_delete_archived_event_csv(tmp_path):
    settings.event_archives_dir = str(tmp_path / "event_archives")

    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "ARCDEL", "nome": "Usuario Delete", "chave": "DL90", "projeto": "P80"},
        )
        assert save_user.status_code == 200

        archive_res = client.post("/api/admin/events/archive")
        assert archive_res.status_code == 200
        file_name = archive_res.json()["archive"]["file_name"]

        delete_res = client.delete(f"/api/admin/events/archives/{file_name}")
        assert delete_res.status_code == 200
        assert delete_res.json()["ok"] is True

        list_res = client.get("/api/admin/events/archives")
        assert list_res.status_code == 200
        assert all(item["file_name"] != file_name for item in list_res.json()["items"])

        missing_res = client.get(f"/api/admin/events/archives/{file_name}")
        assert missing_res.status_code == 404


def test_event_archive_operations_are_logged(tmp_path):
    settings.event_archives_dir = str(tmp_path / "event_archives")

    with TestClient(app) as client:
        ensure_admin_session(client)
        save_user = client.post(
            "/api/admin/users",
            json={"rfid": "ARCAUD1", "nome": "Usuario Auditoria", "chave": "AU90", "projeto": "P80"},
        )
        assert save_user.status_code == 200

        archive_res = client.post("/api/admin/events/archive")
        assert archive_res.status_code == 200
        file_name = archive_res.json()["archive"]["file_name"]

        single_download = client.get(f"/api/admin/events/archives/{file_name}")
        assert single_download.status_code == 200

        download_all = client.get("/api/admin/events/archives/download-all")
        assert download_all.status_code == 200

        delete_res = client.delete(f"/api/admin/events/archives/{file_name}")
        assert delete_res.status_code == 200

        missing_res = client.get(f"/api/admin/events/archives/{file_name}")
        assert missing_res.status_code == 404

        events_res = client.get("/api/admin/events")
        assert events_res.status_code == 200
        assert events_res.json() == []

        with SessionLocal() as db:
            archive_events = db.execute(
                select(CheckEvent)
                .where(CheckEvent.action == "event_archive")
                .order_by(CheckEvent.id)
            ).scalars().all()

        assert any(
            event.status == "created"
            and event.request_path == "/api/admin/events/archive"
            for event in archive_events
        )
        assert any(
            event.status == "downloaded"
            and event.request_path == f"/api/admin/events/archives/{file_name}"
            for event in archive_events
        )
        assert any(
            event.status == "downloaded"
            and event.request_path == "/api/admin/events/archives/download-all"
            for event in archive_events
        )
        assert any(
            event.status == "removed"
            and event.request_path == f"/api/admin/events/archives/{file_name}"
            for event in archive_events
        )
        assert any(
            event.status == "failed"
            and event.http_status == 404
            and event.request_path == f"/api/admin/events/archives/{file_name}"
            for event in archive_events
        )


def test_archive_events_without_current_rows_returns_existing_archives(tmp_path):
    settings.event_archives_dir = str(tmp_path / "event_archives")

    with TestClient(app) as client:
        ensure_admin_session(client)
        first = client.post(
            "/api/admin/users",
            json={"rfid": "ARCEMPTY", "nome": "Usuario Empty", "chave": "EM90", "projeto": "P83"},
        )
        assert first.status_code == 200

        archive_first = client.post("/api/admin/events/archive")
        assert archive_first.status_code == 200
        assert archive_first.json()["created"] is True

        archive_second = client.post("/api/admin/events/archive")
        assert archive_second.status_code == 200
        assert archive_second.json()["created"] is False
        assert archive_second.json()["cleared_count"] == 0
        assert archive_second.json()["archives"]["total"] == 1


def test_list_event_archives_supports_backend_filter_and_pagination(tmp_path):
    settings.event_archives_dir = str(tmp_path / "event_archives")
    archive_dir = tmp_path / "event_archives"
    archive_dir.mkdir(parents=True, exist_ok=True)

    for index in range(12):
        file_path = archive_dir / f"2026-03-{index + 1:02d} 10-00-00 a 2026-03-{index + 1:02d} 11-00-00.csv"
        file_path.write_text("id,event_time\n1,2026-03-01T10:00:00\n", encoding="utf-8-sig")

    with TestClient(app) as client:
        ensure_admin_session(client)
        paged = client.get("/api/admin/events/archives?page=2&page_size=5")
        assert paged.status_code == 200
        payload = paged.json()
        assert payload["total"] == 12
        assert payload["page"] == 2
        assert payload["page_size"] == 5
        assert payload["total_pages"] == 3
        assert payload["total_size_bytes"] > 0
        assert len(payload["items"]) == 5

        filtered = client.get("/api/admin/events/archives?q=2026-03-01&page=1&page_size=5")
        assert filtered.status_code == 200
        filtered_payload = filtered.json()
        assert filtered_payload["total"] >= 1
        assert filtered_payload["query"] == "2026-03-01"
        assert all("2026-03-01" in item["period"] for item in filtered_payload["items"])


def test_admin_login_session_and_logout_flow():
    with TestClient(app) as client:
        session_before = client.get("/api/admin/auth/session")
        assert session_before.status_code == 200
        assert session_before.json()["authenticated"] is False

        login_response = login_admin(client)
        assert login_response.status_code == 200
        assert login_response.json()["ok"] is True

        session_after = client.get("/api/admin/auth/session")
        assert session_after.status_code == 200
        assert session_after.json()["authenticated"] is True
        assert session_after.json()["admin"]["chave"] == "HR70"
        assert session_after.json()["admin"]["can_view_activity_time"] is True
        assert session_after.json()["admin"]["access_scope"] == "full"
        assert session_after.json()["admin"]["allowed_tabs"] == [
            "checkin",
            "checkout",
            "forms",
            "inactive",
            "cadastro",
            "relatorios",
            "eventos",
            "banco-dados",
        ]

        admins = client.get("/api/admin/administrators")
        assert admins.status_code == 200
        assert any(
            row["chave"] == "HR70" and row["status"] == "active" and row["perfil"] == 9
            for row in admins.json()
        )

        logout_response = client.post("/api/admin/auth/logout")
        assert logout_response.status_code == 200
        assert logout_response.json()["ok"] is True

        session_final = client.get("/api/admin/auth/session")
        assert session_final.status_code == 200
        assert session_final.json()["authenticated"] is False


def test_admin_perfil_zero_session_is_limited_to_checkin_and_checkout():
    with SessionLocal() as db:
        user = find_user_by_chave(db, "PZ00")
        if user is None:
            user = User(
                rfid=None,
                nome="Perfil Zero",
                chave="PZ00",
                projeto="P80",
                senha=hash_password("lim123"),
                perfil=0,
                workplace=None,
                placa=None,
                end_rua=None,
                zip=None,
                email=None,
                local=None,
                checkin=None,
                time=None,
                last_active_at=now_sgt(),
                inactivity_days=0,
            )
            db.add(user)
        else:
            user.nome = "Perfil Zero"
            user.projeto = "P80"
            user.senha = hash_password("lim123")
            user.perfil = 0
            user.last_active_at = now_sgt()
            user.inactivity_days = 0
        db.commit()

    with TestClient(app) as client:
        login_response = login_admin(client, chave="PZ00", senha="lim123")
        assert login_response.status_code == 200
        assert login_response.json()["ok"] is True

        session_after = client.get("/api/admin/auth/session")
        assert session_after.status_code == 200
        payload = session_after.json()
        assert payload["authenticated"] is True
        assert payload["admin"]["chave"] == "PZ00"
        assert payload["admin"]["perfil"] == 0
        assert payload["admin"]["can_view_activity_time"] is False
        assert payload["admin"]["access_scope"] == "limited"
        assert payload["admin"]["allowed_tabs"] == ["checkin", "checkout"]

        assert client.get("/api/admin/checkin").status_code == 200
        assert client.get("/api/admin/checkout").status_code == 200

        for path in [
            "/api/admin/administrators",
            "/api/admin/projects",
            "/api/admin/forms",
            "/api/admin/inactive",
            "/api/admin/pending",
            "/api/admin/locations",
            "/api/admin/users",
            "/api/admin/events",
        ]:
            denied = client.get(path)
            assert denied.status_code == 403, path
            assert denied.json()["detail"] == "Este usuario nao possui permissao para esta area do Admin."

        denied_reports = client.get("/api/admin/reports/events", params={"chave": "RP41"})
        assert denied_reports.status_code == 403, denied_reports.text
        assert denied_reports.json()["detail"] == "Este usuario nao possui permissao para esta area do Admin."

        denied_reports_export = client.get("/api/admin/reports/events/export", params={"chave": "RP41"})
        assert denied_reports_export.status_code == 403, denied_reports_export.text
        assert denied_reports_export.json()["detail"] == "Este usuario nao possui permissao para esta area do Admin."

        denied_reports_export_all = client.get("/api/admin/reports/events/export-all")
        assert denied_reports_export_all.status_code == 403, denied_reports_export_all.text
        assert denied_reports_export_all.json()["detail"] == "Este usuario nao possui permissao para esta area do Admin."


def test_admin_checkin_and_checkout_keep_raw_and_safe_activity_time_for_profile_nine(monkeypatch):
    fixed_now = datetime(2026, 4, 26, 16, 30, 45, tzinfo=ZoneInfo(settings.tz_name))
    checkin_time = fixed_now - timedelta(hours=1, minutes=10, seconds=12)
    checkout_time = fixed_now - timedelta(minutes=25, seconds=3)
    monkeypatch.setattr(admin_router, "now_sgt", lambda: fixed_now)

    with SessionLocal() as db:
        for chave, nome, local, is_checkin, event_time in [
            ("P39C", "Perfil Nove Checkin", "Porta 9", True, checkin_time),
            ("P39O", "Perfil Nove Checkout", "Saida 9", False, checkout_time),
        ]:
            user = find_user_by_chave(db, chave)
            if user is None:
                user = User(
                    rfid=None,
                    nome=nome,
                    chave=chave,
                    projeto="P80",
                    local=local,
                    checkin=is_checkin,
                    time=event_time,
                    last_active_at=event_time,
                    inactivity_days=0,
                )
                db.add(user)
            else:
                user.nome = nome
                user.projeto = "P80"
                user.local = local
                user.checkin = is_checkin
                user.time = event_time
                user.last_active_at = event_time
                user.inactivity_days = 0
        db.commit()

    with TestClient(app) as client:
        ensure_admin_session(client)

        checkin_response = client.get("/api/admin/checkin")
        checkout_response = client.get("/api/admin/checkout")

    assert checkin_response.status_code == 200, checkin_response.text
    assert checkout_response.status_code == 200, checkout_response.text

    timezone = ZoneInfo(settings.tz_name)
    checkin_row = next(row for row in checkin_response.json() if row["chave"] == "P39C")
    assert normalize_event_time(datetime.fromisoformat(checkin_row["time"])) == checkin_time
    assert checkin_row["activity_date_label"] == checkin_time.astimezone(timezone).strftime("%d/%m/%Y")
    assert checkin_row["activity_time_label"] == checkin_time.astimezone(timezone).strftime("%H:%M:%S")
    assert checkin_row["activity_day_key"] == checkin_time.astimezone(timezone).strftime("%Y-%m-%d")

    checkout_row = next(row for row in checkout_response.json() if row["chave"] == "P39O")
    assert normalize_event_time(datetime.fromisoformat(checkout_row["time"])) == checkout_time
    assert checkout_row["activity_date_label"] == checkout_time.astimezone(timezone).strftime("%d/%m/%Y")
    assert checkout_row["activity_time_label"] == checkout_time.astimezone(timezone).strftime("%H:%M:%S")
    assert checkout_row["activity_day_key"] == checkout_time.astimezone(timezone).strftime("%Y-%m-%d")


def test_admin_checkin_and_checkout_hide_raw_activity_time_for_profile_zero(monkeypatch):
    fixed_now = datetime(2026, 4, 26, 16, 30, 45, tzinfo=ZoneInfo(settings.tz_name))
    checkin_time = fixed_now - timedelta(hours=2, minutes=5)
    checkout_time = fixed_now - timedelta(minutes=40)
    monkeypatch.setattr(admin_router, "now_sgt", lambda: fixed_now)

    with SessionLocal() as db:
        admin = find_user_by_chave(db, "P3L0")
        if admin is None:
            admin = User(
                rfid=None,
                nome="Perfil Zero Fase 3",
                chave="P3L0",
                projeto="P80",
                senha=hash_password("lim123"),
                perfil=0,
                workplace=None,
                placa=None,
                end_rua=None,
                zip=None,
                email=None,
                local=None,
                checkin=None,
                time=None,
                last_active_at=fixed_now,
                inactivity_days=0,
            )
            db.add(admin)
        else:
            admin.nome = "Perfil Zero Fase 3"
            admin.projeto = "P80"
            admin.senha = hash_password("lim123")
            admin.perfil = 0
            admin.last_active_at = fixed_now
            admin.inactivity_days = 0

        for chave, nome, local, is_checkin, event_time in [
            ("P30C", "Perfil Zero Checkin", "Porta 0", True, checkin_time),
            ("P30O", "Perfil Zero Checkout", "Saida 0", False, checkout_time),
        ]:
            user = find_user_by_chave(db, chave)
            if user is None:
                user = User(
                    rfid=None,
                    nome=nome,
                    chave=chave,
                    projeto="P80",
                    local=local,
                    checkin=is_checkin,
                    time=event_time,
                    last_active_at=event_time,
                    inactivity_days=0,
                )
                db.add(user)
            else:
                user.nome = nome
                user.projeto = "P80"
                user.local = local
                user.checkin = is_checkin
                user.time = event_time
                user.last_active_at = event_time
                user.inactivity_days = 0
            db.flush()
            grant_user_project_memberships(db, admin, ["P80"])
        db.commit()

    with TestClient(app) as client:
        login_response = login_admin(client, chave="P3L0", senha="lim123")
        assert login_response.status_code == 200, login_response.text

        checkin_response = client.get("/api/admin/checkin")
        checkout_response = client.get("/api/admin/checkout")

    assert checkin_response.status_code == 200, checkin_response.text
    assert checkout_response.status_code == 200, checkout_response.text

    timezone = ZoneInfo(settings.tz_name)
    checkin_row = next(row for row in checkin_response.json() if row["chave"] == "P30C")
    assert checkin_row["time"] is None
    assert checkin_row["activity_date_label"] == checkin_time.astimezone(timezone).strftime("%d/%m/%Y")
    assert checkin_row["activity_time_label"] is None
    assert checkin_row["activity_day_key"] == checkin_time.astimezone(timezone).strftime("%Y-%m-%d")

    checkout_row = next(row for row in checkout_response.json() if row["chave"] == "P30O")
    assert checkout_row["time"] is None
    assert checkout_row["activity_date_label"] == checkout_time.astimezone(timezone).strftime("%d/%m/%Y")
    assert checkout_row["activity_time_label"] is None
    assert checkout_row["activity_day_key"] == checkout_time.astimezone(timezone).strftime("%Y-%m-%d")


def test_mobile_sync_keeps_existing_memberships_across_sequential_legacy_project_switches():
    ensure_web_user_exists(chave="MP80", projeto="P83", nome="Usuario Mobile Multi Projeto")

    with SessionLocal() as db:
        user = get_user_by_chave(db, "MP80")
        ensure_user_active_project_is_member(db, user)
        add_user_project_membership(db, user, "P80")
        db.commit()

    with TestClient(app) as client:
        first_response = client.post(
            "/api/mobile/events/sync",
            headers=MOBILE_HEADERS,
            json={
                "chave": "MP80",
                "projeto": "P80",
                "action": "checkin",
                "event_time": now_sgt().isoformat(),
                "client_event_id": f"android-{uuid.uuid4().hex}",
            },
        )

    assert first_response.status_code == 200, first_response.text

    with SessionLocal() as db:
        user = get_user_by_chave(db, "MP80")
        assert user.projeto == "P80"
        assert get_materialized_user_project_names(db, user.id) == ["P80", "P83"]

    with TestClient(app) as client:
        second_response = client.post(
            "/api/mobile/events/sync",
            headers=MOBILE_HEADERS,
            json={
                "chave": "MP80",
                "projeto": "P83",
                "action": "checkout",
                "event_time": now_sgt().isoformat(),
                "client_event_id": f"android-{uuid.uuid4().hex}",
            },
        )

    assert second_response.status_code == 200, second_response.text

    with SessionLocal() as db:
        user = get_user_by_chave(db, "MP80")
        active_project = user.projeto
        materialized_names = get_materialized_user_project_names(db, user.id)

    assert active_project == "P83"
    assert materialized_names == ["P80", "P83"]


def test_mobile_sync_creates_materialized_membership_for_new_placeholder_user():
    with TestClient(app) as client:
        response = client.post(
            "/api/mobile/events/sync",
            headers=MOBILE_HEADERS,
            json={
                "chave": "MM82",
                "projeto": "P82",
                "action": "checkin",
                "event_time": now_sgt().isoformat(),
                "client_event_id": f"android-{uuid.uuid4().hex}",
            },
        )

    assert response.status_code == 200, response.text

    with SessionLocal() as db:
        user = get_user_by_chave(db, "MM82")
        active_project = user.projeto
        materialized_names = get_materialized_user_project_names(db, user.id)

    assert active_project == "P82"
    assert materialized_names == ["P82"]


def test_provider_submit_creates_materialized_membership_for_new_user():
    with TestClient(app) as client:
        response = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": "PV82",
                "nome": "Usuario Provider Membership",
                "projeto": "P82",
                "atividade": "check-in",
                "informe": "normal",
                "data": "18/04/2026",
                "hora": "08:00:00",
            },
        )

    assert response.status_code == 200, response.text

    with SessionLocal() as db:
        user = get_user_by_chave(db, "PV82")
        active_project = user.projeto
        materialized_names = get_materialized_user_project_names(db, user.id)

    assert active_project == "P82"
    assert materialized_names == ["P82"]


def test_provider_submit_keeps_existing_memberships_when_legacy_event_changes_active_project():
    ensure_web_user_exists(chave="PV83", projeto="P80", nome="Usuario Provider Multi Projeto")

    with SessionLocal() as db:
        user = get_user_by_chave(db, "PV83")
        ensure_user_active_project_is_member(db, user)
        add_user_project_membership(db, user, "P83")
        db.commit()

    with TestClient(app) as client:
        response = client.post(
            "/api/provider/updaterecords",
            headers=PROVIDER_HEADERS,
            json={
                "chave": "PV83",
                "nome": "Usuario Provider Multi Projeto",
                "projeto": "P83",
                "atividade": "check-in",
                "informe": "normal",
                "data": "18/04/2026",
                "hora": "08:00:00",
            },
        )

    assert response.status_code == 200, response.text
    assert response.json()["updated_project"] is True

    with SessionLocal() as db:
        user = get_user_by_chave(db, "PV83")
        active_project = user.projeto
        materialized_names = get_materialized_user_project_names(db, user.id)

    assert active_project == "P83"
    assert materialized_names == ["P80", "P83"]


def test_mobile_forms_submit_keeps_plural_memberships_while_queueing_single_operational_project():
    ensure_web_user_exists(chave="MF83", projeto="P83", nome="Usuario Forms Multi Projeto")

    with SessionLocal() as db:
        user = get_user_by_chave(db, "MF83")
        ensure_user_active_project_is_member(db, user)
        add_user_project_membership(db, user, "P80")
        db.commit()

    with TestClient(app) as client:
        response = client.post(
            "/api/mobile/events/forms-submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "MF83",
                "projeto": "P80",
                "action": "checkin",
                "informe": "normal",
                "event_time": now_sgt().isoformat(),
                "client_event_id": f"android-forms-{uuid.uuid4().hex}",
            },
        )

    assert response.status_code == 200, response.text
    assert response.json()["queued_forms"] is True

    with SessionLocal() as db:
        user = get_user_by_chave(db, "MF83")
        queued = db.execute(
            select(FormsSubmission)
            .where(FormsSubmission.chave == "MF83")
            .order_by(FormsSubmission.id.desc())
        ).scalars().first()
        sync_event = db.execute(
            select(UserSyncEvent)
            .where(UserSyncEvent.chave == "MF83")
            .order_by(UserSyncEvent.id.desc())
        ).scalars().first()

        assert user.projeto == "P80"
        assert get_materialized_user_project_names(db, user.id) == ["P80", "P83"]
        assert queued is not None and queued.projeto == "P80"
        assert sync_event is not None and sync_event.projeto == "P80"


def test_web_self_registration_creates_materialized_membership_for_new_user():
    with TestClient(app) as client:
        response = client.post(
            "/api/web/auth/register-user",
            json={
                "chave": "WM82",
                "nome": "usuario web membership",
                "projetos": ["P82"],
                "email": "wm82@example.com",
                "senha": "cad123",
                "confirmar_senha": "cad123",
            },
        )

    assert response.status_code == 201, response.text

    with SessionLocal() as db:
        user = get_user_by_chave(db, "WM82")
        active_project = user.projeto
        materialized_names = get_materialized_user_project_names(db, user.id)

    assert active_project == "P82"
    assert materialized_names == ["P82"]


def test_admin_user_create_materializes_membership_row():
    with TestClient(app) as client:
        ensure_admin_session(client)
        response = client.post(
            "/api/admin/users",
            json={
                "rfid": "ADM8201",
                "nome": "Usuario Admin Membership",
                "chave": "AM82",
                "projeto": "P82",
            },
        )

    assert response.status_code == 200, response.text

    with SessionLocal() as db:
        user = get_user_by_chave(db, "AM82")
        active_project = user.projeto
        materialized_names = get_materialized_user_project_names(db, user.id)

    assert active_project == "P82"
    assert materialized_names == ["P82"]


def test_admin_legacy_single_project_update_replaces_memberships_with_explicit_single_selection():
    user_key = make_test_key("A")
    with TestClient(app) as client:
        ensure_admin_session(client)
        created = client.post(
            "/api/admin/users",
            json={
                "rfid": "ADM8301",
                "nome": "Usuario Admin Multi Projeto",
                "chave": user_key,
                "projeto": "P82",
            },
        )
        assert created.status_code == 200, created.text

    with SessionLocal() as db:
        user = get_user_by_chave(db, user_key)
        user_id = user.id
        ensure_user_active_project_is_member(db, user)
        add_user_project_membership(db, user, "P83")
        db.commit()

    with TestClient(app) as client:
        ensure_admin_session(client)
        updated = client.post(
            "/api/admin/users",
            json={
                "user_id": user_id,
                "rfid": "ADM8301",
                "nome": "Usuario Admin Multi Projeto",
                "chave": user_key,
                "projeto": "P83",
            },
        )

    assert updated.status_code == 200, updated.text

    with SessionLocal() as db:
        user = get_user_by_chave(db, user_key)
        active_project = user.projeto
        materialized_names = get_materialized_user_project_names(db, user.id)

    assert active_project == "P83"
    assert materialized_names == ["P83"]


def test_scoped_admin_user_update_preserves_memberships_outside_visible_scope():
    with SessionLocal() as db:
        scoped_admin = User(
            rfid=None,
            nome="Admin Escopo P80",
            chave="SM80",
            projeto="P80",
            senha=hash_password("scope123"),
            perfil=1,
            workplace=None,
            placa=None,
            end_rua=None,
            zip=None,
            email=None,
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
        )
        target_user = User(
            rfid="SCP8001",
            nome="Usuario Multi Escopo",
            chave="SU80",
            projeto="P83",
            workplace=None,
            placa=None,
            end_rua=None,
            zip=None,
            email=None,
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
        )
        db.add_all([scoped_admin, target_user])
        db.flush()
        grant_user_project_memberships(db, scoped_admin, ["P80"])
        grant_user_project_memberships(db, target_user, ["P80", "P83"])
        db.commit()
        target_user_id = target_user.id

    with TestClient(app) as client:
        login_response = login_admin(client, chave="SM80", senha="scope123")
        assert login_response.status_code == 200, login_response.text

        updated = client.post(
            "/api/admin/users",
            json={
                "user_id": target_user_id,
                "rfid": "SCP8001",
                "nome": "Usuario Multi Escopo Atualizado",
                "chave": "SU80",
                "projeto": "P80",
                "projetos": ["P80"],
            },
        )

    assert updated.status_code == 200, updated.text

    with SessionLocal() as db:
        target_user = get_user_by_chave(db, "SU80")
        materialized_names = get_materialized_user_project_names(db, target_user.id)

    assert target_user.nome == "Usuario Multi Escopo Atualizado"
    assert target_user.projeto == "P80"
    assert materialized_names == ["P80", "P83"]


def test_admin_runtime_scope_requires_materialized_memberships_and_ignores_legacy_scope_fields():
    recent_time = now_sgt() - timedelta(hours=1)

    with SessionLocal() as db:
        admin = User(
            rfid=None,
            nome="Admin Sem Membership Materializada",
            chave="AM00",
            projeto="P80",
            senha=hash_password("scope123"),
            perfil=1,
            admin_monitored_projects_json=dump_admin_monitored_projects(["P83"]),
            workplace=None,
            placa=None,
            end_rua=None,
            zip=None,
            email=None,
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
        )
        legacy_project_user = User(
            rfid=None,
            nome="Usuario Projeto Legado P80",
            chave="AM80",
            projeto="P80",
            local="Porta 80",
            checkin=True,
            time=recent_time,
            last_active_at=recent_time,
            inactivity_days=0,
        )
        legacy_scope_user = User(
            rfid=None,
            nome="Usuario Escopo Legado P83",
            chave="AM83",
            projeto="P83",
            local="Porta 83",
            checkin=True,
            time=recent_time,
            last_active_at=recent_time,
            inactivity_days=0,
        )
        db.add_all([admin, legacy_project_user, legacy_scope_user])
        db.flush()
        grant_user_project_memberships(db, legacy_project_user, ["P80"])
        grant_user_project_memberships(db, legacy_scope_user, ["P83"])
        db.commit()

    with TestClient(app) as client:
        login_response = login_admin(client, chave="AM00", senha="scope123")
        assert login_response.status_code == 200, login_response.text

        users_response = client.get("/api/admin/users")
        checkin_response = client.get("/api/admin/checkin")
        blocked_create = client.post(
            "/api/admin/users",
            json={
                "rfid": "AMNEW",
                "nome": "Bloqueado Sem Membership",
                "chave": "AMN0",
                "projeto": "P80",
            },
        )

    assert users_response.status_code == 200, users_response.text
    assert users_response.json() == []

    assert checkin_response.status_code == 200, checkin_response.text
    assert checkin_response.json() == []

    assert blocked_create.status_code == 403, blocked_create.text
    assert blocked_create.json()["detail"] == "Administrador sem projetos vinculados nao pode alterar usuarios."

    with SessionLocal() as db:
        admin = get_user_by_chave(db, "AM00")
        assert get_materialized_user_project_names(db, admin.id) == []


def test_profile_nine_runtime_scope_uses_memberships_for_users_events_reports_and_database_events():
    event_time = datetime(2026, 4, 25, 7, 30, 0, tzinfo=ZoneInfo(settings.tz_name))

    with SessionLocal() as db:
        admin = User(
            rfid=None,
            nome="Perfil Nove Escopo Limitado",
            chave="P9S1",
            projeto="P80",
            senha=hash_password("adm123"),
            perfil=9,
            workplace=None,
            placa=None,
            end_rua=None,
            zip=None,
            email=None,
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
        )
        visible_user = User(
            rfid="P9RF80",
            nome="Usuario Escopo P80",
            chave="P980",
            projeto="P80",
            workplace=None,
            placa=None,
            end_rua=None,
            zip=None,
            email=None,
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
        )
        hidden_user = User(
            rfid="P9RF83",
            nome="Usuario Escopo P83",
            chave="P983",
            projeto="P83",
            workplace=None,
            placa=None,
            end_rua=None,
            zip=None,
            email=None,
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
        )
        db.add_all([admin, visible_user, hidden_user])
        db.flush()
        grant_user_project_memberships(db, admin, ["P80"])
        grant_user_project_memberships(db, visible_user, ["P80"])
        grant_user_project_memberships(db, hidden_user, ["P83"])
        db.add_all(
            [
                UserSyncEvent(
                    user_id=visible_user.id,
                    chave=visible_user.chave,
                    rfid=visible_user.rfid,
                    source="web",
                    action="checkin",
                    projeto="P80",
                    local="main",
                    ontime=True,
                    event_time=event_time,
                    created_at=now_sgt(),
                    source_request_id=f"profile-nine-visible-{uuid.uuid4().hex}",
                    device_id=None,
                ),
                UserSyncEvent(
                    user_id=hidden_user.id,
                    chave=hidden_user.chave,
                    rfid=hidden_user.rfid,
                    source="provider",
                    action="checkout",
                    projeto="P83",
                    local="Forms",
                    ontime=False,
                    event_time=event_time,
                    created_at=now_sgt(),
                    source_request_id=f"profile-nine-hidden-{uuid.uuid4().hex}",
                    device_id="provider",
                ),
            ]
        )
        visible_event = CheckEvent(
            idempotency_key=f"profile-nine-db-visible-{uuid.uuid4().hex}",
            source="device",
            rfid=visible_user.rfid,
            action="checkin",
            status="success",
            message="Entrada P80",
            details="scope=P80",
            project="P80",
            device_id="ESP32-P80",
            local="Portaria 80",
            request_path="/api/scan",
            http_status=200,
            ontime=True,
            event_time=event_time,
            submitted_at=event_time,
            retry_count=0,
        )
        hidden_event = CheckEvent(
            idempotency_key=f"profile-nine-db-hidden-{uuid.uuid4().hex}",
            source="device",
            rfid=hidden_user.rfid,
            action="checkout",
            status="success",
            message="Saida P83",
            details="scope=P83",
            project="P83",
            device_id="ESP32-P83",
            local="Portaria 83",
            request_path="/api/scan",
            http_status=200,
            ontime=False,
            event_time=event_time,
            submitted_at=event_time,
            retry_count=0,
        )
        db.add_all([visible_event, hidden_event])
        db.commit()
        visible_event_id = visible_event.id
        hidden_event_id = hidden_event.id

    with TestClient(app) as client:
        login_response = login_admin(client, chave="P9S1", senha="adm123")
        assert login_response.status_code == 200, login_response.text

        users_response = client.get("/api/admin/users")
        events_response = client.get("/api/admin/events")
        database_events_response = client.get("/api/admin/database-events")
        visible_report_response = client.get("/api/admin/reports/events", params={"chave": "P980"})
        hidden_report_response = client.get("/api/admin/reports/events", params={"chave": "P983"})
        export_all_response = client.get("/api/admin/reports/events/export-all")

    assert users_response.status_code == 200, users_response.text
    visible_user_keys = {row["chave"] for row in users_response.json()}
    assert "P980" in visible_user_keys
    assert "P983" not in visible_user_keys

    assert events_response.status_code == 200, events_response.text
    visible_event_ids = {row["id"] for row in events_response.json()}
    assert visible_event_id in visible_event_ids
    assert hidden_event_id not in visible_event_ids

    assert database_events_response.status_code == 200, database_events_response.text
    database_payload = database_events_response.json()
    database_event_ids = {row["id"] for row in database_payload["items"]}
    assert visible_event_id in database_event_ids
    assert hidden_event_id not in database_event_ids
    assert database_payload["filter_options"]["project"] == ["P80"]

    assert visible_report_response.status_code == 200, visible_report_response.text
    assert visible_report_response.json()["person"]["chave"] == "P980"

    assert hidden_report_response.status_code == 404, hidden_report_response.text

    assert export_all_response.status_code == 200, export_all_response.text
    workbook = load_workbook(io.BytesIO(export_all_response.content))
    worksheet = workbook.active
    exported_names = {
        row[0]
        for row in worksheet.iter_rows(min_row=2, max_col=1, values_only=True)
        if row[0]
    }
    workbook.close()

    assert "Usuario Escopo P80" in exported_names
    assert "Usuario Escopo P83" not in exported_names


def test_admin_perfil_one_session_keeps_full_scope_without_activity_time_visibility():
    with SessionLocal() as db:
        user = find_user_by_chave(db, "P100")
        if user is None:
            user = User(
                rfid=None,
                nome="Perfil Um",
                chave="P100",
                projeto="P80",
                senha=hash_password("adm123"),
                perfil=1,
                workplace=None,
                placa=None,
                end_rua=None,
                zip=None,
                email=None,
                local=None,
                checkin=None,
                time=None,
                last_active_at=now_sgt(),
                inactivity_days=0,
            )
            db.add(user)
        else:
            user.nome = "Perfil Um"
            user.projeto = "P80"
            user.senha = hash_password("adm123")
            user.perfil = 1
            user.last_active_at = now_sgt()
            user.inactivity_days = 0
        db.commit()

    with TestClient(app) as client:
        login_response = login_admin(client, chave="P100", senha="adm123")
        assert login_response.status_code == 200
        assert login_response.json()["ok"] is True

        session_after = client.get("/api/admin/auth/session")
        assert session_after.status_code == 200
        payload = session_after.json()
        assert payload["authenticated"] is True
        assert payload["admin"]["chave"] == "P100"
        assert payload["admin"]["perfil"] == 1
        assert payload["admin"]["can_view_activity_time"] is False
        assert payload["admin"]["access_scope"] == "full"
        assert payload["admin"]["allowed_tabs"] == [
            "checkin",
            "checkout",
            "forms",
            "inactive",
            "cadastro",
            "relatorios",
            "eventos",
            "banco-dados",
        ]


def test_admin_request_access_and_approval_flow():
    with TestClient(app) as client:
        request_response = client.post(
            "/api/admin/auth/request-access",
            json={
                "chave": "TS11",
                "nome_completo": "Teste Solicitante",
                "senha": "Senha@123",
            },
        )
        assert request_response.status_code == 200
        assert request_response.json()["ok"] is True

        duplicate_response = client.post(
            "/api/admin/auth/request-access",
            json={
                "chave": "TS11",
                "nome_completo": "Teste Solicitante",
                "senha": "Senha@123",
            },
        )
        assert duplicate_response.status_code == 409

        login_response = login_admin(client)
        assert login_response.status_code == 200

        list_response = client.get("/api/admin/administrators")
        assert list_response.status_code == 200
        pending_row = next(row for row in list_response.json() if row["chave"] == "TS11")
        assert pending_row["status"] == "pending"

        approve_response = client.post(f"/api/admin/administrators/requests/{pending_row['id']}/approve")
        assert approve_response.status_code == 200
        assert approve_response.json()["ok"] is True

        relogin = client.post("/api/admin/auth/logout")
        assert relogin.status_code == 200

        new_admin_login = login_admin(client, chave="TS11", senha="Senha@123")
        assert new_admin_login.status_code == 200

        with SessionLocal() as db:
            admin = db.execute(select(User).where(User.chave == "TS11")).scalar_one_or_none()
            pending = db.execute(select(AdminAccessRequest).where(AdminAccessRequest.chave == "TS11")).scalar_one_or_none()
            assert admin is not None
            assert admin.perfil == 1
            assert admin.senha is not None
            assert pending is None


def test_admin_self_service_request_registers_unknown_user_and_allows_profile_override_on_approval():
    with TestClient(app) as client:
        status_response = client.get("/api/admin/auth/request-access/status", params={"chave": "NR11"})
        assert status_response.status_code == 200
        assert status_response.json()["found"] is False

        request_response = client.post(
            "/api/admin/auth/request-access/self-service",
            json={
                "chave": "NR11",
                "nome_completo": "Novo Requisitante",
                "projeto": "P80",
                "senha": "cad123",
                "confirmar_senha": "cad123",
            },
        )
        assert request_response.status_code == 200
        assert request_response.json()["ok"] is True

        with SessionLocal() as db:
            user = db.execute(select(User).where(User.chave == "NR11")).scalar_one_or_none()
            pending = db.execute(select(AdminAccessRequest).where(AdminAccessRequest.chave == "NR11")).scalar_one_or_none()
            assert user is not None
            assert user.perfil == 0
            assert user.projeto == "P80"
            assert user.senha is not None
            assert verify_password("cad123", user.senha) is True
            assert pending is not None
            assert pending.requested_profile == 1

        login_response = login_admin(client)
        assert login_response.status_code == 200

        rows_response = client.get("/api/admin/administrators")
        assert rows_response.status_code == 200
        pending_row = next(row for row in rows_response.json() if row["chave"] == "NR11")
        assert pending_row["row_type"] == "request"
        assert pending_row["perfil"] == 1

        approve_response = client.post(
            f"/api/admin/administrators/requests/{pending_row['id']}/approve",
            json={"perfil": 9},
        )
        assert approve_response.status_code == 200
        assert approve_response.json()["ok"] is True

        with SessionLocal() as db:
            user = db.execute(select(User).where(User.chave == "NR11")).scalar_one_or_none()
            pending = db.execute(select(AdminAccessRequest).where(AdminAccessRequest.chave == "NR11")).scalar_one_or_none()
            assert user is not None
            assert user.perfil == 9
            assert pending is None


def test_admin_self_service_request_uses_registered_transport_user_and_keeps_revoke_available():
    ensure_web_user_exists(chave="RK12", nome="Requisitante Conhecido")

    with SessionLocal() as db:
        user = get_user_by_chave(db, "RK12")
        user.perfil = 2
        db.commit()

    with TestClient(app) as client:
        registration = register_web_password(client, chave="RK12", senha="rk1234", ensure_user_exists=False)
        assert registration.status_code == 200

        status_response = client.get("/api/admin/auth/request-access/status", params={"chave": "RK12"})
        assert status_response.status_code == 200
        assert status_response.json()["found"] is True
        assert status_response.json()["has_password"] is True
        assert status_response.json()["is_admin"] is False
        assert status_response.json()["has_pending_request"] is False

        request_response = client.post(
            "/api/admin/auth/request-access/self-service",
            json={"chave": "RK12"},
        )
        assert request_response.status_code == 200
        assert request_response.json()["ok"] is True

        login_response = login_admin(client)
        assert login_response.status_code == 200

        rows_response = client.get("/api/admin/administrators")
        assert rows_response.status_code == 200
        pending_row = next(row for row in rows_response.json() if row["chave"] == "RK12")
        assert pending_row["row_type"] == "request"

        approve_response = client.post(
            f"/api/admin/administrators/requests/{pending_row['id']}/approve",
            json={"perfil": 2},
        )
        assert approve_response.status_code == 200

        admins_response = client.get("/api/admin/administrators")
        assert admins_response.status_code == 200
        approved_row = next(
            row for row in admins_response.json() if row["chave"] == "RK12" and row["row_type"] == "admin"
        )
        assert approved_row["perfil"] == 3
        assert approved_row["can_revoke"] is True
        profile_update_response = client.post(
            f"/api/admin/administrators/{approved_row['id']}/profile",
            json={"perfil": 2},
        )
        assert profile_update_response.status_code == 200
        assert profile_update_response.json()["ok"] is True

        refreshed_admins_response = client.get("/api/admin/administrators")
        assert refreshed_admins_response.status_code == 200
        refreshed_row = next(
            row for row in refreshed_admins_response.json() if row["chave"] == "RK12" and row["row_type"] == "admin"
        )
        assert refreshed_row["perfil"] == 3
        assert refreshed_row["can_revoke"] is True

        with SessionLocal() as db:
            user = db.execute(select(User).where(User.chave == "RK12")).scalar_one_or_none()
            assert user is not None
            assert user.perfil == 3
            assert user.senha is not None
            assert verify_password("rk1234", user.senha) is True


def test_admin_request_access_frontend_e2e_handles_existing_and_unknown_keys():
    existing_key = make_test_key("E")
    unknown_key = make_test_key("N")
    existing_password = "ex1234"
    unknown_password = "cad123"
    existing_name = "Usuario Existente Front"
    unknown_name = "Usuario Novo Front"

    ensure_web_user_exists(chave=existing_key, projeto="P80", nome=existing_name)
    with TestClient(app) as client:
        register_response = register_web_password(
            client,
            chave=existing_key,
            senha=existing_password,
            projeto="P80",
            ensure_user_exists=False,
        )
        assert register_response.status_code == 200, register_response.text

    with live_app_server() as base_url:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            try:
                page.goto(f"{base_url}/admin", wait_until="domcontentloaded")
                page.locator("#requestAdminButton").wait_for(state="visible")

                page.get_by_role("button", name="Solicitar Administração").click()
                page.locator("#requestAdminModal").wait_for(state="visible")
                page.locator("#requestAdminChave").fill(unknown_key)
                page.locator("#requestAdminRegistrationModal").wait_for(state="visible")
                assert page.input_value("#requestAdminRegistrationChave") == unknown_key
                assert "Chave nao cadastrada" in (page.text_content("#requestAdminRegistrationStatus") or "")

                page.locator("#requestAdminRegistrationNome").fill(unknown_name)
                page.locator("#requestAdminRegistrationProjeto").select_option("P80")
                page.locator("#requestAdminRegistrationSenha").fill(unknown_password)
                page.locator("#requestAdminRegistrationConfirm").fill(unknown_password)
                page.locator("#requestAdminRegistrationSaveButton").click()
                page.wait_for_function(
                    "() => document.querySelector('#authStatus').textContent.includes('Solicitacao enviada')"
                )
                page.wait_for_function(
                    "() => document.querySelector('#requestAdminRegistrationModal').classList.contains('hidden')"
                )

                page.get_by_role("button", name="Solicitar Administração").click()
                page.locator("#requestAdminModal").wait_for(state="visible")
                page.locator("#requestAdminChave").fill(existing_key)
                page.wait_for_function(
                    "() => document.querySelector('#requestAdminStatus').textContent.includes('Solicitacao enviada')"
                )
                assert page.locator("#requestAdminRegistrationModal").is_hidden()
                page.wait_for_function(
                    "() => document.querySelector('#requestAdminModal').classList.contains('hidden')"
                )
            finally:
                browser.close()

    with SessionLocal() as db:
        existing_user = db.execute(select(User).where(User.chave == existing_key)).scalar_one_or_none()
        unknown_user = db.execute(select(User).where(User.chave == unknown_key)).scalar_one_or_none()
        existing_request = db.execute(
            select(AdminAccessRequest).where(AdminAccessRequest.chave == existing_key)
        ).scalar_one_or_none()
        unknown_request = db.execute(
            select(AdminAccessRequest).where(AdminAccessRequest.chave == unknown_key)
        ).scalar_one_or_none()

        assert existing_user is not None
        assert existing_user.nome == existing_name
        assert existing_user.perfil == 0
        assert existing_user.senha is not None
        assert verify_password(existing_password, existing_user.senha) is True

        assert unknown_user is not None
        assert unknown_user.nome == unknown_name
        assert unknown_user.projeto == "P80"
        assert unknown_user.perfil == 0
        assert unknown_user.senha is not None
        assert verify_password(unknown_password, unknown_user.senha) is True

        assert existing_request is not None
        assert existing_request.nome_completo == existing_name
        assert existing_request.requested_profile == 1

        assert unknown_request is not None
        assert unknown_request.nome_completo == unknown_name
        assert unknown_request.requested_profile == 1


def test_stale_checkin_rows_leave_checkin_and_move_to_inactive():
    ensure_web_user_exists(chave="CI11", nome="Checkin Antigo", projeto="P80")
    stale_time = now_sgt() - timedelta(days=1, minutes=5)
    set_user_checkin_state(chave="CI11", event_time=stale_time, local="Web")

    with TestClient(app) as client:
        ensure_admin_session(client)

        checkin_response = client.get("/api/admin/checkin")
        assert checkin_response.status_code == 200
        checkin_rows = checkin_response.json()
        assert all(row["chave"] != "CI11" for row in checkin_rows)

        inactive_response = client.get("/api/admin/inactive")
        assert inactive_response.status_code == 200
        inactive_row = next(row for row in inactive_response.json() if row["chave"] == "CI11")
        assert inactive_row["latest_action"] == "checkin"
        assert inactive_row["inactivity_days"] >= 1


def test_admin_time_based_rows_include_timezone_metadata_and_system_event_fallback():
    reference_now = now_sgt()
    active_time = reference_now - timedelta(hours=1)
    inactive_time = reference_now - timedelta(hours=26)
    active_key = f"T{uuid.uuid4().hex[:3].upper()}"
    inactive_key = f"T{uuid.uuid4().hex[:3].upper()}"
    project_name = f"J{uuid.uuid4().hex[:3].upper()}"

    with SessionLocal() as db:
        project = db.execute(select(Project).where(Project.name == project_name)).scalar_one_or_none()
        if project is None:
            db.add(Project(name=project_name, **build_project_fields_for_country("JP")))

        db.add(
            User(
                rfid=None,
                chave=active_key,
                nome="Usuario Horario Ativo",
                projeto=project_name,
                local="Porta A",
                checkin=True,
                time=active_time,
                last_active_at=active_time,
                inactivity_days=0,
            )
        )
        db.add(
            User(
                rfid=None,
                chave=inactive_key,
                nome="Usuario Horario Inativo",
                projeto=project_name,
                local="Porta B",
                checkin=False,
                time=inactive_time,
                last_active_at=inactive_time,
                inactivity_days=1,
            )
        )
        db.add(
            CheckEvent(
                idempotency_key=uuid.uuid4().hex,
                source="admin",
                action="info",
                status="done",
                message="Evento tecnico sem projeto",
                details=None,
                project=None,
                device_id=None,
                local=None,
                request_path="/api/admin/test-timezone",
                http_status=200,
                ontime=None,
                event_time=reference_now,
                submitted_at=None,
                retry_count=0,
                rfid=None,
            )
        )
        db.commit()

    with TestClient(app) as client:
        ensure_admin_session(client)

        checkin_response = client.get("/api/admin/checkin")
        assert checkin_response.status_code == 200
        active_row = next(row for row in checkin_response.json() if row["chave"] == active_key)
        assert active_row["timezone_name"] == "Asia/Tokyo"
        assert active_row["timezone_label"] == "Japão (+9)"

        inactive_response = client.get("/api/admin/inactive")
        assert inactive_response.status_code == 200
        inactive_row = next(row for row in inactive_response.json() if row["chave"] == inactive_key)
        assert inactive_row["timezone_name"] == "Asia/Tokyo"
        assert inactive_row["timezone_label"] == "Japão (+9)"

        events_response = client.get("/api/admin/events")
        assert events_response.status_code == 200
        technical_event = next(row for row in events_response.json() if row["message"] == "Evento tecnico sem projeto")
        assert technical_event["project"] is None
        assert technical_event["timezone_name"] == str(settings.tz_name)
        assert technical_event["timezone_label"] == build_timezone_label(
            country_name="Sistema",
            timezone_name=str(settings.tz_name),
            reference_time=reference_now,
        )


def test_admin_login_rejects_transport_only_profile():
    ensure_web_user_exists(chave="TT20", nome="Transport Only")
    with TestClient(app) as client:
        registration = register_web_password(client, chave="TT20", senha="tt2024", ensure_user_exists=False)
        assert registration.status_code == 200

    with SessionLocal() as db:
        transport_only_user = get_user_by_chave(db, "TT20")
        transport_only_user.perfil = 2
        db.commit()

    with TestClient(app) as client:
        login_response = login_admin(client, chave="TT20", senha="tt2024")
        assert login_response.status_code == 403


def test_administrators_endpoint_lists_admin_profiles_only():
    ensure_web_user_exists(chave="UTO9", nome="Admin Perfil")
    ensure_web_user_exists(chave="TP22", nome="Transport Perfil")
    ensure_web_user_exists(chave="ZZ00", nome="Sem Perfil")

    with SessionLocal() as db:
        get_user_by_chave(db, "UTO9").perfil = 1
        get_user_by_chave(db, "TP22").perfil = 2
        get_user_by_chave(db, "ZZ00").perfil = 0
        db.commit()

    with TestClient(app) as client:
        ensure_admin_session(client)
        rows = client.get("/api/admin/administrators")
        assert rows.status_code == 200
        rows_by_key = {row["chave"]: row for row in rows.json() if row["row_type"] == "admin"}
        assert rows_by_key["UTO9"]["perfil"] == 1
        assert rows_by_key["UTO9"]["can_revoke"] is True
        assert "TP22" not in rows_by_key
        assert "ZZ00" not in rows_by_key


def test_admin_password_reset_and_redefine_flow():
    with TestClient(app) as client:
        login_response = login_admin(client)
        assert login_response.status_code == 200

        request_new_admin = client.post(
            "/api/admin/auth/request-access",
            json={
                "chave": "AB12",
                "nome_completo": "Admin Auxiliar",
                "senha": "SenhaNova1",
            },
        )
        assert request_new_admin.status_code == 200

        admins_before = client.get("/api/admin/administrators")
        pending_row = next(row for row in admins_before.json() if row["chave"] == "AB12")
        approve_response = client.post(f"/api/admin/administrators/requests/{pending_row['id']}/approve")
        assert approve_response.status_code == 200

        reset_response = client.post("/api/admin/auth/request-password-reset", json={"chave": "AB12"})
        assert reset_response.status_code == 200
        assert "Outro administrador" in reset_response.json()["message"]

        blocked_login = login_admin(client, chave="AB12", senha="SenhaNova1")
        assert blocked_login.status_code == 403

        admins_after_reset = client.get("/api/admin/administrators")
        reset_row = next(row for row in admins_after_reset.json() if row["chave"] == "AB12")
        assert reset_row["status"] == "password_reset_requested"

        set_password_response = client.post(
            f"/api/admin/administrators/{reset_row['id']}/set-password",
            json={"nova_senha": "SenhaFinal2"},
        )
        assert set_password_response.status_code == 200

        client.post("/api/admin/auth/logout")
        relogin_new_password = login_admin(client, chave="AB12", senha="SenhaFinal2")
        assert relogin_new_password.status_code == 200


def test_admin_self_service_password_change_flow():
    with TestClient(app) as client:
        assert login_admin(client).status_code == 200

        request_new_admin = client.post(
            "/api/admin/auth/request-access",
            json={
                "chave": "SC12",
                "nome_completo": "Senha Própria",
                "senha": "SenhaIni1",
            },
        )
        assert request_new_admin.status_code == 200

        admins_before = client.get("/api/admin/administrators")
        pending_row = next(row for row in admins_before.json() if row["chave"] == "SC12")
        approve_response = client.post(f"/api/admin/administrators/requests/{pending_row['id']}/approve")
        assert approve_response.status_code == 200

        verify_ok = client.post(
            "/api/admin/auth/verify-current-password",
            json={"chave": "SC12", "senha_atual": "SenhaIni1"},
        )
        assert verify_ok.status_code == 200
        assert verify_ok.json()["valid"] is True

        verify_invalid = client.post(
            "/api/admin/auth/verify-current-password",
            json={"chave": "SC12", "senha_atual": "SenhaErr1"},
        )
        assert verify_invalid.status_code == 200
        assert verify_invalid.json()["valid"] is False

        invalid_change = client.post(
            "/api/admin/auth/change-password",
            json={
                "chave": "SC12",
                "senha_atual": "SenhaErr1",
                "nova_senha": "SenhaSC2",
                "confirmar_senha": "SenhaSC2",
            },
        )
        assert invalid_change.status_code == 401

        change_response = client.post(
            "/api/admin/auth/change-password",
            json={
                "chave": "SC12",
                "senha_atual": "SenhaIni1",
                "nova_senha": "SenhaSC2",
                "confirmar_senha": "SenhaSC2",
            },
        )
        assert change_response.status_code == 200
        assert "Senha alterada com sucesso." in change_response.json()["message"]

        client.post("/api/admin/auth/logout")
        blocked_old_login = login_admin(client, chave="SC12", senha="SenhaIni1")
        assert blocked_old_login.status_code == 401

        relogin_new_password = login_admin(client, chave="SC12", senha="SenhaSC2")
        assert relogin_new_password.status_code == 200

        events_response = client.get("/api/admin/events")
        assert events_response.status_code == 200
        events = events_response.json()
        assert any(
            event["action"] == "password"
            and event["status"] == "updated"
            and event["request_path"] == "/api/admin/auth/change-password"
            and "chave=SC12" in (event["details"] or "")
            for event in events
        )


def test_admin_event_audit_covers_new_auth_lifecycle():
    with TestClient(app) as client:
        first_request = client.post(
            "/api/admin/auth/request-access",
            json={
                "chave": "EV12",
                "nome_completo": "Evento Auditoria",
                "senha": "SenhaEvt1",
            },
        )
        assert first_request.status_code == 200

        duplicate_request = client.post(
            "/api/admin/auth/request-access",
            json={
                "chave": "EV12",
                "nome_completo": "Evento Auditoria",
                "senha": "SenhaEvt1",
            },
        )
        assert duplicate_request.status_code == 409

        assert login_admin(client).status_code == 200

        administrators = client.get("/api/admin/administrators")
        pending_row = next(row for row in administrators.json() if row["chave"] == "EV12")

        approve_response = client.post(f"/api/admin/administrators/requests/{pending_row['id']}/approve")
        assert approve_response.status_code == 200

        first_reset = client.post("/api/admin/auth/request-password-reset", json={"chave": "EV12"})
        assert first_reset.status_code == 200

        duplicate_reset = client.post("/api/admin/auth/request-password-reset", json={"chave": "EV12"})
        assert duplicate_reset.status_code == 409

        blocked_login = login_admin(client, chave="EV12", senha="SenhaEvt1")
        assert blocked_login.status_code == 403

        administrators_after_reset = client.get("/api/admin/administrators")
        reset_row = next(row for row in administrators_after_reset.json() if row["chave"] == "EV12")

        set_password = client.post(
            f"/api/admin/administrators/{reset_row['id']}/set-password",
            json={"nova_senha": "SenhaEvt2"},
        )
        assert set_password.status_code == 200

        revoke_response = client.post(f"/api/admin/administrators/{reset_row['id']}/revoke")
        assert revoke_response.status_code == 200

        events_response = client.get("/api/admin/events")
        assert events_response.status_code == 200
        events = events_response.json()

        assert any(
            event["action"] == "admin_request"
            and event["status"] == "pending"
            and event["request_path"] == "/api/admin/auth/request-access"
            and "chave=EV12" in (event["details"] or "")
            for event in events
        )
        assert any(
            event["action"] == "admin_request"
            and event["status"] == "failed"
            and event["http_status"] == 409
            and event["request_path"] == "/api/admin/auth/request-access"
            and "chave=EV12" in (event["details"] or "")
            for event in events
        )
        assert any(
            event["action"] == "admin_request"
            and event["status"] == "approved"
            and event["request_path"] == f"/api/admin/administrators/requests/{pending_row['id']}/approve"
            for event in events
        )
        assert any(
            event["action"] == "password"
            and event["status"] == "pending"
            and event["request_path"] == "/api/admin/auth/request-password-reset"
            for event in events
        )
        assert any(
            event["action"] == "password"
            and event["status"] == "failed"
            and event["http_status"] == 409
            and event["request_path"] == "/api/admin/auth/request-password-reset"
            for event in events
        )
        assert any(
            event["action"] == "login"
            and event["status"] == "blocked"
            and event["request_path"] == "/api/admin/auth/login"
            and "chave=EV12" in (event["details"] or "")
            for event in events
        )
        assert any(
            event["action"] == "password"
            and event["status"] == "updated"
            and event["request_path"] == f"/api/admin/administrators/{reset_row['id']}/set-password"
            for event in events
        )
        assert any(
            event["action"] == "admin_access"
            and event["status"] == "removed"
            and event["request_path"] == f"/api/admin/administrators/{reset_row['id']}/revoke"
            for event in events
        )


def test_transport_inline_auth_respects_user_profile():
    ensure_web_user_exists(chave="TRP2", nome="Transport Access")
    ensure_web_user_exists(chave="AD11", nome="Admin Only")
    with TestClient(app) as client:
        registration = register_web_password(client, chave="TRP2", senha="tp1234", ensure_user_exists=False)
        assert registration.status_code == 200
        admin_registration = register_web_password(client, chave="AD11", senha="ad1111", ensure_user_exists=False)
        assert admin_registration.status_code == 200

    with SessionLocal() as db:
        transport_user = get_user_by_chave(db, "TRP2")
        transport_user.perfil = 2
        admin_only_user = get_user_by_chave(db, "AD11")
        admin_only_user.perfil = 1
        db.commit()

    with TestClient(app) as client:
        denied = client.post("/api/transport/auth/verify", json={"chave": "AD11", "senha": "ad1111"})
        assert denied.status_code == 200
        assert denied.json()["authenticated"] is False
        assert "transport access" in denied.json()["message"].lower()
        assert denied.json()["message_key"] == "auth.noAccess"
        assert denied.json()["error_code"] == "transport_auth_access_denied"

        granted = client.post("/api/transport/auth/verify", json={"chave": "TRP2", "senha": "tp1234"})
        assert granted.status_code == 200
        assert granted.json()["authenticated"] is True
        assert granted.json()["user"]["perfil"] == 2
        assert granted.json()["message_key"] == "status.accessGranted"

        dashboard = client.get("/api/transport/dashboard")
        assert dashboard.status_code == 200

        logout = client.post("/api/transport/auth/logout")
        assert logout.status_code == 200

        dashboard_after_logout = client.get("/api/transport/dashboard")
        assert dashboard_after_logout.status_code == 401


def test_admin_locations_crud_and_mobile_catalog_sync():
    with TestClient(app) as client:
        ensure_admin_session(client)
        reset_location_settings = client.post(
            "/api/admin/locations/settings",
            json={
                "location_accuracy_threshold_meters": 30,
            },
        )
        assert reset_location_settings.status_code == 200

        create_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Base P80",
                "coordinates": build_rectangle_coordinates(1.255936, 103.611066),
                "projects": ["P80", "P82"],
                "tolerance_meters": 150,
            },
        )
        assert create_location.status_code == 200
        assert create_location.json()["ok"] is True

        locations = client.get("/api/admin/locations")
        assert locations.status_code == 200
        assert locations.json()["location_accuracy_threshold_meters"] == 30
        base_p80 = next(row for row in locations.json()["items"] if row["local"] == "Base P80")
        assert base_p80["coordinates"] == build_rectangle_coordinates(1.255936, 103.611066)
        assert base_p80["projects"] == ["P80", "P82"]
        assert base_p80["tolerance_meters"] == 150

        update_location_settings = client.post(
            "/api/admin/locations/settings",
            json={
                "location_accuracy_threshold_meters": 45,
            },
        )
        assert update_location_settings.status_code == 200
        assert update_location_settings.json()["ok"] is True
        assert update_location_settings.json()["location_accuracy_threshold_meters"] == 45

        events = client.get("/api/admin/events")
        assert events.status_code == 200
        location_settings_event = next(
            event
            for event in events.json()
            if event["action"] == "location_config" and event["request_path"] == "/api/admin/locations/settings"
        )
        assert location_settings_event["chave"] == ADMIN_LOGIN_CHAVE
        assert location_settings_event["message"] == (
            "O valor do erro máximo para considerar a coordenada do usuário foi ajustado para 45 metros."
        )

        update_location = client.post(
            "/api/admin/locations",
            json={
                "location_id": base_p80["id"],
                "local": "Base P80",
                "coordinates": build_rectangle_coordinates(1.255936, 103.611066, latitude_delta=0.004065, longitude_delta=0.000936),
                "projects": ["P82"],
                "tolerance_meters": 250,
            },
        )
        assert update_location.status_code == 200
        assert update_location.json()["ok"] is True

        updated_locations = client.get("/api/admin/locations")
        assert updated_locations.status_code == 200
        updated_base_p80 = next(row for row in updated_locations.json()["items"] if row["local"] == "Base P80")
        assert updated_base_p80["coordinates"] == build_rectangle_coordinates(
            1.255936,
            103.611066,
            latitude_delta=0.004065,
            longitude_delta=0.000936,
        )
        assert updated_base_p80["projects"] == ["P82"]
        assert updated_base_p80["latitude"] == 1.255936
        assert updated_base_p80["longitude"] == 103.611066

        mobile_catalog = client.get("/api/mobile/locations", headers=MOBILE_HEADERS)
        assert mobile_catalog.status_code == 200
        assert mobile_catalog.json()["location_accuracy_threshold_meters"] == 45
        assert "mixed_zone_interval_minutes" not in mobile_catalog.json()
        assert "coordinate_update_frequency_headers" not in mobile_catalog.json()
        assert "coordinate_update_frequency_rows" not in mobile_catalog.json()
        synced_row = next(row for row in mobile_catalog.json()["items"] if row["local"] == "Base P80")
        assert synced_row["tolerance_meters"] == 250
        assert synced_row["coordinates"] == build_rectangle_coordinates(
            1.255936,
            103.611066,
            latitude_delta=0.004065,
            longitude_delta=0.000936,
        )
        assert synced_row["latitude"] == 1.255936
        assert synced_row["longitude"] == 103.611066

        remove_location = client.delete(f"/api/admin/locations/{base_p80['id']}")
        assert remove_location.status_code == 200
        assert remove_location.json()["ok"] is True


def test_admin_project_auto_checkout_distance_settings_sync_to_mobile_catalog():
    with TestClient(app) as client:
        ensure_admin_session(client)

        initial_settings = client.get("/api/admin/locations/auto-checkout-distances")
        assert initial_settings.status_code == 200
        initial_items = initial_settings.json()["items"]
        assert next(item for item in initial_items if item["project_name"] == "P80")["minimum_checkout_distance_meters"] == 2000

        update_settings = client.post(
            "/api/admin/locations/auto-checkout-distances",
            json={
                "items": [
                    {"project_name": "p80", "minimum_checkout_distance_meters": 1400},
                    {"project_name": "P82", "minimum_checkout_distance_meters": 2600},
                ]
            },
        )
        assert update_settings.status_code == 200
        payload = update_settings.json()
        assert payload["ok"] is True
        assert next(item for item in payload["items"] if item["project_name"] == "P80")["minimum_checkout_distance_meters"] == 1400
        assert next(item for item in payload["items"] if item["project_name"] == "P82")["minimum_checkout_distance_meters"] == 2600

        mobile_catalog = client.get("/api/mobile/locations", headers=MOBILE_HEADERS)
        assert mobile_catalog.status_code == 200
        mobile_thresholds = mobile_catalog.json()["minimum_checkout_distance_meters_by_project"]
        assert mobile_thresholds["P80"] == 1400
        assert mobile_thresholds["P82"] == 2600
        assert mobile_thresholds["P83"] == 2000


def test_admin_locations_audit_endpoint_lists_flagged_rows_and_can_include_valid_rows():
    with TestClient(app) as client:
        ensure_admin_session(client)
        ensure_project_exists("P95")

        create_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Audit Base P95",
                "coordinates": build_rectangle_coordinates(1.275936, 103.621066),
                "projects": ["P95"],
                "tolerance_meters": 150,
            },
        )
        assert create_location.status_code == 200

    with SessionLocal() as db:
        timestamp = now_sgt()
        db.add(
            ManagedLocation(
                local="Audit Legacy P95",
                latitude=1.275936,
                longitude=103.621066,
                coordinates_json=None,
                projects_json=dump_location_projects(["P95"]),
                tolerance_meters=150,
                created_at=timestamp,
                updated_at=timestamp,
            )
        )
        db.commit()

    with TestClient(app) as client:
        ensure_admin_session(client)

        audit_response = client.get("/api/admin/locations/audit")
        assert audit_response.status_code == 200
        payload = audit_response.json()
        assert payload["summary"]["locations_with_errors"] >= 1
        assert any(row["local"] == "Audit Legacy P95" for row in payload["rows"])
        assert all(row["local"] != "Audit Base P95" for row in payload["rows"])

        audit_with_valid = client.get("/api/admin/locations/audit", params={"include_valid": "true"})
        assert audit_with_valid.status_code == 200
        payload_with_valid = audit_with_valid.json()
        assert any(row["local"] == "Audit Base P95" for row in payload_with_valid["rows"])
        legacy_row = next(row for row in payload_with_valid["rows"] if row["local"] == "Audit Legacy P95")
        issue_codes = {issue["code"] for issue in legacy_row["issues"]}
        assert "too_few_coordinates" in issue_codes


def test_admin_locations_list_contract_preserves_expected_shape():
    with TestClient(app) as client:
        ensure_admin_session(client)
        ensure_project_exists("P98")

        create_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Contract Base P98",
                "coordinates": build_rectangle_coordinates(1.295936, 103.641066),
                "projects": ["P98"],
                "tolerance_meters": 180,
            },
        )
        assert create_location.status_code == 200

        locations = client.get("/api/admin/locations")
        assert locations.status_code == 200
        payload = locations.json()
        assert set(payload.keys()) == {"items", "location_accuracy_threshold_meters"}

        location_row = next(row for row in payload["items"] if row["local"] == "Contract Base P98")
        assert set(location_row.keys()) == {
            "id",
            "local",
            "latitude",
            "longitude",
            "coordinates",
            "projects",
            "tolerance_meters",
        }
        assert location_row["coordinates"] == build_rectangle_coordinates(1.295936, 103.641066)
        assert location_row["projects"] == ["P98"]
        assert location_row["tolerance_meters"] == 180


def test_web_locations_catalog_includes_accuracy_threshold_for_lifecycle_capture():
    ensure_project_exists("P82")
    web_key = make_test_key("W")
    ensure_web_user_exists(chave=web_key, projeto="P82", nome="Web Location Catalog")

    with TestClient(app) as client:
        ensure_admin_session(client)

        update_location_settings = client.post(
            "/api/admin/locations/settings",
            json={
                "location_accuracy_threshold_meters": 25,
            },
        )
        assert update_location_settings.status_code == 200

        create_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Web Catalog Base P82",
                "coordinates": build_rectangle_coordinates(1.265936, 103.621066),
                "projects": ["P82"],
                "tolerance_meters": 150,
            },
        )
        assert create_location.status_code == 200

    with TestClient(app) as client:
        registered = register_web_password(client, chave=web_key, senha="web123", projeto="P82", ensure_user_exists=False)
        assert registered.status_code == 200

        locations = client.get("/api/web/check/locations")
        assert locations.status_code == 200
        payload = locations.json()
        assert set(payload.keys()) == {"items", "location_accuracy_threshold_meters", "mixed_zone_interval_minutes"}
        assert payload["location_accuracy_threshold_meters"] == 25
        assert payload["mixed_zone_interval_minutes"] == 30
        assert "Web Catalog Base P82" in payload["items"]


def test_admin_location_save_contract_returns_simple_action_response():
    with TestClient(app) as client:
        ensure_admin_session(client)
        ensure_project_exists("P99")

        create_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Contract Save P99",
                "coordinates": build_rectangle_coordinates(1.305936, 103.651066),
                "projects": ["P99"],
                "tolerance_meters": 190,
            },
        )
        assert create_location.status_code == 200
        assert create_location.json() == {
            "ok": True,
            "message": "Localizacao salva com sucesso.",
        }


def test_admin_location_validation_failure_is_logged_in_events():
    with TestClient(app) as client:
        ensure_admin_session(client)

        create_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Audit Fail P96",
                "coordinates": [
                    {"latitude": 1.255800, "longitude": 103.611000},
                    {"latitude": 1.256100, "longitude": 103.611400},
                    {"latitude": 1.255800, "longitude": 103.611400},
                    {"latitude": 1.256100, "longitude": 103.611000},
                ],
                "projects": ["P80"],
                "tolerance_meters": 150,
            },
        )
        assert create_location.status_code == 422

        events = client.get("/api/admin/events")
        assert events.status_code == 200
        failed_event = next(
            event
            for event in events.json()
            if event["action"] == "location"
            and event["status"] == "failed"
            and event["local"] == "Audit Fail P96"
            and event["request_path"] == "/api/admin/locations"
        )
        assert failed_event["message"] == "Location validation failed via admin"
        assert "updated_by=HR70" in failed_event["details"]
        assert "coordinate_count=4" in failed_event["details"]
        assert "validation_errors=A localizacao precisa formar um poligono valido sem auto-interseccao" in failed_event["details"]


def test_admin_location_update_logs_geometry_change_details():
    with TestClient(app) as client:
        ensure_admin_session(client)
        ensure_project_exists("P97")

        create_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Audit Geometry P97",
                "coordinates": build_rectangle_coordinates(1.285936, 103.631066),
                "projects": ["P97"],
                "tolerance_meters": 150,
            },
        )
        assert create_location.status_code == 200

        locations = client.get("/api/admin/locations")
        assert locations.status_code == 200
        geometry_row = next(row for row in locations.json()["items"] if row["local"] == "Audit Geometry P97")

        update_location = client.post(
            "/api/admin/locations",
            json={
                "location_id": geometry_row["id"],
                "local": "Audit Geometry P97",
                "coordinates": build_rectangle_coordinates(
                    1.285936,
                    103.631066,
                    latitude_delta=0.00035,
                    longitude_delta=0.00025,
                ),
                "projects": ["P97"],
                "tolerance_meters": 240,
            },
        )
        assert update_location.status_code == 200

        events = client.get("/api/admin/events")
        assert events.status_code == 200
        updated_event = next(
            event
            for event in events.json()
            if event["action"] == "location"
            and event["status"] == "updated"
            and event["local"] == "Audit Geometry P97"
        )
        assert updated_event["message"] == "Location geometry updated via admin"
        assert "geometry_changed=yes" in updated_event["details"]
        assert "previous_tolerance_meters=150" in updated_event["details"]
        assert "tolerance_meters=240" in updated_event["details"]
        assert "previous_coordinates=" in updated_event["details"]
        assert "coordinates=" in updated_event["details"]


def test_web_location_options_contract_returns_only_name_list():
    project_name = f"Q{uuid.uuid4().hex[:3].upper()}"
    user_key = make_test_key("W")
    location_name = f"Lista Contrato {project_name}"

    with TestClient(app) as client:
        ensure_admin_session(client)
        ensure_project_exists(project_name)
        auth_response = register_web_password(client, chave=user_key, senha="loc123", projeto=project_name)
        assert auth_response.status_code == 200

        create_location = client.post(
            "/api/admin/locations",
            json={
                "local": location_name,
                "coordinates": build_rectangle_coordinates(1.315936, 103.661066),
                "projects": [project_name],
                "tolerance_meters": 200,
            },
        )
        assert create_location.status_code == 200

        locations_response = client.get("/api/web/check/locations")
        assert locations_response.status_code == 200
        payload = locations_response.json()
        assert set(payload.keys()) == {"items", "location_accuracy_threshold_meters", "mixed_zone_interval_minutes"}
        assert payload["items"] == [location_name]
        assert isinstance(payload["location_accuracy_threshold_meters"], int)
        assert isinstance(payload["mixed_zone_interval_minutes"], int)
        assert all(isinstance(item, str) for item in payload["items"])


def test_web_location_match_contract_preserves_expected_response_shape():
    with TestClient(app) as client:
        ensure_admin_session(client)
        ensure_project_exists("Q02")
        auth_response = register_web_password(client, chave="WQ02", senha="loc123", projeto="Q02")
        assert auth_response.status_code == 200

        create_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Match Contrato Q02",
                "coordinates": build_rectangle_coordinates(1.325936, 103.671066),
                "projects": ["Q02"],
                "tolerance_meters": 210,
            },
        )
        assert create_location.status_code == 200

        update_settings = client.post(
            "/api/admin/locations/settings",
            json={"location_accuracy_threshold_meters": 25},
        )
        assert update_settings.status_code == 200

        match_response = client.post(
            "/api/web/check/location",
            json={
                "latitude": 1.325936,
                "longitude": 103.671066,
                "accuracy_meters": 8,
            },
        )
        assert match_response.status_code == 200
        payload = match_response.json()
        assert set(payload.keys()) == {
            "matched",
            "resolved_local",
            "label",
            "status",
            "message",
            "accuracy_meters",
            "accuracy_threshold_meters",
            "minimum_checkout_distance_meters",
            "nearest_workplace_distance_meters",
        }
        assert payload["matched"] is True
        assert payload["resolved_local"] == "Match Contrato Q02"
        assert payload["label"] == "Match Contrato Q02"
        assert payload["status"] == "matched"
        assert payload["accuracy_meters"] == 8
        assert payload["accuracy_threshold_meters"] == 25
        assert payload["minimum_checkout_distance_meters"] == 2000


def test_admin_locations_allow_same_name_with_different_coordinates():
    with TestClient(app) as client:
        ensure_admin_session(client)

        first_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Base Compartilhada",
                "coordinates": build_rectangle_coordinates(1.255936, 103.611066),
                "projects": ["P80"],
                "tolerance_meters": 150,
            },
        )
        assert first_location.status_code == 200
        assert first_location.json()["ok"] is True

        second_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Base Compartilhada",
                "coordinates": build_rectangle_coordinates(1.266001, 103.622002),
                "projects": ["P82"],
                "tolerance_meters": 220,
            },
        )
        assert second_location.status_code == 200
        assert second_location.json()["ok"] is True

        locations = client.get("/api/admin/locations")
        assert locations.status_code == 200

        duplicated_rows = [row for row in locations.json()["items"] if row["local"] == "Base Compartilhada"]
        assert len(duplicated_rows) == 2
        assert {row["projects"][0] for row in duplicated_rows} == {"P80", "P82"}
        assert {
            (row["latitude"], row["longitude"], row["tolerance_meters"])
            for row in duplicated_rows
        } == {
            (1.255936, 103.611066, 150),
            (1.266001, 103.622002, 220),
        }


def test_admin_location_rejects_zero_tolerance():
    with TestClient(app) as client:
        ensure_admin_session(client)

        create_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Base Tolerancia Zero",
                "coordinates": build_rectangle_coordinates(1.255936, 103.611066),
                "projects": ["P80"],
                "tolerance_meters": 0,
            },
        )
        assert create_location.status_code == 422
        assert "greater than or equal to 1" in create_location.text


def test_admin_location_rejects_payload_with_fewer_than_three_coordinates():
    with TestClient(app) as client:
        ensure_admin_session(client)

        create_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Base Invalida",
                "coordinates": [
                    {"latitude": 1.255936, "longitude": 103.611066},
                    {"latitude": 1.256136, "longitude": 103.611266},
                ],
                "projects": ["P80"],
                "tolerance_meters": 150,
            },
        )
        assert create_location.status_code == 422
        assert "Informe ao menos 3 coordenadas distintas" in create_location.text


def test_admin_location_rejects_self_intersecting_polygon():
    with TestClient(app) as client:
        ensure_admin_session(client)

        create_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Base Laco",
                "coordinates": [
                    {"latitude": 1.255800, "longitude": 103.611000},
                    {"latitude": 1.256100, "longitude": 103.611400},
                    {"latitude": 1.255800, "longitude": 103.611400},
                    {"latitude": 1.256100, "longitude": 103.611000},
                ],
                "projects": ["P80"],
                "tolerance_meters": 150,
            },
        )
        assert create_location.status_code == 422
        assert "poligono valido sem auto-interseccao" in create_location.text


def test_admin_location_rejects_zero_area_polygon():
    with TestClient(app) as client:
        ensure_admin_session(client)

        create_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Base Colinear",
                "coordinates": [
                    {"latitude": 1.255800, "longitude": 103.611000},
                    {"latitude": 1.255800, "longitude": 103.611200},
                    {"latitude": 1.255800, "longitude": 103.611400},
                ],
                "projects": ["P80"],
                "tolerance_meters": 150,
            },
        )
        assert create_location.status_code == 422
        assert "area valida" in create_location.text


def test_admin_location_accepts_triangle_polygon_and_persists_vertices():
    with TestClient(app) as client:
        ensure_admin_session(client)
        ensure_project_exists("P96")

        create_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Triangulo P96",
                "coordinates": [
                    {"latitude": 1.255800, "longitude": 103.611000},
                    {"latitude": 1.256100, "longitude": 103.611100},
                    {"latitude": 1.255900, "longitude": 103.611350},
                ],
                "projects": ["P96"],
                "tolerance_meters": 120,
            },
        )
        assert create_location.status_code == 200

        locations = client.get("/api/admin/locations")
        assert locations.status_code == 200
        triangle_row = next(row for row in locations.json()["items"] if row["local"] == "Triangulo P96")

        assert triangle_row["tolerance_meters"] == 120
        assert len(triangle_row["coordinates"]) == 3
        assert triangle_row["coordinates"][0] == {"latitude": 1.2558, "longitude": 103.611}
        assert triangle_row["coordinates"][1] == {"latitude": 1.2561, "longitude": 103.6111}
        assert triangle_row["coordinates"][2] == {"latitude": 1.2559, "longitude": 103.61135}


def test_admin_location_update_persists_reordered_vertices():
    with TestClient(app) as client:
        ensure_admin_session(client)
        ensure_project_exists("P97R")

        original_coordinates = build_rectangle_coordinates(1.295936, 103.641066)
        create_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Reorder Base P97R",
                "coordinates": original_coordinates,
                "projects": ["P97R"],
                "tolerance_meters": 150,
            },
        )
        assert create_location.status_code == 200

        locations = client.get("/api/admin/locations")
        assert locations.status_code == 200
        reorder_row = next(row for row in locations.json()["items"] if row["local"] == "Reorder Base P97R")

        reordered_coordinates = original_coordinates[1:] + original_coordinates[:1]
        update_location = client.post(
            "/api/admin/locations",
            json={
                "location_id": reorder_row["id"],
                "local": "Reorder Base P97R",
                "coordinates": reordered_coordinates,
                "projects": ["P97R"],
                "tolerance_meters": 150,
            },
        )
        assert update_location.status_code == 200

        refreshed_locations = client.get("/api/admin/locations")
        assert refreshed_locations.status_code == 200
        updated_row = next(row for row in refreshed_locations.json()["items"] if row["local"] == "Reorder Base P97R")

        assert updated_row["coordinates"] == reordered_coordinates


def test_admin_location_update_allows_existing_detached_project_assignments():
    with TestClient(app) as client:
        ensure_admin_session(client)

        create_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Detached Project Base",
                "coordinates": build_rectangle_coordinates(1.255936, 103.611066),
                "projects": ["P80"],
                "tolerance_meters": 150,
            },
        )
        assert create_location.status_code == 200

        locations = client.get("/api/admin/locations")
        assert locations.status_code == 200
        detached_location = next(row for row in locations.json()["items"] if row["local"] == "Detached Project Base")

    with SessionLocal() as db:
        location = db.get(ManagedLocation, detached_location["id"])
        assert location is not None
        location.projects_json = dump_location_projects(["LEGACY"])
        db.commit()

    with TestClient(app) as client:
        ensure_admin_session(client)

        update_location = client.post(
            "/api/admin/locations",
            json={
                "location_id": detached_location["id"],
                "local": "Detached Project Base Updated",
                "coordinates": build_rectangle_coordinates(1.255936, 103.611066),
                "projects": ["LEGACY"],
                "tolerance_meters": 200,
            },
        )
        assert update_location.status_code == 200
        assert update_location.json()["ok"] is True

        updated_locations = client.get("/api/admin/locations")
        assert updated_locations.status_code == 200
        updated_row = next(row for row in updated_locations.json()["items"] if row["id"] == detached_location["id"])
        assert updated_row["local"] == "Detached Project Base Updated"
        assert updated_row["projects"] == ["LEGACY"]
        assert updated_row["tolerance_meters"] == 200


def test_removing_project_reassigns_linked_location_projects():
    with TestClient(app) as client:
        ensure_admin_session(client)
        ensure_project_exists("Q03")

        create_location = client.post(
            "/api/admin/locations",
            json={
                "local": "Project Reassign Base",
                "coordinates": build_rectangle_coordinates(1.255936, 103.611066),
                "projects": ["P80", "Q03"],
                "tolerance_meters": 150,
            },
        )
        assert create_location.status_code == 200

        projects_response = client.get("/api/admin/projects")
        assert projects_response.status_code == 200
        fallback_project = next(
            project["name"]
            for project in sorted(projects_response.json(), key=lambda row: row["name"])
            if project["name"] != "Q03"
        )
        project_q03 = next(project for project in projects_response.json() if project["name"] == "Q03")

        remove_project = client.delete(f"/api/admin/projects/{project_q03['id']}")
        assert remove_project.status_code == 200
        assert remove_project.json()["ok"] is True

        updated_locations = client.get("/api/admin/locations")
        assert updated_locations.status_code == 200
        updated_row = next(row for row in updated_locations.json()["items"] if row["local"] == "Project Reassign Base")
        expected_projects = ["P80"]
        if fallback_project != "P80":
            expected_projects.append(fallback_project)
        assert updated_row["projects"] == expected_projects


def test_mobile_forms_submit_uses_default_and_custom_local():
    with TestClient(app) as client:
        first_event = client.post(
            "/api/mobile/events/forms-submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "LX01",
                "projeto": "P83",
                "action": "checkin",
                "informe": "normal",
                "event_time": now_sgt().isoformat(),
                "client_event_id": f"mobile-{uuid.uuid4().hex}",
            },
        )
        assert first_event.status_code == 200
        assert first_event.json()["ok"] is True

        with SessionLocal() as db:
            user = get_user_by_chave(db, "LX01")
            queued = db.execute(
                select(FormsSubmission)
                .where(FormsSubmission.chave == "LX01")
                .order_by(FormsSubmission.id.desc())
            ).scalars().first()
            sync_event = db.execute(
                select(UserSyncEvent)
                .where(UserSyncEvent.chave == "LX01")
                .order_by(UserSyncEvent.id.desc())
            ).scalars().first()

            assert user.local == "Aplicativo"
            assert queued is not None and queued.local == "Aplicativo"
            assert sync_event is not None and sync_event.local == "Aplicativo"

        second_event = client.post(
            "/api/mobile/events/forms-submit",
            headers=MOBILE_HEADERS,
            json={
                "chave": "LX01",
                "projeto": "P83",
                "action": "checkout",
                "local": "Base P80",
                "informe": "retroativo",
                "event_time": now_sgt().isoformat(),
                "client_event_id": f"mobile-{uuid.uuid4().hex}",
            },
        )
        assert second_event.status_code == 200
        assert second_event.json()["ok"] is True

        with SessionLocal() as db:
            user = get_user_by_chave(db, "LX01")
            queued = db.execute(
                select(FormsSubmission)
                .where(FormsSubmission.chave == "LX01")
                .order_by(FormsSubmission.id.desc())
            ).scalars().first()
            sync_event = db.execute(
                select(UserSyncEvent)
                .where(UserSyncEvent.chave == "LX01")
                .order_by(UserSyncEvent.id.desc())
            ).scalars().first()

            assert user.local == "Base P80"
            assert queued is not None and queued.local == "Base P80"
            assert sync_event is not None and sync_event.local == "Base P80"


def _transport_ai_api_regression_timestamp() -> datetime:
    return datetime(2026, 5, 6, 8, 0, 0, tzinfo=ZoneInfo(settings.tz_name))


def _configure_transport_ai_api_regression_runtime(monkeypatch) -> None:
    """Run transport AI API regressions with deterministic planning and the fake route provider."""

    monkeypatch.setattr(settings, "transport_ai_enabled", True)
    monkeypatch.setattr(settings, "transport_ai_agent_mode", "deterministic")
    monkeypatch.setattr(settings, "transport_ai_route_provider", "fake")
    monkeypatch.setattr(settings, "transport_ai_operational_approval_evidence", "phase8-loadtest-2026-05-05")
    monkeypatch.setattr(settings, "transport_ai_max_concurrent_runs", 1)
    monkeypatch.setattr(settings, "here_api_key", "test-here-api-key")


def _create_transport_ai_api_regression_vehicle_candidate(
    db,
    *,
    plate: str,
    service_date: date,
) -> tuple[Vehicle, list[TransportVehicleSchedule]]:
    timestamp = _transport_ai_api_regression_timestamp()
    vehicle = Vehicle(
        placa=plate,
        tipo="carro",
        color="White",
        lugares=4,
        tolerance=0,
        service_scope="extra",
    )
    db.add(vehicle)
    db.flush()

    schedule = TransportVehicleSchedule(
        vehicle_id=vehicle.id,
        service_scope="extra",
        route_kind="home_to_work",
        recurrence_kind="single_date",
        service_date=service_date,
        weekday=None,
        departure_time="07:30",
        is_active=True,
        created_at=timestamp,
        updated_at=timestamp,
    )
    db.add(schedule)
    db.flush()
    return vehicle, [schedule]


def _create_transport_ai_api_regression_fixture(
    db,
    *,
    service_date: date,
    seed_existing_assignment: bool,
) -> dict[str, object]:
    """Build a documented AI API fixture that mirrors the proven isolated transport AI test data."""

    from sistema.app.models import MobileAppSettings
    from sistema.app.services.transport_ai_runs import ensure_transport_ai_actor_admin_user

    timestamp = _transport_ai_api_regression_timestamp()
    scenario_token = uuid.uuid4().hex[:6].upper()
    settings_context = {
        "previous": clone_transport_settings_payload(db),
        "created_currency_code": None,
    }

    settings_row = db.get(MobileAppSettings, 1)
    if settings_row is None:
        settings_row = MobileAppSettings(
            id=1,
            created_at=timestamp,
            updated_at=timestamp,
        )
        db.add(settings_row)

    settings_row.transport_work_to_home_time = "16:45"
    settings_row.transport_last_update_time = "16:00"
    settings_row.transport_default_car_seats = 4
    settings_row.transport_default_minivan_seats = 6
    settings_row.transport_default_van_seats = 10
    settings_row.transport_default_bus_seats = 40
    settings_row.transport_default_tolerance_minutes = 5
    settings_row.transport_price_currency_code = "SGD"
    settings_row.transport_price_rate_unit = "day"
    settings_row.transport_default_car_price = 15
    settings_row.transport_default_minivan_price = 30
    settings_row.transport_default_van_price = 45
    settings_row.transport_default_bus_price = 70
    settings_row.updated_at = timestamp
    db.flush()

    project = Project(
        name=make_test_project_name(f"AI{scenario_token[:4]}"),
        country_code="SG",
        country_name="Singapore",
        timezone_name="Asia/Singapore",
        address="1 Marina Boulevard",
        zip_code="018989",
    )
    db.add(project)
    db.flush()

    rider = User(
        rfid=None,
        chave=make_test_key("A"),
        senha=None,
        perfil=0,
        admin_monitored_projects_json=None,
        nome=f"Transport AI Regression Rider {scenario_token}",
        projeto=project.name,
        workplace=None,
        vehicle_id=None,
        placa=None,
        end_rua="10 Bayfront Avenue",
        zip="018956",
        email=None,
        local=None,
        checkin=None,
        time=None,
        last_active_at=timestamp,
        inactivity_days=0,
    )
    db.add(rider)
    db.flush()

    request_row = TransportRequest(
        user_id=rider.id,
        request_kind="extra",
        recurrence_kind="single_date",
        requested_time="08:00",
        selected_weekdays_json=None,
        single_date=service_date,
        created_via="admin",
        status="active",
        created_at=timestamp,
        updated_at=timestamp,
        cancelled_at=None,
    )
    db.add(request_row)
    db.flush()

    fixture_bundle: dict[str, object] = {
        "project_ids": [project.id],
        "user_ids": [rider.id],
        "request_ids": [request_row.id],
        "vehicle_ids": [],
        "schedule_ids": [],
        "projects": {
            "primary": {
                "id": project.id,
                "name": project.name,
            }
        },
        "requests": {
            "primary": {
                "id": request_row.id,
                "chave": rider.chave,
                "nome": rider.nome,
                "projeto": rider.projeto,
                "service_date": service_date.isoformat(),
            }
        },
        "settings_context": settings_context,
        "assignment_id": None,
        "seed_vehicle_id": None,
    }

    if not seed_existing_assignment:
        return fixture_bundle

    actor_admin = ensure_transport_ai_actor_admin_user(
        db,
        chave=ADMIN_LOGIN_CHAVE,
        nome_completo=str(settings.bootstrap_admin_name),
        ensured_at=timestamp,
    )
    vehicle, schedules = _create_transport_ai_api_regression_vehicle_candidate(
        db,
        plate=f"AI{scenario_token[:4]}01",
        service_date=service_date,
    )
    assignment = TransportAssignment(
        request_id=request_row.id,
        service_date=service_date,
        route_kind="home_to_work",
        vehicle_id=vehicle.id,
        status="confirmed",
        response_message="confirmed-for-transport-ai-api-regression",
        acknowledged_by_user=False,
        acknowledged_at=None,
        assigned_by_admin_id=actor_admin.id,
        created_at=timestamp,
        updated_at=timestamp,
        notified_at=None,
    )
    db.add(assignment)
    db.flush()

    fixture_bundle["vehicle_ids"] = [vehicle.id]
    fixture_bundle["schedule_ids"] = [schedule.id for schedule in schedules]
    fixture_bundle["assignment_id"] = assignment.id
    fixture_bundle["seed_vehicle_id"] = vehicle.id
    return fixture_bundle


def _cleanup_transport_ai_api_regression_fixture(
    db,
    fixture_bundle: dict[str, object] | None,
    *,
    run_key: str | None,
) -> None:
    """Delete persisted AI artifacts before restoring the shared transport planning fixtures."""

    from sistema.app.models import TransportAIAppliedRouteStop, TransportAIRun, TransportAISuggestion

    if fixture_bundle is None:
        db.commit()
        transport_reevaluation_module.clear_transport_reevaluation_events()
        return

    bundle = dict(fixture_bundle)
    request_ids = [int(value) for value in bundle.get("request_ids") or []]
    static_vehicle_ids = {int(value) for value in bundle.get("vehicle_ids") or []}
    dynamic_vehicle_ids: set[int] = set()

    if request_ids:
        assignments = db.execute(
            select(TransportAssignment).where(TransportAssignment.request_id.in_(request_ids))
        ).scalars().all()
        for assignment in assignments:
            if assignment.vehicle_id is None or assignment.vehicle_id in static_vehicle_ids:
                continue
            dynamic_vehicle_ids.add(int(assignment.vehicle_id))

    if dynamic_vehicle_ids:
        bundle["vehicle_ids"] = [*bundle.get("vehicle_ids", []), *sorted(dynamic_vehicle_ids)]
        extra_schedule_ids = [
            schedule.id
            for schedule in db.execute(
                select(TransportVehicleSchedule).where(TransportVehicleSchedule.vehicle_id.in_(sorted(dynamic_vehicle_ids)))
            ).scalars().all()
        ]
        bundle["schedule_ids"] = [*bundle.get("schedule_ids", []), *extra_schedule_ids]

    normalized_run_key = str(run_key or "").strip()
    if normalized_run_key:
        run = db.execute(
            select(TransportAIRun).where(TransportAIRun.run_key == normalized_run_key)
        ).scalar_one_or_none()
        if run is not None:
            suggestions = db.execute(
                select(TransportAISuggestion).where(TransportAISuggestion.run_id == run.id)
            ).scalars().all()
            for suggestion in suggestions:
                for applied_stop in db.execute(
                    select(TransportAIAppliedRouteStop).where(TransportAIAppliedRouteStop.suggestion_id == suggestion.id)
                ).scalars().all():
                    db.delete(applied_stop)
                db.delete(suggestion)
            db.delete(run)

    cleanup_transport_planning_fixture_bundle(db, bundle)
    transport_reevaluation_module.clear_transport_reevaluation_events()


def _isolate_transport_ai_api_regression_requests(
    db,
    *,
    keep_request_ids: list[int],
) -> None:
    timestamp = _transport_ai_api_regression_timestamp()
    preserved_request_ids = {int(value) for value in keep_request_ids}

    for request_row in db.execute(select(TransportRequest)).scalars().all():
        if request_row.id in preserved_request_ids or request_row.status != "active":
            continue
        request_row.status = "cancelled"
        request_row.cancelled_at = timestamp
        request_row.updated_at = timestamp


def _start_transport_ai_api_regression_run(
    client: TestClient,
    *,
    service_date: date,
) -> tuple[dict[str, object], dict[str, object]]:
    start_response = client.post(
        "/api/transport/ai/route-calculations",
        json={
            "service_date": service_date.isoformat(),
            "route_kind": "home_to_work",
            "earliest_boarding_time": "06:50",
            "arrival_at_work_time": "07:45",
        },
    )
    assert start_response.status_code == 201, start_response.text
    start_payload = start_response.json()
    assert start_payload["ok"] is True
    assert start_payload["status"] == "proposed"
    assert start_payload["suggestion_ready"] is True
    assert start_payload["run_key"]
    assert start_payload["suggestion_key"]

    status_response = client.get(
        f"/api/transport/ai/route-calculations/{start_payload['run_key']}"
    )
    assert status_response.status_code == 200, status_response.text
    status_payload = status_response.json()
    assert status_payload["ok"] is True
    assert status_payload["status"] == "proposed"
    assert status_payload["suggestion_ready"] is True
    assert status_payload["suggestion_key"] == start_payload["suggestion_key"]
    assert status_payload["suggestion"] is not None
    assert status_payload["suggestion"]["status"] == "shown"
    return start_payload, status_payload


def test_transport_ai_api_flow_start_suggestion_save_latest_apply(monkeypatch):
    from sistema.app.models import TransportAIAppliedRouteStop, TransportAIRun, TransportAISuggestion

    _configure_transport_ai_api_regression_runtime(monkeypatch)
    transport_reevaluation_module.clear_transport_reevaluation_events()

    service_date = date(2035, 2, 12)
    fixture_bundle: dict[str, object] | None = None
    run_key: str | None = None

    try:
        with SessionLocal() as db:
            fixture_bundle = _create_transport_ai_api_regression_fixture(
                db,
                service_date=service_date,
                seed_existing_assignment=False,
            )
            db.commit()

        request_id = int(fixture_bundle["requests"]["primary"]["id"])

        with SessionLocal() as db:
            _isolate_transport_ai_api_regression_requests(db, keep_request_ids=[request_id])
            db.commit()

        with TestClient(app) as admin_client:
            ensure_admin_session(admin_client)
            start_payload, status_payload = _start_transport_ai_api_regression_run(
                admin_client,
                service_date=service_date,
            )
            run_key = str(start_payload["run_key"])
            suggestion_key = str(status_payload["suggestion_key"])

            save_response = admin_client.post(f"/api/transport/ai/suggestions/{suggestion_key}/save")
            assert save_response.status_code == 200, save_response.text
            save_payload = save_response.json()
            assert save_payload["status"] == "saved"
            assert save_payload["suggestion"]["status"] == "saved"
            assert save_payload["can_apply"] is True
            assert save_payload["can_cancel_restore"] is True

            latest_response = admin_client.get(
                "/api/transport/ai/suggestions/latest",
                params={"service_date": service_date.isoformat(), "route_kind": "home_to_work"},
            )
            assert latest_response.status_code == 200, latest_response.text
            latest_payload = latest_response.json()
            assert latest_payload["run_key"] == run_key
            assert latest_payload["suggestion_key"] == suggestion_key
            assert latest_payload["status"] == "saved"
            assert latest_payload["suggestion"]["status"] == "saved"

            apply_response = admin_client.post(f"/api/transport/ai/suggestions/{suggestion_key}/apply")
            assert apply_response.status_code == 200, apply_response.text
            apply_payload = apply_response.json()
            assert apply_payload["status"] == "applied"
            assert apply_payload["suggestion"]["status"] == "applied"
            assert apply_payload["can_apply"] is False
            assert apply_payload["can_cancel_restore"] is False

            latest_after_apply = admin_client.get(
                "/api/transport/ai/suggestions/latest",
                params={"service_date": service_date.isoformat(), "route_kind": "home_to_work"},
            )
            assert latest_after_apply.status_code == 404, latest_after_apply.text

        with SessionLocal() as db:
            run = db.execute(select(TransportAIRun).where(TransportAIRun.run_key == run_key)).scalar_one()
            suggestion = db.execute(
                select(TransportAISuggestion).where(TransportAISuggestion.suggestion_key == suggestion_key)
            ).scalar_one()
            assignment = db.execute(
                select(TransportAssignment).where(
                    TransportAssignment.request_id == request_id,
                    TransportAssignment.service_date == service_date,
                    TransportAssignment.route_kind == "home_to_work",
                )
            ).scalar_one()
            vehicle = db.get(Vehicle, assignment.vehicle_id)
            applied_route_stops = db.execute(
                select(TransportAIAppliedRouteStop).where(TransportAIAppliedRouteStop.suggestion_id == suggestion.id)
            ).scalars().all()

        assert run.status == "applied"
        assert suggestion.status == "applied"
        assert assignment.status == "confirmed"
        assert vehicle is not None
        assert applied_route_stops
    finally:
        with SessionLocal() as db:
            _cleanup_transport_ai_api_regression_fixture(db, fixture_bundle, run_key=run_key)


def test_transport_ai_api_flow_start_suggestion_cancel_restore(monkeypatch):
    from sistema.app.models import TransportAIRun, TransportAISuggestion

    _configure_transport_ai_api_regression_runtime(monkeypatch)
    transport_reevaluation_module.clear_transport_reevaluation_events()

    service_date = date(2035, 2, 13)
    fixture_bundle: dict[str, object] | None = None
    run_key: str | None = None

    try:
        with SessionLocal() as db:
            fixture_bundle = _create_transport_ai_api_regression_fixture(
                db,
                service_date=service_date,
                seed_existing_assignment=True,
            )
            db.commit()

        seeded_assignment_id = int(fixture_bundle["assignment_id"])
        seeded_vehicle_id = int(fixture_bundle["seed_vehicle_id"])
        request_id = int(fixture_bundle["requests"]["primary"]["id"])

        with SessionLocal() as db:
            _isolate_transport_ai_api_regression_requests(db, keep_request_ids=[request_id])
            db.commit()

        with TestClient(app) as admin_client:
            ensure_admin_session(admin_client)
            start_payload, status_payload = _start_transport_ai_api_regression_run(
                admin_client,
                service_date=service_date,
            )
            run_key = str(start_payload["run_key"])
            suggestion_key = str(status_payload["suggestion_key"])

            with SessionLocal() as db:
                pending_assignment = db.get(TransportAssignment, seeded_assignment_id)

            assert pending_assignment is not None
            assert pending_assignment.status == "pending"
            assert pending_assignment.vehicle_id is None

            cancel_response = admin_client.post(f"/api/transport/ai/suggestions/{suggestion_key}/cancel")
            assert cancel_response.status_code == 200, cancel_response.text
            cancel_payload = cancel_response.json()
            assert cancel_payload["status"] == "cancelled"
            assert cancel_payload["suggestion"]["status"] == "discarded"
            assert cancel_payload["can_apply"] is False
            assert cancel_payload["can_cancel_restore"] is False

            latest_after_cancel = admin_client.get(
                "/api/transport/ai/suggestions/latest",
                params={"service_date": service_date.isoformat(), "route_kind": "home_to_work"},
            )
            assert latest_after_cancel.status_code == 404, latest_after_cancel.text

        with SessionLocal() as db:
            run = db.execute(select(TransportAIRun).where(TransportAIRun.run_key == run_key)).scalar_one()
            suggestion = db.execute(
                select(TransportAISuggestion).where(TransportAISuggestion.suggestion_key == suggestion_key)
            ).scalar_one()
            restored_assignment = db.get(TransportAssignment, seeded_assignment_id)

        assert run.status == "cancelled"
        assert suggestion.status == "discarded"
        assert restored_assignment is not None
        assert restored_assignment.status == "confirmed"
        assert restored_assignment.vehicle_id == seeded_vehicle_id
    finally:
        with SessionLocal() as db:
            _cleanup_transport_ai_api_regression_fixture(db, fixture_bundle, run_key=run_key)


def test_transport_ai_api_flow_apply_blocks_when_request_drifts(monkeypatch):
    from sistema.app.models import TransportAIRun, TransportAISuggestion

    _configure_transport_ai_api_regression_runtime(monkeypatch)
    transport_reevaluation_module.clear_transport_reevaluation_events()

    service_date = date(2035, 2, 14)
    fixture_bundle: dict[str, object] | None = None
    run_key: str | None = None

    try:
        with SessionLocal() as db:
            fixture_bundle = _create_transport_ai_api_regression_fixture(
                db,
                service_date=service_date,
                seed_existing_assignment=False,
            )
            db.commit()

        request_id = int(fixture_bundle["requests"]["primary"]["id"])

        with SessionLocal() as db:
            _isolate_transport_ai_api_regression_requests(db, keep_request_ids=[request_id])
            db.commit()

        with TestClient(app) as admin_client:
            ensure_admin_session(admin_client)
            start_payload, status_payload = _start_transport_ai_api_regression_run(
                admin_client,
                service_date=service_date,
            )
            run_key = str(start_payload["run_key"])
            suggestion_key = str(status_payload["suggestion_key"])

            with SessionLocal() as db:
                drift_vehicle, _ = _create_transport_ai_api_regression_vehicle_candidate(
                    db,
                    plate=f"DR{uuid.uuid4().hex[:6].upper()}",
                    service_date=service_date,
                )
                pending_assignment = db.execute(
                    select(TransportAssignment).where(
                        TransportAssignment.request_id == request_id,
                        TransportAssignment.service_date == service_date,
                        TransportAssignment.route_kind == "home_to_work",
                    )
                ).scalar_one()
                pending_assignment.status = "confirmed"
                pending_assignment.vehicle_id = drift_vehicle.id
                pending_assignment.updated_at = now_sgt()
                db.commit()

            apply_response = admin_client.post(f"/api/transport/ai/suggestions/{suggestion_key}/apply")
            assert apply_response.status_code == 409, apply_response.text
            apply_payload = apply_response.json()
            assert apply_payload["status"] == "proposed"
            assert apply_payload["suggestion"]["status"] == "shown"
            assert any(issue["code"] == "request_not_pending" for issue in apply_payload["issues"])

        with SessionLocal() as db:
            run = db.execute(select(TransportAIRun).where(TransportAIRun.run_key == run_key)).scalar_one()
            suggestion = db.execute(
                select(TransportAISuggestion).where(TransportAISuggestion.suggestion_key == suggestion_key)
            ).scalar_one()
            drifted_assignment = db.execute(
                select(TransportAssignment).where(
                    TransportAssignment.request_id == request_id,
                    TransportAssignment.service_date == service_date,
                    TransportAssignment.route_kind == "home_to_work",
                )
            ).scalar_one()

        assert run.status == "proposed"
        assert suggestion.status == "shown"
        assert drifted_assignment.status == "confirmed"
        assert drifted_assignment.vehicle_id is not None
    finally:
        with SessionLocal() as db:
            _cleanup_transport_ai_api_regression_fixture(db, fixture_bundle, run_key=run_key)
