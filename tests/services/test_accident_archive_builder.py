"""Tests for Task F2 — accident_archive_builder service."""
from __future__ import annotations

import json
import zipfile
from datetime import datetime
from io import BytesIO
from pathlib import Path
from unittest.mock import MagicMock, call, patch

import sqlalchemy as sa
from openpyxl import load_workbook
from sqlalchemy.orm import Session, sessionmaker

from sistema.app.database import Base
from sistema.app.models import (
    Accident,
    AccidentArchive,
    AccidentUserReport,
    AccidentVideoUpload,
    AdminUser,
    Project,
    User,
)
from sistema.app.services.accident_archive_builder import (
    COLUMN_ORDER,
    _build_xlsx,
    _slugify,
    build_and_attach_archive_for_accident,
)
from sistema.app.services.accident_numbering import format_accident_number


# ---------------------------------------------------------------------------
# Shared test infrastructure
# ---------------------------------------------------------------------------

_NOW = datetime(2026, 3, 15, 10, 0, 0)


def _make_session(tmp_path: Path) -> Session:
    engine = sa.create_engine(
        f"sqlite+pysqlite:///{(tmp_path / 'test.db').as_posix()}"
    )
    Base.metadata.create_all(bind=engine)
    factory = sessionmaker(
        bind=engine, autocommit=False, autoflush=False, expire_on_commit=False
    )
    return factory()


def _make_project(db: Session) -> Project:
    p = Project(
        name="ARCHTEST",
        country_code="SG",
        country_name="Singapore",
        timezone_name="Asia/Singapore",
        address="1 Arch St",
        zip_code="099999",
    )
    db.add(p)
    db.flush()
    return p


def _make_admin(db: Session) -> AdminUser:
    a = AdminUser(
        chave="ADMA1",
        nome_completo="Admin Arch",
        created_at=_NOW,
        updated_at=_NOW,
    )
    db.add(a)
    db.flush()
    return a


def _make_accident(
    db: Session,
    proj: Project,
    admin: AdminUser,
    *,
    accident_number: int = 0,
) -> Accident:
    a = Accident(
        accident_number=accident_number,
        project_id=proj.id,
        project_name_snapshot=proj.name,
        location_name_snapshot="Sala A",
        location_is_registered=False,
        origin="admin",
        opened_by_admin_id=admin.id,
        opened_at=_NOW,
        created_at=_NOW,
        updated_at=_NOW,
    )
    db.add(a)
    db.flush()
    return a


def _make_user(db: Session, chave: str) -> User:
    u = User(
        chave=chave,
        nome=f"User {chave}",
        projeto="ARCHTEST",
        checkin=False,
        local="Sala A",
        last_active_at=_NOW,
        inactivity_days=0,
    )
    db.add(u)
    db.flush()
    return u


def _make_report(db: Session, accident: Accident, user: User) -> AccidentUserReport:
    r = AccidentUserReport(
        accident_id=accident.id,
        user_id=user.id,
        user_chave_snapshot=user.chave,
        user_name_snapshot=user.nome,
        user_phone_snapshot=None,
        user_projects_snapshot=json.dumps(["ARCHTEST"]),
        user_local_snapshot="Sala A",
        zone="safety",
        status="ok",
        reported_at=_NOW,
        created_at=_NOW,
        updated_at=_NOW,
    )
    db.add(r)
    db.flush()
    return r


def _make_video(
    db: Session,
    accident: Accident,
    user: User,
    *,
    idempotency_key: str = "idem-001",
    content_type: str = "video/mp4",
) -> AccidentVideoUpload:
    v = AccidentVideoUpload(
        idempotency_key=idempotency_key,
        accident_id=accident.id,
        user_id=user.id,
        object_key=f"accidents/0000/{user.id}/{idempotency_key}.mp4",
        public_url=f"https://example.com/{idempotency_key}.mp4",
        content_type=content_type,
        size_bytes=512,
        captured_at=_NOW,
        created_at=_NOW,
    )
    db.add(v)
    db.flush()
    return v


def _local_settings(tmp_path: Path) -> MagicMock:
    return MagicMock(
        event_archives_dir=str(tmp_path),
        do_spaces_bucket=None,
        do_spaces_access_key=None,
        do_spaces_secret_key=None,
        tz_name="UTC",
    )


# ---------------------------------------------------------------------------
# test_archive_zip_contains_xlsx
# ---------------------------------------------------------------------------


def test_archive_zip_contains_xlsx(tmp_path: Path):
    """The generated ZIP must contain <NNNN>.xlsx at the root."""
    db = _make_session(tmp_path)
    proj = _make_project(db)
    admin = _make_admin(db)
    accident = _make_accident(db, proj, admin, accident_number=7)
    user = _make_user(db, "U101")
    _make_report(db, accident, user)
    db.commit()

    with (
        patch("sistema.app.services.accident_archive_builder.SessionLocal", return_value=db),
        patch("sistema.app.services.object_storage.settings", _local_settings(tmp_path)),
        patch("sistema.app.services.accident_archive_builder._use_remote", return_value=False),
        patch("sistema.app.services.accident_archive_builder.notify_admin_data_changed"),
    ):
        build_and_attach_archive_for_accident(accident.id)

    zip_path = (
        tmp_path
        / "accidents_local_storage"
        / "accidents"
        / "0007"
        / "archive"
        / "0007.zip"
    )
    assert zip_path.exists(), "ZIP file not created"
    with zipfile.ZipFile(zip_path) as zf:
        names = zf.namelist()
    assert "0007.xlsx" in names, f"0007.xlsx not in ZIP: {names}"


# ---------------------------------------------------------------------------
# test_archive_zip_contains_videos_subfolder
# ---------------------------------------------------------------------------


def test_archive_zip_contains_videos_subfolder(tmp_path: Path):
    """The generated ZIP must contain Registros/<filename> for each video."""
    db = _make_session(tmp_path)
    proj = _make_project(db)
    admin = _make_admin(db)
    accident = _make_accident(db, proj, admin, accident_number=8)
    user = _make_user(db, "U102")
    _make_report(db, accident, user)
    video = _make_video(db, accident, user, idempotency_key="idem-abc")
    db.commit()

    # Write fake video bytes to local storage
    fake_video_path = (
        tmp_path / "accidents_local_storage" / video.object_key
    )
    fake_video_path.parent.mkdir(parents=True, exist_ok=True)
    fake_video_path.write_bytes(b"fake-video-data")

    with (
        patch("sistema.app.services.accident_archive_builder.SessionLocal", return_value=db),
        patch("sistema.app.services.object_storage.settings", _local_settings(tmp_path)),
        patch("sistema.app.services.accident_archive_builder._use_remote", return_value=False),
        patch("sistema.app.services.accident_archive_builder._local_root", return_value=tmp_path / "accidents_local_storage"),
        patch("sistema.app.services.accident_archive_builder.notify_admin_data_changed"),
    ):
        build_and_attach_archive_for_accident(accident.id)

    zip_path = (
        tmp_path
        / "accidents_local_storage"
        / "accidents"
        / "0008"
        / "archive"
        / "0008.zip"
    )
    assert zip_path.exists()
    with zipfile.ZipFile(zip_path) as zf:
        names = zf.namelist()

    registros = [n for n in names if n.startswith("Registros/")]
    assert len(registros) == 1, f"Expected 1 Registros/ entry, got {names}"
    expected_slug = _slugify("idem-abc")
    assert registros[0] == f"Registros/{user.chave}/01_{expected_slug}.mp4", registros[0]


# ---------------------------------------------------------------------------
# test_xlsx_columns_match_spec
# ---------------------------------------------------------------------------


def test_xlsx_columns_match_spec():
    """The XLSX first row must exactly match COLUMN_ORDER."""
    from types import SimpleNamespace
    from sistema.app.schemas import SituacaoPessoalRow

    accident_stub = SimpleNamespace(
        accident_number=1,
        project_name_snapshot="Test Project",
        location_name_snapshot="Test Location",
        opened_at=_NOW,
        description="",
    )
    row = SituacaoPessoalRow(
        user_id=1,
        event_time=_NOW,
        name="Alice",
        chave="A001",
        projects=["P1"],
        local="Office",
        zone="Segurança",
        status="OK",
        phone=None,
        videos=[],
        priority=4,
        row_color="light-green",
    )
    buf = _build_xlsx(accident_stub, [row], video_files_by_user_chave={})
    wb = load_workbook(buf)
    ws = wb.active
    # 5 metadata rows + 1 blank separator = 6 rows before the column header row (row 7)
    header = [ws.cell(row=7, column=i + 1).value for i in range(len(COLUMN_ORDER))]
    assert header == COLUMN_ORDER


# ---------------------------------------------------------------------------
# test_xlsx_handles_zero_videos
# ---------------------------------------------------------------------------


def test_xlsx_handles_zero_videos():
    """XLSX built without any videos must still open correctly with empty Registros cell."""
    from types import SimpleNamespace
    from sistema.app.schemas import SituacaoPessoalRow

    accident_stub = SimpleNamespace(
        accident_number=2,
        project_name_snapshot="Test Project",
        location_name_snapshot="Test Location",
        opened_at=_NOW,
        description="",
    )
    row = SituacaoPessoalRow(
        user_id=2,
        event_time=_NOW,
        name="Bob",
        chave="B002",
        projects=["P2"],
        local=None,
        zone="Aguardando",
        status="Aguardando",
        phone=None,
        videos=[],
        priority=3,
        row_color="light-blue",
    )
    buf = _build_xlsx(accident_stub, [row], video_files_by_user_chave={})
    wb = load_workbook(buf)
    ws = wb.active
    # 5 metadata rows + 1 blank separator + 1 column header = data starts at row 8
    registros_col = COLUMN_ORDER.index("Registros") + 1
    registros_cell = ws.cell(row=8, column=registros_col).value
    assert registros_cell == "" or registros_cell is None


# ---------------------------------------------------------------------------
# test_xlsx_filename_uses_4_digit_format
# ---------------------------------------------------------------------------


def test_xlsx_filename_uses_4_digit_format(tmp_path: Path):
    """XLSX filename inside the ZIP must use the zero-padded 4-digit accident number."""
    db = _make_session(tmp_path)
    proj = _make_project(db)
    admin = _make_admin(db)
    accident = _make_accident(db, proj, admin, accident_number=42)
    user = _make_user(db, "U103")
    _make_report(db, accident, user)
    db.commit()

    with (
        patch("sistema.app.services.accident_archive_builder.SessionLocal", return_value=db),
        patch("sistema.app.services.object_storage.settings", _local_settings(tmp_path)),
        patch("sistema.app.services.accident_archive_builder._use_remote", return_value=False),
        patch("sistema.app.services.accident_archive_builder.notify_admin_data_changed"),
    ):
        build_and_attach_archive_for_accident(accident.id)

    zip_path = (
        tmp_path
        / "accidents_local_storage"
        / "accidents"
        / "0042"
        / "archive"
        / "0042.zip"
    )
    assert zip_path.exists()
    with zipfile.ZipFile(zip_path) as zf:
        names = zf.namelist()
    assert "0042.xlsx" in names, f"Expected 0042.xlsx in {names}"


# ---------------------------------------------------------------------------
# test_archive_record_persists
# ---------------------------------------------------------------------------


def test_archive_record_persists(tmp_path: Path):
    """After build, AccidentArchive row must exist and accident.archive_object_key set."""
    # Use separate sessions: one for setup, one injected into builder, one for assertions
    engine = sa.create_engine(
        f"sqlite+pysqlite:///{(tmp_path / 'test.db').as_posix()}"
    )
    Base.metadata.create_all(bind=engine)
    factory = sessionmaker(bind=engine, autocommit=False, autoflush=False, expire_on_commit=False)

    setup_db = factory()
    proj = _make_project(setup_db)
    admin = _make_admin(setup_db)
    accident = _make_accident(setup_db, proj, admin, accident_number=3)
    user = _make_user(setup_db, "U104")
    _make_report(setup_db, accident, user)
    setup_db.commit()
    accident_id = accident.id
    setup_db.close()

    # Wrapper: provides a session to builder without closing it on __exit__
    class _CommitOnlySession:
        def __init__(self):
            self._db = factory()
        def __enter__(self):
            return self._db
        def __exit__(self, *a):
            self._db.commit()

    builder_cm = _CommitOnlySession()
    mock_session_local = MagicMock(return_value=builder_cm)

    with (
        patch("sistema.app.services.accident_archive_builder.SessionLocal", mock_session_local),
        patch("sistema.app.services.object_storage.settings", _local_settings(tmp_path)),
        patch("sistema.app.services.accident_archive_builder._use_remote", return_value=False),
        patch("sistema.app.services.accident_archive_builder.notify_admin_data_changed"),
    ):
        build_and_attach_archive_for_accident(accident_id)

    check_db = factory()
    archive = check_db.query(AccidentArchive).filter_by(accident_id=accident_id).first()
    assert archive is not None, "AccidentArchive row not created"
    assert archive.xlsx_object_key == "accidents/0003/archive/0003.xlsx"
    assert archive.zip_object_key == "accidents/0003/archive/0003.zip"
    assert archive.size_bytes > 0
    assert archive.snapshot_json is not None

    refreshed_accident = check_db.get(Accident, accident_id)
    assert refreshed_accident.archive_object_key == "accidents/0003/archive/0003.zip"
    check_db.close()


# ---------------------------------------------------------------------------
# test_archive_publishes_ready_event
# ---------------------------------------------------------------------------


def test_archive_publishes_ready_event(tmp_path: Path):
    """build_and_attach_archive_for_accident must call notify_admin_data_changed with archive_ready=True."""
    db = _make_session(tmp_path)
    proj = _make_project(db)
    admin = _make_admin(db)
    accident = _make_accident(db, proj, admin, accident_number=5)
    user = _make_user(db, "U105")
    _make_report(db, accident, user)
    db.commit()

    mock_notify = MagicMock()

    with (
        patch("sistema.app.services.accident_archive_builder.SessionLocal", return_value=db),
        patch("sistema.app.services.object_storage.settings", _local_settings(tmp_path)),
        patch("sistema.app.services.accident_archive_builder._use_remote", return_value=False),
        patch("sistema.app.services.accident_archive_builder.notify_admin_data_changed", mock_notify),
    ):
        build_and_attach_archive_for_accident(accident.id)

    mock_notify.assert_called_once_with(
        "accident_closed",
        metadata={"accident_id": accident.id, "archive_ready": True},
    )


# ---------------------------------------------------------------------------
# Phase 7 / prompt 7.1 — Accident.description in the generated XLSX
# ---------------------------------------------------------------------------


def _read_xlsx_from_local_zip(tmp_path: Path, accident_label: str):
    """Extract the XLSX from the locally-stored archive ZIP and load it with openpyxl."""
    zip_path = (
        tmp_path
        / "accidents_local_storage"
        / "accidents"
        / accident_label
        / "archive"
        / f"{accident_label}.zip"
    )
    assert zip_path.exists(), f"archive ZIP not generated at {zip_path}"
    with zipfile.ZipFile(zip_path) as zf:
        names = zf.namelist()
        xlsx_name = f"{accident_label}.xlsx"
        assert xlsx_name in names, f"{xlsx_name} missing from ZIP: {names}"
        with zf.open(xlsx_name) as xlsx_fp:
            buffer = BytesIO(xlsx_fp.read())
    return load_workbook(buffer), zip_path


def test_xlsx_header_row5_includes_accident_description_with_accents_and_emoji(tmp_path: Path):
    """Phase 7 / item 2.3: the generated XLSX header row 5 must contain the accident description.

    UTF-8 accents and emoji must survive round-tripping through the archive
    pipeline (openpyxl → zip → openpyxl). We verify against the exact prefix
    "Descrição: <text>" that _build_xlsx writes at line 79.
    """
    description = "Descrição com acentuação áéíóú, ç e emoji 🚨 — relatório completo."

    db = _make_session(tmp_path)
    proj = _make_project(db)
    admin = _make_admin(db)
    accident = _make_accident(db, proj, admin, accident_number=11)
    accident.description = description
    user = _make_user(db, "U201")
    _make_report(db, accident, user)
    db.commit()
    accident_label = format_accident_number(accident.accident_number)

    with (
        patch("sistema.app.services.accident_archive_builder.SessionLocal", return_value=db),
        patch("sistema.app.services.object_storage.settings", _local_settings(tmp_path)),
        patch("sistema.app.services.accident_archive_builder._use_remote", return_value=False),
        patch("sistema.app.services.accident_archive_builder.notify_admin_data_changed"),
    ):
        build_and_attach_archive_for_accident(accident.id)

    wb, _zip_path = _read_xlsx_from_local_zip(tmp_path, accident_label)
    ws = wb.active

    # Header rows in _build_xlsx (linha 74-81):
    #   row 1: Acidente N.º
    #   row 2: Projeto
    #   row 3: Local
    #   row 4: Data abertura
    #   row 5: Descrição
    #   row 6: blank
    a5 = ws.cell(row=5, column=1).value
    assert a5 == f"Descrição: {description}", (
        f"Expected description on A5, got {a5!r}"
    )
    # Belt and suspenders — emoji and accents must be exactly preserved.
    assert "🚨" in a5
    assert "áéíóú" in a5
    assert "ç" in a5


def test_xlsx_header_row5_uses_placeholder_when_description_is_empty(tmp_path: Path):
    """When accident.description == '' (default), the XLSX shows 'Descrição: (sem descrição)'.

    Pins the fallback in _build_xlsx so closing an accident without ever filling
    the description still produces a readable header.
    """
    db = _make_session(tmp_path)
    proj = _make_project(db)
    admin = _make_admin(db)
    accident = _make_accident(db, proj, admin, accident_number=12)
    # description defaults to "" via the column default; do not override.
    user = _make_user(db, "U202")
    _make_report(db, accident, user)
    db.commit()
    accident_label = format_accident_number(accident.accident_number)

    with (
        patch("sistema.app.services.accident_archive_builder.SessionLocal", return_value=db),
        patch("sistema.app.services.object_storage.settings", _local_settings(tmp_path)),
        patch("sistema.app.services.accident_archive_builder._use_remote", return_value=False),
        patch("sistema.app.services.accident_archive_builder.notify_admin_data_changed"),
    ):
        build_and_attach_archive_for_accident(accident.id)

    wb, _zip_path = _read_xlsx_from_local_zip(tmp_path, accident_label)
    ws = wb.active
    a5 = ws.cell(row=5, column=1).value
    assert a5 == "Descrição: (sem descrição)", f"Expected fallback placeholder, got {a5!r}"


def test_xlsx_description_does_not_disturb_column_header_row7(tmp_path: Path):
    """Pinning that adding the Descrição header row did not push COLUMN_ORDER off row 7.

    test_xlsx_columns_match_spec already asserts this for the synthetic accident
    stub; here we assert it again for the full archive-builder pipeline so a
    regression in either layer is caught.
    """
    db = _make_session(tmp_path)
    proj = _make_project(db)
    admin = _make_admin(db)
    accident = _make_accident(db, proj, admin, accident_number=13)
    accident.description = "qualquer texto aqui"
    user = _make_user(db, "U203")
    _make_report(db, accident, user)
    db.commit()
    accident_label = format_accident_number(accident.accident_number)

    with (
        patch("sistema.app.services.accident_archive_builder.SessionLocal", return_value=db),
        patch("sistema.app.services.object_storage.settings", _local_settings(tmp_path)),
        patch("sistema.app.services.accident_archive_builder._use_remote", return_value=False),
        patch("sistema.app.services.accident_archive_builder.notify_admin_data_changed"),
    ):
        build_and_attach_archive_for_accident(accident.id)

    wb, _zip_path = _read_xlsx_from_local_zip(tmp_path, accident_label)
    ws = wb.active
    header_row = [ws.cell(row=7, column=i + 1).value for i in range(len(COLUMN_ORDER))]
    assert header_row == COLUMN_ORDER, (
        f"Column header row 7 drifted from COLUMN_ORDER: {header_row!r}"
    )
