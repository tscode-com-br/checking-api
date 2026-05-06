from __future__ import annotations

from collections import Counter
from datetime import date, datetime
from io import BytesIO
import json
from pathlib import Path

from sqlalchemy import select
from sqlalchemy.orm import Session

from ..core.config import settings
from ..models import TransportAIRun, TransportAISuggestion, TransportAssignment, TransportRequest, User, Vehicle
from ..schemas import TransportOperationalProposal, TransportOperationalSnapshot
from .time_utils import now_sgt
from .transport_ai_runs import resolve_transport_ai_run_llm_snapshot_fields


_PAIRED_ROUTE_KIND = {
    "home_to_work": "work_to_home",
    "work_to_home": "home_to_work",
}


def _build_transport_export_file_name(timestamp: datetime) -> str:
    return f"Transport List - {timestamp:%Y%m%d - %H%M%S}.xlsx"


def _build_transport_operational_plan_file_name(timestamp: datetime) -> str:
    return f"Transport Operational Plan - {timestamp:%Y%m%d - %H%M%S}.xlsx"


def _resolve_transport_export_path(file_name: str) -> Path:
    export_dir = Path(settings.transport_exports_dir)
    export_dir.mkdir(parents=True, exist_ok=True)

    candidate = export_dir / file_name
    if not candidate.exists():
        return candidate

    counter = 2
    while True:
        deduplicated = export_dir / f"{candidate.stem} ({counter}){candidate.suffix}"
        if not deduplicated.exists():
            return deduplicated
        counter += 1


def build_transport_list_export(
    db: Session,
    *,
    service_date: date,
    selected_route_kind: str,
) -> tuple[str, bytes]:
    return build_transport_operational_plan_export(
        db,
        service_date=service_date,
        selected_route_kind=selected_route_kind,
        proposal=None,
    )


def _collect_transport_list_rows(
    db: Session,
    *,
    service_date: date,
    selected_route_kind: str,
) -> list[list[str | None]]:
    from .transport_vehicle_operations import find_transport_vehicle_schedule

    route_priority = {
        selected_route_kind: 0,
        _PAIRED_ROUTE_KIND.get(selected_route_kind, "work_to_home"): 1,
    }

    exported_rows: list[tuple[tuple[str, str, int], list[str | None]]] = []
    assignments = db.execute(
        select(TransportAssignment, TransportRequest, User, Vehicle)
        .join(TransportRequest, TransportRequest.id == TransportAssignment.request_id)
        .join(User, User.id == TransportRequest.user_id)
        .join(Vehicle, Vehicle.id == TransportAssignment.vehicle_id)
        .where(
            TransportAssignment.service_date == service_date,
            TransportAssignment.status == "confirmed",
            TransportAssignment.vehicle_id.is_not(None),
        )
    ).all()

    assignments_by_request: dict[int, tuple[int, TransportAssignment, User, Vehicle]] = {}
    for assignment, transport_request, user, vehicle in assignments:
        candidate_priority = route_priority.get(assignment.route_kind, 2)
        current = assignments_by_request.get(transport_request.id)
        if current is None or candidate_priority < current[0]:
            assignments_by_request[transport_request.id] = (candidate_priority, assignment, user, vehicle)

    for _priority, assignment, user, vehicle in assignments_by_request.values():
        schedule = find_transport_vehicle_schedule(
            db,
            vehicle=vehicle,
            service_date=service_date,
            route_kind=assignment.route_kind,
        )
        exported_rows.append(
            (
                (
                    user.nome.lower(),
                    user.chave,
                    assignment.id,
                ),
                [
                    user.nome,
                    user.chave,
                    user.projeto,
                    user.end_rua,
                    service_date.isoformat(),
                    schedule.departure_time if schedule is not None else None,
                ],
            )
        )

    return [row for _, row in sorted(exported_rows, key=lambda item: item[0])]


def _snapshot_request_rows(snapshot: TransportOperationalSnapshot):
    return snapshot.regular_requests + snapshot.weekend_requests + snapshot.extra_requests


def _snapshot_vehicle_registry_rows(snapshot: TransportOperationalSnapshot):
    return (
        snapshot.regular_vehicle_registry
        + snapshot.weekend_vehicle_registry
        + snapshot.extra_vehicle_registry
    )


def _append_sheet_rows(worksheet, rows: list[list[object]]) -> None:
    for row in rows:
        worksheet.append(row)


def _build_executive_summary_rows(
    *,
    snapshot: TransportOperationalSnapshot,
    generated_at: datetime,
    proposal: TransportOperationalProposal | None,
) -> list[list[object]]:
    request_status_counts = Counter(request.assignment_status for request in _snapshot_request_rows(snapshot))
    summary_rows: list[list[object]] = [
        ["Campo/Field", "Valor/Value"],
        ["Modo/Mode", "proposal_review" if proposal is not None else "current_state"],
        ["Gerado em/Generated At", generated_at.isoformat()],
        ["Data/Service Date", snapshot.service_date.isoformat()],
        ["Rota/Route", snapshot.route_kind],
        ["Saida Work to Home/Work to Home Departure", snapshot.work_to_home_departure_time],
        ["Total Requests", len(_snapshot_request_rows(snapshot))],
        ["Pending Requests", request_status_counts.get("pending", 0)],
        ["Confirmed Requests", request_status_counts.get("confirmed", 0)],
        ["Rejected Requests", request_status_counts.get("rejected", 0)],
        ["Cancelled Requests", request_status_counts.get("cancelled", 0)],
        ["Total Vehicles", len(_snapshot_vehicle_registry_rows(snapshot))],
    ]

    if proposal is None:
        return summary_rows

    summary_rows.extend(
        [
            ["Proposal Key", proposal.proposal_key],
            ["Proposal Status", proposal.proposal_status],
            ["Proposal Origin", proposal.origin],
            ["Proposal Created At", proposal.created_at.isoformat()],
            ["Proposal Expires At", proposal.expires_at.isoformat() if proposal.expires_at else None],
            ["Total Decisions", proposal.summary.total_decisions],
            ["Confirmed Decisions", proposal.summary.confirmed_decisions],
            ["Rejected Decisions", proposal.summary.rejected_decisions],
            ["Pending Decisions", proposal.summary.pending_decisions],
            ["Validation Issues", len(proposal.validation_issues)],
            ["Audit Entries", len(proposal.audit_trail)],
        ]
    )
    return summary_rows


def _build_vehicle_load_rows(
    *,
    snapshot: TransportOperationalSnapshot,
    proposal: TransportOperationalProposal | None,
) -> list[list[object]]:
    projected_confirmations = Counter(
        decision.vehicle_id
        for decision in (proposal.decisions if proposal is not None else [])
        if decision.suggested_status == "confirmed" and decision.vehicle_id is not None
    )
    rows: list[list[object]] = [[
        "Placa/Plate",
        "Tipo/Type",
        "Capacidade/Capacity",
        "Alocados Atuais/Current Assigned",
        "Alocados Projetados/Projected Assigned",
        "Saldo Projetado/Projected Remaining",
        "Data/Date",
        "Rota/Route",
        "Partida/Departure",
    ]]

    for registry in sorted(
        _snapshot_vehicle_registry_rows(snapshot),
        key=lambda row: (row.placa or "", row.route_kind or "", row.vehicle_id),
    ):
        projected_assigned = registry.assigned_count + projected_confirmations.get(registry.vehicle_id, 0)
        projected_remaining = None if registry.lugares is None else registry.lugares - projected_assigned
        rows.append(
            [
                registry.placa,
                registry.tipo,
                registry.lugares,
                registry.assigned_count,
                projected_assigned,
                projected_remaining,
                registry.service_date.isoformat() if registry.service_date else None,
                registry.route_kind,
                registry.departure_time,
            ]
        )

    return rows


def _build_snapshot_request_rows(snapshot: TransportOperationalSnapshot) -> list[list[object]]:
    rows: list[list[object]] = [[
        "Request ID",
        "Tipo/Kind",
        "Status",
        "Horario/Time",
        "Data/Date",
        "Chave/Key",
        "Nome/Name",
        "Projeto/Project",
        "Workplace",
        "Endereco/Address",
        "Veiculo/Vehicle",
        "Resposta/Response",
    ]]

    for request in sorted(
        _snapshot_request_rows(snapshot),
        key=lambda row: (row.request_kind, row.nome.lower(), row.chave, row.id),
    ):
        rows.append(
            [
                request.id,
                request.request_kind,
                request.assignment_status,
                request.requested_time,
                request.service_date.isoformat(),
                request.chave,
                request.nome,
                request.projeto,
                request.workplace,
                request.end_rua,
                request.assigned_vehicle.placa if request.assigned_vehicle is not None else None,
                request.response_message,
            ]
        )

    return rows


def _build_proposed_decision_rows(
    *,
    snapshot: TransportOperationalSnapshot,
    proposal: TransportOperationalProposal,
) -> list[list[object]]:
    request_index = {request.id: request for request in _snapshot_request_rows(snapshot)}
    vehicle_index = {registry.vehicle_id: registry for registry in _snapshot_vehicle_registry_rows(snapshot)}
    rows: list[list[object]] = [[
        "Request ID",
        "Tipo/Kind",
        "Chave/Key",
        "Nome/Name",
        "Status Sugerido/Suggested Status",
        "Veiculo Sugerido/Suggested Vehicle",
        "Mensagem/Response",
        "Justificativa/Rationale",
    ]]

    for decision in proposal.decisions:
        request = request_index.get(decision.request_id)
        vehicle = vehicle_index.get(decision.vehicle_id) if decision.vehicle_id is not None else None
        rows.append(
            [
                decision.request_id,
                decision.request_kind,
                request.chave if request is not None else None,
                request.nome if request is not None else None,
                decision.suggested_status,
                vehicle.placa if vehicle is not None else decision.vehicle_id,
                decision.response_message,
                decision.rationale,
            ]
        )

    return rows


def _build_exception_rows(
    *,
    snapshot: TransportOperationalSnapshot,
    proposal: TransportOperationalProposal | None,
) -> list[list[object]]:
    rows: list[list[object]] = [["Tipo/Type", "Referencia/Reference", "Status ou Codigo/Status or Code", "Mensagem/Message"]]

    for request in sorted(
        _snapshot_request_rows(snapshot),
        key=lambda row: (row.assignment_status, row.nome.lower(), row.id),
    ):
        if request.assignment_status == "confirmed":
            continue
        rows.append(
            [
                "snapshot_request",
                f"request:{request.id}",
                request.assignment_status,
                f"{request.nome} ({request.chave}) remains {request.assignment_status}.",
            ]
        )

    if proposal is not None:
        for issue in proposal.validation_issues:
            reference_parts = []
            if issue.request_id is not None:
                reference_parts.append(f"request:{issue.request_id}")
            if issue.vehicle_id is not None:
                reference_parts.append(f"vehicle:{issue.vehicle_id}")
            rows.append(
                [
                    "validation_issue",
                    ", ".join(reference_parts) or proposal.proposal_key,
                    issue.code,
                    issue.message,
                ]
            )

        for decision in proposal.decisions:
            if decision.suggested_status == "confirmed":
                continue
            rows.append(
                [
                    "proposal_decision",
                    f"request:{decision.request_id}",
                    decision.suggested_status,
                    decision.response_message or decision.rationale,
                ]
            )

    return rows


def _build_audit_rows(proposal: TransportOperationalProposal) -> list[list[object]]:
    rows: list[list[object]] = [[
        "Acao/Action",
        "Resultado/Outcome",
        "Ator/Actor Key",
        "Nome/Actor Name",
        "Instante/Occurred At",
        "Mensagem/Message",
    ]]
    for entry in proposal.audit_trail:
        rows.append(
            [
                entry.action,
                entry.outcome,
                entry.actor.chave,
                entry.actor.nome_completo,
                entry.occurred_at.isoformat(),
                entry.message,
            ]
        )
    return rows


def _load_transport_ai_export_dict(raw_value: str | None) -> dict[str, object]:
    if not raw_value:
        return {}
    try:
        payload = json.loads(raw_value)
    except Exception:
        return {}
    return payload if isinstance(payload, dict) else {}


def _load_transport_ai_export_list(raw_value: str | None) -> list[dict[str, object]]:
    if not raw_value:
        return []
    try:
        payload = json.loads(raw_value)
    except Exception:
        return []
    if not isinstance(payload, list):
        return []
    return [item for item in payload if isinstance(item, dict)]


def _serialize_transport_export_value(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=True, separators=(",", ":"), sort_keys=True)
    return value


def _load_transport_ai_export_suggestion(
    db: Session,
    *,
    proposal: TransportOperationalProposal | None,
) -> TransportAISuggestion | None:
    if proposal is None or proposal.origin != "agent":
        return None

    return db.execute(
        select(TransportAISuggestion)
        .where(TransportAISuggestion.proposal_key == proposal.proposal_key)
        .order_by(TransportAISuggestion.updated_at.desc(), TransportAISuggestion.id.desc())
        .limit(1)
    ).scalar_one_or_none()


def _load_transport_ai_export_run(db: Session, *, suggestion: TransportAISuggestion | None) -> TransportAIRun | None:
    if suggestion is None:
        return None

    return db.execute(
        select(TransportAIRun)
        .where(TransportAIRun.id == suggestion.run_id)
        .limit(1)
    ).scalar_one_or_none()


def _resolve_transport_ai_export_llm_fields(run: TransportAIRun | None) -> tuple[str | None, str | None, str | None]:
    if run is None:
        return None, None, None

    llm_fields = resolve_transport_ai_run_llm_snapshot_fields(run)
    return (
        str(llm_fields["llm_provider"] or "").strip() or None,
        str(llm_fields["llm_model"] or "").strip() or None,
        str(llm_fields["llm_reasoning_effort"] or "").strip() or None,
    )


def _build_transport_ai_summary_rows(suggestion: TransportAISuggestion, *, run: TransportAIRun | None = None) -> list[list[object]]:
    plan_payload = _load_transport_ai_export_dict(suggestion.agent_plan_json)
    cost_summary = _load_transport_ai_export_dict(suggestion.cost_summary_json)
    change_summary = _load_transport_ai_export_dict(suggestion.change_summary_json)
    passenger_allocations = _load_transport_ai_export_list(suggestion.assignment_actions_json)
    route_itineraries = _load_transport_ai_export_list(suggestion.route_itineraries_json)
    validation_issues = _load_transport_ai_export_list(suggestion.validation_issues_json)
    llm_provider, llm_model, llm_reasoning_effort = _resolve_transport_ai_export_llm_fields(run)

    return [
        ["Campo/Field", "Valor/Value"],
        ["Suggestion Key", suggestion.suggestion_key],
        ["Suggestion Status", suggestion.status],
        ["Proposal Key", suggestion.proposal_key],
        ["Prompt Version", suggestion.prompt_version],
        ["LLM Provider", llm_provider],
        ["LLM Model", llm_model],
        ["LLM Reasoning Effort", llm_reasoning_effort],
        ["Suggestion Created At", suggestion.created_at.isoformat()],
        ["Suggestion Updated At", suggestion.updated_at.isoformat()],
        ["Suggestion Saved At", suggestion.saved_at.isoformat() if suggestion.saved_at is not None else None],
        ["Suggestion Applied At", suggestion.applied_at.isoformat() if suggestion.applied_at is not None else None],
        ["Suggestion Discarded At", suggestion.discarded_at.isoformat() if suggestion.discarded_at is not None else None],
        ["Objective Summary", plan_payload.get("objective_summary")],
        ["Current Estimated Cost", cost_summary.get("current_total_estimated_cost")],
        ["Suggested Estimated Cost", cost_summary.get("suggested_total_estimated_cost")],
        ["Estimated Cost Delta", cost_summary.get("estimated_cost_delta")],
        ["Price Currency Code", cost_summary.get("price_currency_code")],
        ["Price Rate Unit", cost_summary.get("price_rate_unit")],
        ["Current Vehicle Count", cost_summary.get("current_vehicle_count")],
        ["Suggested Vehicle Count", cost_summary.get("suggested_vehicle_count")],
        ["Total Vehicle Actions", change_summary.get("total_vehicle_actions")],
        ["Keep Actions", change_summary.get("keep_count")],
        ["Create Actions", change_summary.get("create_count")],
        ["Update Actions", change_summary.get("update_count")],
        ["Remove From Day Actions", change_summary.get("remove_from_day_count")],
        ["Passenger Allocations", len(passenger_allocations)],
        ["Route Itineraries", len(route_itineraries)],
        ["Validation Issues", len(validation_issues)],
    ]


def _build_transport_ai_vehicle_action_rows(suggestion: TransportAISuggestion) -> list[list[object]]:
    rows: list[list[object]] = [[
        "Action Key",
        "Action Type",
        "Scope",
        "Vehicle ID",
        "Schedule ID",
        "Client Vehicle Key",
        "Cost Delta",
        "Rationale",
        "Before",
        "After",
    ]]
    vehicle_actions = _load_transport_ai_export_list(suggestion.vehicle_actions_json)
    for action in sorted(vehicle_actions, key=lambda row: (str(row.get("action_type") or ""), str(row.get("action_key") or ""))):
        rows.append(
            [
                action.get("action_key"),
                action.get("action_type"),
                action.get("service_scope"),
                action.get("vehicle_id"),
                action.get("schedule_id"),
                action.get("client_vehicle_key"),
                action.get("cost_delta"),
                action.get("rationale"),
                _serialize_transport_export_value(action.get("before")),
                _serialize_transport_export_value(action.get("after")),
            ]
        )
    return rows


def _build_transport_ai_itinerary_rows(suggestion: TransportAISuggestion) -> list[list[object]]:
    rows: list[list[object]] = [[
        "Route Key",
        "Vehicle Ref",
        "Plate",
        "Vehicle Type",
        "Scope",
        "Project",
        "Country",
        "Projected Arrival",
        "Estimated Cost",
        "Stop Order",
        "Stop Type",
        "Request ID",
        "Passenger",
        "Address",
        "Scheduled Time",
        "Duration From Previous",
        "Distance From Previous",
    ]]
    itineraries = _load_transport_ai_export_list(suggestion.route_itineraries_json)
    for itinerary in sorted(itineraries, key=lambda row: (str(row.get("route_key") or ""), str(row.get("vehicle_ref") or ""))):
        stops = itinerary.get("stops") if isinstance(itinerary.get("stops"), list) else []
        stop_rows = [stop for stop in stops if isinstance(stop, dict)] or [{}]
        stop_rows.sort(key=lambda row: (int(row.get("stop_order") or 0), str(row.get("stop_type") or "")))
        for stop in stop_rows:
            rows.append(
                [
                    itinerary.get("route_key"),
                    itinerary.get("vehicle_ref"),
                    itinerary.get("plate"),
                    itinerary.get("vehicle_type"),
                    itinerary.get("service_scope"),
                    itinerary.get("project_name"),
                    itinerary.get("country_code"),
                    itinerary.get("projected_arrival_time"),
                    itinerary.get("estimated_cost"),
                    stop.get("stop_order"),
                    stop.get("stop_type"),
                    stop.get("request_id"),
                    stop.get("passenger_name"),
                    stop.get("address"),
                    stop.get("scheduled_time"),
                    stop.get("duration_from_previous_seconds"),
                    stop.get("distance_from_previous_meters"),
                ]
            )
    return rows


def _build_transport_ai_issue_rows(suggestion: TransportAISuggestion) -> list[list[object]]:
    rows: list[list[object]] = [["Code", "Blocking", "Request ID", "Vehicle ID", "Message"]]
    validation_issues = _load_transport_ai_export_list(suggestion.validation_issues_json)
    for issue in validation_issues:
        rows.append(
            [
                issue.get("code"),
                issue.get("blocking"),
                issue.get("request_id"),
                issue.get("vehicle_id"),
                issue.get("message"),
            ]
        )
    return rows


def build_transport_operational_plan_export(
    db: Session,
    *,
    service_date: date,
    selected_route_kind: str,
    proposal: TransportOperationalProposal | None,
) -> tuple[str, bytes]:
    from openpyxl import Workbook

    from .transport import now_sgt as transport_now_sgt
    from .transport_proposals import build_transport_operational_snapshot

    timestamp = transport_now_sgt()
    file_name = (
        _build_transport_operational_plan_file_name(timestamp)
        if proposal is not None
        else _build_transport_export_file_name(timestamp)
    )
    snapshot = proposal.snapshot if proposal is not None else build_transport_operational_snapshot(
        db,
        service_date=service_date,
        route_kind=selected_route_kind,
        captured_at=timestamp,
    )
    ai_suggestion = _load_transport_ai_export_suggestion(db, proposal=proposal)
    ai_run = _load_transport_ai_export_run(db, suggestion=ai_suggestion)
    export_rows = _collect_transport_list_rows(
        db,
        service_date=service_date,
        selected_route_kind=selected_route_kind,
    )

    workbook = Workbook()
    transport_list_sheet = workbook.active
    transport_list_sheet.title = "Transport List"
    transport_list_sheet.append([
        "Nome/Name",
        "Chave/Key",
        "Projeto/Project",
        "Endereço/Address",
        "Data/Date",
        "Partida/Departure",
    ])
    for row in export_rows:
        transport_list_sheet.append(row)

    executive_summary_sheet = workbook.create_sheet("Executive Summary")
    _append_sheet_rows(
        executive_summary_sheet,
        _build_executive_summary_rows(snapshot=snapshot, generated_at=timestamp, proposal=proposal),
    )

    vehicle_load_sheet = workbook.create_sheet("Vehicle Load")
    _append_sheet_rows(vehicle_load_sheet, _build_vehicle_load_rows(snapshot=snapshot, proposal=proposal))

    snapshot_requests_sheet = workbook.create_sheet("Snapshot Requests")
    _append_sheet_rows(snapshot_requests_sheet, _build_snapshot_request_rows(snapshot))

    if proposal is not None:
        proposed_decisions_sheet = workbook.create_sheet("Proposed Decisions")
        _append_sheet_rows(
            proposed_decisions_sheet,
            _build_proposed_decision_rows(snapshot=snapshot, proposal=proposal),
        )

    exceptions_sheet = workbook.create_sheet("Exceptions")
    _append_sheet_rows(exceptions_sheet, _build_exception_rows(snapshot=snapshot, proposal=proposal))

    if proposal is not None:
        audit_trail_sheet = workbook.create_sheet("Audit Trail")
        _append_sheet_rows(audit_trail_sheet, _build_audit_rows(proposal))

    if ai_suggestion is not None:
        ai_summary_sheet = workbook.create_sheet("AI Summary")
        _append_sheet_rows(ai_summary_sheet, _build_transport_ai_summary_rows(ai_suggestion, run=ai_run))

        ai_vehicle_actions_sheet = workbook.create_sheet("AI Vehicle Actions")
        _append_sheet_rows(ai_vehicle_actions_sheet, _build_transport_ai_vehicle_action_rows(ai_suggestion))

        ai_itineraries_sheet = workbook.create_sheet("AI Itineraries")
        _append_sheet_rows(ai_itineraries_sheet, _build_transport_ai_itinerary_rows(ai_suggestion))

        ai_issues_sheet = workbook.create_sheet("AI Issues")
        _append_sheet_rows(ai_issues_sheet, _build_transport_ai_issue_rows(ai_suggestion))

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    content = output.getvalue()
    output.close()

    export_path = _resolve_transport_export_path(file_name)
    export_path.write_bytes(content)
    return export_path.name, content
