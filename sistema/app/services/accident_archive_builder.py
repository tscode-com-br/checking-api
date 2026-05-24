"""Archive builder — generates XLSX + ZIP for a closed accident and stores in object storage.

Called as a background task after an accident is closed.  The resulting archive is
attached to the Accident record (archive_object_key) and an AccidentArchive row is
created with metadata.
"""

from __future__ import annotations

import json
import re
import zipfile
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from sqlalchemy import select

from ..database import SessionLocal
from ..models import Accident, AccidentArchive, AccidentVideoUpload
from .accident_numbering import format_accident_number
from .accident_situation_table import build_situation_rows
from .admin_updates import notify_admin_data_changed
from .object_storage import _local_root, _use_remote, upload_stream
from .time_utils import now_sgt

# Column order in the XLSX (A=1, B=2, …)
# A  Horário
# B  Atividade/Local
# C  Nome
# D  Chave
# E  Projetos
# F  Local
# G  Ciência
# H  Zona de
# I  Situação
# J  Contato
# K  Registros

COLUMN_ORDER = [
    "Horário",
    "Atividade/Local",
    "Nome",
    "Chave",
    "Projetos",
    "Local",
    "Ciência",
    "Zona de",
    "Situação",
    "Contato",
    "Registros",
]

_COL_REGISTROS = len(COLUMN_ORDER)  # 1-based index of "Registros" column


def _slugify(value: str) -> str:
    """Convert an arbitrary string to a safe filename segment."""
    return re.sub(r"[^A-Za-z0-9_-]+", "_", value)[:60]


def _build_xlsx(
    accident: Accident,
    snapshot_rows,
    video_files_by_user_chave: dict[str, list[str]],
) -> BytesIO:
    """Build the 'Situação de Pessoal' spreadsheet and return it as a BytesIO."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Situacao de Pessoal"

    # Metadata header rows
    bold = Font(bold=True)
    header_rows = [
        (f"Acidente N.º: {format_accident_number(accident.accident_number)}", ""),
        (f"Projeto: {accident.project_name_snapshot}", ""),
        (f"Local: {accident.location_name_snapshot}", ""),
        (f"Data abertura: {accident.opened_at.strftime('%d/%m/%Y %H:%M') if accident.opened_at else ''}", ""),
        (f"Descrição: {accident.description or '(sem descrição)'}", ""),
        ("", ""),  # blank separator
    ]
    for label, _ in header_rows:
        ws.append([label])
        ws.cell(row=ws.max_row, column=1).font = bold

    # Column headers
    ws.append(COLUMN_ORDER)
    for col_idx in range(1, len(COLUMN_ORDER) + 1):
        ws.cell(row=ws.max_row, column=col_idx).font = bold

    for row in snapshot_rows:
        user_chave = row.chave
        videos = video_files_by_user_chave.get(user_chave, [])
        registros_text = "\n".join(f"Registros/{user_chave}/{filename}" for filename in videos)

        ciencia = "Ciente" if row.awareness_status == "acknowledged" else "Aguardando"

        ws.append([
            row.event_time.isoformat(),
            row.activity_local or "",
            row.name,
            row.chave,
            ", ".join(row.projects),
            row.local or "",
            ciencia,
            row.zone,
            row.status,
            row.phone or "",
            registros_text,
        ])
        cell = ws.cell(row=ws.max_row, column=_COL_REGISTROS)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if videos:
            cell.hyperlink = f"Registros/{user_chave}/{videos[0]}"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _read_video_bytes(object_key: str) -> bytes:
    """Fetch raw video bytes from storage (local or remote)."""
    if _use_remote():
        from ..core.config import settings
        from .object_storage import _make_boto3_client

        client = _make_boto3_client()
        result = client.get_object(Bucket=settings.do_spaces_bucket, Key=object_key)
        return result["Body"].read()

    target = _local_root() / object_key
    return target.read_bytes() if target.exists() else b""


def build_and_attach_archive_for_accident(accident_id: int) -> None:
    """Build XLSX + ZIP archive for *accident_id*, upload to storage, persist metadata."""
    with SessionLocal() as db:
        accident = db.get(Accident, accident_id)
        if accident is None:
            return

        snapshot_rows = build_situation_rows(db, accident=accident)

        # Build chave lookup from snapshot_rows
        chave_by_user_id: dict[int, str] = {row.user_id: row.chave for row in snapshot_rows}

        videos = (
            db.execute(
                select(AccidentVideoUpload).where(
                    AccidentVideoUpload.accident_id == accident.id
                )
            )
            .scalars()
            .all()
        )

        # Build user_chave -> [filename, ...] and ZIP payloads
        # Filenames use zero-padded index within each user's videos
        video_files_by_user_chave: dict[str, list[str]] = {}
        video_payloads: dict[str, bytes] = {}  # key = full zip path

        # Group videos by user to generate per-user sequential index
        videos_by_user: dict[int, list[AccidentVideoUpload]] = {}
        for video in videos:
            videos_by_user.setdefault(video.user_id, []).append(video)

        for user_id, user_videos in videos_by_user.items():
            user_chave = chave_by_user_id.get(user_id) or str(user_id)
            for idx, video in enumerate(user_videos, start=1):
                ext = video.content_type.split("/")[-1]
                if ext == "quicktime":
                    ext = "mov"
                filename = f"{idx:02d}_{_slugify(video.idempotency_key)}.{ext}"
                zip_path = f"Registros/{user_chave}/{filename}"
                video_files_by_user_chave.setdefault(user_chave, []).append(filename)
                video_payloads[zip_path] = _read_video_bytes(video.object_key)

        xlsx_buffer = _build_xlsx(accident, snapshot_rows, video_files_by_user_chave)

        # Build ZIP
        zip_buffer = BytesIO()
        xlsx_name = f"{format_accident_number(accident.accident_number)}.xlsx"
        with zipfile.ZipFile(zip_buffer, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(xlsx_name, xlsx_buffer.getvalue())
            for zip_path, payload in video_payloads.items():
                zf.writestr(zip_path, payload)
        zip_buffer.seek(0)

        # Upload XLSX and ZIP to storage
        acc_label = format_accident_number(accident.accident_number)
        xlsx_key = f"accidents/{acc_label}/archive/{xlsx_name}"
        zip_key = f"accidents/{acc_label}/archive/{acc_label}.zip"

        upload_stream(
            object_key=xlsx_key,
            stream=BytesIO(xlsx_buffer.getvalue()),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        upload_stream(
            object_key=zip_key,
            stream=zip_buffer,
            content_type="application/zip",
        )

        size_bytes = zip_buffer.seek(0, 2) or 0
        zip_buffer.seek(0)

        archive = AccidentArchive(
            accident_id=accident.id,
            snapshot_json=json.dumps(
                [row.model_dump() for row in snapshot_rows], default=str
            ),
            xlsx_object_key=xlsx_key,
            zip_object_key=zip_key,
            size_bytes=size_bytes,
            generated_at=now_sgt(),
        )
        accident.archive_object_key = zip_key
        db.add(archive)
        db.commit()

    notify_admin_data_changed(
        "accident_closed",
        metadata={"accident_id": accident_id, "archive_ready": True},
    )
