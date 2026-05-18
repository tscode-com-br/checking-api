import asyncio
import json
from io import BytesIO
from datetime import date, datetime, time as dt_time, timedelta
from uuid import uuid4

from fastapi import APIRouter, BackgroundTasks, Body, Depends, HTTPException, Query, Request
from fastapi.exceptions import RequestValidationError
from fastapi.responses import FileResponse, RedirectResponse, Response, StreamingResponse
from pydantic import ValidationError
from sqlalchemy import asc, delete, desc, func, or_, select, update
from sqlalchemy.orm import Session

from ..database import get_db
from ..database import get_database_diagnostics
from ..models import (
    Accident,
    AccidentArchive,
    AdminAccessRequest,
    AdminUser,
    CheckEvent,
    CheckingHistory,
    ManagedLocation,
    PendingRegistration,
    Project,
    ProjectAutoCheckoutDistance,
    TransportAssignment,
    TransportRequest,
    Workplace,
    User,
    UserSyncEvent,
    Vehicle,
)
from ..schemas import (
    AccidentClosedListResponse,
    AccidentClosedRow,
    AccidentLocationOption,
    AccidentProjectOption,
    AccidentSummary,
    AdminAccidentOpenRequest,
    AdminAccidentStateResponse,
    AdminAccessRequestCreate,
    AdminActionResponse,
    DatabaseDiagnosticsResponse,
    FormsQueueDiagnosticsResponse,
    AdminPasswordVerifyResponse,
    AdminProfileUpdateRequest,
    AdminProjectMinimumCheckoutDistanceSaveResponse,
    AdminProjectMinimumCheckoutDistanceUpdate,
    AdminProjectMinimumCheckoutDistanceListResponse,
    AdminProjectMinimumCheckoutDistanceRow,
    AdminSelfAccessRequest,
    AdminSelfAccessStatusResponse,
    DatabaseEventFilterOptions,
    DatabaseEventListResponse,
    AdminIdentity,
    AdminLocationAuditResponse,
    AdminLocationsResponse,
    AdminLocationSettingsResponse,
    AdminLocationSettingsUpdate,
    AdminLocationUpsert,
    AdminLoginRequest,
    AdminManagementRow,
    AdminPasswordResetRequest,
    AdminSelfPasswordChangeRequest,
    AdminSelfPasswordVerifyRequest,
    AdminPasswordSetRequest,
    ProjectCreate,
    ProjectRow,
    ProjectUpdate,
    AdminSessionResponse,
    AdminUserListRow,
    AdminUserUpsert,
    EventArchiveCreateResponse,
    EventArchiveListResponse,
    EventArchiveRow,
    EventRow,
    InactiveUserRow,
    LocationRow,
    PendingRow,
    ProviderFormRow,
    ReportEventRow,
    ReportEventsResponse,
    ReportPersonRow,
    UserRow,
)
from ..services.admin_auth import (
    ADMIN_ACCESS_SCOPE_FULL,
    ADMIN_ACCESS_SCOPE_LIMITED,
    ADMIN_ACCESS_DIGIT,
    TRANSPORT_ACCESS_DIGIT,
    add_profile_access,
    clear_admin_session,
    describe_user_profile,
    get_admin_access_scope,
    get_admin_allowed_tabs,
    get_authenticated_admin_from_session,
    hash_password,
    normalize_admin_key,
    normalize_user_profile,
    remove_profile_access,
    require_full_admin_session,
    require_admin_session,
    require_admin_stream_session,
    user_has_admin_access,
    user_can_access_admin_panel,
    user_can_view_activity_time,
    user_profile_has_access,
    verify_password,
)
from ..services.admin_updates import admin_updates_broker, notify_admin_data_changed, notify_web_check_data_changed
from ..services.admin_project_scope import (
    location_matches_effective_admin_scope,
    project_matches_effective_admin_scope,
    extract_admin_monitored_projects,
    resolve_effective_admin_project_names,
    user_matches_effective_admin_scope,
)
from ..services.event_archives import (
    build_event_archives_zip,
    create_event_archive,
    delete_event_archive,
    get_event_archive_path,
    list_event_archives_page,
)
from ..services.event_logger import log_event
from ..services.forms_queue import get_forms_queue_diagnostics
from ..services.managed_locations import (
    dump_location_coordinates,
    dump_location_projects,
    extract_location_coordinates,
    extract_location_projects,
)
from ..services.location_audit import audit_locations_from_db
from ..services.location_settings import (
    get_location_accuracy_threshold_meters,
    get_mixed_zone_interval_minutes,
    list_project_minimum_checkout_distance_rows,
    upsert_project_minimum_checkout_distance_rows,
    upsert_location_settings,
)
from ..services.project_catalog import (
    build_project_fields,
    ensure_known_project,
    ensure_known_projects,
    list_project_names,
    list_projects,
    resolve_default_project_name,
)
from ..services.time_utils import build_timezone_context, build_timezone_label, now_sgt
from ..services.transport_vehicle_base import (
    resolve_vehicle_for_user_transport_link,
    sync_user_vehicle_reference,
)
from ..services.user_activity import (
    calculate_inactivity_days,
    has_exceeded_continuous_inactivity_window,
    is_user_inactive,
    sync_user_inactivity,
)
from ..services.user_projects import (
    add_user_project_membership,
    ensure_user_active_project_is_member,
    list_materialized_user_project_names,
    list_user_project_names,
    list_user_project_names_map,
    normalize_user_project_names,
    replace_user_project_memberships,
    resolve_user_active_project,
)
from ..services.user_sync import find_user_by_chave, find_user_by_rfid, resolve_latest_user_activities, resolve_latest_user_activity
from ..services.accident_lifecycle import (
    AccidentAlreadyActiveError,
    InvalidAccidentLocationError,
    NoActiveAccidentError,
    close_accident,
    list_active_accident,
    open_accident,
)
from ..services.accident_numbering import format_accident_number
from ..services.accident_archive_builder import build_and_attach_archive_for_accident
from ..services.accident_situation_table import build_situation_rows

router = APIRouter(prefix="/api/admin", tags=["admin"])

EVENT_KEY_FIELDS = ("approved_by", "rejected_by", "revoked_by", "updated_by", "chave")
DATABASE_EVENT_ACTIONS = ("checkin", "checkout")
PROVIDER_FORMS_REQUEST_PATH = "/api/provider/updaterecords"
DATABASE_EVENT_PAGE_SIZE = 50
DATABASE_EVENT_DEFAULT_SORT_BY = "event_time"
DATABASE_EVENT_DEFAULT_SORT_DIRECTION = "desc"
DATABASE_EVENT_SQL_SORT_FIELDS = {
    "id": CheckEvent.id,
    "event_time": CheckEvent.event_time,
    "action": func.lower(func.coalesce(CheckEvent.action, "")),
    "rfid": func.lower(func.coalesce(CheckEvent.rfid, "")),
    "project": func.lower(func.coalesce(CheckEvent.project, "")),
    "local": func.lower(func.coalesce(CheckEvent.local, "")),
    "source": func.lower(func.coalesce(CheckEvent.source, "")),
    "status": func.lower(func.coalesce(CheckEvent.status, "")),
    "http_status": func.coalesce(CheckEvent.http_status, -1),
    "device_id": func.lower(func.coalesce(CheckEvent.device_id, "")),
    "message": func.lower(func.coalesce(CheckEvent.message, "")),
    "details": func.lower(func.coalesce(CheckEvent.details, "")),
}
DATABASE_EVENT_SORTABLE_FIELDS = frozenset((*DATABASE_EVENT_SQL_SORT_FIELDS.keys(), "chave"))
DATABASE_EVENT_SORT_DIRECTIONS = frozenset(("asc", "desc"))
REPORT_EXPORT_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
REPORT_EXPORT_COLUMNS_WITH_TIME = (
    "Data",
    "Horário",
    "Ação",
    "Origem",
    "Local",
    "Projeto",
    "Fuso Horário",
    "Assiduidade",
)
REPORT_EXPORT_COLUMNS_WITHOUT_TIME = (
    "Data",
    "Ação",
    "Origem",
    "Local",
    "Projeto",
    "Fuso Horário",
    "Assiduidade",
)
REPORT_EXPORT_ALL_COLUMNS_WITH_TIME = (
    "Nome",
    *REPORT_EXPORT_COLUMNS_WITH_TIME,
)
REPORT_EXPORT_ALL_COLUMNS_WITHOUT_TIME = (
    "Nome",
    *REPORT_EXPORT_COLUMNS_WITHOUT_TIME,
)
REPORT_SOURCE_LABELS = {
    "web": "Aplicativo",
    "device": "Box ESP32-0001",
    "provider": "Forms",
}
REPORT_ACTION_LABELS = {
    "checkin": "Check-In",
    "checkout": "Check-Out",
    "register": "Cadastro",
    "admin_request": "Solicitação Admin",
    "admin_access": "Admin",
}
REPORT_LOCAL_LABELS = {
    "main": "Escritório Principal",
    "co80": "Escritório Avançado P80",
    "un80": "A bordo da P80",
    "co83": "Escritório Avançado P83",
    "un83": "A bordo da P83",
}


def format_assiduidade_label(ontime: bool | None) -> str:
    return "Retroativo" if ontime is False else "Normal"


def build_presence_activity_fields(
    *,
    event_time: datetime,
    timezone_context,
    can_view_activity_time: bool,
) -> tuple[datetime | None, str, str | None, str]:
    localized_event_time = event_time.astimezone(timezone_context.timezone)
    return (
        event_time if can_view_activity_time else None,
        localized_event_time.strftime("%d/%m/%Y"),
        localized_event_time.strftime("%H:%M:%S") if can_view_activity_time else None,
        localized_event_time.strftime("%Y-%m-%d"),
    )


def build_report_event_time_fields(
    *,
    event_time: datetime,
    timezone_context,
    can_view_activity_time: bool,
) -> tuple[datetime | None, str | None, str]:
    localized_event_time = event_time.astimezone(timezone_context.timezone)
    return (
        event_time if can_view_activity_time else None,
        localized_event_time.strftime("%H:%M:%S") if can_view_activity_time else None,
        localized_event_time.strftime("%d/%m/%Y"),
    )


def build_database_event_time_fields(
    *,
    event_time: datetime,
    timezone_context,
    can_view_activity_time: bool,
) -> tuple[datetime | None, str, str | None]:
    localized_event_time = event_time.astimezone(timezone_context.timezone)
    return (
        event_time if can_view_activity_time else None,
        localized_event_time.strftime("%d/%m/%Y"),
        localized_event_time.strftime("%H:%M:%S") if can_view_activity_time else None,
    )


def format_quantity(value: int, singular: str, plural: str) -> str:
    return f"{value} {singular if value == 1 else plural}"


def format_report_source_label(source: str | None) -> str:
    normalized = str(source or "").strip().lower()
    if not normalized:
        return "-"
    return REPORT_SOURCE_LABELS.get(normalized, source or "-")


def format_report_action_label(action: str | None) -> str:
    normalized = str(action or "").strip().lower()
    if not normalized:
        return "-"
    return REPORT_ACTION_LABELS.get(normalized, action or "-")


def format_report_local_label(local: str | None) -> str:
    normalized = str(local or "").strip().lower()
    if not normalized:
        return "-"
    return REPORT_LOCAL_LABELS.get(normalized, local or "-")


def build_report_export_file_name(*, user: User, timestamp: datetime) -> str:
    return f"Relatorio - {user.chave} - {timestamp:%Y%m%d - %H%M%S}.xlsx"


def build_report_export_all_file_name(*, timestamp: datetime) -> str:
    return f"Relatorio - Todos - {timestamp:%Y%m%d - %H%M%S}.xlsx"


def build_report_export_metadata(*, report: ReportEventsResponse) -> str:
    person = report.person
    events_count = len(report.events)
    events_label = "1 evento" if events_count == 1 else f"{events_count} eventos"
    return (
        f"Projeto atual: {person.projeto or '-'} | RFID: {person.rfid or '-'} | "
        f"Fuso horário: {person.timezone_label or '-'} | {events_label}"
    )


def build_report_export_columns(*, can_view_activity_time: bool) -> tuple[str, ...]:
    return REPORT_EXPORT_COLUMNS_WITH_TIME if can_view_activity_time else REPORT_EXPORT_COLUMNS_WITHOUT_TIME


def build_report_export_all_columns(*, can_view_activity_time: bool) -> tuple[str, ...]:
    return REPORT_EXPORT_ALL_COLUMNS_WITH_TIME if can_view_activity_time else REPORT_EXPORT_ALL_COLUMNS_WITHOUT_TIME


def build_report_projects_by_name(db: Session, project_names: set[str]) -> dict[str, Project]:
    normalized_project_names = sorted(
        {
            str(project_name or "").strip().upper()
            for project_name in project_names
            if str(project_name or "").strip()
        }
    )
    if not normalized_project_names:
        return {}

    return {
        project.name: project
        for project in db.execute(select(Project).where(Project.name.in_(normalized_project_names))).scalars().all()
    }


def resolve_report_event_project_name(event: UserSyncEvent, *, user: User | None = None) -> str:
    return str(event.projeto or (user.projeto if user is not None else None) or "").strip().upper() or "-"


def build_report_event_row(
    event: UserSyncEvent,
    *,
    user: User | None = None,
    projects_by_name: dict[str, Project],
    can_view_activity_time: bool = True,
) -> ReportEventRow:
    project_name = resolve_report_event_project_name(event, user=user)
    project = projects_by_name.get(project_name) if project_name != "-" else None
    timezone_context = build_timezone_context(
        project_name=project.name if project is not None else (None if project_name == "-" else project_name),
        country_name=project.country_name if project is not None else None,
        timezone_name=project.timezone_name if project is not None else None,
        reference_time=event.event_time,
    )
    raw_event_time, event_time_label, event_date = build_report_event_time_fields(
        event_time=event.event_time,
        timezone_context=timezone_context,
        can_view_activity_time=can_view_activity_time,
    )
    return ReportEventRow(
        id=event.id,
        source=event.source,
        source_label=format_report_source_label(event.source),
        action=event.action,
        action_label=format_report_action_label(event.action),
        projeto=project_name,
        local=event.local,
        local_label=format_report_local_label(event.local),
        ontime=event.ontime,
        assiduidade=format_assiduidade_label(event.ontime),
        event_time=raw_event_time,
        event_time_label=event_time_label,
        timezone_name=timezone_context.timezone_name,
        timezone_label=timezone_context.timezone_label,
        event_date=event_date,
    )


def build_report_event_export_values(
    row: ReportEventRow,
    *,
    can_view_activity_time: bool,
) -> tuple[str, ...]:
    values = [row.event_date or "-"]
    if can_view_activity_time:
        values.append(row.event_time_label or "-")
    values.extend(
        [
            row.action_label or format_report_action_label(row.action),
            row.source_label or format_report_source_label(row.source),
            row.local_label or format_report_local_label(row.local),
            row.projeto or "-",
            row.timezone_label or "-",
            row.assiduidade or "Normal",
        ]
    )
    return tuple(values)


def build_report_events_export(
    *,
    user: User,
    report: ReportEventsResponse,
    can_view_activity_time: bool,
) -> tuple[str, bytes]:
    from openpyxl import Workbook

    timestamp = now_sgt()
    file_name = build_report_export_file_name(user=user, timestamp=timestamp)
    columns = build_report_export_columns(can_view_activity_time=can_view_activity_time)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Relatório"

    worksheet.append([f"{report.person.nome or '-'} ({report.person.chave or '-'})"])
    worksheet.append([build_report_export_metadata(report=report)])
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    worksheet.append([])
    worksheet.append(list(columns))

    for row in report.events:
        worksheet.append(list(build_report_event_export_values(row, can_view_activity_time=can_view_activity_time)))

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    content = output.getvalue()
    output.close()
    return file_name, content


def build_all_report_events_export(
    db: Session,
    *,
    can_view_activity_time: bool,
    current_admin: User | None = None,
) -> tuple[str, bytes]:
    from openpyxl import Workbook

    rows = db.execute(
        select(UserSyncEvent, User)
        .join(User, User.id == UserSyncEvent.user_id, isouter=True)
        .order_by(
            asc(func.lower(func.coalesce(User.nome, ""))),
            asc(func.lower(func.coalesce(User.chave, ""))),
            desc(UserSyncEvent.event_time),
            desc(UserSyncEvent.id),
        )
    ).all()

    effective_admin_projects = (
        resolve_effective_admin_project_names(db, current_admin)
        if current_admin is not None
        else None
    )
    if current_admin is not None:
        if not effective_admin_projects:
            rows = []
        else:
            report_users = [user for _, user in rows if user is not None]
            user_project_names_by_user_id = list_user_project_names_map(db, report_users)
            rows = [
                (event, user)
                for event, user in rows
                if (
                    user is not None
                    and user_matches_effective_admin_scope(
                        db,
                        current_admin,
                        user,
                        admin_project_names=effective_admin_projects,
                        user_project_names=user_project_names_by_user_id.get(user.id, []),
                    )
                )
                or (
                    user is None
                    and project_matches_effective_admin_scope(
                        db,
                        current_admin,
                        event.projeto,
                        admin_project_names=effective_admin_projects,
                        allow_blank=False,
                    )
                )
            ]

    projects_by_name = build_report_projects_by_name(
        db,
        {
            *(event.projeto for event, _ in rows),
            *((user.projeto if user is not None else None) for _, user in rows),
        },
    )

    timestamp = now_sgt()
    file_name = build_report_export_all_file_name(timestamp=timestamp)
    columns = build_report_export_all_columns(can_view_activity_time=can_view_activity_time)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Relatório"
    worksheet.append(list(columns))

    for event, user in rows:
        report_row = build_report_event_row(
            event,
            user=user,
            projects_by_name=projects_by_name,
            can_view_activity_time=can_view_activity_time,
        )
        worksheet.append([
            (user.nome if user is not None and user.nome else "-"),
            *build_report_event_export_values(report_row, can_view_activity_time=can_view_activity_time),
        ])

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    content = output.getvalue()
    output.close()
    return file_name, content


def parse_event_details(details: str | None) -> dict[str, str]:
    if not details:
        return {}

    parsed: dict[str, str] = {}
    for part in str(details).split(";"):
        if "=" not in part:
            continue
        key, value = part.split("=", 1)
        normalized_key = key.strip()
        normalized_value = value.strip()
        if normalized_key and normalized_value:
            parsed[normalized_key] = normalized_value
    return parsed


def build_provider_forms_rows(
    db: Session,
    *,
    current_admin: User | None = None,
) -> list[ProviderFormRow]:
    rows = db.execute(
        select(UserSyncEvent, User)
        .join(User, User.id == UserSyncEvent.user_id, isouter=True)
        .where(
            UserSyncEvent.source == "provider",
            UserSyncEvent.action.in_(DATABASE_EVENT_ACTIONS),
        )
        .order_by(desc(UserSyncEvent.id))
    ).all()

    project_names = sorted(
        {
            (event.projeto or (user.projeto if user is not None else None) or "").strip().upper()
            for event, user in rows
            if (event.projeto or (user.projeto if user is not None else None))
        }
    )
    projects_by_name = {
        project.name: project
        for project in db.execute(select(Project).where(Project.name.in_(project_names))).scalars().all()
    } if project_names else {}

    can_view_activity_time = True if current_admin is None else user_can_view_activity_time(current_admin)
    payload: list[ProviderFormRow] = []
    for event, user in rows:
        project_name = (event.projeto or (user.projeto if user is not None else None) or "-").strip().upper()
        project = projects_by_name.get(project_name) if project_name != "-" else None
        timezone_context = build_timezone_context(
            project_name=project.name if project is not None else (None if project_name == "-" else project_name),
            country_name=project.country_name if project is not None else None,
            timezone_name=project.timezone_name if project is not None else None,
            reference_time=event.event_time,
        )
        localized_event_time = event.event_time.astimezone(timezone_context.timezone)
        payload.append(
            ProviderFormRow(
                recebimento=event.event_time if can_view_activity_time else None,
                recebimento_date_label=localized_event_time.strftime("%d/%m/%Y"),
                recebimento_time_label=localized_event_time.strftime("%H:%M:%S") if can_view_activity_time else None,
                chave=(event.chave or "-").upper(),
                nome=((user.nome if user is not None else "-") or "-").upper(),
                projeto=project_name,
                timezone_name=timezone_context.timezone_name,
                timezone_label=timezone_context.timezone_label,
                atividade="check-in" if event.action == "checkin" else "check-out",
                informe="retroativo" if event.ontime is False else "normal",
                data=localized_event_time.strftime("%d/%m/%Y"),
                hora=localized_event_time.strftime("%H:%M:%S") if can_view_activity_time else None,
            )
        )

    return payload


def _normalize_report_name_query(value: str | None) -> str | None:
    normalized = " ".join(str(value or "").strip().split())
    return normalized or None


def resolve_report_user(
    db: Session,
    *,
    chave: str | None,
    nome: str | None,
    current_admin: User | None = None,
) -> User:
    normalized_key = str(chave or "").strip().upper() or None
    normalized_name = _normalize_report_name_query(nome)

    if normalized_key and normalized_name:
        raise HTTPException(status_code=400, detail="Informe apenas chave ou nome para consultar o relatorio.")
    if not normalized_key and not normalized_name:
        raise HTTPException(status_code=400, detail="Informe chave ou nome para consultar o relatorio.")

    if normalized_key:
        user = find_user_by_chave(db, normalized_key)
        if user is None or not user_matches_effective_admin_scope(db, current_admin, user):
            raise HTTPException(status_code=404, detail="Usuario nao encontrado para a chave informada.")
        return user

    assert normalized_name is not None
    matches = db.execute(
        select(User)
        .where(func.lower(User.nome) == normalized_name.lower())
        .order_by(User.nome, User.id)
    ).scalars().all()
    if current_admin is not None:
        matches = [user for user in matches if user_matches_effective_admin_scope(db, current_admin, user)]
    if not matches:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado para o nome informado.")
    if len(matches) > 1:
        raise HTTPException(
            status_code=409,
            detail="Mais de um usuario encontrado para o nome informado. Use a chave.",
        )
    return matches[0]


def build_report_events_response(db: Session, *, user: User, can_view_activity_time: bool = True) -> ReportEventsResponse:
    rows = db.execute(
        select(UserSyncEvent)
        .where(UserSyncEvent.user_id == user.id)
        .order_by(desc(UserSyncEvent.event_time), desc(UserSyncEvent.id))
    ).scalars().all()

    projects_by_name = build_report_projects_by_name(db, {user.projeto, *(row.projeto for row in rows)})
    person_project_names = list_user_project_names(db, user)

    person_project_name = str(user.projeto or "").strip().upper() or "-"
    person_project = projects_by_name.get(person_project_name) if person_project_name != "-" else None
    person_reference_time = rows[0].event_time if rows else now_sgt()
    person_timezone_context = build_timezone_context(
        project_name=person_project.name if person_project is not None else (None if person_project_name == "-" else person_project_name),
        country_name=person_project.country_name if person_project is not None else None,
        timezone_name=person_project.timezone_name if person_project is not None else None,
        reference_time=person_reference_time,
    )

    payload: list[ReportEventRow] = []
    for row in rows:
        payload.append(
            build_report_event_row(
                row,
                user=user,
                projects_by_name=projects_by_name,
                can_view_activity_time=can_view_activity_time,
            )
        )

    return ReportEventsResponse(
        person=ReportPersonRow(
            id=user.id,
            rfid=user.rfid,
            nome=user.nome,
            chave=user.chave,
            projeto=person_project_name,
            projetos=normalize_user_project_names(person_project_names or [person_project_name]),
            timezone_name=person_timezone_context.timezone_name,
            timezone_label=person_timezone_context.timezone_label,
        ),
        events=payload,
    )


def delete_provider_forms_rows(db: Session) -> int:
    result = db.execute(
        delete(UserSyncEvent).where(
            UserSyncEvent.source == "provider",
            UserSyncEvent.action.in_(DATABASE_EVENT_ACTIONS),
        )
    )
    return max(int(result.rowcount or 0), 0)


def resolve_event_key(event: CheckEvent, *, user_keys_by_rfid: dict[str, str]) -> str | None:
    details_map = parse_event_details(event.details)
    for field_name in EVENT_KEY_FIELDS:
        field_value = details_map.get(field_name)
        if field_value:
            return field_value.upper()
    if event.rfid:
        return user_keys_by_rfid.get(event.rfid)
    return None


def build_event_row_payload(
    rows: list[CheckEvent],
    db: Session,
    *,
    can_view_activity_time: bool = True,
) -> list[EventRow]:
    rfids = sorted({row.rfid for row in rows if row.rfid})
    project_names = sorted({row.project for row in rows if row.project})
    user_keys_by_rfid: dict[str, str] = {}
    projects_by_name: dict[str, Project] = {}
    if rfids:
        user_keys_by_rfid = {
            rfid: chave
            for rfid, chave in db.execute(select(User.rfid, User.chave).where(User.rfid.in_(rfids))).all()
            if rfid is not None
        }
    if project_names:
        projects_by_name = {
            project.name: project
            for project in db.execute(select(Project).where(Project.name.in_(project_names))).scalars().all()
        }

    payload: list[EventRow] = []
    for row in rows:
        project = projects_by_name.get(row.project) if row.project else None
        timezone_context = build_timezone_context(
            project_name=project.name if project is not None else row.project,
            country_name=project.country_name if project is not None else None,
            timezone_name=project.timezone_name if project is not None else None,
            reference_time=row.event_time,
        )
        raw_event_time, event_date_label, event_time_label = build_database_event_time_fields(
            event_time=row.event_time,
            timezone_context=timezone_context,
            can_view_activity_time=can_view_activity_time,
        )
        payload.append(
            EventRow(
                id=row.id,
                source=row.source,
                rfid=row.rfid,
                chave=resolve_event_key(row, user_keys_by_rfid=user_keys_by_rfid),
                device_id=row.device_id,
                local=row.local,
                action=row.action,
                status=row.status,
                message=row.message,
                details=row.details,
                project=row.project,
                ontime=row.ontime,
                request_path=row.request_path,
                http_status=row.http_status,
                retry_count=row.retry_count,
                event_time=raw_event_time,
                event_date_label=event_date_label,
                event_time_label=event_time_label,
                timezone_name=timezone_context.timezone_name,
                timezone_label=timezone_context.timezone_label,
            )
        )
    return payload


def get_database_event_sort_value(row: EventRow, sort_by: str) -> object:
    if sort_by == "id":
        return row.id
    if sort_by == "event_time":
        return row.event_time
    if sort_by == "http_status":
        return row.http_status if row.http_status is not None else -1
    if sort_by == "chave":
        return (row.chave or "").upper()
    if sort_by == "action":
        return row.action or ""
    if sort_by == "rfid":
        return row.rfid or ""
    if sort_by == "project":
        return row.project or ""
    if sort_by == "local":
        return row.local or ""
    if sort_by == "source":
        return row.source or ""
    if sort_by == "status":
        return row.status or ""
    if sort_by == "device_id":
        return row.device_id or ""
    if sort_by == "message":
        return row.message or ""
    if sort_by == "details":
        return row.details or ""
    return row.event_time


def sort_database_event_payload(items: list[EventRow], sort_by: str, sort_direction: str) -> list[EventRow]:
    reverse = sort_direction == "desc"
    return sorted(
        items,
        key=lambda row: (get_database_event_sort_value(row, sort_by), row.id),
        reverse=reverse,
    )


def build_database_event_filter_options(
    db: Session,
    *,
    allowed_project_names: list[str] | None = None,
) -> DatabaseEventFilterOptions:
    option_rows = db.execute(
        select(
            CheckEvent.rfid,
            CheckEvent.details,
            CheckEvent.action,
            CheckEvent.project,
            CheckEvent.source,
            CheckEvent.status,
        ).where(CheckEvent.action.in_(DATABASE_EVENT_ACTIONS))
    ).all()

    rfids = sorted({rfid for rfid, *_ in option_rows if rfid})
    user_keys_by_rfid: dict[str, str] = {}
    if rfids:
        user_keys_by_rfid = {
            rfid: chave
            for rfid, chave in db.execute(select(User.rfid, User.chave).where(User.rfid.in_(rfids))).all()
            if rfid is not None
        }

    actions: set[str] = set()
    keys: set[str] = set()
    seen_rfids: set[str] = set()
    projects: set[str] = set()
    sources: set[str] = set()
    statuses: set[str] = set()

    for rfid, details, action, project, source, status in option_rows:
        if action:
            actions.add(action)
        if rfid:
            seen_rfids.add(rfid)
        if project:
            projects.add(project)
        if source:
            sources.add(source)
        if status:
            statuses.add(status)

        resolved_key = None
        details_map = parse_event_details(details)
        for field_name in EVENT_KEY_FIELDS:
            field_value = details_map.get(field_name)
            if field_value:
                resolved_key = field_value.upper()
                break
        if resolved_key is None and rfid:
            resolved_key = user_keys_by_rfid.get(rfid)
        if resolved_key:
            keys.add(resolved_key)

    return DatabaseEventFilterOptions(
        action=sorted(actions),
        chave=sorted(keys),
        rfid=sorted(seen_rfids),
        project=(
            [project_name for project_name in sorted(projects) if project_name in set(allowed_project_names)]
            if allowed_project_names is not None
            else sorted(projects)
        ),
        source=sorted(sources),
        status=sorted(statuses),
    )


def build_location_settings_log_message(
    *,
    previous_accuracy_threshold_meters: int,
    current_accuracy_threshold_meters: int,
    previous_mixed_zone_interval_minutes: int,
    current_mixed_zone_interval_minutes: int,
) -> str:
    changes: list[str] = []
    if previous_accuracy_threshold_meters != current_accuracy_threshold_meters:
        changes.append(
            "O valor do erro máximo para considerar a coordenada do usuário foi ajustado para "
            f"{format_quantity(current_accuracy_threshold_meters, 'metro', 'metros')}."
        )
    if previous_mixed_zone_interval_minutes != current_mixed_zone_interval_minutes:
        changes.append(
            "O intervalo de tempo para Zona Mista foi ajustado para "
            f"{format_quantity(current_mixed_zone_interval_minutes, 'minuto', 'minutos')}."
        )
    if changes:
        return " ".join(changes)
    return "As configurações de localização foram salvas sem alterações."


def build_presence_rows(
    db: Session,
    *,
    action: str,
    current_admin: User | None = None,
    reference_time=None,
) -> list[UserRow]:
    all_projects = list_projects(db)
    current_time = reference_time or now_sgt()
    can_view_activity_time = True if current_admin is None else user_can_view_activity_time(current_admin)
    rows = db.execute(select(User).order_by(User.nome, User.id)).scalars().all()
    rows, _, effective_admin_projects = filter_users_for_admin_scope(
        db,
        rows,
        current_admin=current_admin,
    )
    if current_admin is not None and effective_admin_projects == []:
        return []

    latest_activities = resolve_latest_user_activities(db, users=rows)
    project_names_by_user_id = list_user_project_names_map(db, rows)
    projects_by_name = {project.name: project for project in all_projects}
    payload: list[tuple[datetime, UserRow]] = []

    for user in rows:
        latest_activity = latest_activities.get(user.id)
        if latest_activity is None or latest_activity.action != action:
            continue
        if is_user_inactive(latest_activity.event_time, reference_time=current_time):
            continue
        project = projects_by_name.get(user.projeto)
        timezone_context = build_timezone_context(
            project_name=project.name if project is not None else user.projeto,
            country_name=project.country_name if project is not None else None,
            timezone_name=project.timezone_name if project is not None else None,
            reference_time=latest_activity.event_time,
        )
        raw_time, activity_date_label, activity_time_label, activity_day_key = build_presence_activity_fields(
            event_time=latest_activity.event_time,
            timezone_context=timezone_context,
            can_view_activity_time=can_view_activity_time,
        )

        payload.append(
            (
                latest_activity.event_time,
                UserRow(
                id=user.id,
                rfid=user.rfid,
                nome=user.nome,
                chave=user.chave,
                projeto=user.projeto,
                projetos=normalize_user_project_names(project_names_by_user_id.get(user.id, [user.projeto])),
                timezone_name=timezone_context.timezone_name,
                timezone_label=timezone_context.timezone_label,
                local=latest_activity.local if latest_activity.local is not None else user.local,
                checkin=action == "checkin",
                time=raw_time,
                activity_date_label=activity_date_label,
                activity_time_label=activity_time_label,
                activity_day_key=activity_day_key,
                assiduidade=format_assiduidade_label(latest_activity.ontime),
                ),
            )
        )

    payload.sort(key=lambda item: item[0], reverse=True)
    return [row for _, row in payload]


def build_missing_checkout_rows(db: Session, *, reference_time=None) -> list[UserRow]:
    return []


def build_inactive_rows(
    db: Session,
    *,
    current_admin: User | None = None,
    reference_time=None,
) -> list[InactiveUserRow]:
    rows = db.execute(select(User).order_by(User.nome, User.id)).scalars().all()
    payload: list[InactiveUserRow] = []
    current_time = reference_time or now_sgt()
    rows, _, effective_admin_projects = filter_users_for_admin_scope(
        db,
        rows,
        current_admin=current_admin,
    )
    effective_admin_project_set = set(effective_admin_projects) if effective_admin_projects is not None else None
    project_names_by_user_id = list_user_project_names_map(db, rows)
    project_names = sorted({user.projeto for user in rows if user.projeto})
    projects_by_name = {
        project.name: project
        for project in db.execute(select(Project).where(Project.name.in_(project_names))).scalars().all()
    } if project_names else {}

    for user in rows:
        latest_activity = resolve_latest_user_activity(db, user=user)
        if latest_activity is None:
            continue
        if not is_user_inactive(latest_activity.event_time, reference_time=current_time):
            continue

        inactivity_days = calculate_inactivity_days(latest_activity.event_time, reference_time=current_time)
        if has_exceeded_continuous_inactivity_window(latest_activity.event_time, reference_time=current_time) and inactivity_days < 1:
            inactivity_days = 1
        project = projects_by_name.get(user.projeto)
        timezone_context = build_timezone_context(
            project_name=project.name if project is not None else user.projeto,
            country_name=project.country_name if project is not None else None,
            timezone_name=project.timezone_name if project is not None else None,
            reference_time=latest_activity.event_time,
        )

        payload.append(
            InactiveUserRow(
                id=user.id,
                rfid=user.rfid,
                nome=user.nome,
                chave=user.chave,
                projeto=user.projeto,
                projetos=normalize_user_project_names(project_names_by_user_id.get(user.id, [user.projeto])),
                timezone_name=timezone_context.timezone_name,
                timezone_label=timezone_context.timezone_label,
                latest_action=latest_activity.action,
                latest_time=latest_activity.event_time,
                inactivity_days=inactivity_days,
            )
        )

    payload.sort(key=lambda row: (-row.inactivity_days, row.nome, row.chave))
    return payload


def notify_admin_views(*reasons: str) -> None:
    for reason in dict.fromkeys(reasons):
        notify_admin_data_changed(reason)


def encode_sse(payload: dict[str, str]) -> str:
    return f"data: {json.dumps(payload)}\n\n"


def build_admin_identity(admin: User) -> AdminIdentity:
    access_scope = get_admin_access_scope(admin)
    if access_scope not in {ADMIN_ACCESS_SCOPE_LIMITED, ADMIN_ACCESS_SCOPE_FULL}:
        access_scope = ADMIN_ACCESS_SCOPE_FULL
    return AdminIdentity(
        id=admin.id,
        chave=admin.chave,
        nome_completo=admin.nome,
        perfil=admin.perfil,
        can_view_activity_time=user_can_view_activity_time(admin),
        access_scope=access_scope,
        allowed_tabs=list(get_admin_allowed_tabs(admin)),
    )


def build_location_row(location: ManagedLocation) -> LocationRow:
    coordinates = extract_location_coordinates(location)
    primary_coordinate = coordinates[0]
    return LocationRow(
        id=location.id,
        local=location.local,
        latitude=primary_coordinate["latitude"],
        longitude=primary_coordinate["longitude"],
        coordinates=coordinates,
        projects=extract_location_projects(location),
        tolerance_meters=location.tolerance_meters,
    )


def format_location_coordinates_details(coordinates: list[dict[str, float]]) -> str:
    if not coordinates:
        return "-"
    return " | ".join(
        f"{coordinate['latitude']:.6f},{coordinate['longitude']:.6f}"
        for coordinate in coordinates
    )


def build_location_upsert_event_message(*, created: bool, geometry_changed: bool) -> str:
    if created:
        return "Location created via admin"
    if geometry_changed:
        return "Location geometry updated via admin"
    return "Location metadata updated via admin"


def build_location_upsert_event_details(
    *,
    current_admin: User,
    location_id: int | None,
    coordinates: list[dict[str, float]],
    projects: list[str],
    tolerance_meters: int,
    geometry_changed: bool,
    previous_coordinates: list[dict[str, float]] | None = None,
    previous_projects: list[str] | None = None,
    previous_tolerance_meters: int | None = None,
) -> str:
    details = [f"updated_by={current_admin.chave}"]
    if location_id is not None:
        details.append(f"location_id={location_id}")
    if previous_coordinates is not None:
        details.append(f"previous_coordinates={format_location_coordinates_details(previous_coordinates)}")
        details.append(f"previous_vertex_count={len(previous_coordinates)}")
    details.append(f"coordinates={format_location_coordinates_details(coordinates)}")
    details.append(f"vertex_count={len(coordinates)}")
    if previous_projects is not None:
        details.append(f"previous_projects={'|'.join(previous_projects) or '-'}")
    details.append(f"projects={'|'.join(projects) or '-'}")
    if previous_tolerance_meters is not None:
        details.append(f"previous_tolerance_meters={previous_tolerance_meters}")
    details.append(f"tolerance_meters={tolerance_meters}")
    details.append(f"geometry_changed={'yes' if geometry_changed else 'no'}")
    return "; ".join(details)


def normalize_location_validation_messages(errors: list[dict[str, object]]) -> str:
    messages: list[str] = []
    for error in errors:
        raw_message = str(error.get("msg") or error.get("message") or "Erro de validacao")
        normalized_message = raw_message.removeprefix("Value error, ").strip()
        if normalized_message and normalized_message not in messages:
            messages.append(normalized_message)
    return " | ".join(messages) if messages else "Erro de validacao"


def build_request_validation_errors(validation_error: ValidationError) -> list[dict[str, object]]:
    normalized_errors: list[dict[str, object]] = []
    for error in validation_error.errors():
        normalized_error = dict(error)
        location = normalized_error.get("loc")
        if isinstance(location, tuple):
            normalized_error["loc"] = ("body", *location)
        elif isinstance(location, list):
            normalized_error["loc"] = ["body", *location]
        normalized_errors.append(normalized_error)
    return normalized_errors


def log_location_validation_failure(
    db: Session,
    *,
    current_admin: User,
    payload: object,
    validation_message: str,
) -> None:
    payload_dict = payload if isinstance(payload, dict) else {}
    raw_local = str(payload_dict.get("local") or "").strip() or None
    raw_projects = payload_dict.get("projects")
    project_values = (
        [str(item).strip().upper() for item in raw_projects if str(item).strip()]
        if isinstance(raw_projects, list)
        else []
    )
    raw_coordinates = payload_dict.get("coordinates")
    coordinate_count = len(raw_coordinates) if isinstance(raw_coordinates, list) else 0
    tolerance_meters = payload_dict.get("tolerance_meters")
    details = [
        f"updated_by={current_admin.chave}",
        f"projects={'|'.join(project_values) or '-'}",
        f"coordinate_count={coordinate_count}",
        f"tolerance_meters={tolerance_meters if tolerance_meters is not None else '-'}",
        f"validation_errors={validation_message}",
    ]
    log_event(
        db,
        source="admin",
        action="location",
        status="failed",
        message="Location validation failed via admin",
        local=raw_local,
        request_path="/api/admin/locations",
        http_status=422,
        details="; ".join(details),
        commit=True,
    )


def build_project_row(project: Project) -> ProjectRow:
    return ProjectRow(
        id=project.id,
        name=project.name,
        country_code=project.country_code,
        country_name=project.country_name,
        timezone_name=project.timezone_name,
        timezone_label=build_timezone_label(
            country_name=project.country_name,
            timezone_name=project.timezone_name,
        ),
        address=str(project.address or "").strip(),
        zip_code=str(project.zip_code or "").strip(),
    )


def resolve_location_projects_for_upsert(
    db: Session,
    *,
    project_names: list[str],
    existing_location: ManagedLocation | None,
) -> list[str]:
    known_project_names = set(list_project_names(db))
    existing_location_projects = set(extract_location_projects(existing_location)) if existing_location is not None else set()

    resolved_projects: list[str] = []
    seen_projects: set[str] = set()
    for project_name in project_names:
        normalized_name = str(project_name or "").strip().upper()
        if not normalized_name or normalized_name in seen_projects:
            continue
        if normalized_name not in known_project_names and normalized_name not in existing_location_projects:
            raise HTTPException(status_code=422, detail="Projeto nao encontrado para a localizacao.")
        seen_projects.add(normalized_name)
        resolved_projects.append(normalized_name)

    if not resolved_projects:
        raise HTTPException(status_code=422, detail="Selecione ao menos um projeto para a localizacao.")

    return resolved_projects


def filter_users_for_admin_scope(
    db: Session,
    users: list[User],
    *,
    current_admin: User | None,
) -> tuple[list[User], dict[int, list[str]], list[str] | None]:
    project_names_by_user_id = list_user_project_names_map(db, users)
    effective_admin_projects = (
        resolve_effective_admin_project_names(db, current_admin)
        if current_admin is not None
        else None
    )
    if effective_admin_projects is None:
        return users, project_names_by_user_id, None
    if not effective_admin_projects:
        return [], project_names_by_user_id, []

    visible_users = [
        user
        for user in users
        if user_matches_effective_admin_scope(
            db,
            current_admin,
            user,
            admin_project_names=effective_admin_projects,
            user_project_names=project_names_by_user_id.get(user.id, []),
        )
    ]
    return visible_users, project_names_by_user_id, effective_admin_projects


def filter_check_events_for_admin_scope(
    db: Session,
    rows: list[CheckEvent],
    *,
    current_admin: User | None,
) -> list[CheckEvent]:
    effective_admin_projects = (
        resolve_effective_admin_project_names(db, current_admin)
        if current_admin is not None
        else None
    )
    if effective_admin_projects is None:
        return rows
    if not effective_admin_projects:
        return []

    rfids_without_project = sorted({row.rfid for row in rows if row.rfid and not row.project})
    users_by_rfid: dict[str, User] = {}
    user_project_names_by_user_id: dict[int, list[str]] = {}
    if rfids_without_project:
        scoped_users = db.execute(select(User).where(User.rfid.in_(rfids_without_project))).scalars().all()
        users_by_rfid = {user.rfid: user for user in scoped_users if user.rfid is not None}
        user_project_names_by_user_id = list_user_project_names_map(db, scoped_users)

    filtered_rows: list[CheckEvent] = []
    for row in rows:
        if project_matches_effective_admin_scope(
            db,
            current_admin,
            row.project,
            admin_project_names=effective_admin_projects,
            allow_blank=False,
        ):
            filtered_rows.append(row)
            continue

        if row.project:
            continue

        matched_user = users_by_rfid.get(str(row.rfid or "").strip())
        if matched_user is not None and user_matches_effective_admin_scope(
            db,
            current_admin,
            matched_user,
            admin_project_names=effective_admin_projects,
            user_project_names=user_project_names_by_user_id.get(matched_user.id, []),
        ):
            filtered_rows.append(row)
            continue

        if row.rfid is None:
            filtered_rows.append(row)

    return filtered_rows


def build_pending_registration_scope_maps(
    db: Session,
    pending_rows: list[PendingRegistration],
) -> tuple[dict[str, str], dict[str, list[str]]]:
    rfids = sorted({row.rfid for row in pending_rows if row.rfid})
    latest_scan_local_by_rfid: dict[str, str] = {}
    if rfids:
        scan_rows = db.execute(
            select(CheckEvent.rfid, CheckEvent.local)
            .where(CheckEvent.rfid.in_(rfids), CheckEvent.request_path == "/api/scan")
            .order_by(desc(CheckEvent.id))
        ).all()
        for rfid, local in scan_rows:
            normalized_rfid = str(rfid or "").strip()
            if not normalized_rfid or normalized_rfid in latest_scan_local_by_rfid:
                continue
            normalized_local = str(local or "").strip()
            if normalized_local:
                latest_scan_local_by_rfid[normalized_rfid] = normalized_local

    location_names = sorted({local for local in latest_scan_local_by_rfid.values() if local})
    location_projects_by_local: dict[str, list[str]] = {}
    if location_names:
        locations = db.execute(select(ManagedLocation).where(ManagedLocation.local.in_(location_names))).scalars().all()
        location_projects_by_local = {
            location.local: extract_location_projects(location)
            for location in locations
        }

    return latest_scan_local_by_rfid, location_projects_by_local


def pending_matches_admin_scope(
    db: Session,
    pending_row: PendingRegistration,
    *,
    current_admin: User | None,
    admin_project_names: list[str] | None = None,
    latest_scan_local_by_rfid: dict[str, str] | None = None,
    location_projects_by_local: dict[str, list[str]] | None = None,
) -> bool:
    effective_admin_projects = (
        admin_project_names
        if current_admin is not None
        else None
    )
    if current_admin is not None and effective_admin_projects is None:
        effective_admin_projects = resolve_effective_admin_project_names(db, current_admin) or []
    if effective_admin_projects is None:
        return True
    if not effective_admin_projects:
        return False

    latest_scan_local_by_rfid = latest_scan_local_by_rfid or {}
    location_projects_by_local = location_projects_by_local or {}
    pending_local = latest_scan_local_by_rfid.get(pending_row.rfid)
    if not pending_local:
        return True

    return location_matches_effective_admin_scope(
        db,
        current_admin,
        location_projects_by_local.get(pending_local, []),
        admin_project_names=effective_admin_projects,
        allow_global_locations=True,
    )


def normalize_administrator_profile(value: int | str | None) -> int:
    normalized = normalize_user_profile(value)
    if normalized == 9:
        return 9

    digits = {character for character in str(normalized) if character.isdigit() and character != "0"}
    digits.add(ADMIN_ACCESS_DIGIT)
    return int("".join(sorted(digits))) if digits else int(ADMIN_ACCESS_DIGIT)


def merge_user_profile_values(base_value: int | str | None, extra_value: int | str | None) -> int:
    normalized_base = normalize_user_profile(base_value)
    normalized_extra = normalize_user_profile(extra_value)
    if normalized_base == 9 or normalized_extra == 9:
        return 9

    digits = {
        character
        for character in f"{normalized_base}{normalized_extra}"
        if character.isdigit() and character != "0"
    }
    return int("".join(sorted(digits))) if digits else 0


def list_admin_rows(db: Session) -> list[AdminManagementRow]:
    admin_candidates = db.execute(
        select(User)
        .where(User.perfil != 0)
        .order_by(User.nome, User.chave)
    ).scalars().all()
    admins = [admin for admin in admin_candidates if user_has_admin_access(admin)]
    requests = db.execute(select(AdminAccessRequest).order_by(AdminAccessRequest.requested_at.desc())).scalars().all()
    all_project_names = list_project_names(db)

    rows: list[AdminManagementRow] = []
    for admin in admins:
        status = "password_reset_requested" if admin.senha is None else "active"
        status_label = describe_user_profile(admin.perfil)
        admin_project_names = list_materialized_user_project_names(db, admin)
        if not admin_project_names and admin.admin_monitored_projects_json is None:
            admin_project_names = all_project_names
        if admin.senha is None:
            status_label = f"{status_label} | senha pendente"
        rows.append(
            AdminManagementRow(
                id=admin.id,
                row_type="admin",
                chave=admin.chave,
                nome=admin.nome,
                perfil=admin.perfil,
                projects=admin_project_names,
                status=status,
                status_label=status_label,
                can_revoke=True,
                can_approve=False,
                can_reject=False,
                can_set_password=admin.senha is None,
            )
        )

    for request_row in requests:
        requested_profile = request_row.requested_profile or 1
        rows.append(
            AdminManagementRow(
                id=request_row.id,
                row_type="request",
                chave=request_row.chave,
                nome=request_row.nome_completo,
                perfil=requested_profile,
                status="pending",
                status_label=f"Solicitacao Pendente | {describe_user_profile(requested_profile)}",
                can_revoke=False,
                can_approve=True,
                can_reject=True,
                can_set_password=False,
            )
        )

    return rows


@router.post("/auth/login", response_model=AdminActionResponse)
def admin_login(payload: AdminLoginRequest, request: Request, db: Session = Depends(get_db)) -> AdminActionResponse:
    key = normalize_admin_key(payload.chave)
    admin = db.execute(select(User).where(User.chave == key)).scalar_one_or_none()

    if admin is None:
        log_event(
            db,
            source="admin",
            action="login",
            status="failed",
            message="Administrative login rejected",
            request_path="/api/admin/auth/login",
            http_status=401,
            details=f"chave={key}",
            commit=True,
        )
        raise HTTPException(status_code=401, detail="Chave ou senha invalida")

    if not user_can_access_admin_panel(admin):
        log_event(
            db,
            source="admin",
            action="login",
            status="blocked",
            message="Administrative login blocked due to missing admin panel access",
            request_path="/api/admin/auth/login",
            http_status=403,
            details=f"chave={admin.chave}",
            commit=True,
        )
        raise HTTPException(
            status_code=403,
            detail="Este usuario nao possui acesso ao painel Admin.",
        )

    if admin.senha is None:
        log_event(
            db,
            source="admin",
            action="login",
            status="blocked",
            message="Administrative login blocked due to missing user password",
            request_path="/api/admin/auth/login",
            http_status=403,
            details=f"chave={admin.chave}",
            commit=True,
        )
        raise HTTPException(status_code=403, detail="Este usuario ainda nao possui senha cadastrada.")

    if not verify_password(payload.senha, admin.senha):
        log_event(
            db,
            source="admin",
            action="login",
            status="failed",
            message="Administrative login rejected",
            request_path="/api/admin/auth/login",
            http_status=401,
            details=f"chave={key}",
            commit=True,
        )
        raise HTTPException(status_code=401, detail="Chave ou senha invalida")

    clear_admin_session(request)
    request.session["admin_user_id"] = admin.id
    log_event(
        db,
        source="admin",
        action="login",
        status="done",
        message="Administrative login completed",
        request_path="/api/admin/auth/login",
        http_status=200,
        details=f"chave={admin.chave}",
        commit=True,
    )
    return AdminActionResponse(ok=True, message="Login realizado com sucesso.")


@router.post("/auth/logout", response_model=AdminActionResponse)
def admin_logout(request: Request, db: Session = Depends(get_db)) -> AdminActionResponse:
    admin = get_authenticated_admin_from_session(request, db)
    if admin is not None:
        log_event(
            db,
            source="admin",
            action="logout",
            status="done",
            message="Administrative logout completed",
            request_path="/api/admin/auth/logout",
            http_status=200,
            details=f"chave={admin.chave}",
            commit=True,
        )
    clear_admin_session(request)
    return AdminActionResponse(ok=True, message="Sessao encerrada com sucesso.")


@router.get("/auth/session", response_model=AdminSessionResponse)
def admin_session(request: Request, db: Session = Depends(get_db)) -> AdminSessionResponse:
    admin = get_authenticated_admin_from_session(request, db)
    if admin is None:
        return AdminSessionResponse(authenticated=False)
    return AdminSessionResponse(authenticated=True, admin=build_admin_identity(admin))


@router.get("/auth/request-access/status", response_model=AdminSelfAccessStatusResponse)
def get_admin_request_access_status(
    chave: str = Query(min_length=4, max_length=4),
    db: Session = Depends(get_db),
) -> AdminSelfAccessStatusResponse:
    key = normalize_admin_key(chave)
    existing_user = db.execute(select(User).where(User.chave == key)).scalar_one_or_none()
    pending_request = db.execute(select(AdminAccessRequest).where(AdminAccessRequest.chave == key)).scalar_one_or_none()
    is_admin = user_has_admin_access(existing_user)
    has_password = existing_user is not None and bool(existing_user.senha)

    if pending_request is not None:
        message = "Ja existe uma solicitacao pendente para essa chave."
    elif is_admin:
        message = "Esta chave ja possui acesso administrativo."
    elif existing_user is None:
        message = "Chave nao cadastrada. Continue para registrar o usuario."
    elif not has_password:
        message = "Esta chave ja existe, mas ainda nao possui senha cadastrada."
    else:
        message = "Chave cadastrada. A solicitacao pode ser enviada."

    return AdminSelfAccessStatusResponse(
        found=existing_user is not None,
        chave=key,
        has_password=has_password,
        is_admin=is_admin,
        has_pending_request=pending_request is not None,
        message=message,
    )


@router.post("/auth/request-access/self-service", response_model=AdminActionResponse)
def request_admin_access_self_service(
    payload: AdminSelfAccessRequest,
    db: Session = Depends(get_db),
) -> AdminActionResponse:
    key = normalize_admin_key(payload.chave)
    existing_user = db.execute(select(User).where(User.chave == key)).scalar_one_or_none()
    if existing_user is not None and user_has_admin_access(existing_user):
        log_event(
            db,
            source="admin",
            action="admin_request",
            status="failed",
            message="Administrative self-service request rejected because key already belongs to an admin",
            request_path="/api/admin/auth/request-access/self-service",
            http_status=409,
            details=f"chave={key}",
            commit=True,
        )
        raise HTTPException(status_code=409, detail="Ja existe um administrador com essa chave.")

    pending_request = db.execute(select(AdminAccessRequest).where(AdminAccessRequest.chave == key)).scalar_one_or_none()
    if pending_request is not None:
        log_event(
            db,
            source="admin",
            action="admin_request",
            status="failed",
            message="Administrative self-service request rejected because another request is already pending",
            request_path="/api/admin/auth/request-access/self-service",
            http_status=409,
            details=f"chave={key}",
            commit=True,
        )
        raise HTTPException(status_code=409, detail="Ja existe uma solicitacao pendente para essa chave.")

    if existing_user is not None:
        if existing_user.senha is None:
            log_event(
                db,
                source="admin",
                action="admin_request",
                status="failed",
                message="Administrative self-service request rejected because the registered user has no password",
                request_path="/api/admin/auth/request-access/self-service",
                http_status=409,
                details=f"chave={key}",
                commit=True,
            )
            raise HTTPException(status_code=409, detail="Esta chave ainda nao possui senha cadastrada.")
        request_name = existing_user.nome
        request_password_hash = existing_user.senha
    else:
        if not payload.nome_completo:
            raise HTTPException(status_code=422, detail="Informe o nome completo para cadastrar o usuario.")
        if not payload.projeto:
            raise HTTPException(status_code=422, detail="Selecione um projeto para cadastrar o usuario.")
        if not payload.senha:
            raise HTTPException(status_code=422, detail="Informe uma senha para cadastrar o usuario.")

        timestamp = now_sgt()
        request_password_hash = hash_password(payload.senha)
        request_name = payload.nome_completo
        db.add(
            User(
                rfid=None,
                chave=key,
                senha=request_password_hash,
                perfil=0,
                nome=request_name,
                projeto=ensure_known_project(db, payload.projeto),
                workplace=None,
                placa=None,
                end_rua=None,
                zip=None,
                cargo=None,
                email=None,
                local=None,
                checkin=None,
                time=None,
                last_active_at=timestamp,
                inactivity_days=0,
            )
        )

    db.add(
        AdminAccessRequest(
            chave=key,
            nome_completo=request_name,
            password_hash=request_password_hash,
            requested_profile=1,
            requested_at=now_sgt(),
        )
    )
    log_event(
        db,
        source="admin",
        action="admin_request",
        status="pending",
        message="Administrative self-service access request created",
        request_path="/api/admin/auth/request-access/self-service",
        http_status=200,
        details=f"chave={key}; nome={request_name}; registered_user={existing_user is not None}",
    )
    db.commit()
    notify_admin_views("admin", "event")
    return AdminActionResponse(ok=True, message="Solicitacao enviada para aprovacao de um administrador.")


@router.post("/auth/request-access", response_model=AdminActionResponse)
def request_admin_access(payload: AdminAccessRequestCreate, db: Session = Depends(get_db)) -> AdminActionResponse:
    key = normalize_admin_key(payload.chave)
    existing_admin = db.execute(select(User).where(User.chave == key)).scalar_one_or_none()
    if existing_admin is not None and user_has_admin_access(existing_admin):
        log_event(
            db,
            source="admin",
            action="admin_request",
            status="failed",
            message="Administrative access request rejected because key already belongs to an admin",
            request_path="/api/admin/auth/request-access",
            http_status=409,
            details=f"chave={key}",
            commit=True,
        )
        raise HTTPException(status_code=409, detail="Ja existe um administrador com essa chave.")

    pending_request = db.execute(select(AdminAccessRequest).where(AdminAccessRequest.chave == key)).scalar_one_or_none()
    if pending_request is not None:
        log_event(
            db,
            source="admin",
            action="admin_request",
            status="failed",
            message="Administrative access request rejected because another request is already pending",
            request_path="/api/admin/auth/request-access",
            http_status=409,
            details=f"chave={key}",
            commit=True,
        )
        raise HTTPException(status_code=409, detail="Ja existe uma solicitacao pendente para essa chave.")

    db.add(
        AdminAccessRequest(
            chave=key,
            nome_completo=payload.nome_completo.strip(),
            password_hash=hash_password(payload.senha),
            requested_profile=1,
            requested_at=now_sgt(),
        )
    )
    log_event(
        db,
        source="admin",
        action="admin_request",
        status="pending",
        message="Administrative access request created",
        request_path="/api/admin/auth/request-access",
        http_status=200,
        details=f"chave={key}; nome={payload.nome_completo.strip()}",
    )
    db.commit()
    notify_admin_views("admin", "event")
    return AdminActionResponse(ok=True, message="Solicitacao enviada para aprovacao de um administrador.")


@router.post("/auth/request-password-reset", response_model=AdminActionResponse)
def request_password_reset(payload: AdminPasswordResetRequest, db: Session = Depends(get_db)) -> AdminActionResponse:
    key = normalize_admin_key(payload.chave)
    admin = db.execute(select(User).where(User.chave == key)).scalar_one_or_none()
    if admin is None or not user_has_admin_access(admin):
        log_event(
            db,
            source="admin",
            action="password",
            status="failed",
            message="Administrative password reset request failed because admin was not found",
            request_path="/api/admin/auth/request-password-reset",
            http_status=404,
            details=f"chave={key}",
            commit=True,
        )
        raise HTTPException(status_code=404, detail="Administrador nao encontrado para a chave informada.")
    if admin.senha is None:
        log_event(
            db,
            source="admin",
            action="password",
            status="failed",
            message="Administrative password reset request rejected because a reset is already pending",
            request_path="/api/admin/auth/request-password-reset",
            http_status=409,
            details=f"chave={admin.chave}",
            commit=True,
        )
        raise HTTPException(status_code=409, detail="Ja existe um pedido de recadastro de senha para esta chave.")

    admin.senha = None
    log_event(
        db,
        source="admin",
        action="password",
        status="pending",
        message="Administrative password reset requested",
        request_path="/api/admin/auth/request-password-reset",
        http_status=200,
        details=f"chave={admin.chave}",
    )
    db.commit()
    notify_admin_views("admin", "event")
    return AdminActionResponse(
        ok=True,
        message="Sua senha foi removida. Outro administrador devera cadastrar uma nova senha.",
    )


@router.post("/auth/verify-current-password", response_model=AdminPasswordVerifyResponse)
def verify_current_admin_password(
    payload: AdminSelfPasswordVerifyRequest,
    db: Session = Depends(get_db),
) -> AdminPasswordVerifyResponse:
    key = normalize_admin_key(payload.chave)
    admin = db.execute(select(User).where(User.chave == key)).scalar_one_or_none()
    if admin is None:
        raise HTTPException(status_code=404, detail="Administrador nao encontrado para a chave informada.")
    if not user_has_admin_access(admin):
        raise HTTPException(status_code=403, detail="Este usuario nao possui acesso ao Admin.")
    if admin.senha is None:
        raise HTTPException(status_code=403, detail="Este usuario ainda nao possui senha cadastrada.")
    if not verify_password(payload.senha_atual, admin.senha):
        return AdminPasswordVerifyResponse(ok=True, valid=False, message="A senha atual nao confere.")

    return AdminPasswordVerifyResponse(ok=True, valid=True, message="Senha atual confirmada.")


@router.post("/auth/change-password", response_model=AdminActionResponse)
def change_own_admin_password(
    payload: AdminSelfPasswordChangeRequest,
    db: Session = Depends(get_db),
) -> AdminActionResponse:
    key = normalize_admin_key(payload.chave)
    admin = db.execute(select(User).where(User.chave == key)).scalar_one_or_none()

    if admin is None:
        log_event(
            db,
            source="admin",
            action="password",
            status="failed",
            message="Administrative self-service password change failed because admin was not found",
            request_path="/api/admin/auth/change-password",
            http_status=404,
            details=f"chave={key}",
            commit=True,
        )
        raise HTTPException(status_code=404, detail="Administrador nao encontrado para a chave informada.")

    if not user_has_admin_access(admin):
        log_event(
            db,
            source="admin",
            action="password",
            status="blocked",
            message="Administrative self-service password change blocked due to missing admin profile",
            request_path="/api/admin/auth/change-password",
            http_status=403,
            details=f"chave={admin.chave}",
            commit=True,
        )
        raise HTTPException(status_code=403, detail="Este usuario nao possui acesso ao Admin.")

    if admin.senha is None:
        log_event(
            db,
            source="admin",
            action="password",
            status="blocked",
            message="Administrative self-service password change blocked due to missing current password",
            request_path="/api/admin/auth/change-password",
            http_status=403,
            details=f"chave={admin.chave}",
            commit=True,
        )
        raise HTTPException(status_code=403, detail="Este usuario ainda nao possui senha cadastrada.")

    if not verify_password(payload.senha_atual, admin.senha):
        log_event(
            db,
            source="admin",
            action="password",
            status="failed",
            message="Administrative self-service password change failed due to invalid current password",
            request_path="/api/admin/auth/change-password",
            http_status=401,
            details=f"chave={admin.chave}",
            commit=True,
        )
        raise HTTPException(status_code=401, detail="A senha atual nao confere.")

    admin.senha = hash_password(payload.nova_senha)
    log_event(
        db,
        source="admin",
        action="password",
        status="updated",
        message="Administrative self-service password changed",
        request_path="/api/admin/auth/change-password",
        http_status=200,
        details=f"chave={admin.chave}",
    )
    db.commit()
    notify_admin_views("event")
    return AdminActionResponse(ok=True, message="Senha alterada com sucesso.")


@router.get("/stream", dependencies=[Depends(require_admin_stream_session)])
async def stream_updates(request: Request) -> StreamingResponse:
    subscriber_id, queue = admin_updates_broker.subscribe()

    async def event_generator():
        try:
            yield encode_sse({"reason": "connected"})
            while True:
                if await request.is_disconnected():
                    break

                try:
                    payload = await asyncio.wait_for(queue.get(), timeout=15)
                    yield f"data: {payload}\n\n"
                except asyncio.TimeoutError:
                    yield ": keep-alive\n\n"
        finally:
            admin_updates_broker.unsubscribe(subscriber_id)

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


def _accident_summary(db: Session, accident: Accident) -> AccidentSummary:
    opened_by_label = "—"
    if accident.opened_by_admin_id:
        admin = db.get(AdminUser, accident.opened_by_admin_id)
        if admin:
            opened_by_label = admin.nome_completo
    elif accident.opened_by_user_id:
        user = db.get(User, accident.opened_by_user_id)
        if user:
            opened_by_label = user.nome
    return AccidentSummary(
        id=accident.id,
        accident_number=accident.accident_number,
        accident_number_label=format_accident_number(accident.accident_number),
        project_name=accident.project_name_snapshot,
        location_name=accident.location_name_snapshot,
        location_is_registered=accident.location_is_registered,
        origin=accident.origin,
        opened_by_label=opened_by_label,
        opened_at=accident.opened_at,
        closed_at=accident.closed_at,
    )


@router.get("/accidents/active", response_model=AdminAccidentStateResponse, dependencies=[Depends(require_admin_session)])
def get_active_accident_state(db: Session = Depends(get_db)) -> AdminAccidentStateResponse:
    active = list_active_accident(db)
    if active is None:
        return AdminAccidentStateResponse(is_active=False)
    return AdminAccidentStateResponse(
        is_active=True,
        accident=_accident_summary(db, active),
        situation_rows=build_situation_rows(db, accident=active),
    )


@router.post("/accidents/open", response_model=AdminAccidentStateResponse, dependencies=[Depends(require_full_admin_session)])
def open_admin_accident(
    payload: AdminAccidentOpenRequest,
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> AdminAccidentStateResponse:
    try:
        accident = open_accident(
            db,
            origin="admin",
            project_id=payload.project_id,
            location_id=payload.location_id,
            custom_location_name=payload.custom_location_name,
            opened_by_admin_id=current_admin.id,
            opened_by_user_id=None,
        )
    except AccidentAlreadyActiveError:
        raise HTTPException(status_code=409, detail="Ja existe um acidente em curso.")
    except InvalidAccidentLocationError:
        raise HTTPException(status_code=422, detail="O local selecionado nao pertence ao projeto.")

    log_event(
        db,
        source="admin",
        action="accident_open",
        status="done",
        message="Accident opened by admin",
        request_path="/api/admin/accidents/open",
        http_status=200,
        details=f"accident_number={accident.accident_number} project_id={payload.project_id}",
        commit=True,
    )
    return AdminAccidentStateResponse(
        is_active=True,
        accident=_accident_summary(db, accident),
        situation_rows=build_situation_rows(db, accident=accident),
    )




@router.post("/accidents/close", response_model=AdminAccidentStateResponse, dependencies=[Depends(require_full_admin_session)])
def close_admin_accident(
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> AdminAccidentStateResponse:
    active = list_active_accident(db)
    if active is None:
        raise HTTPException(status_code=409, detail="Nenhum acidente em curso.")

    closed = close_accident(db, accident=active, closed_by_admin_id=current_admin.id)
    background_tasks.add_task(build_and_attach_archive_for_accident, closed.id)

    log_event(
        db,
        source="admin",
        action="accident_close",
        status="done",
        message="Accident closed by admin",
        request_path="/api/admin/accidents/close",
        http_status=200,
        details=f"accident_id={closed.id}",
        commit=True,
    )
    return AdminAccidentStateResponse(is_active=False)


def generate_presigned_url(object_key: str, expires_in_seconds: int = 300) -> str:
    from ..services.object_storage import generate_presigned_url as _gen_url
    return _gen_url(object_key=object_key, expires_in_seconds=expires_in_seconds)


@router.get("/accidents", response_model=AccidentClosedListResponse, dependencies=[Depends(require_full_admin_session)])
def list_closed_accidents_endpoint(
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> AccidentClosedListResponse:
    rows = []
    accidents = db.execute(
        select(Accident).where(Accident.closed_at.is_not(None)).order_by(Accident.accident_number.desc())
    ).scalars().all()
    for accident in accidents:
        archive = db.execute(
            select(AccidentArchive).where(AccidentArchive.accident_id == accident.id)
        ).scalar_one_or_none()
        opened_by_label = "—"
        if accident.opened_by_admin_id:
            admin = db.get(AdminUser, accident.opened_by_admin_id)
            if admin:
                opened_by_label = admin.nome_completo
        elif accident.opened_by_user_id:
            user = db.get(User, accident.opened_by_user_id)
            if user:
                opened_by_label = user.nome
        rows.append(AccidentClosedRow(
            id=accident.id,
            accident_number_label=format_accident_number(accident.accident_number),
            project_name=accident.project_name_snapshot,
            author_label=opened_by_label,
            opened_at=accident.opened_at,
            closed_at=accident.closed_at,
            download_url=f"/api/admin/accidents/{accident.id}/archive",
            download_ready=archive is not None,
            can_delete=(current_admin.perfil == 9),
        ))
    return AccidentClosedListResponse(rows=rows)


@router.get("/accidents/{accident_id}/archive", dependencies=[Depends(require_full_admin_session)])
def download_accident_archive(
    accident_id: int,
    db: Session = Depends(get_db),
) -> Response:
    archive = db.execute(
        select(AccidentArchive).where(AccidentArchive.accident_id == accident_id)
    ).scalar_one_or_none()
    if archive is None:
        raise HTTPException(status_code=404, detail="Arquivo do acidente ainda nao esta pronto.")
    presigned_url = generate_presigned_url(object_key=archive.zip_object_key, expires_in_seconds=300)
    return RedirectResponse(url=presigned_url, status_code=307)


@router.get("/accidents/wizard/projects", response_model=list[AccidentProjectOption], dependencies=[Depends(require_full_admin_session)])
def list_accident_wizard_projects(db: Session = Depends(get_db)) -> list[AccidentProjectOption]:
    return [AccidentProjectOption(id=p.id, name=p.name) for p in list_projects(db)]


@router.get("/accidents/wizard/locations", response_model=list[AccidentLocationOption], dependencies=[Depends(require_full_admin_session)])
def list_accident_wizard_locations(
    project_id: int = Query(...),
    db: Session = Depends(get_db),
) -> list[AccidentLocationOption]:
    project = db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Projeto nao encontrado.")
    options = []
    for loc in db.execute(select(ManagedLocation)).scalars().all():
        try:
            projects = json.loads(loc.projects_json or "[]")
        except Exception:
            projects = []
        if project.name in projects:
            options.append(AccidentLocationOption(id=loc.id, name=loc.local, registered=True))
    return options


def delete_prefix(prefix: str) -> None:
    from ..services.object_storage import delete_prefix as _del_prefix
    _del_prefix(prefix=prefix)


@router.get("/accidents/local-asset/{path:path}")
def serve_local_asset(path: str) -> Response:
    """Serve locally-stored accident assets (dev only). Returns 404 in production."""
    from ..services.object_storage import _use_remote, _local_root
    if _use_remote():
        raise HTTPException(status_code=404, detail="Not available in production.")
    target = _local_root() / path
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="Asset nao encontrado.")
    return FileResponse(str(target))


@router.delete("/accidents/{accident_id}", response_model=AdminActionResponse, dependencies=[Depends(require_full_admin_session)])
def delete_accident_endpoint(
    accident_id: int,
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> AdminActionResponse:
    if current_admin.perfil != 9:
        raise HTTPException(status_code=403, detail="Apenas perfil 9 pode remover acidentes.")
    accident = db.get(Accident, accident_id)
    if accident is None:
        raise HTTPException(status_code=404, detail="Acidente nao encontrado.")
    if accident.closed_at is None:
        raise HTTPException(status_code=409, detail="Nao e possivel remover um acidente em curso. Encerre o Modo Acidente primeiro.")

    accident_number = accident.accident_number
    db.delete(accident)  # cascade removes reports, videos, archive
    db.commit()

    delete_prefix(prefix=f"accidents/{format_accident_number(accident_number)}/")
    log_event(
        db,
        source="admin",
        action="accident_delete",
        status="done",
        message=f"Accident {accident_number} deleted",
        details=f"by admin={current_admin.chave}",
        commit=True,
    )

    notify_admin_data_changed("accident_closed", metadata={"deleted_accident_id": accident_id})
    notify_web_check_data_changed("accident_closed", metadata={"deleted_accident_id": accident_id})

    return AdminActionResponse(ok=True, message="Acidente removido com sucesso.")


@router.get("/administrators", response_model=list[AdminManagementRow], dependencies=[Depends(require_full_admin_session)])
def list_administrators(db: Session = Depends(get_db)) -> list[AdminManagementRow]:
    return list_admin_rows(db)


@router.get("/projects", response_model=list[ProjectRow], dependencies=[Depends(require_full_admin_session)])
def list_admin_projects(db: Session = Depends(get_db)) -> list[ProjectRow]:
    return [build_project_row(project) for project in list_projects(db)]


@router.post("/projects", response_model=ProjectRow, dependencies=[Depends(require_full_admin_session)])
def create_admin_project(
    payload: ProjectCreate,
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> ProjectRow:
    existing_project = db.execute(select(Project).where(Project.name == payload.name)).scalar_one_or_none()
    if existing_project is not None:
        raise HTTPException(status_code=409, detail="Ja existe um projeto com esse nome.")

    project = Project(
        name=payload.name,
        address=payload.address,
        zip_code=payload.zip_code,
        **build_project_fields(
            country_code=payload.country_code,
            country_name=payload.country_name,
            timezone_name=payload.timezone_name,
        ),
    )
    db.add(project)
    db.flush()
    log_event(
        db,
        source="admin",
        action="register",
        status="done",
        message="Project created via admin",
        request_path="/api/admin/projects",
        http_status=200,
        details=(
            f"updated_by={current_admin.chave}; project_name={payload.name}; country_code={payload.country_code}; "
            f"country_name={payload.country_name}; timezone_name={payload.timezone_name}; "
            f"address={payload.address or '-'}; zip_code={payload.zip_code or '-'}"
        ),
    )
    add_user_project_membership(db, current_admin, payload.name)
    db.commit()
    db.refresh(project)
    notify_admin_views("register", "event")
    return build_project_row(project)


@router.put("/projects/{project_id}", response_model=ProjectRow, dependencies=[Depends(require_full_admin_session)])
def update_admin_project(
    project_id: int,
    payload: ProjectUpdate,
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> ProjectRow:
    project = db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Projeto nao encontrado.")

    previous_name = project.name
    previous_country_code = project.country_code
    previous_country_name = project.country_name
    previous_timezone_name = project.timezone_name
    previous_address = str(project.address or "").strip()
    previous_zip_code = str(project.zip_code or "").strip()
    updated_fields = build_project_fields(
        country_code=payload.country_code,
        country_name=payload.country_name,
        timezone_name=payload.timezone_name,
    )
    updated_user_links = 0
    updated_location_links = 0
    updated_admin_scopes = 0
    recreated_minimum_checkout_distances = 0

    if payload.name != previous_name:
        existing_project = db.execute(
            select(Project).where(Project.name == payload.name, Project.id != project_id)
        ).scalar_one_or_none()
        if existing_project is not None:
            raise HTTPException(status_code=409, detail="Ja existe um projeto com esse nome.")

        linked_users = db.execute(select(User).where(User.projeto == previous_name).order_by(User.id)).scalars().all()
        updated_user_links = len(linked_users)
        for linked_user in linked_users:
            linked_user.projeto = payload.name

        linked_locations = db.execute(select(ManagedLocation).order_by(ManagedLocation.id)).scalars().all()
        for linked_location in linked_locations:
            existing_projects = extract_location_projects(linked_location)
            if previous_name not in existing_projects:
                continue

            linked_location.projects_json = dump_location_projects(
                [payload.name if project_name == previous_name else project_name for project_name in existing_projects]
            )
            linked_location.updated_at = now_sgt()
            updated_location_links += 1

        for administrator in db.execute(select(User).order_by(User.id)).scalars().all():
            if not user_has_admin_access(administrator):
                continue

            explicit_projects = extract_admin_monitored_projects(administrator)
            if explicit_projects is None or previous_name not in explicit_projects:
                continue

            administrator.admin_monitored_projects_json = None
            updated_admin_scopes += 1

        existing_distance_rows = db.execute(
            select(ProjectAutoCheckoutDistance)
            .where(ProjectAutoCheckoutDistance.project_name == previous_name)
            .order_by(ProjectAutoCheckoutDistance.id)
        ).scalars().all()
        preserved_distance_rows = [
            {
                "minimum_checkout_distance_meters": row.minimum_checkout_distance_meters,
                "created_at": row.created_at,
            }
            for row in existing_distance_rows
        ]
        for row in existing_distance_rows:
            db.delete(row)
        if existing_distance_rows:
            db.flush()

        project.name = payload.name
        if preserved_distance_rows:
            db.flush()
            current_time = now_sgt()
            for row in preserved_distance_rows:
                db.add(
                    ProjectAutoCheckoutDistance(
                        project_name=payload.name,
                        minimum_checkout_distance_meters=row["minimum_checkout_distance_meters"],
                        created_at=row["created_at"],
                        updated_at=current_time,
                    )
                )
            recreated_minimum_checkout_distances = len(preserved_distance_rows)

    project.country_code = updated_fields["country_code"]
    project.country_name = updated_fields["country_name"]
    project.timezone_name = updated_fields["timezone_name"]
    project.address = payload.address
    project.zip_code = payload.zip_code

    log_event(
        db,
        source="admin",
        action="update",
        status="done",
        message="Project updated via admin",
        request_path=f"/api/admin/projects/{project_id}",
        http_status=200,
        details=(
            f"updated_by={current_admin.chave}; project_id={project_id}; project_name={project.name}; "
            f"project_name_old={previous_name}; "
            f"country_code={previous_country_code}->{project.country_code}; "
            f"country_name={previous_country_name}->{project.country_name}; "
            f"timezone_name={previous_timezone_name}->{project.timezone_name}; "
            f"address={previous_address or '-'}->{project.address or '-'}; "
            f"zip_code={previous_zip_code or '-'}->{project.zip_code or '-'}; "
            f"updated_user_links={updated_user_links}; "
            f"updated_location_links={updated_location_links}; "
            f"updated_admin_scopes={updated_admin_scopes}; "
            f"recreated_minimum_checkout_distances={recreated_minimum_checkout_distances}"
        ),
    )
    db.commit()
    db.refresh(project)
    notify_admin_views("register", "event")
    return build_project_row(project)


@router.delete("/projects/{project_id}", response_model=AdminActionResponse, dependencies=[Depends(require_full_admin_session)])
def remove_admin_project(
    project_id: int,
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> AdminActionResponse:
    project = db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Projeto nao encontrado.")

    all_projects = list_projects(db)
    if len(all_projects) <= 1:
        raise HTTPException(status_code=409, detail="Nao e possivel remover o ultimo projeto cadastrado.")

    remaining_project_names = sorted(row.name for row in all_projects if row.id != project.id)
    fallback_project = next(iter(remaining_project_names), None)
    all_users = db.execute(select(User).order_by(User.id)).scalars().all()
    project_names_by_user_id = list_user_project_names_map(db, all_users)
    linked_users = [
        user
        for user in all_users
        if project.name in project_names_by_user_id.get(user.id, [])
    ]

    blocked_users: list[User] = []
    reassignment_targets: list[tuple[User, list[str]]] = []
    for linked_user in linked_users:
        current_project_names = project_names_by_user_id.get(linked_user.id, [])
        next_project_names = [project_name for project_name in current_project_names if project_name != project.name]
        if not next_project_names:
            if not user_has_admin_access(linked_user):
                blocked_users.append(linked_user)
                continue
            next_project_names = [fallback_project or resolve_default_project_name(db)]
        reassignment_targets.append((linked_user, next_project_names))

    if blocked_users:
        raise HTTPException(status_code=409, detail="Nao e possivel remover um projeto com usuarios vinculados.")

    reassigned_user_count = 0
    for linked_user, next_project_names in reassignment_targets:
        replace_user_project_memberships(db, linked_user, next_project_names)
        reassigned_user_count += 1

    linked_locations = db.execute(select(ManagedLocation).order_by(ManagedLocation.id)).scalars().all()
    reassigned_location_count = 0
    for linked_location in linked_locations:
        existing_projects = extract_location_projects(linked_location)
        if project.name not in existing_projects:
            continue

        next_projects = [
            fallback_project if project_name == project.name else project_name
            for project_name in existing_projects
        ]
        if not any(next_projects):
            next_projects = [fallback_project or resolve_default_project_name(db)]

        linked_location.projects_json = dump_location_projects(next_projects)
        linked_location.updated_at = now_sgt()
        reassigned_location_count += 1

    updated_admin_scope_count = 0
    for administrator in db.execute(select(User).order_by(User.id)).scalars().all():
        if not user_has_admin_access(administrator):
            continue

        explicit_projects = extract_admin_monitored_projects(administrator)
        if explicit_projects is None or project.name not in explicit_projects:
            continue

        remaining_explicit_projects = [
            project_name for project_name in explicit_projects if project_name != project.name
        ]
        if remaining_explicit_projects:
            ensure_user_active_project_is_member(db, administrator)

        administrator.admin_monitored_projects_json = None
        updated_admin_scope_count += 1

    db.delete(project)
    log_event(
        db,
        source="admin",
        action="register",
        status="removed",
        message="Project removed via admin",
        request_path=f"/api/admin/projects/{project_id}",
        http_status=200,
        details=(
            f"updated_by={current_admin.chave}; project_name={project.name}; project_id={project_id}; "
            f"reassigned_users={reassigned_user_count}; reassigned_locations={reassigned_location_count}; "
            f"updated_admin_scopes={updated_admin_scope_count}"
        ),
    )
    db.commit()
    notify_admin_views("register", "event")
    return AdminActionResponse(ok=True, message="Projeto removido com sucesso.")


@router.post(
    "/administrators/requests/{request_id}/approve",
    response_model=AdminActionResponse,
)
def approve_administrator_request(
    request_id: int,
    payload: AdminProfileUpdateRequest | None = None,
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> AdminActionResponse:
    access_request = db.get(AdminAccessRequest, request_id)
    if access_request is None:
        log_event(
            db,
            source="admin",
            action="admin_request",
            status="failed",
            message="Administrative access request approval failed because request was not found",
            request_path=f"/api/admin/administrators/requests/{request_id}/approve",
            http_status=404,
            details=f"request_id={request_id}; approved_by={current_admin.chave}",
            commit=True,
        )
        raise HTTPException(status_code=404, detail="Solicitacao de administrador nao encontrada.")

    existing_admin = db.execute(select(User).where(User.chave == access_request.chave)).scalar_one_or_none()
    default_project_name = resolve_default_project_name(db)
    known_project_names = set(list_project_names(db))
    requested_profile = normalize_administrator_profile(
        payload.perfil if payload is not None else (access_request.requested_profile or 1)
    )
    if existing_admin is not None and user_has_admin_access(existing_admin):
        log_event(
            db,
            source="admin",
            action="admin_request",
            status="failed",
            message="Administrative access request approval failed because target key is already assigned",
            request_path=f"/api/admin/administrators/requests/{request_id}/approve",
            http_status=409,
            details=f"chave={access_request.chave}; approved_by={current_admin.chave}",
            commit=True,
        )
        raise HTTPException(status_code=409, detail="Ja existe um administrador com essa chave.")

    timestamp = now_sgt()
    approved_admin: User
    if existing_admin is None:
        approved_admin = User(
            rfid=None,
            chave=access_request.chave,
            senha=access_request.password_hash,
            perfil=requested_profile,
            nome=access_request.nome_completo,
            projeto=default_project_name,
            admin_monitored_projects_json=None,
            workplace=None,
            placa=None,
            end_rua=None,
            zip=None,
            cargo=None,
            email=None,
            local=None,
            checkin=None,
            time=None,
            last_active_at=timestamp,
            inactivity_days=0,
        )
        db.add(approved_admin)
        db.flush()
    else:
        approved_admin = existing_admin
        existing_admin.nome = access_request.nome_completo
        if requested_profile != TRANSPORT_ACCESS_DIGIT or existing_admin.senha is None:
            existing_admin.senha = access_request.password_hash
        existing_admin.perfil = merge_user_profile_values(existing_admin.perfil, requested_profile)
        existing_admin.admin_monitored_projects_json = None

    approved_project_names = list_user_project_names(db, approved_admin)
    if not approved_project_names:
        seed_project_name = str(approved_admin.projeto or "").strip().upper() or default_project_name
        if seed_project_name not in known_project_names:
            seed_project_name = default_project_name
        approved_project_names = [seed_project_name]
    replace_user_project_memberships(db, approved_admin, approved_project_names)

    log_event(
        db,
        source="admin",
        action="admin_request",
        status="approved",
        message="Administrative access request approved",
        request_path=f"/api/admin/administrators/requests/{request_id}/approve",
        http_status=200,
        details=f"chave={access_request.chave}; approved_by={current_admin.chave}",
    )
    db.delete(access_request)
    db.commit()
    notify_admin_views("admin", "event")
    return AdminActionResponse(ok=True, message="Administrador aprovado com sucesso.")


@router.post(
    "/administrators/{admin_id}/profile",
    response_model=AdminActionResponse,
)
def update_administrator_profile(
    admin_id: int,
    payload: AdminProfileUpdateRequest,
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> AdminActionResponse:
    try:
        payload = AdminProfileUpdateRequest.model_validate(
            payload.model_dump(),
            context={"allowed_project_names": list_project_names(db)},
        )
    except ValidationError as error:
        request_validation_errors = build_request_validation_errors(error)
        raise RequestValidationError(request_validation_errors) from error

    administrator = db.get(User, admin_id)
    if administrator is None or not user_has_admin_access(administrator):
        log_event(
            db,
            source="admin",
            action="admin_access",
            status="failed",
            message="Administrator profile update failed because target admin was not found",
            request_path=f"/api/admin/administrators/{admin_id}/profile",
            http_status=404,
            details=f"admin_id={admin_id}; updated_by={current_admin.chave}",
            commit=True,
        )
        raise HTTPException(status_code=404, detail="Administrador nao encontrado.")

    previous_profile = normalize_user_profile(administrator.perfil)
    next_profile = normalize_administrator_profile(payload.perfil)
    previous_project_names = list_user_project_names(db, administrator)
    previous_active_project = administrator.projeto
    next_project_names = previous_project_names
    administrator.admin_monitored_projects_json = None

    if payload.projects is not None:
        replace_user_project_memberships(db, administrator, payload.projects)
        next_project_names = list_user_project_names(db, administrator)

    administrator.perfil = next_profile
    next_active_project = administrator.projeto
    log_event(
        db,
        source="admin",
        action="admin_access",
        status="updated",
        message="Administrator configuration updated",
        request_path=f"/api/admin/administrators/{admin_id}/profile",
        http_status=200,
        details=(
            f"chave={administrator.chave}; old_profile={previous_profile}; "
            f"new_profile={next_profile}; "
            f"old_projects={json.dumps(previous_project_names, ensure_ascii=True, separators=(",", ":"))}; "
            f"new_projects={json.dumps(next_project_names, ensure_ascii=True, separators=(",", ":"))}; "
            f"old_active_project={previous_active_project or '-'}; "
            f"new_active_project={next_active_project or '-'}; "
            f"updated_by={current_admin.chave}"
        ),
    )
    db.commit()
    notify_admin_views("admin", "event")
    return AdminActionResponse(ok=True, message="Configuracoes do administrador atualizadas com sucesso.")


@router.post(
    "/administrators/requests/{request_id}/reject",
    response_model=AdminActionResponse,
)
def reject_administrator_request(
    request_id: int,
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> AdminActionResponse:
    access_request = db.get(AdminAccessRequest, request_id)
    if access_request is None:
        log_event(
            db,
            source="admin",
            action="admin_request",
            status="failed",
            message="Administrative access request rejection failed because request was not found",
            request_path=f"/api/admin/administrators/requests/{request_id}/reject",
            http_status=404,
            details=f"request_id={request_id}; rejected_by={current_admin.chave}",
            commit=True,
        )
        raise HTTPException(status_code=404, detail="Solicitacao de administrador nao encontrada.")

    log_event(
        db,
        source="admin",
        action="admin_request",
        status="rejected",
        message="Administrative access request rejected",
        request_path=f"/api/admin/administrators/requests/{request_id}/reject",
        http_status=200,
        details=f"chave={access_request.chave}; rejected_by={current_admin.chave}",
    )
    db.delete(access_request)
    db.commit()
    notify_admin_views("admin", "event")
    return AdminActionResponse(ok=True, message="Solicitacao rejeitada com sucesso.")


@router.post("/administrators/{admin_id}/revoke", response_model=AdminActionResponse)
def revoke_administrator(
    admin_id: int,
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> AdminActionResponse:
    admin = db.get(User, admin_id)
    if admin is None or not user_has_admin_access(admin):
        log_event(
            db,
            source="admin",
            action="admin_access",
            status="failed",
            message="Administrator revocation failed because target admin was not found",
            request_path=f"/api/admin/administrators/{admin_id}/revoke",
            http_status=404,
            details=f"admin_id={admin_id}; revoked_by={current_admin.chave}",
            commit=True,
        )
        raise HTTPException(status_code=404, detail="Administrador nao encontrado.")
    if admin.id == current_admin.id:
        log_event(
            db,
            source="admin",
            action="admin_access",
            status="failed",
            message="Administrator revocation rejected because self-revocation is not allowed",
            request_path=f"/api/admin/administrators/{admin_id}/revoke",
            http_status=409,
            details=f"chave={admin.chave}; revoked_by={current_admin.chave}",
            commit=True,
        )
        raise HTTPException(status_code=409, detail="Voce nao pode revogar seu proprio acesso.")

    total_admins = sum(
        1
        for row in db.execute(select(User).where(User.perfil != 0)).scalars().all()
        if user_has_admin_access(row)
    )
    if total_admins <= 1:
        log_event(
            db,
            source="admin",
            action="admin_access",
            status="failed",
            message="Administrator revocation rejected because the last active admin cannot be removed",
            request_path=f"/api/admin/administrators/{admin_id}/revoke",
            http_status=409,
            details=f"chave={admin.chave}; revoked_by={current_admin.chave}",
            commit=True,
        )
        raise HTTPException(status_code=409, detail="Nao e possivel revogar o unico administrador ativo do sistema.")

    admin.perfil = remove_profile_access(admin.perfil, ADMIN_ACCESS_DIGIT)
    log_event(
        db,
        source="admin",
        action="admin_access",
        status="removed",
        message="Administrator access revoked",
        request_path=f"/api/admin/administrators/{admin_id}/revoke",
        http_status=200,
        details=f"chave={admin.chave}; revoked_by={current_admin.chave}",
    )
    db.commit()
    notify_admin_views("admin", "event")
    return AdminActionResponse(ok=True, message="Administrador revogado com sucesso.")


@router.post("/administrators/{admin_id}/set-password", response_model=AdminActionResponse)
def set_administrator_password(
    admin_id: int,
    payload: AdminPasswordSetRequest,
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> AdminActionResponse:
    admin = db.get(User, admin_id)
    if admin is None or not user_has_admin_access(admin):
        log_event(
            db,
            source="admin",
            action="password",
            status="failed",
            message="Administrative password update failed because target admin was not found",
            request_path=f"/api/admin/administrators/{admin_id}/set-password",
            http_status=404,
            details=f"admin_id={admin_id}; updated_by={current_admin.chave}",
            commit=True,
        )
        raise HTTPException(status_code=404, detail="Administrador nao encontrado.")
    if admin.senha is not None:
        log_event(
            db,
            source="admin",
            action="password",
            status="failed",
            message="Administrative password update rejected because no reset is pending",
            request_path=f"/api/admin/administrators/{admin_id}/set-password",
            http_status=409,
            details=f"chave={admin.chave}; updated_by={current_admin.chave}",
            commit=True,
        )
        raise HTTPException(status_code=409, detail="Esse administrador nao possui recadastro de senha pendente.")

    admin.senha = hash_password(payload.nova_senha)
    log_event(
        db,
        source="admin",
        action="password",
        status="updated",
        message="Administrative password updated",
        request_path=f"/api/admin/administrators/{admin_id}/set-password",
        http_status=200,
        details=f"chave={admin.chave}; updated_by={current_admin.chave}",
    )
    db.commit()
    notify_admin_views("admin", "event")
    return AdminActionResponse(ok=True, message="Nova senha cadastrada com sucesso.")


@router.get("/checkin", response_model=list[UserRow])
def list_checkin(
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_admin_session),
) -> list[UserRow]:
    reference_time = now_sgt()
    if sync_user_inactivity(db, reference_time=reference_time):
        db.commit()
    return build_presence_rows(db, action="checkin", current_admin=current_admin, reference_time=reference_time)


@router.get("/checkout", response_model=list[UserRow])
def list_checkout(
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_admin_session),
) -> list[UserRow]:
    reference_time = now_sgt()
    if sync_user_inactivity(db, reference_time=reference_time):
        db.commit()
    return build_presence_rows(db, action="checkout", current_admin=current_admin, reference_time=reference_time)


@router.get("/forms", response_model=list[ProviderFormRow])
def list_provider_forms(
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> list[ProviderFormRow]:
    return build_provider_forms_rows(db, current_admin=current_admin)


@router.get("/forms/queue/diagnostics", response_model=FormsQueueDiagnosticsResponse)
def get_forms_queue_diagnostics_view(
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> FormsQueueDiagnosticsResponse:
    return FormsQueueDiagnosticsResponse.model_validate(get_forms_queue_diagnostics(db=db))


@router.get("/diagnostics/database", response_model=DatabaseDiagnosticsResponse)
def get_database_diagnostics_view(
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> DatabaseDiagnosticsResponse:
    return DatabaseDiagnosticsResponse.model_validate(get_database_diagnostics(db=db))


@router.get("/reports/events", response_model=ReportEventsResponse)
def get_report_events(
    chave: str | None = Query(default=None),
    nome: str | None = Query(default=None),
    current_admin: User = Depends(require_full_admin_session),
    db: Session = Depends(get_db),
) -> ReportEventsResponse:
    user = resolve_report_user(db, chave=chave, nome=nome, current_admin=current_admin)
    return build_report_events_response(
        db,
        user=user,
        can_view_activity_time=user_can_view_activity_time(current_admin),
    )


@router.get("/reports/events/export")
def export_report_events(
    chave: str | None = Query(default=None),
    nome: str | None = Query(default=None),
    current_admin: User = Depends(require_full_admin_session),
    db: Session = Depends(get_db),
) -> Response:
    can_view_activity_time = user_can_view_activity_time(current_admin)
    user = resolve_report_user(db, chave=chave, nome=nome, current_admin=current_admin)
    report = build_report_events_response(db, user=user, can_view_activity_time=can_view_activity_time)
    file_name, content = build_report_events_export(
        user=user,
        report=report,
        can_view_activity_time=can_view_activity_time,
    )
    return Response(
        content=content,
        media_type=REPORT_EXPORT_CONTENT_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
    )


@router.get("/reports/events/export-all")
def export_all_report_events(
    current_admin: User = Depends(require_full_admin_session),
    db: Session = Depends(get_db),
) -> Response:
    file_name, content = build_all_report_events_export(
        db,
        can_view_activity_time=user_can_view_activity_time(current_admin),
        current_admin=current_admin,
    )
    return Response(
        content=content,
        media_type=REPORT_EXPORT_CONTENT_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
    )


@router.delete("/forms", response_model=AdminActionResponse, dependencies=[Depends(require_full_admin_session)])
def clear_provider_forms(db: Session = Depends(get_db)) -> AdminActionResponse:
    cleared_count = delete_provider_forms_rows(db)
    db.commit()
    notify_admin_views("event")
    if cleared_count == 0:
        return AdminActionResponse(ok=True, message="Nao havia registros de Forms para remover.")
    return AdminActionResponse(ok=True, message=f"{cleared_count} registro(s) de Forms removido(s) com sucesso.")


@router.get("/missing-checkout", response_model=list[UserRow], dependencies=[Depends(require_full_admin_session)])
def list_missing_checkout(db: Session = Depends(get_db)) -> list[UserRow]:
    reference_time = now_sgt()
    if sync_user_inactivity(db, reference_time=reference_time):
        db.commit()
    return build_missing_checkout_rows(db, reference_time=reference_time)


@router.get("/inactive", response_model=list[InactiveUserRow])
def list_inactive(
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> list[InactiveUserRow]:
    reference_time = now_sgt()
    if sync_user_inactivity(db, reference_time=reference_time):
        db.commit()
    return build_inactive_rows(db, current_admin=current_admin, reference_time=reference_time)


@router.get("/pending", response_model=list[PendingRow], dependencies=[Depends(require_full_admin_session)])
def list_pending(
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> list[PendingRow]:
    rows = db.execute(select(PendingRegistration).order_by(desc(PendingRegistration.last_seen_at))).scalars().all()
    effective_admin_projects = resolve_effective_admin_project_names(db, current_admin) or []
    if not effective_admin_projects:
        return []

    latest_scan_local_by_rfid, location_projects_by_local = build_pending_registration_scope_maps(db, rows)
    rows = [
        row
        for row in rows
        if pending_matches_admin_scope(
            db,
            row,
            current_admin=current_admin,
            admin_project_names=effective_admin_projects,
            latest_scan_local_by_rfid=latest_scan_local_by_rfid,
            location_projects_by_local=location_projects_by_local,
        )
    ]
    return [
        PendingRow(
            id=r.id,
            rfid=r.rfid,
            first_seen_at=r.first_seen_at,
            last_seen_at=r.last_seen_at,
            attempts=r.attempts,
        )
        for r in rows
    ]


@router.get("/locations", response_model=AdminLocationsResponse, dependencies=[Depends(require_full_admin_session)])
def list_locations(
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> AdminLocationsResponse:
    rows = db.execute(select(ManagedLocation).order_by(ManagedLocation.local, ManagedLocation.id)).scalars().all()
    effective_admin_projects = resolve_effective_admin_project_names(db, current_admin) or []
    known_project_names = set(list_project_names(db))
    rows = [
        row
        for row in rows
        if (
            location_matches_effective_admin_scope(
                db,
                current_admin,
                extract_location_projects(row),
                admin_project_names=effective_admin_projects,
                allow_global_locations=True,
            )
            or (
                bool(extract_location_projects(row))
                and not any(project_name in known_project_names for project_name in extract_location_projects(row))
            )
        )
    ]
    return AdminLocationsResponse(
        items=[build_location_row(row) for row in rows],
        location_accuracy_threshold_meters=get_location_accuracy_threshold_meters(db),
        mixed_zone_interval_minutes=get_mixed_zone_interval_minutes(db),
    )


@router.get(
    "/locations/auto-checkout-distances",
    response_model=AdminProjectMinimumCheckoutDistanceListResponse,
    dependencies=[Depends(require_full_admin_session)],
)
def list_project_minimum_checkout_distances(
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> AdminProjectMinimumCheckoutDistanceListResponse:
    rows = list_project_minimum_checkout_distance_rows(db)
    effective_admin_projects = resolve_effective_admin_project_names(db, current_admin) or []
    rows = [row for row in rows if row.project_name in set(effective_admin_projects)]
    return AdminProjectMinimumCheckoutDistanceListResponse(
        items=[
            AdminProjectMinimumCheckoutDistanceRow(
                project_name=row.project_name,
                minimum_checkout_distance_meters=row.minimum_checkout_distance_meters,
            )
            for row in rows
        ]
    )


@router.post(
    "/locations/auto-checkout-distances",
    response_model=AdminProjectMinimumCheckoutDistanceSaveResponse,
    dependencies=[Depends(require_full_admin_session)],
)
def update_project_minimum_checkout_distances(
    payload: AdminProjectMinimumCheckoutDistanceUpdate,
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> AdminProjectMinimumCheckoutDistanceSaveResponse:
    effective_admin_projects = resolve_effective_admin_project_names(db, current_admin) or []
    effective_admin_project_set = set(effective_admin_projects)
    if not effective_admin_project_set:
        raise HTTPException(status_code=403, detail="Administrador sem projetos vinculados nao pode alterar configuracoes por projeto.")

    current_rows = list_project_minimum_checkout_distance_rows(db)
    previous_values = {
        row.project_name: row.minimum_checkout_distance_meters
        for row in current_rows
    }
    normalized_items = [
        (
            ensure_known_project(db, item.project_name),
            item.minimum_checkout_distance_meters,
        )
        for item in payload.items
    ]
    if any(project_name not in effective_admin_project_set for project_name, _ in normalized_items):
        raise HTTPException(status_code=403, detail="Nao e possivel alterar projetos fora do seu escopo.")

    upsert_project_minimum_checkout_distance_rows(db, normalized_items)
    refreshed_rows = list_project_minimum_checkout_distance_rows(db)
    refreshed_rows = [row for row in refreshed_rows if row.project_name in effective_admin_project_set]
    changed_rows = [
        f"{row.project_name}:{previous_values.get(row.project_name)}->{row.minimum_checkout_distance_meters}"
        for row in refreshed_rows
        if previous_values.get(row.project_name) != row.minimum_checkout_distance_meters
    ]

    log_event(
        db,
        source="admin",
        action="location_config",
        status="updated",
        message="Project automatic checkout distances updated",
        request_path="/api/admin/locations/auto-checkout-distances",
        http_status=200,
        details=(
            f"updated_by={current_admin.chave}; "
            f"updated_projects={len(changed_rows)}; "
            f"changes={'; '.join(changed_rows) if changed_rows else 'none'}"
        ),
    )
    db.commit()
    notify_admin_views("location", "event")
    return AdminProjectMinimumCheckoutDistanceSaveResponse(
        ok=True,
        message="Distancias minimas para check-out automatico salvas com sucesso.",
        items=[
            AdminProjectMinimumCheckoutDistanceRow(
                project_name=row.project_name,
                minimum_checkout_distance_meters=row.minimum_checkout_distance_meters,
            )
            for row in refreshed_rows
        ],
    )


@router.get("/locations/audit", response_model=AdminLocationAuditResponse, dependencies=[Depends(require_full_admin_session)])
def audit_locations(
    include_valid: bool = Query(default=False),
    db: Session = Depends(get_db),
) -> AdminLocationAuditResponse:
    report = audit_locations_from_db(db)
    rows_payload = [
        row.to_dict()
        for row in report.rows
        if include_valid or row.issues
    ]
    return AdminLocationAuditResponse(
        summary=report.summary.to_dict(),
        rows=rows_payload,
    )


@router.post("/locations", response_model=AdminActionResponse, dependencies=[Depends(require_full_admin_session)])
def upsert_location(
    payload: object = Body(...),
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> AdminActionResponse:
    try:
        validated_payload = AdminLocationUpsert.model_validate(payload)
    except ValidationError as error:
        request_validation_errors = build_request_validation_errors(error)
        log_location_validation_failure(
            db,
            current_admin=current_admin,
            payload=payload,
            validation_message=normalize_location_validation_messages(request_validation_errors),
        )
        raise RequestValidationError(request_validation_errors) from error

    location = db.get(ManagedLocation, validated_payload.location_id) if validated_payload.location_id is not None else None
    if validated_payload.location_id is not None and location is None:
        raise HTTPException(status_code=404, detail="Localizacao nao encontrada.")

    effective_admin_projects = resolve_effective_admin_project_names(db, current_admin) or []
    effective_admin_project_set = set(effective_admin_projects)
    if not effective_admin_project_set:
        raise HTTPException(status_code=403, detail="Administrador sem projetos vinculados nao pode alterar localizacoes.")

    timestamp = now_sgt()
    coordinates = [
        {"latitude": coordinate.latitude, "longitude": coordinate.longitude}
        for coordinate in (validated_payload.coordinates or [])
    ]
    previous_coordinates = extract_location_coordinates(location) if location is not None else None
    previous_projects = extract_location_projects(location) if location is not None else None
    previous_tolerance_meters = int(location.tolerance_meters) if location is not None else None

    try:
        location_projects = resolve_location_projects_for_upsert(
            db,
            project_names=validated_payload.projects,
            existing_location=location,
        )
    except HTTPException as error:
        if error.status_code == 422:
            log_location_validation_failure(
                db,
                current_admin=current_admin,
                payload={
                    "local": validated_payload.local,
                    "projects": validated_payload.projects,
                    "coordinates": coordinates,
                    "tolerance_meters": validated_payload.tolerance_meters,
                },
                validation_message=str(error.detail),
            )
        raise

    existing_location_project_set = set(previous_projects or [])
    requested_location_project_set = set(location_projects)
    preserves_existing_detached_projects = (
        location is not None
        and requested_location_project_set == existing_location_project_set
        and bool(requested_location_project_set)
        and requested_location_project_set.isdisjoint(effective_admin_project_set)
    )

    if location is not None and not preserves_existing_detached_projects and not location_matches_effective_admin_scope(
        db,
        current_admin,
        existing_location_project_set,
        admin_project_names=effective_admin_projects,
        allow_global_locations=False,
    ):
        raise HTTPException(status_code=403, detail="Localizacao fora do escopo do administrador.")

    if not preserves_existing_detached_projects and any(
        project_name not in effective_admin_project_set for project_name in location_projects
    ):
        raise HTTPException(status_code=403, detail="Nao e possivel salvar localizacoes com projetos fora do seu escopo.")

    primary_coordinate = coordinates[0]
    coordinates_json = dump_location_coordinates(coordinates)
    projects_json = dump_location_projects(location_projects)
    created = False
    if location is None:
        location = ManagedLocation(
            local=validated_payload.local,
            latitude=primary_coordinate["latitude"],
            longitude=primary_coordinate["longitude"],
            coordinates_json=coordinates_json,
            projects_json=projects_json,
            tolerance_meters=validated_payload.tolerance_meters,
            created_at=timestamp,
            updated_at=timestamp,
        )
        db.add(location)
        created = True
    else:
        location.local = validated_payload.local
        location.latitude = primary_coordinate["latitude"]
        location.longitude = primary_coordinate["longitude"]
        location.coordinates_json = coordinates_json
        location.projects_json = projects_json
        location.tolerance_meters = validated_payload.tolerance_meters
        location.updated_at = timestamp

    geometry_changed = created or (
        previous_coordinates != coordinates
        or previous_tolerance_meters != validated_payload.tolerance_meters
    )

    log_event(
        db,
        source="admin",
        action="location",
        status="created" if created else "updated",
        message=build_location_upsert_event_message(created=created, geometry_changed=geometry_changed),
        local=validated_payload.local,
        request_path="/api/admin/locations",
        http_status=200,
        details=build_location_upsert_event_details(
            current_admin=current_admin,
            location_id=None if created else int(location.id or 0),
            coordinates=coordinates,
            projects=location_projects,
            tolerance_meters=validated_payload.tolerance_meters,
            geometry_changed=geometry_changed,
            previous_coordinates=previous_coordinates,
            previous_projects=previous_projects,
            previous_tolerance_meters=previous_tolerance_meters,
        ),
    )
    db.commit()
    notify_admin_views("location", "event")
    return AdminActionResponse(ok=True, message="Localizacao salva com sucesso.")


@router.post("/locations/settings", response_model=AdminLocationSettingsResponse, dependencies=[Depends(require_full_admin_session)])
def update_location_settings(
    payload: AdminLocationSettingsUpdate,
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> AdminLocationSettingsResponse:
    previous_accuracy_threshold_meters = get_location_accuracy_threshold_meters(db)
    previous_mixed_zone_interval_minutes = get_mixed_zone_interval_minutes(db)
    settings = upsert_location_settings(
        db,
        accuracy_threshold_meters=payload.location_accuracy_threshold_meters,
        mixed_zone_interval_minutes=payload.mixed_zone_interval_minutes,
    )
    log_message = build_location_settings_log_message(
        previous_accuracy_threshold_meters=previous_accuracy_threshold_meters,
        current_accuracy_threshold_meters=settings.location_accuracy_threshold_meters,
        previous_mixed_zone_interval_minutes=previous_mixed_zone_interval_minutes,
        current_mixed_zone_interval_minutes=settings.mixed_zone_interval_minutes,
    )
    log_event(
        db,
        source="admin",
        action="location_config",
        status="updated",
        message=log_message,
        request_path="/api/admin/locations/settings",
        http_status=200,
        details=(
            f"updated_by={current_admin.chave}; "
            f"previous_location_accuracy_threshold_meters={previous_accuracy_threshold_meters}; "
            f"location_accuracy_threshold_meters={settings.location_accuracy_threshold_meters}; "
            f"previous_mixed_zone_interval_minutes={previous_mixed_zone_interval_minutes}; "
            f"mixed_zone_interval_minutes={settings.mixed_zone_interval_minutes}"
        ),
    )
    db.commit()
    notify_admin_views("location", "event")
    return AdminLocationSettingsResponse(
        ok=True,
        message="Configuracoes de localizacao salvas com sucesso.",
        location_accuracy_threshold_meters=settings.location_accuracy_threshold_meters,
        mixed_zone_interval_minutes=settings.mixed_zone_interval_minutes,
    )


@router.delete("/locations/{location_id}", response_model=AdminActionResponse, dependencies=[Depends(require_full_admin_session)])
def remove_location(
    location_id: int,
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> AdminActionResponse:
    location = db.get(ManagedLocation, location_id)
    if location is None:
        log_event(
            db,
            source="admin",
            action="location",
            status="failed",
            message="Location not found for removal",
            request_path=f"/api/admin/locations/{location_id}",
            http_status=404,
            details=f"updated_by={current_admin.chave}; location_id={location_id}",
            commit=True,
        )
        raise HTTPException(status_code=404, detail="Localizacao nao encontrada.")

    effective_admin_projects = resolve_effective_admin_project_names(db, current_admin) or []
    if not location_matches_effective_admin_scope(
        db,
        current_admin,
        extract_location_projects(location),
        admin_project_names=effective_admin_projects,
        allow_global_locations=False,
    ):
        raise HTTPException(status_code=403, detail="Localizacao fora do escopo do administrador.")

    location_name = location.local
    db.delete(location)
    log_event(
        db,
        source="admin",
        action="location",
        status="removed",
        message="Location removed via admin",
        local=location_name,
        request_path=f"/api/admin/locations/{location_id}",
        http_status=200,
        details=f"updated_by={current_admin.chave}; location_id={location_id}",
    )
    db.commit()
    notify_admin_views("location", "event")
    return AdminActionResponse(ok=True, message="Localizacao removida com sucesso.")


@router.get("/users", response_model=list[AdminUserListRow], dependencies=[Depends(require_full_admin_session)])
def list_users(
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> list[AdminUserListRow]:
    rows = db.execute(select(User).order_by(User.nome, User.rfid)).scalars().all()
    rows, project_names_by_user_id, _ = filter_users_for_admin_scope(
        db,
        rows,
        current_admin=current_admin,
    )
    payload_rows: list[AdminUserListRow] = []
    for row in rows:
        row_projects = project_names_by_user_id.get(row.id, []) if row.id is not None else []
        active_project = resolve_user_active_project(row, row_projects)
        payload_rows.append(
            AdminUserListRow(
                id=row.id,
                rfid=row.rfid,
                nome=row.nome,
                chave=row.chave,
                perfil=row.perfil,
                projeto=active_project,
                projeto_ativo=active_project,
                projetos=row_projects,
                vehicle_id=row.vehicle_id,
                workplace=row.workplace,
                placa=row.placa,
                end_rua=row.end_rua,
                zip=row.zip,
                cargo=row.cargo,
                email=row.email,
            )
        )
    return payload_rows


@router.post("/users", dependencies=[Depends(require_full_admin_session)])
def upsert_user(
    payload: AdminUserUpsert,
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> dict:
    requested_project_names = ensure_known_projects(
        db,
        payload.projetos or ([payload.projeto] if payload.projeto is not None else []),
        detail="Um ou mais projetos do usuário nao existem no catalogo atual.",
    )
    payload_fields = set(getattr(payload, "model_fields_set", set()))
    placa_was_provided = "placa" in payload_fields
    vehicle_id_was_provided = "vehicle_id" in payload_fields
    vehicle_link_was_provided = placa_was_provided or vehicle_id_was_provided
    effective_admin_projects = resolve_effective_admin_project_names(db, current_admin) or []
    effective_admin_project_set = set(effective_admin_projects)
    if not effective_admin_project_set:
        raise HTTPException(status_code=403, detail="Administrador sem projetos vinculados nao pode alterar usuarios.")

    requested_active_project = ensure_known_project(db, payload.projeto) if payload.projeto is not None else None

    user = None
    linked_existing_user = False
    if payload.user_id is not None:
        user = db.get(User, payload.user_id)
        if user is None:
            raise HTTPException(status_code=404, detail="User not found")
    elif payload.rfid:
        user = find_user_by_rfid(db, payload.rfid)
        if user is None:
            user = find_user_by_chave(db, payload.chave)
            if user is not None and user.rfid is None:
                linked_existing_user = True

    conflicting_user = find_user_by_chave(db, payload.chave)
    if conflicting_user is not None and (user is None or conflicting_user.id != user.id):
        if user is None and conflicting_user.rfid is None and payload.rfid is not None:
            user = conflicting_user
            linked_existing_user = True
        else:
            raise HTTPException(status_code=409, detail="Ja existe um usuario cadastrado com essa chave")

    if payload.rfid is not None:
        conflicting_rfid_user = find_user_by_rfid(db, payload.rfid)
        if conflicting_rfid_user is not None and (user is None or conflicting_rfid_user.id != user.id):
            raise HTTPException(status_code=409, detail="Ja existe um usuario cadastrado com esse RFID")

    linked_vehicle = None
    if vehicle_link_was_provided:
        try:
            linked_vehicle = resolve_vehicle_for_user_transport_link(
                db,
                vehicle_id=payload.vehicle_id if vehicle_id_was_provided else None,
                plate=payload.placa if placa_was_provided else None,
            )
        except ValueError as exc:
            if str(exc) == "Vehicle not found for the provided id.":
                raise HTTPException(status_code=404, detail="Veiculo nao encontrado para o identificador informado") from exc
            if str(exc) == "Vehicle not found for the provided plate.":
                raise HTTPException(status_code=404, detail="Veiculo nao encontrado para a placa informada") from exc
            if str(exc) == "The provided vehicle_id does not match the provided plate.":
                raise HTTPException(status_code=409, detail="O veiculo informado nao corresponde a placa enviada") from exc
            raise

    if payload.workplace is not None:
        workplace = db.execute(select(Workplace).where(Workplace.workplace == payload.workplace)).scalar_one_or_none()
        if workplace is None:
            raise HTTPException(status_code=404, detail="Workplace nao encontrado para o nome informado")

    if user is not None and user_has_admin_access(user) and not user_profile_has_access(payload.perfil, ADMIN_ACCESS_DIGIT):
        total_admins = sum(
            1
            for row in db.execute(select(User).where(User.perfil != 0)).scalars().all()
            if user_has_admin_access(row)
        )
        if total_admins <= 1:
            raise HTTPException(status_code=409, detail="Nao e possivel remover o unico administrador ativo do sistema.")

    if user is not None and not user_matches_effective_admin_scope(db, current_admin, user, admin_project_names=effective_admin_projects):
        raise HTTPException(status_code=404, detail="User not found")

    previous_project_names = list_user_project_names(db, user) if user is not None else []
    previous_active_project = user.projeto if user is not None else None

    if user is None:
        unauthorized_projects = [
            project_name
            for project_name in requested_project_names
            if project_name not in effective_admin_project_set
        ]
        if unauthorized_projects:
            raise HTTPException(status_code=403, detail="Nao e possivel vincular o usuario a projetos fora do seu escopo.")
    else:
        preserved_projects_outside_scope = [
            project_name
            for project_name in previous_project_names
            if project_name not in effective_admin_project_set
        ]
        unauthorized_projects = [
            project_name
            for project_name in requested_project_names
            if project_name not in effective_admin_project_set and project_name not in preserved_projects_outside_scope
        ]
        if unauthorized_projects:
            raise HTTPException(status_code=403, detail="Nao e possivel adicionar projetos fora do seu escopo ao usuario.")

        scoped_requested_projects = [
            project_name
            for project_name in requested_project_names
            if project_name in effective_admin_project_set
        ]
        requested_project_names = normalize_user_project_names(
            [*scoped_requested_projects, *preserved_projects_outside_scope]
        )

    if user:
        previous_key = user.chave
        user.nome = payload.nome
        user.chave = payload.chave
        user.perfil = normalize_user_profile(payload.perfil)
        user.workplace = payload.workplace
        user.rfid = payload.rfid
        if vehicle_link_was_provided:
            sync_user_vehicle_reference(user, linked_vehicle)
        user.end_rua = payload.end_rua
        user.zip = payload.zip
        user.cargo = payload.cargo
        user.email = payload.email
        replace_user_project_memberships(db, user, requested_project_names)
        if (
            requested_active_project is not None
            and requested_active_project in requested_project_names
            and requested_active_project in effective_admin_project_set
        ):
            user.projeto = requested_active_project
        if previous_key != user.chave:
            db.execute(
                update(UserSyncEvent)
                .where(UserSyncEvent.user_id == user.id)
                .values(chave=user.chave)
            )
            db.execute(
                update(CheckingHistory)
                .where(CheckingHistory.chave == previous_key)
                .values(chave=user.chave)
            )
    else:
        if payload.rfid is None:
            raise HTTPException(status_code=400, detail="RFID is required for new users")
        user = User(
            rfid=payload.rfid,
            nome=payload.nome,
            chave=payload.chave,
            perfil=normalize_user_profile(payload.perfil),
            projeto=requested_project_names[0],
            workplace=payload.workplace,
            vehicle_id=linked_vehicle.id if linked_vehicle is not None else None,
            placa=linked_vehicle.placa if linked_vehicle is not None else None,
            end_rua=payload.end_rua,
            zip=payload.zip,
            cargo=payload.cargo,
            email=payload.email,
            local=None,
            checkin=None,
            time=None,
            last_active_at=now_sgt(),
            inactivity_days=0,
        )
        db.add(user)
        db.flush()
        replace_user_project_memberships(db, user, requested_project_names)

    current_project_names = list_user_project_names(db, user)
    current_active_project = user.projeto

    pending = None
    if payload.rfid is not None:
        pending = db.execute(select(PendingRegistration).where(PendingRegistration.rfid == payload.rfid)).scalar_one_or_none()
    if pending:
        db.delete(pending)

    log_event(
        db,
        idempotency_key=f"register-{uuid4()}",
        source="admin",
        action="register",
        status="done",
        message="User registered via admin",
        rfid=payload.rfid,
        project=current_active_project,
        request_path="/api/admin/users",
        http_status=200,
        submitted_at=now_sgt(),
        details=(
            f"updated_by={current_admin.chave}; chave={payload.chave}; "
            f"nome={payload.nome}; perfil={normalize_user_profile(payload.perfil)}; linked_existing_user={linked_existing_user}; "
            f"previous_projects={json.dumps(previous_project_names, ensure_ascii=True)}; "
            f"current_projects={json.dumps(current_project_names, ensure_ascii=True)}; "
            f"previous_active_project={previous_active_project or '-'}; "
            f"current_active_project={current_active_project or '-'}; "
            f"placa={(linked_vehicle.placa if vehicle_link_was_provided and linked_vehicle is not None else user.placa) or '-'}; "
            f"vehicle_id={(linked_vehicle.id if vehicle_link_was_provided and linked_vehicle is not None else user.vehicle_id) or '-'}"
        ),
    )
    db.commit()
    notify_admin_views("register", "event")

    return {
        "ok": True,
        "rfid": user.rfid,
        "user_id": user.id,
        "linked_existing_user": linked_existing_user,
    }


@router.delete("/pending/{pending_id}", dependencies=[Depends(require_full_admin_session)])
def remove_pending(
    pending_id: int,
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> dict:
    pending = db.get(PendingRegistration, pending_id)
    if pending is None:
        log_event(
            db,
            source="admin",
            action="pending",
            status="failed",
            message="Pending registration not found for removal",
            request_path=f"/api/admin/pending/{pending_id}",
            http_status=404,
            details=f"updated_by={current_admin.chave}; pending_id={pending_id}",
            commit=True,
        )
        raise HTTPException(status_code=404, detail="Pending registration not found")

    effective_admin_projects = resolve_effective_admin_project_names(db, current_admin) or []
    latest_scan_local_by_rfid, location_projects_by_local = build_pending_registration_scope_maps(db, [pending])
    if not pending_matches_admin_scope(
        db,
        pending,
        current_admin=current_admin,
        admin_project_names=effective_admin_projects,
        latest_scan_local_by_rfid=latest_scan_local_by_rfid,
        location_projects_by_local=location_projects_by_local,
    ):
        raise HTTPException(status_code=404, detail="Pending registration not found")

    pending_rfid = pending.rfid
    db.delete(pending)
    log_event(
        db,
        source="admin",
        action="pending",
        status="removed",
        message="Pending registration removed via admin",
        rfid=pending_rfid,
        request_path=f"/api/admin/pending/{pending_id}",
        http_status=200,
        details=f"updated_by={current_admin.chave}; pending_id={pending_id}",
    )
    db.commit()
    notify_admin_views("pending", "event")
    return {"ok": True, "id": pending_id}


@router.delete("/users/{user_id}", dependencies=[Depends(require_full_admin_session)])
def remove_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> dict:
    user = db.get(User, user_id)
    if user is None:
        log_event(
            db,
            source="admin",
            action="register",
            status="failed",
            message="User not found for removal",
            request_path=f"/api/admin/users/{user_id}",
            http_status=404,
            details=f"updated_by={current_admin.chave}; user_id={user_id}",
            commit=True,
        )
        raise HTTPException(status_code=404, detail="User not found")

    effective_admin_projects = resolve_effective_admin_project_names(db, current_admin) or []
    effective_admin_project_set = set(effective_admin_projects)
    if not effective_admin_project_set:
        raise HTTPException(status_code=403, detail="Administrador sem projetos vinculados nao pode remover usuarios.")

    user_project_names = list_user_project_names(db, user)
    if not user_matches_effective_admin_scope(
        db,
        current_admin,
        user,
        admin_project_names=effective_admin_projects,
        user_project_names=user_project_names,
    ):
        raise HTTPException(status_code=404, detail="User not found")

    if any(project_name not in effective_admin_project_set for project_name in user_project_names):
        raise HTTPException(
            status_code=403,
            detail="Nao e possivel remover um usuario com projetos fora do seu escopo.",
        )

    if user_has_admin_access(user):
        total_admins = sum(
            1
            for row in db.execute(select(User).where(User.perfil != 0)).scalars().all()
            if user_has_admin_access(row)
        )
        if total_admins <= 1:
            raise HTTPException(status_code=409, detail="Nao e possivel remover o unico administrador ativo do sistema.")

    user_rfid = user.rfid
    user_key = user.chave
    pending = None
    if user_rfid is not None:
        pending = db.execute(select(PendingRegistration).where(PendingRegistration.rfid == user_rfid)).scalar_one_or_none()
    if pending is not None:
        db.delete(pending)

    # Delete transport data (no cascade on transport_requests.user_id)
    user_request_ids = db.execute(
        select(TransportRequest.id).where(TransportRequest.user_id == user.id)
    ).scalars().all()
    if user_request_ids:
        db.execute(delete(TransportAssignment).where(TransportAssignment.request_id.in_(user_request_ids)))
        db.execute(delete(TransportRequest).where(TransportRequest.user_id == user.id))

    db.execute(delete(UserSyncEvent).where(UserSyncEvent.user_id == user.id))
    db.delete(user)
    log_event(
        db,
        source="admin",
        action="register",
        status="removed",
        message="User removed via admin",
        rfid=user_rfid,
        request_path=f"/api/admin/users/{user_id}",
        http_status=200,
        details=(
            f"updated_by={current_admin.chave}; chave={user_key}; user_id={user_id}; "
            f"pending_removed={pending is not None}"
        ),
    )
    db.commit()
    notify_admin_views("register", "event")
    return {"ok": True, "user_id": user_id}


@router.post("/users/{user_id}/reset-password", response_model=AdminActionResponse)
def reset_user_password(
    user_id: int,
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> AdminActionResponse:
    user = db.get(User, user_id)
    if user is None:
        log_event(
            db,
            source="admin",
            action="password",
            status="failed",
            message="User password reset failed because target user was not found",
            request_path=f"/api/admin/users/{user_id}/reset-password",
            http_status=404,
            details=f"updated_by={current_admin.chave}; user_id={user_id}",
            commit=True,
        )
        raise HTTPException(status_code=404, detail="Usuario nao encontrado.")

    had_password = bool(user.senha)
    if had_password:
        user.senha = None

    log_event(
        db,
        source="admin",
        action="password",
        status="removed" if had_password else "noop",
        message="Web user password removed via admin" if had_password else "Web user password reset requested via admin but user already had no password",
        request_path=f"/api/admin/users/{user_id}/reset-password",
        http_status=200,
        details=(
            f"updated_by={current_admin.chave}; chave={user.chave}; "
            f"user_id={user_id}; had_password={had_password}"
        ),
    )
    db.commit()
    notify_admin_views("register", "event")
    if had_password:
        return AdminActionResponse(
            ok=True,
            message="Senha removida com sucesso. O usuario podera cadastrar uma nova senha.",
        )
    return AdminActionResponse(
        ok=True,
        message="Esse usuario ja esta sem senha cadastrada e ja pode cadastrar uma nova senha.",
    )


@router.get("/events", response_model=list[EventRow])
def list_events(
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> list[EventRow]:
    rows = db.execute(
        select(CheckEvent)
        .where(CheckEvent.action != "event_archive")
        .order_by(desc(CheckEvent.id))
        .limit(200)
    ).scalars().all()
    rows = filter_check_events_for_admin_scope(db, rows, current_admin=current_admin)
    return build_event_row_payload(
        rows,
        db,
        can_view_activity_time=user_can_view_activity_time(current_admin),
    )


@router.get(
    "/database-events",
    response_model=DatabaseEventListResponse,
    dependencies=[Depends(require_full_admin_session)],
)
def list_database_events(
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
    action: str | None = Query(default=None),
    project: str | None = Query(default=None),
    source: str | None = Query(default=None),
    status: str | None = Query(default=None),
    chave: str | None = Query(default=None, min_length=1, max_length=4),
    rfid: str | None = Query(default=None, min_length=1, max_length=64),
    search: str | None = Query(default=None, min_length=1, max_length=120),
    from_date: date | None = Query(default=None),
    to_date: date | None = Query(default=None),
    sort_by: str = Query(default=DATABASE_EVENT_DEFAULT_SORT_BY),
    sort_direction: str = Query(default=DATABASE_EVENT_DEFAULT_SORT_DIRECTION),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=DATABASE_EVENT_PAGE_SIZE, ge=1, le=200),
) -> DatabaseEventListResponse:
    normalized_action = str(action or "").strip().lower() or None
    if normalized_action and normalized_action not in DATABASE_EVENT_ACTIONS:
        raise HTTPException(status_code=400, detail="Acao invalida para a consulta de eventos do banco de dados.")

    normalized_project = str(project or "").strip().upper() or None
    normalized_source = str(source or "").strip().lower() or None
    normalized_status = str(status or "").strip().lower() or None
    normalized_key = str(chave or "").strip().upper() or None
    normalized_rfid = str(rfid or "").strip() or None
    normalized_search = str(search or "").strip().lower() or None
    normalized_sort_by = str(sort_by or "").strip().lower() or DATABASE_EVENT_DEFAULT_SORT_BY
    normalized_sort_direction = str(sort_direction or "").strip().lower() or DATABASE_EVENT_DEFAULT_SORT_DIRECTION

    if normalized_sort_by not in DATABASE_EVENT_SORTABLE_FIELDS:
        raise HTTPException(status_code=400, detail="Coluna invalida para ordenacao de eventos.")
    if normalized_sort_direction not in DATABASE_EVENT_SORT_DIRECTIONS:
        raise HTTPException(status_code=400, detail="Direcao invalida para ordenacao de eventos.")

    effective_admin_projects = resolve_effective_admin_project_names(db, current_admin) or []
    effective_admin_project_set = set(effective_admin_projects)
    filter_options = build_database_event_filter_options(db, allowed_project_names=effective_admin_projects)

    if not effective_admin_project_set:
        return DatabaseEventListResponse(
            items=[],
            total=0,
            page=1,
            page_size=page_size,
            total_pages=1,
            filter_options=filter_options,
        )

    if normalized_project and normalized_project not in effective_admin_project_set:
        return DatabaseEventListResponse(
            items=[],
            total=0,
            page=1,
            page_size=page_size,
            total_pages=1,
            filter_options=filter_options,
        )

    if from_date and to_date and from_date > to_date:
        raise HTTPException(status_code=400, detail="Intervalo de datas invalido para a consulta de eventos.")

    query = select(CheckEvent).where(
        CheckEvent.action.in_(DATABASE_EVENT_ACTIONS),
        CheckEvent.project.in_(effective_admin_projects),
    )

    if normalized_action:
        query = query.where(CheckEvent.action == normalized_action)
    if normalized_project:
        query = query.where(CheckEvent.project == normalized_project)
    if normalized_source:
        query = query.where(func.lower(CheckEvent.source) == normalized_source)
    if normalized_status:
        query = query.where(func.lower(CheckEvent.status) == normalized_status)
    if normalized_rfid:
        query = query.where(CheckEvent.rfid == normalized_rfid)
    if normalized_search:
        like_pattern = f"%{normalized_search}%"
        query = query.where(
            or_(
                func.lower(func.coalesce(CheckEvent.rfid, "")).like(like_pattern),
                func.lower(func.coalesce(CheckEvent.source, "")).like(like_pattern),
                func.lower(func.coalesce(CheckEvent.device_id, "")).like(like_pattern),
                func.lower(func.coalesce(CheckEvent.local, "")).like(like_pattern),
                func.lower(func.coalesce(CheckEvent.status, "")).like(like_pattern),
                func.lower(func.coalesce(CheckEvent.message, "")).like(like_pattern),
                func.lower(func.coalesce(CheckEvent.details, "")).like(like_pattern),
                func.lower(func.coalesce(CheckEvent.project, "")).like(like_pattern),
                func.lower(func.coalesce(CheckEvent.request_path, "")).like(like_pattern),
            )
        )
    if from_date:
        from_datetime = datetime.combine(from_date, dt_time.min, tzinfo=now_sgt().tzinfo)
        query = query.where(CheckEvent.event_time >= from_datetime)
    if to_date:
        to_datetime = datetime.combine(to_date + timedelta(days=1), dt_time.min, tzinfo=now_sgt().tzinfo)
        query = query.where(CheckEvent.event_time < to_datetime)

    if normalized_key:
        key_match = db.execute(select(User.rfid).where(User.chave == normalized_key, User.rfid.is_not(None))).scalars().all()
        details_pattern = f"%{normalized_key}%"
        key_conditions = [func.upper(func.coalesce(CheckEvent.details, "")).like(details_pattern)]
        if key_match:
            key_conditions.append(CheckEvent.rfid.in_(key_match))
        query = query.where(or_(*key_conditions))

    count_query = select(func.count()).select_from(query.subquery())
    total = db.execute(count_query).scalar_one()
    total_pages = max(1, (total + page_size - 1) // page_size)
    current_page = min(page, total_pages)

    if normalized_sort_by == "chave":
        rows = db.execute(query).scalars().all()
        sorted_items = sort_database_event_payload(
            build_event_row_payload(rows, db),
            sort_by=normalized_sort_by,
            sort_direction=normalized_sort_direction,
        )
        offset = (current_page - 1) * page_size
        paginated_items = sorted_items[offset: offset + page_size]
        return DatabaseEventListResponse(
            items=paginated_items,
            total=total,
            page=current_page,
            page_size=page_size,
            total_pages=total_pages,
            filter_options=filter_options,
        )

    sort_expression = DATABASE_EVENT_SQL_SORT_FIELDS[normalized_sort_by]
    sort_function = asc if normalized_sort_direction == "asc" else desc
    rows = db.execute(
        query
        .order_by(sort_function(sort_expression), sort_function(CheckEvent.id))
        .offset((current_page - 1) * page_size)
        .limit(page_size)
    ).scalars().all()

    return DatabaseEventListResponse(
        items=build_event_row_payload(rows, db),
        total=total,
        page=current_page,
        page_size=page_size,
        total_pages=total_pages,
        filter_options=filter_options,
    )


@router.post("/events/archive", response_model=EventArchiveCreateResponse, dependencies=[Depends(require_full_admin_session)])
def archive_events(
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> EventArchiveCreateResponse:
    current_rows = db.execute(
        select(CheckEvent)
        .order_by(CheckEvent.event_time, CheckEvent.id)
    ).scalars().all()
    rows = [row for row in current_rows if row.action != "event_archive"]
    archive = create_event_archive(rows)
    pruned_archive_rows = len(current_rows) - len(rows)

    if current_rows:
        db.execute(delete(CheckEvent))
        db.commit()

    if archive is not None:
        log_event(
            db,
            source="admin",
            action="event_archive",
            status="created",
            message="Event log archive created",
            request_path="/api/admin/events/archive",
            http_status=200,
            details=(
                f"file_name={archive.file_name}; period={archive.period}; "
                f"record_count={archive.record_count}; pruned_archive_rows={pruned_archive_rows}; "
                f"created_by={current_admin.chave}"
            )[:1000],
            commit=True,
        )
    else:
        log_event(
            db,
            source="admin",
            action="event_archive",
            status="noop",
            message="Event log archive requested but there were no current events to archive",
            request_path="/api/admin/events/archive",
            http_status=200,
            details=f"pruned_archive_rows={pruned_archive_rows}; created_by={current_admin.chave}",
            commit=True,
        )

    archives_page = list_event_archives_page()

    return EventArchiveCreateResponse(
        created=archive is not None,
        cleared_count=len(rows) if archive is not None else 0,
        archive=EventArchiveRow(**archive.__dict__) if archive is not None else None,
        archives=EventArchiveListResponse(
            items=[EventArchiveRow(**item.__dict__) for item in archives_page.items],
            total=archives_page.total,
            total_size_bytes=archives_page.total_size_bytes,
            page=archives_page.page,
            page_size=archives_page.page_size,
            total_pages=archives_page.total_pages,
            query=archives_page.query,
        ),
    )


@router.get("/events/archives", response_model=EventArchiveListResponse, dependencies=[Depends(require_full_admin_session)])
def get_event_archives(
    q: str = Query(default=""),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=8, ge=1, le=100),
) -> EventArchiveListResponse:
    archives_page = list_event_archives_page(query=q, page=page, page_size=page_size)
    return EventArchiveListResponse(
        items=[EventArchiveRow(**item.__dict__) for item in archives_page.items],
        total=archives_page.total,
        total_size_bytes=archives_page.total_size_bytes,
        page=archives_page.page,
        page_size=archives_page.page_size,
        total_pages=archives_page.total_pages,
        query=archives_page.query,
    )


@router.get("/events/archives/download-all", dependencies=[Depends(require_full_admin_session)])
def download_all_event_archives(
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> Response:
    try:
        file_name, payload = build_event_archives_zip()
    except FileNotFoundError as exc:
        log_event(
            db,
            source="admin",
            action="event_archive",
            status="failed",
            message="Download of all archived event logs failed because there are no archives",
            request_path="/api/admin/events/archives/download-all",
            http_status=404,
            details=f"downloaded_by={current_admin.chave}",
            commit=True,
        )
        raise HTTPException(status_code=404, detail="No archived event logs found") from exc

    log_event(
        db,
        source="admin",
        action="event_archive",
        status="downloaded",
        message="All archived event logs downloaded as zip",
        request_path="/api/admin/events/archives/download-all",
        http_status=200,
        details=f"file_name={file_name}; downloaded_by={current_admin.chave}",
        commit=True,
    )

    return Response(
        content=payload,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
    )


@router.get("/events/archives/{file_name}", dependencies=[Depends(require_full_admin_session)])
def download_event_archive(
    file_name: str,
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> FileResponse:
    try:
        archive_path = get_event_archive_path(file_name)
    except FileNotFoundError as exc:
        log_event(
            db,
            source="admin",
            action="event_archive",
            status="failed",
            message="Archived event log download failed because file was not found",
            request_path=f"/api/admin/events/archives/{file_name}",
            http_status=404,
            details=f"file_name={file_name}; downloaded_by={current_admin.chave}",
            commit=True,
        )
        raise HTTPException(status_code=404, detail="Archived event log not found") from exc

    log_event(
        db,
        source="admin",
        action="event_archive",
        status="downloaded",
        message="Archived event log downloaded",
        request_path=f"/api/admin/events/archives/{file_name}",
        http_status=200,
        details=f"file_name={file_name}; downloaded_by={current_admin.chave}",
        commit=True,
    )

    return FileResponse(path=archive_path, media_type="text/csv", filename=archive_path.name)


@router.delete("/events/archives/{file_name}", dependencies=[Depends(require_full_admin_session)])
def remove_event_archive(
    file_name: str,
    db: Session = Depends(get_db),
    current_admin: User = Depends(require_full_admin_session),
) -> dict:
    try:
        delete_event_archive(file_name)
    except FileNotFoundError as exc:
        log_event(
            db,
            source="admin",
            action="event_archive",
            status="failed",
            message="Archived event log removal failed because file was not found",
            request_path=f"/api/admin/events/archives/{file_name}",
            http_status=404,
            details=f"file_name={file_name}; removed_by={current_admin.chave}",
            commit=True,
        )
        raise HTTPException(status_code=404, detail="Archived event log not found") from exc

    log_event(
        db,
        source="admin",
        action="event_archive",
        status="removed",
        message="Archived event log removed",
        request_path=f"/api/admin/events/archives/{file_name}",
        http_status=200,
        details=f"file_name={file_name}; removed_by={current_admin.chave}",
        commit=True,
    )

    return {"ok": True, "file_name": file_name}
